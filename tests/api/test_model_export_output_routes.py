"""Integration tests for money-output routes in app/api/routers/ui_model_outputs.py.

Covers:
  - GET  /ui/models/{id}/investor-export/preflight       (resend eligibility)
  - GET  /ui/models/{id}/investor-export.xlsx            (workbook bytes)
  - POST /ui/models/{id}/investor-export/async           (ExportJob row + enqueue)
  - GET  /ui/exports/{job_id}/status                     (poll contract)
  - POST /ui/exports/{job_id}/resend                     (cached-bytes copy job)
  - POST /ui/models/{id}/noi-inputs                      (OperationalInputs write)
  - POST /ui/models/{id}/history/{snapshot_id}/revert    (snapshot restore)
  - POST /ui/models/{id}/save-as-template                (ScenarioTemplate row)

Celery note: the async-export and resend handlers commit an ExportJob row,
then enqueue via ``celery_app.send_task``. The broker is unreachable in this
test environment, so ``send_task`` is monkeypatched to a recorder — the
tests still assert the full synchronous contract (job row persisted with
the right profile/user/bytes, response JSON, and that the correct task name
+ job id were dispatched). Actual worker execution needs a live worker and
is out of scope here.

Auth: preflight/async/status/resend/history/save-as-template resolve the
user from the session cookie (``_get_user``); the ``require_auth_for_ui``
middleware additionally 303-redirects unauthenticated non-HTMX /ui/
requests — so authenticated tests use ``set_client_auth`` and the 401
tests send ``HX-Request: true`` to reach the handler's own auth guard.
"""

from __future__ import annotations

import uuid
from decimal import Decimal
from io import BytesIO

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.deal import OperationalInputs, ScenarioSnapshot, UseLine
from app.models.export_job import ExportJob, ExportJobStatus
from app.models.project import Project

from tests.conftest import (
    seed_deal_model,
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
    set_client_auth,
)

pytestmark = pytest.mark.asyncio

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


async def _seed_model(session: AsyncSession):
    """Financial-seeded scenario; returns plain ids captured pre-request."""
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user)
    deal_model, inputs, _income, _opex = await seed_deal_model_with_financials(
        session, opp, user
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()
    await session.commit()
    return deal_model.id, project.id, user.id, inputs.id


@pytest.fixture
def celery_recorder(monkeypatch):
    """Replace celery_app.send_task with a call recorder (broker unreachable
    in tests; the enqueue itself is the synchronous contract we assert)."""
    from app.tasks.celery_app import celery_app

    calls: list[tuple[str, list]] = []

    def _fake_send_task(name, args=None, kwargs=None, **_kw):
        calls.append((name, list(args or [])))
        return None

    monkeypatch.setattr(celery_app, "send_task", _fake_send_task)
    return calls


# ---------------------------------------------------------------------------
# GET /ui/models/{id}/investor-export.xlsx
# ---------------------------------------------------------------------------


async def test_investor_export_xlsx_returns_openable_workbook(
    client: AsyncClient, session: AsyncSession
) -> None:
    model_id, _project_id, user_id, _inputs_id = await _seed_model(session)
    set_client_auth(client, user_id)

    resp = await client.get(f"/ui/models/{model_id}/investor-export.xlsx")
    assert resp.status_code == 200, resp.text[:500]
    assert resp.headers["content-type"].startswith(_XLSX_MIME)
    assert "attachment" in resp.headers.get("content-disposition", "")

    wb = load_workbook(BytesIO(resp.content))
    # Core money sheets must exist (roster asserted in full by the exporter
    # suite; here we prove the HTTP route streams a real workbook).
    for sheet in ("Cover", "Underwriting Summary", "Sources & Uses",
                  "Glossary & Methodology"):
        assert sheet in wb.sheetnames, f"missing sheet {sheet!r}: {wb.sheetnames}"


async def test_investor_export_xlsx_404_unknown_model(
    client: AsyncClient, session: AsyncSession
) -> None:
    _model_id, _project_id, user_id, _inputs_id = await _seed_model(session)
    set_client_auth(client, user_id)

    resp = await client.get(f"/ui/models/{uuid.uuid4()}/investor-export.xlsx")
    assert resp.status_code == 404


# ---------------------------------------------------------------------------
# GET /ui/models/{id}/investor-export/preflight
# ---------------------------------------------------------------------------


async def test_preflight_not_eligible_without_prior_export(
    client: AsyncClient, session: AsyncSession
) -> None:
    model_id, _project_id, user_id, _inputs_id = await _seed_model(session)
    set_client_auth(client, user_id)

    resp = await client.get(f"/ui/models/{model_id}/investor-export/preflight")
    assert resp.status_code == 200
    body = resp.json()
    assert body == {"resend_eligible": False, "resend_job_id": None}


async def test_preflight_eligible_with_cached_sent_job(
    client: AsyncClient, session: AsyncSession
) -> None:
    model_id, _project_id, user_id, _inputs_id = await _seed_model(session)

    job = ExportJob(
        scenario_id=model_id,
        user_id=user_id,
        recipient_email="lp@example.com",
        status=ExportJobStatus.sent,
        xlsx_bytes=b"PK-fake-workbook",
        export_profile="internal",
    )
    session.add(job)
    await session.commit()
    job_id = job.id

    set_client_auth(client, user_id)
    resp = await client.get(f"/ui/models/{model_id}/investor-export/preflight")
    assert resp.status_code == 200
    body = resp.json()
    # No OperationalOutputs exist → the cached sent job is newer than any
    # compute, so a resend is offered.
    assert body["resend_eligible"] is True
    assert body["resend_job_id"] == str(job_id)


async def test_preflight_profile_mismatch_not_eligible(
    client: AsyncClient, session: AsyncSession
) -> None:
    model_id, _project_id, user_id, _inputs_id = await _seed_model(session)
    session.add(ExportJob(
        scenario_id=model_id,
        user_id=user_id,
        recipient_email="lp@example.com",
        status=ExportJobStatus.sent,
        xlsx_bytes=b"PK-fake-workbook",
        export_profile="lp",
    ))
    await session.commit()

    set_client_auth(client, user_id)
    resp = await client.get(
        f"/ui/models/{model_id}/investor-export/preflight",
        params={"profile": "internal"},
    )
    assert resp.status_code == 200
    assert resp.json()["resend_eligible"] is False


async def test_preflight_unauthenticated_401(
    client: AsyncClient, session: AsyncSession
) -> None:
    model_id, _project_id, _user_id, _inputs_id = await _seed_model(session)

    # HX-Request bypasses the login-redirect middleware so the handler's own
    # auth guard answers.
    resp = await client.get(
        f"/ui/models/{model_id}/investor-export/preflight",
        headers={"HX-Request": "true"},
    )
    assert resp.status_code == 401


# ---------------------------------------------------------------------------
# POST /ui/models/{id}/investor-export/async + GET /ui/exports/{job_id}/status
# ---------------------------------------------------------------------------


async def test_async_export_persists_job_and_enqueues(
    client: AsyncClient, session: AsyncSession, celery_recorder
) -> None:
    from app.tasks.export import RUN_EXPORT_TASK

    model_id, _project_id, user_id, _inputs_id = await _seed_model(session)
    set_client_auth(client, user_id)

    resp = await client.post(
        f"/ui/models/{model_id}/investor-export/async",
        json={"profile": "lp"},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["status"] == "queued"
    job_id = uuid.UUID(body["job_id"])

    session.expire_all()
    job = await session.get(ExportJob, job_id)
    assert job is not None
    assert job.scenario_id == model_id
    assert job.user_id == user_id
    assert job.status == ExportJobStatus.queued
    assert job.export_profile == "lp"

    assert celery_recorder == [(RUN_EXPORT_TASK, [str(job_id)])]


async def test_async_export_unknown_profile_falls_back_to_internal(
    client: AsyncClient, session: AsyncSession, celery_recorder
) -> None:
    model_id, _project_id, user_id, _inputs_id = await _seed_model(session)
    set_client_auth(client, user_id)

    resp = await client.post(
        f"/ui/models/{model_id}/investor-export/async",
        json={"profile": "definitely-not-a-profile"},
    )
    assert resp.status_code == 200
    job_id = uuid.UUID(resp.json()["job_id"])

    session.expire_all()
    job = await session.get(ExportJob, job_id)
    assert job.export_profile == "internal"


async def test_export_status_roundtrip(
    client: AsyncClient, session: AsyncSession
) -> None:
    model_id, _project_id, user_id, _inputs_id = await _seed_model(session)
    job = ExportJob(
        scenario_id=model_id,
        user_id=user_id,
        recipient_email="lp@example.com",
        status=ExportJobStatus.failed,
        error_message="smtp exploded",
    )
    session.add(job)
    await session.commit()
    job_id = job.id

    set_client_auth(client, user_id)
    resp = await client.get(f"/ui/exports/{job_id}/status")
    assert resp.status_code == 200
    body = resp.json()
    assert body["job_id"] == str(job_id)
    assert body["status"] == "failed"
    assert body["error_message"] == "smtp exploded"
    assert body["created_at"] is not None
    assert body["completed_at"] is None


async def test_export_status_404_for_unknown_job(
    client: AsyncClient, session: AsyncSession
) -> None:
    _model_id, _project_id, user_id, _inputs_id = await _seed_model(session)
    set_client_auth(client, user_id)

    resp = await client.get(f"/ui/exports/{uuid.uuid4()}/status")
    assert resp.status_code == 404


# ---------------------------------------------------------------------------
# POST /ui/exports/{job_id}/resend
# ---------------------------------------------------------------------------


async def test_resend_spawns_copy_job_from_cached_bytes(
    client: AsyncClient, session: AsyncSession, celery_recorder
) -> None:
    from app.tasks.export import RESEND_EXPORT_TASK

    model_id, _project_id, user_id, _inputs_id = await _seed_model(session)
    src = ExportJob(
        scenario_id=model_id,
        user_id=user_id,
        recipient_email="lp@example.com",
        status=ExportJobStatus.sent,
        xlsx_bytes=b"PK-cached-build",
        filename="deal-export.xlsx",
    )
    session.add(src)
    await session.commit()
    src_id = src.id

    set_client_auth(client, user_id)
    resp = await client.post(f"/ui/exports/{src_id}/resend")
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["status"] == "queued"
    new_id = uuid.UUID(body["job_id"])
    assert new_id != src_id

    session.expire_all()
    new_job = await session.get(ExportJob, new_id)
    assert new_job.scenario_id == model_id
    assert new_job.xlsx_bytes == b"PK-cached-build"
    assert new_job.filename == "deal-export.xlsx"
    assert new_job.status == ExportJobStatus.queued

    assert celery_recorder == [(RESEND_EXPORT_TASK, [str(new_id)])]


async def test_resend_409_when_no_cached_bytes(
    client: AsyncClient, session: AsyncSession, celery_recorder
) -> None:
    model_id, _project_id, user_id, _inputs_id = await _seed_model(session)
    src = ExportJob(
        scenario_id=model_id,
        user_id=user_id,
        recipient_email="lp@example.com",
        status=ExportJobStatus.failed,
        xlsx_bytes=None,
    )
    session.add(src)
    await session.commit()
    src_id = src.id

    set_client_auth(client, user_id)
    resp = await client.post(f"/ui/exports/{src_id}/resend")
    assert resp.status_code == 409
    assert celery_recorder == []


# ---------------------------------------------------------------------------
# POST /ui/models/{id}/noi-inputs
# ---------------------------------------------------------------------------


async def test_noi_inputs_persist_to_operational_inputs(
    client: AsyncClient, session: AsyncSession
) -> None:
    model_id, _project_id, user_id, inputs_id = await _seed_model(session)
    set_client_auth(client, user_id)

    resp = await client.post(
        f"/ui/models/{model_id}/noi-inputs",
        data={
            # Display formatting ($ + commas) must be stripped server-side.
            "noi_stabilized_input": "$650,000",
            "noi_escalation_rate_pct": "2.5",
        },
    )
    assert resp.status_code == 200, resp.text
    assert "NOI inputs saved" in resp.text

    session.expire_all()
    inputs = await session.get(OperationalInputs, inputs_id)
    assert inputs.noi_stabilized_input == Decimal("650000")
    assert inputs.noi_escalation_rate_pct == Decimal("2.5")
    assert inputs.noi_auto_seeded is False


async def test_noi_inputs_blank_clears_value(
    client: AsyncClient, session: AsyncSession
) -> None:
    model_id, _project_id, user_id, inputs_id = await _seed_model(session)
    set_client_auth(client, user_id)

    resp = await client.post(
        f"/ui/models/{model_id}/noi-inputs",
        data={"noi_stabilized_input": "", "noi_escalation_rate_pct": ""},
    )
    assert resp.status_code == 200

    session.expire_all()
    inputs = await session.get(OperationalInputs, inputs_id)
    assert inputs.noi_stabilized_input is None
    # Empty escalation falls back to the 3% default.
    assert inputs.noi_escalation_rate_pct == Decimal("3")


async def test_noi_inputs_400_when_scenario_has_no_project(
    client: AsyncClient, session: AsyncSession
) -> None:
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user)
    bare_model = await seed_deal_model(session, opp, user)
    await session.commit()
    model_id, user_id = bare_model.id, user.id

    set_client_auth(client, user_id)
    resp = await client.post(
        f"/ui/models/{model_id}/noi-inputs",
        data={"noi_stabilized_input": "100000"},
    )
    assert resp.status_code == 400


# ---------------------------------------------------------------------------
# POST /ui/models/{id}/history/{snapshot_id}/revert
# ---------------------------------------------------------------------------


async def test_revert_restores_use_lines_from_snapshot(
    client: AsyncClient, session: AsyncSession
) -> None:
    """End-to-end input restore: snapshot a $500k Hard Costs line, mutate it
    to $999,999, revert via the route, and verify the DB row is back at the
    snapshotted amount (rows are re-created, so match by label)."""
    from app.exporters.snapshot import capture_snapshot

    model_id, project_id, user_id, _inputs_id = await _seed_model(session)

    session.add(UseLine(
        project_id=project_id,
        label="Hard Costs",
        phase="construction",
        amount=Decimal("500000"),
        cost_category="hard",
        timing_type="first_day",
    ))
    await session.flush()

    snap = await capture_snapshot(session, model_id, triggered_by="compute")
    snap_id = snap.id
    await session.commit()

    # Mutate post-snapshot state.
    use_line = (
        await session.execute(
            select(UseLine).where(
                UseLine.project_id == project_id,
                UseLine.label == "Hard Costs",
            )
        )
    ).scalar_one()
    use_line.amount = Decimal("999999")
    await session.commit()

    set_client_auth(client, user_id)
    resp = await client.post(f"/ui/models/{model_id}/history/{snap_id}/revert")
    assert resp.status_code == 303, resp.text
    assert resp.headers["location"] == f"/models/{model_id}/builder"

    session.expire_all()
    restored = (
        await session.execute(
            select(UseLine).where(
                UseLine.project_id == project_id,
                UseLine.label == "Hard Costs",
            )
        )
    ).scalar_one()
    assert restored.amount == Decimal("500000"), (
        "revert must restore the snapshotted use-line amount"
    )


async def test_revert_404_for_unknown_snapshot(
    client: AsyncClient, session: AsyncSession
) -> None:
    model_id, _project_id, user_id, _inputs_id = await _seed_model(session)
    set_client_auth(client, user_id)

    resp = await client.post(
        f"/ui/models/{model_id}/history/{uuid.uuid4()}/revert"
    )
    assert resp.status_code == 404


async def test_revert_403_for_foreign_org(
    client: AsyncClient, session: AsyncSession
) -> None:
    model_id, _project_id, _user_id, _inputs_id = await _seed_model(session)
    snap = ScenarioSnapshot(
        scenario_id=model_id,
        version=1,
        triggered_by="compute",
        inputs_json={},
        outputs_json={},
    )
    session.add(snap)
    # A user from a different org must not be able to revert this model.
    _other_org, other_user = await seed_org(session)
    await session.commit()
    snap_id, other_user_id = snap.id, other_user.id

    set_client_auth(client, other_user_id)
    resp = await client.post(f"/ui/models/{model_id}/history/{snap_id}/revert")
    assert resp.status_code == 403


async def test_revert_unauthenticated_401(
    client: AsyncClient, session: AsyncSession
) -> None:
    model_id, _project_id, _user_id, _inputs_id = await _seed_model(session)

    resp = await client.post(
        f"/ui/models/{model_id}/history/{uuid.uuid4()}/revert",
        headers={"HX-Request": "true"},
    )
    assert resp.status_code == 401


# ---------------------------------------------------------------------------
# POST /ui/models/{id}/save-as-template
# ---------------------------------------------------------------------------


async def test_save_as_template_persists_scenario_template(
    client: AsyncClient, session: AsyncSession
) -> None:
    from app.models.scenario_template import ScenarioTemplate

    model_id, _project_id, user_id, _inputs_id = await _seed_model(session)
    set_client_auth(client, user_id)

    resp = await client.post(
        f"/ui/models/{model_id}/save-as-template",
        data={"name": "Value-Add 8plex", "description": "Base template"},
    )
    assert resp.status_code == 200, resp.text
    assert "Value-Add 8plex" in resp.text

    session.expire_all()
    tmpl = (
        await session.execute(
            select(ScenarioTemplate).where(
                ScenarioTemplate.source_scenario_id == model_id
            )
        )
    ).scalar_one()
    assert tmpl.name == "Value-Add 8plex"
    assert tmpl.description == "Base template"
    assert tmpl.created_by_user_id == user_id
    assert isinstance(tmpl.template_json, dict)


async def test_save_as_template_requires_name(
    client: AsyncClient, session: AsyncSession
) -> None:
    model_id, _project_id, user_id, _inputs_id = await _seed_model(session)
    set_client_auth(client, user_id)

    resp = await client.post(
        f"/ui/models/{model_id}/save-as-template",
        data={"name": "   "},
    )
    assert resp.status_code == 400
