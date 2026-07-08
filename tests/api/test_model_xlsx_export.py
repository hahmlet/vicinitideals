"""Tests for ``GET /api/models/{model_id}/export/xlsx``.

The endpoint now serves the investor workbook (the legacy round-trip
``excel_export`` module was deleted 2026-07 — its promised re-importer was
never built). ``profile`` selects the sheet set; default is ``internal``.
The deprecated UI download route ``/ui/models/{id}/export.xlsx`` is gone.
"""

from __future__ import annotations

from io import BytesIO
from uuid import uuid4

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

pytestmark = pytest.mark.asyncio


async def _seed_model(session: AsyncSession):
    from tests.conftest import seed_deal_model_with_financials, seed_opportunity, seed_org

    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user)
    deal_model, _, _, _ = await seed_deal_model_with_financials(session, opp, user)
    await session.commit()
    return deal_model


async def test_api_xlsx_export_serves_investor_workbook(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    deal_model = await _seed_model(session)

    resp = await client.get(f"/api/models/{deal_model.id}/export/xlsx")
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    disposition = resp.headers["content-disposition"]
    assert "attachment; filename=" in disposition
    # Default profile is "internal" → "-underwriting" filename suffix.
    assert "-underwriting.xlsx" in disposition

    wb = load_workbook(BytesIO(resp.content), data_only=True)
    expected = {
        "Cover",
        "Underwriting Summary",
        "Underwriting Pro Forma",
        "Underwriting Cash Flow",
        "Sources & Uses",
        "Investor Returns",
        "Waterfall",
        "Unit Mix",
        "Assumptions",
        "Glossary & Methodology",
    }
    assert expected <= set(wb.sheetnames), set(wb.sheetnames)


async def test_api_xlsx_export_profile_param_gates_sheets(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    deal_model = await _seed_model(session)

    resp = await client.get(f"/api/models/{deal_model.id}/export/xlsx?profile=lender")
    assert resp.status_code == 200, resp.text
    assert "-lender.xlsx" in resp.headers["content-disposition"]

    wb = load_workbook(BytesIO(resp.content), data_only=True)
    # Lender package drops the investor-returns sheets, keeps debt.
    assert "Investor Returns" not in wb.sheetnames
    assert "Waterfall" not in wb.sheetnames
    assert "Debt Schedule" in wb.sheetnames


async def test_api_xlsx_export_404_for_missing_model(client: AsyncClient) -> None:
    resp = await client.get(f"/api/models/{uuid4()}/export/xlsx")
    assert resp.status_code == 404


async def test_deprecated_ui_export_route_is_gone(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    from tests.conftest import set_client_auth

    deal_model = await _seed_model(session)
    # Authenticate so the check reaches routing: an unauthenticated request
    # now gets 401 from the auth middleware (HTMX no longer bypasses it),
    # which would vacuously pass a "route is gone" 404 assertion.
    set_client_auth(client, deal_model.created_by_user_id)

    resp = await client.get(
        f"/ui/models/{deal_model.id}/export.xlsx",
        headers={"hx-request": "true"},
    )
    assert resp.status_code == 404
