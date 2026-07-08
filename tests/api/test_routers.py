"""Broad router smoke suite — runs against the shared Postgres conftest.

Migrated off an inline in-memory SQLite engine (2026-07): all fixtures
(`client`, `session`, `auth_headers`) now come from tests/conftest.py, and
seeding goes through the shared seed helpers so the schema matches production.
"""

from __future__ import annotations

import asyncio
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, cast
from unittest.mock import patch
from uuid import UUID, uuid4

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.main import create_app
from app.models.capital import (
    CapitalModule,
    WaterfallResult,
    WaterfallTier,
    WaterfallTierType,
)
from app.models.cashflow import CashFlow, OperationalOutputs, PeriodType
from app.models.deal import (
    Deal,
    IncomeStream,
    IncomeStreamType,
    OperatingExpenseLine,
    OperationalInputs,
    ProjectType,
    Scenario,
)
from app.models.ingestion import DedupCandidate, DedupStatus, IngestJob, RecordType
from app.models.manifest import WorkflowRunManifest
from app.models.opportunity import Opportunity
from app.models.org import Organization, User
from app.models.portfolio import Portfolio, PortfolioProject
from app.models.project import Project
from app.models.scenario import Sensitivity, SensitivityResult, SensitivityStatus
from app.tasks.scenario import run_scenario
from tests.conftest import seed_opportunity, seed_org


def _as_user(auth_headers: dict[str, str], user: User) -> dict[str, str]:
    """Auth headers with X-User-ID pointing at a real seeded user.

    Several routes are org-scoped (they look up the current user and compare
    org_id) or persist current_user_id into FK columns — a random UUID header
    404s or violates the FK on Postgres.
    """
    return {**auth_headers, "X-User-ID": str(user.id)}


async def _seed_stack(
    session: AsyncSession,
    *,
    opp_name: str,
    deal_name: str,
    project_type: ProjectType = ProjectType.acquisition,
    version: int = 1,
    is_active: bool = True,
) -> tuple[Organization, User, Opportunity, Scenario, Project]:
    """Seed the canonical lineage: Org → Opportunity + Deal → Scenario → Project."""
    org, user = await seed_org(session)
    opportunity = await seed_opportunity(session, org, user, name=opp_name)

    top_deal = Deal(org_id=org.id, name=deal_name, created_by_user_id=user.id)
    session.add(top_deal)
    await session.flush()

    model = Scenario(
        deal_id=top_deal.id,
        created_by_user_id=user.id,
        name=deal_name,
        project_type=project_type,
        version=version,
        is_active=is_active,
    )
    session.add(model)
    await session.flush()

    dev_project = Project(
        scenario_id=model.id,
        opportunity_id=opportunity.id,
        name="Default Project",
    )
    session.add(dev_project)
    await session.flush()
    return org, user, opportunity, model, dev_project


async def _seed_model_for_run_tests(
    session: AsyncSession,
    *,
    include_capital: bool = False,
) -> tuple[UUID, User]:
    _, user, _, model, dev_project = await _seed_stack(
        session,
        opp_name="Workflow Run Project",
        deal_name="Workflow Run Deal",
        project_type=ProjectType.acquisition,
    )

    session.add(
        OperationalInputs(
            project_id=dev_project.id,
            purchase_price=Decimal("800000"),
            closing_costs_pct=Decimal("2.5"),
            renovation_cost_total=Decimal("120000"),
            renovation_months=2,
            lease_up_months=2,
            unit_count_existing=8,
            expense_growth_rate_pct_annual=Decimal("3.0"),
            exit_cap_rate_pct=Decimal("5.5"),
        )
    )
    session.add(
        IncomeStream(
            project_id=dev_project.id,
            stream_type=IncomeStreamType.residential_rent,
            label="Unit Rent",
            unit_count=8,
            amount_per_unit_monthly=Decimal("1550"),
            stabilized_occupancy_pct=Decimal("95"),
            escalation_rate_pct_annual=Decimal("2.5"),
            active_in_phases=["lease_up", "stabilized", "exit"],
        )
    )
    session.add(
        OperatingExpenseLine(
            project_id=dev_project.id,
            label="Utilities",
            annual_amount=Decimal("1800"),
            escalation_rate_pct_annual=Decimal("3.0"),
            active_in_phases=["lease_up", "stabilized", "exit"],
        )
    )

    if include_capital:
        capital_module = CapitalModule(
            scenario_id=model.id,
            label="Sponsor Equity",
            vehicle_type="equity",
            stack_position=1,
            source={"amount": 2000000},
            carry={"carry_type": "none", "payment_frequency": "at_exit"},
            exit_terms={"exit_type": "profit_share", "trigger": "sale", "profit_share_pct": 100},
            active_phase_start="acquisition",
            active_phase_end="exit",
        )
        session.add(capital_module)
        await session.flush()
        session.add(
            WaterfallTier(
                scenario_id=model.id,
                capital_module_id=capital_module.id,
                priority=1,
                tier_type=WaterfallTierType.return_of_equity,
                lp_split_pct=Decimal("100"),
                gp_split_pct=Decimal("0"),
                description="Return sponsor equity",
            )
        )

    await session.flush()
    return model.id, user


async def test_health_returns_structured_ok_payload(
    client: AsyncClient,
    auth_headers: dict[str, str],
) -> None:
    response = await client.get("/health", headers=auth_headers)

    assert response.status_code == 200
    body = response.json()
    assert body["code"] == "ok"
    assert body["message"] == "re-modeling API is healthy"
    assert body["detail"]["status"] == "ok"
    assert body["detail"]["version"]
    assert body["detail"]["started_at"] > 0
    assert response.headers["X-Trace-ID"]
    assert int(response.headers["X-Process-Time-Ms"]) >= 0


async def test_unhandled_exception_returns_structured_500_payload(
    auth_headers: dict[str, str],
) -> None:
    app = create_app()

    # /api/* is exempt from the UI login-redirect middleware, so the request
    # reaches the route handler and exercises the 500 envelope.
    @app.get("/api/boom")
    async def boom() -> None:
        raise RuntimeError("boom")

    transport = ASGITransport(app=app, raise_app_exceptions=False)
    async with AsyncClient(transport=transport, base_url="http://testserver") as local_client:
        response = await local_client.get("/api/boom", headers=auth_headers)

    assert response.status_code == 500
    assert response.json() == {
        "code": "internal_server_error",
        "message": "An unexpected server error occurred.",
        "detail": None,
    }


async def test_missing_api_key_returns_403(client: AsyncClient) -> None:
    # Use a non-UI path so the API key middleware actually fires (all /api/* and
    # /health are exempt). A session cookie satisfies the UI login-redirect
    # middleware (hx-request no longer bypasses it — 2026-07-08 fix) so the
    # request reaches the API-key gate; hx-request keeps the onboarding guard
    # (which opens a prod-engine DB session) out of the path.
    from app.api.auth import COOKIE_NAME, create_session_token

    client.cookies.set(COOKIE_NAME, create_session_token(uuid4()))
    response = await client.get(
        "/projects",
        headers={"X-User-ID": str(uuid4()), "hx-request": "true"},
    )

    assert response.status_code == 403
    assert response.json() == {
        "code": "forbidden",
        "message": "Invalid API key",
        "detail": None,
    }


async def test_mock_billing_ui_path_bypasses_api_key_middleware(client: AsyncClient) -> None:
    response = await client.get("/mock/billing/embedded")

    assert response.status_code == 303
    assert response.headers.get("location") == "/login?next=/mock/billing/embedded"


async def test_not_found_http_exception_returns_structured_404_payload(
    client: AsyncClient,
    auth_headers: dict[str, str],
) -> None:
    response = await client.get(f"/api/projects/{uuid4()}", headers=auth_headers)

    assert response.status_code == 404
    assert response.json() == {
        "code": "not_found",
        "message": "Project not found",
        "detail": None,
    }


async def test_post_projects_missing_fields_returns_structured_422_payload(
    client: AsyncClient,
    auth_headers: dict[str, str],
) -> None:
    response = await client.post("/api/projects", json={"status": "active"}, headers=auth_headers)

    assert response.status_code == 422
    body = response.json()
    assert body["code"] == "validation_error"
    assert body["message"] == "Request validation failed"
    # name became optional on ProjectCreate — only org_id is required now
    assert any(error["field"] == "org_id" and "Field required" in error["message"] for error in body["detail"])


async def test_get_projects_returns_empty_list(
    client: AsyncClient,
    auth_headers: dict[str, str],
) -> None:
    response = await client.get("/api/projects", headers=auth_headers)

    assert response.status_code == 200
    assert response.json() == []


async def test_post_projects_creates_project(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    org, _user = await seed_org(session)
    await session.commit()

    payload = {
        "org_id": str(org.id),
        "name": "12-Unit Townhome Deal",
        "status": "active",
        "project_category": "proposed",
        "source": "manual",
    }

    response = await client.post("/api/projects", json=payload, headers=auth_headers)

    assert response.status_code == 201
    data = response.json()
    assert data["name"] == payload["name"]
    assert data["org_id"] == str(org.id)
    assert data["status"] == "active"


async def test_get_project_summary_returns_active_model_rollup(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    _, _, opportunity, model, _dev_project = await _seed_stack(
        session,
        opp_name="Project Summary Deal",
        deal_name="Summary Base Case",
        project_type=ProjectType.acquisition,
    )

    session.add(
        CapitalModule(
            scenario_id=model.id,
            label="Senior Loan",
            vehicle_type="debt",
            stack_position=1,
            source={"amount": 500000, "interest_rate_pct": 5.0},
            carry={"carry_type": "pi", "payment_frequency": "monthly"},
            exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
            active_phase_start="acquisition",
            active_phase_end="exit",
        )
    )
    await session.commit()
    project_id = opportunity.id
    model_id = model.id

    inputs_response = await client.put(
        f"/api/models/{model_id}/inputs",
        json={
            "unit_count_new": 12,
            "purchase_price": "800000",
            "renovation_cost_total": "120000",
            "renovation_months": 3,
            "lease_up_months": 2,
            "expense_growth_rate_pct_annual": "3.0",
            "exit_cap_rate_pct": "5.5",
        },
        headers=auth_headers,
    )
    assert inputs_response.status_code == 200

    income_response = await client.post(
        f"/api/models/{model_id}/income-streams",
        json={
            "stream_type": "residential_rent",
            "label": "Unit Rent",
            "unit_count": 12,
            "amount_per_unit_monthly": "1450",
            "stabilized_occupancy_pct": "95",
            "escalation_rate_pct_annual": "2.5",
            "active_in_phases": ["lease_up", "stabilized", "exit"],
        },
        headers=auth_headers,
    )
    assert income_response.status_code == 201

    expense_response = await client.post(
        f"/api/models/{model_id}/expense-lines",
        json={
            "label": "Utilities",
            "annual_amount": "2400",
            "escalation_rate_pct_annual": "3.0",
            "active_in_phases": ["lease_up", "stabilized", "exit"],
        },
        headers=auth_headers,
    )
    assert expense_response.status_code == 201

    compute_response = await client.post(f"/api/models/{model_id}/compute", headers=auth_headers)
    assert compute_response.status_code == 200
    compute_payload = compute_response.json()
    assert compute_payload["observability"]["run_type"] == "cashflow"
    assert compute_payload["observability"]["trace_id"] == compute_response.headers["X-Trace-ID"]

    summary_response = await client.get(f"/api/projects/{project_id}/summary", headers=auth_headers)
    assert summary_response.status_code == 200

    summary = summary_response.json()
    assert summary["id"] == str(project_id)
    assert summary["name"] == "Project Summary Deal"
    assert summary["active_deal_model_id"] == str(model_id)
    assert summary["income_stream_count"] == 1
    assert summary["expense_line_count"] == 1
    # 2 = seeded Senior Loan + the equity module the compute auto-creates to
    # balance Sources = Uses when no equity is configured
    assert summary["capital_module_count"] == 2
    assert summary["outputs"] is not None
    assert Decimal(str(summary["outputs"]["total_project_cost"])) > Decimal("0")
    assert int(summary["outputs"]["total_timeline_months"]) > 0
    assert Decimal(str(summary["outputs"]["project_irr_unlevered"])) != Decimal("0")


async def test_get_project_summary_returns_null_outputs_when_project_has_no_model(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    org, user = await seed_org(session)
    opportunity = await seed_opportunity(session, org, user, name="No Model Project")
    await session.commit()

    response = await client.get(f"/api/projects/{opportunity.id}/summary", headers=auth_headers)

    assert response.status_code == 200
    summary = response.json()
    assert summary["id"] == str(opportunity.id)
    assert summary["active_deal_model_id"] is None
    assert summary["income_stream_count"] == 0
    assert summary["expense_line_count"] == 0
    assert summary["capital_module_count"] == 0
    assert summary["outputs"] is None


async def test_get_model_runs_returns_cashflow_manifest_after_compute(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    model_id, user = await _seed_model_for_run_tests(session)
    await session.commit()
    headers = _as_user(auth_headers, user)

    compute_response = await client.post(f"/api/models/{model_id}/compute", headers=headers)

    assert compute_response.status_code == 200

    manifests = list(
        (
            await session.execute(
                select(WorkflowRunManifest)
                .where(WorkflowRunManifest.scenario_id == model_id)
                .order_by(WorkflowRunManifest.created_at.desc())
            )
        ).scalars()
    )

    assert len(manifests) == 1
    assert manifests[0].engine == "cashflow"
    assert manifests[0].run_id
    assert manifests[0].inputs_json == {
        "model_id": str(model_id),
        "project_type": "acquisition",
        "unit_count": 8,
        "income_stream_count": 1,
    }

    runs_response = await client.get(f"/api/models/{model_id}/runs", headers=headers)

    assert runs_response.status_code == 200
    runs = runs_response.json()
    assert len(runs) == 1
    assert runs[0]["engine"] == "cashflow"
    assert runs[0]["scenario_id"] == str(model_id)
    assert runs[0]["outputs_json"]["cash_flow_count"] == compute_response.json()["cash_flow_count"]


async def test_replay_model_run_recomputes_waterfall_and_persists_new_manifest(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    model_id, user = await _seed_model_for_run_tests(session, include_capital=True)
    await session.commit()
    headers = _as_user(auth_headers, user)

    cashflow_response = await client.post(f"/api/models/{model_id}/compute", headers=headers)
    assert cashflow_response.status_code == 200

    waterfall_response = await client.post(f"/api/models/{model_id}/waterfall/compute", headers=headers)
    assert waterfall_response.status_code == 200
    waterfall_payload = waterfall_response.json()
    assert waterfall_payload["waterfall_result_count"] > 0
    assert waterfall_payload["observability"]["run_type"] == "waterfall"
    assert waterfall_payload["observability"]["trace_id"] == waterfall_response.headers["X-Trace-ID"]

    runs_response = await client.get(f"/api/models/{model_id}/runs", headers=headers)
    assert runs_response.status_code == 200
    runs = runs_response.json()
    # With capital modules configured, /compute also runs the waterfall inside
    # its convergence loop, so the exact manifest count varies — assert both
    # engines are represented and use count deltas around the replay instead.
    assert {run["engine"] for run in runs} == {"cashflow", "waterfall"}
    runs_before = len(runs)
    waterfall_runs_before = sum(1 for run in runs if run["engine"] == "waterfall")

    waterfall_run_id = next(run["run_id"] for run in runs if run["engine"] == "waterfall")
    replay_response = await client.post(
        f"/api/models/{model_id}/runs/{waterfall_run_id}/replay",
        headers=headers,
    )

    assert replay_response.status_code == 200
    replay_data = replay_response.json()
    assert replay_data["deal_model_id"] == str(model_id)
    assert replay_data["waterfall_result_count"] == waterfall_response.json()["waterfall_result_count"]

    replay_runs_response = await client.get(f"/api/models/{model_id}/runs", headers=headers)
    assert replay_runs_response.status_code == 200
    replay_runs = replay_runs_response.json()
    assert len(replay_runs) == runs_before + 1
    assert sum(1 for run in replay_runs if run["engine"] == "waterfall") == waterfall_runs_before + 1


@pytest.mark.parametrize(
    ("payload_overrides", "expected_detail"),
    [
        ({"variable": "bad.key"}, "Invalid scenario variable"),
        ({"range_steps": 0}, "range_steps must be at least 1"),
        (
            {"range_min": "10", "range_max": "5"},
            "range_min must be less than or equal to range_max",
        ),
    ],
)
async def test_post_project_scenarios_rejects_invalid_input(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
    payload_overrides: dict[str, str | int],
    expected_detail: str,
) -> None:
    _, _, opportunity, model, _dev_project = await _seed_stack(
        session,
        opp_name="Scenario Guard Project",
        deal_name="Scenario Guard Deal",
        project_type=ProjectType.new_construction,
    )
    await session.commit()

    payload = {
        "scenario_id": str(model.id),
        "variable": "operational.exit_cap_rate_pct",
        "range_min": "4.5",
        "range_max": "6.0",
        "range_steps": 4,
    }
    payload.update(payload_overrides)

    response = await client.post(
        f"/api/projects/{opportunity.id}/scenarios",
        json=payload,
        headers=auth_headers,
    )

    assert response.status_code == 400
    assert response.json()["code"] == "bad_request"
    assert expected_detail in response.json()["message"]
    assert response.json()["detail"] is None


async def test_post_project_scenarios_accepts_valid_range(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    _, user, opportunity, model, _dev_project = await _seed_stack(
        session,
        opp_name="Scenario Valid Project",
        deal_name="Scenario Valid Deal",
        project_type=ProjectType.new_construction,
    )
    await session.commit()

    with patch("app.api.routers.scenarios.sweep_variable.apply_async") as mocked_apply_async:
        mocked_apply_async.return_value.id = "scenario-task-123"

        response = await client.post(
            f"/api/projects/{opportunity.id}/scenarios",
            json={
                "scenario_id": str(model.id),
                "variable": "operational.exit_cap_rate_pct",
                "range_min": "4.5",
                "range_max": "6.0",
                "range_steps": 4,
            },
            headers=_as_user(auth_headers, user),
        )

    assert response.status_code == 201
    payload = response.json()
    assert payload["status"] == "queued"
    assert payload["task_id"] == "scenario-task-123"
    assert payload["trace_id"] == response.headers["X-Trace-ID"]
    assert datetime.fromisoformat(payload["queued_at"])
    mocked_apply_async.assert_called_once_with(
        args=[payload["scenario_id"]],
        kwargs={"trace_id": payload["trace_id"]},
        queue="analysis",
    )

    sensitivities = list((await session.execute(select(Sensitivity))).scalars())
    assert len(sensitivities) == 1
    assert sensitivities[0].variable == "operational.exit_cap_rate_pct"


async def test_get_scenario_status_includes_model_version_snapshot_after_run(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    _, user, opportunity, model, dev_project = await _seed_stack(
        session,
        opp_name="Scenario Status Project",
        deal_name="Scenario Status Deal",
        project_type=ProjectType.new_construction,
        version=3,
    )

    session.add(
        OperationalInputs(
            project_id=dev_project.id,
            unit_count_new=12,
            exit_cap_rate_pct=Decimal("5.500000"),
            lease_up_months=6,
            expense_growth_rate_pct_annual=Decimal("3.000000"),
        )
    )
    await session.commit()

    with patch("app.api.routers.scenarios.sweep_variable.apply_async") as mocked_apply_async:
        mocked_apply_async.return_value.id = "scenario-task-456"
        create_response = await client.post(
            f"/api/projects/{opportunity.id}/scenarios",
            json={
                "scenario_id": str(model.id),
                "variable": "operational.exit_cap_rate_pct",
                "range_min": "4.5",
                "range_max": "6.0",
                "range_steps": 4,
            },
            headers=_as_user(auth_headers, user),
        )

    assert create_response.status_code == 201
    scenario_id = create_response.json()["scenario_id"]
    # Capture before expire_all — expired attribute access raises MissingGreenlet
    user_headers = _as_user(auth_headers, user)
    model_id = model.id

    async def fake_compute_cash_flows(*args, **kwargs):  # type: ignore[no-untyped-def]
        return {
            "project_irr_levered": Decimal("14.500000"),
            "total_project_cost": Decimal("1500000.000000"),
            "equity_required": Decimal("500000.000000"),
            "noi_stabilized": Decimal("60000.000000"),
        }

    async def fake_compute_waterfall(*args, **kwargs):  # type: ignore[no-untyped-def]
        return {
            "lp_irr_pct": Decimal("12.250000"),
            "gp_irr_pct": Decimal("16.750000"),
        }

    # app.db.AsyncSessionLocal is already rebound to the test DB by the
    # session-scoped _rebind_app_db fixture, so the task hits the test DB.
    with (
        patch("app.tasks.scenario.compute_cash_flows", fake_compute_cash_flows),
        patch("app.tasks.scenario.compute_waterfall", fake_compute_waterfall),
    ):
        await asyncio.to_thread(cast(Any, run_scenario), scenario_id)

    # The task committed via its own connections — drop stale identity-map state.
    session.expire_all()

    status_response = await client.get(
        f"/api/scenarios/{scenario_id}/status", headers=user_headers
    )

    assert status_response.status_code == 200
    snapshot = status_response.json()["model_version_snapshot"]
    assert snapshot is not None
    assert snapshot["deal_model_id"] == str(model_id)
    assert snapshot["deal_model_version"] == 3
    assert snapshot["project_type"] == ProjectType.new_construction.value
    assert snapshot["unit_count_new"] == 12
    assert snapshot["exit_cap_rate_pct"] == "5.500000"
    assert isinstance(snapshot["captured_at"], str)
    assert datetime.fromisoformat(snapshot["captured_at"])
    assert status_response.json()["run_count"] == 1


async def test_get_scenario_results_defaults_to_latest_run_and_supports_run_query(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    _, user, opportunity, model, _dev_project = await _seed_stack(
        session,
        opp_name="Scenario Results Project",
        deal_name="Scenario Results Deal",
        project_type=ProjectType.new_construction,
    )

    sensitivity = Sensitivity(
        opportunity_id=opportunity.id,
        scenario_id=model.id,
        created_by_user_id=user.id,
        variable="operational.exit_cap_rate_pct",
        range_min=Decimal("4.500000"),
        range_max=Decimal("6.000000"),
        range_steps=4,
        run_count=2,
        status=SensitivityStatus.complete,
    )
    session.add(sensitivity)
    await session.flush()

    session.add_all(
        [
            SensitivityResult(
                sensitivity_id=sensitivity.id,
                run_number=1,
                variable_value=Decimal("4.500000"),
                project_irr_pct=Decimal("15.000000"),
            ),
            SensitivityResult(
                sensitivity_id=sensitivity.id,
                run_number=1,
                variable_value=Decimal("5.000000"),
                project_irr_pct=Decimal("14.750000"),
            ),
            SensitivityResult(
                sensitivity_id=sensitivity.id,
                run_number=2,
                variable_value=Decimal("5.500000"),
                project_irr_pct=Decimal("14.500000"),
            ),
            SensitivityResult(
                sensitivity_id=sensitivity.id,
                run_number=2,
                variable_value=Decimal("6.000000"),
                project_irr_pct=Decimal("14.250000"),
            ),
        ]
    )
    await session.commit()
    scenario_id = sensitivity.id

    latest_response = await client.get(
        f"/api/scenarios/{scenario_id}/results", headers=_as_user(auth_headers, user)
    )
    assert latest_response.status_code == 200
    latest_payload = latest_response.json()
    assert [item["run_number"] for item in latest_payload] == [2, 2]
    assert [Decimal(item["variable_value"]) for item in latest_payload] == [
        Decimal("5.500000"),
        Decimal("6.000000"),
    ]

    historical_response = await client.get(
        f"/api/scenarios/{scenario_id}/results?run=1",
        headers=_as_user(auth_headers, user),
    )
    assert historical_response.status_code == 200
    historical_payload = historical_response.json()
    assert [item["run_number"] for item in historical_payload] == [1, 1]
    assert [Decimal(item["variable_value"]) for item in historical_payload] == [
        Decimal("4.500000"),
        Decimal("5.000000"),
    ]


async def test_get_scenario_compare_returns_variance_and_attribution_contract(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    _, user, opportunity, model, dev_project = await _seed_stack(
        session,
        opp_name="Scenario Compare Project",
        deal_name="Scenario Compare Deal",
        project_type=ProjectType.new_construction,
    )

    session.add(
        OperationalInputs(
            project_id=dev_project.id,
            unit_count_new=12,
            exit_cap_rate_pct=Decimal("5.500000"),
            expense_growth_rate_pct_annual=Decimal("3.000000"),
        )
    )

    sensitivity = Sensitivity(
        opportunity_id=opportunity.id,
        scenario_id=model.id,
        created_by_user_id=user.id,
        variable="operational.exit_cap_rate_pct",
        range_min=Decimal("4.500000"),
        range_max=Decimal("6.000000"),
        range_steps=3,
        status=SensitivityStatus.complete,
    )
    session.add(sensitivity)
    await session.flush()

    session.add_all(
        [
            SensitivityResult(
                sensitivity_id=sensitivity.id,
                variable_value=Decimal("4.500000"),
                project_irr_pct=Decimal("16.000000"),
                lp_irr_pct=Decimal("12.500000"),
                gp_irr_pct=Decimal("17.500000"),
                equity_multiple=Decimal("3.200000"),
                cash_on_cash_year1_pct=Decimal("11.500000"),
            ),
            SensitivityResult(
                sensitivity_id=sensitivity.id,
                variable_value=Decimal("5.500000"),
                project_irr_pct=Decimal("15.000000"),
                lp_irr_pct=Decimal("12.000000"),
                gp_irr_pct=Decimal("17.000000"),
                equity_multiple=Decimal("3.000000"),
                cash_on_cash_year1_pct=Decimal("11.000000"),
            ),
            SensitivityResult(
                sensitivity_id=sensitivity.id,
                variable_value=Decimal("6.000000"),
                project_irr_pct=Decimal("14.500000"),
                lp_irr_pct=Decimal("11.750000"),
                gp_irr_pct=Decimal("16.500000"),
                equity_multiple=Decimal("2.900000"),
                cash_on_cash_year1_pct=Decimal("10.750000"),
            ),
        ]
    )
    await session.commit()
    scenario_id = sensitivity.id

    response = await client.get(f"/api/scenarios/{scenario_id}/compare", headers=auth_headers)

    assert response.status_code == 200
    payload = response.json()
    assert payload["sensitivity_id"] == str(scenario_id)
    assert payload["variable"]["key"] == "operational.exit_cap_rate_pct"
    assert payload["variable"]["label"] == "Exit Cap Rate"
    assert payload["variable"]["unit"] == "pct"
    assert payload["baseline"]["variable_value"] == "5.500000"

    comparison_by_value = {
        item["variable_value"]: item for item in payload["comparisons"]
    }
    higher_cap = comparison_by_value["6.000000"]
    assert higher_cap["attribution"]["direction"] == "increase"
    assert higher_cap["attribution"]["delta"] == "0.500000"
    assert higher_cap["project_irr_pct"]["value"] == "14.500000"
    assert higher_cap["project_irr_pct"]["delta"] == "-0.500000"
    assert higher_cap["project_irr_pct"]["delta_pct"] == "-3.333333"

    baseline = comparison_by_value["5.500000"]
    assert baseline["attribution"]["direction"] == "no_change"
    assert baseline["equity_multiple"]["delta"] == "0.000000"


async def test_list_portfolios_supports_pagination_counts_and_total_header(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    org, user = await seed_org(session)

    alpha = Portfolio(org_id=org.id, name="Alpha Portfolio")
    bravo = Portfolio(org_id=org.id, name="Bravo Portfolio")
    charlie = Portfolio(org_id=org.id, name="Charlie Portfolio")
    session.add_all([alpha, bravo, charlie])
    await session.flush()

    project_one = await seed_opportunity(session, org, user, name="Alpha One")
    project_two = await seed_opportunity(session, org, user, name="Alpha Two")
    project_three = await seed_opportunity(session, org, user, name="Bravo One")

    session.add_all(
        [
            PortfolioProject(portfolio_id=alpha.id, project_id=project_one.id),
            PortfolioProject(portfolio_id=alpha.id, project_id=project_two.id),
            PortfolioProject(portfolio_id=bravo.id, project_id=project_three.id),
        ]
    )
    await session.commit()
    org_id = org.id

    first_page = await client.get(
        f"/api/portfolios?org_id={org_id}&limit=2&offset=0",
        headers=auth_headers,
    )

    assert first_page.status_code == 200
    assert first_page.headers["X-Total-Count"] == "3"
    first_items = first_page.json()
    assert len(first_items) == 2

    second_page = await client.get(
        f"/api/portfolios?org_id={org_id}&limit=2&offset=2",
        headers=auth_headers,
    )

    assert second_page.status_code == 200
    assert second_page.headers["X-Total-Count"] == "3"
    second_items = second_page.json()
    assert len(second_items) == 1

    all_items = {item["name"]: item for item in [*first_items, *second_items]}
    assert set(all_items) == {"Alpha Portfolio", "Bravo Portfolio", "Charlie Portfolio"}
    assert all_items["Alpha Portfolio"]["project_count"] == 2
    assert all_items["Bravo Portfolio"]["project_count"] == 1
    assert all_items["Charlie Portfolio"]["project_count"] == 0


async def test_compute_portfolio_gantt_and_summary_include_project_rollups(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    def make_cash_flow(
        deal_model_id, period: int, period_type: PeriodType, noi: str, net_cash_flow: str
    ) -> CashFlow:
        noi_value = Decimal(noi)
        net_value = Decimal(net_cash_flow)
        return CashFlow(
            scenario_id=deal_model_id,
            period=period,
            period_type=period_type,
            gross_revenue=Decimal("10000.000000"),
            vacancy_loss=Decimal("500.000000"),
            effective_gross_income=Decimal("9500.000000"),
            operating_expenses=Decimal("3000.000000"),
            capex_reserve=Decimal("250.000000"),
            noi=noi_value,
            debt_service=Decimal("0.000000"),
            net_cash_flow=net_value,
            cumulative_cash_flow=net_value,
        )

    org, user = await seed_org(session)

    portfolio = Portfolio(org_id=org.id, name="Portfolio Rollup")
    session.add(portfolio)
    project_a = await seed_opportunity(session, org, user, name="Burnside Apartments")
    project_b = await seed_opportunity(session, org, user, name="Division Cottages")
    await session.flush()

    top_deal_a = Deal(org_id=org.id, name="Burnside Base Case", created_by_user_id=user.id)
    top_deal_b = Deal(org_id=org.id, name="Division Base Case", created_by_user_id=user.id)
    session.add_all([top_deal_a, top_deal_b])
    await session.flush()

    model_a = Scenario(
        deal_id=top_deal_a.id,
        created_by_user_id=user.id,
        name="Burnside Base Case",
        project_type=ProjectType.new_construction,
        is_active=True,
    )
    model_b = Scenario(
        deal_id=top_deal_b.id,
        created_by_user_id=user.id,
        name="Division Base Case",
        project_type=ProjectType.value_add,
        is_active=True,
    )
    session.add_all([model_a, model_b])
    await session.flush()

    dev_project_a = Project(
        scenario_id=model_a.id,
        opportunity_id=project_a.id,
        name="Default Project",
    )
    dev_project_b = Project(
        scenario_id=model_b.id,
        opportunity_id=project_b.id,
        name="Default Project",
    )
    session.add_all([dev_project_a, dev_project_b])
    await session.flush()

    session.add_all(
        [
            OperationalInputs(
                project_id=dev_project_a.id,
                unit_count_new=24,
                milestone_dates={
                    "pre_construction_start": "2026-02-01",
                    "construction_start": "2026-04-01",
                    "lease_up_start": "2026-06-01",
                    "stabilized_start": "2026-08-01",
                    "exit_date": "2026-10-01",
                },
            ),
            OperationalInputs(
                project_id=dev_project_b.id,
                unit_count_new=12,
            ),
            PortfolioProject(
                portfolio_id=portfolio.id,
                project_id=project_a.id,
                scenario_id=model_a.id,
                start_date=date(2026, 1, 1),
                capital_contribution=Decimal("500000.000000"),
            ),
            PortfolioProject(
                portfolio_id=portfolio.id,
                project_id=project_b.id,
                scenario_id=model_b.id,
                start_date=date(2026, 3, 1),
                capital_contribution=Decimal("350000.000000"),
            ),
        ]
    )

    session.add_all(
        [
            make_cash_flow(model_a.id, 0, PeriodType.acquisition, "0.000000", "-100000.000000"),
            make_cash_flow(model_a.id, 1, PeriodType.pre_construction, "0.000000", "-15000.000000"),
            make_cash_flow(model_a.id, 2, PeriodType.pre_construction, "0.000000", "-15000.000000"),
            make_cash_flow(model_a.id, 3, PeriodType.construction, "0.000000", "-50000.000000"),
            make_cash_flow(model_a.id, 4, PeriodType.construction, "0.000000", "-50000.000000"),
            make_cash_flow(model_a.id, 5, PeriodType.lease_up, "12000.000000", "6000.000000"),
            make_cash_flow(model_a.id, 6, PeriodType.stabilized, "18000.000000", "9000.000000"),
            make_cash_flow(model_a.id, 7, PeriodType.stabilized, "18000.000000", "9000.000000"),
            make_cash_flow(model_a.id, 8, PeriodType.exit, "18000.000000", "250000.000000"),
            make_cash_flow(model_b.id, 0, PeriodType.acquisition, "0.000000", "-75000.000000"),
            make_cash_flow(model_b.id, 1, PeriodType.hold, "6000.000000", "2500.000000"),
            make_cash_flow(model_b.id, 2, PeriodType.major_renovation, "0.000000", "-20000.000000"),
            make_cash_flow(model_b.id, 3, PeriodType.major_renovation, "0.000000", "-20000.000000"),
            make_cash_flow(model_b.id, 4, PeriodType.lease_up, "8000.000000", "3200.000000"),
            make_cash_flow(model_b.id, 5, PeriodType.stabilized, "11000.000000", "4800.000000"),
            make_cash_flow(model_b.id, 6, PeriodType.exit, "11000.000000", "125000.000000"),
            OperationalOutputs(
                scenario_id=model_a.id,
                total_project_cost=Decimal("1500000.000000"),
                equity_required=Decimal("450000.000000"),
                noi_stabilized=Decimal("216000.000000"),
                project_irr_levered=Decimal("18.250000"),
            ),
            OperationalOutputs(
                scenario_id=model_b.id,
                total_project_cost=Decimal("900000.000000"),
                equity_required=Decimal("275000.000000"),
                noi_stabilized=Decimal("132000.000000"),
                project_irr_levered=Decimal("14.500000"),
            ),
        ]
    )
    await session.commit()
    portfolio_id = portfolio.id
    project_a_id = project_a.id
    project_b_id = project_b.id

    compute_response = await client.post(
        f"/api/portfolios/{portfolio_id}/gantt/compute",
        headers=auth_headers,
    )

    assert compute_response.status_code == 200
    computed_entries = compute_response.json()
    assert len(computed_entries) == 12

    burnside_entries = [
        entry for entry in computed_entries if entry["project_id"] == str(project_a_id)
    ]
    division_entries = [
        entry for entry in computed_entries if entry["project_id"] == str(project_b_id)
    ]
    assert [entry["phase"] for entry in burnside_entries] == [
        "acquisition",
        "pre_construction",
        "construction",
        "lease_up",
        "stabilized",
        "exit",
    ]
    assert [entry["phase"] for entry in division_entries] == [
        "acquisition",
        "hold",
        "major_renovation",
        "lease_up",
        "stabilized",
        "exit",
    ]

    precon_entry = next(entry for entry in burnside_entries if entry["phase"] == "pre_construction")
    hold_entry = next(entry for entry in division_entries if entry["phase"] == "hold")
    assert precon_entry["start_date"] == "2026-02-01"
    assert precon_entry["end_date"] == "2026-03-31"
    assert hold_entry["start_date"] == "2026-04-01"
    assert hold_entry["end_date"] == "2026-04-30"

    gantt_response = await client.get(f"/api/portfolios/{portfolio_id}/gantt", headers=auth_headers)
    assert gantt_response.status_code == 200
    assert len(gantt_response.json()) == 12

    summary_response = await client.get(f"/api/portfolios/{portfolio_id}/summary", headers=auth_headers)
    assert summary_response.status_code == 200
    summary = summary_response.json()
    assert summary["project_count"] == 2
    assert summary["gantt_entry_count"] == 12
    assert Decimal(str(summary["total_capital_contribution"])) == Decimal("850000.000000")
    assert Decimal(str(summary["total_project_cost"])) == Decimal("2400000.000000")
    assert Decimal(str(summary["total_equity_required"])) == Decimal("725000.000000")
    assert Decimal(str(summary["total_noi_stabilized"])) == Decimal("348000.000000")

    project_rollups = {project["project_id"]: project for project in summary["projects"]}
    assert set(project_rollups) == {str(project_a_id), str(project_b_id)}
    assert project_rollups[str(project_a_id)]["project_name"] == "Burnside Apartments"
    assert Decimal(str(project_rollups[str(project_a_id)]["project_irr_levered"])) == Decimal(
        "18.250000"
    )
    assert Decimal(str(project_rollups[str(project_b_id)]["equity_required"])) == Decimal(
        "275000.000000"
    )


async def test_post_and_get_model_expense_lines_round_trip(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    _, _, _, model, dev_project = await _seed_stack(
        session,
        opp_name="Expense Line Project",
        deal_name="Expense Line Deal",
        project_type=ProjectType.acquisition,
        is_active=False,
    )
    dev_project.unit_mix = [
        {"label": "1BR", "unit_count": 12, "sqft": 700},
    ]
    await session.commit()
    model_id = model.id

    response = await client.post(
        f"/api/models/{model_id}/expense-lines",
        json={
            "label": "Electric",
            "annual_amount": "1200",
            "escalation_rate_pct_annual": "5.0",
            "active_in_phases": ["lease_up", "stabilized", "exit"],
            "notes": "Owner-paid common area electric",
        },
        headers=auth_headers,
    )

    assert response.status_code == 201
    created = response.json()
    assert "project_id" in created
    assert created["label"] == "Electric"
    assert Decimal(str(created["annual_amount"])) == Decimal("1200")
    assert created["per_type"] == "flat"

    listed = await client.get(f"/api/models/{model_id}/expense-lines", headers=auth_headers)

    assert listed.status_code == 200
    rows = listed.json()
    assert len(rows) == 1
    assert rows[0]["label"] == "Electric"
    assert rows[0]["per_type"] == "flat"
    assert rows[0]["active_in_phases"] == ["lease_up", "stabilized", "exit"]
    assert rows[0]["notes"] == "Owner-paid common area electric"

    per_unit_response = await client.post(
        f"/api/models/{model_id}/expense-lines",
        json={
            "label": "Telephone / Internet",
            "per_type": "per_unit",
            "per_value": "75",
            "escalation_rate_pct_annual": "3.0",
        },
        headers=auth_headers,
    )

    assert per_unit_response.status_code == 201
    per_unit_created = per_unit_response.json()
    assert Decimal(str(per_unit_created["annual_amount"])) == Decimal("900")
    assert per_unit_created["active_in_phases"] == ["lease_up", "stabilized"]


async def test_update_and_delete_model_income_stream_and_expense_line(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    _, _, _, model, _dev_project = await _seed_stack(
        session,
        opp_name="Editable Assumptions Project",
        deal_name="Editable Assumptions Deal",
        project_type=ProjectType.acquisition,
        is_active=False,
    )
    await session.commit()
    model_id = model.id

    income_response = await client.post(
        f"/api/models/{model_id}/income-streams",
        json={
            "stream_type": "residential_rent",
            "label": "Studios",
            "unit_count": 8,
            "amount_per_unit_monthly": "1450",
            "stabilized_occupancy_pct": "94",
            "escalation_rate_pct_annual": "2.5",
            "active_in_phases": ["lease_up", "stabilized", "exit"],
            "notes": "Initial lease-up rent",
        },
        headers=auth_headers,
    )
    assert income_response.status_code == 201
    income_stream_id = income_response.json()["id"]

    expense_response = await client.post(
        f"/api/models/{model_id}/expense-lines",
        json={
            "label": "Repairs",
            "annual_amount": "3600",
            "escalation_rate_pct_annual": "3.0",
            "active_in_phases": ["stabilized", "exit"],
            "notes": "Recurring repairs reserve",
        },
        headers=auth_headers,
    )
    assert expense_response.status_code == 201
    expense_line_id = expense_response.json()["id"]

    update_income = await client.patch(
        f"/api/models/{model_id}/income-streams/{income_stream_id}",
        json={
            "label": "Studios Renovated",
            "amount_per_unit_monthly": "1525",
            "notes": "Updated after unit upgrades",
        },
        headers=auth_headers,
    )
    assert update_income.status_code == 200
    assert update_income.json()["label"] == "Studios Renovated"
    assert Decimal(str(update_income.json()["amount_per_unit_monthly"])) == Decimal("1525")
    assert update_income.json()["notes"] == "Updated after unit upgrades"

    update_expense = await client.patch(
        f"/api/models/{model_id}/expense-lines/{expense_line_id}",
        json={
            "annual_amount": "4200",
            "active_in_phases": ["lease_up", "stabilized", "exit"],
            "notes": "Expanded reserve after inspection",
        },
        headers=auth_headers,
    )
    assert update_expense.status_code == 200
    assert Decimal(str(update_expense.json()["annual_amount"])) == Decimal("4200")
    assert update_expense.json()["active_in_phases"] == ["lease_up", "stabilized", "exit"]
    assert update_expense.json()["notes"] == "Expanded reserve after inspection"

    delete_income = await client.delete(
        f"/api/models/{model_id}/income-streams/{income_stream_id}",
        headers=auth_headers,
    )
    assert delete_income.status_code == 204

    delete_expense = await client.delete(
        f"/api/models/{model_id}/expense-lines/{expense_line_id}",
        headers=auth_headers,
    )
    assert delete_expense.status_code == 204

    listed_income = await client.get(f"/api/models/{model_id}/income-streams", headers=auth_headers)
    listed_expenses = await client.get(f"/api/models/{model_id}/expense-lines", headers=auth_headers)

    assert listed_income.status_code == 200
    assert listed_income.json() == []
    assert listed_expenses.status_code == 200
    assert listed_expenses.json() == []


async def test_update_and_delete_capital_modules_and_waterfall_tiers(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    _, user, _, model, _dev_project = await _seed_stack(
        session,
        opp_name="Capital Editor Project",
        deal_name="Capital Editor Deal",
        project_type=ProjectType.new_construction,
        is_active=False,
    )
    await session.commit()
    model_id = model.id
    auth_headers = _as_user(auth_headers, user)  # capital routes are org-scoped

    capital_response = await client.post(
        f"/api/models/{model_id}/capital-modules",
        json={
            "label": "Senior Loan",
            "vehicle_type": "debt",
            "stack_position": 1,
            "source": {"amount": 700000, "interest_rate_pct": 6.1, "notes": "Initial term sheet", "hold_term_years": 10},
            "carry": {"carry_type": "io_only", "payment_frequency": "monthly", "capitalized": False},
            "exit_terms": {"exit_type": "full_payoff", "trigger": "sale", "notes": "Repay at exit"},
            "active_phase_start": "acquisition",
            "active_phase_end": "exit",
        },
        headers=auth_headers,
    )
    assert capital_response.status_code == 201
    capital_module_id = capital_response.json()["id"]

    tier_response = await client.post(
        f"/api/models/{model_id}/waterfall-tiers",
        json={
            "priority": 1,
            "tier_type": "pref_return",
            "irr_hurdle_pct": "8.0",
            "lp_split_pct": "90",
            "gp_split_pct": "10",
            "description": "Preferred return tier",
            "capital_module_id": capital_module_id,
        },
        headers=auth_headers,
    )
    assert tier_response.status_code == 201
    tier_id = tier_response.json()["id"]

    update_capital = await client.patch(
        f"/api/models/{model_id}/capital-modules/{capital_module_id}",
        json={
            "stack_position": 2,
            "source": {"amount": 725000, "interest_rate_pct": 5.9, "notes": "Updated lender quote"},
            "active_phase_end": "stabilized",
        },
        headers=auth_headers,
    )
    assert update_capital.status_code == 200
    assert update_capital.json()["stack_position"] == 2
    assert update_capital.json()["source"]["amount"] == "725000"
    assert update_capital.json()["active_phase_end"] == "stabilized"

    update_tier = await client.patch(
        f"/api/models/{model_id}/waterfall-tiers/{tier_id}",
        json={
            "priority": 2,
            "lp_split_pct": "85",
            "gp_split_pct": "15",
            "description": "Promote after pref",
        },
        headers=auth_headers,
    )
    assert update_tier.status_code == 200
    assert update_tier.json()["priority"] == 2
    assert Decimal(str(update_tier.json()["lp_split_pct"])) == Decimal("85")
    assert update_tier.json()["description"] == "Promote after pref"

    delete_tier = await client.delete(
        f"/api/models/{model_id}/waterfall-tiers/{tier_id}",
        headers=auth_headers,
    )
    assert delete_tier.status_code == 204

    delete_capital = await client.delete(
        f"/api/models/{model_id}/capital-modules/{capital_module_id}",
        headers=auth_headers,
    )
    assert delete_capital.status_code == 204

    listed_capital = await client.get(f"/api/models/{model_id}/capital-modules", headers=auth_headers)
    listed_tiers = await client.get(f"/api/models/{model_id}/waterfall-tiers", headers=auth_headers)

    assert listed_capital.status_code == 200
    assert listed_capital.json() == []
    assert listed_tiers.status_code == 200
    assert listed_tiers.json() == []


async def test_dedup_review_endpoints_list_pending_and_resolve_candidates(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    # resolved_by_user_id is an FK to users — the resolver must be a real user
    _, resolver = await seed_org(session)
    auth_headers = _as_user(auth_headers, resolver)
    resolver_user_id = str(resolver.id)

    ingest_job = IngestJob(source="crexi", triggered_by="pytest", status="completed")
    session.add(ingest_job)
    await session.flush()

    # Scraped listings are Opportunity rows post-0067 (scraped_listings → opportunities)
    listing_a = Opportunity(
        ingest_job_id=ingest_job.id,
        source="crexi",
        source_id="dedup-a",
        source_url="https://example.com/listings/dedup-a",
        address_normalized="123 MAIN ST GRESHAM OR 97030",
        address_raw="123 Main St, Gresham, OR 97030",
        is_new=True,
        matches_saved_criteria=False,
    )
    listing_b = Opportunity(
        ingest_job_id=ingest_job.id,
        source="crexi",
        source_id="dedup-b",
        source_url="https://example.com/listings/dedup-b",
        address_normalized="123 MAIN ST GRESHAM OR",
        address_raw="123 Main St, Gresham, OR",
        is_new=True,
        matches_saved_criteria=False,
    )
    listing_c = Opportunity(
        ingest_job_id=ingest_job.id,
        source="crexi",
        source_id="dedup-c",
        source_url="https://example.com/listings/dedup-c",
        address_normalized="500 OAK ST GRESHAM OR 97030",
        address_raw="500 Oak St, Gresham, OR 97030",
        is_new=True,
        matches_saved_criteria=False,
    )
    listing_d = Opportunity(
        ingest_job_id=ingest_job.id,
        source="crexi",
        source_id="dedup-d",
        source_url="https://example.com/listings/dedup-d",
        address_normalized="500 OAK ST GRESHAM OR",
        address_raw="500 Oak St, Gresham, OR",
        is_new=True,
        matches_saved_criteria=False,
    )
    session.add_all([listing_a, listing_b, listing_c, listing_d])
    await session.flush()

    merge_candidate = DedupCandidate(
        ingest_job_id=ingest_job.id,
        record_a_type=RecordType.listing,
        record_a_id=listing_a.id,
        record_b_type=RecordType.listing,
        record_b_id=listing_b.id,
        confidence_score=0.76,
        match_signals={"address_fuzzy": 0.76},
        status=DedupStatus.pending,
    )
    swap_candidate = DedupCandidate(
        ingest_job_id=ingest_job.id,
        record_a_type=RecordType.listing,
        record_a_id=listing_c.id,
        record_b_type=RecordType.listing,
        record_b_id=listing_d.id,
        confidence_score=0.78,
        match_signals={"address_fuzzy": 0.78},
        status=DedupStatus.pending,
    )
    keep_candidate = DedupCandidate(
        ingest_job_id=ingest_job.id,
        record_a_type=RecordType.listing,
        record_a_id=listing_a.id,
        record_b_type=RecordType.listing,
        record_b_id=listing_d.id,
        confidence_score=0.62,
        match_signals={"address_fuzzy": 0.62},
        status=DedupStatus.pending,
    )
    session.add_all([merge_candidate, swap_candidate, keep_candidate])
    await session.commit()

    pending_response = await client.get("/api/dedup/pending", headers=auth_headers)
    assert pending_response.status_code == 200
    assert {row["id"] for row in pending_response.json()} == {
        str(merge_candidate.id),
        str(swap_candidate.id),
        str(keep_candidate.id),
    }

    keep_response = await client.patch(
        f"/api/dedup/{keep_candidate.id}/keep-separate",
        headers=auth_headers,
    )
    assert keep_response.status_code == 200
    assert keep_response.json()["status"] == "kept_separate"
    assert keep_response.json()["resolved_by_user_id"] == resolver_user_id
    assert keep_response.json()["resolved_at"] is not None

    merge_response = await client.patch(
        f"/api/dedup/{merge_candidate.id}/merge",
        headers=auth_headers,
    )
    assert merge_response.status_code == 200
    assert merge_response.json()["status"] == "merged"

    swap_response = await client.patch(
        f"/api/dedup/{swap_candidate.id}/swap",
        headers=auth_headers,
    )
    assert swap_response.status_code == 200
    assert swap_response.json()["status"] == "swapped"

    # Capture ids before expire_all — expired attribute access on an
    # AsyncSession-bound instance raises MissingGreenlet.
    listing_a_id, listing_b_id, listing_c_id, listing_d_id = (
        listing_a.id,
        listing_b.id,
        listing_c.id,
        listing_d.id,
    )
    session.expire_all()
    merged_listing = await session.get(Opportunity, listing_b_id)
    swapped_listing = await session.get(Opportunity, listing_c_id)

    assert merged_listing is not None
    assert merged_listing.canonical_id == listing_a_id
    assert merged_listing.is_new is False

    assert swapped_listing is not None
    assert swapped_listing.canonical_id == listing_d_id
    assert swapped_listing.is_new is False


async def test_get_model_json_export_returns_portable_payload(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    _, _, opportunity, model, dev_project = await _seed_stack(
        session,
        opp_name="JSON Export Project",
        deal_name="JSON Export Deal",
        project_type=ProjectType.acquisition,
        is_active=False,
    )
    opportunity.apn = "R765432100"
    opportunity.address_normalized = "123 Main St, Gresham, OR 97030"
    opportunity.address_raw = "123 Main St"
    opportunity.lot_sqft = Decimal("7405")
    opportunity.year_built = 1998
    opportunity.gba_sqft = Decimal("2400")
    opportunity.property_type = "Multifamily"

    session.add(
        OperationalInputs(
            project_id=dev_project.id,
            unit_count_existing=8,
            exit_cap_rate_pct=Decimal("5.5"),
        )
    )
    session.add(
        IncomeStream(
            project_id=dev_project.id,
            stream_type=IncomeStreamType.residential_rent,
            label="Unit Rent",
            unit_count=8,
            amount_per_unit_monthly=Decimal("1750"),
            stabilized_occupancy_pct=Decimal("94"),
            active_in_phases=["hold", "stabilized"],
        )
    )
    capital_module = CapitalModule(
        id=uuid4(),
        scenario_id=model.id,
        label="Common Equity",
        vehicle_type="equity",
        stack_position=2,
        source={"amount": 325000},
        carry={"carry_type": "none", "payment_frequency": "at_exit"},
        exit_terms={"exit_type": "profit_share", "trigger": "sale", "profit_share_pct": 70},
        active_phase_start="acquisition",
        active_phase_end="exit",
    )
    session.add(capital_module)
    await session.flush()

    session.add(
        WaterfallTier(
            scenario_id=model.id,
            capital_module_id=capital_module.id,
            priority=1,
            tier_type=WaterfallTierType.residual,
            lp_split_pct=Decimal("70"),
            gp_split_pct=Decimal("30"),
            description="Residual split",
        )
    )
    await session.commit()

    response = await client.get(f"/api/models/{model.id}/export/json", headers=auth_headers)

    assert response.status_code == 200
    data = response.json()
    assert data["schema_version"] == "deal-json-v3"
    assert data["project"]["Name"] == "JSON Export Project"
    assert data["project"]["UnparsedAddress"] == "123 Main St, Gresham, OR 97030"
    assert data["project"]["City"] == "Gresham"
    assert data["project"]["StateOrProvince"] == "OR"
    assert data["project"]["PostalCode"] == "97030"
    assert data["project"]["ParcelNumber"] == "R765432100"
    assert data["project"]["LotSizeSquareFeet"] == 7405
    assert data["project"]["YearBuilt"] == 1998
    assert data["project"]["BuildingAreaTotal"] == 2400
    assert data["project"]["PropertyType"] == "Multifamily"
    assert data["deal_model"]["name"] == "JSON Export Deal"
    assert data["income_streams"][0]["label"] == "Unit Rent"
    assert data["capital_modules"][0]["label"] == "Common Equity"
    assert data["waterfall_tiers"][0]["tier_type"] == "residual"


async def test_get_deals_schema_returns_agent_facing_schema(
    client: AsyncClient,
    auth_headers: dict[str, str],
) -> None:
    response = await client.get("/api/deals/schema", headers=auth_headers)

    assert response.status_code == 200
    data = response.json()
    assert data["type"] == "object"
    assert data["properties"]["project"]["type"] == "object"
    assert "ParcelNumber" in data["properties"]["project"]["properties"]
    assert "deal_model" in data["required"]


async def test_post_model_import_validate_reports_schema_issues(
    client: AsyncClient,
    auth_headers: dict[str, str],
) -> None:
    response = await client.post(
        "/api/models/import/validate",
        json={
            "schema_version": "legacy-v0",
            "deal_model": {"name": "Legacy Payload", "project_type": "acquisition"},
        },
        headers=auth_headers,
    )

    assert response.status_code == 200
    data = response.json()
    assert data["valid"] is False
    assert any("Unsupported schema_version" in error for error in data["errors"])


async def test_post_project_model_import_creates_nested_records(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    org, user = await seed_org(session)
    opportunity = await seed_opportunity(session, org, user, name="JSON Import Project")
    await session.commit()
    # import persists created_by_user_id (FK to users) from the X-User-ID header
    auth_headers = _as_user(auth_headers, user)

    source_module_id = str(uuid4())
    payload = {
        "schema_version": "deal-json-v1",
        "deal_model": {
            "name": "Imported Deal",
            "project_type": "acquisition",
            "version": 2,
            "is_active": True,
        },
        "operational_inputs": {
            "unit_count_existing": 10,
            "exit_cap_rate_pct": "5.25",
        },
        "income_streams": [
            {
                "stream_type": "residential_rent",
                "label": "Renovated Units",
                "unit_count": 10,
                "amount_per_unit_monthly": "2100",
                "stabilized_occupancy_pct": "95",
                "active_in_phases": ["hold", "stabilized"],
            }
        ],
        "capital_modules": [
            {
                "id": source_module_id,
                "label": "Preferred Equity",
                "vehicle_type": "equity",
                "equity_role": "lp",
                "stack_position": 1,
                "source": {"amount": 400000, "interest_rate_pct": 10.0},
                "carry": {"carry_type": "accruing", "payment_frequency": "at_exit"},
                "exit_terms": {"exit_type": "full_payoff", "trigger": "sale"},
                "active_phase_start": "acquisition",
                "active_phase_end": "exit",
            }
        ],
        "waterfall_tiers": [
            {
                "priority": 1,
                "tier_type": "pref_return",
                "capital_module_id": source_module_id,
                "irr_hurdle_pct": "8.0",
                "lp_split_pct": "90",
                "gp_split_pct": "10",
                "description": "Preferred return",
            }
        ],
    }

    response = await client.post(
        f"/api/projects/{opportunity.id}/models/import",
        json=payload,
        headers=auth_headers,
    )

    assert response.status_code == 201
    data = response.json()
    assert data["model"]["name"] == "Imported Deal"
    assert data["counts"]["income_streams"] == 1
    assert data["counts"]["capital_modules"] == 1
    assert data["counts"]["waterfall_tiers"] == 1

    models = list((await session.execute(select(Scenario))).scalars())
    inputs = list((await session.execute(select(OperationalInputs))).scalars())
    streams = list((await session.execute(select(IncomeStream))).scalars())
    modules = list((await session.execute(select(CapitalModule))).scalars())
    tiers = list((await session.execute(select(WaterfallTier))).scalars())

    assert len(models) == 1
    assert len(inputs) == 1
    assert len(streams) == 1
    assert len(modules) == 1
    assert len(tiers) == 1


async def test_post_deals_import_creates_project_from_payload(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    await seed_org(session)
    await session.commit()

    payload = {
        "schema_version": "deal-json-v1",
        "project": {
            "Name": "Agent Imported Project",
            "UnparsedAddress": "456 Oak Ave, Portland, OR 97201",
            "City": "Portland",
            "StateOrProvince": "OR",
            "PostalCode": "97201",
            "ParcelNumber": "RIMPORT123",
            "LotSizeSquareFeet": "8800",
            "YearBuilt": 2005,
            "BuildingAreaTotal": "6200",
            "PropertyType": "Multifamily",
        },
        "deal_model": {
            "name": "Imported Agent Deal",
            "project_type": "acquisition",
            "version": 1,
            "is_active": True,
            "operational_inputs": {
                "unit_count_existing": 9,
                "exit_cap_rate_pct": "5.4",
            },
            "income_streams": [
                {
                    "stream_type": "residential_rent",
                    "label": "Imported Units",
                    "unit_count": 9,
                    "amount_per_unit_monthly": "2050",
                    "stabilized_occupancy_pct": "95",
                    "active_in_phases": ["hold", "stabilized"],
                }
            ],
            "capital_stack": [
                {
                    "label": "Imported Equity",
                    "vehicle_type": "equity",
                    "stack_position": 1,
                    "source": {"amount": 450000},
                    "carry": {"carry_type": "none", "payment_frequency": "at_exit"},
                    "exit_terms": {"exit_type": "profit_share", "trigger": "sale", "profit_share_pct": 100},
                    "active_phase_start": "acquisition",
                    "active_phase_end": "exit",
                }
            ],
            "waterfall_tiers": [
                {
                    "priority": 1,
                    "tier_type": "residual",
                    "lp_split_pct": "90",
                    "gp_split_pct": "10",
                    "description": "Residual split",
                }
            ],
        },
    }

    response = await client.post("/api/deals/import", json=payload, headers=auth_headers)

    assert response.status_code == 201
    data = response.json()
    assert data["project_id"]
    assert data["deal_model_id"]

    project_response = await client.get(f"/api/projects/{data['project_id']}", headers=auth_headers)
    assert project_response.status_code == 200
    assert project_response.json()["name"] == "Agent Imported Project"

    models_response = await client.get(f"/api/projects/{data['project_id']}/models", headers=auth_headers)
    assert models_response.status_code == 200
    assert len(models_response.json()) == 1
    assert models_response.json()[0]["name"] == "Imported Agent Deal"


async def test_exported_model_json_round_trips_through_validation_and_import(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    org, user = await seed_org(session)
    source_project = await seed_opportunity(session, org, user, name="Round Trip Source")
    target_project = await seed_opportunity(session, org, user, name="Round Trip Target")
    # import persists created_by_user_id (FK to users) from the X-User-ID header
    auth_headers = _as_user(auth_headers, user)

    top_deal = Deal(org_id=org.id, name="Round Trip Deal", created_by_user_id=user.id)
    session.add(top_deal)
    await session.flush()

    model = Scenario(
        deal_id=top_deal.id,
        created_by_user_id=user.id,
        name="Round Trip Deal",
        project_type=ProjectType.acquisition,
        is_active=True,
    )
    session.add(model)
    await session.flush()

    dev_project = Project(
        scenario_id=model.id,
        opportunity_id=source_project.id,
        name="Default Project",
    )
    session.add(dev_project)
    await session.flush()

    session.add(
        OperationalInputs(
            project_id=dev_project.id,
            unit_count_existing=6,
            exit_cap_rate_pct=Decimal("5.25"),
        )
    )
    session.add(
        IncomeStream(
            project_id=dev_project.id,
            stream_type=IncomeStreamType.residential_rent,
            label="Round Trip Rent",
            unit_count=6,
            amount_per_unit_monthly=Decimal("1750"),
            stabilized_occupancy_pct=Decimal("95"),
            active_in_phases=["hold", "stabilized"],
        )
    )

    capital_module = CapitalModule(
        id=uuid4(),
        scenario_id=model.id,
        label="Round Trip Equity",
        vehicle_type="equity",
        stack_position=1,
        source={"amount": 300000},
        carry={"carry_type": "none", "payment_frequency": "at_exit"},
        exit_terms={"exit_type": "profit_share", "trigger": "sale", "profit_share_pct": 100},
        active_phase_start="acquisition",
        active_phase_end="exit",
    )
    session.add(capital_module)
    await session.flush()
    session.add(
        WaterfallTier(
            scenario_id=model.id,
            capital_module_id=capital_module.id,
            priority=1,
            tier_type=WaterfallTierType.residual,
            lp_split_pct=Decimal("90"),
            gp_split_pct=Decimal("10"),
            description="Residual split",
        )
    )
    await session.commit()

    export_response = await client.get(f"/api/models/{model.id}/export/json", headers=auth_headers)
    assert export_response.status_code == 200
    exported_payload = export_response.json()

    validate_response = await client.post(
        "/api/models/import/validate",
        json=exported_payload,
        headers=auth_headers,
    )
    assert validate_response.status_code == 200
    assert validate_response.json()["valid"] is True

    import_response = await client.post(
        f"/api/projects/{target_project.id}/models/import",
        json=exported_payload,
        headers=auth_headers,
    )
    assert import_response.status_code == 201
    counts = import_response.json()["counts"]
    assert counts["income_streams"] == 1
    assert counts["capital_modules"] == 1
    assert counts["waterfall_tiers"] == 1


async def test_get_model_excel_export_returns_multisheet_workbook(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    _, _, _, model, dev_project = await _seed_stack(
        session,
        opp_name="Excel Export Project",
        deal_name="Tower Test Deal",
        project_type=ProjectType.acquisition,
        is_active=False,
    )

    session.add(
        OperationalInputs(
            project_id=dev_project.id,
            unit_count_existing=12,
            unit_count_new=2,
            expense_growth_rate_pct_annual=Decimal("3.0"),
            exit_cap_rate_pct=Decimal("5.25"),
        )
    )
    session.add(
        IncomeStream(
            project_id=dev_project.id,
            stream_type=IncomeStreamType.residential_rent,
            label="Unit Rent",
            unit_count=12,
            amount_per_unit_monthly=Decimal("1850"),
            stabilized_occupancy_pct=Decimal("95"),
            escalation_rate_pct_annual=Decimal("2.5"),
            active_in_phases=["hold", "stabilized"],
        )
    )
    session.add(
        CashFlow(
            scenario_id=model.id,
            period=1,
            period_type=PeriodType.acquisition,
            gross_revenue=Decimal("22200"),
            vacancy_loss=Decimal("1110"),
            effective_gross_income=Decimal("21090"),
            operating_expenses=Decimal("4800"),
            capex_reserve=Decimal("350"),
            noi=Decimal("15940"),
            debt_service=Decimal("7200"),
            net_cash_flow=Decimal("8740"),
            cumulative_cash_flow=Decimal("8740"),
        )
    )
    session.add(
        OperationalOutputs(
            scenario_id=model.id,
            total_project_cost=Decimal("1650000"),
            equity_required=Decimal("550000"),
            total_timeline_months=18,
            noi_stabilized=Decimal("192000"),
            cap_rate_on_cost_pct=Decimal("6.4"),
            dscr=Decimal("1.31"),
            project_irr_levered=Decimal("15.8"),
            project_irr_unlevered=Decimal("12.1"),
            computed_at=datetime(2026, 4, 2, 12, 0, tzinfo=UTC),
        )
    )
    capital_module = CapitalModule(
        scenario_id=model.id,
        label="Senior Loan",
        vehicle_type="debt",
        stack_position=1,
        source={"amount": 1100000, "interest_rate_pct": 6.1},
        carry={"carry_type": "io_only", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition",
        active_phase_end="exit",
    )
    session.add(capital_module)
    await session.flush()

    tier = WaterfallTier(
        scenario_id=model.id,
        capital_module_id=capital_module.id,
        priority=1,
        tier_type=WaterfallTierType.pref_return,
        irr_hurdle_pct=Decimal("8.0"),
        lp_split_pct=Decimal("90"),
        gp_split_pct=Decimal("10"),
        description="Preferred return tier",
    )
    session.add(tier)
    await session.flush()

    session.add(
        WaterfallResult(
            scenario_id=model.id,
            period=1,
            tier_id=tier.id,
            capital_module_id=capital_module.id,
            cash_distributed=Decimal("5000"),
            cumulative_distributed=Decimal("5000"),
            party_irr_pct=Decimal("8.4"),
        )
    )
    await session.commit()

    response = await client.get(f"/api/models/{model.id}/export/xlsx", headers=auth_headers)

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment; filename=" in response.headers["content-disposition"]

    # The export is now the multi-sheet investor workbook (Cover-first layout);
    # most sheets are conditional on model contents, so assert the stable spine.
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames[0] == "Cover"
    assert "Glossary & Methodology" in workbook.sheetnames
    assert len(workbook.sheetnames) >= 3

    cover_values = {
        str(cell.value)
        for row in workbook["Cover"].iter_rows()
        for cell in row
        if cell.value is not None
    }
    assert any("Tower Test Deal" in value for value in cover_values)


async def test_get_waterfall_report_returns_investor_timelines_and_summary(
    client: AsyncClient,
    session: AsyncSession,
    auth_headers: dict[str, str],
) -> None:
    _, user, _, model, _dev_project = await _seed_stack(
        session,
        opp_name="Investor Report Project",
        deal_name="Investor Distribution Deal",
        project_type=ProjectType.acquisition,
        is_active=False,
    )
    auth_headers = _as_user(auth_headers, user)  # capital routes are org-scoped

    lp_module = CapitalModule(
        scenario_id=model.id,
        label="LP Equity",
        vehicle_type="equity",
        equity_role="lp",
        stack_position=1,
        source={"amount": 40000},
        carry={"carry_type": "none", "payment_frequency": "at_exit"},
        exit_terms={"exit_type": "profit_share", "trigger": "sale", "profit_share_pct": 90},
        active_phase_start="acquisition",
        active_phase_end="exit",
    )
    gp_module = CapitalModule(
        scenario_id=model.id,
        label="GP Promote",
        vehicle_type="equity",
        equity_role="gp",
        stack_position=2,
        source={"amount": 10000},
        carry={"carry_type": "none", "payment_frequency": "at_exit"},
        exit_terms={"exit_type": "profit_share", "trigger": "sale", "profit_share_pct": 100},
        active_phase_start="acquisition",
        active_phase_end="exit",
    )
    session.add_all([lp_module, gp_module])
    await session.flush()

    pref_tier = WaterfallTier(
        scenario_id=model.id,
        capital_module_id=lp_module.id,
        priority=1,
        tier_type=WaterfallTierType.pref_return,
        lp_split_pct=Decimal("100"),
        gp_split_pct=Decimal("0"),
        description="LP pref return",
    )
    residual_tier = WaterfallTier(
        scenario_id=model.id,
        capital_module_id=None,
        priority=2,
        tier_type=WaterfallTierType.residual,
        lp_split_pct=Decimal("70"),
        gp_split_pct=Decimal("30"),
        description="Residual split",
    )
    session.add_all([pref_tier, residual_tier])
    await session.flush()

    session.add_all(
        [
            WaterfallResult(
                scenario_id=model.id,
                period=1,
                tier_id=pref_tier.id,
                capital_module_id=lp_module.id,
                cash_distributed=Decimal("5000"),
                cumulative_distributed=Decimal("5000"),
                party_irr_pct=None,
            ),
            WaterfallResult(
                scenario_id=model.id,
                period=2,
                tier_id=pref_tier.id,
                capital_module_id=lp_module.id,
                cash_distributed=Decimal("7000"),
                cumulative_distributed=Decimal("12000"),
                party_irr_pct=None,
            ),
            WaterfallResult(
                scenario_id=model.id,
                period=3,
                tier_id=pref_tier.id,
                capital_module_id=lp_module.id,
                cash_distributed=Decimal("6000"),
                cumulative_distributed=Decimal("18000"),
                party_irr_pct=Decimal("14.25"),
            ),
            WaterfallResult(
                scenario_id=model.id,
                period=3,
                tier_id=residual_tier.id,
                capital_module_id=lp_module.id,
                cash_distributed=Decimal("2000"),
                cumulative_distributed=Decimal("2000"),
                party_irr_pct=None,
            ),
            WaterfallResult(
                scenario_id=model.id,
                period=3,
                tier_id=residual_tier.id,
                capital_module_id=gp_module.id,
                cash_distributed=Decimal("7000"),
                cumulative_distributed=Decimal("7000"),
                party_irr_pct=Decimal("19.5"),
            ),
        ]
    )
    await session.commit()
    model_id = model.id
    lp_module_id = lp_module.id
    gp_module_id = gp_module.id

    response = await client.get(f"/api/models/{model_id}/waterfall/report", headers=auth_headers)

    assert response.status_code == 200
    data = response.json()
    assert data["scenario_id"] == str(model_id)
    assert data["investor_count"] == 2
    assert Decimal(str(data["total_cash_distributed"])) == Decimal("27000")

    assert [item["investor_name"] for item in data["investors"]] == ["LP Equity", "GP Promote"]

    lp_summary = next(item for item in data["investors"] if item["capital_module_id"] == str(lp_module_id))
    gp_summary = next(item for item in data["investors"] if item["capital_module_id"] == str(gp_module_id))

    assert lp_summary["vehicle_type"] == "equity"
    assert Decimal(str(lp_summary["committed_capital"])) == Decimal("40000")
    assert Decimal(str(lp_summary["total_cash_distributed"])) == Decimal("20000")
    assert Decimal(str(lp_summary["ending_cumulative_distributed"])) == Decimal("20000")
    assert Decimal(str(lp_summary["equity_multiple"])) == Decimal("0.5")
    assert Decimal(str(lp_summary["cash_on_cash_year_1_pct"])) == Decimal("50")
    assert Decimal(str(lp_summary["latest_party_irr_pct"])) == Decimal("14.25")
    assert lp_summary["timeline"] == [
        {"period": 1, "cash_distributed": "5000.000000", "cumulative_distributed": "5000.000000"},
        {"period": 2, "cash_distributed": "7000.000000", "cumulative_distributed": "12000.000000"},
        {"period": 3, "cash_distributed": "8000.000000", "cumulative_distributed": "20000.000000"},
    ]

    assert gp_summary["vehicle_type"] == "equity"
    assert Decimal(str(gp_summary["committed_capital"])) == Decimal("10000")
    assert Decimal(str(gp_summary["total_cash_distributed"])) == Decimal("7000")
    assert Decimal(str(gp_summary["ending_cumulative_distributed"])) == Decimal("7000")
    assert Decimal(str(gp_summary["equity_multiple"])) == Decimal("0.7")
    assert Decimal(str(gp_summary["cash_on_cash_year_1_pct"])) == Decimal("70")
    assert Decimal(str(gp_summary["latest_party_irr_pct"])) == Decimal("19.5")


async def test_post_scraper_run_enqueues_crexi_task(
    client: AsyncClient,
    auth_headers: dict[str, str],
) -> None:
    observed: dict[str, str] = {}

    class _FakeAsyncResult:
        id = "celery-task-123"

    class _FakeControl:
        def ping(self, timeout: float = 0.5) -> list[dict[str, str]]:
            return [{"worker@fake": "pong"}]

    class _FakeApp:
        control = _FakeControl()

    class _FakeTask:
        app = _FakeApp()

        def apply_async(self, *, kwargs=None, queue=None):
            assert kwargs is not None
            # triggered_by is hardcoded to "ui" in the route (no user attribution)
            assert kwargs["triggered_by"] == "ui"
            assert kwargs["trace_id"]
            observed["trace_id"] = kwargs["trace_id"]
            assert queue == "scraping"
            return _FakeAsyncResult()

    with patch("app.api.routers.ingest.scrape_crexi", new=_FakeTask()):
        response = await client.post("/api/scraper/run", headers=auth_headers)

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "queued"
    assert payload["task_id"] == "celery-task-123"
    assert payload["source"] == "crexi"
    assert payload["trace_id"] == observed["trace_id"] == response.headers["X-Trace-ID"]
    assert datetime.fromisoformat(payload["queued_at"])


async def test_get_model_excel_export_returns_404_for_missing_model(
    client: AsyncClient,
    auth_headers: dict[str, str],
) -> None:
    response = await client.get(f"/api/models/{uuid4()}/export/xlsx", headers=auth_headers)

    assert response.status_code == 404
    assert response.json() == {
        "code": "not_found",
        "message": "Scenario not found",
        "detail": None,
    }
