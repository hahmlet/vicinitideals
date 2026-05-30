"""Engine-vs-Excel parity on the S&U Operating Reserve UseLine.

The cashflow engine sizes the Operating Reserve as
``max(opex_monthly, ds_monthly) * reserve_months`` (revenue/opex mode).
For any debt-heavy stack, ``ds_monthly`` >> ``opex_monthly``, so the
reserve is debt-service-driven.

The S&U formula previously read ``=s_operating_reserve_months *
s_y1_opex / 12`` — opex-only. For the Brittany / anemic-construction
scenarios it under-stated the reserve by 4-5×, leaving Sources < Uses
in Excel even though the engine balanced.

This test seeds a scenario whose ``ds_monthly`` clearly exceeds
``opex_monthly``, recalcs the workbook, and asserts the S&U Operating
Reserve cell equals the engine-persisted UseLine amount within $1.
"""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from app.models.capital import (
    CapitalModule,
    CapitalModuleProject,
    EquityRole,
    VehicleType,
)
from app.models.deal import UseLine
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)
from tests.exporters._parity_helpers import (
    RecalcUnavailableError,
    recalc_workbook,
)


pytestmark = pytest.mark.asyncio


def _find_row_with_label_prefix(ws, prefix: str) -> int | None:
    needle = prefix.strip().lower()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip().lower().startswith(needle):
            return r
    return None


async def _seed_debt_heavy_scenario(session: AsyncSession):
    """Seed: small OpEx footprint + sizable PI-amortizing debt module so
    engine's `max(opex, ds) * months` lands on the debt-service branch."""
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="OpReserve Parity")
    deal_model, _inputs, seeded_stream, seeded_opex = (
        await seed_deal_model_with_financials(session, opp, user)
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()

    seeded_stream.active_in_phases = ["stabilized"]
    seeded_opex.active_in_phases = ["stabilized"]

    debt = CapitalModule(
        scenario_id=deal_model.id,
        label="Senior PI Loan",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={
            "amount": "5000000",
            "interest_rate_pct": 6.5,
            "amort_term_years": 30,
            "hold_term_years": 30,
            "auto_size": True,
            "binding_constraint": "gap_fill",
        },
        carry={
            "phases": [
                {"name": "construction", "carry_type": "pi"},
                {"name": "operation", "carry_type": "pi"},
            ]
        },
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition",
        active_phase_end="exit",
    )
    equity = CapitalModule(
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type=VehicleType.equity.value,
        equity_role=EquityRole.lp.value,
        stack_position=2,
        source={"amount": "1500000"},
        carry={"carry_type": "none", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition",
        active_phase_end="exit",
    )
    session.add_all([debt, equity])

    # Auto-sizer needs Uses to size against. One acquisition use line creates
    # a funding gap that the auto_size debt module fills, which in turn
    # triggers the Operating Reserve UseLine creation.
    session.add(
        UseLine(
            project_id=project.id,
            label="Land Acquisition",
            amount=Decimal("6000000"),
            phase="acquisition",
            cost_category="hard",
        )
    )
    await session.flush()

    session.add_all([
        CapitalModuleProject(
            capital_module_id=debt.id,
            project_id=project.id,
            amount=Decimal("5000000"),
        ),
        CapitalModuleProject(
            capital_module_id=equity.id,
            project_id=project.id,
            amount=Decimal("1500000"),
        ),
    ])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model, project


async def test_operating_reserve_excel_matches_engine_amount(
    session: AsyncSession, tmp_path: Path
) -> None:
    """S&U Operating Reserve cell (post-recalc) must match the engine's
    persisted UseLine.amount within $1 for a debt-heavy stack."""
    scenario, project = await _seed_debt_heavy_scenario(session)

    op_reserve = (
        await session.execute(
            select(UseLine).where(
                UseLine.project_id == project.id,
                UseLine.label == "Operating Reserve",
            )
        )
    ).scalar_one_or_none()
    assert op_reserve is not None, (
        "fixture should produce an Operating Reserve UseLine"
    )
    engine_amount = float(op_reserve.amount)
    assert engine_amount > 0

    blob = await export_investor_workbook(scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)
    try:
        recalc_workbook(path)
    except RecalcUnavailableError as exc:
        pytest.skip(f"no recalc backend: {exc}")

    wb = load_workbook(path, data_only=True)
    ws = wb["Sources & Uses"]
    reserve_row = _find_row_with_label_prefix(ws, "  Operating Reserve")
    assert reserve_row is not None, "Operating Reserve row not found in S&U"

    excel_amount = ws.cell(row=reserve_row, column=2).value
    assert isinstance(excel_amount, (int, float)), (
        f"Operating Reserve cell not numeric after recalc: {excel_amount!r}"
    )

    diff = abs(float(excel_amount) - engine_amount)
    assert diff < 1.0, (
        f"S&U Operating Reserve parity broken: engine={engine_amount}, "
        f"excel={excel_amount}, diff={diff}. Formula likely needs "
        f"MAX(opex, debt_service) instead of opex-only."
    )
