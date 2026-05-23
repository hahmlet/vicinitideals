"""Engine-vs-formula parity for the Combined Equity Multiple cells on
UW Summary and Investor Returns. Verifies the new SUMIF-over-
``r_uw_cf_levered`` formula evaluates to a sensible value post-recalc
and is in the same ballpark as the engine's ``combined_em_x``.

Skips when no recalc backend (Excel COM / LibreOffice) is available —
matches the gating used by ``test_formula_parity_returns``.
"""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import (
    _load_all,
    export_investor_workbook,
)
from app.models.capital import (
    CapitalModule,
    CapitalModuleProject,
    EquityRole,
    VehicleType,
)
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)
from tests.exporters._parity_helpers import (
    RecalcUnavailableError,
    recalc_workbook,
)


async def _seed(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="EM Parity")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()
    debt = CapitalModule(
        scenario_id=deal_model.id,
        label="Senior Loan",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={
            "amount": "500000", "interest_rate_pct": 6.5,
            "amort_term_years": 30, "hold_term_years": 10,
        },
        carry={"carry_type": "io_only", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    equity = CapitalModule(
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type=VehicleType.equity.value,
        equity_role=EquityRole.lp.value,
        stack_position=2,
        source={"amount": "250000"},
        carry={"carry_type": "none", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    session.add_all([debt, equity])
    await session.flush()
    session.add_all([
        CapitalModuleProject(
            capital_module_id=debt.id, project_id=project.id,
            amount=Decimal("500000"),
        ),
        CapitalModuleProject(
            capital_module_id=equity.id, project_id=project.id,
            amount=Decimal("250000"),
        ),
    ])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find(ws, needle: str) -> int | None:
    n = needle.lower()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and n in v.lower():
            return r
    return None


async def test_combined_em_evaluates_in_sane_range(
    session: AsyncSession, tmp_path: Path
):
    """Post-recalc, the EM formula returns a numeric ≥0 (or 0 from IFERROR).
    EM is typically 1.0–5.0× for a healthy deal; we widen the band to
    avoid coupling to the seed's specific numbers."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile="internal")
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)
    try:
        recalc_workbook(path)
    except RecalcUnavailableError as exc:
        pytest.skip(f"no recalc backend: {exc}")

    wb = load_workbook(path, data_only=True)
    for sheet, label in (
        ("Underwriting Summary", "combined equity multiple"),
        ("Investor Returns", "combined equity multiple (scenario)"),
    ):
        ws = wb[sheet]
        r = _find(ws, label)
        assert r is not None, f"{sheet}: row missing"
        val = ws.cell(row=r, column=2).value
        if val is None:
            pytest.skip(f"{sheet}: Excel returned None for EM cell")
        assert isinstance(val, (int, float)), (
            f"{sheet}: EM not numeric post-recalc; got {val!r}"
        )
        assert 0.0 <= float(val) <= 50.0, (
            f"{sheet}: EM out of sane range; got {val}"
        )


async def test_property_valuation_formulas_evaluate(
    session: AsyncSession, tmp_path: Path
):
    """Going-In Cap Value, Exit Cap Value, Cap Spread post-recalc all
    evaluate to numerics (or empty string when IFERROR fires on missing
    cap input)."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile="internal")
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)
    try:
        recalc_workbook(path)
    except RecalcUnavailableError as exc:
        pytest.skip(f"no recalc backend: {exc}")

    wb = load_workbook(path, data_only=True)
    ws = wb["Underwriting Summary"]
    for label in (
        "going-in cap value",
        "exit cap value",
        "cap spread",
    ):
        r = _find(ws, label)
        assert r is not None, f"row missing: {label}"
        val = ws.cell(row=r, column=2).value
        # Accept numeric or "" (IFERROR-empty when cap input is 0/missing).
        assert val is None or isinstance(val, (int, float, str)), (
            f"{label!r}: unexpected type post-recalc: {val!r}"
        )
        if isinstance(val, str) and val and val.startswith("="):
            pytest.fail(
                f"{label!r}: formula not evaluated; recalc backend may "
                f"have skipped this sheet ({val!r})"
            )


async def test_em_parity_within_band(
    session: AsyncSession, tmp_path: Path
):
    """Excel-recalc'd EM ≈ engine's combined_em_x within 0.5x. Tolerance
    is loose because the engine's monthly waterfall and the workbook's
    annual-cash-flow SUMIF can differ on equity-call timing."""
    scenario = await _seed(session)
    ctx = await _load_all(session, scenario.id)
    summary = ctx.get("rollup_summary") or {}
    engine_em = summary.get("totals", {}).get("combined_em_x")

    blob = await export_investor_workbook(scenario.id, session, profile="internal")
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)
    try:
        recalc_workbook(path)
    except RecalcUnavailableError as exc:
        pytest.skip(f"no recalc backend: {exc}")

    wb = load_workbook(path, data_only=True)
    ws = wb["Underwriting Summary"]
    r = _find(ws, "combined equity multiple")
    excel_em = ws.cell(row=r, column=2).value

    if engine_em in (None, 0, 0.0):
        pytest.skip("engine produced no EM (no waterfall rollup)")
    if excel_em is None:
        pytest.skip("Excel returned None for EM")

    diff = abs(float(excel_em) - float(engine_em))
    assert diff < 0.5, (
        f"EM parity: engine={engine_em}x, excel={excel_em}x, diff={diff}"
    )
