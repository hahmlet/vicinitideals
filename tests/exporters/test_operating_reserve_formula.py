"""Operating Reserve UseLine on the S&U sheet renders the engine's
per-project static amount (``use_lines.amount``), NOT a scenario-level
formula.

History: the row briefly carried
``=s_operating_reserve_months*MAX(s_y1_opex,s_pf_debt_service_y1)/12``,
but those are scenario-level named cells, so every project's Operating
Reserve row computed the same pool-wide value. Commit 6cec7f3 reverted
to the engine-stored per-project amount.

Guards three contracts:

  1. The Operating Reserve cell is the per-project static amount (the
     seeded UseLine amount), not a formula
  2. It must NOT reference the scenario-level named cells (regression
     guard against re-introducing the pool-wide-value bug)
  3. ``s_y1_opex`` / ``s_operating_reserve_months`` stay registered on
     every profile (other formulas still consume them)
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.exporters.investor_export import export_investor_workbook
from app.models.deal import UseLine
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


async def _seed_with_op_reserve(session: AsyncSession):
    """Seed a scenario + add an explicit Operating Reserve UseLine."""
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Op-Reserve Smoke")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()
    session.add(
        UseLine(
            id=uuid4(),
            project_id=project.id,
            label="Operating Reserve",
            cost_category="soft",
            amount=Decimal("48000"),
        )
    )
    await session.flush()
    return deal_model


def _find_op_reserve_cell(ws):
    """Walk col A for a row labeled like ``Operating Reserve``; return col B value."""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and "operating reserve" in v.lower():
            return ws.cell(row=r, column=2).value
    return None


@pytest.mark.parametrize("profile", ["internal", "lp", "lender", "proforma"])
async def test_operating_reserve_is_static_per_project_amount(
    session: AsyncSession, profile: str
):
    """Every profile that ships S&U renders Operating Reserve as the
    engine-stored per-project amount (commit 6cec7f3 revert)."""
    scenario = await _seed_with_op_reserve(session)
    blob = await export_investor_workbook(
        scenario.id, session, profile=profile,
    )
    wb = load_workbook(BytesIO(blob), data_only=False)

    # Locate the S&U sheet (may be named "Sources & Uses" or similar)
    su_sheet = None
    for name in wb.sheetnames:
        if "Sources" in name and "Uses" in name:
            su_sheet = wb[name]
            break
    assert su_sheet is not None, f"profile={profile} missing S&U sheet"

    value = _find_op_reserve_cell(su_sheet)
    assert value is not None, (
        f"profile={profile}: Operating Reserve row missing on S&U sheet"
    )
    if isinstance(value, str) and value.startswith("="):
        # Regression guard: scenario-level formula gave every project the
        # same pool-wide value — must not come back.
        assert "s_operating_reserve_months" not in value, (
            f"profile={profile}: Operating Reserve must not compute from "
            f"scenario-level named cells; got {value!r}"
        )
        assert "s_y1_opex" not in value, (
            f"profile={profile}: Operating Reserve must not compute from "
            f"scenario-level named cells; got {value!r}"
        )
        pytest.fail(
            f"profile={profile}: Operating Reserve must be the static "
            f"per-project amount; got formula {value!r}"
        )
    # The seeded UseLine amount must round-trip to the cell.
    assert abs(float(value) - 48000.0) < 0.01, (
        f"profile={profile}: Operating Reserve must equal the per-project "
        f"UseLine amount (48000); got {value!r}"
    )


@pytest.mark.parametrize("profile", ["internal", "lp", "lender", "proforma"])
async def test_y1_opex_and_reserve_months_names_registered(
    session: AsyncSession, profile: str
):
    """Both operand named ranges must be registered so the formula resolves."""
    scenario = await _seed_with_op_reserve(session)
    blob = await export_investor_workbook(
        scenario.id, session, profile=profile,
    )
    wb = load_workbook(BytesIO(blob), data_only=False)

    assert "s_y1_opex" in wb.defined_names, (
        f"profile={profile} missing s_y1_opex defined name "
        f"(formula would dangle to #NAME?)"
    )
    assert "s_operating_reserve_months" in wb.defined_names, (
        f"profile={profile} missing s_operating_reserve_months defined name"
    )
