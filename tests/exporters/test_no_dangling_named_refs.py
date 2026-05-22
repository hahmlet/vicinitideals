"""No formula in any exported profile may reference an unregistered
defined name. A dangling reference shows up as ``#NAME?`` when Excel
opens the workbook — silently broken until the LP scrolls past.

The bug we're guarding against: commit 2's S&U Sources rows emitted
``=s_module_<n>_principal`` formulas. The defined name only gets
registered inside ``_build_assumptions`` Block C. On the proforma
profile, Assumptions wasn't rendered, so the formulas referenced a
name that didn't exist. Fixed by adding proforma to ``_HAS_ASSUMPT``.
"""
from __future__ import annotations

import re
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.exporters.investor_export import export_investor_workbook
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


_NAMED_REF_RE = re.compile(r"\b([srp][0-9_a-z]*_[a-z0-9_]+)\b")


async def _seed(session: AsyncSession):
    org, user = await seed_org(session)
    opportunity = await seed_opportunity(
        session, org, user, name="Dangling-Ref Guard"
    )
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opportunity, user
    )
    return deal_model


@pytest.mark.parametrize("profile", ["internal", "lp", "lender", "proforma"])
async def test_every_named_ref_in_formula_resolves(
    session: AsyncSession, profile: str,
):
    """For every defined-name token referenced in a formula cell on any
    sheet, the workbook must have a matching DefinedName entry."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)

    registered = set(wb.defined_names)
    # Built-in / Excel-function names that look like our prefix but aren't.
    # Keep tight — anything matching our regex that isn't in `registered`
    # should be flagged.
    _SAFE = {"sum", "if", "iferror", "round", "max", "min"}

    dangling: list[tuple[str, str, str]] = []  # (sheet, cell, name)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                # Skip HYPERLINK formulas — they use sheet refs, not names.
                if "HYPERLINK(" in v:
                    continue
                for match in _NAMED_REF_RE.findall(v):
                    if match.lower() in _SAFE:
                        continue
                    if match in registered:
                        continue
                    dangling.append((sheet, cell.coordinate, match))

    assert not dangling, (
        f"profile={profile!r}: formulas reference unregistered names:\n"
        + "\n".join(
            f"  {s}!{c} -> {n}" for s, c, n in dangling[:20]
        )
    )
