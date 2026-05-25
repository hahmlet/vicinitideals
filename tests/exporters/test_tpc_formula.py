"""Phase 4 KPI-tail: Total Project Cost hero KPI becomes a formula.

TPC = Total Uses - balance-only Uses (reserves + capitalized-interest
stubs). Matches the engine's `_calculate_total_project_cost`, which
sums capital_event outflows and excludes the same `_BALANCE_ONLY_LABELS`
set the S&U sheet now tracks per-row.

Contract:
  1. Every Use-line row on S&U whose label is in `_BALANCE_ONLY_LABELS`
     contributes its B-column cell ref to a single
     ``s_su_balance_only_total`` named cell rendered as the
     "Balance-Only Subtotal" row below "Debt Sources Subtotal".
  2. UW Summary "Total Project Cost" cell is the formula
     ``=IFERROR(s_su_uses_total - s_su_balance_only_total, <fallback>)``.
  3. Numeric parity with the engine value within 0.01 once LibreOffice
     recalculates the workbook (separate gate — this file verifies
     formula shape + named-cell wiring only).
"""
from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows, _BALANCE_ONLY_LABELS
from app.exporters.investor_export import export_investor_workbook
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


pytestmark = pytest.mark.asyncio


async def _seed(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="TPC Formula Test")
    deal_model, _, stream, opex = await seed_deal_model_with_financials(
        session, opp, user
    )
    stream.active_in_phases = ["stabilized"]
    opex.active_in_phases = ["stabilized"]
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find_row(ws, label_exact: str) -> int | None:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip() == label_exact:
            return r
    return None


async def test_su_balance_only_total_registered(session: AsyncSession) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "s_su_balance_only_total" in set(wb.defined_names)


async def test_su_balance_only_subtotal_row_exists(session: AsyncSession) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Sources & Uses"]
    row = _find_row(ws, "Balance-Only Subtotal")
    assert row is not None, "Balance-Only Subtotal row missing on S&U"
    v = ws.cell(row=row, column=2).value
    assert isinstance(v, str) and v.startswith("="), (
        f"Balance-Only Subtotal must be a formula; got {v!r}"
    )


async def test_tpc_is_formula(session: AsyncSession) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Summary"]
    row = _find_row(ws, "Total Project Cost")
    assert row is not None
    v = ws.cell(row=row, column=2).value
    assert isinstance(v, str) and v.startswith("=IFERROR("), f"got {v!r}"
    assert "s_su_uses_total" in v and "s_su_balance_only_total" in v, (
        f"TPC formula must reference both S&U named totals; got {v!r}"
    )
    assert "s_su_uses_total-s_su_balance_only_total" in v, (
        f"TPC = uses - balance_only; got {v!r}"
    )


async def test_balance_only_subtotal_picks_up_balance_only_labels(
    session: AsyncSession,
) -> None:
    """Every Use-line row on S&U whose label is in `_BALANCE_ONLY_LABELS`
    must contribute its B-column cell ref to the Balance-Only Subtotal
    formula. When the seeded fixture has no such Use lines (default
    seed_deal_model_with_financials path doesn't add reserves), the
    formula must be the literal ``=0`` so TPC = Total Uses − 0 still
    resolves cleanly."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Sources & Uses"]
    balance_rows = [
        r for r in range(1, ws.max_row + 1)
        if isinstance((v := ws.cell(row=r, column=1).value), str)
        and any(label.lower() in v.lower() for label in _BALANCE_ONLY_LABELS)
    ]
    subtotal_row = _find_row(ws, "Balance-Only Subtotal")
    assert subtotal_row is not None
    subtotal_formula = ws.cell(row=subtotal_row, column=2).value
    if not balance_rows:
        assert subtotal_formula == "=0", (
            f"no balance-only Use rows seeded; formula must be '=0', got {subtotal_formula!r}"
        )
        return
    for br in balance_rows:
        assert f"B{br}" in subtotal_formula, (
            f"Balance-Only Subtotal must reference row B{br}; "
            f"got {subtotal_formula!r}"
        )


@pytest.mark.parametrize("profile", ["internal", "lp", "lender"])
async def test_tpc_formula_resolves_on_profile(
    session: AsyncSession, profile: str,
) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    if "Underwriting Summary" not in wb.sheetnames:
        pytest.skip(f"profile {profile!r} doesn't render UW Summary")
    names = set(wb.defined_names)
    assert "s_su_uses_total" in names
    assert "s_su_balance_only_total" in names
    assert "s_total_project_cost" in names
