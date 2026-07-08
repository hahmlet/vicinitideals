"""Phase C: Asset Mgmt Fee row on Pro Forma sheets must be a formula
``=IFERROR(-EGI_cell*s_asset_mgmt_fee,0)`` so LP edits to the Asset Mgmt
Fee Assumption immediately re-flow through every year.

Contract:

  1. The Asset Mgmt Fee row exists on Pro Forma for every profile that
     ships a Pro Forma sheet.
  2. Every year cell is a formula referencing ``s_asset_mgmt_fee`` and
     the EGI row on the same sheet (negative sign).
  3. ``s_asset_mgmt_fee`` defined name is registered on Assumptions.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from app.models.capital import (
    CapitalModule,
    CapitalModuleProject,
    EquityRole,
    VehicleType,
)
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)
from tests.exporters._parity_helpers import find_label_row, proforma_layout


async def _seed(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Asset Mgmt Fee Smoke")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()
    debt = CapitalModule(
        scenario_id=deal_model.id,
        label="Senior Loan",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={
            "amount": "500000", "interest_rate_pct": 6.5,
            "amort_term_years": 30, "hold_term_years": 10,
        },
        carry={"carry_type": "pi", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    equity = CapitalModule(
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type=VehicleType.equity.value,
        equity_role=EquityRole.lp.value,
        stack_position=2,
        source={"amount": "250000"},
        carry={"carry_type": "none", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    session.add_all([debt, equity])
    await session.flush()
    session.add_all([
        CapitalModuleProject(
            capital_module_id=debt.id, project_id=project.id,
            amount=Decimal("500000"),
        ),
        CapitalModuleProject(
            capital_module_id=equity.id, project_id=project.id,
            amount=Decimal("250000"),
        ),
    ])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find_row(ws, label_substr: str) -> int | None:
    label_col, _ = proforma_layout(ws)
    return find_label_row(ws, label_substr, col=label_col)


def _find_pf_sheet(wb):
    for name in wb.sheetnames:
        low = name.lower()
        if "underwriting pro forma" in low:
            return wb[name]
    for name in wb.sheetnames:
        low = name.lower()
        if "pro forma" in low or "proforma" in low:
            return wb[name]
    return None


@pytest.mark.parametrize("profile", ["internal", "lp", "lender", "proforma"])
async def test_asset_mgmt_fee_row_is_formula(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)

    ws = _find_pf_sheet(wb)
    assert ws is not None, f"profile={profile} missing Pro Forma sheet"

    fee_row = _find_row(ws, "asset mgmt fee")
    assert fee_row is not None, (
        f"profile={profile} missing Asset Mgmt Fee row on Pro Forma sheet"
    )

    egi_row = _find_row(ws, "effective gross income")
    assert egi_row is not None

    # Check Y0 + Y1 (layout-aware year columns).
    _, y0_col = proforma_layout(ws)
    for col in (y0_col, y0_col + 1):
        val = ws.cell(row=fee_row, column=col).value
        assert isinstance(val, str) and val.startswith("="), (
            f"profile={profile} col={col}: Asset Mgmt Fee must be formula; "
            f"got {val!r}"
        )
        assert "s_asset_mgmt_fee" in val, (
            f"profile={profile} col={col}: formula must reference "
            f"s_asset_mgmt_fee; got {val!r}"
        )
        assert f"{egi_row}" in val, (
            f"profile={profile} col={col}: formula must reference EGI row "
            f"{egi_row}; got {val!r}"
        )


@pytest.mark.parametrize("profile", ["internal", "lp", "lender", "proforma"])
async def test_asset_mgmt_fee_named_range_registered(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "s_asset_mgmt_fee" in wb.defined_names, (
        f"profile={profile} missing s_asset_mgmt_fee defined name"
    )
