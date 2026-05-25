"""Phase 4 hero-KPI conversions on the Underwriting Summary sheet.

Four Primary-KPI cells flip from engine-computed scalars to formulas
that resolve back to LP-editable Assumptions Block F/G inputs:

  - Combined Stabilized NOI  -> =IF(s_pf_noi_y1>0,s_pf_noi_y1,<engine fallback>)
  - Stabilized DSCR          -> =IF(AND(s_pf_noi_y1>0,s_pf_debt_service_y1>0),
                                    s_pf_noi_y1/s_pf_debt_service_y1, <fallback>)
  - Combined Levered IRR     -> =IFERROR(s_returns_combined_irr,0)
  - Total Modeled Duration   -> =IFERROR(MAX(p1_total_horizon_months,...),0)
                                (or single-ref / fallback for 1 / 0 projects)

Underlying named cells (s_pf_noi_y1, s_pf_debt_service_y1,
s_returns_combined_irr, p{idx}_total_horizon_months) must be registered
on their respective sheets; otherwise the formulas resolve to ``#NAME?``.
"""
from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


pytestmark = pytest.mark.asyncio


async def _seed(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Phase 4 Hero KPIs")
    deal_model, _, stream, opex = await seed_deal_model_with_financials(
        session, opp, user
    )
    stream.active_in_phases = ["stabilized"]
    opex.active_in_phases = ["stabilized"]
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find_row(ws, label_exact: str) -> int | None:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip() == label_exact:
            return r
    return None


async def test_pf_noi_y1_named_cell_registered(session: AsyncSession) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "s_pf_noi_y1" in set(wb.defined_names), (
        "s_pf_noi_y1 must be registered on the Pro Forma NOI row "
        "so UW Summary's Combined NOI / DSCR formulas resolve"
    )


async def test_pf_debt_service_y1_named_cell_registered(
    session: AsyncSession,
) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "s_pf_debt_service_y1" in set(wb.defined_names), (
        "s_pf_debt_service_y1 must be registered on the Pro Forma debt "
        "service row so the DSCR formula resolves without #NAME?"
    )


async def test_pf_gross_revenue_y1_named_cell_registered(
    session: AsyncSession,
) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "s_pf_gross_revenue_y1" in set(wb.defined_names)


async def test_combined_noi_is_hybrid_if_formula(session: AsyncSession) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Summary"]
    row = _find_row(ws, "Combined Stabilized NOI (DSCR basis)")
    assert row is not None, "Combined Stabilized NOI row missing"
    v = ws.cell(row=row, column=2).value
    assert isinstance(v, str) and v.startswith("=IF(s_pf_noi_y1>0,s_pf_noi_y1,"), (
        f"Combined NOI must be hybrid IF on s_pf_noi_y1; got {v!r}"
    )


async def test_dscr_is_formula_on_pro_forma_cells(session: AsyncSession) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Summary"]
    row = _find_row(ws, "Stabilized DSCR (combined)")
    assert row is not None, "Stabilized DSCR row missing"
    v = ws.cell(row=row, column=2).value
    assert isinstance(v, str) and v.startswith("=IF(AND("), f"got {v!r}"
    assert "s_pf_noi_y1>0" in v and "s_pf_debt_service_y1>0" in v, (
        f"DSCR formula must guard both NOI and DS > 0; got {v!r}"
    )
    assert "s_pf_noi_y1/s_pf_debt_service_y1" in v, (
        f"DSCR numerator/denominator wrong; got {v!r}"
    )


async def test_combined_levered_irr_computes_against_uw_cf_levered(
    session: AsyncSession,
) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Summary"]
    row = _find_row(ws, "Combined Levered IRR")
    assert row is not None
    v = ws.cell(row=row, column=2).value
    assert v == "=IFERROR(IRR(r_uw_cf_levered),0)", f"got {v!r}"
    assert "r_uw_cf_levered" in set(wb.defined_names), (
        "r_uw_cf_levered must be registered on the Underwriting Cash Flow "
        "sheet for the Combined Levered IRR formula to resolve"
    )


async def test_combined_levered_irr_lender_profile(
    session: AsyncSession,
) -> None:
    """Lender profile renders UW Summary but not Investor Returns. The
    Combined Levered IRR cell must compute directly against
    r_uw_cf_levered, not alias an Investor Returns named range that
    doesn't exist on this profile."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile="lender")
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "Investor Returns" not in wb.sheetnames
    ws = wb["Underwriting Summary"]
    row = _find_row(ws, "Combined Levered IRR")
    assert row is not None
    v = ws.cell(row=row, column=2).value
    assert v == "=IFERROR(IRR(r_uw_cf_levered),0)"


async def test_modeled_duration_is_max_over_phase_plans(
    session: AsyncSession,
) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Summary"]
    row = _find_row(ws, "Total Modeled Duration (months)")
    assert row is not None
    v = ws.cell(row=row, column=2).value
    assert isinstance(v, str) and v.startswith("=IFERROR("), f"got {v!r}"
    # Single-project seed: formula collapses to direct reference (no MAX).
    assert "p1_total_horizon_months" in v, (
        f"Duration formula must reference per-project horizon; got {v!r}"
    )
    assert "p1_total_horizon_months" in set(wb.defined_names), (
        "p1_total_horizon_months must be registered on Assumptions phase plan"
    )
