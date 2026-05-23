"""Formula-conversion §4.3: the Pro Forma Debt Service formula must gate
each loan by its `active_start_month` as well as `term_months`. A perm
loan in a construction-to-perm stack starts at month N (post-construction)
and should contribute $0 to debt service in years that end before month N.

Contract:

  1. Debt Schedule registers `s_loan_{n}_active_start_month` for every loan.
  2. Pro Forma Debt Service formula wraps each loan in
     `IF(AND(s_loan_{i}_active_start_month<=Y*12, s_loan_{i}_term_months>=Y*12),
         s_loan_{i}_annual_pi, 0)`.
  3. A loan with `active_phase_start="acquisition"` resolves to start_month=0
     so the gate is permissive (preserves pre-gating behavior).
  4. A loan with `active_phase_start="stabilized"` resolves to start_month
     equal to the cumulative pre-stabilization phase months (acquisition +
     hold + pre_construction + construction + lease_up).
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from app.models.capital import (
    CapitalModule,
    CapitalModuleProject,
    EquityRole,
    VehicleType,
)
from app.models.deal import DealModel, OperationalInputs
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


async def _seed_perm_only(session: AsyncSession):
    """Single perm loan, active from acquisition — gate should be permissive."""
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Perm-Only Gate")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()
    perm = CapitalModule(
        scenario_id=deal_model.id,
        label="Perm Loan",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={
            "amount": "500000", "interest_rate_pct": 6.5,
            "amort_term_years": 30, "hold_term_years": 10,
        },
        carry={"carry_type": "pi", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    equity = CapitalModule(
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type=VehicleType.equity.value,
        equity_role=EquityRole.lp.value,
        stack_position=2,
        source={"amount": "250000"},
        carry={"carry_type": "none", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    session.add_all([perm, equity])
    await session.flush()
    session.add_all([
        CapitalModuleProject(
            capital_module_id=perm.id, project_id=project.id,
            amount=Decimal("500000"),
        ),
        CapitalModuleProject(
            capital_module_id=equity.id, project_id=project.id,
            amount=Decimal("250000"),
        ),
    ])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


async def _seed_construction_to_perm(session: AsyncSession):
    """Construction-to-perm stack: bridge loan during construction, perm
    loan active from stabilized. Perm loan should not contribute to debt
    service until the stabilized phase begins.
    """
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="C2P Gate")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )
    # Force new_construction project type so the phase plan includes
    # pre_construction + construction phases ahead of stabilized.
    deal_model.project_type = "new_construction"
    await session.flush()
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()
    inputs = (
        await session.execute(
            select(OperationalInputs).where(
                OperationalInputs.project_id == project.id
            )
        )
    ).scalar_one()
    inputs.entitlement_months = 6
    inputs.construction_months = 18
    inputs.lease_up_months = 0
    inputs.hold_phase_enabled = False
    await session.flush()

    bridge = CapitalModule(
        scenario_id=deal_model.id,
        label="Construction Loan",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={
            "amount": "750000", "interest_rate_pct": 8.0,
            "amort_term_years": 30, "hold_term_years": 3,
        },
        carry={"carry_type": "pi", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="stabilized",
    )
    perm = CapitalModule(
        scenario_id=deal_model.id,
        label="Perm Loan",
        vehicle_type=VehicleType.debt.value,
        stack_position=2,
        source={
            "amount": "500000", "interest_rate_pct": 6.5,
            "amort_term_years": 30, "hold_term_years": 10,
        },
        carry={"carry_type": "pi", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="stabilized", active_phase_end="exit",
    )
    equity = CapitalModule(
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type=VehicleType.equity.value,
        equity_role=EquityRole.lp.value,
        stack_position=3,
        source={"amount": "250000"},
        carry={"carry_type": "none", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    session.add_all([bridge, perm, equity])
    await session.flush()
    session.add_all([
        CapitalModuleProject(
            capital_module_id=bridge.id, project_id=project.id,
            amount=Decimal("750000"),
        ),
        CapitalModuleProject(
            capital_module_id=perm.id, project_id=project.id,
            amount=Decimal("500000"),
        ),
        CapitalModuleProject(
            capital_module_id=equity.id, project_id=project.id,
            amount=Decimal("250000"),
        ),
    ])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find_row(ws, label_substr: str) -> int | None:
    needle = label_substr.lower()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and needle in v.lower():
            return r
    return None


@pytest.mark.parametrize("profile", ["internal", "lender"])
async def test_active_start_month_named_range_registered(
    session: AsyncSession, profile: str
):
    """Every loan on the Debt Schedule registers
    `s_loan_{n}_active_start_month` so the formula gate resolves."""
    scenario = await _seed_perm_only(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "s_loan_1_active_start_month" in wb.defined_names, (
        f"profile={profile} missing s_loan_1_active_start_month"
    )


async def test_active_start_month_zero_for_acquisition_start(
    session: AsyncSession,
):
    """A loan with active_phase_start='acquisition' originates at month 0."""
    scenario = await _seed_perm_only(session)
    blob = await export_investor_workbook(scenario.id, session, profile="internal")
    wb = load_workbook(BytesIO(blob), data_only=True)
    ws = wb["Debt Schedule"]
    row = _find_row(ws, "Perm Loan")
    assert row is not None
    # Active Start (mo) is col 12.
    assert ws.cell(row=row, column=12).value == 0


async def test_active_start_month_positive_for_stabilized_start(
    session: AsyncSession,
):
    """A perm loan with active_phase_start='stabilized' starts after
    acquisition + pre_construction + construction months (1 + 6 + 18 = 25)."""
    scenario = await _seed_construction_to_perm(session)
    blob = await export_investor_workbook(scenario.id, session, profile="internal")
    wb = load_workbook(BytesIO(blob), data_only=True)
    ws = wb["Debt Schedule"]
    perm_row = _find_row(ws, "Perm Loan")
    bridge_row = _find_row(ws, "Construction Loan")
    assert perm_row is not None and bridge_row is not None
    # Bridge (active_phase_start=acquisition): month 0.
    assert ws.cell(row=bridge_row, column=12).value == 0
    # Perm (active_phase_start=stabilized): 1 (acquisition) + 6
    # (pre_construction) + 18 (construction) = 25 months.
    assert ws.cell(row=perm_row, column=12).value == 25


@pytest.mark.parametrize("profile", ["internal", "lender"])
async def test_debt_service_formula_gates_by_active_start(
    session: AsyncSession, profile: str
):
    """Pro Forma Debt Service formula combines active-start + term gates
    via AND() so a perm loan whose active_start_month exceeds Y*12
    contributes 0 in that year."""
    scenario = await _seed_perm_only(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Pro Forma"]
    ds_row = _find_row(ws, "debt service")
    assert ds_row is not None, f"profile={profile}: debt service row missing"

    # Y1 (col 3) — first formula year.
    y1 = ws.cell(row=ds_row, column=3).value
    assert isinstance(y1, str) and y1.startswith("="), (
        f"profile={profile}: Y1 DS not a formula; got {y1!r}"
    )
    assert "AND(" in y1, (
        f"profile={profile}: DS formula must AND() the two gates; got {y1!r}"
    )
    assert "s_loan_1_active_start_month<=12" in y1, (
        f"profile={profile}: Y1 DS missing active-start gate; got {y1!r}"
    )
    assert "s_loan_1_term_months>=12" in y1, (
        f"profile={profile}: Y1 DS still needs term gate; got {y1!r}"
    )

    # Y3 scales the threshold to 36.
    y3 = ws.cell(row=ds_row, column=5).value
    assert isinstance(y3, str) and "s_loan_1_active_start_month<=36" in y3, (
        f"profile={profile}: Y3 DS missing 36-month active-start gate; got {y3!r}"
    )
    assert "s_loan_1_annual_pi" in y3, (
        f"profile={profile}: DS formula must still reference loan P&I; got {y3!r}"
    )
