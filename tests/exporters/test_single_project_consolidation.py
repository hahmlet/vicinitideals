"""Single-project consolidation: P1 sheet suppressed when only one project.

Commit 8 of docs/feature-plans/investor-excel-formula-conversion.md §5.

When a scenario has exactly one project, the per-project sheet
duplicates content already on Underwriting Pro Forma / Cash Flow. The
exporter now skips creating it and the UW Summary mini-table
HYPERLINK points at Underwriting Pro Forma instead.
"""
from __future__ import annotations

from io import BytesIO
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.exporters.investor_export import export_investor_workbook
from app.models.deal import OperationalInputs
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


async def _seed_minimal_scenario(session: AsyncSession):
    org, user = await seed_org(session)
    opportunity = await seed_opportunity(
        session, org, user, name="Consolidation Smoke"
    )
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opportunity, user
    )
    return deal_model


async def test_single_project_omits_p1_sheet(session: AsyncSession):
    scenario = await _seed_minimal_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    p_sheets = [s for s in wb.sheetnames if s.startswith("P") and " " in s]
    assert p_sheets == [], (
        f"single-project should omit per-project sheets; got {p_sheets!r}"
    )


async def test_multi_project_still_renders_per_project_sheets(
    session: AsyncSession,
):
    """Regression guard: multi-project keeps the P{n} sheets."""
    scenario = await _seed_minimal_scenario(session)
    from decimal import Decimal

    extra = Project(
        id=uuid4(),
        scenario_id=scenario.id,
        opportunity_id=None,
        name="Second Project",
    )
    session.add(extra)
    await session.flush()
    session.add(
        OperationalInputs(
            id=uuid4(), project_id=extra.id,
            unit_count_new=4,
            exit_cap_rate_pct=Decimal("5.5"),
        )
    )
    await session.flush()

    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    p_sheets = sorted(
        s for s in wb.sheetnames if s.startswith("P") and " " in s
    )
    assert len(p_sheets) == 2, (
        f"multi-project should keep P1 + P2 sheets; got {p_sheets!r}"
    )


async def test_uw_summary_hyperlink_targets_pro_forma_when_single(
    session: AsyncSession,
):
    """Single-project: per-project mini-table HYPERLINK points at
    Underwriting Pro Forma instead of a non-existent P1 sheet."""
    scenario = await _seed_minimal_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Summary"]

    hyperlink_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("=HYPERLINK("):
                hyperlink_cells.append(cell.value)

    # At least one HYPERLINK formula must target Underwriting Pro Forma.
    proforma_links = [v for v in hyperlink_cells if "Underwriting Pro Forma" in v]
    assert proforma_links, (
        f"expected ≥1 HYPERLINK to Underwriting Pro Forma; got {hyperlink_cells!r}"
    )
    # No HYPERLINK formula should target a P{n} sheet that no longer exists.
    p_sheet_links = [v for v in hyperlink_cells if "#'P1 " in v]
    assert not p_sheet_links, (
        f"single-project should not link to P1 sheet; got {p_sheet_links!r}"
    )
