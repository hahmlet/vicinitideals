"""Tests for the Assumptions-sheet expansion in commit 1 of the
formula-conversion plan.

Plan: docs/feature-plans/investor-excel-formula-conversion.md §3.

Asserts the new input surface is emitted and that every new range
resolves to a real cell on the Assumptions sheet. Does *not* assert
downstream formula behaviour — formulas land in subsequent commits.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.exporters.investor_export import (
    CAPITAL_STACK_HEADERS,
    export_investor_workbook,
)
from app.models.capital import (
    CapitalModule,
    WaterfallTier,
    WaterfallTierType,
)
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


# ── Helpers ───────────────────────────────────────────────────────────────────


async def _seed_scenario(session: AsyncSession):
    org, user = await seed_org(session)
    opportunity = await seed_opportunity(
        session, org, user, name="Assumptions-Expansion Smoke"
    )
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opportunity, user
    )
    return deal_model


def _load_assumptions(blob: bytes):
    wb = load_workbook(BytesIO(blob), data_only=False)
    return wb, wb["Assumptions"]


# ── Block A — new input rows ─────────────────────────────────────────────────


# Plan §3.1 — Block A additions. Each is a user-editable assumption
# that drives a downstream formula (Pro Forma vacancy, CapEx reserve
# line, exit selling-cost deduction, DCF discount rate).
NEW_BLOCK_A_NAMES = (
    "s_vacancy_pct",
    "s_capex_reserve_per_unit",
    "s_selling_costs_pct",
    "s_discount_rate",
)


async def test_block_a_new_inputs_registered(session: AsyncSession):
    """Every new Block A range exists and lives on the Assumptions sheet."""
    scenario = await _seed_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb, _ws = _load_assumptions(blob)

    missing = [n for n in NEW_BLOCK_A_NAMES if n not in wb.defined_names]
    assert not missing, f"missing Block A named ranges: {missing}"

    for name in NEW_BLOCK_A_NAMES:
        defined = wb.defined_names[name]
        destinations = list(defined.destinations)
        assert destinations, f"named range {name} has no destinations"
        sheet, _ref = destinations[0]
        assert sheet == "Assumptions", (
            f"named range {name} points at {sheet}, not Assumptions"
        )


# ── Block C — expanded debt-assumption columns ──────────────────────────────


# Plan §3.1 — the 8 new debt-assumption columns added per module so
# carry-cost formulas in commit 4 can reference them by name.
NEW_BLOCK_C_PER_MODULE_SUFFIXES = (
    "term_years",
    "amort_years",
    "io_months",
    "carry_type",
    "day_count",
    "dscr_min",
    "ltv_pct",
    "prepay_pct",
)


async def test_block_c_header_row_includes_all_debt_fields(session: AsyncSession):
    """Block C header reflects the full 14-column post-expansion layout."""
    scenario = await _seed_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    _wb, ws = _load_assumptions(blob)

    # Block C section label lives on the row of its own; the header row is
    # the row immediately below. Find by scanning column 1 for the
    # section text.
    block_c_label_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "C. Capital Stack":
            block_c_label_row = r
            break
    assert block_c_label_row is not None, "Block C section label not found"

    header_row = block_c_label_row + 1
    actual = tuple(
        ws.cell(row=header_row, column=c).value
        for c in range(1, len(CAPITAL_STACK_HEADERS) + 1)
    )
    assert actual == CAPITAL_STACK_HEADERS, (
        f"Block C header mismatch:\n  expected {CAPITAL_STACK_HEADERS}\n  got      {actual}"
    )


async def test_block_c_module_rows_register_all_debt_inputs(session: AsyncSession):
    """Per-module named ranges exist for every new debt-assumption column."""
    scenario = await _seed_scenario(session)
    # Ensure at least one CapitalModule exists with populated source/carry so
    # the seed-default fixture's debt module gets a row.
    debt = CapitalModule(
        id=uuid4(),
        scenario_id=scenario.id,
        label="Perm Debt (test)",
        stack_position=1,
        vehicle_type="debt",
        equity_role=None,
        source={
            "amount": "5000000",
            "interest_rate_pct": 6.5,
            "hold_term_years": 10,
            "dscr_min": 1.25,
            "ltv_pct": 65.0,
            "prepay_penalty_pct": 2.0,
            "auto_size": False,
        },
        carry={
            "carry_type": "pi",
            "amort_term_years": 30,
            "io_period_months": 12,
            "day_count": "30_360",
            "io_rate_pct": 6.5,
        },
    )
    session.add(debt)
    await session.flush()

    blob = await export_investor_workbook(scenario.id, session)
    wb, _ws = _load_assumptions(blob)

    # The seeder may already have created one or more modules. Walk module
    # ordinals 1..N (until a gap) and assert all 8 new suffixes registered.
    found_any_module = False
    for m_idx in range(1, 20):
        sentinel = f"s_module_{m_idx}_principal"
        if sentinel not in wb.defined_names:
            break
        found_any_module = True
        for suffix in NEW_BLOCK_C_PER_MODULE_SUFFIXES:
            name = f"s_module_{m_idx}_{suffix}"
            assert name in wb.defined_names, (
                f"module {m_idx} missing expected debt-input range {name}"
            )

    assert found_any_module, "no capital modules registered for the test fixture"


# ── Block D — Waterfall hurdles ─────────────────────────────────────────────


async def test_block_d_waterfall_hurdles_registered(session: AsyncSession):
    """Each WaterfallTier gets per-tier inputs registered as named ranges."""
    scenario = await _seed_scenario(session)
    # Seed two tiers so we can verify the per-tier suffix pattern.
    for priority, hurdle, lp, gp in [
        (1, Decimal("8.0"), Decimal("100"), Decimal("0")),
        (2, Decimal("12.0"), Decimal("80"), Decimal("20")),
    ]:
        session.add(
            WaterfallTier(
                id=uuid4(),
                scenario_id=scenario.id,
                priority=priority,
                tier_type=WaterfallTierType.irr_hurdle_split,
                irr_hurdle_pct=hurdle,
                lp_split_pct=lp,
                gp_split_pct=gp,
            )
        )
    await session.flush()

    blob = await export_investor_workbook(scenario.id, session)
    wb, _ws = _load_assumptions(blob)

    expected = {
        "s_tier_1_irr_hurdle",
        "s_tier_1_lp_split",
        "s_tier_1_gp_split",
        "s_tier_2_irr_hurdle",
        "s_tier_2_lp_split",
        "s_tier_2_gp_split",
    }
    missing = expected - set(wb.defined_names)
    assert not missing, f"missing Block D ranges: {missing}"

    for name in expected:
        destinations = list(wb.defined_names[name].destinations)
        sheet, _ref = destinations[0]
        assert sheet == "Assumptions"


# ── Sanity — total named-range count grew ────────────────────────────────────


async def test_assumptions_expansion_added_input_ranges(session: AsyncSession):
    """All new Block A ranges register on a default-seeded scenario.

    Block C expansion ranges are conditional on the scenario having
    CapitalModule rows; the default seeder may produce zero modules,
    so Block C is asserted explicitly in
    ``test_block_c_module_rows_register_all_debt_inputs``.
    """
    scenario = await _seed_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb, _ws = _load_assumptions(blob)

    new_block_a = sum(1 for n in NEW_BLOCK_A_NAMES if n in wb.defined_names)
    assert new_block_a == len(NEW_BLOCK_A_NAMES)
