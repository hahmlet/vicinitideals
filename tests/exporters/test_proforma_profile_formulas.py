"""Proforma profile Pro Forma sheet must use the same EGI / NOI
formulas as the internal-profile Underwriting Pro Forma. ``_write_pf_table``
is shared by both `_build_proforma_combined` and `_build_proforma_project_sheet`
so a single test on the combined sheet guards both.

Prior bug: commit 3 wired formulas into `_build_uw_proforma` only;
`_write_pf_table` still emitted scalar values. Proforma-profile exports
shipped a Pro Forma sheet with hard-coded numbers that ignored
Assumptions edits.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.exporters.investor_export import export_investor_workbook
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


async def _seed(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Proforma Smoke")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )
    return deal_model


def _find_row(ws, label_prefix: str) -> int | None:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.startswith(label_prefix):
            return r
    return None


async def test_proforma_egi_row_is_formula(session: AsyncSession):
    scenario = await _seed(session)
    blob = await export_investor_workbook(
        scenario.id, session, profile="proforma",
    )
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "Pro Forma" in wb.sheetnames
    ws = wb["Pro Forma"]

    egi_row = _find_row(ws, "Effective Gross Income")
    assert egi_row is not None

    formula_count = 0
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=egi_row, column=c).value
        if v is None:
            break
        assert isinstance(v, str) and v.startswith("="), (
            f"Pro Forma EGI Y{c - 2} should be a formula; got {v!r}"
        )
        formula_count += 1
    assert formula_count > 0


async def test_proforma_noi_row_is_formula(session: AsyncSession):
    scenario = await _seed(session)
    blob = await export_investor_workbook(
        scenario.id, session, profile="proforma",
    )
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Pro Forma"]

    noi_row = _find_row(ws, "NOI")
    assert noi_row is not None

    formula_count = 0
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=noi_row, column=c).value
        if v is None:
            break
        assert isinstance(v, str) and v.startswith("="), (
            f"Pro Forma NOI Y{c - 2} should be a formula; got {v!r}"
        )
        assert "-" in v, (
            f"NOI formula expected to contain '-' (EGI - OpEx - CapEx); "
            f"got {v!r}"
        )
        formula_count += 1
    assert formula_count > 0


async def test_proforma_profile_includes_assumptions_sheet(
    session: AsyncSession,
):
    """Regression guard for the S&U dangling ``s_module_*_principal``
    fix — the proforma profile must render Assumptions so the names
    are registered."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(
        scenario.id, session, profile="proforma",
    )
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "Assumptions" in wb.sheetnames, (
        f"proforma profile missing Assumptions sheet; got {wb.sheetnames}"
    )
