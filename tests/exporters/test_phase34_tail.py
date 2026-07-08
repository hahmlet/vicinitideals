"""Phase 3 + 4 tail items:

  Phase 3 (revenue side): Pro Forma renders one indented breakout
  row per IncomeStream grouped with the Gross Revenue total,
  mirroring the OpEx breakout shipped earlier in Phase 3. Each row
  references Assumptions Block F cells so an LP can trace any
  single revenue stream back to its input.

  Phase 4 (Total Uses formula): the Underwriting Summary "Total
  Uses" KPI becomes ``=s_su_uses_total`` instead of an engine
  scalar, so a Use-line edit on Sources & Uses ripples through
  immediately.

Contract (post-consolidation layout, commit e7ba809 — bullets emit
directly ABOVE the total, which SUMs the bullet column range):

  1. Total "Gross Revenue" row stays formula-driven off Block F —
     a SUM over the bullet rows.
  2. One indented breakout row per IncomeStream appears directly
     above the total, inside the total's SUM range.
  3. Each breakout's Y0 is blank, Y1 = ``=s_rev_<slug>_y1_monthly*12``,
     Y2 = prior-year × (1 + per-stream escalation_pct).
  4. UW Summary "Total Uses" cell is the literal formula
     ``=s_su_uses_total``.
"""
from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)
from tests.exporters._parity_helpers import (
    find_label_row,
    parse_sum_range,
    proforma_layout,
)


pytestmark = pytest.mark.asyncio


async def _seed(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Phase 3+4 Tail")
    deal_model, _, stream, opex = await seed_deal_model_with_financials(
        session, opp, user
    )
    stream.active_in_phases = ["stabilized"]
    opex.active_in_phases = ["stabilized"]
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find_row(ws, label_exact: str) -> int | None:
    label_col, _ = proforma_layout(ws)
    return find_label_row(ws, label_exact, col=label_col, exact=True)


def _find_indented_row(ws, label_substr: str) -> int | None:
    label_col, _ = proforma_layout(ws)
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=label_col).value
        if (
            isinstance(v, str)
            and "•" in v
            and label_substr.lower() in v.lower()
        ):
            return r
    return None


# ── Revenue breakout (Phase 3 revenue side) ──────────────────────────


async def test_gross_revenue_total_row_unchanged(
    session: AsyncSession,
) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Pro Forma"]
    row = _find_row(ws, "Gross Revenue")
    assert row is not None
    _, y0_col = proforma_layout(ws)
    y1 = ws.cell(row=row, column=y0_col + 1).value
    assert parse_sum_range(y1) is not None, (
        f"Total Gross Revenue Y1 must SUM the Block-F bullet rows; got {y1!r}"
    )


async def test_revenue_breakout_row_appears_below_total(
    session: AsyncSession,
) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Pro Forma"]
    total_row = _find_row(ws, "Gross Revenue")
    assert total_row is not None
    breakout = _find_indented_row(ws, "1BR Units")
    assert breakout is not None, (
        "1BR Units breakout row missing for Gross Revenue"
    )
    _, y0_col = proforma_layout(ws)
    parsed = parse_sum_range(ws.cell(row=total_row, column=y0_col + 1).value)
    assert parsed is not None, "total must SUM the bullet rows"
    _, first, last = parsed
    assert first <= breakout <= last < total_row, (
        f"breakout row ({breakout}) must sit inside the total's SUM range "
        f"({first}..{last}) above the total ({total_row})"
    )


async def test_revenue_breakout_y0_blank(session: AsyncSession) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Pro Forma"]
    breakout = _find_indented_row(ws, "1BR Units")
    assert breakout is not None
    _, y0_col = proforma_layout(ws)
    y0 = ws.cell(row=breakout, column=y0_col).value
    assert y0 in (None, ""), f"got {y0!r}"


async def test_revenue_breakout_y1_references_block_f(
    session: AsyncSession,
) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Pro Forma"]
    breakout = _find_indented_row(ws, "1BR Units")
    assert breakout is not None
    _, y0_col = proforma_layout(ws)
    y1 = ws.cell(row=breakout, column=y0_col + 1).value
    assert isinstance(y1, str) and y1.startswith("="), f"got {y1!r}"
    assert "s_rev_" in y1 and "_y1_monthly" in y1 and "*12" in y1, (
        f"breakout Y1 must annualize the Block F monthly cell; got {y1!r}"
    )


async def test_revenue_breakout_y2_uses_per_stream_escalation(
    session: AsyncSession,
) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Pro Forma"]
    breakout = _find_indented_row(ws, "1BR Units")
    assert breakout is not None
    _, y0_col = proforma_layout(ws)
    y1_col, y2_col = y0_col + 1, y0_col + 2
    y2 = ws.cell(row=breakout, column=y2_col).value
    assert isinstance(y2, str) and y2.startswith("="), f"got {y2!r}"
    assert "s_rev_" in y2 and "_escalation_pct" in y2, (
        f"breakout Y2 must reference per-stream escalation_pct; got {y2!r}"
    )
    prev_ref = f"{get_column_letter(y1_col)}{breakout}"
    assert prev_ref in y2, (
        f"breakout Y2 must reference its own prior-year cell {prev_ref}; "
        f"got {y2!r}"
    )


# ── Total Uses formula (Phase 4 tail) ────────────────────────────────


async def test_total_uses_is_formula_referencing_su_total(
    session: AsyncSession,
) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Summary"]
    row = _find_row(ws, "Total Uses")
    assert row is not None, "Total Uses row not found"
    v = ws.cell(row=row, column=2).value
    assert v == "=s_su_uses_total", (
        f"Total Uses must be =s_su_uses_total; got {v!r}"
    )
    # Defined-name registration preserved so downstream sheets can keep
    # referencing s_total_uses.
    assert "s_total_uses" in set(wb.defined_names)
