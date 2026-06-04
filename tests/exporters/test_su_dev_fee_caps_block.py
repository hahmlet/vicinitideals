"""Slice 3 (Export v3): Dev Fee Caps block on S&U renders one row per
CapitalModule with non-empty fee_terms, and marks the binding row.

Wiring contract:
  1. ``r_su_dev_fee_caps`` named range exists with the expected span.
  2. Each module with fee_terms gets a row carrying its label, max %,
     per-unit cap, absolute cap, and a binding indicator (✓ or —).
  3. The binding row (matching ``dev_fee_binding_context.binding_source_id``)
     uses the bold label font; non-binding rows use the value font.
  4. When no modules carry fee_terms and no Dev Fee row exists, the
     block is omitted entirely.
"""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.exporters.investor_export import export_investor_workbook
from app.models.capital import CapitalModule
from app.models.deal import UseLine
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


pytestmark = pytest.mark.asyncio


async def _seed_with_two_capped_modules(session: AsyncSession):
    """Seed a scenario with auto Dev Fee + two CapitalModules carrying
    different fee_terms. Module B is the binding source.
    """
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Dev Fee Caps")
    deal_model, _inputs, _stream, _opex = (
        await seed_deal_model_with_financials(session, opp, user)
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()

    # Acquisition + soft cost lines to give the Dev Fee something
    # to compute against.
    session.add(
        UseLine(
            project_id=project.id,
            label="Land Acquisition",
            amount=Decimal("4000000"),
            phase="acquisition",
            cost_category="hard",
        )
    )

    module_a = CapitalModule(
        scenario_id=deal_model.id,
        label="Source A — Loose Cap",
        vehicle_type="debt",
        stack_position=1,
        fee_terms={
            "max_pct": "15.0",
            "absolute_cap": "1500000",
        },
        fee_terms_inherited_from_type=False,
    )
    module_b = CapitalModule(
        scenario_id=deal_model.id,
        label="Source B — Tight Cap",
        vehicle_type="debt",
        stack_position=2,
        fee_terms={
            "max_pct": "8.0",
            "per_unit_cap": "5000",
            "absolute_cap": "750000",
        },
        fee_terms_inherited_from_type=False,
    )
    session.add_all([module_a, module_b])
    await session.flush()

    # Auto Dev Fee row with binding context pointing at module_b.
    dev_fee = UseLine(
        project_id=project.id,
        label="Developer Fee",
        amount=Decimal("750000"),
        phase="soft_costs",
        cost_category="soft",
        is_auto_dev_fee=True,
        dev_fee_pct=Decimal("8.0"),
        dev_fee_binding_context={
            "binding_source_id": str(module_b.id),
            "per_source_allocation": [
                {
                    "capital_module_id": str(module_a.id),
                    "allowable": "1500000",
                },
                {
                    "capital_module_id": str(module_b.id),
                    "allowable": "750000",
                },
            ],
        },
    )
    session.add(dev_fee)
    await session.flush()
    await session.commit()
    return deal_model, module_a, module_b


def _find_row(ws, needle: str, *, after: int = 0) -> int | None:
    needle_norm = needle.strip().lower()
    for r in range(max(1, after + 1), ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and needle_norm in v.strip().lower():
            return r
    return None


async def test_caps_block_renders_both_modules_with_binding_marker(
    session: AsyncSession, tmp_path: Path
) -> None:
    scenario, mod_a, mod_b = await _seed_with_two_capped_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)

    wb = load_workbook(path, data_only=False)
    ws = wb["Sources & Uses"]

    caps_header = _find_row(ws, "Dev Fee Caps (per Source)")
    assert caps_header is not None, "Dev Fee Caps section label missing"

    # The named range r_su_dev_fee_caps must exist.
    assert "r_su_dev_fee_caps" in wb.defined_names, (
        "r_su_dev_fee_caps named range missing"
    )

    # Find the two module rows by label — must search AFTER the caps
    # header because each module also appears in the Sources block above.
    row_a = _find_row(ws, "Source A — Loose Cap", after=caps_header)
    row_b = _find_row(ws, "Source B — Tight Cap", after=caps_header)
    assert row_a is not None and row_b is not None
    assert row_a > caps_header and row_b > caps_header

    # Column layout: 1=label, 2=max%, 3=per_unit, 4=abs cap, 5=allowable, 6=binding
    # Module A: max_pct=15.0% → 0.15
    assert float(ws.cell(row=row_a, column=2).value) == pytest.approx(0.15)
    # Module A has no per_unit_cap.
    assert ws.cell(row=row_a, column=3).value in (None, "", 0)
    # Module A absolute_cap = 1_500_000.
    assert float(ws.cell(row=row_a, column=4).value) == pytest.approx(1_500_000)
    # Module A allowable.
    assert float(ws.cell(row=row_a, column=5).value) == pytest.approx(1_500_000)
    # Module A is NOT binding.
    assert ws.cell(row=row_a, column=6).value == "—"

    # Module B: max_pct=8.0% → 0.08, per_unit=5000, abs=750_000.
    assert float(ws.cell(row=row_b, column=2).value) == pytest.approx(0.08)
    assert float(ws.cell(row=row_b, column=3).value) == pytest.approx(5_000)
    assert float(ws.cell(row=row_b, column=4).value) == pytest.approx(750_000)
    assert float(ws.cell(row=row_b, column=5).value) == pytest.approx(750_000)
    # Module B IS binding.
    assert ws.cell(row=row_b, column=6).value == "✓"

    # Binding row font is bold (FONT_LABEL); non-binding uses FONT_VALUE.
    assert ws.cell(row=row_b, column=1).font.bold is True
    assert ws.cell(row=row_a, column=1).font.bold in (False, None)


async def test_caps_block_omitted_when_no_fee_terms(
    session: AsyncSession, tmp_path: Path
) -> None:
    """No fee_terms on any module → block must not render."""
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="No Caps")
    deal_model, _i, _s, _o = (
        await seed_deal_model_with_financials(session, opp, user)
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()
    session.add(
        UseLine(
            project_id=project.id,
            label="Land Acquisition",
            amount=Decimal("4000000"),
            phase="acquisition",
            cost_category="hard",
        )
    )
    await session.flush()
    await session.commit()

    blob = await export_investor_workbook(deal_model.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)
    wb = load_workbook(path, data_only=False)
    ws = wb["Sources & Uses"]

    assert _find_row(ws, "Dev Fee Caps (per Source)") is None
    assert "r_su_dev_fee_caps" not in wb.defined_names
