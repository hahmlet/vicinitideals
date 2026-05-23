"""Phase F: Cover hero block must cross-ref UW Summary named ranges for
profiles that render UW Summary, so edits to revenue/OpEx/exit assumptions
ripple straight to the LP-facing front page.

Contract:

  1. Profiles with UW Summary (internal/lp/lender): Cover "Stabilized NOI"
     == "=s_combined_noi"; "Levered IRR" == "=s_combined_irr"; "Cap Rate on
     Cost" == "=IFERROR(s_combined_noi/s_su_uses_total,...)".
  2. Profile without UW Summary (proforma): cells stay numeric so the
     workbook still opens (referenced names wouldn't exist).
  3. The named operands must actually be registered on the rendered
     profiles so the Cover formulas resolve, not dangle to #NAME?.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from app.models.capital import (
    CapitalModule,
    CapitalModuleProject,
    EquityRole,
    VehicleType,
)
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


async def _seed(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Cover Hero Smoke")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()
    debt = CapitalModule(
        scenario_id=deal_model.id,
        label="Senior Loan",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={
            "amount": "500000", "interest_rate_pct": 6.5,
            "amort_term_years": 30, "hold_term_years": 10,
        },
        carry={"carry_type": "pi", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    equity = CapitalModule(
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type=VehicleType.equity.value,
        equity_role=EquityRole.lp.value,
        stack_position=2,
        source={"amount": "250000"},
        carry={"carry_type": "none", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    session.add_all([debt, equity])
    await session.flush()
    session.add_all([
        CapitalModuleProject(
            capital_module_id=debt.id, project_id=project.id,
            amount=Decimal("500000"),
        ),
        CapitalModuleProject(
            capital_module_id=equity.id, project_id=project.id,
            amount=Decimal("250000"),
        ),
    ])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find_cover_value(wb, label_substr: str):
    cover = wb["Cover"]
    needle = label_substr.lower()
    for r in range(1, cover.max_row + 1):
        v = cover.cell(row=r, column=1).value
        if isinstance(v, str) and needle in v.lower():
            return cover.cell(row=r, column=2).value
    return None


@pytest.mark.parametrize("profile", ["internal", "lp", "lender"])
async def test_cover_noi_and_irr_are_cross_sheet_formulas(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)

    noi = _find_cover_value(wb, "stabilized noi")
    irr = _find_cover_value(wb, "levered irr")
    cap_rate = _find_cover_value(wb, "cap rate on cost")

    assert noi == "=s_combined_noi", (
        f"profile={profile}: Cover Stabilized NOI must be formula; got {noi!r}"
    )
    assert irr == "=s_combined_irr", (
        f"profile={profile}: Cover Levered IRR must be formula; got {irr!r}"
    )
    assert (
        isinstance(cap_rate, str)
        and "s_combined_noi" in cap_rate
        and "s_su_uses_total" in cap_rate
    ), (
        f"profile={profile}: Cover Cap Rate on Cost must reference both "
        f"named operands; got {cap_rate!r}"
    )


@pytest.mark.parametrize("profile", ["internal", "lp", "lender"])
async def test_cover_formula_operands_registered(
    session: AsyncSession, profile: str
):
    """``s_combined_noi``, ``s_combined_irr``, ``s_su_uses_total`` must all
    be defined names; otherwise the Cover formulas would resolve to #NAME?
    in Excel."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)

    for required in ("s_combined_noi", "s_combined_irr", "s_su_uses_total"):
        assert required in wb.defined_names, (
            f"profile={profile} missing {required} defined name"
        )


async def test_cover_proforma_profile_keeps_numeric_seeds(
    session: AsyncSession,
):
    """Proforma profile omits UW Summary — the cross-sheet operand cells
    don't exist, so the Cover hero must stay numeric to keep the workbook
    opening cleanly."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile="proforma")
    wb = load_workbook(BytesIO(blob), data_only=False)

    noi = _find_cover_value(wb, "stabilized noi")
    irr = _find_cover_value(wb, "levered irr")
    assert not (isinstance(noi, str) and noi.startswith("=")), (
        f"proforma: Cover Stabilized NOI must be numeric (no UW Summary "
        f"to ref); got formula {noi!r}"
    )
    assert not (isinstance(irr, str) and irr.startswith("=")), (
        f"proforma: Cover Levered IRR must be numeric; got formula {irr!r}"
    )
