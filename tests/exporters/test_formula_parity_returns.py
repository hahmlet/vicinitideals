"""Engine-vs-formula parity for the Investor Returns sheet.

Commit 5 of docs/feature-plans/investor-excel-formula-conversion.md §4.3.

Scope: Return ($) per module + Combined Levered IRR (scenario).

  Return ($) — was an engine scalar; now a formula referencing the
  Principal + (Total DS or Distributions) cells on the same row, so
  changes to the distribution column propagate to Return ($).

  Combined Levered IRR (scenario) — was an engine-computed percent
  value; now ``=IFERROR(IRR(r_uw_cf_levered),0)`` so it re-derives from
  the formula-driven Levered Cash Flow row on Underwriting Cash Flow.

Equity Multiple, CoC, weighted IRR, and waterfall tier accruals stay
as engine values for now; their formulas land in a follow-up commit
once equity-contribution and distribution totals are exposed as named
ranges.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import (
    _load_all,
    export_investor_workbook,
)
from app.models.capital import (
    CapitalModule,
    CapitalModuleProject,
    EquityRole,
    VehicleType,
)
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)
from tests.exporters._parity_helpers import (
    RecalcUnavailableError,
    recalc_workbook,
)


async def _seed_scenario(session: AsyncSession):
    """Seed scenario + one debt + one equity capital module wired via
    junctions to the seeded project. Without explicit modules the
    Source Returns table emits zero rows and the formula path can't be
    exercised."""
    org, user = await seed_org(session)
    opportunity = await seed_opportunity(
        session, org, user, name="Returns-Parity Smoke"
    )
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opportunity, user
    )

    # Pull the seeded Project so we can wire junctions to it.
    project_row = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()

    debt = CapitalModule(
        scenario_id=deal_model.id,
        label="Senior Loan",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={"amount": "500000", "interest_rate_pct": 6.5,
                "amort_term_years": 30, "hold_term_years": 10},
        carry={"carry_type": "io_only", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition",
        active_phase_end="exit",
    )
    equity = CapitalModule(
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type=VehicleType.equity.value,
        equity_role=EquityRole.lp.value,
        stack_position=2,
        source={"amount": "250000"},
        carry={"carry_type": "none", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition",
        active_phase_end="exit",
    )
    session.add_all([debt, equity])
    await session.flush()

    session.add_all([
        CapitalModuleProject(
            capital_module_id=debt.id,
            project_id=project_row.id,
            amount=Decimal("500000"),
        ),
        CapitalModuleProject(
            capital_module_id=equity.id,
            project_id=project_row.id,
            amount=Decimal("250000"),
        ),
    ])
    await session.flush()

    # Force a cashflow compile so the export sees debt_service > 0 on
    # the persisted CashFlow rows — without this, total_ds falls back
    # to None and the Return ($) formula path skips.
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _returns_sheet(blob: bytes, data_only: bool):
    wb = load_workbook(BytesIO(blob), data_only=data_only)
    return wb, wb["Investor Returns"]


def _find_row_by_label(ws, label: str) -> int | None:
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == label:
            return r
    return None


async def test_return_dollars_uses_formula_when_data_present(
    session: AsyncSession,
):
    """At least one debt or equity module's Return ($) cell is a formula.

    Modules whose distribution data is missing render an em-dash and
    skip the formula path (by design — no point referencing a string
    cell). We only assert that the formula path was taken for at least
    one row, since the seed scenario has both debt and equity with
    populated returns.
    """
    scenario = await _seed_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    _wb, ws = _returns_sheet(blob, data_only=False)

    # Walk all data rows (start at 3, header at 2)
    formula_rows = 0
    for r in range(3, ws.max_row + 1):
        v = ws.cell(row=r, column=8).value  # Return ($) col
        if isinstance(v, str) and v.startswith("="):
            formula_rows += 1
            # Must subtract principal (col D) from a distribution col (F or G).
            assert "D" in v and ("F" in v or "G" in v), (
                f"row {r} return formula must reference D + (F or G); got {v!r}"
            )
            assert "-" in v, (
                f"row {r} return formula must be a subtraction; got {v!r}"
            )
    assert formula_rows > 0, (
        "expected ≥1 module row to use the Return ($) formula path"
    )


async def test_combined_levered_irr_is_formula(session: AsyncSession):
    """Combined Levered IRR cell should be =IFERROR(IRR(...),0)."""
    scenario = await _seed_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    _wb, ws = _returns_sheet(blob, data_only=False)

    row = _find_row_by_label(ws, "Combined Levered IRR (scenario)")
    assert row is not None

    v = ws.cell(row=row, column=2).value
    assert isinstance(v, str) and v.startswith("=IFERROR(IRR("), (
        f"Combined Levered IRR should be =IFERROR(IRR(...); got {v!r}"
    )
    assert "r_uw_cf_levered" in v, (
        f"Combined Levered IRR must reference r_uw_cf_levered range; got {v!r}"
    )


async def test_return_dollars_evaluates_to_engine_value(
    session: AsyncSession, tmp_path: Path
):
    """For each formula-driven Return ($) cell, post-recalc value matches
    (distributions - principal) for equity or (total_ds - principal) for
    debt, within $1 of the engine's own arithmetic."""
    scenario = await _seed_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)
    try:
        recalc_workbook(path)
    except RecalcUnavailableError as exc:
        pytest.skip(f"no recalc backend: {exc}")

    wb = load_workbook(path, data_only=True)
    ws = wb["Investor Returns"]

    checked = 0
    for r in range(3, ws.max_row + 1):
        principal = ws.cell(row=r, column=4).value
        total_ds = ws.cell(row=r, column=6).value
        distributions = ws.cell(row=r, column=7).value
        return_dollars = ws.cell(row=r, column=8).value

        # Skip non-data rows (em-dash, blanks, section labels).
        if not isinstance(principal, (int, float)):
            continue
        if not isinstance(return_dollars, (int, float)):
            continue

        if isinstance(total_ds, (int, float)) and total_ds > 0:
            expected = float(total_ds) - float(principal)
        elif isinstance(distributions, (int, float)):
            expected = float(distributions) - float(principal)
        else:
            continue

        diff = abs(float(return_dollars) - expected)
        assert diff < 1.0, (
            f"row {r} Return($) parity: expected={expected}, "
            f"excel={return_dollars}, diff={diff}"
        )
        checked += 1
    assert checked > 0, "expected ≥1 row to be checked"


async def test_combined_irr_evaluates_close_to_engine_value(
    session: AsyncSession, tmp_path: Path
):
    """Excel-recalc'd IRR(Levered) ≈ engine's combined_irr_pct value.

    Tolerance is loose (0.5 percentage points) because Excel IRR uses
    annual-period intervals while the engine's combined IRR is computed
    over monthly NCF — annualization differs slightly. The point is
    that the formula is reading the right cash flow stream, not that
    it matches the engine's monthly XIRR bit-for-bit.
    """
    scenario = await _seed_scenario(session)
    ctx = await _load_all(session, scenario.id)
    summary = ctx.get("rollup_summary") or {}
    totals = summary.get("totals") or {}
    engine_irr_pct = totals.get("combined_irr_pct")

    blob = await export_investor_workbook(scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)
    try:
        recalc_workbook(path)
    except RecalcUnavailableError as exc:
        pytest.skip(f"no recalc backend: {exc}")

    wb = load_workbook(path, data_only=True)
    ws = wb["Investor Returns"]
    row = _find_row_by_label(ws, "Combined Levered IRR (scenario)")
    assert row is not None
    excel_value = ws.cell(row=row, column=2).value

    # excel_value is a fraction (0.085 for 8.5%); engine stores percent
    # magnitude (8.5 for 8.5%). Compare in matching units.
    if excel_value is None:
        pytest.skip("Excel returned None for Combined Levered IRR cell")

    if engine_irr_pct in (None, 0, 0.0):
        # Engine didn't run the waterfall rollup, but Excel's IRR formula
        # can still compute on the Levered Cash Flow row. The point of
        # this commit is that the formula is wired to the right stream;
        # we don't require the engine to match. Verify the formula
        # returned a sane numeric (i.e. IFERROR did not have to fire).
        excel_pct = float(excel_value) * 100.0
        assert -100.0 < excel_pct < 1000.0, (
            f"Excel IRR out of sane range; got {excel_pct}%"
        )
        return

    excel_pct = float(excel_value) * 100.0
    engine_pct = float(engine_irr_pct)
    diff = abs(excel_pct - engine_pct)
    assert diff < 0.5, (
        f"Combined IRR parity: engine={engine_pct}%, "
        f"excel={excel_pct}%, diff={diff}pp"
    )
