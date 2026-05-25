"""Phase 4 KPI-tail: Equity Required hero KPI becomes a formula.

Equity Required = max(0, Total Uses - sum of debt source principals).
Both operands are S&U named cells whose own formulas chain back to
Assumptions Block C debt principals and per-project Use lines, so any
LP edit ripples to the hero KPI without re-running the engine.

Contract:
  1. S&U sheet registers ``s_su_debt_sources_total`` summing only debt
     source rows (excluding the implied-equity gap row).
  2. UW Summary "Equity Required" cell is the formula
     ``=IFERROR(MAX(0, s_su_uses_total - s_su_debt_sources_total), <fallback>)``.
  3. Both referenced names resolve in every profile that renders UW
     Summary + S&U (internal, lp, lender).
"""
from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


pytestmark = pytest.mark.asyncio


async def _seed(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Equity Required Formula")
    deal_model, _, stream, opex = await seed_deal_model_with_financials(
        session, opp, user
    )
    stream.active_in_phases = ["stabilized"]
    opex.active_in_phases = ["stabilized"]
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find_row(ws, label_exact: str) -> int | None:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip() == label_exact:
            return r
    return None


async def test_su_debt_sources_total_registered(session: AsyncSession) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "s_su_debt_sources_total" in set(wb.defined_names)


async def test_su_debt_sources_subtotal_row_exists(session: AsyncSession) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Sources & Uses"]
    row = _find_row(ws, "Debt Sources Subtotal")
    assert row is not None, "Debt Sources Subtotal row missing on S&U"
    v = ws.cell(row=row, column=2).value
    assert isinstance(v, str) and v.startswith("="), (
        f"Debt Sources Subtotal must be a formula; got {v!r}"
    )


async def test_equity_required_is_formula(session: AsyncSession) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Summary"]
    row = _find_row(ws, "Equity Required")
    assert row is not None
    v = ws.cell(row=row, column=2).value
    assert isinstance(v, str) and v.startswith("=IFERROR(MAX(0,"), f"got {v!r}"
    assert "s_su_uses_total" in v and "s_su_debt_sources_total" in v, (
        f"Equity Required formula must reference both S&U named totals; got {v!r}"
    )


@pytest.mark.parametrize("profile", ["internal", "lp", "lender"])
async def test_equity_required_formula_resolves_on_profile(
    session: AsyncSession, profile: str,
) -> None:
    """Every profile that renders UW Summary also renders S&U, so both
    operands must be registered. Guards against future profile gating
    that might drop S&U while keeping UW Summary."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    if "Underwriting Summary" not in wb.sheetnames:
        pytest.skip(f"profile {profile!r} doesn't render UW Summary")
    names = set(wb.defined_names)
    assert "s_su_uses_total" in names
    assert "s_su_debt_sources_total" in names
    assert "s_equity_required" in names
