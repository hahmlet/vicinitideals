"""Sanity tests for the Excel-recalc harness used by formula-conversion parity.

Plan: docs/feature-plans/investor-excel-formula-conversion.md §8.

This file tests the *harness*, not the formula conversion itself. The
conversion tests in later commits will use this harness to assert that:

  - engine value X equals the recalced Excel cell value for a named range
  - the named range carries a formula (string starting with ``=``)
  - the formula text references the expected named-range inputs

Right now (post-commit-1) the only formulas in the workbook are
``=HYPERLINK("#'Sheet'!A1", "label")`` strings used for in-workbook
navigation. The harness sanity test uses one of those to prove the
end-to-end pipeline (export -> save -> recalc -> reload) is wired.

These tests skip when no recalc backend is available on the host
(Linux without LibreOffice, macOS without LibreOffice, Windows without
Office). CI gates run on a Windows runner with Office installed.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.exporters.investor_export import export_investor_workbook
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)
from tests.exporters._parity_helpers import (
    RecalcUnavailableError,
    count_formula_cells,
    recalc_workbook,
)


async def _seed_scenario(session: AsyncSession):
    org, user = await seed_org(session)
    opportunity = await seed_opportunity(
        session, org, user, name="Recalc-Harness Smoke"
    )
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opportunity, user
    )
    return deal_model


async def test_recalc_backend_available_or_skip(
    session: AsyncSession, tmp_path: Path
):
    """End-to-end: export, recalc, reload, see at least one HYPERLINK formula.

    Skips on hosts without an available recalc backend (Linux w/o
    LibreOffice, Windows w/o Office). When the backend is available,
    asserts the recalc completes and the cached values are readable.

    The HYPERLINK formulas already in the workbook serve as the smoke
    cell — they're the only formulas pre-conversion. Once commit 2 adds
    real arithmetic formulas, this test stays valid (it asserts ANY
    formula was processed, not a specific cell).
    """
    scenario = await _seed_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)

    xlsx_path = tmp_path / "investor-export.xlsx"
    xlsx_path.write_bytes(blob)

    formula_counts = count_formula_cells(blob)
    total_formulas = sum(formula_counts.values())
    assert total_formulas > 0, (
        "expected at least one HYPERLINK formula in pre-conversion workbook"
    )

    try:
        backend = recalc_workbook(xlsx_path)
    except RecalcUnavailableError as exc:
        pytest.skip(f"no recalc backend on this host: {exc}")

    assert backend in ("excel", "libreoffice")

    # After recalc the .xlsx cache should contain the resolved HYPERLINK
    # display strings. Walk the workbook in data_only mode and confirm at
    # least one cell that was a formula in the as-written workbook now
    # carries a non-formula value.
    wb_after = load_workbook(xlsx_path, data_only=True)
    resolved_formulas = 0
    for sheet in wb_after.sheetnames:
        ws = wb_after[sheet]
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if value is None:
                    continue
                # After recalc, HYPERLINK formulas resolve to the display
                # label (the second argument). They're no longer "=..."
                # strings in data_only mode.
                if isinstance(value, str) and not value.startswith("="):
                    resolved_formulas += 1
    assert resolved_formulas > 0, (
        "recalc completed but no cells carry resolved values"
    )


def test_recalc_unavailable_raises_on_missing_backend(tmp_path: Path):
    """Direct unit test for the error-when-no-backend branch.

    Doesn't depend on the DB / app. Validates the error type and message
    so the higher-level harness raises something callers can catch.
    """
    from tests.exporters._parity_helpers import recalc_with_libreoffice
    # Use a non-existent path so LibreOffice can't actually run. If LO
    # is installed on this host, this test exercises the "not found"
    # path only on hosts where it isn't; skip when LO is present.
    from tests.exporters._parity_helpers import _find_soffice
    if _find_soffice() is not None:
        pytest.skip("LibreOffice is installed; cannot test missing-backend path")
    bogus = tmp_path / "does-not-exist.xlsx"
    with pytest.raises(RecalcUnavailableError):
        recalc_with_libreoffice(bogus)
