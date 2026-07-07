"""Draw Schedule sheet tests (investor Excel export).

Covers the lender-package funding matrix added by the cross-surface sync
audit ("lenders often want a month-by-month funding schedule"):

- profile gating: sheet renders for ``internal`` + ``lender`` only
- empty state: sheet is skipped entirely when the scenario has no
  DrawSource rows (documented convention in export_investor_workbook)
- roster fallback: sources exist but milestones are undated → the sheet
  renders the source roster + hint instead of the monthly matrix
- row math: each source's monthly matrix cells sum to the engine's
  per-source total drawn (Source Summary block on the same sheet), and
  the grand total covers at least the seeded Uses

Seeding note: only CapitalModules are created here — the export's data
loader reuses the draw-schedule panel's engine path, whose reconciliation
backfills DrawSource rows from CapitalModules exactly like the UI panel
does on every load. That keeps these tests on the same code path a real
export uses.
"""
from __future__ import annotations

import uuid
from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.exporters.investor_export import export_investor_workbook
from app.models.capital import CapitalModule
from app.models.deal import UseLine
from app.models.milestone import Milestone, MilestoneType
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)

_SHEET = "Draw Schedule"
_MATRIX_TITLE = "Draw Schedule — Monthly Funding by Source"
_ROSTER_TITLE = "Draw Schedule — Funding Sources"

# Seeded Uses total (Land 800k + Hard Costs 1.2M). The engine's total
# sources required is Uses + debt carry, so the matrix grand total must
# be at least this.
_TOTAL_USES = Decimal("2000000")


async def _seed_scenario(session: AsyncSession, *, with_milestones: bool = True):
    """Scenario with Uses + capital stack (equity + debt) and dated milestones."""
    org, user = await seed_org(session)
    opportunity = await seed_opportunity(session, org, user, name="Draw Sched Opp")
    deal_model, inputs, _, _ = await seed_deal_model_with_financials(
        session, opportunity, user
    )
    project_id = inputs.project_id

    if with_milestones:
        # Anchor milestones (target_date set, no trigger chain) so
        # Milestone.computed_start resolves without a trigger walk.
        milestone_specs = [
            (MilestoneType.close, date(2026, 1, 15), 45, 1),
            (MilestoneType.pre_development, date(2026, 3, 1), 90, 2),
            (MilestoneType.construction, date(2026, 6, 1), 180, 3),
            (MilestoneType.operation_stabilized, date(2026, 12, 1), 365, 4),
        ]
        for ms_type, start, duration, seq in milestone_specs:
            session.add(
                Milestone(
                    id=uuid.uuid4(),
                    project_id=project_id,
                    milestone_type=ms_type,
                    target_date=start,
                    duration_days=duration,
                    sequence_order=seq,
                )
            )

    session.add(
        UseLine(
            id=uuid.uuid4(),
            project_id=project_id,
            label="Land Acquisition",
            phase="acquisition",
            amount=Decimal("800000"),
            timing_type="first_day",
        )
    )
    session.add(
        UseLine(
            id=uuid.uuid4(),
            project_id=project_id,
            label="Hard Costs",
            phase="construction",
            amount=Decimal("1200000"),
            timing_type="spread",
        )
    )

    session.add(
        CapitalModule(
            id=uuid.uuid4(),
            scenario_id=deal_model.id,
            label="LP Equity",
            vehicle_type="equity",
            equity_role="lp",
            stack_position=0,
            source={"amount": 800000},
        )
    )
    session.add(
        CapitalModule(
            id=uuid.uuid4(),
            scenario_id=deal_model.id,
            label="Senior Construction Loan",
            vehicle_type="debt",
            stack_position=1,
            source={"amount": 1200000, "interest_rate_pct": 8.0},
        )
    )
    await session.flush()
    return deal_model


async def _export(session: AsyncSession, scenario_id, profile: str):
    blob = await export_investor_workbook(scenario_id, session, profile=profile)
    return load_workbook(BytesIO(blob), data_only=False)


# ── Profile gating ────────────────────────────────────────────────────────────


async def test_draw_schedule_present_for_internal_and_lender(session: AsyncSession):
    scenario = await _seed_scenario(session)
    for profile in ("internal", "lender"):
        wb = await _export(session, scenario.id, profile)
        assert _SHEET in wb.sheetnames, (
            f"{profile} profile missing {_SHEET}; got {wb.sheetnames}"
        )
        # Sheet order: Draw Schedule immediately follows Debt Schedule.
        names = wb.sheetnames
        assert names.index(_SHEET) == names.index("Debt Schedule") + 1


async def test_draw_schedule_absent_for_lp_and_proforma(session: AsyncSession):
    scenario = await _seed_scenario(session)
    for profile in ("lp", "proforma"):
        wb = await _export(session, scenario.id, profile)
        assert _SHEET not in wb.sheetnames, (
            f"{profile} profile should not render {_SHEET}; got {wb.sheetnames}"
        )


# ── Empty state ───────────────────────────────────────────────────────────────


async def test_draw_schedule_skipped_when_no_draw_sources(session: AsyncSession):
    """No CapitalModules → reconciliation seeds no DrawSources → sheet skipped."""
    org, user = await seed_org(session)
    opportunity = await seed_opportunity(session, org, user)
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opportunity, user
    )
    wb = await _export(session, deal_model.id, "internal")
    assert _SHEET not in wb.sheetnames


async def test_draw_schedule_roster_fallback_without_milestones(session: AsyncSession):
    """Sources exist but no dated milestones → roster + hint, no matrix."""
    scenario = await _seed_scenario(session, with_milestones=False)
    wb = await _export(session, scenario.id, "lender")
    assert _SHEET in wb.sheetnames
    ws = wb[_SHEET]
    assert ws.cell(row=1, column=1).value == _ROSTER_TITLE
    # Both capital-stack sources appear in the roster.
    labels = {ws.cell(row=r, column=2).value for r in range(3, 6)}
    assert "LP Equity" in labels
    assert "Senior Construction Loan" in labels
    # Hint about the missing matrix is present somewhere below the roster.
    hints = [
        ws.cell(row=r, column=1).value
        for r in range(3, 12)
        if isinstance(ws.cell(row=r, column=1).value, str)
    ]
    assert any("monthly draw matrix" in h for h in hints)


# ── Row math ──────────────────────────────────────────────────────────────────


def _parse_matrix(ws):
    """Return (per-source matrix sums by label, summary total-drawn by label)."""
    header = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    assert "Total Drawn" in header, f"matrix header missing Total Drawn: {header}"
    total_col = header.index("Total Drawn") + 1
    month_cols = list(range(4, total_col))
    assert month_cols, "matrix has no month columns"

    matrix_sums: dict[str, Decimal] = {}
    r = 3
    while ws.cell(row=r, column=1).value != "Total":
        label = ws.cell(row=r, column=2).value
        assert label, f"unexpected blank source label at row {r}"
        cells = [ws.cell(row=r, column=c).value for c in month_cols]
        matrix_sums[label] = sum(
            (Decimal(str(v)) for v in cells if isinstance(v, (int, float, Decimal))),
            Decimal(0),
        )
        r += 1
        assert r < 100, "ran off the matrix without finding the Total row"

    # Locate the Source Summary block and read its static Total Drawn col.
    summary_sums: dict[str, Decimal] = {}
    summary_header_row = None
    for row in range(r, r + 20):
        if ws.cell(row=row, column=1).value == "Source Summary":
            summary_header_row = row + 1
            break
    assert summary_header_row is not None, "Source Summary section not found"
    row = summary_header_row + 1
    while ws.cell(row=row, column=1).value:
        label = ws.cell(row=row, column=1).value
        drawn = ws.cell(row=row, column=6).value
        summary_sums[label] = Decimal(str(drawn or 0))
        row += 1
    return matrix_sums, summary_sums


async def test_draw_schedule_row_math(session: AsyncSession):
    scenario = await _seed_scenario(session)
    wb = await _export(session, scenario.id, "internal")
    ws = wb[_SHEET]
    assert ws.cell(row=1, column=1).value == _MATRIX_TITLE

    matrix_sums, summary_sums = _parse_matrix(ws)
    assert matrix_sums, "no source rows in the funding matrix"
    assert set(matrix_sums) == set(summary_sums)

    # Each matrix row's monthly cells sum to the engine's per-source total.
    for label, total in matrix_sums.items():
        assert abs(total - summary_sums[label]) < Decimal("0.01"), (
            f"{label}: matrix sum {total} != engine total {summary_sums[label]}"
        )

    # Grand total covers Uses (equity 800k + auto-sized debt >= 1.2M + carry).
    grand_total = sum(matrix_sums.values(), Decimal(0))
    assert grand_total >= _TOTAL_USES - Decimal("0.01"), (
        f"grand total {grand_total} < seeded uses {_TOTAL_USES}"
    )
