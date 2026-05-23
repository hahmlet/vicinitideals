"""Graceful degradation: Assumptions Block A includes an editable
``Anchor Date (Y0 as-of)`` cell named ``s_anchor_date`` so the LP can
overlay their reporting calendar on the relative Y0/Y1/Y2 grid.

Contract:

  1. ``s_anchor_date`` defined name exists on every profile that ships
     the Assumptions sheet.
  2. Defaulted to scenario.created_at (ISO date string).
"""
from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.exporters.investor_export import export_investor_workbook
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


async def _seed(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Anchor Date Smoke")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )
    return deal_model


@pytest.mark.parametrize("profile", ["internal", "lp", "lender", "proforma"])
async def test_anchor_date_named_range_registered(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "s_anchor_date" in wb.defined_names, (
        f"profile={profile} missing s_anchor_date defined name"
    )


@pytest.mark.parametrize("profile", ["internal", "lp", "lender", "proforma"])
async def test_anchor_date_default_is_iso_date_string(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Assumptions"]
    # Walk Block A for the labeled row
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and "anchor date" in v.lower():
            value = ws.cell(row=r, column=2).value
            assert isinstance(value, str), (
                f"profile={profile}: anchor date should be ISO string; got {value!r}"
            )
            # ISO date matches YYYY-MM-DD
            assert len(value) == 10 and value[4] == "-" and value[7] == "-", (
                f"profile={profile}: not an ISO date: {value!r}"
            )
            return
    pytest.fail(f"profile={profile}: Anchor Date row not found on Assumptions")
