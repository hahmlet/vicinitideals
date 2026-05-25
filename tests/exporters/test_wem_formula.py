"""Phase 4 KPI-tail: Weighted Equity Multiple becomes a formula.

WEM = (PV of equity distributions @ hurdle rate) / equity_required.

Excel formula uses SUMPRODUCT over ``r_uw_cf_levered`` (annual levered
net cash flow on the Underwriting Cash Flow sheet, which is the
net-to-equity series after debt service), discounted by
``(1+s_discount_rate)^year_offset`` where ``year_offset`` =
``COLUMN(r_uw_cf_levered) - MIN(COLUMN(r_uw_cf_levered))``.

Parity vs engine: engine uses monthly-period discounting against
equity-tier-only waterfall distributions; Excel formula uses annual
periods and the aggregate levered CF. Same approximation envelope as
the Combined Equity Multiple formula already shipped on this row
group. Numeric parity is covered by the LibreOffice-recalc gate; this
file verifies formula shape + named-cell wiring only.
"""
from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


pytestmark = pytest.mark.asyncio


async def _seed(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="WEM Formula")
    deal_model, _, stream, opex = await seed_deal_model_with_financials(
        session, opp, user
    )
    stream.active_in_phases = ["stabilized"]
    opex.active_in_phases = ["stabilized"]
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find_row_prefix(ws, label_prefix: str) -> int | None:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip().startswith(label_prefix):
            return r
    return None


async def test_wem_is_sumproduct_formula(session: AsyncSession) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Summary"]
    row = _find_row_prefix(ws, "Weighted Equity Multiple")
    assert row is not None, "WEM row missing on UW Summary"
    v = ws.cell(row=row, column=2).value
    assert isinstance(v, str) and v.startswith("=IFERROR(SUMPRODUCT("), f"got {v!r}"


async def test_wem_formula_references_required_named_ranges(
    session: AsyncSession,
) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Summary"]
    row = _find_row_prefix(ws, "Weighted Equity Multiple")
    v = ws.cell(row=row, column=2).value
    for ref in ("r_uw_cf_levered", "s_discount_rate", "s_equity_required"):
        assert ref in v, f"WEM formula must reference {ref}; got {v!r}"


async def test_wem_named_range_registered(session: AsyncSession) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    names = set(wb.defined_names)
    assert "s_weighted_equity_multiple" in names
    # All three upstream operands must be registered for the formula
    # to resolve at open-time (not #NAME?).
    assert "r_uw_cf_levered" in names
    assert "s_discount_rate" in names
    assert "s_equity_required" in names


@pytest.mark.parametrize("profile", ["internal", "lp", "lender"])
async def test_wem_formula_resolves_on_profile(
    session: AsyncSession, profile: str,
) -> None:
    """Every profile that renders UW Summary must also register the
    three operand named ranges (Underwriting Cash Flow + Assumptions
    Block A + the Equity Required cell on UW Summary itself)."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    if "Underwriting Summary" not in wb.sheetnames:
        pytest.skip(f"profile {profile!r} doesn't render UW Summary")
    names = set(wb.defined_names)
    assert "r_uw_cf_levered" in names
    assert "s_discount_rate" in names
    assert "s_equity_required" in names
    assert "s_weighted_equity_multiple" in names
