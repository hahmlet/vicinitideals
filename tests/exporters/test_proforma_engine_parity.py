"""Engine-vs-Excel parity on the Underwriting Pro Forma Y1 Gross Revenue.

Catches a class of bugs where an engine field has multiple input paths
(``unit_count × amount_per_unit_monthly`` OR a flat
``amount_fixed_monthly``) but the exporter Assumptions block only
writes one path. The Excel formula references the named cells we did
write — which compute to 0 for streams using the other path — and the
Pro Forma Y1 number ends up wrong while every formula-structure test
still passes.

This test seeds a scenario whose revenue mixes both paths, recalcs the
exported workbook, and asserts the Pro Forma Y1 Gross Revenue cell
equals the expected ``Σ(stream stabilized monthly) × 12`` within $1.

We compare to the stabilized × 12 input total (not the engine's
``cash_flows[year=1].gross_revenue`` aggregate), because the engine
buckets revenue by calendar year and Y1 typically includes only the
operation-phase months of that year — so engine[Y1] is a partial-year
number while Excel Y1 is "stabilized monthly × 12" by design. They
diverge for any scenario with pre-op months, even though both are
correct under their own convention. Stabilized × 12 is the right
reference for this Y1 cell.
"""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from app.models.deal import IncomeStream, IncomeStreamType
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)
from tests.exporters._parity_helpers import (
    RecalcUnavailableError,
    recalc_workbook,
)


pytestmark = pytest.mark.asyncio


def _find_row(ws, label_text: str) -> int | None:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip() == label_text:
            return r
    return None


def _expected_y1_gross_revenue(streams: list[IncomeStream]) -> Decimal:
    """Mirror of engine ``_stream_base_amount`` for the Y1 stabilized
    annualization: prefer ``amount_fixed_monthly`` when set, else
    ``unit_count * amount_per_unit_monthly``."""
    total = Decimal(0)
    for s in streams:
        if s.amount_fixed_monthly is not None:
            total += Decimal(s.amount_fixed_monthly)
        else:
            count = Decimal(s.unit_count if s.unit_count is not None else 1)
            rent = Decimal(s.amount_per_unit_monthly or 0)
            total += count * rent
    return total * Decimal(12)


async def _seed_mixed_revenue_streams(session: AsyncSession):
    """Seed: residential rent (unit × per-unit) + fixed-amount stream
    (water reimbursement) on the same project so engine gross_revenue
    blends both paths."""
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="PF Parity Mixed")
    deal_model, _inputs, seeded_stream, seeded_opex = (
        await seed_deal_model_with_financials(session, opp, user)
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()

    seeded_stream.active_in_phases = ["stabilized"]
    seeded_opex.active_in_phases = ["stabilized"]

    fixed_stream = IncomeStream(
        id=uuid4(),
        project_id=project.id,
        stream_type=IncomeStreamType.utility_water,
        label="Water Reimbursement",
        unit_count=None,
        amount_per_unit_monthly=None,
        amount_fixed_monthly=Decimal("2323.000000"),
        stabilized_occupancy_pct=Decimal("100"),
        escalation_rate_pct_annual=Decimal("3.0"),
        active_in_phases=["stabilized"],
    )
    session.add(fixed_stream)
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()

    streams = list(
        (
            await session.execute(
                select(IncomeStream).where(IncomeStream.project_id == project.id)
            )
        ).scalars().all()
    )
    return deal_model, streams


async def test_y1_gross_revenue_excel_matches_stabilized_annualized(
    session: AsyncSession, tmp_path: Path
) -> None:
    """Pro Forma Y1 Gross Revenue (post-recalc) must equal
    ``Σ(stream stabilized monthly) × 12`` within $1.

    Before the exporter learned to read ``amount_fixed_monthly``, the
    fixed-amount stream's named cell was 0, so Excel under-reported Y1
    Gross Revenue by ``amount_fixed_monthly × 12``."""
    _scenario, streams = await _seed_mixed_revenue_streams(session)
    expected = float(_expected_y1_gross_revenue(streams))
    assert expected > 0

    blob = await export_investor_workbook(_scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)
    try:
        recalc_workbook(path)
    except RecalcUnavailableError as exc:
        pytest.skip(f"no recalc backend: {exc}")

    wb = load_workbook(path, data_only=True)
    ws = wb["Underwriting Pro Forma"]
    gr_row = _find_row(ws, "Gross Revenue")
    assert gr_row is not None
    excel_y1 = ws.cell(row=gr_row, column=3).value
    assert isinstance(excel_y1, (int, float)), (
        f"Excel Y1 Gross Revenue not numeric after recalc: {excel_y1!r}"
    )

    diff = abs(float(excel_y1) - expected)
    assert diff < 1.0, (
        f"Y1 Gross Revenue parity broken: expected={expected}, "
        f"excel={excel_y1}, diff={diff}. "
        f"Likely cause: an income-stream input path "
        f"(`amount_fixed_monthly` vs `unit_count × amount_per_unit_monthly`) "
        f"is not being read into the Assumptions Block F Y1-monthly cell."
    )
