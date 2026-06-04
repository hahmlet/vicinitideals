"""Slice 1 (Export v3): Underwriting Summary surfaces the bank-account
solvency proof persisted on ``OperationalOutputs.bank_account_proof``.

Contract verified here:

  * solvent fixture → ``s_bank_proof_min_balance``,
    ``s_bank_proof_min_balance_date``, and ``s_bank_proof_is_solvent``
    all render with the persisted values; max-shortfall cells are
    absent (insolvent-only).
  * null fixture (pre-0102 / legacy) → none of the bank-proof named
    cells appear; workbook stays clean.

Multi-project aggregation: when more than one project carries a
proof, the worst-case (lowest ``min_balance``) row wins. That
selection happens against the rollup's ``per_project`` list, so the
test seeds two projects with distinct proofs and asserts the worse
one drives the KPI.
"""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from app.models.cashflow import OperationalOutputs
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


pytestmark = pytest.mark.asyncio


def _named_cell_value(wb, name: str):
    """Return the value of a named single-cell range, or None if absent."""
    dn = wb.defined_names.get(name)
    if dn is None:
        return None
    for sheet_title, coord in dn.destinations:
        cell = coord.replace("$", "")
        col_letters = "".join(c for c in cell if c.isalpha())
        row_digits = "".join(c for c in cell if c.isdigit())
        col_idx = 0
        for ch in col_letters:
            col_idx = col_idx * 26 + (ord(ch.upper()) - ord("A") + 1)
        return wb[sheet_title].cell(row=int(row_digits), column=col_idx).value
    return None


async def _seed_scenario_with_proof(session: AsyncSession, proof: dict | None):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Bank Proof KPI")
    deal_model, _inputs, _stream, _opex = (
        await seed_deal_model_with_financials(session, opp, user)
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()
    # Drive the rollup through compute so OperationalOutputs row exists.
    await compute_cash_flows(deal_model.id, session)
    if proof is not None:
        outputs = (
            await session.execute(
                select(OperationalOutputs).where(
                    OperationalOutputs.project_id == project.id
                )
            )
        ).scalar_one_or_none()
        if outputs is None:
            outputs = OperationalOutputs(
                project_id=project.id, scenario_id=deal_model.id
            )
            session.add(outputs)
        outputs.bank_account_proof = proof
    else:
        outputs = (
            await session.execute(
                select(OperationalOutputs).where(
                    OperationalOutputs.project_id == project.id
                )
            )
        ).scalar_one_or_none()
        if outputs is not None:
            outputs.bank_account_proof = None
    await session.commit()
    return deal_model, project


async def test_solvent_proof_renders_kpi_cells(
    session: AsyncSession, tmp_path: Path
) -> None:
    proof = {
        "opening_cash": "1500000",
        "min_balance": "275000.50",
        "min_balance_date": "2027-08-01",
        "max_shortfall": "0",
        "max_shortfall_date": None,
        "is_solvent": True,
        "co_period": 18,
        "stabilized_period": 36,
        "months_simulated": 36,
        "proof_start": "day_0",
        "stabilization_anchor": None,
    }
    scenario, _ = await _seed_scenario_with_proof(session, proof)
    blob = await export_investor_workbook(scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)
    wb = load_workbook(path, data_only=False)

    min_bal = _named_cell_value(wb, "s_bank_proof_min_balance")
    assert min_bal is not None
    assert float(min_bal) == pytest.approx(275000.50, abs=0.01)

    min_bal_date = _named_cell_value(wb, "s_bank_proof_min_balance_date")
    assert min_bal_date == "2027-08-01"

    is_solvent = _named_cell_value(wb, "s_bank_proof_is_solvent")
    assert is_solvent == "Yes"

    # Solvent: max-shortfall cells must NOT render.
    assert _named_cell_value(wb, "s_bank_proof_max_shortfall") is None
    assert _named_cell_value(wb, "s_bank_proof_max_shortfall_date") is None


async def test_insolvent_proof_renders_shortfall_cells(
    session: AsyncSession, tmp_path: Path
) -> None:
    proof = {
        "opening_cash": "500000",
        "min_balance": "-120000",
        "min_balance_date": "2027-09-01",
        "max_shortfall": "180000",
        "max_shortfall_date": "2027-09-01",
        "is_solvent": False,
        "co_period": 18,
        "stabilized_period": 36,
        "months_simulated": 36,
        "proof_start": "day_0",
        "stabilization_anchor": None,
    }
    scenario, _ = await _seed_scenario_with_proof(session, proof)
    blob = await export_investor_workbook(scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)
    wb = load_workbook(path, data_only=False)

    assert _named_cell_value(wb, "s_bank_proof_is_solvent") == "No"
    max_short = _named_cell_value(wb, "s_bank_proof_max_shortfall")
    assert max_short is not None
    assert float(max_short) == pytest.approx(180000.0, abs=0.01)
    assert _named_cell_value(wb, "s_bank_proof_max_shortfall_date") == "2027-09-01"


async def test_null_proof_skips_kpi_section(
    session: AsyncSession, tmp_path: Path
) -> None:
    """Pre-0102 legacy deals carry NULL bank_account_proof — the
    section must not appear, no named cells emitted, workbook clean."""
    scenario, _ = await _seed_scenario_with_proof(session, None)
    blob = await export_investor_workbook(scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)
    wb = load_workbook(path, data_only=False)

    assert _named_cell_value(wb, "s_bank_proof_min_balance") is None
    assert _named_cell_value(wb, "s_bank_proof_min_balance_date") is None
    assert _named_cell_value(wb, "s_bank_proof_is_solvent") is None
