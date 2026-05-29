"""Assumptions sheet Blocks F (Revenue inputs) + G (OpEx inputs).

These blocks are the foundation of the engine-to-formula migration: every
revenue stream and every operating expense line in the scenario lands on
the Assumptions sheet as a named input cell, so downstream Pro Forma /
Cash Flow / Unit Mix / UW Summary cells can reference them via
``=s_rev_<slug>_*`` or ``=s_opex_<slug>_*`` instead of hardcoding.

Contract:

  1. One row per IncomeStream — registers ``s_rev_<slug>_unit_count``,
     ``s_rev_<slug>_rent_per_unit_monthly``, ``s_rev_<slug>_occupancy_pct``,
     ``s_rev_<slug>_escalation_pct``, and the computed
     ``s_rev_<slug>_y1_monthly``.
  2. One row per OperatingExpenseLine — registers ``s_opex_<slug>_annual``,
     ``s_opex_<slug>_escalation_pct``, ``s_opex_<slug>_monthly``.
  3. Slug collisions resolve with ``_2`` / ``_3`` suffixes so two streams
     (or two lines) with the same label produce distinct named cells.
  4. The blocks render on every workbook profile that builds the
     Assumptions sheet (internal / lp / lender / proforma).
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import (
    _opex_slugs,
    _slugify_simple,
    _stream_slugs,
    export_investor_workbook,
)
from app.models.deal import IncomeStream, IncomeStreamType, OperatingExpenseLine
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


pytestmark = pytest.mark.asyncio


def test_slugify_simple_lowercases_strips_punctuation():
    assert _slugify_simple("1BR Units") == "1br_units"
    assert _slugify_simple("Property Mgmt!") == "property_mgmt"
    assert _slugify_simple("  -- Foo Bar -- ") == "foo_bar"
    assert _slugify_simple(None) == ""
    assert _slugify_simple("") == ""


def test_stream_slugs_dedupe_collisions():
    class _S:
        def __init__(self, id_, label):
            self.id = id_
            self.label = label

    s1, s2, s3 = _S("a", "1BR Units"), _S("b", "1BR Units"), _S("c", None)
    slugs = _stream_slugs([s1, s2, s3])
    assert slugs == {"a": "1br_units", "b": "1br_units_2", "c": "stream_3"}


def test_opex_slugs_dedupe_collisions():
    class _L:
        def __init__(self, id_, label):
            self.id = id_
            self.label = label

    l1, l2, l3 = _L("a", "Property Mgmt"), _L("b", "Property Mgmt"), _L("c", "")
    slugs = _opex_slugs([l1, l2, l3])
    assert slugs == {"a": "property_mgmt", "b": "property_mgmt_2", "c": "opex_3"}


async def _seed_with_extra_streams_and_opex(session: AsyncSession):
    """Seed minimal scenario, then add two extra income streams + two
    extra OpEx lines so we can verify per-row cell registration."""
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Rev/OpEx Smoke")
    deal_model, _inputs, _income, _opex = await seed_deal_model_with_financials(
        session, opp, user
    )
    primary = (await session.execute(
        select(Project).where(
            Project.scenario_id == deal_model.id,
            Project.name == "Main Project",
        )
    )).scalar_one()

    # Two extra streams (one collides with the seeded label, one unique)
    session.add_all([
        IncomeStream(
            id=uuid4(), project_id=primary.id,
            stream_type=IncomeStreamType.residential_rent,
            label="1BR Units",  # collides with seeded "1BR Units"
            unit_count=4,
            amount_per_unit_monthly=Decimal("1500"),
            stabilized_occupancy_pct=Decimal("95"),
            escalation_rate_pct_annual=Decimal("3.0"),
        ),
        IncomeStream(
            id=uuid4(), project_id=primary.id,
            stream_type=IncomeStreamType.residential_rent,
            label="2BR Units",
            unit_count=6,
            amount_per_unit_monthly=Decimal("1800"),
            stabilized_occupancy_pct=Decimal("95"),
            escalation_rate_pct_annual=Decimal("3.0"),
        ),
    ])
    # Two extra OpEx lines (one collides, one unique)
    session.add_all([
        OperatingExpenseLine(
            id=uuid4(), project_id=primary.id,
            label="Property Management",  # collides with seeded line
            annual_amount=Decimal("5000"),
            escalation_rate_pct_annual=Decimal("3.0"),
        ),
        OperatingExpenseLine(
            id=uuid4(), project_id=primary.id,
            label="Insurance",
            annual_amount=Decimal("12000"),
            escalation_rate_pct_annual=Decimal("4.0"),
        ),
    ])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


async def test_revenue_block_registers_one_cell_set_per_stream(
    session: AsyncSession,
) -> None:
    scenario = await _seed_with_extra_streams_and_opex(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    defined = {name for name in wb.defined_names}

    # Seed adds one "1BR Units" stream; we add another → slug collision
    # resolves with _2. Plus "2BR Units" unique.
    required = {
        "s_rev_1br_units_unit_count",
        "s_rev_1br_units_rent_per_unit_monthly",
        "s_rev_1br_units_occupancy_pct",
        "s_rev_1br_units_escalation_pct",
        "s_rev_1br_units_y1_monthly",
        "s_rev_1br_units_2_unit_count",
        "s_rev_2br_units_unit_count",
    }
    missing = required - defined
    assert not missing, f"revenue block missing named cells: {sorted(missing)}"


async def test_revenue_block_y1_monthly_is_formula_referencing_inputs(
    session: AsyncSession,
) -> None:
    """The Y1 monthly cell must be a formula consuming the count and
    rent input cells — true gross potential rent (pre-vacancy). Edit
    count/rent on Assumptions, see Y1 monthly update in Excel.

    Note: occupancy is intentionally NOT referenced. It applies on the
    Pro Forma via the Vacancy Loss row so EGI = Gross − Vacancy matches
    the engine's `gross_revenue` field (pre-vacancy by convention).
    """
    scenario = await _seed_with_extra_streams_and_opex(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    dn = wb.defined_names.get("s_rev_1br_units_y1_monthly")
    assert dn is not None
    for sheet_name, cell_ref in dn.destinations:
        formula = wb[sheet_name][cell_ref].value
        assert isinstance(formula, str) and formula.startswith("=")
        assert "s_rev_1br_units_unit_count" in formula
        assert "s_rev_1br_units_rent_per_unit_monthly" in formula
        assert "occupancy_pct" not in formula, (
            "Y1 monthly must be pre-vacancy gross; occupancy belongs "
            "on the Pro Forma Vacancy Loss row"
        )
        return
    pytest.fail("s_rev_1br_units_y1_monthly had no destination")


async def test_opex_block_registers_one_cell_set_per_line(
    session: AsyncSession,
) -> None:
    scenario = await _seed_with_extra_streams_and_opex(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    defined = {name for name in wb.defined_names}

    required = {
        "s_opex_property_management_annual",
        "s_opex_property_management_escalation_pct",
        "s_opex_property_management_monthly",
        "s_opex_property_management_2_annual",
        "s_opex_insurance_annual",
        "s_opex_insurance_escalation_pct",
    }
    missing = required - defined
    assert not missing, f"opex block missing named cells: {sorted(missing)}"


async def test_opex_block_monthly_is_formula_dividing_annual(
    session: AsyncSession,
) -> None:
    scenario = await _seed_with_extra_streams_and_opex(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    dn = wb.defined_names.get("s_opex_insurance_monthly")
    assert dn is not None
    for sheet_name, cell_ref in dn.destinations:
        formula = wb[sheet_name][cell_ref].value
        assert isinstance(formula, str) and formula.startswith("=")
        assert "s_opex_insurance_annual" in formula
        assert "/12" in formula
        return
    pytest.fail("s_opex_insurance_monthly had no destination")


async def test_blocks_land_on_assumptions_sheet(
    session: AsyncSession,
) -> None:
    """Both blocks must live on the Assumptions sheet — that's the
    single source of truth contract."""
    scenario = await _seed_with_extra_streams_and_opex(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    for name in ("s_rev_1br_units_unit_count", "s_opex_insurance_annual"):
        dn = wb.defined_names.get(name)
        assert dn is not None, f"{name} not registered"
        sheets = {sheet for sheet, _ in dn.destinations}
        assert "Assumptions" in sheets, (
            f"expected {name} on Assumptions, got {sheets}"
        )
