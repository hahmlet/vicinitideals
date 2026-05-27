"""Phase 5d+5e+5f: LP/GP Distribution Totals + Returns Summary + CF Series.

Contract:
  Phase 5d —
  1. Investor Returns sheet writes s_lp_distributions_total and
     s_gp_distributions_total as IFERROR-wrapped SUM formulas:
       s_lp_distributions_total =IFERROR(
           IFERROR(s_waterfall_tier_1_lp_amt,0)+...+IFERROR(s_waterfall_tier_7_lp_amt,0),
           {engine_fallback})
     and mirror for GP.
  2. Formulas reference all 7 canonical tier LP/GP named cells, so Slice 5c's
     per-tier formulas propagate up to the total without re-running the engine.

  Phase 5e —
  3. Investor Returns sheet writes:
       s_committed_lp_equity   — scalar: sum of LP module committed principals
       s_committed_gp_equity   — scalar: sum of GP module committed principals
       s_lp_em                 — formula: =IFERROR(s_lp_distributions_total/s_committed_lp_equity,…)
       s_gp_em                 — formula: =IFERROR(s_gp_distributions_total/s_committed_gp_equity,…)
  4. LP/GP EM formulas reference s_lp_distributions_total and s_committed_lp_equity,
     so editing split inputs + F9 reflows EM cells.

  Phase 5f —
  5. Investor Returns sheet emits annual LP/GP cash flow rows:
       s_returns_lp_y0 = −committed_lp (initial investment, negative)
       s_returns_lp_y1..yN = waterfall distributions bucketed by year
       r_returns_lp_cf = range covering all LP CF cells (Y0..YN)
       (mirror for GP: s_returns_gp_*, r_returns_gp_cf)
  6. s_lp_irr / s_gp_irr converted to live formula:
       =IFERROR(IRR(r_returns_lp_cf), {engine_fallback})
  7. s_lp_coc_y1 / s_gp_coc_y1 added as live formula:
       =IFERROR(s_returns_lp_y1/s_committed_lp_equity, {fallback})
  8. For unconfigured scenarios (no WaterfallTier / no LP-GP equity modules),
     all named cells are still registered (formula or em-dash) so cross-sheet
     refs resolve without #NAME?.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.exporters.investor_export import export_investor_workbook
from app.models.capital import CapitalModule, WaterfallTier, WaterfallTierType
import uuid
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)

pytestmark = pytest.mark.asyncio

_MAX_CANONICAL_TIERS = 7


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _defined_names(wb) -> set[str]:
    return {dn for dn in wb.defined_names}


def _get_named_value(wb, name: str):
    """Return the cell value for a workbook-scoped defined name."""
    try:
        defn = wb.defined_names[name]
    except KeyError:
        return None
    for _title, coord in defn.destinations:
        ws = wb[_title]
        return ws[coord].value
    return None


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

async def _seed_with_lp_gp_modules(session: AsyncSession):
    """Seed a scenario with LP equity + GP equity modules and two WaterfallTiers."""
    from app.engines.cashflow import compute_cash_flows

    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Phase5d LP GP Returns Test")
    deal_model, _inputs, stream, opex = await seed_deal_model_with_financials(
        session, opp, user
    )
    stream.active_in_phases = ["stabilized"]
    opex.active_in_phases = ["stabilized"]

    # LP equity module
    lp_module = CapitalModule(
        id=uuid.uuid4(),
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type="equity",
        equity_role="lp",
        stack_position=1,
        source={"amount": 1_500_000},
        carry={"carry_type": "none", "payment_frequency": "at_exit"},
        exit_terms={"exit_type": "profit_share", "trigger": "sale", "profit_share_pct": 100},
        active_phase_start="acquisition",
        active_phase_end="exit",
    )
    # GP equity module
    gp_module = CapitalModule(
        id=uuid.uuid4(),
        scenario_id=deal_model.id,
        label="GP Equity",
        vehicle_type="equity",
        equity_role="gp",
        stack_position=2,
        source={"amount": 300_000},
        carry={"carry_type": "none", "payment_frequency": "at_exit"},
        exit_terms={"exit_type": "profit_share", "trigger": "sale", "profit_share_pct": 100},
        active_phase_start="acquisition",
        active_phase_end="exit",
    )
    session.add_all([lp_module, gp_module])
    await session.flush()

    # Pref return (LP=100%, GP=0%) + Residual (LP=70%, GP=30%)
    tier1 = WaterfallTier(
        scenario_id=deal_model.id,
        capital_module_id=lp_module.id,
        priority=1,
        tier_type=WaterfallTierType.pref_return,
        irr_hurdle_pct=Decimal("8"),
        lp_split_pct=Decimal("100"),
        gp_split_pct=Decimal("0"),
        description="Pref return LP 100%",
    )
    tier2 = WaterfallTier(
        scenario_id=deal_model.id,
        capital_module_id=lp_module.id,
        priority=2,
        tier_type=WaterfallTierType.residual,
        irr_hurdle_pct=None,
        lp_split_pct=Decimal("70"),
        gp_split_pct=Decimal("30"),
        description="Residual 70/30",
    )
    session.add_all([tier1, tier2])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


async def _seed_no_equity(session: AsyncSession):
    """Seed a scenario with no equity modules and no WaterfallTier rows."""
    from app.engines.cashflow import compute_cash_flows

    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Phase5d No-Equity Test")
    deal_model, _cap, stream, opex = await seed_deal_model_with_financials(
        session, opp, user
    )
    stream.active_in_phases = ["stabilized"]
    opex.active_in_phases = ["stabilized"]
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


# ---------------------------------------------------------------------------
# Phase 5d: Distribution Total named cells
# ---------------------------------------------------------------------------

async def test_lp_distributions_total_registered(session: AsyncSession) -> None:
    """s_lp_distributions_total named cell exists on Investor Returns."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "s_lp_distributions_total" in _defined_names(wb)


async def test_gp_distributions_total_registered(session: AsyncSession) -> None:
    """s_gp_distributions_total named cell exists on Investor Returns."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "s_gp_distributions_total" in _defined_names(wb)


async def test_lp_distributions_total_is_iferror_formula(session: AsyncSession) -> None:
    """s_lp_distributions_total must be an IFERROR formula, not a scalar."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    val = _get_named_value(wb, "s_lp_distributions_total")
    assert isinstance(val, str) and val.startswith("=IFERROR("), (
        f"s_lp_distributions_total must be IFERROR formula; got {val!r}"
    )


async def test_gp_distributions_total_is_iferror_formula(session: AsyncSession) -> None:
    """s_gp_distributions_total must be an IFERROR formula, not a scalar."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    val = _get_named_value(wb, "s_gp_distributions_total")
    assert isinstance(val, str) and val.startswith("=IFERROR("), (
        f"s_gp_distributions_total must be IFERROR formula; got {val!r}"
    )


async def test_lp_distributions_formula_references_all_canonical_tiers(
    session: AsyncSession,
) -> None:
    """LP distributions formula references all 7 canonical tier LP amt cells."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    val = _get_named_value(wb, "s_lp_distributions_total")
    assert isinstance(val, str)
    for i in range(1, _MAX_CANONICAL_TIERS + 1):
        assert f"s_waterfall_tier_{i}_lp_amt" in val, (
            f"LP distributions formula must reference s_waterfall_tier_{i}_lp_amt; got {val!r}"
        )


async def test_gp_distributions_formula_references_all_canonical_tiers(
    session: AsyncSession,
) -> None:
    """GP distributions formula references all 7 canonical tier GP amt cells."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    val = _get_named_value(wb, "s_gp_distributions_total")
    assert isinstance(val, str)
    for i in range(1, _MAX_CANONICAL_TIERS + 1):
        assert f"s_waterfall_tier_{i}_gp_amt" in val, (
            f"GP distributions formula must reference s_waterfall_tier_{i}_gp_amt; got {val!r}"
        )


async def test_lp_gp_distributions_total_registered_for_no_equity_scenario(
    session: AsyncSession,
) -> None:
    """Unconfigured scenario: both distribution total cells still registered."""
    scenario = await _seed_no_equity(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    names = _defined_names(wb)
    assert "s_lp_distributions_total" in names
    assert "s_gp_distributions_total" in names


# ---------------------------------------------------------------------------
# Phase 5e: LP / GP Returns Summary named cells
# ---------------------------------------------------------------------------

async def test_lp_em_named_cell_registered(session: AsyncSession) -> None:
    """s_lp_em named cell exists when LP equity present."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "s_lp_em" in _defined_names(wb)


async def test_gp_em_named_cell_registered(session: AsyncSession) -> None:
    """s_gp_em named cell exists when GP equity present."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "s_gp_em" in _defined_names(wb)


async def test_lp_em_is_iferror_formula(session: AsyncSession) -> None:
    """s_lp_em must be a live IFERROR formula referencing distributions / equity."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    val = _get_named_value(wb, "s_lp_em")
    assert isinstance(val, str) and val.startswith("=IFERROR("), (
        f"s_lp_em must be IFERROR formula; got {val!r}"
    )
    assert "s_lp_distributions_total" in val, (
        f"s_lp_em must reference s_lp_distributions_total; got {val!r}"
    )
    assert "s_committed_lp_equity" in val, (
        f"s_lp_em must reference s_committed_lp_equity; got {val!r}"
    )


async def test_gp_em_is_iferror_formula(session: AsyncSession) -> None:
    """s_gp_em must be a live IFERROR formula referencing distributions / equity."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    val = _get_named_value(wb, "s_gp_em")
    assert isinstance(val, str) and val.startswith("=IFERROR("), (
        f"s_gp_em must be IFERROR formula; got {val!r}"
    )
    assert "s_gp_distributions_total" in val, (
        f"s_gp_em must reference s_gp_distributions_total; got {val!r}"
    )
    assert "s_committed_gp_equity" in val, (
        f"s_gp_em must reference s_committed_gp_equity; got {val!r}"
    )


async def test_lp_irr_named_cell_registered(session: AsyncSession) -> None:
    """s_lp_irr named cell exists on Investor Returns."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "s_lp_irr" in _defined_names(wb)


async def test_gp_irr_named_cell_registered(session: AsyncSession) -> None:
    """s_gp_irr named cell exists on Investor Returns."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "s_gp_irr" in _defined_names(wb)


async def test_committed_lp_equity_registered(session: AsyncSession) -> None:
    """s_committed_lp_equity named cell exists and is a numeric scalar."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "s_committed_lp_equity" in _defined_names(wb)
    val = _get_named_value(wb, "s_committed_lp_equity")
    assert isinstance(val, (int, float)), (
        f"s_committed_lp_equity must be numeric; got {val!r}"
    )
    assert val > 0, f"LP equity should be positive; got {val!r}"


async def test_committed_gp_equity_registered(session: AsyncSession) -> None:
    """s_committed_gp_equity named cell exists and is a numeric scalar."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "s_committed_gp_equity" in _defined_names(wb)
    val = _get_named_value(wb, "s_committed_gp_equity")
    assert isinstance(val, (int, float)), (
        f"s_committed_gp_equity must be numeric; got {val!r}"
    )
    assert val > 0, f"GP equity should be positive; got {val!r}"


async def test_returns_summary_cells_registered_for_no_equity_scenario(
    session: AsyncSession,
) -> None:
    """Unconfigured scenario: all LP/GP returns cells still registered (formula or em-dash)."""
    scenario = await _seed_no_equity(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    names = _defined_names(wb)
    for cell_name in (
        "s_committed_lp_equity", "s_committed_gp_equity",
        "s_lp_irr", "s_gp_irr",
        "s_lp_em", "s_gp_em",
        # Phase 5f cells — must be registered even with no equity
        "s_returns_lp_y0", "s_returns_lp_y1",
        "s_returns_gp_y0", "s_returns_gp_y1",
        "r_returns_lp_cf", "r_returns_gp_cf",
        "s_lp_coc_y1", "s_gp_coc_y1",
    ):
        assert cell_name in names, f"{cell_name} missing in unconfigured scenario"


# ---------------------------------------------------------------------------
# Phase 5f: Cash Flow Series + IRR formula + CoC Year 1
# ---------------------------------------------------------------------------

async def test_lp_cf_y0_and_y1_cells_registered(session: AsyncSession) -> None:
    """s_returns_lp_y0 and s_returns_lp_y1 exist on Investor Returns."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    names = _defined_names(wb)
    assert "s_returns_lp_y0" in names, "s_returns_lp_y0 not registered"
    assert "s_returns_lp_y1" in names, "s_returns_lp_y1 not registered"


async def test_gp_cf_y0_and_y1_cells_registered(session: AsyncSession) -> None:
    """s_returns_gp_y0 and s_returns_gp_y1 exist on Investor Returns."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    names = _defined_names(wb)
    assert "s_returns_gp_y0" in names, "s_returns_gp_y0 not registered"
    assert "s_returns_gp_y1" in names, "s_returns_gp_y1 not registered"


async def test_lp_cf_range_registered(session: AsyncSession) -> None:
    """r_returns_lp_cf range name exists on Investor Returns."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "r_returns_lp_cf" in _defined_names(wb), "r_returns_lp_cf range not registered"


async def test_gp_cf_range_registered(session: AsyncSession) -> None:
    """r_returns_gp_cf range name exists on Investor Returns."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "r_returns_gp_cf" in _defined_names(wb), "r_returns_gp_cf range not registered"


async def test_lp_cf_y0_is_negative_committed_equity(session: AsyncSession) -> None:
    """s_returns_lp_y0 = −committed LP equity (initial investment, negative)."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    y0 = _get_named_value(wb, "s_returns_lp_y0")
    committed = _get_named_value(wb, "s_committed_lp_equity")
    assert isinstance(y0, (int, float)), f"s_returns_lp_y0 must be numeric; got {y0!r}"
    assert isinstance(committed, (int, float)), (
        f"s_committed_lp_equity must be numeric; got {committed!r}"
    )
    assert y0 < 0, f"LP CF Y0 must be negative (investment outflow); got {y0}"
    assert abs(y0) == pytest.approx(committed, rel=1e-4), (
        f"LP CF Y0 should equal −committed_lp_equity; Y0={y0}, committed={committed}"
    )


async def test_lp_irr_is_irr_formula(session: AsyncSession) -> None:
    """s_lp_irr must be an IFERROR(IRR(r_returns_lp_cf),…) live formula."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    val = _get_named_value(wb, "s_lp_irr")
    assert isinstance(val, str), f"s_lp_irr must be a formula string; got {val!r}"
    assert "IRR(r_returns_lp_cf)" in val, (
        f"s_lp_irr must contain IRR(r_returns_lp_cf); got {val!r}"
    )
    assert val.startswith("=IFERROR("), (
        f"s_lp_irr must start with =IFERROR(; got {val!r}"
    )


async def test_gp_irr_is_irr_formula(session: AsyncSession) -> None:
    """s_gp_irr must be an IFERROR(IRR(r_returns_gp_cf),…) live formula."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    val = _get_named_value(wb, "s_gp_irr")
    assert isinstance(val, str), f"s_gp_irr must be a formula string; got {val!r}"
    assert "IRR(r_returns_gp_cf)" in val, (
        f"s_gp_irr must contain IRR(r_returns_gp_cf); got {val!r}"
    )
    assert val.startswith("=IFERROR("), (
        f"s_gp_irr must start with =IFERROR(; got {val!r}"
    )


async def test_lp_coc_y1_registered(session: AsyncSession) -> None:
    """s_lp_coc_y1 named cell exists on Investor Returns."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "s_lp_coc_y1" in _defined_names(wb), "s_lp_coc_y1 not registered"


async def test_gp_coc_y1_registered(session: AsyncSession) -> None:
    """s_gp_coc_y1 named cell exists on Investor Returns."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "s_gp_coc_y1" in _defined_names(wb), "s_gp_coc_y1 not registered"


async def test_lp_coc_y1_is_iferror_formula(session: AsyncSession) -> None:
    """s_lp_coc_y1 must be an IFERROR formula referencing s_returns_lp_y1."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    val = _get_named_value(wb, "s_lp_coc_y1")
    assert isinstance(val, str) and val.startswith("=IFERROR("), (
        f"s_lp_coc_y1 must be IFERROR formula; got {val!r}"
    )
    assert "s_returns_lp_y1" in val, (
        f"s_lp_coc_y1 must reference s_returns_lp_y1; got {val!r}"
    )
    assert "s_committed_lp_equity" in val, (
        f"s_lp_coc_y1 must reference s_committed_lp_equity; got {val!r}"
    )


async def test_gp_coc_y1_is_iferror_formula(session: AsyncSession) -> None:
    """s_gp_coc_y1 must be an IFERROR formula referencing s_returns_gp_y1."""
    scenario = await _seed_with_lp_gp_modules(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    val = _get_named_value(wb, "s_gp_coc_y1")
    assert isinstance(val, str) and val.startswith("=IFERROR("), (
        f"s_gp_coc_y1 must be IFERROR formula; got {val!r}"
    )
    assert "s_returns_gp_y1" in val, (
        f"s_gp_coc_y1 must reference s_returns_gp_y1; got {val!r}"
    )
    assert "s_committed_gp_equity" in val, (
        f"s_gp_coc_y1 must reference s_committed_gp_equity; got {val!r}"
    )
