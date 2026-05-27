"""Phase 5c: Waterfall LP/GP Amount cells become formula-driven named cells.

Contract:
  1. For each configured WaterfallTier, Waterfall sheet writes:
       s_waterfall_tier_{n}_lp_amt  =IFERROR(s_waterfall_tier_{n}_distributed
                                             *s_tier_{n}_lp_split/100, fallback)
       s_waterfall_tier_{n}_gp_amt  =IFERROR(s_waterfall_tier_{n}_distributed
                                             *s_tier_{n}_gp_split/100, fallback)
  2. Formula references s_waterfall_tier_{n}_distributed (engine-written
     named cell for Total Distributed) and s_tier_{n}_lp_split / s_tier_{n}_gp_split
     (Assumptions Block D blue-input cells). Editing the split inputs + F9
     reflows LP/GP amount columns without re-running the engine.
  3. For unconfigured scenarios (no WaterfallTier rows), all 7 canonical
     tier slots register $0 named cells so Slice 5d's SUM formula can
     reference all tier names without #NAME? errors.
  4. Numeric parity with engine values is covered by the LibreOffice
     recalc gate — this file verifies formula shape + named-cell wiring only.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.exporters.investor_export import export_investor_workbook
from app.models.capital import CapitalModule, WaterfallTier, WaterfallTierType
import uuid
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)

pytestmark = pytest.mark.asyncio

_MAX_CANONICAL_TIERS = 7  # matches len(_CANONICAL_WATERFALL_TIERS)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _defined_names(wb) -> set[str]:
    return {dn for dn in wb.defined_names}


def _get_named_value(wb, name: str):
    """Return the cell value for a workbook-scoped defined name."""
    try:
        defn = wb.defined_names[name]
    except KeyError:
        return None
    for _title, coord in defn.destinations:
        ws = wb[_title]
        return ws[coord].value
    return None


# ---------------------------------------------------------------------------
# Fixtures / seed helpers
# ---------------------------------------------------------------------------

async def _seed_with_tiers(session: AsyncSession):
    """Seed a scenario with two configured WaterfallTier rows."""
    from app.engines.cashflow import compute_cash_flows

    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Phase5c Waterfall Formula Test")
    deal_model, _inputs, stream, opex = await seed_deal_model_with_financials(
        session, opp, user
    )
    stream.active_in_phases = ["stabilized"]
    opex.active_in_phases = ["stabilized"]

    # Capital module required as FK for WaterfallTier
    cap_module = CapitalModule(
        id=uuid.uuid4(),
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type="equity",
        stack_position=1,
        source={"amount": 1500000},
        carry={"carry_type": "none", "payment_frequency": "at_exit"},
        exit_terms={"exit_type": "profit_share", "trigger": "sale", "profit_share_pct": 100},
        active_phase_start="acquisition",
        active_phase_end="exit",
    )
    session.add(cap_module)
    await session.flush()

    # Two WaterfallTier rows: pref-return (LP=100%) + residual (LP=70%)
    tier1 = WaterfallTier(
        scenario_id=deal_model.id,
        capital_module_id=cap_module.id,
        priority=1,
        tier_type=WaterfallTierType.pref_return,
        irr_hurdle_pct=Decimal("8"),
        lp_split_pct=Decimal("100"),
        gp_split_pct=Decimal("0"),
        description="Pref return LP 100%",
    )
    tier2 = WaterfallTier(
        scenario_id=deal_model.id,
        capital_module_id=cap_module.id,
        priority=2,
        tier_type=WaterfallTierType.residual,
        irr_hurdle_pct=None,
        lp_split_pct=Decimal("70"),
        gp_split_pct=Decimal("30"),
        description="Residual 70/30",
    )
    session.add_all([tier1, tier2])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


async def _seed_no_tiers(session: AsyncSession):
    """Seed a scenario with no WaterfallTier rows (unconfigured path)."""
    from app.engines.cashflow import compute_cash_flows

    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Phase5c Unconfigured Test")
    deal_model, _cap, stream, opex = await seed_deal_model_with_financials(
        session, opp, user
    )
    stream.active_in_phases = ["stabilized"]
    opex.active_in_phases = ["stabilized"]
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


# ---------------------------------------------------------------------------
# Tests: configured tiers
# ---------------------------------------------------------------------------

async def test_lp_amt_named_cells_exist_for_configured_tiers(
    session: AsyncSession,
) -> None:
    """s_waterfall_tier_{n}_lp_amt registered for each configured tier."""
    scenario = await _seed_with_tiers(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    names = _defined_names(wb)
    assert "s_waterfall_tier_1_lp_amt" in names, "tier 1 LP amt named cell missing"
    assert "s_waterfall_tier_2_lp_amt" in names, "tier 2 LP amt named cell missing"


async def test_gp_amt_named_cells_exist_for_configured_tiers(
    session: AsyncSession,
) -> None:
    """s_waterfall_tier_{n}_gp_amt registered for each configured tier."""
    scenario = await _seed_with_tiers(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    names = _defined_names(wb)
    assert "s_waterfall_tier_1_gp_amt" in names, "tier 1 GP amt named cell missing"
    assert "s_waterfall_tier_2_gp_amt" in names, "tier 2 GP amt named cell missing"


async def test_lp_amt_formula_references_distributed_and_split(
    session: AsyncSession,
) -> None:
    """LP amount formula references both distributed total and lp_split input."""
    scenario = await _seed_with_tiers(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    for idx in (1, 2):
        val = _get_named_value(wb, f"s_waterfall_tier_{idx}_lp_amt")
        assert isinstance(val, str) and val.startswith("="), (
            f"tier {idx} LP amt must be a formula; got {val!r}"
        )
        assert f"s_waterfall_tier_{idx}_distributed" in val, (
            f"tier {idx} LP formula must reference s_waterfall_tier_{idx}_distributed; "
            f"got {val!r}"
        )
        assert f"s_tier_{idx}_lp_split" in val, (
            f"tier {idx} LP formula must reference s_tier_{idx}_lp_split; "
            f"got {val!r}"
        )


async def test_gp_amt_formula_references_distributed_and_split(
    session: AsyncSession,
) -> None:
    """GP amount formula references both distributed total and gp_split input."""
    scenario = await _seed_with_tiers(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    for idx in (1, 2):
        val = _get_named_value(wb, f"s_waterfall_tier_{idx}_gp_amt")
        assert isinstance(val, str) and val.startswith("="), (
            f"tier {idx} GP amt must be a formula; got {val!r}"
        )
        assert f"s_waterfall_tier_{idx}_distributed" in val, (
            f"tier {idx} GP formula must reference s_waterfall_tier_{idx}_distributed; "
            f"got {val!r}"
        )
        assert f"s_tier_{idx}_gp_split" in val, (
            f"tier {idx} GP formula must reference s_tier_{idx}_gp_split; "
            f"got {val!r}"
        )


async def test_lp_amt_formula_has_iferror_wrapper(
    session: AsyncSession,
) -> None:
    """LP amount formula uses IFERROR so workbook opens even with missing named ranges."""
    scenario = await _seed_with_tiers(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    val = _get_named_value(wb, "s_waterfall_tier_1_lp_amt")
    assert "IFERROR" in val.upper(), (
        f"LP amt formula must use IFERROR; got {val!r}"
    )


# ---------------------------------------------------------------------------
# Tests: unconfigured tiers (canonical structure, no WaterfallTier rows)
# ---------------------------------------------------------------------------

async def test_unconfigured_lp_amt_cells_registered_for_all_canonical_tiers(
    session: AsyncSession,
) -> None:
    """Unconfigured scenario: all 7 canonical tier LP amt cells registered as $0."""
    scenario = await _seed_no_tiers(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    names = _defined_names(wb)
    for idx in range(1, _MAX_CANONICAL_TIERS + 1):
        assert f"s_waterfall_tier_{idx}_lp_amt" in names, (
            f"Unconfigured: s_waterfall_tier_{idx}_lp_amt missing"
        )


async def test_unconfigured_gp_amt_cells_registered_for_all_canonical_tiers(
    session: AsyncSession,
) -> None:
    """Unconfigured scenario: all 7 canonical tier GP amt cells registered as $0."""
    scenario = await _seed_no_tiers(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    names = _defined_names(wb)
    for idx in range(1, _MAX_CANONICAL_TIERS + 1):
        assert f"s_waterfall_tier_{idx}_gp_amt" in names, (
            f"Unconfigured: s_waterfall_tier_{idx}_gp_amt missing"
        )


async def test_unconfigured_lp_amt_values_are_zero(
    session: AsyncSession,
) -> None:
    """Unconfigured LP amt cells contain $0, not formulas or _DASH strings."""
    scenario = await _seed_no_tiers(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    for idx in range(1, _MAX_CANONICAL_TIERS + 1):
        val = _get_named_value(wb, f"s_waterfall_tier_{idx}_lp_amt")
        assert val == 0 or val == 0.0, (
            f"Unconfigured tier {idx} LP amt should be 0; got {val!r}"
        )
