"""Smoke + parity tests for the LP investor Excel export.

Sheet roster grows commit-by-commit: commit 1 added Cover/Assumptions/
Glossary; commit 2 inserts the four Underwriting rollup sheets in the
§2 final order; commit 3 will splice per-project sheets between
Assumptions and Glossary. Tests assert exact roster + grow with the build.

Tests use the in-memory async SQLite + ``seed_deal_model_with_financials``
fixture from ``tests/conftest.py``; the exporter's data loader is exercised
end-to-end with real ORM rows.
"""
from __future__ import annotations

import re
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.exporters._doc_validator import name_lookup, parse_doc, slugs_for_name
from app.exporters.investor_export import (
    MAX_PROJECTS_PER_SCENARIO,
    PROJECT_SHEET_NAME_BUDGET,
    export_investor_workbook,
    make_investor_filename,
)
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


def _commit_3_sheet_order(num_projects: int) -> tuple[str, ...]:
    """Sheet roster after commit 3: per-project sheets sit between
    Assumptions and Glossary. Project sheet names are ``P{n} {Name}``.
    """
    base_pre = (
        "Cover",
        "Underwriting Summary",
        "Underwriting Pro Forma",
        "Underwriting Cash Flow",
        "Investor Returns",
        "Assumptions",
    )
    return base_pre + tuple(
        # Match the seeder's "Main Project" name; tests with custom names
        # build the expected roster themselves.
        f"P{i + 1} Main Project" for i in range(num_projects)
    ) + ("Glossary & Methodology",)

# Named-range prefixes the validator recognises. Anything else is treated as a
# free-form name (debug / one-off) and is allowed but doesn't get
# bidirectional-validated against the doc glossary.
_NAMED_RANGE_RE = re.compile(r"^(s|p\d+|r)_")

# Cells that are intentionally not "metrics" in the FINANCIAL_MODEL.md sense.
# Each falls into one of:
#   - Cover/scenario meta: sponsor, deal name, dates — context for the LP
#   - Assumption inputs: hold years, growth rates, occupancy starting point —
#     drive metrics rather than being them
#   - Aggregate/derived view of a metric (combined/worst across projects) —
#     not its own doc header but a slice of an existing metric
#   - Capital stack rows: per-module principal/rate (line items not metrics)
#   - Waterfall tier sums: per-tier-type aggregates (not metrics in the doc
#     sense, the metrics are LP/GP IRR/EM/etc. computed from these)
#   - S&U panel totals: scoped variants of "Total Uses" / "Total Sources" /
#     "Sources Gap" — these are doc-tagged metrics already, but the workbook
#     emits per-scope variants (s_su_*, p<n>_uw_*) that don't match the
#     bare doc name. These could be aliased; for now they're allow-listed.
_NON_METRIC_NAMES = frozenset({
    # Cover sheet meta
    "s_sponsor_name", "s_deal_name", "s_scenario_name", "s_snapshot_date",
    "s_project_count", "s_noi_basis",
    # Assumptions Block A — meta + inputs
    "s_assumptions_scenario_name", "s_assumptions_noi_basis",
    "s_assumptions_project_type",
    "s_hold_years", "s_opex_growth_rate", "s_initial_occupancy",
    "s_operating_reserve_months",
    # Per-project meta (header cells)
    "s_returns_combined_irr",  # alias for s_combined_irr
})

# Workbook ranges with these prefixes are not validated. Each prefix corresponds
# to a structural section (capital stack rows, waterfall tier sums, assumption
# inputs, S&U totals) where the per-instance name doesn't map 1:1 to a doc
# metric — the doc-level metric is what's tagged.
_NON_METRIC_PREFIXES = (
    "s_module_",       # capital stack rows (per-module principal/rate)
    "s_waterfall_",    # waterfall tier sums (per-tier-type cash distributed)
    "s_assumptions_",  # assumption inputs
    "s_su_",           # scenario S&U panel totals (alias of Total Uses/Sources/Gap)
)


def _is_non_metric(name: str) -> bool:
    """True for named ranges that are intentionally not metric outputs.

    Per-project Block-B inputs (``p<n>_acquisition_price``, ``p<n>_unit_count_*``,
    ``p<n>_avg_*_rent``, etc.) and per-project sheet meta (``p<n>_uw_*``) follow
    structural patterns rather than living in the explicit allow-list.
    """
    if name in _NON_METRIC_NAMES:
        return True
    if any(name.startswith(p) for p in _NON_METRIC_PREFIXES):
        return True
    # Per-project assumption inputs — match by suffix on the p<n>_ prefix.
    per_project_input_suffixes = {
        "project_name", "project_type",
        "acquisition_price", "unit_count_existing", "unit_count_new",
        "avg_in_place_rent", "avg_market_rent",
        "stabilized_occupancy", "going_in_cap_rate", "exit_cap_rate",
        "construction_months", "lease_up_months", "hold_years",
    }
    m = re.match(r"^p\d+_(.+)$", name)
    if m and m.group(1) in per_project_input_suffixes:
        return True
    # Per-project sheet meta + S&U totals (scoped variants of doc metrics).
    if m and m.group(1).startswith("uw_"):
        return True
    return False


async def _seed_minimal_scenario(session: AsyncSession):
    """Smallest reproducible Scenario for export smoke tests."""
    org, user = await seed_org(session)
    opportunity = await seed_opportunity(session, org, user, name="Smoke-Test Opportunity")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opportunity, user
    )
    return deal_model


async def test_workbook_has_expected_sheets(session: AsyncSession):
    """Single-project scenario: roster has exactly one P1 sheet inserted."""
    scenario = await _seed_minimal_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert tuple(wb.sheetnames) == _commit_3_sheet_order(num_projects=1)


async def test_per_project_sheet_per_project(session: AsyncSession):
    """Multi-project scenario: one P{n} sheet per project, in created_at order."""
    scenario = await _seed_minimal_scenario(session)
    # Seed two extra projects on the same scenario so we have N=3 total.
    from decimal import Decimal as _D
    from uuid import uuid4

    from app.models.deal import OperationalInputs as _OI
    from app.models.project import Project as _Project

    for label in ("Liberty", "East 25"):
        proj = _Project(
            id=uuid4(),
            scenario_id=scenario.id,
            opportunity_id=None,
            name=label,
            deal_type="acquisition",
        )
        session.add(proj)
        await session.flush()
        session.add(
            _OI(
                id=uuid4(), project_id=proj.id,
                unit_count_new=4, hold_period_years=5,
                exit_cap_rate_pct=_D("5.5"),
            )
        )
        await session.flush()

    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    sheets = tuple(wb.sheetnames)

    project_sheets = [s for s in sheets if s.startswith("P") and " " in s]
    assert len(project_sheets) == 3
    assert project_sheets[0].startswith("P1 ")
    assert project_sheets[1] == "P2 Liberty"
    assert project_sheets[2] == "P3 East 25"


async def test_long_project_name_truncated_to_27_chars(session: AsyncSession):
    """Per plan §2: P{n} prefix (4 chars) + ≤27 chars of name = 31 cap."""
    scenario = await _seed_minimal_scenario(session)
    from uuid import uuid4

    from app.models.project import Project as _Project

    long_name = "Northwest Freeway Industrial Park Phase II Buildings A through G"
    long_proj = _Project(
        id=uuid4(),
        scenario_id=scenario.id,
        opportunity_id=None,
        name=long_name,
        deal_type="acquisition",
    )
    session.add(long_proj)
    await session.flush()

    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    p2_sheets = [s for s in wb.sheetnames if s.startswith("P2 ")]
    assert len(p2_sheets) == 1
    sheet_name = p2_sheets[0]
    # 31-char Excel ceiling
    assert len(sheet_name) <= 31, sheet_name
    # Prefix is exactly "P2 " (3 chars), then up to 27 of the name
    assert sheet_name.startswith("P2 ")
    truncated_name_part = sheet_name[3:]
    assert truncated_name_part == long_name[: PROJECT_SHEET_NAME_BUDGET].rstrip()


async def test_named_ranges_resolve_to_existing_cells(session: AsyncSession):
    """Every defined name resolves to a real cell on the named sheet."""
    scenario = await _seed_minimal_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    assert len(wb.defined_names) > 0, "expected at least one defined name"
    for name in wb.defined_names:
        defined = wb.defined_names[name]
        destinations = list(defined.destinations)
        assert destinations, f"defined name {name!r} has no destinations"
        for sheet_title, ref in destinations:
            assert sheet_title in wb.sheetnames, (
                f"defined name {name!r} points to missing sheet {sheet_title!r}"
            )
            assert ref, f"defined name {name!r} has an empty cell ref"


async def test_no_sheet_protection(session: AsyncSession):
    """Investor export is read-only by convention (LPs need to copy values out).

    Plan §10: do not set Protection() on any sheet — the audience expectation
    is that the file is a values dump that can be reformatted, copied into
    the LP's own model, etc.
    """
    scenario = await _seed_minimal_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        assert not ws.protection.sheet, (
            f"sheet {sheet_name!r} has protection enabled — investor export should be read-only by convention only"
        )


async def test_glossary_sheet_has_investor_metrics(session: AsyncSession):
    """The Glossary sheet sources its term list from FINANCIAL_MODEL.md.

    Every metric tagged ``investor`` in the doc should land as a row.
    """
    scenario = await _seed_minimal_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    glossary = wb["Glossary & Methodology"]
    glossary_terms = {
        cell.value for cell in glossary["A"]
        if cell.value and isinstance(cell.value, str)
    }

    expected = {m.name for m in parse_doc().for_audience("investor")}
    missing = expected - glossary_terms
    assert not missing, f"investor-tagged metrics missing from Glossary sheet: {sorted(missing)}"


# Aggregate-variant patterns: workbook names that compose a base metric with
# a scope qualifier (combined / worst / lp / gp / scenario-level snapshot).
# Each pattern maps to the canonical doc-entry name it traces to.
_NAMED_RANGE_ALIASES: tuple[tuple[re.Pattern[str], str], ...] = (
    # Aggregate views (combined / worst across projects)
    (re.compile(r"^s_combined_irr$"), "LP IRR"),
    (re.compile(r"^s_returns_combined_irr$"), "LP IRR"),
    (re.compile(r"^s_combined_noi$"), "Stabilized NOI"),
    (re.compile(r"^s_worst_dscr$"), "DSCR"),
    (re.compile(r"^s_combined_dscr$"), "DSCR"),
    (re.compile(r"^s_hold_months$"), "Hold Period"),
    (re.compile(r"^s_modeled_duration_months$"), "Hold Period"),
    # Phase F additions — combined / scenario-level variants of base metrics
    (re.compile(r"^s_combined_unlevered_irr$"), "Unlevered IRR"),
    (re.compile(r"^s_combined_equity_multiple$"), "Equity Multiple"),
    (re.compile(r"^s_coc_year_one$"), "Cash-on-Cash Year 1"),
    # Valuation-delta is a derived display row (Modeled − Direct Cap), not its
    # own metric in the doc; trace it to the Direct Cap Value entry it nets
    # against rather than adding a "Δ" header.
    (re.compile(r"^s_valuation_delta$"), "Direct Cap Value"),
    # LP/GP-scoped variants of "Equity Multiple"
    (re.compile(r"^s_(lp|gp)_equity_multiple$"), "Equity Multiple"),
    # Asset Mgmt Fee — input on Assumptions, also a metric on Investor Returns
    (re.compile(r"^s_asset_mgmt_fee$"), "Asset Management Fee"),
    # Per-project sheet alias: workbook calls it timeline_months, doc Hold Period
    (re.compile(r"^p\d+_timeline_months$"), "Hold Period"),
    # Per-project levered/unlevered IRR — match by suffix
    (re.compile(r"^p\d+_levered_irr$"), "Levered IRR"),
    (re.compile(r"^p\d+_unlevered_irr$"), "Unlevered IRR"),
    # Per-project NOI variant (workbook reorders "noi_stabilized" vs
    # doc "Stabilized NOI"; sorted-tokens slug catches it, but be explicit)
    (re.compile(r"^p\d+_noi_stabilized$"), "Stabilized NOI"),
    # Net Cash Flow == Levered Cash Flow in our cashflow engine — the engine
    # writes both labels to different columns but they're computed identically
    # (NOI − DS + capital flows). Doc tag is "Levered Cash Flow".
    (re.compile(r"^r_(p\d+_)?(uw_)?(cf_)?net_cash_flow$"), "Levered Cash Flow"),
    # GP promote total — workbook adds a "_dollars" suffix; doc is "GP Promote"
    (re.compile(r"^s_gp_promote_dollars$"), "GP Promote"),
)


def _candidate_slugs(workbook_name: str) -> set[str]:
    """Generate every candidate slug a workbook named range could match against.

    Strips: ``s_`` / ``p<n>_`` / ``r_`` outer prefix, then any nested ``p<n>_``
    (for ``r_p<n>_*`` ranges), then ``uw_`` / ``cf_`` / ``matrix_`` section
    markers (Underwriting / CashFlow sheet / Per-Year Matrix view tags).
    Adds " cash flow" expansion when the range was tagged ``_cf_`` so e.g.
    ``r_p1_cf_levered`` matches doc "Levered Cash Flow".
    """
    bare = re.sub(r"^(s|p\d+|r)_", "", workbook_name)
    bare = re.sub(r"^p\d+_", "", bare)
    has_cf_marker = "_cf_" in workbook_name
    # ``matrix_`` is the per-year-matrix scope marker. Each matrix row is a
    # year-by-year view of an existing tagged metric (NOI, OER, IRR, etc.);
    # stripping the marker lets the row's name slug match the base metric.
    bare = re.sub(r"^(uw_|cf_|matrix_)+", "", bare)
    bare = re.sub(r"_(uw|cf|matrix)_", "_", bare)
    text = bare.replace("_", " ")
    aliases = slugs_for_name(text)
    if has_cf_marker:
        aliases.update(slugs_for_name(text + " cash flow"))
    # Apply explicit aggregate-variant aliases on top of the parsed slugs.
    for pattern, canonical in _NAMED_RANGE_ALIASES:
        if pattern.match(workbook_name):
            aliases.update(slugs_for_name(canonical))
    return aliases


async def test_every_named_range_traces_to_doc_entry(session: AsyncSession):
    """Bidirectional validator (workbook → doc direction) per plan §3.8 + §7.

    For every named range matching ``^(s|p\\d+|r)_`` in the workbook that
    purports to be a metric, assert it traces to a tagged doc entry in
    ``investor`` ∪ ``lender`` ∪ ``app`` (any user-facing audience). Cells
    that are intentionally not metrics — scenario meta, assumption inputs,
    capital-stack line items, waterfall tier sums — are excluded via
    ``_is_non_metric``.

    This is a **strict** gate: any new named range that doesn't trace
    must either (a) get a doc entry added in FINANCIAL_MODEL.md, (b) be
    added to ``_NON_METRIC_NAMES`` / ``_NON_METRIC_PREFIXES`` /
    ``per_project_input_suffixes`` with a one-line rationale, or (c) get
    an explicit alias added to ``_NAMED_RANGE_ALIASES``.
    """
    scenario = await _seed_minimal_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    lookup = name_lookup()  # default audiences = investor ∪ lender ∪ app

    orphans: list[str] = []
    for name in wb.defined_names:
        if not _NAMED_RANGE_RE.match(name):
            continue
        if _is_non_metric(name):
            continue
        if any(slug in lookup for slug in _candidate_slugs(name)):
            continue
        orphans.append(name)

    assert not orphans, (
        "named ranges with no doc entry — either tag the metric in "
        "FINANCIAL_MODEL.md, list in _NON_METRIC_NAMES/_PREFIXES, or "
        "add to _NAMED_RANGE_ALIASES:\n  "
        + "\n  ".join(sorted(orphans))
    )


async def test_uw_summary_kpis_match_engine_outputs(session: AsyncSession):
    """Parity test: hero KPIs on the Underwriting Summary sheet match the
    rollup engine's outputs to within rounding tolerance.
    """
    from app.engines.underwriting_rollup import rollup_summary

    scenario = await _seed_minimal_scenario(session)
    expected = await rollup_summary(scenario.id, session)
    expected_totals = expected["totals"]

    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    uw = wb["Underwriting Summary"]

    for key, name in (
        ("total_project_cost", "s_total_project_cost"),
        ("total_uses", "s_total_uses"),
        ("equity_required", "s_equity_required"),
    ):
        cell_value = _resolve_named_cell(wb, uw.title, name)
        engine_value = float(expected_totals.get(key) or 0)
        assert cell_value is not None, f"named range {name} not found"
        assert abs(float(cell_value or 0) - engine_value) < 0.01, (
            f"{name}: cell={cell_value} vs engine={engine_value}"
        )


async def test_export_handles_string_stored_enum_columns(session: AsyncSession):
    """Regression: production reads ``Scenario.project_type`` as a bare string,
    not the ``ProjectType`` enum instance.

    The Mapped column is typed ``Mapped[ProjectType]`` but the underlying
    column is plain ``String(60)`` — no enum adapter. SQLAlchemy stores
    the enum's ``.value`` on write but doesn't coerce back to the enum on
    read. The in-memory test fixture's session keeps the original Python
    enum reference, so prior tests didn't catch ``scenario.project_type.value``
    crashing with AttributeError in production. This test forces the
    string-stored case by reassigning the attribute to a plain string
    before invoking the export.
    """
    scenario = await _seed_minimal_scenario(session)
    scenario.project_type = "acquisition"  # simulate post-roundtrip read
    await session.flush()

    blob = await export_investor_workbook(scenario.id, session)
    assert blob, "expected workbook bytes — exporter must tolerate string-stored enums"
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "Assumptions" in wb.sheetnames


async def test_no_accidental_formula_strings(session: AsyncSession):
    """Regression: any string-typed cell value that starts with ``=`` is
    interpreted by Excel as a formula. If the string isn't valid formula
    syntax, Excel strips it on open with a "Removed Records: Formula"
    warning — silently destroying the cell content. We hit this once with
    Phase F4's hint cells (``"= Σ Stab NOI / 5.5% Exit Cap"``) which Excel
    tried to parse as a formula and failed.

    This test scans every populated cell across every sheet, asserts that
    any value starting with ``=`` is one of the export's known formula
    types (currently only ``HYPERLINK(...)``). Anything else is a bug.
    """
    scenario = await _seed_minimal_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    bad_cells: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                # Allowed: real HYPERLINK formulas the export generates
                # for in-workbook navigation + GitHub anchor links on the
                # Glossary sheet.
                if v.startswith("=HYPERLINK("):
                    continue
                bad_cells.append(f"{sheet_name}!{cell.coordinate}: {v!r}")

    assert not bad_cells, (
        "Cells whose string value starts with '=' but isn't a recognised "
        "formula — Excel will strip these on open with a Removed Records "
        "warning. Either drop the leading '=' or wrap as a real formula:\n  "
        + "\n  ".join(bad_cells)
    )


async def test_filename_slugged(session: AsyncSession):
    scenario = await _seed_minimal_scenario(session)
    from app.models.deal import Deal

    deal = await session.get(Deal, scenario.deal_id)
    name = make_investor_filename(scenario, deal)
    assert name.endswith("-investor.xlsx")
    assert " " not in name


def test_max_project_constants_are_consistent():
    """Plan §2: 5 max projects → 4-char prefix `P{n} ` + 27-char name = 31 (Excel ceiling)."""
    assert MAX_PROJECTS_PER_SCENARIO <= 99, "ordinal would need 3 digits"
    assert PROJECT_SHEET_NAME_BUDGET == 27
    assert len("P99 ") + PROJECT_SHEET_NAME_BUDGET == 31  # 2-digit ordinals fit too


# ── Helpers ───────────────────────────────────────────────────────────────────


def _slug(text: str) -> str:
    """Lowercase + alphanum-only slug for matching named-range fragments to
    metric titles like 'Total Project Cost (TPC)'."""
    return re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")


def _resolve_named_cell(wb, sheet_title: str, name: str):
    """Read the value at the cell registered under ``name`` on ``sheet_title``."""
    if name not in wb.defined_names:
        return None
    defined = wb.defined_names[name]
    for resolved_sheet, ref in defined.destinations:
        if resolved_sheet != sheet_title:
            continue
        sheet = wb[resolved_sheet]
        ref_clean = ref.replace("$", "")
        return sheet[ref_clean].value
    return None


_ = pytest  # silence "imported but unused" — pytest's asyncio_mode=auto handles invocation
