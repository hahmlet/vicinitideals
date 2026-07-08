"""Integration test: per-project phase plan block in investor export.

Exports a workbook for a multi-project scenario and verifies the
per-project sheet renders the Phase Plan produced by
:func:`app.engines.phase_plan.build_project_phase_windows`.

Contract (post-e7ba809): the per-phase KV named cells
(``p<n>_phase_<phase>_{start,end,duration}_month``) were replaced by a
display table (Phase | Start | Days | Months). The named cells that
downstream formulas actually consume are:

  - ``p<n>_total_horizon_months`` — sum of all phase durations (always)
  - ``p<n>_perm_origination_month`` — construction-to-perm conversion
    month (only when a construction-side phase exists); feeds the
    perm-gated Debt Service formulas.

So the tests assert: named cells resolve to sensible positive values,
and the Phase Plan table shows every phase with positive durations that
sum to the horizon.
"""
from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.exporters.investor_export import export_investor_workbook
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


pytestmark = pytest.mark.asyncio


async def _seed_multi_project_scenario(session: AsyncSession):
    """Per-project sheets are skipped when len(projects)==1, so seed a
    second project so ``_build_project_sheet`` actually fires."""
    from decimal import Decimal as _D
    from uuid import uuid4

    from app.models.deal import OperationalInputs as _OI
    from app.models.project import Project as _Project

    org, user = await seed_org(session)
    opportunity = await seed_opportunity(session, org, user, name="Phase-Plan-Smoke")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opportunity, user
    )
    extra = _Project(
        id=uuid4(),
        scenario_id=deal_model.id,
        opportunity_id=None,
        name="Second",
    )
    session.add(extra)
    await session.flush()
    session.add(
        _OI(
            id=uuid4(), project_id=extra.id,
            unit_count_new=4,
            exit_cap_rate_pct=_D("5.5"),
        )
    )
    await session.flush()
    return deal_model


def _resolve_named(wb, name: str) -> object | None:
    dn = wb.defined_names.get(name)
    if dn is None:
        return None
    for sheet_name, cell_ref in dn.destinations:
        sheet = wb[sheet_name]
        return sheet[cell_ref.replace("$", "")].value
    return None


def _phase_plan_table(wb, sheet_name_prefix: str) -> list[tuple[str, float]]:
    """Return ``[(phase_label, duration_months), ...]`` from the Phase
    Plan table on the first sheet whose name starts with the prefix."""
    ws = None
    for name in wb.sheetnames:
        if name.startswith(sheet_name_prefix):
            ws = wb[name]
            break
    assert ws is not None, f"no sheet starting with {sheet_name_prefix!r}"

    section_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Phase Plan":
            section_row = r
            break
    assert section_row is not None, f"Phase Plan section missing on {ws.title}"

    header = [ws.cell(row=section_row + 1, column=c).value for c in range(1, 5)]
    assert header == ["Phase", "Start", "Days", "Months"], (
        f"unexpected Phase Plan header: {header}"
    )

    rows: list[tuple[str, float]] = []
    r = section_row + 2
    while r <= ws.max_row:
        label = ws.cell(row=r, column=1).value
        if not isinstance(label, str) or label in (
            "Perm origination month", "Total horizon (months)",
        ):
            break
        months = ws.cell(row=r, column=4).value
        rows.append((label, float(months or 0)))
        r += 1
    return rows


async def test_phase_plan_named_cells_present_and_resolve(session: AsyncSession) -> None:
    scenario = await _seed_multi_project_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    defined = {name for name in wb.defined_names}
    # Both projects have OperationalInputs, so both get a horizon cell.
    required = {"p1_total_horizon_months", "p2_total_horizon_months"}
    missing = required - defined
    assert not missing, f"phase plan cells missing from workbook: {sorted(missing)}"

    horizon = _resolve_named(wb, "p1_total_horizon_months")
    assert isinstance(horizon, (int, float)) and float(horizon) > 0, (
        f"p1_total_horizon_months must resolve to a positive number; "
        f"got {horizon!r}"
    )


async def test_phase_plan_cell_values_are_positive_integers_in_order(
    session: AsyncSession,
) -> None:
    """The Phase Plan table must list every phase with positive
    durations, in engine order (acquisition first, exit last), and the
    durations must sum to ``p1_total_horizon_months``."""
    scenario = await _seed_multi_project_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    table = _phase_plan_table(wb, "P1 ")
    assert table, "Phase Plan table has no phase rows"

    labels = [label for label, _ in table]
    # Acquisition + Stabilized + Exit exist on any project type.
    for required in ("Acquisition", "Stabilized", "Exit"):
        assert required in labels, f"{required} phase missing: {labels}"
    assert labels[0] == "Acquisition", (
        f"acquisition must open the plan; got {labels}"
    )
    assert labels.index("Stabilized") < labels.index("Exit"), (
        f"stabilized must precede exit; got {labels}"
    )

    for label, months in table:
        assert months > 0, f"phase {label!r} has non-positive duration {months}"

    horizon = _resolve_named(wb, "p1_total_horizon_months")
    assert horizon is not None
    total = sum(months for _, months in table)
    assert abs(float(horizon) - total) < 0.01, (
        f"horizon ({horizon}) must equal the sum of phase durations ({total})"
    )
