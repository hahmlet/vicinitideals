"""Phase 4 KPI-tail: Cash-on-Cash (Year 1) becomes a formula.

CoC Y1 = first-year levered cash flow ÷ equity required.

Excel formula uses ``INDEX(r_uw_cf_levered, 1, 1)`` to pick the first
year cell of the annual Levered Cash Flow row on the Underwriting
Cash Flow sheet. ``year_cols`` on that sheet starts at 1 (Y0 stub is
intentionally skipped), so INDEX position 1 = Y1. ``MAX(0, ...)``
clamps the negative-Y1 case (equity call in year one) to 0 so a deal
still ramping shows 0% CoC rather than a misleading negative.

Parity vs engine: engine sums per-period waterfall ``cash_distributed``
across periods 1-12 for the equity tiers; Excel formula uses annual
aggregated Levered CF (sum of all tier distributions). Same
approximation envelope as the Combined EM / WEM formulas already
shipped on this row group.
"""
from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


pytestmark = pytest.mark.asyncio


async def _seed(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="CoC Y1 Formula")
    deal_model, _, stream, opex = await seed_deal_model_with_financials(
        session, opp, user
    )
    stream.active_in_phases = ["stabilized"]
    opex.active_in_phases = ["stabilized"]
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find_row(ws, label_exact: str) -> int | None:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip() == label_exact:
            return r
    return None


async def test_coc_y1_is_formula(session: AsyncSession) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Summary"]
    row = _find_row(ws, "Cash-on-Cash (Year 1)")
    assert row is not None, "CoC Y1 row missing on UW Summary"
    v = ws.cell(row=row, column=2).value
    assert isinstance(v, str) and v.startswith("=IFERROR(MAX(0,INDEX("), f"got {v!r}"


async def test_coc_y1_formula_references_required_named_ranges(
    session: AsyncSession,
) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Summary"]
    row = _find_row(ws, "Cash-on-Cash (Year 1)")
    v = ws.cell(row=row, column=2).value
    for ref in ("r_uw_cf_levered", "s_equity_required"):
        assert ref in v, f"CoC Y1 formula must reference {ref}; got {v!r}"
    # INDEX position 1,1 picks Y1 — Y0 is skipped on UW Cash Flow.
    assert "INDEX(r_uw_cf_levered,1,1)" in v, (
        f"CoC Y1 must INDEX the first column of r_uw_cf_levered; got {v!r}"
    )


async def test_coc_y1_named_range_registered(session: AsyncSession) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    names = set(wb.defined_names)
    assert "s_coc_year_one" in names
    assert "r_uw_cf_levered" in names
    assert "s_equity_required" in names


@pytest.mark.parametrize("profile", ["internal", "lp", "lender"])
async def test_coc_y1_formula_resolves_on_profile(
    session: AsyncSession, profile: str,
) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    if "Underwriting Summary" not in wb.sheetnames:
        pytest.skip(f"profile {profile!r} doesn't render UW Summary")
    names = set(wb.defined_names)
    assert "r_uw_cf_levered" in names
    assert "s_equity_required" in names
    assert "s_coc_year_one" in names
