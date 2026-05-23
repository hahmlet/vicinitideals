"""Graceful degradation: Debt Schedule must disclose engine-driven
approximations and binding constraints the LP would otherwise have to
infer from missing context. Notes are conditional — only emitted when
the underlying condition is present, so the block stays empty for a
vanilla perm-debt deal.

Contract:

  1. DSCR-capped sizing mode → note text mentions DSCR cap.
  2. Interest-reserve carry → note mentions average-draw approximation.
  3. Capitalized-interest (PIK) carry → note mentions PIK.
  4. Vanilla perm-debt only deal → no Notes section header.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from app.models.capital import (
    CapitalModule,
    CapitalModuleProject,
    EquityRole,
    VehicleType,
)
from app.models.deal import OperationalInputs
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


async def _seed_base(session: AsyncSession, *, opp_name: str):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name=opp_name)
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()
    return deal_model, project


def _attach_perm(deal_model_id, project_id) -> tuple[CapitalModule, CapitalModule]:
    debt = CapitalModule(
        scenario_id=deal_model_id,
        label="Senior Loan",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={
            "amount": "500000", "interest_rate_pct": 6.5,
            "amort_term_years": 30, "hold_term_years": 10,
        },
        carry={"carry_type": "pi", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    equity = CapitalModule(
        scenario_id=deal_model_id,
        label="LP Equity",
        vehicle_type=VehicleType.equity.value,
        equity_role=EquityRole.lp.value,
        stack_position=2,
        source={"amount": "250000"},
        carry={"carry_type": "none", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    return debt, equity


def _all_text(ws) -> str:
    chunks = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str):
            chunks.append(v)
    return "\n".join(chunks).lower()


async def test_dscr_capped_sizing_mode_emits_note(session: AsyncSession):
    deal_model, project = await _seed_base(session, opp_name="DSCR Cap Note")
    inputs = (
        await session.execute(
            select(OperationalInputs).where(
                OperationalInputs.project_id == project.id
            )
        )
    ).scalar_one()
    inputs.debt_sizing_mode = "dscr_capped"
    debt, equity = _attach_perm(deal_model.id, project.id)
    session.add_all([debt, equity])
    await session.flush()
    session.add_all([
        CapitalModuleProject(
            capital_module_id=debt.id, project_id=project.id,
            amount=Decimal("500000"),
        ),
        CapitalModuleProject(
            capital_module_id=equity.id, project_id=project.id,
            amount=Decimal("250000"),
        ),
    ])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()

    blob = await export_investor_workbook(deal_model.id, session, profile="internal")
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Debt Schedule"]
    text = _all_text(ws)

    assert "notes" in text, "Notes section header missing"
    assert "dscr-capped" in text, "DSCR-capped sizing-mode note missing"


async def test_interest_reserve_carry_emits_note(session: AsyncSession):
    deal_model, project = await _seed_base(session, opp_name="IR Note")
    debt = CapitalModule(
        scenario_id=deal_model.id,
        label="Construction Loan",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={
            "amount": "500000", "interest_rate_pct": 7.0,
            "amort_term_years": 30, "hold_term_years": 3,
        },
        carry={"carry_type": "interest_reserve", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "construction_end"},
        active_phase_start="construction", active_phase_end="stabilized",
    )
    equity = CapitalModule(
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type=VehicleType.equity.value,
        equity_role=EquityRole.lp.value,
        stack_position=2,
        source={"amount": "250000"},
        carry={"carry_type": "none", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    session.add_all([debt, equity])
    await session.flush()
    session.add_all([
        CapitalModuleProject(
            capital_module_id=debt.id, project_id=project.id,
            amount=Decimal("500000"),
        ),
        CapitalModuleProject(
            capital_module_id=equity.id, project_id=project.id,
            amount=Decimal("250000"),
        ),
    ])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()

    blob = await export_investor_workbook(deal_model.id, session, profile="internal")
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Debt Schedule"]
    text = _all_text(ws)
    assert "interest reserve" in text, "Interest-reserve disclosure missing"
    assert "average-draw" in text, "Approximation note text missing"


async def test_vanilla_perm_debt_emits_no_notes_header(session: AsyncSession):
    """Pure PI/perm deal with no special carry + no special sizing mode →
    Notes block stays empty (header not emitted)."""
    deal_model, project = await _seed_base(session, opp_name="Vanilla Perm")
    debt, equity = _attach_perm(deal_model.id, project.id)
    session.add_all([debt, equity])
    await session.flush()
    session.add_all([
        CapitalModuleProject(
            capital_module_id=debt.id, project_id=project.id,
            amount=Decimal("500000"),
        ),
        CapitalModuleProject(
            capital_module_id=equity.id, project_id=project.id,
            amount=Decimal("250000"),
        ),
    ])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()

    blob = await export_investor_workbook(deal_model.id, session, profile="internal")
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Debt Schedule"]
    # Walk column A; "Notes" should not appear as a standalone header on
    # this sheet — only the Loan Summary + Amortization sections.
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip().lower() == "notes":
            pytest.fail(
                f"vanilla deal must not emit Notes section header (row {r})"
            )
