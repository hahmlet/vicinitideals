"""Parity helpers for the investor-export formula conversion.

Each commit in the formula-conversion plan
(``docs/feature-plans/investor-excel-formula-conversion.md``) flips
hard-coded cell values to formulas. To prevent silent regressions the
test suite captures a "before" workbook, then after the conversion
captures an "after" workbook and asserts the *computed* values are
unchanged (within tolerance) — only the cell *type* changes
(``value`` → ``formula``).

This module provides the diff primitive used by those tests.

Two distinct workbook readers are needed:

- ``data_only=False`` — returns the cell as-written. For a value cell
  this is the value; for a formula cell this is the formula string
  starting with ``=``. Used to assert formulas were emitted.
- ``data_only=True`` — returns the *cached* computed value openpyxl
  stored when the workbook was last saved by a formula engine
  (Excel / LibreOffice). For workbooks openpyxl wrote without
  recalc the cached values are stale, so the parity helper here
  compares cell-as-written for value cells, and accepts that
  formula cells need a recalc step (provided separately) before they
  can be value-compared.

The helper is intentionally light: workbook iteration + cell coercion +
tolerance comparison. Nothing here knows about specific sheets or
named ranges.
"""
from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# Tolerance for numeric comparison. 0.01 absolute covers Decimal/float
# round-trip noise; 1e-6 relative covers very large values like Total
# Project Cost where 0.01 absolute is meaninglessly tight.
ABSOLUTE_TOLERANCE = 0.01
RELATIVE_TOLERANCE = 1e-6


@dataclass(frozen=True)
class CellLocation:
    """Sheet + 1-based row/col, stringified for human-readable diff output."""

    sheet: str
    row: int
    col: int

    def __str__(self) -> str:
        return f"{self.sheet}!{get_column_letter(self.col)}{self.row}"


@dataclass(frozen=True)
class CellDelta:
    """A single cell whose value differs between two workbook snapshots."""

    location: CellLocation
    before: Any
    after: Any

    def __str__(self) -> str:
        return f"{self.location}: {self.before!r} -> {self.after!r}"


def _coerce(value: Any) -> Any:
    """Normalize a cell value for comparison.

    - ``None`` and ``""`` collapse to ``None`` (both render blank).
    - ``bool`` stays as-is.
    - ``Decimal`` -> ``float`` (openpyxl writes Decimals as floats).
    - Everything else passes through.
    """
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, Decimal):
        return float(value)
    return value


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _values_match(a: Any, b: Any) -> bool:
    """Compare two coerced cell values with numeric tolerance.

    Strings (incl. formulas) compare exactly. Numerics compare within
    the absolute-OR-relative tolerance defined at module top.
    """
    if a is None and b is None:
        return True
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        if abs(a - b) <= ABSOLUTE_TOLERANCE:
            return True
        denom = max(abs(a), abs(b))
        return denom > 0 and abs(a - b) / denom <= RELATIVE_TOLERANCE
    return a == b


def _iter_cells(blob: bytes, *, data_only: bool):
    """Yield (sheet, row, col, value) for every cell in the workbook.

    Only iterates the populated bounding box per sheet — empty rows
    past ``ws.max_row`` and empty cols past ``ws.max_column`` are
    skipped, which is what openpyxl already does for ``ws.iter_rows``.
    """
    wb = load_workbook(BytesIO(blob), data_only=data_only)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                yield sheet, cell.row, cell.column, cell.value


def workbook_cell_map(
    blob: bytes, *, data_only: bool = False
) -> dict[CellLocation, Any]:
    """Build a {CellLocation: coerced value} map for a workbook blob.

    ``data_only=False`` returns formula strings for formula cells;
    ``data_only=True`` returns the cached computed value (or ``None``
    if openpyxl has no cache, which is the common case for workbooks
    written by openpyxl without subsequent recalc).
    """
    return {
        CellLocation(sheet=s, row=r, col=c): _coerce(v)
        for s, r, c, v in _iter_cells(blob, data_only=data_only)
        if _coerce(v) is not None
    }


def diff_workbook_values(
    before_blob: bytes, after_blob: bytes
) -> list[CellDelta]:
    """Return cells whose values differ between two workbook snapshots.

    Compares both in ``data_only=False`` mode (cell-as-written). Diff
    includes:

    - Cells present in ``before`` but missing in ``after`` (or vice versa)
    - Cells whose value/formula changed beyond tolerance

    Order: sorted by sheet, then row, then column for stable test output.
    """
    before = workbook_cell_map(before_blob, data_only=False)
    after = workbook_cell_map(after_blob, data_only=False)

    deltas: list[CellDelta] = []
    all_locations = set(before) | set(after)
    for loc in sorted(all_locations, key=lambda l: (l.sheet, l.row, l.col)):
        b = before.get(loc)
        a = after.get(loc)
        if _values_match(b, a):
            continue
        deltas.append(CellDelta(location=loc, before=b, after=a))
    return deltas


def diff_only_value_cells(
    before_blob: bytes, after_blob: bytes
) -> list[CellDelta]:
    """Diff that ignores cells that became formulas.

    Use this when you've intentionally converted some cells to formulas
    and want to assert *no other cell* changed. A cell that was a value
    in ``before`` and is a formula in ``after`` is skipped (its
    computed value is asserted separately via a recalc-based test).
    """
    deltas = diff_workbook_values(before_blob, after_blob)
    return [d for d in deltas if not _is_formula(d.after)]


def count_formula_cells(blob: bytes) -> dict[str, int]:
    """Count formula cells per sheet. Used to show conversion progress."""
    counts: dict[str, int] = {}
    for sheet, _row, _col, value in _iter_cells(blob, data_only=False):
        if _is_formula(value):
            counts[sheet] = counts.get(sheet, 0) + 1
    return counts


# ── Excel recalc + computed-value access ──────────────────────────────────────
#
# openpyxl writes formulas as-text but cannot evaluate them. To compare an
# engine-computed value (e.g. DSCR = 1.15) against the Excel cell that should
# show that same value via a formula, the workbook must be opened by a real
# formula engine that recalcs every cell and writes the computed values back
# to the .xlsx cache. openpyxl then reads those cached values via
# ``data_only=True``.
#
# The harness uses Excel COM on Windows (Office must be installed). A future
# CI gate will swap in headless LibreOffice via the same interface
# (``recalc_workbook(path)``) so non-Windows runs can execute the parity
# tests too. The plan calls this out in §8 / §9.


import platform  # noqa: E402 — kept near the COM helpers, not the top imports.
import subprocess  # noqa: E402
from pathlib import Path  # noqa: E402


class RecalcUnavailableError(RuntimeError):
    """Raised when no formula-evaluation engine is available on this host."""


def recalc_with_excel_com(path: Path) -> None:
    """Force-recalc a workbook in place via Excel COM (Windows + Office).

    Opens ``path`` in a hidden Excel instance, runs ``Application.Calculate``
    + ``Workbook.Save`` (which embeds the computed values into the .xlsx
    cache), then closes Excel. After this call, ``load_workbook(path,
    data_only=True)`` returns Excel-evaluated cell values for every formula
    in the workbook.

    Raises ``RecalcUnavailableError`` if pywin32 is missing or Excel is
    not installed. Callers should catch and skip the test on those hosts.
    """
    if platform.system() != "Windows":
        raise RecalcUnavailableError(
            "Excel COM recalc is Windows-only; install LibreOffice or use "
            "a different recalc backend on this host"
        )
    try:
        import win32com.client  # type: ignore[import-not-found]
    except ImportError as exc:
        raise RecalcUnavailableError(
            "pywin32 not installed in this environment"
        ) from exc

    abs_path = str(Path(path).resolve())
    # ``gencache.EnsureDispatch`` builds the COM type library cache so
    # property setters and methods resolve through static dispatch. Plain
    # ``Dispatch`` returns a dynamic proxy that on newer Office installs
    # rejects both property setters (``excel.DisplayAlerts = False``) and
    # methods (``excel.Quit()``) with ``AttributeError`` because the
    # type library can't be auto-generated. When EnsureDispatch fails
    # too (``This COM object can not automate the makepy process``),
    # there's nothing we can do — surface as RecalcUnavailableError so
    # tests skip cleanly instead of failing with a misleading trace.
    try:
        excel = win32com.client.gencache.EnsureDispatch("Excel.Application")
    except (TypeError, AttributeError) as exc:
        raise RecalcUnavailableError(
            f"Excel COM type library unavailable on this host "
            f"(makepy automation failed: {exc!r}). Use LibreOffice or "
            f"prime the gen_py cache manually."
        ) from exc

    try:
        excel.DisplayAlerts = False
        excel.Visible = False
        wb = excel.Workbooks.Open(abs_path)
        try:
            excel.CalculateFull()
            wb.Save()
        finally:
            wb.Close(SaveChanges=False)
    finally:
        try:
            excel.Quit()
        except AttributeError:
            pass


def recalc_with_libreoffice(path: Path) -> None:
    """Headless LibreOffice recalc (Linux/macOS/Windows-with-LO).

    Roundtrips the workbook through ``soffice --headless --convert-to xlsx``
    which evaluates formulas and writes cached values into the output. The
    resulting .xlsx replaces the original at ``path``.

    Used as the CI backend on Linux runners where Excel COM is unavailable.
    """
    soffice = _find_soffice()
    if soffice is None:
        raise RecalcUnavailableError(
            "LibreOffice (soffice) not found on PATH"
        )
    src = Path(path).resolve()
    out_dir = src.parent
    subprocess.run(
        [soffice, "--headless", "--calc", "--convert-to", "xlsx",
         "--outdir", str(out_dir), str(src)],
        check=True, capture_output=True, timeout=120,
    )
    # LibreOffice writes <stem>.xlsx into out_dir; if it equals src we're
    # done. If LO chose a different name (rare), surface that as an error.
    if not src.exists():
        raise RecalcUnavailableError(
            f"LibreOffice convert did not produce {src.name}"
        )


def _find_soffice() -> str | None:
    """Locate the ``soffice`` executable across platforms."""
    import shutil
    for name in ("soffice", "libreoffice"):
        path = shutil.which(name)
        if path:
            return path
    return None


def recalc_workbook(path: Path) -> str:
    """Recalc ``path`` with whichever engine is available.

    Returns the backend name ("excel" or "libreoffice") on success.
    Tries Excel COM first (faster, more accurate), then LibreOffice.
    Raises ``RecalcUnavailableError`` if neither works.
    """
    if platform.system() == "Windows":
        try:
            recalc_with_excel_com(path)
            return "excel"
        except RecalcUnavailableError:
            pass
    recalc_with_libreoffice(path)
    return "libreoffice"


# Excel's seven legacy error sentinels plus the two dynamic-array-era
# ones. Excel writes these as the literal cell value when a formula
# fails to resolve. Spotting them post-recalc is the cheapest way to
# catch a whole class of dangling references / type mismatches the
# engine-parity tests would otherwise pass right over.
EXCEL_ERROR_VALUES: frozenset[str] = frozenset({
    "#NAME?", "#REF!", "#VALUE!", "#DIV/0!", "#NUM!",
    "#N/A", "#NULL!", "#SPILL!", "#CALC!",
})


def find_error_cells(path: Path) -> list[tuple[str, str, str]]:
    """Scan every sheet for cells whose post-recalc value is an Excel
    error sentinel (``#NAME?``, ``#REF!``, etc.).

    Returns a list of ``(sheet, coordinate, error_value)`` tuples —
    empty when the workbook is clean. Call ``recalc_workbook`` first so
    the error values are actually written to the cells; openpyxl-only
    loads won't surface them.
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    errors: list[tuple[str, str, str]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v in EXCEL_ERROR_VALUES:
                    errors.append((sheet_name, cell.coordinate, v))
    return errors


def read_named_value(path: Path, name: str, *, data_only: bool = True):
    """Read the value of a defined name from a workbook.

    With ``data_only=True``, returns the cached computed value (run
    ``recalc_workbook`` first to make this meaningful). With
    ``data_only=False``, returns the formula text (string starting with
    ``=``) or the literal value if the cell isn't a formula.

    Raises ``KeyError`` if the defined name doesn't exist.
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=data_only)
    if name not in wb.defined_names:
        raise KeyError(f"defined name {name!r} not in workbook")
    destinations = list(wb.defined_names[name].destinations)
    if not destinations:
        raise KeyError(f"defined name {name!r} has no destinations")
    sheet, ref = destinations[0]
    # ref is "$A$3" style; openpyxl accepts it directly.
    return wb[sheet][ref.replace("$", "")].value


def read_formula_text(path: Path, name: str) -> str | None:
    """Return the formula text behind a defined name, or None if not a formula.

    Convenience wrapper around ``read_named_value(..., data_only=False)``
    that filters to formula strings only. Used by parity tests that assert
    a cell is formula-driven, not a hard-coded value.
    """
    raw = read_named_value(path, name, data_only=False)
    if isinstance(raw, str) and raw.startswith("="):
        return raw
    return None
