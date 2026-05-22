"""Engine-vs-formula parity for the Debt Schedule sheet.

Commit 6 of docs/feature-plans/investor-excel-formula-conversion.md §4.4.

Scope: Loan Summary "Annual P&I" cell for ``pi`` carry-type rows.
Other carry types and the amort table conversion stay as engine
values for now (CUMIPMT / CUMPRINC wiring lands in a later commit).
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import (
    export_investor_workbook,
)
from app.models.capital import CapitalModule, CapitalModuleProject, VehicleType
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)
from tests.exporters._parity_helpers import (
    RecalcUnavailableError,
    recalc_workbook,
)


async def _seed_pi_loan_scenario(session: AsyncSession):
    """Seed scenario with a ``pi`` carry-type debt module so the PMT
    formula path fires. Without ``pi`` carry the Annual P&I cell stays
    em-dash and there is nothing to test."""
    org, user = await seed_org(session)
    opportunity = await seed_opportunity(
        session, org, user, name="Debt-Parity Smoke"
    )
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opportunity, user
    )
    project_row = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()

    perm = CapitalModule(
        scenario_id=deal_model.id,
        label="Perm Loan",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={
            "amount": "1000000",
            "interest_rate_pct": 6.0,
            "amort_term_years": 30,
            "hold_term_years": 10,
            "io_months": 0,
        },
        carry={"carry_type": "pi", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition",
        active_phase_end="exit",
    )
    session.add(perm)
    await session.flush()
    session.add(
        CapitalModuleProject(
            capital_module_id=perm.id,
            project_id=project_row.id,
            amount=Decimal("1000000"),
        )
    )
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find_row_by_label(ws, label: str) -> int | None:
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == label:
            return r
    return None


async def test_annual_pi_is_pmt_formula(session: AsyncSession):
    """Annual P&I cell on the Perm Loan row carries a PMT formula."""
    scenario = await _seed_pi_loan_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)

    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Debt Schedule"]
    row = _find_row_by_label(ws, "Perm Loan")
    assert row is not None
    v = ws.cell(row=row, column=10).value
    assert isinstance(v, str) and "PMT(" in v, (
        f"Annual P&I should be a PMT formula; got {v!r}"
    )
    # Must reference principal (col C), rate (col D), amort (col F) on
    # the same row so LP edits to any of them recompute the cell.
    for col_letter in ("C", "D", "F"):
        assert f"{col_letter}{row}" in v, (
            f"PMT formula must reference {col_letter}{row}; got {v!r}"
        )


async def test_annual_pi_evaluates_close_to_engine_value(
    session: AsyncSession, tmp_path: Path
):
    """Excel-recalc'd PMT*12 matches the engine's annual P&I within $1."""
    scenario = await _seed_pi_loan_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)
    try:
        recalc_workbook(path)
    except RecalcUnavailableError as exc:
        pytest.skip(f"no recalc backend: {exc}")

    # Engine reference: monthly_pmt * 12 for principal=1e6, rate=6%, n=30y
    from app.engines.cashflow import _monthly_pmt
    expected = float(_monthly_pmt(Decimal("1000000"), 6.0, 30) * Decimal(12))

    wb = load_workbook(path, data_only=True)
    ws = wb["Debt Schedule"]
    row = _find_row_by_label(ws, "Perm Loan")
    assert row is not None
    excel_value = ws.cell(row=row, column=10).value
    assert excel_value is not None
    diff = abs(float(excel_value) - expected)
    assert diff < 1.0, (
        f"Annual P&I parity: engine={expected}, excel={excel_value}, "
        f"diff={diff}"
    )
