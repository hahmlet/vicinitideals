"""Phase A: Gross Revenue Y2+ on Pro Forma sheets must be growth-chain driven
so a single LP edit to a growth-rate input ripples through every downstream
year on the pro forma.

Two layouts implement that contract today:

  - Legacy layout (proforma-profile "Pro Forma" sheet): the Gross Revenue
    total row itself chains ``=prev_year_cell * (1 + s_revenue_growth_rate)``
    from Y2 on; Y0/Y1 stay numeric engine seeds.
  - Consolidated layout ("Underwriting Pro Forma", post-e7ba809): the total
    row is a column SUM over per-stream bullet rows, and the growth chain
    lives on each bullet — Y1 = ``=s_rev_<slug>_y1_monthly*12`` (numeric
    input cell on Assumptions Block F), Y2+ =
    ``=prev_cell*(1+s_rev_<slug>_escalation_pct)``.

Guards three contracts:

  1. ``s_revenue_growth_rate`` named range is registered on the Assumptions
     sheet (so legacy-layout formulas aren't dangling).
  2. The Gross Revenue Y2 column is formula-driven and traces back to a
     growth-rate input (sheet-wide rate or per-stream escalation).
  3. The chain is anchored on a numeric base (Y0/Y1 engine seeds on the
     legacy layout; the Block F Y1 input cell on the consolidated layout).
"""
from __future__ import annotations

import re
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from app.models.capital import (
    CapitalModule,
    CapitalModuleProject,
    EquityRole,
    VehicleType,
)
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)
from tests.exporters._parity_helpers import (
    find_label_row,
    parse_sum_range,
    proforma_layout,
)


async def _seed(session: AsyncSession):
    """Seed a scenario rich enough to produce >= 3 annual columns so the
    Y2+ revenue growth chain has cells to inspect."""
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Revenue Growth Smoke")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()

    debt = CapitalModule(
        scenario_id=deal_model.id,
        label="Senior Loan",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={
            "amount": "500000", "interest_rate_pct": 6.5,
            "amort_term_years": 30, "hold_term_years": 10,
        },
        carry={"carry_type": "pi", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    equity = CapitalModule(
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type=VehicleType.equity.value,
        equity_role=EquityRole.lp.value,
        stack_position=2,
        source={"amount": "250000"},
        carry={"carry_type": "none", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    session.add_all([debt, equity])
    await session.flush()
    session.add_all([
        CapitalModuleProject(
            capital_module_id=debt.id, project_id=project.id,
            amount=Decimal("500000"),
        ),
        CapitalModuleProject(
            capital_module_id=equity.id, project_id=project.id,
            amount=Decimal("250000"),
        ),
    ])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find_pro_forma_sheet(wb):
    for name in wb.sheetnames:
        low = name.lower()
        if "pro forma" in low or "proforma" in low:
            return wb[name]
    return None


def _gross_revenue_row(ws) -> int | None:
    label_col, _ = proforma_layout(ws)
    return find_label_row(ws, "gross revenue", col=label_col)


def _bullet_rows_for_total(ws, total_row: int, y1_col: int) -> list[int]:
    """Rows covered by the total row's ``=SUM(col_first:col_last)`` formula.

    The consolidated layout emits per-stream bullet rows directly above
    the total; the total's Y1 SUM range is the authoritative bullet span.
    """
    parsed = parse_sum_range(ws.cell(row=total_row, column=y1_col).value)
    if parsed is None:
        return []
    _, first, last = parsed
    return list(range(first, last + 1))


@pytest.mark.parametrize("profile", ["internal", "lp", "lender", "proforma"])
async def test_revenue_growth_named_range_registered(
    session: AsyncSession, profile: str
):
    """``s_revenue_growth_rate`` must be defined on every profile that
    ships the Pro Forma sheet, or the Y2+ formulas would dangle to #NAME?."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "s_revenue_growth_rate" in wb.defined_names, (
        f"profile={profile} missing s_revenue_growth_rate defined name"
    )


@pytest.mark.parametrize("profile", ["internal", "lp", "lender", "proforma"])
async def test_gross_revenue_y2_plus_is_growth_formula(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)

    ws = _find_pro_forma_sheet(wb)
    assert ws is not None, f"profile={profile} missing Pro Forma sheet"

    gr_row = _gross_revenue_row(ws)
    assert gr_row is not None, (
        f"profile={profile} missing Gross Revenue row on Pro Forma sheet"
    )
    label_col, y0_col = proforma_layout(ws)
    y1_col, y2_col = y0_col + 1, y0_col + 2

    y2 = ws.cell(row=gr_row, column=y2_col).value
    assert isinstance(y2, str) and y2.startswith("="), (
        f"profile={profile}: Gross Revenue Y2 must be formula; got {y2!r}"
    )

    if label_col == 1:
        # Legacy layout: sheet-wide growth chain on the total row itself.
        assert "s_revenue_growth_rate" in y2, (
            f"profile={profile}: Y2 formula missing s_revenue_growth_rate "
            f"ref; got {y2!r}"
        )
        prev_col = get_column_letter(y1_col)
        assert f"{prev_col}{gr_row}" in y2, (
            f"profile={profile}: Y2 formula must reference prior-year cell "
            f"{prev_col}{gr_row}; got {y2!r}"
        )
        return

    # Consolidated layout: total = SUM over bullet rows; the growth chain
    # lives on each per-stream bullet via its escalation input.
    bullet_rows = _bullet_rows_for_total(ws, gr_row, y1_col)
    assert bullet_rows, (
        f"profile={profile}: Gross Revenue total must SUM over per-stream "
        f"bullet rows; Y1 cell was "
        f"{ws.cell(row=gr_row, column=y1_col).value!r}"
    )
    assert parse_sum_range(y2) is not None, (
        f"profile={profile}: Gross Revenue Y2 must SUM the bullet column; "
        f"got {y2!r}"
    )
    prev_col = get_column_letter(y1_col)
    for br in bullet_rows:
        bullet_y2 = ws.cell(row=br, column=y2_col).value
        assert isinstance(bullet_y2, str) and bullet_y2.startswith("="), (
            f"profile={profile}: bullet row {br} Y2 must be a growth "
            f"formula; got {bullet_y2!r}"
        )
        assert "s_rev_" in bullet_y2 and "_escalation_pct" in bullet_y2, (
            f"profile={profile}: bullet row {br} Y2 must reference the "
            f"per-stream s_rev_<slug>_escalation_pct input; got {bullet_y2!r}"
        )
        assert f"{prev_col}{br}" in bullet_y2, (
            f"profile={profile}: bullet row {br} Y2 must chain off its own "
            f"prior-year cell {prev_col}{br}; got {bullet_y2!r}"
        )


@pytest.mark.parametrize("profile", ["internal", "lp", "lender", "proforma"])
async def test_gross_revenue_y0_y1_remain_numeric_seeds(
    session: AsyncSession, profile: str
):
    """Growth chain needs a numeric base.

    Legacy layout: Y0+Y1 on the total row stay engine-seeded numerics.
    Consolidated layout: the chain anchors on each bullet's Y1 formula
    ``=s_rev_<slug>_y1_monthly*12`` whose named cell is a numeric input
    on Assumptions Block F (the Y1 bullet must not already escalate).
    """
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)

    ws = _find_pro_forma_sheet(wb)
    assert ws is not None

    gr_row = _gross_revenue_row(ws)
    assert gr_row is not None
    label_col, y0_col = proforma_layout(ws)
    y1_col = y0_col + 1

    if label_col == 1:
        y0 = ws.cell(row=gr_row, column=y0_col).value
        y1 = ws.cell(row=gr_row, column=y1_col).value
        for label, val in (("Y0", y0), ("Y1", y1)):
            assert not (isinstance(val, str) and val.startswith("=")), (
                f"profile={profile}: Gross Revenue {label} must be numeric "
                f"seed, got formula {val!r}"
            )
        return

    bullet_rows = _bullet_rows_for_total(ws, gr_row, y1_col)
    assert bullet_rows, f"profile={profile}: no revenue bullet rows found"
    for br in bullet_rows:
        y1 = ws.cell(row=br, column=y1_col).value
        assert isinstance(y1, str) and y1.startswith("="), (
            f"profile={profile}: bullet row {br} Y1 must anchor the chain; "
            f"got {y1!r}"
        )
        assert "_y1_monthly" in y1 and "*12" in y1, (
            f"profile={profile}: bullet row {br} Y1 must annualize the "
            f"Block F monthly input; got {y1!r}"
        )
        assert "_escalation_pct" not in y1, (
            f"profile={profile}: bullet row {br} Y1 must NOT already "
            f"escalate (chain starts at Y2); got {y1!r}"
        )
        # The anchor input itself must be numeric on Assumptions.
        m = re.search(r"(s_rev_\w+_y1_monthly)", y1)
        assert m is not None, f"got {y1!r}"
        name = m.group(1)
        assert name in wb.defined_names, (
            f"profile={profile}: {name} not registered — bullet Y1 would "
            f"dangle to #NAME?"
        )
        sheet, ref = next(iter(wb.defined_names[name].destinations))
        anchor = wb[sheet][ref.replace("$", "")].value
        assert isinstance(anchor, (int, float)) or (
            isinstance(anchor, str) and anchor.startswith("=")
        ), (
            f"profile={profile}: {name} anchor cell empty; got {anchor!r}"
        )
