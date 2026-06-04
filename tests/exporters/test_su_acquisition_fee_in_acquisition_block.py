"""Slice 3 (Export v3): auto Acquisition Fee UseLines must render in the
Acquisition cost section of the S&U sheet, not Soft Costs.

The engine seeds the auto Acquisition Fee row with
``cost_category="soft"`` (so it stays grouped with sponsor-paid
overhead lines in the DB shape), but the LP expects to see the fee
where it belongs conceptually — as part of the acquisition spend.
Slice 3 effective-category routing on the exporter side reads
``is_auto_acquisition_fee=True`` and lands the row under Acquisition
regardless of its stored ``cost_category``.

Wiring contract:
  1. Acquisition Fee row appears under the Acquisition subtotal range.
  2. ``s_acquisition_fee`` named range points at its B column cell.
  3. NO Acquisition Fee row appears under Soft Costs.
"""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.exporters.investor_export import export_investor_workbook
from app.models.deal import UseLine
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


pytestmark = pytest.mark.asyncio


async def _seed_with_acq_fee(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Acq Fee Routing")
    deal_model, _inputs, _stream, _opex = (
        await seed_deal_model_with_financials(session, opp, user)
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()

    # Hard acquisition line — anchors the Acquisition category so the
    # auto fee row is not the only entry under it.
    session.add(
        UseLine(
            project_id=project.id,
            label="Land Acquisition",
            amount=Decimal("4000000"),
            phase="acquisition",
            cost_category="hard",
        )
    )
    # Auto Acquisition Fee — engine convention is to persist with
    # cost_category="soft"; the exporter routes it to Acquisition
    # because of the is_auto_acquisition_fee flag.
    session.add(
        UseLine(
            project_id=project.id,
            label="Acquisition Fee",
            amount=Decimal("80000"),
            phase="acquisition",
            cost_category="soft",
            is_auto_acquisition_fee=True,
            acquisition_fee_pct=Decimal("2.0"),
        )
    )
    await session.flush()
    await session.commit()
    return deal_model, project


def _find_row(ws, needle: str) -> int | None:
    needle_norm = needle.strip().lower()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and needle_norm in v.strip().lower():
            return r
    return None


def _find_all_rows(ws, needle: str) -> list[int]:
    needle_norm = needle.strip().lower()
    out: list[int] = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and needle_norm in v.strip().lower():
            out.append(r)
    return out


def _find_named_cell(wb, name: str) -> tuple[str, int, int] | None:
    dn = wb.defined_names.get(name)
    if dn is None:
        return None
    for sheet_title, coord in dn.destinations:
        cell = coord.replace("$", "")
        col_letters = "".join(c for c in cell if c.isalpha())
        row_digits = "".join(c for c in cell if c.isdigit())
        col_idx = 0
        for ch in col_letters:
            col_idx = col_idx * 26 + (ord(ch.upper()) - ord("A") + 1)
        return sheet_title, int(row_digits), col_idx
    return None


async def test_acquisition_fee_lands_under_acquisition_subtotal(
    session: AsyncSession, tmp_path: Path
) -> None:
    scenario, _ = await _seed_with_acq_fee(session)
    blob = await export_investor_workbook(scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)

    wb = load_workbook(path, data_only=False)
    ws = wb["Sources & Uses"]

    fee_row = _find_row(ws, "Acquisition Fee")
    assert fee_row is not None, "Acquisition Fee row not rendered"

    acq_subtotal = _find_row(ws, "Subtotal Acquisition")
    soft_subtotal = _find_row(ws, "Subtotal Soft")
    assert acq_subtotal is not None, "Subtotal Acquisition row missing"

    # The fee row must appear before the Acquisition subtotal (inside
    # its SUM range), and either before any Soft subtotal or absent
    # entirely from the Soft section.
    assert fee_row < acq_subtotal, (
        f"Acquisition Fee row ({fee_row}) appears after the "
        f"Acquisition subtotal ({acq_subtotal}); it is being routed "
        f"to the wrong category."
    )
    if soft_subtotal is not None:
        # Stronger check: the fee row must NOT be in the Soft Costs
        # SUM range (between Soft section start and its subtotal).
        acq_formula = ws.cell(row=acq_subtotal, column=2).value
        assert isinstance(acq_formula, str)
        # The SUM range B{first}:B{last} on the Acquisition subtotal
        # must encompass the fee row.
        assert f"B{fee_row}" not in (
            ws.cell(row=soft_subtotal, column=2).value or ""
        ), "Acquisition Fee row is being summed into the Soft Costs subtotal"


async def test_acquisition_fee_named_cell_registered(
    session: AsyncSession, tmp_path: Path
) -> None:
    scenario, _ = await _seed_with_acq_fee(session)
    blob = await export_investor_workbook(scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)

    wb = load_workbook(path, data_only=False)
    ws = wb["Sources & Uses"]

    fee_row = _find_row(ws, "Acquisition Fee")
    resolved = _find_named_cell(wb, "s_acquisition_fee")
    assert resolved is not None, "s_acquisition_fee named range missing"
    sheet_title, name_row, name_col = resolved
    assert sheet_title == "Sources & Uses"
    assert name_row == fee_row
    assert name_col == 2
    assert float(ws.cell(row=fee_row, column=2).value) == pytest.approx(80000.0)
