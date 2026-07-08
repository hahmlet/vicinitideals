"""Pro Forma Y1 Gross Revenue and Y1 OpEx must be formulas that trace back
to Assumptions Block F / G cells, not engine-computed hardcoded values.

Post-consolidation layout (commit e7ba809): the "Underwriting Pro Forma"
sheet emits one indented bullet row per stream / expense line ABOVE the
total row, and the total is a column-range SUM over those bullets. The
Block F / G references live on the bullet rows.

Contract:

  1. Y1 Gross Revenue total = ``=SUM(<col>{first_bullet}:<col>{last_bullet})``
     where each bullet's Y1 = ``=s_rev_<slug>_y1_monthly*12``.
  2. Y1 OpEx total = SUM over bullets whose Y1 = ``=s_opex_<slug>_annual``.
  3. Y0 must NOT use the stabilized Y1 input formula (construction-phase
     math differs from stabilized inputs).
  4. Y2+ keeps a growth chain off the prior year's cell — per-stream
     escalation on the bullet rows.
  5. Multi-project scenarios with duplicate stream labels get
     globally-resolved slugs (no silent collision overwrite).
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from app.models.deal import IncomeStream, IncomeStreamType, OperatingExpenseLine
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)
from tests.exporters._parity_helpers import (
    find_label_row,
    parse_sum_range,
    proforma_layout,
)


pytestmark = pytest.mark.asyncio


def _find_row(ws, label_text: str) -> int | None:
    label_col, _ = proforma_layout(ws)
    return find_label_row(ws, label_text, col=label_col, exact=True)


def _bullet_span(ws, total_row: int, y1_col: int) -> list[int]:
    """Bullet rows covered by the total row's Y1 ``=SUM(D4:D7)`` formula."""
    parsed = parse_sum_range(ws.cell(row=total_row, column=y1_col).value)
    if parsed is None:
        return []
    _, first, last = parsed
    return list(range(first, last + 1))


async def _seed_with_one_stream_one_opex(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="PF Y1 Formula")
    deal_model, _, stream, opex = await seed_deal_model_with_financials(
        session, opp, user
    )
    # Activate the seeded stream + opex in the stabilized phase so the
    # engine produces non-zero Y1 revenue/opex. Without this, engine
    # Y1 = 0 and the phase-gating in `_build_uw_proforma` (rightly)
    # suppresses the input-formula override, which would break the
    # assertions below that the Y1 cell IS a formula.
    stream.active_in_phases = ["stabilized"]
    opex.active_in_phases = ["stabilized"]
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


async def test_y1_gross_revenue_is_sum_of_block_f_y1_monthly_times_12(
    session: AsyncSession,
) -> None:
    scenario = await _seed_with_one_stream_one_opex(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    ws = wb["Underwriting Pro Forma"]
    gr_row = _find_row(ws, "Gross Revenue")
    assert gr_row is not None, "Gross Revenue row not found"
    _, y0_col = proforma_layout(ws)
    y1_col = y0_col + 1

    # Total = SUM over the per-stream bullet rows emitted above it.
    y1 = ws.cell(row=gr_row, column=y1_col).value
    assert isinstance(y1, str) and y1.startswith("="), (
        f"Y1 Gross Revenue must be a formula, got {y1!r}"
    )
    bullets = _bullet_span(ws, gr_row, y1_col)
    assert bullets, f"Y1 total must SUM the bullet rows; got {y1!r}"

    # Each bullet's Y1 annualizes its Block F monthly input cell.
    for br in bullets:
        bullet_y1 = ws.cell(row=br, column=y1_col).value
        assert isinstance(bullet_y1, str) and bullet_y1.startswith("="), (
            f"bullet row {br} Y1 must be a formula; got {bullet_y1!r}"
        )
        assert "s_rev_" in bullet_y1, (
            f"bullet Y1 must reference Block F cells: {bullet_y1}"
        )
        assert "_y1_monthly" in bullet_y1, (
            f"bullet Y1 must use _y1_monthly cells: {bullet_y1}"
        )
        assert "*12" in bullet_y1, (
            f"bullet Y1 must annualize via *12: {bullet_y1}"
        )


async def test_y1_opex_is_sum_of_block_g_annual_cells(
    session: AsyncSession,
) -> None:
    scenario = await _seed_with_one_stream_one_opex(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    ws = wb["Underwriting Pro Forma"]
    opex_row = _find_row(ws, "Operating Expenses")
    assert opex_row is not None, "Operating Expenses row not found"
    _, y0_col = proforma_layout(ws)
    y1_col = y0_col + 1

    y1 = ws.cell(row=opex_row, column=y1_col).value
    assert isinstance(y1, str) and y1.startswith("="), (
        f"Y1 OpEx must be a formula, got {y1!r}"
    )
    bullets = _bullet_span(ws, opex_row, y1_col)
    assert bullets, f"Y1 OpEx total must SUM the bullet rows; got {y1!r}"
    for br in bullets:
        bullet_y1 = ws.cell(row=br, column=y1_col).value
        assert isinstance(bullet_y1, str) and bullet_y1.startswith("="), (
            f"bullet row {br} Y1 must be a formula; got {bullet_y1!r}"
        )
        assert "s_opex_" in bullet_y1, (
            f"bullet Y1 must reference Block G: {bullet_y1}"
        )
        assert "_annual" in bullet_y1, (
            f"bullet Y1 must use _annual cells: {bullet_y1}"
        )


async def test_y0_gross_revenue_stays_engine_value(
    session: AsyncSession,
) -> None:
    """Y0 (pre-op) is construction-phase math — must NOT use the
    stabilized Y1 input formula."""
    scenario = await _seed_with_one_stream_one_opex(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    ws = wb["Underwriting Pro Forma"]
    gr_row = _find_row(ws, "Gross Revenue")
    _, y0_col = proforma_layout(ws)
    y1_col = y0_col + 1

    # Y0 must not be seeded from the stabilized Y1 input chain — neither
    # on the total row nor on any bullet row (bullet Y0 stays blank).
    y0 = ws.cell(row=gr_row, column=y0_col).value
    if isinstance(y0, str) and "_y1_monthly" in y0:
        pytest.fail(f"Y0 Gross Revenue must not use the Y1 input formula: {y0}")
    for br in _bullet_span(ws, gr_row, y1_col):
        bullet_y0 = ws.cell(row=br, column=y0_col).value
        assert bullet_y0 in (None, ""), (
            f"bullet row {br} Y0 must stay blank (construction-phase math "
            f"is engine-governed); got {bullet_y0!r}"
        )


async def test_y2_gross_revenue_growth_chain_still_works(
    session: AsyncSession,
) -> None:
    """Growth chain (Y2+ = prior * (1 + escalation)) must keep working
    on top of the Y1 input formula — post-consolidation the chain lives
    on each per-stream bullet row via ``s_rev_<slug>_escalation_pct``."""
    scenario = await _seed_with_one_stream_one_opex(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    ws = wb["Underwriting Pro Forma"]
    gr_row = _find_row(ws, "Gross Revenue")
    _, y0_col = proforma_layout(ws)
    y1_col, y2_col = y0_col + 1, y0_col + 2

    y2 = ws.cell(row=gr_row, column=y2_col).value
    assert isinstance(y2, str) and y2.startswith("="), (
        f"Y2 Gross Revenue must be a formula, got {y2!r}"
    )
    bullets = _bullet_span(ws, gr_row, y1_col)
    assert bullets, "Y2 total must aggregate the bullet rows"
    prev_col = get_column_letter(y1_col)
    for br in bullets:
        bullet_y2 = ws.cell(row=br, column=y2_col).value
        assert isinstance(bullet_y2, str) and bullet_y2.startswith("="), (
            f"bullet row {br} Y2 must be a growth formula; got {bullet_y2!r}"
        )
        assert "s_rev_" in bullet_y2 and "_escalation_pct" in bullet_y2, (
            f"bullet Y2 must reference per-stream escalation: {bullet_y2}"
        )
        assert f"{prev_col}{br}" in bullet_y2, (
            f"bullet Y2 must chain off its prior-year cell "
            f"{prev_col}{br}: {bullet_y2}"
        )


async def _seed_multi_project_duplicate_stream_labels(session: AsyncSession):
    """Two projects, each with a stream labeled '1BR Units' — verifies
    cross-project slug collision resolves with _2 suffix."""
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Multi-Proj Dup Labels")
    deal_model, _, stream, opex = await seed_deal_model_with_financials(
        session, opp, user
    )
    # Activate primary stream + opex in stabilized phase (see
    # _seed_with_one_stream_one_opex for the why).
    stream.active_in_phases = ["stabilized"]
    opex.active_in_phases = ["stabilized"]
    await session.flush()

    extra = Project(
        id=uuid4(), scenario_id=deal_model.id,
        opportunity_id=None, name="Second",
    )
    session.add(extra)
    await session.flush()
    from app.models.deal import OperationalInputs as _OI
    session.add(_OI(
        id=uuid4(), project_id=extra.id,
        unit_count_new=4, exit_cap_rate_pct=Decimal("5.5"),
    ))
    await session.flush()
    # The seed already added "1BR Units" on the primary; add the same
    # label on Second so the two collide cross-project.
    session.add(IncomeStream(
        id=uuid4(), project_id=extra.id,
        stream_type=IncomeStreamType.residential_rent,
        label="1BR Units",
        unit_count=4, amount_per_unit_monthly=Decimal("1500"),
        stabilized_occupancy_pct=Decimal("95"),
        escalation_rate_pct_annual=Decimal("3.0"),
        active_in_phases=["stabilized"],
    ))
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


async def test_cross_project_duplicate_stream_labels_dedupe_globally(
    session: AsyncSession,
) -> None:
    """Two streams on different projects with the same label must
    resolve to distinct slugs (``1br_units`` + ``1br_units_2``) — not
    silently shadow each other."""
    scenario = await _seed_multi_project_duplicate_stream_labels(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    defined = {name for name in wb.defined_names}
    assert "s_rev_1br_units_unit_count" in defined
    assert "s_rev_1br_units_2_unit_count" in defined, (
        "cross-project label collision must resolve with _2 suffix"
    )

    # Pro Forma Y1 total must aggregate bullet rows referencing BOTH
    # dedup'd cells (one bullet per stream, distinct slugs).
    ws = wb["Underwriting Pro Forma"]
    gr_row = _find_row(ws, "Gross Revenue")
    _, y0_col = proforma_layout(ws)
    y1_col = y0_col + 1
    bullets = _bullet_span(ws, gr_row, y1_col)
    assert bullets, "Gross Revenue total must SUM the bullet rows"
    bullet_formulas = "|".join(
        str(ws.cell(row=br, column=y1_col).value) for br in bullets
    )
    assert "s_rev_1br_units_y1_monthly" in bullet_formulas
    assert "s_rev_1br_units_2_y1_monthly" in bullet_formulas
