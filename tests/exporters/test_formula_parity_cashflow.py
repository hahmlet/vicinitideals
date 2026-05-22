"""Engine-vs-formula parity for the Underwriting Cash Flow sheet.

Commit 4 of docs/feature-plans/investor-excel-formula-conversion.md §4.2.

Scope: derived rows that are clean arithmetic of other rows on the same
sheet — Levered CF, Unlevered CF, DSCR, Cumulative CF. NOI / Capital
Events / Debt Proceeds / Debt Service stay as engine values; their
formulas land later when the Pro Forma → Cash Flow cross-sheet wiring
is built (commit 4 narrow scope, plan §7).

Parity loop per row:
  1. Cell is a formula (string starting with ``=``) for every Y0..Yn col
  2. Formula references the right operand cells on the sheet
  3. After Excel recalc, the cell value matches the engine within $1
     (or 0.001 for DSCR ratios)
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.exporters.investor_export import (
    _aggregate_scenario_annual,
    _coerce_decimal,
    _funder_class,
    _load_all,
    _signed_capital_events_by_year,
    export_investor_workbook,
)
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)
from tests.exporters._parity_helpers import (
    RecalcUnavailableError,
    recalc_workbook,
)


async def _seed_scenario(session: AsyncSession):
    org, user = await seed_org(session)
    opportunity = await seed_opportunity(
        session, org, user, name="CashFlow-Parity Smoke"
    )
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opportunity, user
    )
    return deal_model


def _cashflow_sheet(blob: bytes, data_only: bool):
    wb = load_workbook(BytesIO(blob), data_only=data_only)
    return wb, wb["Underwriting Cash Flow"]


def _find_row_by_label(ws, label: str) -> int | None:
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == label:
            return r
    return None


async def test_levered_row_is_formula_each_year(session: AsyncSession):
    """Levered CF cells must be formulas, not engine scalars."""
    scenario = await _seed_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    _wb, ws = _cashflow_sheet(blob, data_only=False)

    lev_row = _find_row_by_label(ws, "Levered Cash Flow")
    assert lev_row is not None

    formula_count = 0
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=lev_row, column=c).value
        if v is None:
            break
        assert isinstance(v, str) and v.startswith("="), (
            f"Levered Y{c - 2} should be a formula; got {v!r}"
        )
        # Levered = NOI + CapEvt [+ DebtProceeds] - DebtSvc, so the
        # formula must contain a minus sign.
        assert "-" in v, f"Levered formula expected to contain '-'; got {v!r}"
        formula_count += 1
    assert formula_count > 0


async def test_dscr_row_is_formula_with_iferror(session: AsyncSession):
    """DSCR cells must use IFERROR(NOI/DS,0) to avoid #DIV/0! at Y0."""
    scenario = await _seed_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    _wb, ws = _cashflow_sheet(blob, data_only=False)

    dscr_row = _find_row_by_label(ws, "DSCR (annual)")
    assert dscr_row is not None

    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=dscr_row, column=c).value
        if v is None:
            break
        assert isinstance(v, str) and v.startswith("=IFERROR("), (
            f"DSCR Y{c - 2} should use =IFERROR; got {v!r}"
        )
        assert "/" in v, f"DSCR formula expected to contain '/'; got {v!r}"


async def test_cumulative_row_is_running_sum_formula(session: AsyncSession):
    """Cumulative CF cells must be SUM($B$lev_row:<col><lev_row>) formulas."""
    scenario = await _seed_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    _wb, ws = _cashflow_sheet(blob, data_only=False)

    cum_row = _find_row_by_label(ws, "Cumulative Cash Flow")
    assert cum_row is not None

    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=cum_row, column=c).value
        if v is None:
            break
        assert isinstance(v, str) and v.startswith("=SUM("), (
            f"Cumulative Y{c - 2} should be =SUM(...); got {v!r}"
        )
        # Must be a running sum anchored at $B$<row> — the dollar-sign
        # anchor proves the range start is fixed (won't drift on copy).
        assert "$B$" in v, (
            f"Cumulative formula must anchor at $B$<row>; got {v!r}"
        )


async def test_unlevered_row_is_formula_each_year(session: AsyncSession):
    """Unlevered cells must be ``=Levered + DebtService`` formulas."""
    scenario = await _seed_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    _wb, ws = _cashflow_sheet(blob, data_only=False)

    unl_row = _find_row_by_label(ws, "Unlevered Cash Flow")
    assert unl_row is not None

    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=unl_row, column=c).value
        if v is None:
            break
        assert isinstance(v, str) and v.startswith("="), (
            f"Unlevered Y{c - 2} should be a formula; got {v!r}"
        )
        assert "+" in v, f"Unlevered formula expected to contain '+'; got {v!r}"


def _expected_levered(annual, capital_events_by_year, debt_y0, year_cols):
    """Replay the engine arithmetic so test assertions don't depend on
    whichever row in ``annual`` holds the cached engine NCF."""
    out: dict[int, Decimal] = {}
    for y in year_cols:
        noi = annual.get(y, {}).get("noi", Decimal(0))
        cap = capital_events_by_year.get(y, Decimal(0))
        dp = debt_y0 if y == 0 else Decimal(0)
        ds = annual.get(y, {}).get("debt_service", Decimal(0))
        out[y] = noi + cap + dp - ds
    return out


async def test_levered_evaluates_to_engine_value(
    session: AsyncSession, tmp_path: Path
):
    """Excel-recalc'd Levered CF == NOI + CapEvt + DebtProceeds - DebtSvc."""
    scenario = await _seed_scenario(session)
    ctx = await _load_all(session, scenario.id)
    annual = _aggregate_scenario_annual(ctx["cash_flows"])
    capital_events_by_year = _signed_capital_events_by_year(ctx["cash_flow_items"])

    _junc_dp: dict = {}
    for _j in ctx.get("junctions", []):
        _junc_dp[_j.capital_module_id] = _junc_dp.get(
            _j.capital_module_id, Decimal(0)
        ) + _coerce_decimal(_j.amount or 0)
    debt_y0 = Decimal(0)
    for _m in ctx.get("capital_modules", []):
        if _funder_class(_m) == "Debt":
            debt_y0 += _junc_dp.get(_m.id) or _coerce_decimal(
                (_m.source or {}).get("amount") or 0
            )
    if debt_y0 <= Decimal(1):
        debt_y0 = Decimal(0)

    max_year = min(max(annual) if annual else 0, 10)
    year_cols = list(range(0, max(max_year, 1) + 1))
    expected = _expected_levered(annual, capital_events_by_year, debt_y0, year_cols)

    blob = await export_investor_workbook(scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)
    try:
        recalc_workbook(path)
    except RecalcUnavailableError as exc:
        pytest.skip(f"no recalc backend: {exc}")

    wb = load_workbook(path, data_only=True)
    ws = wb["Underwriting Cash Flow"]
    lev_row = _find_row_by_label(ws, "Levered Cash Flow")
    assert lev_row is not None

    for col_offset, year in enumerate(year_cols):
        engine_value = float(expected.get(year, Decimal(0)))
        excel_value = ws.cell(row=lev_row, column=2 + col_offset).value
        if excel_value is None:
            excel_value = 0
        diff = abs(float(excel_value) - engine_value)
        assert diff < 1.0, (
            f"Levered Y{year} parity: engine={engine_value}, "
            f"excel={excel_value}, diff={diff}"
        )


async def test_dscr_evaluates_to_engine_value(
    session: AsyncSession, tmp_path: Path
):
    """Excel-recalc'd DSCR == NOI / DS per year (within 0.001)."""
    scenario = await _seed_scenario(session)
    ctx = await _load_all(session, scenario.id)
    annual = _aggregate_scenario_annual(ctx["cash_flows"])

    blob = await export_investor_workbook(scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)
    try:
        recalc_workbook(path)
    except RecalcUnavailableError as exc:
        pytest.skip(f"no recalc backend: {exc}")

    wb = load_workbook(path, data_only=True)
    ws = wb["Underwriting Cash Flow"]
    dscr_row = _find_row_by_label(ws, "DSCR (annual)")
    assert dscr_row is not None

    max_year = min(max(annual) if annual else 0, 10)
    year_cols = list(range(0, max(max_year, 1) + 1))

    for col_offset, year in enumerate(year_cols):
        noi = annual.get(year, {}).get("noi", Decimal(0))
        ds = annual.get(year, {}).get("debt_service", Decimal(0))
        if ds and ds != 0:
            expected = float(noi / ds)
        else:
            expected = 0.0
        excel_value = ws.cell(row=dscr_row, column=2 + col_offset).value
        if excel_value is None:
            excel_value = 0
        diff = abs(float(excel_value) - expected)
        assert diff < 1e-3, (
            f"DSCR Y{year} parity: engine={expected}, "
            f"excel={excel_value}, diff={diff}"
        )


async def test_cumulative_evaluates_to_running_sum(
    session: AsyncSession, tmp_path: Path
):
    """Excel-recalc'd Cumulative[n] == sum(Levered[0..n])."""
    scenario = await _seed_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)
    try:
        recalc_workbook(path)
    except RecalcUnavailableError as exc:
        pytest.skip(f"no recalc backend: {exc}")

    wb = load_workbook(path, data_only=True)
    ws = wb["Underwriting Cash Flow"]
    lev_row = _find_row_by_label(ws, "Levered Cash Flow")
    cum_row = _find_row_by_label(ws, "Cumulative Cash Flow")
    assert lev_row is not None and cum_row is not None

    running = 0.0
    for c in range(2, ws.max_column + 1):
        lev_val = ws.cell(row=lev_row, column=c).value
        cum_val = ws.cell(row=cum_row, column=c).value
        if lev_val is None and cum_val is None:
            break
        running += float(lev_val or 0)
        excel_cum = float(cum_val or 0)
        diff = abs(excel_cum - running)
        assert diff < 1.0, (
            f"Cumulative col {c} parity: expected running={running}, "
            f"excel={excel_cum}, diff={diff}"
        )
