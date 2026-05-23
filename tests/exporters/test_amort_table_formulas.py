"""Phase D commit 7: Debt Schedule Amortization table is wired as
CUMIPMT / CUMPRINC formulas over the Loan Summary's principal/rate/
amort/IO cells so LP edits flow through.

Contract:

  1. Year-1 Beg Balance cell pulls principal absolute-ref ($C${perm_row}).
  2. Later years' Beg Balance reaches up to prior row's End Balance.
  3. Annual Payment formula uses PMT (or IO branch when IO covers year).
  4. Interest formula uses CUMIPMT (or IO branch).
  5. Principal = AnnualPmt − Interest; End = Beg − Principal.
  6. No bare numeric cells in the amort table (every value is a formula).
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from app.models.capital import (
    CapitalModule,
    CapitalModuleProject,
    EquityRole,
    VehicleType,
)
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


async def _seed(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Amort Formula Smoke")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()
    debt = CapitalModule(
        scenario_id=deal_model.id,
        label="Senior Loan",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={
            "amount": "500000", "interest_rate_pct": 6.5,
            "amort_term_years": 30, "hold_term_years": 10,
        },
        carry={"carry_type": "pi", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    equity = CapitalModule(
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type=VehicleType.equity.value,
        equity_role=EquityRole.lp.value,
        stack_position=2,
        source={"amount": "250000"},
        carry={"carry_type": "none", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    session.add_all([debt, equity])
    await session.flush()
    session.add_all([
        CapitalModuleProject(
            capital_module_id=debt.id, project_id=project.id,
            amount=Decimal("500000"),
        ),
        CapitalModuleProject(
            capital_module_id=equity.id, project_id=project.id,
            amount=Decimal("250000"),
        ),
    ])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _amort_rows(ws) -> list[int]:
    """Return row numbers of amort table data rows (after the
    'Amortization —' section header + header row)."""
    header_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.startswith("Amortization"):
            header_row = r
            break
    if header_row is None:
        return []
    data_start = header_row + 2  # +1 section header, +1 col header
    out = []
    for r in range(data_start, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, int):
            out.append(r)
        else:
            break
    return out


@pytest.mark.parametrize("profile", ["internal", "lender"])
async def test_amort_table_all_cells_are_formulas(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Debt Schedule"]

    rows = _amort_rows(ws)
    assert rows, f"profile={profile}: amort table data rows not found"

    for r in rows:
        for col in (2, 3, 4, 5, 6):  # Beg, Pmt, Int, Prin, End
            v = ws.cell(row=r, column=col).value
            assert isinstance(v, str) and v.startswith("="), (
                f"profile={profile}: row {r} col {col} not a formula; got {v!r}"
            )


@pytest.mark.parametrize("profile", ["internal", "lender"])
async def test_amort_table_uses_cumipmt(session: AsyncSession, profile: str):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Debt Schedule"]

    rows = _amort_rows(ws)
    assert rows
    first = rows[0]
    interest_formula = ws.cell(row=first, column=4).value
    assert "CUMIPMT" in interest_formula, (
        f"profile={profile}: interest cell must use CUMIPMT; got {interest_formula!r}"
    )
    payment_formula = ws.cell(row=first, column=3).value
    assert "PMT" in payment_formula, (
        f"profile={profile}: payment cell must use PMT; got {payment_formula!r}"
    )


@pytest.mark.parametrize("profile", ["internal", "lender"])
async def test_amort_beg_balance_chain(session: AsyncSession, profile: str):
    """Year-1 Beg Balance pulls principal; year-N Beg Balance = prior End."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Debt Schedule"]

    rows = _amort_rows(ws)
    assert len(rows) >= 2

    y1_beg = ws.cell(row=rows[0], column=2).value
    assert y1_beg.startswith("=$C$"), (
        f"profile={profile}: Y1 Beg must be absolute principal ref; got {y1_beg!r}"
    )

    y2_beg = ws.cell(row=rows[1], column=2).value
    expected = f"=F{rows[0]}"
    assert y2_beg == expected, (
        f"profile={profile}: Y2 Beg must reach prior End ({expected}); got {y2_beg!r}"
    )


@pytest.mark.parametrize("profile", ["internal", "lender"])
async def test_amort_principal_and_end_are_derivations(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Debt Schedule"]

    rows = _amort_rows(ws)
    assert rows
    first = rows[0]
    principal_cell = ws.cell(row=first, column=5).value
    end_cell = ws.cell(row=first, column=6).value
    assert principal_cell == f"=C{first}-D{first}", (
        f"profile={profile}: principal cell must be Pmt−Int; got {principal_cell!r}"
    )
    assert end_cell == f"=B{first}-E{first}", (
        f"profile={profile}: end cell must be Beg−Principal; got {end_cell!r}"
    )
