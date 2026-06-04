"""Slice 1 (Export v3): Operating Deficit Reserve must render on the S&U
sheet, expose the named cell ``s_odr_amount``, and be flagged as
balance-only so it does NOT inflate Total Project Cost on UW Summary.

The engine's module-level ``_BALANCE_ONLY_LABELS`` set was missing
"Operating Deficit Reserve" through 2026-06-03 — present only in a
local set inside ``_auto_size_debt_modules``. The exporter imports the
module-level set; with ODR missing the exporter:

  - rendered the row (good)
  - did NOT add it to ``balance_only_refs`` (bad — TPC over-stated)

This test seeds an ODR UseLine directly (no curves needed), exports,
and asserts the structural contract:

  1. row exists with the ``Operating Deficit Reserve`` label
  2. ``s_odr_amount`` named range points at its B column cell
  3. the row's B-cell ref appears in the Balance-Only Subtotal
     formula so ``s_su_balance_only_total`` includes it
  4. TPC formula on UW Summary remains
     ``IFERROR(s_su_uses_total-s_su_balance_only_total, ...)`` —
     unchanged, but now picks ODR up via the (4) wiring

No recalc backend is needed; this is a wiring/structure test.
"""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.exporters.investor_export import export_investor_workbook
from app.models.deal import UseLine
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


pytestmark = pytest.mark.asyncio


async def _seed_with_odr(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="ODR Balance-Only")
    deal_model, _inputs, _stream, _opex = (
        await seed_deal_model_with_financials(session, opp, user)
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()

    # Acquisition hard cost so the S&U sheet has a non-trivial Uses base
    # and the balance-only subtotal isn't the whole Uses total.
    session.add(
        UseLine(
            project_id=project.id,
            label="Land Acquisition",
            amount=Decimal("4000000"),
            phase="acquisition",
            cost_category="hard",
        )
    )
    # Operating Deficit Reserve — seeded directly. The engine writes
    # this label during ``_auto_size_debt_modules`` whenever the
    # lease-up curves leave a revenue/OpEx gap. We bypass the engine
    # here because the wiring contract (balance-only treatment) is
    # independent of how the row was created.
    session.add(
        UseLine(
            project_id=project.id,
            label="Operating Deficit Reserve",
            amount=Decimal("250000"),
            phase="lease_up",
            cost_category="soft",
        )
    )
    await session.flush()
    await session.commit()
    return deal_model, project


def _find_row(ws, needle: str) -> int | None:
    needle_norm = needle.strip().lower()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and needle_norm in v.strip().lower():
            return r
    return None


def _find_named_cell(wb, name: str) -> tuple[str, int, int] | None:
    """Resolve a defined name to (sheet, row, col). Returns None if absent."""
    dn = wb.defined_names.get(name)
    if dn is None:
        return None
    for sheet_title, coord in dn.destinations:
        # coord like "'Sources & Uses'!$B$42" — strip $.
        cell = coord.replace("$", "")
        col_letters = "".join(c for c in cell if c.isalpha())
        row_digits = "".join(c for c in cell if c.isdigit())
        col_idx = 0
        for ch in col_letters:
            col_idx = col_idx * 26 + (ord(ch.upper()) - ord("A") + 1)
        return sheet_title, int(row_digits), col_idx
    return None


async def test_odr_row_rendered_and_marked_balance_only(
    session: AsyncSession, tmp_path: Path
) -> None:
    scenario, project = await _seed_with_odr(session)

    blob = await export_investor_workbook(scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)

    wb = load_workbook(path, data_only=False)
    ws = wb["Sources & Uses"]

    # (1) ODR row exists.
    odr_row = _find_row(ws, "Operating Deficit Reserve")
    assert odr_row is not None, "ODR row not rendered on S&U sheet"
    assert float(ws.cell(row=odr_row, column=2).value) == pytest.approx(250000.0)

    # (2) s_odr_amount named range resolves to that cell.
    resolved = _find_named_cell(wb, "s_odr_amount")
    assert resolved is not None, "s_odr_amount named range missing"
    sheet_title, name_row, name_col = resolved
    assert sheet_title == "Sources & Uses"
    assert name_row == odr_row
    assert name_col == 2

    # (3) Balance-Only Subtotal formula references the ODR row's B-cell.
    bo_row = _find_row(ws, "Balance-Only Subtotal")
    assert bo_row is not None, "Balance-Only Subtotal row missing"
    bo_formula = ws.cell(row=bo_row, column=2).value
    assert isinstance(bo_formula, str) and bo_formula.startswith("=")
    assert f"B{odr_row}" in bo_formula, (
        f"Balance-Only Subtotal formula does not include ODR row "
        f"B{odr_row}. Got: {bo_formula}. The engine's "
        f"_BALANCE_ONLY_LABELS set must include 'Operating Deficit "
        f"Reserve' for the exporter to wire this correctly."
    )


async def test_odr_excluded_from_tpc_formula(
    session: AsyncSession, tmp_path: Path
) -> None:
    """TPC formula on UW Summary must reference both Uses and
    Balance-Only Subtotal so ODR is subtracted out."""
    scenario, project = await _seed_with_odr(session)

    blob = await export_investor_workbook(scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)

    wb = load_workbook(path, data_only=False)
    ws_uw = wb["Underwriting Summary"]
    tpc_resolved = _find_named_cell(wb, "s_total_project_cost")
    assert tpc_resolved is not None, "s_total_project_cost named range missing"
    _, tpc_row, tpc_col = tpc_resolved
    tpc_formula = ws_uw.cell(row=tpc_row, column=tpc_col).value
    assert isinstance(tpc_formula, str) and tpc_formula.startswith("=")
    # Must subtract balance-only from total uses for TPC to exclude ODR.
    assert "s_su_uses_total" in tpc_formula
    assert "s_su_balance_only_total" in tpc_formula
