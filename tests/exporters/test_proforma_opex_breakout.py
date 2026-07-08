"""Pro Forma must show one indented breakout row per OpEx line grouped
with the Operating Expenses total. Each breakout row references the
corresponding Assumptions Block G cells so an LP can trace any single
expense back to its input — and the per-line escalation rate carries
the row forward through the Y2+ growth chain.

Contract (post-consolidation layout, commit e7ba809 — bullets emit
directly ABOVE the total, and the total SUMs the bullet column range):

  1. Total "Operating Expenses" row is a formula aggregating the
     Block-G-driven bullet rows (``=SUM(<col>{first}:<col>{last})``).
  2. One breakout row per OperatingExpenseLine appears directly above
     the total, indented (label starts with ``   •``) and inside the
     total's SUM range.
  3. Each breakout row's Y0 cell is blank (construction-phase OpEx
     differs from stabilized inputs).
  4. Each breakout row's Y1 cell = ``=s_opex_<slug>_annual``.
  5. Each breakout row's Y2 cell = the prior-year cell times
     ``(1+s_opex_<slug>_escalation_pct)`` — per-line escalation, not
     the sheet-wide rate, so different expense categories can ramp at
     different speeds.
"""
from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)
from tests.exporters._parity_helpers import (
    find_label_row,
    parse_sum_range,
    proforma_layout,
)


pytestmark = pytest.mark.asyncio


async def _seed(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="PF OpEx Breakout")
    deal_model, _, stream, opex = await seed_deal_model_with_financials(
        session, opp, user
    )
    stream.active_in_phases = ["stabilized"]
    opex.active_in_phases = ["stabilized"]
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find_row(ws, label_exact: str) -> int | None:
    label_col, _ = proforma_layout(ws)
    return find_label_row(ws, label_exact, col=label_col, exact=True)


def _find_indented_row(ws, label_substr: str) -> int | None:
    """Locate a breakout row whose label contains the substring AND
    starts with the indent marker ``•``."""
    label_col, _ = proforma_layout(ws)
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=label_col).value
        if (
            isinstance(v, str)
            and "•" in v
            and label_substr.lower() in v.lower()
        ):
            return r
    return None


async def test_opex_total_row_unchanged_phase_2_formula(
    session: AsyncSession,
) -> None:
    """Total Operating Expenses row stays formula-driven off Block G —
    a SUM over the bullet rows whose Y1 cells reference Block G."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Pro Forma"]
    row = _find_row(ws, "Operating Expenses")
    assert row is not None, "Operating Expenses total row missing"
    _, y0_col = proforma_layout(ws)
    y1 = ws.cell(row=row, column=y0_col + 1).value
    assert parse_sum_range(y1) is not None, (
        f"Total OpEx Y1 must SUM the Block-G bullet rows; got {y1!r}"
    )


async def test_opex_breakout_row_appears_below_total(
    session: AsyncSession,
) -> None:
    """The seeded ``Property Management`` line gets an indented row
    grouped with (directly above) the Operating Expenses total and
    inside the total's SUM range."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Pro Forma"]
    total_row = _find_row(ws, "Operating Expenses")
    assert total_row is not None
    breakout = _find_indented_row(ws, "Property Management")
    assert breakout is not None, (
        "Property Management breakout row missing for Operating Expenses"
    )
    _, y0_col = proforma_layout(ws)
    parsed = parse_sum_range(ws.cell(row=total_row, column=y0_col + 1).value)
    assert parsed is not None, "total must SUM the bullet rows"
    _, first, last = parsed
    assert first <= breakout <= last < total_row, (
        f"breakout row ({breakout}) must sit inside the total's SUM range "
        f"({first}..{last}) above the total ({total_row})"
    )


async def test_breakout_y0_is_blank(session: AsyncSession) -> None:
    """Y0 stays empty — construction-phase OpEx is engine-governed."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Pro Forma"]
    breakout = _find_indented_row(ws, "Property Management")
    assert breakout is not None
    _, y0_col = proforma_layout(ws)
    y0 = ws.cell(row=breakout, column=y0_col).value
    assert y0 in (None, ""), f"breakout Y0 must be blank; got {y0!r}"


async def test_breakout_y1_references_block_g_annual_cell(
    session: AsyncSession,
) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Pro Forma"]
    breakout = _find_indented_row(ws, "Property Management")
    assert breakout is not None
    _, y0_col = proforma_layout(ws)
    y1 = ws.cell(row=breakout, column=y0_col + 1).value
    assert isinstance(y1, str) and y1.startswith("="), f"got {y1!r}"
    assert "s_opex_" in y1 and "_annual" in y1, (
        f"breakout Y1 must reference s_opex_<slug>_annual; got {y1!r}"
    )


async def test_breakout_y2_uses_per_line_escalation(
    session: AsyncSession,
) -> None:
    """Y2 = prior-year cell × (1 + per-line escalation_pct), not the
    sheet-wide s_opex_growth_rate."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Pro Forma"]
    breakout = _find_indented_row(ws, "Property Management")
    assert breakout is not None
    _, y0_col = proforma_layout(ws)
    y1_col, y2_col = y0_col + 1, y0_col + 2
    y2 = ws.cell(row=breakout, column=y2_col).value
    assert isinstance(y2, str) and y2.startswith("="), f"got {y2!r}"
    assert "s_opex_" in y2 and "_escalation_pct" in y2, (
        f"breakout Y2 must reference per-line s_opex_<slug>_escalation_pct; "
        f"got {y2!r}"
    )
    # Must reference the breakout row's own prior column, not the total.
    prev_ref = f"{get_column_letter(y1_col)}{breakout}"
    assert prev_ref in y2, (
        f"breakout Y2 must reference breakout's own prior-year cell "
        f"{prev_ref}; got {y2!r}"
    )
