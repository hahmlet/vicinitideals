"""Phase D: UW Summary "Property Valuation" block must use formulas so
LP edits to NOI / cap rate inputs re-derive Going-In Cap Value, Exit Cap
Value, and Cap Spread without re-running the engine.

Contract:

  1. Going-In Cap Value cell = ``=IFERROR(s_combined_noi/s_going_in_cap_rate,"")``
  2. Exit Cap Value cell      = ``=IFERROR(s_exit_year_noi/s_exit_cap_rate,"")``
  3. Cap Spread cell          = ``=IFERROR(s_yield_on_cost-s_going_in_cap_rate,"")``
  4. All referenced named ranges are registered on the workbook.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from app.models.capital import (
    CapitalModule,
    CapitalModuleProject,
    EquityRole,
    VehicleType,
)
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


async def _seed(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Prop Val Smoke")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()
    debt = CapitalModule(
        scenario_id=deal_model.id,
        label="Senior Loan",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={
            "amount": "500000", "interest_rate_pct": 6.5,
            "amort_term_years": 30, "hold_term_years": 10,
        },
        carry={"carry_type": "pi", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    equity = CapitalModule(
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type=VehicleType.equity.value,
        equity_role=EquityRole.lp.value,
        stack_position=2,
        source={"amount": "250000"},
        carry={"carry_type": "none", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    session.add_all([debt, equity])
    await session.flush()
    session.add_all([
        CapitalModuleProject(
            capital_module_id=debt.id, project_id=project.id,
            amount=Decimal("500000"),
        ),
        CapitalModuleProject(
            capital_module_id=equity.id, project_id=project.id,
            amount=Decimal("250000"),
        ),
    ])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find_row(ws, label_substr: str) -> int | None:
    needle = label_substr.lower()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and needle in v.lower():
            return r
    return None


@pytest.mark.parametrize("profile", ["internal", "lp", "lender"])
async def test_going_in_cap_value_is_formula(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Summary"]
    r = _find_row(ws, "going-in cap value")
    assert r is not None, f"profile={profile}: Going-In Cap Value row missing"
    val = ws.cell(row=r, column=2).value
    assert isinstance(val, str) and val.startswith("="), (
        f"profile={profile}: must be formula; got {val!r}"
    )
    assert "s_combined_noi" in val and "s_going_in_cap_rate" in val, (
        f"profile={profile}: missing operand refs; got {val!r}"
    )


@pytest.mark.parametrize("profile", ["internal", "lp", "lender"])
async def test_exit_cap_value_is_formula(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Summary"]
    r = _find_row(ws, "exit cap value")
    assert r is not None, f"profile={profile}: Exit Cap Value row missing"
    val = ws.cell(row=r, column=2).value
    assert isinstance(val, str) and val.startswith("="), (
        f"profile={profile}: must be formula; got {val!r}"
    )
    assert "s_exit_year_noi" in val and "s_exit_cap_rate" in val, (
        f"profile={profile}: missing operand refs; got {val!r}"
    )


@pytest.mark.parametrize("profile", ["internal", "lp", "lender"])
async def test_cap_spread_is_formula(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Summary"]
    r = _find_row(ws, "cap spread")
    assert r is not None, f"profile={profile}: Cap Spread row missing"
    val = ws.cell(row=r, column=2).value
    assert isinstance(val, str) and val.startswith("="), (
        f"profile={profile}: must be formula; got {val!r}"
    )
    assert "s_yield_on_cost" in val and "s_going_in_cap_rate" in val, (
        f"profile={profile}: missing operand refs; got {val!r}"
    )


@pytest.mark.parametrize("profile", ["internal", "lp", "lender"])
async def test_property_valuation_named_ranges_registered(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    for required in (
        "s_combined_noi",
        "s_going_in_cap_rate",
        "s_exit_year_noi",
        "s_exit_cap_rate",
        "s_yield_on_cost",
        "s_going_in_cap_value",
        "s_direct_cap_value",
        "s_cap_spread",
    ):
        assert required in wb.defined_names, (
            f"profile={profile} missing {required} defined name"
        )
