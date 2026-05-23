"""Phase D: UW Summary "Yield on Cost (NOI ÷ TPC)" must be a formula
``=IFERROR(s_combined_noi/s_su_uses_total,"")`` so it tracks LP edits to
revenue / OpEx (through ``s_combined_noi``) and Use-line edits (through
``s_su_uses_total`` on the S&U sheet).

Contract:

  1. Cell is a formula starting with ``=`` and referencing both operands.
  2. Both operands are registered defined names.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from app.models.capital import (
    CapitalModule,
    CapitalModuleProject,
    EquityRole,
    VehicleType,
)
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


async def _seed(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Yield On Cost Smoke")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()
    debt = CapitalModule(
        scenario_id=deal_model.id,
        label="Senior Loan",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={
            "amount": "500000", "interest_rate_pct": 6.5,
            "amort_term_years": 30, "hold_term_years": 10,
        },
        carry={"carry_type": "pi", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    equity = CapitalModule(
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type=VehicleType.equity.value,
        equity_role=EquityRole.lp.value,
        stack_position=2,
        source={"amount": "250000"},
        carry={"carry_type": "none", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    session.add_all([debt, equity])
    await session.flush()
    session.add_all([
        CapitalModuleProject(
            capital_module_id=debt.id, project_id=project.id,
            amount=Decimal("500000"),
        ),
        CapitalModuleProject(
            capital_module_id=equity.id, project_id=project.id,
            amount=Decimal("250000"),
        ),
    ])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find_row(ws, label_substr: str) -> int | None:
    needle = label_substr.lower()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and needle in v.lower():
            return r
    return None


@pytest.mark.parametrize("profile", ["internal", "lp", "lender"])
async def test_yield_on_cost_is_formula(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)

    ws = wb["Underwriting Summary"]
    yoc_row = _find_row(ws, "yield on cost")
    assert yoc_row is not None, f"profile={profile} missing Yield on Cost row"

    val = ws.cell(row=yoc_row, column=2).value
    assert isinstance(val, str) and val.startswith("="), (
        f"profile={profile}: Yield on Cost must be formula; got {val!r}"
    )
    assert "s_combined_noi" in val and "s_su_uses_total" in val, (
        f"profile={profile}: formula must reference both named operands; "
        f"got {val!r}"
    )


@pytest.mark.parametrize("profile", ["internal", "lp", "lender"])
async def test_yield_on_cost_operands_registered(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    for required in ("s_combined_noi", "s_su_uses_total"):
        assert required in wb.defined_names, (
            f"profile={profile} missing {required} defined name"
        )
