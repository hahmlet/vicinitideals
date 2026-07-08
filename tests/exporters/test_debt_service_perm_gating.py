"""Per-year Debt Service formula must gate construction-to-perm loans
on the loan's perm-origination month so PI doesn't accrue during the
construction window.

Contract:

  1. For a PI loan whose Debt Schedule row registered an
     ``s_loan_<n>_perm_origination_month`` cell, each Y2+ Debt Service
     formula term wraps that loan in
     ``AND(s_loan_{i}_term_months >= Y*12,
           Y*12 >= s_loan_{i}_perm_origination_month)`` —
     not just the legacy term-only gate.

  2. For a loan with no registered perm cell (pure-acquisition scenario
     with no construction-side phase, or a loan funding no eligible
     project), the formula keeps the legacy
     ``IF(term_months >= Y*12, annual_pi, 0)`` form so backward
     compatibility is preserved.

  3. Single-project scenarios also gate on perm origination — the phase
     plan emits onto the Assumptions sheet (Block E) when the
     per-project sheet is suppressed, so single-project value_add deals
     get the same perm-aware formula multi-project does.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from app.models.capital import (
    CapitalModule,
    CapitalModuleProject,
    EquityRole,
    VehicleType,
)
from app.models.deal import OperationalInputs
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)
from tests.exporters._parity_helpers import find_label_row, proforma_layout


pytestmark = pytest.mark.asyncio


async def _seed(session: AsyncSession, *, multi_project: bool):
    """Seed a value_add scenario with a PI debt module. ``multi_project``
    toggles whether a second project is added — only multi-project
    scenarios render per-project sheets and thus register the
    ``p<n>_perm_origination_month`` cells the gating depends on."""
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="DS Perm Gate")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )

    if multi_project:
        extra = Project(
            id=uuid4(),
            scenario_id=deal_model.id,
            opportunity_id=None,
            name="Second",
        )
        session.add(extra)
        await session.flush()
        session.add(
            OperationalInputs(
                id=uuid4(), project_id=extra.id,
                unit_count_new=4,
                exit_cap_rate_pct=Decimal("5.5"),
            )
        )
        await session.flush()

    primary = (await session.execute(
        select(Project).where(
            Project.scenario_id == deal_model.id,
            Project.name == "Main Project",
        )
    )).scalar_one()

    debt = CapitalModule(
        scenario_id=deal_model.id,
        label="Senior",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={"amount": "500000", "interest_rate_pct": 6.5,
                "amort_term_years": 30, "hold_term_years": 10},
        carry={"carry_type": "pi", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    equity = CapitalModule(
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type=VehicleType.equity.value,
        equity_role=EquityRole.lp.value,
        stack_position=2,
        source={"amount": "250000"},
        carry={"carry_type": "none", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    session.add_all([debt, equity])
    await session.flush()
    session.add_all([
        CapitalModuleProject(
            capital_module_id=debt.id, project_id=primary.id,
            amount=Decimal("500000"),
        ),
        CapitalModuleProject(
            capital_module_id=equity.id, project_id=primary.id,
            amount=Decimal("250000"),
        ),
    ])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find_debt_service_row(ws) -> int | None:
    label_col, _ = proforma_layout(ws)
    return find_label_row(ws, "Debt Service", col=label_col, exact=True)


async def test_multi_project_debt_service_gated_on_perm_origination(
    session: AsyncSession,
) -> None:
    scenario = await _seed(session, multi_project=True)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    ws = wb["Underwriting Pro Forma"]
    row = _find_debt_service_row(ws)
    assert row is not None, "Debt Service row not found on Underwriting Pro Forma"

    # Scan every column for a formula that contains the perm-origination
    # guard. Y0/Y1 cells may be plain values; Y2+ should carry the gate.
    perm_gate_seen = False
    for col in range(2, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if not isinstance(cell.value, str) or not cell.value.startswith("="):
            continue
        if (
            "s_loan_1_perm_origination_month" in cell.value
            and "AND(" in cell.value
        ):
            perm_gate_seen = True
            break
    assert perm_gate_seen, (
        "expected at least one Debt Service column to gate on "
        "s_loan_1_perm_origination_month via AND(...) — none found"
    )


async def test_single_project_debt_service_also_gated_on_perm_origination(
    session: AsyncSession,
) -> None:
    """Single-project scenarios with a construction-side phase emit
    p1_* cells on the Assumptions sheet, so the Debt Service formula
    must gate on s_loan_1_perm_origination_month just like multi-project."""
    scenario = await _seed(session, multi_project=False)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    ws = wb["Underwriting Pro Forma"]
    row = _find_debt_service_row(ws)
    assert row is not None

    perm_gate_seen = False
    for col in range(2, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if not isinstance(cell.value, str) or not cell.value.startswith("="):
            continue
        if (
            "s_loan_1_perm_origination_month" in cell.value
            and "AND(" in cell.value
        ):
            perm_gate_seen = True
            break
    assert perm_gate_seen, (
        "single-project Debt Service must gate on s_loan_1_perm_origination_month "
        "(phase plan now emits onto Assumptions for single-project too)"
    )


async def test_c2p_notes_column_consumes_active_in_operations(
    session: AsyncSession,
) -> None:
    """The C2P Status block's 6th column must reference the
    ``s_loan_<n>_active_in_operations`` named range — that's the entire
    point of registering the boolean."""
    scenario = await _seed(session, multi_project=True)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    ws = wb["Debt Schedule"]
    # Find the C2P section header row.
    header_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and "Construction-to-Perm Status" in v:
            header_row = r
            break
    assert header_row is not None, "C2P Status section not found"

    # Notes column = column 6, first data row is header_row + 2.
    notes_cell = ws.cell(row=header_row + 2, column=6).value
    assert isinstance(notes_cell, str) and notes_cell.startswith("=")
    assert "s_loan_1_active_in_operations" in notes_cell
