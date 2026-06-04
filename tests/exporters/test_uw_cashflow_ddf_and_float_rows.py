"""Slice 2 (Export v3): the Underwriting Cash Flow sheet emits three
new rows when the corresponding ``OperationalOutputs`` JSONB series
are populated:

  * Deferred Dev Fee Balance — closing balance per year (forward-fill)
  * Deferred Dev Fee Recovered — paydowns per year (sum of waterfall
    + float topup contributions)
  * Float Earnings (Found Money) — per-year ``found_money_periods``

Each row is named so downstream sheets / future formulas can
reference it: ``r_uw_cf_ddf_balance``, ``r_uw_cf_ddf_recovered``,
``r_uw_cf_float_earnings``.

When neither series is populated (legacy / pre-Phase-B deals), all
three rows are absent and the sheet structure stays clean.
"""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from app.models.cashflow import OperationalOutputs
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


pytestmark = pytest.mark.asyncio


def _resolve_named_range(wb, name: str):
    """Return (sheet, [(row,col)...]) for a defined name, or None."""
    dn = wb.defined_names.get(name)
    if dn is None:
        return None
    out: list[tuple[int, int]] = []
    sheet = None
    for sheet_title, coord in dn.destinations:
        sheet = sheet_title
        coord = coord.replace("$", "")
        if ":" in coord:
            a, b = coord.split(":")
            a_col = "".join(c for c in a if c.isalpha())
            a_row = int("".join(c for c in a if c.isdigit()))
            b_col = "".join(c for c in b if c.isalpha())
            b_row = int("".join(c for c in b if c.isdigit()))

            def _col_idx(letters: str) -> int:
                idx = 0
                for ch in letters:
                    idx = idx * 26 + (ord(ch.upper()) - ord("A") + 1)
                return idx
            for col in range(_col_idx(a_col), _col_idx(b_col) + 1):
                out.append((a_row, col))
        else:
            col_letters = "".join(c for c in coord if c.isalpha())
            row = int("".join(c for c in coord if c.isdigit()))
            col_idx = 0
            for ch in col_letters:
                col_idx = col_idx * 26 + (ord(ch.upper()) - ord("A") + 1)
            out.append((row, col_idx))
    return sheet, out


def _row_values(wb, name: str) -> list:
    """Return the cell values of a named single-row range, in order."""
    resolved = _resolve_named_range(wb, name)
    if resolved is None:
        return []
    sheet, cells = resolved
    ws = wb[sheet]
    return [ws.cell(row=r, column=c).value for r, c in cells]


async def _seed_with_series(
    session: AsyncSession,
    ddf_series: dict | None = None,
    float_series: dict | None = None,
):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Slice2 DDF/Float")
    deal_model, _inputs, _stream, _opex = (
        await seed_deal_model_with_financials(session, opp, user)
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()
    await compute_cash_flows(deal_model.id, session)
    outputs = (
        await session.execute(
            select(OperationalOutputs).where(
                OperationalOutputs.project_id == project.id
            )
        )
    ).scalar_one_or_none()
    if outputs is None:
        outputs = OperationalOutputs(
            project_id=project.id, scenario_id=deal_model.id
        )
        session.add(outputs)
    outputs.dev_fee_balance_series = ddf_series
    outputs.float_earnings_series = float_series
    await session.commit()
    return deal_model, project


async def test_ddf_balance_row_emits_year_end_closing_balance(
    session: AsyncSession, tmp_path: Path
) -> None:
    """A simple two-year paydown: $500k at close, $200k paid in Y1
    (closing $300k), $300k paid in Y2 (closing $0). The Excel row
    must show [500k, 300k, 0] across [Y0, Y1, Y2]."""
    series = {
        "opening_at_close": "500000",
        "fully_paid_period": 18,
        "total_paid": "500000",
        "remaining_at_horizon": "0",
        "periods": [
            # Periods 1-12 = Y1; closing of last period in Y1 = 300000.
            {"period": 1,  "opening_balance": "500000",
             "paydown_from_waterfall": "0",
             "paydown_from_float_topup": "0",
             "closing_balance": "500000"},
            {"period": 12, "opening_balance": "500000",
             "paydown_from_waterfall": "200000",
             "paydown_from_float_topup": "0",
             "closing_balance": "300000"},
            # Periods 13-24 = Y2; closing of last period = 0.
            {"period": 18, "opening_balance": "300000",
             "paydown_from_waterfall": "300000",
             "paydown_from_float_topup": "0",
             "closing_balance": "0"},
        ],
    }
    scenario, _ = await _seed_with_series(session, ddf_series=series)
    blob = await export_investor_workbook(scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)
    wb = load_workbook(path, data_only=False)

    balance_row = _row_values(wb, "r_uw_cf_ddf_balance")
    assert balance_row, "r_uw_cf_ddf_balance must be emitted when series present"
    # Y0 carries opening_at_close.
    assert float(balance_row[0]) == pytest.approx(500000.0, abs=0.01)
    # Y1 = closing of period 12.
    assert float(balance_row[1]) == pytest.approx(300000.0, abs=0.01)
    # Y2 = closing of period 18 (highest period in Y2).
    assert float(balance_row[2]) == pytest.approx(0.0, abs=0.01)

    recovered_row = _row_values(wb, "r_uw_cf_ddf_recovered")
    assert recovered_row, "r_uw_cf_ddf_recovered must be emitted alongside balance"
    # Y0 = no paydown, Y1 = 200000, Y2 = 300000.
    assert float(recovered_row[0]) == pytest.approx(0.0, abs=0.01)
    assert float(recovered_row[1]) == pytest.approx(200000.0, abs=0.01)
    assert float(recovered_row[2]) == pytest.approx(300000.0, abs=0.01)


async def test_float_earnings_row_aggregates_found_money_periods(
    session: AsyncSession, tmp_path: Path
) -> None:
    series = {
        "sources": [
            {
                "float_source_id": "00000000-0000-0000-0000-000000000001",
                "parent_module_id": None,
                "total_earnings": 75000.0,
                "schedule": [],
                "warnings": [],
            }
        ],
        # Periods spread across Y1 + Y2.
        "found_money_periods": {
            "6":  25000.0,
            "12": 25000.0,
            "18": 25000.0,
        },
        "warnings": [],
    }
    scenario, _ = await _seed_with_series(session, float_series=series)
    blob = await export_investor_workbook(scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)
    wb = load_workbook(path, data_only=False)

    float_row = _row_values(wb, "r_uw_cf_float_earnings")
    assert float_row, "r_uw_cf_float_earnings must be emitted when series present"
    # Y0 has no found money.
    assert float(float_row[0]) == pytest.approx(0.0, abs=0.01)
    # Y1 (periods 1-12) gets $50k (6 + 12).
    assert float(float_row[1]) == pytest.approx(50000.0, abs=0.01)
    # Y2 (periods 13-24) gets $25k (period 18).
    assert float(float_row[2]) == pytest.approx(25000.0, abs=0.01)


async def test_no_series_means_no_rows(
    session: AsyncSession, tmp_path: Path
) -> None:
    """Legacy deals (both JSONB columns null) skip all three rows."""
    scenario, _ = await _seed_with_series(
        session, ddf_series=None, float_series=None
    )
    blob = await export_investor_workbook(scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)
    wb = load_workbook(path, data_only=False)

    assert _resolve_named_range(wb, "r_uw_cf_ddf_balance") is None
    assert _resolve_named_range(wb, "r_uw_cf_ddf_recovered") is None
    assert _resolve_named_range(wb, "r_uw_cf_float_earnings") is None
