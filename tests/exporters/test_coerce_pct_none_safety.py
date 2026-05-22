"""Regression: ``_coerce_pct(None)`` used to crash with
``decimal.InvalidOperation``. The Block C debt-assumption row and the
Block D waterfall-hurdle row both call it on optional ORM fields
(``source.dscr_min`` / ``tier.irr_hurdle_pct``), so any seed that left
those blank produced a 500 on export. Now both paths return ``None``
and the cell renders blank / em-dash via ``registry.write``.

Triggered on production by a real WaterfallTier with no IRR hurdle
configured (proforma profile, ``hazelwood-commons`` deal).
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from sqlalchemy.ext.asyncio import AsyncSession

from app.exporters.investor_export import (
    _coerce_decimal,
    _coerce_pct,
    export_investor_workbook,
)
from app.models.capital import (
    CapitalModule,
    EquityRole,
    VehicleType,
    WaterfallTier,
)
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


def test_coerce_decimal_none_returns_none():
    assert _coerce_decimal(None) is None
    assert _coerce_decimal("") is None
    assert _coerce_decimal(0) == Decimal(0)
    assert _coerce_decimal("5.5") == Decimal("5.5")


def test_coerce_pct_none_returns_none():
    assert _coerce_pct(None) is None
    assert _coerce_pct("") is None
    assert _coerce_pct(100) == Decimal(1)
    assert _coerce_pct("6.5") == Decimal("0.065")


async def test_export_survives_waterfall_tier_with_null_irr_hurdle(
    session: AsyncSession,
):
    """End-to-end: scenario with a WaterfallTier whose irr_hurdle_pct
    is None must export without crashing the request. Mirrors the
    production deal that triggered the 500."""
    from openpyxl import load_workbook

    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Null-Hurdle Tier")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )
    equity = CapitalModule(
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type=VehicleType.equity.value,
        equity_role=EquityRole.lp.value,
        stack_position=1,
        source={"amount": "250000"},
        carry={"carry_type": "none", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    session.add(equity)
    await session.flush()

    # Two tiers — one with hurdle, one without — covers both code paths.
    session.add_all([
        WaterfallTier(
            scenario_id=deal_model.id,
            capital_module_id=equity.id,
            priority=1, tier_type="pref_return",
            irr_hurdle_pct=Decimal("8.0"),
            lp_split_pct=Decimal("100"), gp_split_pct=Decimal("0"),
        ),
        WaterfallTier(
            scenario_id=deal_model.id,
            capital_module_id=None,
            priority=2, tier_type="residual",
            irr_hurdle_pct=None,  # the crash trigger
            lp_split_pct=Decimal("70"), gp_split_pct=Decimal("30"),
        ),
    ])
    await session.flush()

    # Both profiles render Assumptions Block D; test both so a regression
    # on either path trips the suite.
    for profile in ("internal", "proforma"):
        blob = await export_investor_workbook(
            deal_model.id, session, profile=profile,
        )
        wb = load_workbook(BytesIO(blob), data_only=False)
        assert "Assumptions" in wb.sheetnames, (
            f"profile={profile} should render Assumptions"
        )
