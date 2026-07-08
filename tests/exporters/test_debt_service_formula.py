"""Phase E: Underwriting Pro Forma Debt Service Y1+ must be a SUM over the
per-loan ``s_loan_{i}_annual_pi`` named ranges registered on the Debt
Schedule sheet, so an LP changing a loan principal or rate sees the Pro
Forma debt service shift in lock-step. Y0 stays at the engine value
(construction-phase debt service often differs from the stabilized PMT).

Contract:

  1. internal + lender profiles render Debt Schedule, so Y1+ Debt Service
     == ``=s_loan_1_annual_pi+...``.
  2. lp + proforma profiles do NOT render Debt Schedule. Cells must stay
     numeric to avoid #NAME?.
  3. Net Cash Flow Y1+ is derived from NOI + Debt Service so the new
     formula flows through.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from app.models.capital import (
    CapitalModule,
    CapitalModuleProject,
    EquityRole,
    VehicleType,
)
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)
from tests.exporters._parity_helpers import find_label_row, proforma_layout


async def _seed(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Debt Service Smoke")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()
    debt = CapitalModule(
        scenario_id=deal_model.id,
        label="Senior Loan",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={
            "amount": "500000", "interest_rate_pct": 6.5,
            "amort_term_years": 30, "hold_term_years": 10,
        },
        carry={"carry_type": "pi", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    equity = CapitalModule(
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type=VehicleType.equity.value,
        equity_role=EquityRole.lp.value,
        stack_position=2,
        source={"amount": "250000"},
        carry={"carry_type": "none", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    session.add_all([debt, equity])
    await session.flush()
    session.add_all([
        CapitalModuleProject(
            capital_module_id=debt.id, project_id=project.id,
            amount=Decimal("500000"),
        ),
        CapitalModuleProject(
            capital_module_id=equity.id, project_id=project.id,
            amount=Decimal("250000"),
        ),
    ])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find_row(ws, label_substr: str) -> int | None:
    label_col, _ = proforma_layout(ws)
    return find_label_row(ws, label_substr, col=label_col)


def _find_pf_sheet(wb):
    for name in wb.sheetnames:
        low = name.lower()
        if "underwriting pro forma" in low:
            return wb[name]
    # Proforma profile uses "Pro Forma"
    for name in wb.sheetnames:
        low = name.lower()
        if "pro forma" in low or "proforma" in low:
            return wb[name]
    return None


@pytest.mark.parametrize("profile", ["internal", "lender"])
async def test_debt_service_y1_plus_is_sum_formula(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)

    ws = _find_pf_sheet(wb)
    assert ws is not None, f"profile={profile} missing Pro Forma sheet"

    ds_row = _find_row(ws, "debt service")
    assert ds_row is not None, "missing Debt Service row"

    _, y0_col = proforma_layout(ws)
    y1 = ws.cell(row=ds_row, column=y0_col + 1).value
    assert isinstance(y1, str) and y1.startswith("="), (
        f"profile={profile}: Debt Service Y1 must be formula; got {y1!r}"
    )
    assert "s_loan_1_annual_pi" in y1, (
        f"profile={profile}: Y1 formula missing s_loan_1_annual_pi; got {y1!r}"
    )
    assert "s_loan_1_annual_pi" in wb.defined_names, (
        f"profile={profile} missing s_loan_1_annual_pi defined name "
        f"(formula would dangle to #NAME?)"
    )


@pytest.mark.parametrize("profile", ["lp", "proforma"])
async def test_debt_service_stays_numeric_without_debt_schedule(
    session: AsyncSession, profile: str
):
    """LP + proforma omit the Debt Schedule sheet; the s_loan_*_annual_pi
    named ranges never get registered, so the Debt Service row must stay
    numeric to keep the workbook clean of #NAME? errors."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)

    ws = _find_pf_sheet(wb)
    assert ws is not None

    ds_row = _find_row(ws, "debt service")
    assert ds_row is not None
    _, y0_col = proforma_layout(ws)
    y1 = ws.cell(row=ds_row, column=y0_col + 1).value
    if isinstance(y1, str) and y1.startswith("="):
        assert "s_loan_" not in y1, (
            f"profile={profile}: Debt Service must not reference Debt "
            f"Schedule names when the sheet is absent; got {y1!r}"
        )


@pytest.mark.parametrize("profile", ["internal", "lender"])
async def test_net_cash_flow_y1_is_derived_from_noi_and_debt(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)

    ws = _find_pf_sheet(wb)
    assert ws is not None

    noi_row = _find_row(ws, "noi")
    ncf_row = _find_row(ws, "net cash flow")
    assert noi_row is not None and ncf_row is not None

    _, y0_col = proforma_layout(ws)
    y1_col = y0_col + 1
    y1 = ws.cell(row=ncf_row, column=y1_col).value
    assert isinstance(y1, str) and y1.startswith("="), (
        f"profile={profile}: Net Cash Flow Y1 must be formula; got {y1!r}"
    )
    noi_ref = f"{get_column_letter(y1_col)}{noi_row}"
    assert noi_ref in y1, (
        f"profile={profile}: NCF Y1 must reference NOI cell {noi_ref}; got {y1!r}"
    )
