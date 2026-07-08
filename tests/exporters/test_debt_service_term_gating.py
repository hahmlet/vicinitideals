"""Phase E refinement: the Pro Forma Debt Service formula must gate each
loan's annual P&I contribution by its hold term — a loan whose
``s_loan_{n}_term_months`` is exhausted before year Y stops contributing.

Contract:

  1. Debt Service formula in each Y2+ column wraps every loan in
     ``IF(s_loan_{i}_term_months>={year_end_months}, s_loan_{i}_annual_pi, 0)``.
  2. The threshold scales with the column's year — Y2's check is 24, Y3
     is 36, etc.
  3. ``s_loan_{n}_term_months`` defined name is registered on the
     workbook for every loan.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from app.models.capital import (
    CapitalModule,
    CapitalModuleProject,
    EquityRole,
    VehicleType,
)
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)
from tests.exporters._parity_helpers import find_label_row, proforma_layout


async def _seed(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Term Gate Smoke")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()
    debt = CapitalModule(
        scenario_id=deal_model.id,
        label="Senior Loan",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={
            "amount": "500000", "interest_rate_pct": 6.5,
            "amort_term_years": 30, "hold_term_years": 10,
        },
        carry={"carry_type": "pi", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    equity = CapitalModule(
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type=VehicleType.equity.value,
        equity_role=EquityRole.lp.value,
        stack_position=2,
        source={"amount": "250000"},
        carry={"carry_type": "none", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    session.add_all([debt, equity])
    await session.flush()
    session.add_all([
        CapitalModuleProject(
            capital_module_id=debt.id, project_id=project.id,
            amount=Decimal("500000"),
        ),
        CapitalModuleProject(
            capital_module_id=equity.id, project_id=project.id,
            amount=Decimal("250000"),
        ),
    ])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find_row(ws, label_substr: str) -> int | None:
    label_col, _ = proforma_layout(ws)
    return find_label_row(ws, label_substr, col=label_col)


@pytest.mark.parametrize("profile", ["internal", "lender"])
async def test_term_months_named_range_registered(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "s_loan_1_term_months" in wb.defined_names, (
        f"profile={profile} missing s_loan_1_term_months"
    )


@pytest.mark.parametrize("profile", ["internal", "lender"])
async def test_debt_service_formula_gates_by_term_months(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Pro Forma"]
    ds_row = _find_row(ws, "debt service")
    assert ds_row is not None, f"profile={profile}: debt service row missing"

    # UW Pro Forma columns (layout-aware): Y0 = engine value, Y1 = first
    # formula (threshold=12), Y2 threshold=24, Y3 threshold=36, etc.
    _, y0_col = proforma_layout(ws)
    y1 = ws.cell(row=ds_row, column=y0_col + 1).value
    assert isinstance(y1, str) and y1.startswith("="), (
        f"profile={profile}: Y1 DS not a formula; got {y1!r}"
    )
    assert "s_loan_1_term_months>=12" in y1, (
        f"profile={profile}: Y1 DS missing 12-month gate; got {y1!r}"
    )

    y2 = ws.cell(row=ds_row, column=y0_col + 2).value
    assert isinstance(y2, str) and "s_loan_1_term_months>=24" in y2, (
        f"profile={profile}: Y2 DS missing 24-month gate; got {y2!r}"
    )

    y3 = ws.cell(row=ds_row, column=y0_col + 3).value
    assert isinstance(y3, str) and "s_loan_1_term_months>=36" in y3, (
        f"profile={profile}: Y3 DS missing 36-month gate; got {y3!r}"
    )


@pytest.mark.parametrize("profile", ["internal", "lender"])
async def test_debt_service_formula_wraps_each_loan_in_if(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Pro Forma"]
    ds_row = _find_row(ws, "debt service")
    assert ds_row is not None
    _, y0_col = proforma_layout(ws)
    y1 = ws.cell(row=ds_row, column=y0_col + 1).value
    assert "IF(" in y1, f"profile={profile}: DS formula must use IF gates"
    assert "s_loan_1_annual_pi" in y1, (
        f"profile={profile}: DS formula must still reference loan P&I; got {y1!r}"
    )
