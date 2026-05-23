"""Construction-to-Perm Status block in the Debt Schedule.

Contract:

  1. ``s_loan_<n>_perm_origination_month`` is registered as a named cell
     for every debt module that funds at least one project with a
     registered ``p<idx>_perm_origination_month`` (i.e. a project whose
     phase plan has a construction-side phase).
  2. ``s_loan_<n>_active_in_operations`` is a boolean formula referencing
     both the term-months cell and the perm-origination cell — it must
     evaluate to TRUE when the loan's term extends past perm origination
     and FALSE otherwise.
  3. The block also renders for single-project scenarios: ``p1_*`` phase
     cells are emitted on the Assumptions sheet (Block E) when the
     per-project sheet is suppressed, so the same Debt Schedule
     references resolve.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from app.models.capital import (
    CapitalModule,
    CapitalModuleProject,
    EquityRole,
    VehicleType,
)
from app.models.deal import OperationalInputs
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


pytestmark = pytest.mark.asyncio


async def _seed_multi_project_with_debt(session: AsyncSession):
    """Two value_add projects (default project_type) so per-project sheets
    render, plus one debt + one equity module wired to the first project."""
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="C2P Smoke")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )

    # Second project so len(projects) > 1 and per-project sheets render.
    extra = Project(
        id=uuid4(),
        scenario_id=deal_model.id,
        opportunity_id=None,
        name="Second",
    )
    session.add(extra)
    await session.flush()
    session.add(
        OperationalInputs(
            id=uuid4(), project_id=extra.id,
            unit_count_new=4,
            exit_cap_rate_pct=Decimal("5.5"),
        )
    )
    await session.flush()

    # Fetch the primary project so we can junction the debt onto it.
    from sqlalchemy import select
    primary = (await session.execute(
        select(Project).where(Project.scenario_id == deal_model.id, Project.name == "Main Project")
    )).scalar_one()

    debt = CapitalModule(
        scenario_id=deal_model.id,
        label="Senior Construction Loan",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={
            "amount": "500000", "interest_rate_pct": 6.5,
            "amort_term_years": 30, "hold_term_years": 10,
        },
        carry={"carry_type": "pi", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    equity = CapitalModule(
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type=VehicleType.equity.value,
        equity_role=EquityRole.lp.value,
        stack_position=2,
        source={"amount": "250000"},
        carry={"carry_type": "none", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    session.add_all([debt, equity])
    await session.flush()
    session.add_all([
        CapitalModuleProject(
            capital_module_id=debt.id, project_id=primary.id,
            amount=Decimal("500000"),
        ),
        CapitalModuleProject(
            capital_module_id=equity.id, project_id=primary.id,
            amount=Decimal("250000"),
        ),
    ])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


async def _seed_single_project_with_debt(session: AsyncSession):
    """Single-project scenario — per-project sheets are skipped so
    no perm cells exist for the C2P block to reference."""
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="C2P Solo")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )

    from sqlalchemy import select
    primary = (await session.execute(
        select(Project).where(Project.scenario_id == deal_model.id, Project.name == "Main Project")
    )).scalar_one()

    debt = CapitalModule(
        scenario_id=deal_model.id,
        label="Senior",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={"amount": "300000", "interest_rate_pct": 6.0,
                "amort_term_years": 30, "hold_term_years": 10},
        carry={"carry_type": "pi", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    session.add(debt)
    await session.flush()
    session.add(CapitalModuleProject(
        capital_module_id=debt.id, project_id=primary.id,
        amount=Decimal("300000"),
    ))
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


async def test_c2p_named_cells_registered_for_multi_project(
    session: AsyncSession,
) -> None:
    scenario = await _seed_multi_project_with_debt(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    defined = {name for name in wb.defined_names}
    assert "s_loan_1_perm_origination_month" in defined
    assert "s_loan_1_active_in_operations" in defined


async def test_c2p_perm_origination_formula_references_project_cell(
    session: AsyncSession,
) -> None:
    """The per-loan perm-origination cell must be a formula that pulls
    from at least one ``p<idx>_perm_origination_month`` named range."""
    scenario = await _seed_multi_project_with_debt(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    dn = wb.defined_names.get("s_loan_1_perm_origination_month")
    assert dn is not None
    for sheet_name, cell_ref in dn.destinations:
        formula = wb[sheet_name][cell_ref].value
        assert isinstance(formula, str) and formula.startswith("=")
        assert "p1_perm_origination_month" in formula
        # IFERROR wrapper keeps the cell readable when no project is eligible.
        assert "IFERROR" in formula
        return
    pytest.fail("named range had no destination")


async def test_c2p_active_in_ops_formula_references_term_and_perm(
    session: AsyncSession,
) -> None:
    """``s_loan_<n>_active_in_operations`` must depend on both the term-
    months cell and the perm-origination cell so an LP edit to either
    flows through."""
    scenario = await _seed_multi_project_with_debt(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    dn = wb.defined_names.get("s_loan_1_active_in_operations")
    assert dn is not None
    for sheet_name, cell_ref in dn.destinations:
        formula = wb[sheet_name][cell_ref].value
        assert isinstance(formula, str) and formula.startswith("=")
        assert "s_loan_1_term_months" in formula
        assert "s_loan_1_perm_origination_month" in formula
        return
    pytest.fail("named range had no destination")


async def test_c2p_block_renders_for_single_project_scenario(
    session: AsyncSession,
) -> None:
    """Single-project workbooks emit the phase plan on the Assumptions
    sheet (Block E) so the Debt Schedule C2P block has p1_* cells to
    reference. Both s_loan_1_perm_origination_month and
    s_loan_1_active_in_operations must register."""
    scenario = await _seed_single_project_with_debt(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    defined = {name for name in wb.defined_names}
    assert "p1_perm_origination_month" in defined, (
        "single-project phase plan must emit p1_perm_origination_month on Assumptions"
    )
    assert "s_loan_1_perm_origination_month" in defined
    assert "s_loan_1_active_in_operations" in defined

    # Verify the perm cell sits on Assumptions (not on a P1 sheet, which
    # is suppressed for single-project).
    dn = wb.defined_names["p1_perm_origination_month"]
    sheets = {sheet for sheet, _ in dn.destinations}
    assert "Assumptions" in sheets, (
        f"expected p1_perm_origination_month on Assumptions, got {sheets}"
    )
