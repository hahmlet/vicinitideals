"""Phase B: Operating Expenses grow via a chain formula so a single edit
to a growth input ripples through every downstream year — and through
NOI/Net Cash Flow because those are derived formulas referencing OpEx.

Layout notes (commit e7ba809):

  - ``_build_uw_proforma`` (internal/lp/lender profiles) — Underwriting
    Pro Forma sheet: the OpEx total row SUMs per-line bullet rows; the
    growth chain lives on each bullet via ``s_opex_<slug>_escalation_pct``.
  - ``_write_pf_table`` (proforma profile) — Pro Forma sheet keeps the
    sheet-wide ``s_opex_growth_rate`` chain on the total row.
  - CapEx Reserve was deliberately REMOVED from the growth chain on both
    sheets: the engine already applies expense_growth per period, and
    anchoring the chain at a construction-phase Y1=0 zeroed every year.
    Y2+ CapEx cells must stay numeric engine values.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from app.models.capital import (
    CapitalModule,
    CapitalModuleProject,
    EquityRole,
    VehicleType,
)
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)
from tests.exporters._parity_helpers import parse_sum_range, proforma_layout


async def _seed(session: AsyncSession):
    """Seed a scenario rich enough to produce >= 3 annual columns so the
    Y2+ growth chain has cells to inspect. Mirrors the full-workbook
    error-scan fixture."""
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="Phase B Growth Chain")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()

    debt = CapitalModule(
        scenario_id=deal_model.id,
        label="Senior Loan",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={
            "amount": "500000", "interest_rate_pct": 6.5,
            "amort_term_years": 30, "hold_term_years": 10,
        },
        carry={"carry_type": "pi", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    equity = CapitalModule(
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type=VehicleType.equity.value,
        equity_role=EquityRole.lp.value,
        stack_position=2,
        source={"amount": "250000"},
        carry={"carry_type": "none", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    session.add_all([debt, equity])
    await session.flush()
    session.add_all([
        CapitalModuleProject(
            capital_module_id=debt.id, project_id=project.id,
            amount=Decimal("500000"),
        ),
        CapitalModuleProject(
            capital_module_id=equity.id, project_id=project.id,
            amount=Decimal("250000"),
        ),
    ])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find_row(ws, label_prefix: str) -> int | None:
    label_col, _ = proforma_layout(ws)
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=label_col).value
        if isinstance(v, str) and v.startswith(label_prefix):
            return r
    return None


def _assert_growth_chain(ws, row: int, growth_name: str) -> None:
    """Y2 onward must be a formula referencing ``growth_name``."""
    _, y0_col = proforma_layout(ws)
    chain_started = False
    for c in range(y0_col + 2, ws.max_column + 1):  # start at Y2
        v = ws.cell(row=row, column=c).value
        if v is None:
            break
        assert isinstance(v, str) and v.startswith("="), (
            f"row {row} col {c} expected formula; got {v!r}"
        )
        assert growth_name in v, (
            f"row {row} col {c} expected growth chain referencing "
            f"{growth_name!r}; got {v!r}"
        )
        chain_started = True
    assert chain_started, (
        f"row {row} has no Y2+ cells — workbook has too few years to "
        f"exercise the growth chain"
    )


def _assert_engine_numeric_y2_plus(ws, row: int) -> None:
    """Y2 onward must stay numeric engine values (no growth chain).

    Guards the e7ba809 fix: chaining CapEx Reserve off a construction-
    phase Y1=0 anchor zeroed every downstream year, so the exporter now
    writes the engine's per-period expense-growth values directly.
    """
    _, y0_col = proforma_layout(ws)
    seen = 0
    for c in range(y0_col + 2, ws.max_column + 1):
        v = ws.cell(row=row, column=c).value
        if v is None:
            break
        assert not (isinstance(v, str) and v.startswith("=")), (
            f"row {row} col {c}: CapEx Reserve must stay an engine value "
            f"(growth chain deliberately removed — Y1=0 anchor bug); "
            f"got formula {v!r}"
        )
        seen += 1
    assert seen > 0, (
        f"row {row} has no Y2+ cells — workbook has too few years"
    )


# ── Internal profile: Underwriting Pro Forma ─────────────────────────────────


async def test_uw_proforma_opex_grows_via_chain(session: AsyncSession):
    """Internal-profile UW Pro Forma OpEx: total Y2+ SUMs the bullet
    rows; each bullet Y2+ chains via its per-line escalation input."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Pro Forma"]

    opex_row = _find_row(ws, "Operating Expenses")
    assert opex_row is not None, "Operating Expenses row not found"

    _, y0_col = proforma_layout(ws)
    y1_col = y0_col + 1
    parsed = parse_sum_range(ws.cell(row=opex_row, column=y1_col).value)
    assert parsed is not None, (
        "OpEx total must SUM the per-line bullet rows; got "
        f"{ws.cell(row=opex_row, column=y1_col).value!r}"
    )
    _, first, last = parsed
    # Total row: every Y2+ column stays a SUM over the bullet column.
    for c in range(y0_col + 2, ws.max_column + 1):
        v = ws.cell(row=opex_row, column=c).value
        if v is None:
            break
        assert parse_sum_range(v) is not None, (
            f"OpEx total col {c} must SUM the bullet column; got {v!r}"
        )
    # Bullet rows carry the growth chain via per-line escalation.
    for br in range(first, last + 1):
        _assert_growth_chain(ws, br, "_escalation_pct")


async def test_uw_proforma_capex_grows_via_chain(session: AsyncSession):
    """Internal-profile UW Pro Forma CapEx Reserve Y2+ = engine values
    (growth chain deliberately removed in e7ba809)."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Pro Forma"]

    capex_row = _find_row(ws, "CapEx Reserve")
    assert capex_row is not None, "CapEx Reserve row not found"

    _assert_engine_numeric_y2_plus(ws, capex_row)


# ── Proforma profile: Pro Forma sheet ─────────────────────────────────────────


async def test_proforma_opex_grows_via_chain(session: AsyncSession):
    """Proforma-profile Pro Forma OpEx Y2+ = chain formula."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(
        scenario.id, session, profile="proforma",
    )
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Pro Forma"]

    opex_row = _find_row(ws, "Operating Expenses")
    assert opex_row is not None

    _assert_growth_chain(ws, opex_row, "s_opex_growth_rate")


async def test_proforma_capex_grows_via_chain(session: AsyncSession):
    """Proforma-profile Pro Forma CapEx Reserve Y2+ = engine values
    (excluded from the chain — see _write_pf_table's rationale)."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(
        scenario.id, session, profile="proforma",
    )
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Pro Forma"]

    capex_row = _find_row(ws, "CapEx Reserve")
    assert capex_row is not None

    _assert_engine_numeric_y2_plus(ws, capex_row)


# ── OER row picks up OpEx changes ────────────────────────────────────────────


async def test_uw_proforma_oer_is_formula(session: AsyncSession):
    """OER row references the OpEx + EGI cells so growth-chain edits flow."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Pro Forma"]

    oer_row = _find_row(ws, "OER (OpEx")
    assert oer_row is not None, "OER row not found"

    _, y0_col = proforma_layout(ws)
    formula_count = 0
    for c in range(y0_col, ws.max_column + 1):
        v = ws.cell(row=oer_row, column=c).value
        if v is None:
            break
        assert isinstance(v, str) and v.startswith("="), (
            f"OER col {c} expected formula; got {v!r}"
        )
        assert "IFERROR" in v, f"OER col {c} should be IFERROR-guarded; got {v!r}"
        formula_count += 1
    assert formula_count > 0


async def test_proforma_oer_is_formula(session: AsyncSession):
    """Proforma profile OER row likewise references the OpEx + EGI cells."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(
        scenario.id, session, profile="proforma",
    )
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Pro Forma"]

    oer_row = _find_row(ws, "OER (OpEx")
    assert oer_row is not None

    _, y0_col = proforma_layout(ws)
    formula_count = 0
    for c in range(y0_col, ws.max_column + 1):
        v = ws.cell(row=oer_row, column=c).value
        if v is None:
            break
        assert isinstance(v, str) and v.startswith("="), (
            f"OER col {c} expected formula; got {v!r}"
        )
        assert "IFERROR" in v
        formula_count += 1
    assert formula_count > 0
