"""Engine-vs-formula parity for the Pro Forma EGI + NOI conversion.

Commit 3 of docs/feature-plans/investor-excel-formula-conversion.md §4.1.

Scope is intentionally narrow: only the rows whose math is a direct
sum/difference of other rows on the same sheet are formula-driven in
this commit (Effective Gross Income, NOI). Gross Revenue, Vacancy
Loss, Operating Expenses, CapEx Reserve, Debt Service, Net Cash Flow
stay as engine values — Debt Service formulas land in commit 4 with
the Debt Schedule conversion, and revenue/opex growth-projection
formulas land in a later commit once an ``s_revenue_growth_rate``
input is wired.

The parity loop:

  1. Cell is a formula (string starting with ``=``) for every Y0..Yn
     column on the EGI + NOI rows
  2. Formula references the GrossRev/Vacancy cells (EGI) or EGI/OpEx
     cells (NOI) on the same column
  3. After Excel recalc, the cell value == engine.effective_gross_income
     (or .noi) for that year within tolerance
"""
from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.exporters.investor_export import (
    _aggregate_scenario_annual,
    _load_all,
    export_investor_workbook,
)
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)
from tests.exporters._parity_helpers import (
    RecalcUnavailableError,
    find_label_row,
    proforma_layout,
    recalc_workbook,
)


async def _seed_scenario(session: AsyncSession):
    org, user = await seed_org(session)
    opportunity = await seed_opportunity(
        session, org, user, name="ProForma-Parity Smoke"
    )
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opportunity, user
    )
    return deal_model


def _proforma_sheet(blob: bytes):
    wb = load_workbook(BytesIO(blob), data_only=False)
    return wb, wb["Underwriting Pro Forma"]


def _find_row_by_label(ws, label: str) -> int | None:
    label_col, _ = proforma_layout(ws)
    return find_label_row(ws, label, col=label_col, exact=True)


_CELL_REF_RE = re.compile(r"[A-Z]+\d+")


async def test_egi_row_is_formula_each_year(session: AsyncSession):
    """Every Y0..Yn cell on the EGI row carries a formula, not a value."""
    scenario = await _seed_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    _wb, ws = _proforma_sheet(blob)

    egi_row = _find_row_by_label(ws, "Effective Gross Income")
    assert egi_row is not None

    _, y0_col = proforma_layout(ws)
    # Walk year columns starting at Y0; stop at the first empty cell.
    formula_count = 0
    for c in range(y0_col, ws.max_column + 1):
        v = ws.cell(row=egi_row, column=c).value
        if v is None:
            break
        assert isinstance(v, str) and v.startswith("="), (
            f"EGI Y{c - y0_col} should be a formula; got {v!r}"
        )
        # Must reference at least two cells (GrossRev + Vacancy) on this
        # column. Cell references look like "C5", "C6", etc.
        assert len(_CELL_REF_RE.findall(v)) >= 2, (
            f"EGI formula should reference 2+ cells; got {v!r}"
        )
        formula_count += 1
    assert formula_count > 0


async def test_noi_row_is_formula_each_year(session: AsyncSession):
    """Every Y0..Yn cell on the NOI row carries a formula."""
    scenario = await _seed_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    _wb, ws = _proforma_sheet(blob)

    noi_row = _find_row_by_label(ws, "NOI")
    assert noi_row is not None

    _, y0_col = proforma_layout(ws)
    formula_count = 0
    for c in range(y0_col, ws.max_column + 1):
        v = ws.cell(row=noi_row, column=c).value
        if v is None:
            break
        assert isinstance(v, str) and v.startswith("="), (
            f"NOI Y{c - y0_col} should be a formula; got {v!r}"
        )
        # Should subtract (EGI - OpEx) so a minus must appear in the formula.
        assert "-" in v, f"NOI formula expected to contain '-'; got {v!r}"
        formula_count += 1
    assert formula_count > 0


async def test_egi_evaluates_to_engine_value(
    session: AsyncSession, tmp_path: Path
):
    """Excel-recalc'd EGI == engine effective_gross_income for each year."""
    scenario = await _seed_scenario(session)
    ctx = await _load_all(session, scenario.id)
    annual = _aggregate_scenario_annual(ctx["cash_flows"])

    blob = await export_investor_workbook(scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)

    try:
        recalc_workbook(path)
    except RecalcUnavailableError as exc:
        pytest.skip(f"no recalc backend: {exc}")

    wb = load_workbook(path, data_only=True)
    ws = wb["Underwriting Pro Forma"]
    egi_row = _find_row_by_label(ws, "Effective Gross Income")
    assert egi_row is not None

    max_year = min(max(annual) if annual else 0, 10)
    year_cols = list(range(0, max(max_year, 1) + 1))

    _, y0_col = proforma_layout(ws)
    for col_offset, year in enumerate(year_cols):
        engine_value = float(annual.get(year, {}).get("effective_gross_income", 0))
        excel_value = ws.cell(row=egi_row, column=y0_col + col_offset).value
        if excel_value is None:
            excel_value = 0
        diff = abs(float(excel_value) - engine_value)
        assert diff < 1.0, (
            f"EGI Y{year} parity break: engine={engine_value}, "
            f"excel={excel_value}, diff={diff}"
        )


async def test_noi_evaluates_to_engine_value(
    session: AsyncSession, tmp_path: Path
):
    """Excel-recalc'd NOI == engine noi for each year."""
    scenario = await _seed_scenario(session)
    ctx = await _load_all(session, scenario.id)
    annual = _aggregate_scenario_annual(ctx["cash_flows"])

    blob = await export_investor_workbook(scenario.id, session)
    path = tmp_path / "wb.xlsx"
    path.write_bytes(blob)

    try:
        recalc_workbook(path)
    except RecalcUnavailableError as exc:
        pytest.skip(f"no recalc backend: {exc}")

    wb = load_workbook(path, data_only=True)
    ws = wb["Underwriting Pro Forma"]
    noi_row = _find_row_by_label(ws, "NOI")
    assert noi_row is not None

    max_year = min(max(annual) if annual else 0, 10)
    year_cols = list(range(0, max(max_year, 1) + 1))

    _, y0_col = proforma_layout(ws)
    for col_offset, year in enumerate(year_cols):
        engine_value = float(annual.get(year, {}).get("noi", 0))
        excel_value = ws.cell(row=noi_row, column=y0_col + col_offset).value
        if excel_value is None:
            excel_value = 0
        diff = abs(float(excel_value) - engine_value)
        assert diff < 1.0, (
            f"NOI Y{year} parity break: engine={engine_value}, "
            f"excel={excel_value}, diff={diff}"
        )


async def test_egi_formula_subtracts_vacancy(session: AsyncSession):
    """EGI must net out Vacancy Loss exactly once (no double-count).

    Sign convention (commit e7ba809): the Vacancy Loss row is written
    as signed NEGATIVE numbers so the LP reads a negative line, and the
    EGI formula ADDS the signed vacancy cell (``=GrossRev + Vacancy``).
    Adding a subtraction on top of the negative values would flip back
    to the original double-count bug, so this test pins both halves:

      1. every EGI cell adds the vacancy cell (``+<vac_ref>``), and
      2. every numeric Vacancy Loss cell is <= 0.

    Gross Revenue stays true GPR (occupancy excluded from the
    rent_y1_monthly formula — see test_rent_y1_monthly_excludes_occupancy).
    """
    from openpyxl.utils import get_column_letter

    scenario = await _seed_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    _wb, ws = _proforma_sheet(blob)

    egi_row = _find_row_by_label(ws, "Effective Gross Income")
    vac_row = _find_row_by_label(ws, "Vacancy Loss")
    assert egi_row is not None and vac_row is not None

    _, y0_col = proforma_layout(ws)
    for c in range(y0_col, ws.max_column + 1):
        v = ws.cell(row=egi_row, column=c).value
        if v is None:
            break
        assert isinstance(v, str) and v.startswith("=")
        vac_ref = f"{get_column_letter(c)}{vac_row}"
        assert f"+{vac_ref}" in v, (
            f"EGI Y{c - y0_col} formula must add the signed vacancy cell "
            f"({vac_ref}); got {v!r}"
        )
        assert f"-{vac_ref}" not in v, (
            f"EGI Y{c - y0_col} must not subtract the already-negative "
            f"vacancy cell (double-count); got {v!r}"
        )
        vac_val = ws.cell(row=vac_row, column=c).value
        if isinstance(vac_val, (int, float)):
            assert vac_val <= 0, (
                f"Vacancy Loss Y{c - y0_col} must be written as a signed "
                f"negative value; got {vac_val!r}"
            )


async def test_rent_y1_monthly_excludes_occupancy(session: AsyncSession):
    """``s_rev_<slug>_y1_monthly`` must be count × rent (pre-vacancy).

    Regression: pre-fix formula multiplied by occupancy_pct, making
    Pro Forma's Gross Revenue line already net of vacancy — which
    then double-counted when the EGI formula added a positive vacancy
    cell on top. Y1 monthly is now the true gross potential rent;
    occupancy applies via the Vacancy Loss row and the EGI subtraction.
    """
    scenario = await _seed_scenario(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)

    rent_names = [
        n for n in wb.defined_names
        if n.startswith("s_rev_") and n.endswith("_y1_monthly")
    ]
    assert rent_names, "no s_rev_*_y1_monthly named ranges emitted"

    for name in rent_names:
        dn = wb.defined_names[name]
        for sheet, coord in dn.destinations:
            cell = wb[sheet][coord]
            formula = cell.value
            assert isinstance(formula, str) and formula.startswith("=")
            assert "occupancy_pct" not in formula, (
                f"{name} must NOT multiply by occupancy "
                f"(applied on Pro Forma instead); got {formula!r}"
            )
