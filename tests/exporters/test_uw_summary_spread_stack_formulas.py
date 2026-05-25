"""Underwriting Summary → Spread Stack cells must be formulas
referencing the named KPI cells above them, not engine-computed
scalars. LP edits to revenue / OpEx / Use lines re-derive NOI + TPC
upstream and the Spread Stack rows follow without re-running the
engine.

Contract:

  1. Cap Rate on Cost = ``=IFERROR(s_combined_noi/s_total_project_cost,0)``
  2. Cap Rate Spread (vs RFR) = ``=IFERROR(s_spread_cap_pct-s_rfr_pct,0)``
  3. Levered IRR Spread (vs Cap Rate) =
     ``=IFERROR(s_combined_irr-s_spread_cap_pct,0)``
  4. Levered IRR Spread (vs RFR) =
     ``=IFERROR(s_combined_irr-s_rfr_pct,0)``

Each of the four named ranges (`s_spread_cap_pct`, `s_cap_rate_spread`,
`s_irr_spread`, `s_irr_rfr_spread`) must still be registered so any
downstream sheet (Cover, Investor Returns) can reference them.
"""
from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


pytestmark = pytest.mark.asyncio


async def _seed(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="UW Summary Spread")
    deal_model, _, stream, opex = await seed_deal_model_with_financials(
        session, opp, user
    )
    stream.active_in_phases = ["stabilized"]
    opex.active_in_phases = ["stabilized"]
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find_row(ws, label_substr: str) -> int | None:
    needle = label_substr.lower()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and needle in v.lower():
            return r
    return None


async def test_cap_rate_on_cost_is_formula(session: AsyncSession) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Summary"]
    row = _find_row(ws, "Cap Rate on Cost")
    assert row is not None
    v = ws.cell(row=row, column=2).value
    assert isinstance(v, str) and v.startswith("="), f"got {v!r}"
    assert "s_combined_noi" in v and "s_total_project_cost" in v


async def test_cap_rate_spread_is_formula(session: AsyncSession) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Summary"]
    row = _find_row(ws, "Cap Rate Spread (vs RFR)")
    assert row is not None
    v = ws.cell(row=row, column=2).value
    assert isinstance(v, str) and v.startswith("="), f"got {v!r}"
    assert "s_spread_cap_pct" in v and "s_rfr_pct" in v


async def test_levered_irr_spread_vs_cap_is_formula(
    session: AsyncSession,
) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Summary"]
    row = _find_row(ws, "Levered IRR Spread (vs Cap Rate)")
    assert row is not None
    v = ws.cell(row=row, column=2).value
    assert isinstance(v, str) and v.startswith("="), f"got {v!r}"
    assert "s_combined_irr" in v and "s_spread_cap_pct" in v


async def test_levered_irr_spread_vs_rfr_is_formula(
    session: AsyncSession,
) -> None:
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Summary"]
    row = _find_row(ws, "Levered IRR Spread (vs RFR)")
    assert row is not None
    v = ws.cell(row=row, column=2).value
    assert isinstance(v, str) and v.startswith("="), f"got {v!r}"
    assert "s_combined_irr" in v and "s_rfr_pct" in v


async def test_spread_stack_named_ranges_registered(
    session: AsyncSession,
) -> None:
    """All four spread cells stay registered as defined names so
    downstream sheets can reference them."""
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session)
    wb = load_workbook(BytesIO(blob), data_only=False)
    defined = set(wb.defined_names)
    for name in (
        "s_spread_cap_pct",
        "s_cap_rate_spread",
        "s_irr_spread",
        "s_irr_rfr_spread",
    ):
        assert name in defined, f"{name} missing from defined names"
