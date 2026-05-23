"""Phase D commit 8: Combined Equity Multiple on both UW Summary and
Returns sheet must be a SUMIF formula over ``r_uw_cf_levered`` so LP
edits to upstream NOI / debt service / capital events propagate without
re-running the engine.

Contract:

  1. UW Summary "Combined Equity Multiple" cell starts with ``=`` and
     references ``r_uw_cf_levered``.
  2. Returns "Combined Equity Multiple (scenario)" cell is a formula
     (when scenario has real equity stack) referencing the same range.
  3. r_uw_cf_levered defined name is registered.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engines.cashflow import compute_cash_flows
from app.exporters.investor_export import export_investor_workbook
from app.models.capital import (
    CapitalModule,
    CapitalModuleProject,
    EquityRole,
    VehicleType,
)
from app.models.project import Project
from tests.conftest import (
    seed_deal_model_with_financials,
    seed_opportunity,
    seed_org,
)


async def _seed(session: AsyncSession):
    org, user = await seed_org(session)
    opp = await seed_opportunity(session, org, user, name="EM Formula Smoke")
    deal_model, _, _, _ = await seed_deal_model_with_financials(
        session, opp, user
    )
    project = (
        await session.execute(
            select(Project).where(Project.scenario_id == deal_model.id)
        )
    ).scalar_one()
    debt = CapitalModule(
        scenario_id=deal_model.id,
        label="Senior Loan",
        vehicle_type=VehicleType.debt.value,
        stack_position=1,
        source={
            "amount": "500000", "interest_rate_pct": 6.5,
            "amort_term_years": 30, "hold_term_years": 10,
        },
        carry={"carry_type": "pi", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    equity = CapitalModule(
        scenario_id=deal_model.id,
        label="LP Equity",
        vehicle_type=VehicleType.equity.value,
        equity_role=EquityRole.lp.value,
        stack_position=2,
        source={"amount": "250000"},
        carry={"carry_type": "none", "payment_frequency": "monthly"},
        exit_terms={"exit_type": "full_payoff", "trigger": "sale"},
        active_phase_start="acquisition", active_phase_end="exit",
    )
    session.add_all([debt, equity])
    await session.flush()
    session.add_all([
        CapitalModuleProject(
            capital_module_id=debt.id, project_id=project.id,
            amount=Decimal("500000"),
        ),
        CapitalModuleProject(
            capital_module_id=equity.id, project_id=project.id,
            amount=Decimal("250000"),
        ),
    ])
    await session.flush()
    await compute_cash_flows(deal_model.id, session)
    await session.commit()
    return deal_model


def _find_row(ws, label_substr: str) -> int | None:
    needle = label_substr.lower()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and needle in v.lower():
            return r
    return None


@pytest.mark.parametrize("profile", ["internal", "lp", "lender"])
async def test_uw_summary_em_is_sumif_formula(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Underwriting Summary"]
    r = _find_row(ws, "combined equity multiple")
    assert r is not None, f"profile={profile}: row missing"
    val = ws.cell(row=r, column=2).value
    assert isinstance(val, str) and val.startswith("="), (
        f"profile={profile}: must be formula; got {val!r}"
    )
    assert "SUMIF" in val and "r_uw_cf_levered" in val, (
        f"profile={profile}: missing SUMIF/range ref; got {val!r}"
    )


@pytest.mark.parametrize("profile", ["internal", "lp"])
async def test_returns_em_is_sumif_formula(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    ws = wb["Investor Returns"]
    r = _find_row(ws, "combined equity multiple (scenario)")
    assert r is not None, f"profile={profile}: row missing"
    val = ws.cell(row=r, column=2).value
    assert isinstance(val, str) and val.startswith("="), (
        f"profile={profile}: must be formula; got {val!r}"
    )
    assert "SUMIF" in val and "r_uw_cf_levered" in val, (
        f"profile={profile}: missing SUMIF/range ref; got {val!r}"
    )


@pytest.mark.parametrize("profile", ["internal", "lp", "lender"])
async def test_cf_levered_named_range_present(
    session: AsyncSession, profile: str
):
    scenario = await _seed(session)
    blob = await export_investor_workbook(scenario.id, session, profile=profile)
    wb = load_workbook(BytesIO(blob), data_only=False)
    assert "r_uw_cf_levered" in wb.defined_names, (
        f"profile={profile} missing r_uw_cf_levered defined name"
    )
