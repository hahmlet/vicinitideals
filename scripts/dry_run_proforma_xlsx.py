"""Dry-run the proforma parser against an xlsx file.

Usage:
    uv run python scripts/dry_run_proforma_xlsx.py <path.xlsx> [sheet_name] [property_column]

Mirrors the production xlsx path:
    openpyxl.load_workbook → _sheet_to_text → build_opex_prompt → LLM → postprocess
"""
from __future__ import annotations

import json
import sys

import openpyxl

from app.config import settings
from app.tasks.proforma_parse import (
    ParsedExpenses,
    _llm_client,
    _postprocess_expense_lines,
    _sheet_to_text,
    build_opex_prompt,
)


def main() -> None:
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    path = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None
    property_column = sys.argv[3] if len(sys.argv) > 3 else None

    print(f"\nFile: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    ws = wb[sheet_name] if sheet_name else wb.active
    print(f"Using sheet: {ws.title!r}")
    if property_column:
        print(f"Property column filter: {property_column!r}")
    print()

    md = _sheet_to_text(ws, property_column=property_column)
    if not md:
        print("ERROR: _sheet_to_text produced no output.")
        sys.exit(2)

    print(f"Markdown length: {len(md):,} chars  |  lines: {md.count(chr(10)):,}")
    print("=" * 60)
    print("FIRST 40 LINES OF MARKDOWN:")
    print("=" * 60)
    for line in md.splitlines()[:40]:
        print(line.encode("ascii", errors="replace").decode())
    print()

    client = _llm_client()

    print("=" * 60)
    print(f"PARSING OPEX (model={settings.ollama_model})…")
    print("=" * 60)
    opex_prompt = build_opex_prompt(md)
    try:
        parsed_exp: ParsedExpenses = client.chat.completions.create(
            model=settings.ollama_model,
            response_format={"type": "json_object"},
            messages=[{"role": "user", "content": opex_prompt}],
        ).choices[0].message.content
        parsed_exp = ParsedExpenses.model_validate_json(parsed_exp)
        raw_lines = [e.model_dump() for e in parsed_exp.expense_lines]
        print("--- RAW LLM OUTPUT (before snap) ---")
        for r in raw_lines:
            print(f"    {r['original_label']!r:40s} raw_cat={r['mapped_category']!r}")
        print("--- AFTER _snap_category POST-PROCESSING ---")
        processed = _postprocess_expense_lines(raw_lines)
        for d in processed:
            conf = d.get("confidence", 0.0) or 0.0
            conf_label = "HIGH" if conf >= 0.85 else ("MID" if conf >= 0.60 else "LOW")
            flag = "  [EXCLUDED]" if not d.get("is_operating_expense", True) else ""
            print(f"  [{conf_label:4s} {conf:.2f}]{flag}")
            print(f"    {d['original_label']!r:40s} -> {d['mapped_category']!r}")
            print(f"    Annual: ${d['annual_amount']:,.0f}")
    except Exception as exc:
        print(f"OpEx parse ERROR: {exc}")


if __name__ == "__main__":
    main()
