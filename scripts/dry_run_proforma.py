"""Dry-run the pro forma parser against a local Excel file.

Usage:
    uv run python scripts/dry_run_proforma.py <xlsx_path> <revenue_sheet> <opex_sheet> [property_column]

Example:
    uv run python scripts/dry_run_proforma.py \
        "docs/models/Revenue - OpEx Examples/Green Seed Revenue and OpEx (1).xlsx" \
        "Revenue" "Adjusted Expenses"

No database or Redis touched — prints parsed results only.
"""
from __future__ import annotations

import json
import sys
from typing import Any

import openpyxl
from pydantic import BaseModel


STANDARD_OPEX_CATEGORIES = (
    "Real Estate Taxes", "Insurance", "Property Management",
    "Utilities — Water/Sewer", "Utilities — Electric", "Utilities — Gas",
    "Utilities — Trash", "Repairs & Maintenance", "Marketing & Leasing",
    "Administrative", "Payroll", "Landscaping & Snow Removal",
    "Pest Control", "Cleaning & Janitorial", "Security",
    "Resident Services", "Compliance & Legal", "Source Compliance",
    "Bank/Software Fees", "Unit Turnover", "Other",
)


class UnitTypeResult(BaseModel):
    name: str
    count: int
    avg_sqft: float
    avg_monthly_rent: float
    confidence: float


class ParsedRevenue(BaseModel):
    unit_types: list[UnitTypeResult]
    warnings: list[str] = []


class ExpenseLineResult(BaseModel):
    original_label: str
    annual_amount: float
    mapped_category: str | None
    confidence: float
    is_operating_expense: bool


class ParsedExpenses(BaseModel):
    expense_lines: list[ExpenseLineResult]
    warnings: list[str] = []


_DEBT_KEYWORDS: frozenset[str] = frozenset({
    "depreciation", "interest", "amortization", "principal",
    "debt service", "mortgage", "loan fee", "income tax",
    "total income", "total expenses", "total other", "net income",
    "net operating", "subtotal", "total revenue", "utilities total",
    "total utilities", "total opex",
})

_EXCLUDE_ROW_PREFIXES: tuple[str, ...] = (
    "total", "subtotal", "net ", "grand total",
)


def _is_debt_or_total(label: str) -> bool:
    lower = label.lower().strip()
    for kw in _DEBT_KEYWORDS:
        if kw in lower:
            return True
    for prefix in _EXCLUDE_ROW_PREFIXES:
        if lower.startswith(prefix):
            return True
    return False


def _snap_category(raw: str | None) -> str | None:
    if raw is None:
        return None
    import re

    def _norm(s: str) -> set[str]:
        s = s.lower()
        s = re.sub(r"[^a-z0-9 ]", " ", s)
        return set(s.split())

    raw_tokens = _norm(raw)
    if not raw_tokens:
        return None

    best_score = 0.0
    best_cat: str | None = None
    for cat in STANDARD_OPEX_CATEGORIES:
        cat_tokens = _norm(cat)
        if not cat_tokens:
            continue
        overlap = len(raw_tokens & cat_tokens)
        score = overlap / max(len(raw_tokens | cat_tokens), 1)
        if score > best_score:
            best_score = score
            best_cat = cat

    return best_cat if best_score >= 0.40 else None


def _postprocess_expense_lines(lines: list[dict]) -> list[dict]:
    out = []
    for line in lines:
        label = line.get("original_label", "")
        if _is_debt_or_total(label):
            line["is_operating_expense"] = False
        line["mapped_category"] = _snap_category(line.get("mapped_category"))
        out.append(line)
    return out


def _sheet_to_text(ws: Any, property_column: str | None = None, max_rows: int = 200) -> str:
    import re

    _TOTAL_HEADERS = {"total", "annual", "annualtotal", "ytd", "fullyear"}

    raw_rows: list[list[str]] = []
    for row in ws.iter_rows(max_row=min(ws.max_row, max_rows), values_only=True):
        if not any(c is not None for c in row):
            continue
        raw_rows.append([str(c).strip() if c is not None else "" for c in row])

    if not raw_rows:
        return ""

    header = raw_rows[0]
    keep: list[int] = []
    if property_column:
        keep = [0] + [i for i, h in enumerate(header) if property_column.lower() in h.lower()]
    else:
        total_col = next(
            (i for i, h in enumerate(header)
             if re.sub(r"[^a-z]", "", h.lower()) in _TOTAL_HEADERS),
            None,
        )
        if total_col is not None and total_col != 0:
            keep = [0, total_col]

    def _filter(cells: list[str]) -> list[str]:
        if keep:
            return [cells[i] for i in keep if i < len(cells)]
        return cells

    filtered = [_filter(r) for r in raw_rows]
    col_count = max(len(r) for r in filtered)
    header_row = filtered[0] + [""] * (col_count - len(filtered[0]))
    sep = ["---"] * col_count
    lines = [
        "| " + " | ".join(header_row) + " |",
        "| " + " | ".join(sep) + " |",
    ]
    for row in filtered[1:]:
        padded = row + [""] * (col_count - len(row))
        lines.append("| " + " | ".join(padded) + " |")
    return "\n".join(lines)


def _llm_client() -> Any:
    import instructor
    from openai import OpenAI
    raw = OpenAI(base_url="http://192.168.1.184:11434/v1", api_key="ollama")
    return instructor.from_openai(raw, mode=instructor.Mode.JSON)


def main() -> None:
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(1)

    xlsx_path = sys.argv[1]
    revenue_sheet = sys.argv[2]
    opex_sheet = sys.argv[3]
    property_column = sys.argv[4] if len(sys.argv) > 4 else None

    print(f"\nFile:          {xlsx_path}")
    print(f"Revenue sheet: {revenue_sheet}")
    print(f"OpEx sheet:    {opex_sheet}")
    if property_column:
        print(f"Property col:  {property_column}")
    print()

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    print(f"Sheets in file: {wb.sheetnames}\n")

    revenue_text = _sheet_to_text(wb[revenue_sheet]) if revenue_sheet in wb.sheetnames else ""
    opex_text = _sheet_to_text(wb[opex_sheet], property_column=property_column) if opex_sheet in wb.sheetnames else ""

    print("=" * 60)
    print("REVENUE SHEET (first 20 lines sent to LLM):")
    print("=" * 60)
    for line in revenue_text.splitlines()[:20]:
        print(line)
    print()

    print("=" * 60)
    print("OPEX SHEET (first 30 lines sent to LLM):")
    print("=" * 60)
    for line in opex_text.splitlines()[:30]:
        print(line)
    print()

    client = _llm_client()
    categories_str = "\n".join(f"- {c}" for c in STANDARD_OPEX_CATEGORIES)

    # --- Revenue ---
    print("=" * 60)
    print("PARSING REVENUE (calling Ollama)…")
    print("=" * 60)
    rev_prompt = (
        "You are a real estate financial analyst. The following is tabular data from a "
        "spreadsheet's revenue or rent-roll sheet. Group the units into distinct unit types "
        "(e.g. Studio, 1BR, 2BR, or developer-named types like 'Tower Small'). "
        "For each type, return the count, average square footage, and average gross monthly rent. "
        "If the sheet is a unit-by-unit roll (individual unit numbers), group similar-sized units. "
        "Only return actual residential or commercial units — exclude parking, storage, laundry, "
        "and ancillary income rows.\n\n"
        f"SHEET DATA:\n{revenue_text}"
    )
    try:
        parsed_rev: ParsedRevenue = client.chat.completions.create(
            model="qwen2.5:7b",
            response_model=ParsedRevenue,
            messages=[{"role": "user", "content": rev_prompt}],
        )
        print(json.dumps([u.model_dump() for u in parsed_rev.unit_types], indent=2))
        if parsed_rev.warnings:
            print("Warnings:", parsed_rev.warnings)
    except Exception as exc:
        print(f"Revenue parse ERROR: {exc}")

    # --- OpEx ---
    print()
    print("=" * 60)
    print("PARSING OPEX (calling Ollama)…")
    print("=" * 60)
    opex_prompt = (
        "You are a real estate financial analyst. The following is tabular data from a "
        "spreadsheet's operating expense sheet.\n\n"
        "RULES -- follow exactly:\n"
        "1. Return EVERY labeled expense row in the sheet, even if you cannot map it. "
        "Do not skip rows or merge rows together.\n"
        "2. For each row, extract the annual dollar amount. "
        "If the sheet shows monthly columns, sum them to get the annual total. "
        "If a TOTAL or annual column is present, use that instead.\n"
        "3. Map each label to EXACTLY one category from the STANDARD CATEGORIES list below. "
        "Copy the category name character-for-character. "
        "Set mapped_category=null ONLY if none of the standard categories fit at all.\n"
        "4. Set is_operating_expense=false for: debt service, mortgage interest, "
        "depreciation, loan fee amortization, principal payments, income tax. "
        "Also exclude subtotal/total rows and income rows.\n"
        "5. Set confidence 0.85-1.0 for obvious matches, 0.60-0.84 for reasonable guesses, "
        "below 0.60 when unsure.\n\n"
        f"STANDARD CATEGORIES (use exact spelling):\n{categories_str}\n\n"
        "Mapping hints:\n"
        "- 'LIFT Monitoring', 'OHCS', 'bond compliance', 'HUD monitoring' -> Source Compliance\n"
        "- 'Prop Mgmt', 'On-Site Mgmt', 'Off-Site Mgmt' -> Property Management\n"
        "- 'RE Taxes', 'Real Estate Tax', 'Property Tax' -> Real Estate Taxes\n"
        "- 'Gresham Police Fire Parks', 'Municipal assessment' -> Real Estate Taxes\n"
        "- 'Accounting', 'CPA', 'Audit', 'Professional Fees', 'Legal', 'Licenses' -> Compliance & Legal\n"
        "- 'Bank', 'NSF', 'Financing charges', 'Computer', 'Software', 'Internet' -> Bank/Software Fees\n"
        "- 'Office Supplies', 'Administrative' -> Administrative\n"
        "- 'Tenant Events', 'Resident Activities' -> Resident Services\n"
        "- 'Fire Monitoring', 'Security System' -> Security\n"
        "- 'Garbage', 'Trash', 'Waste' -> Utilities -- Trash\n"
        "- 'AR Writeoffs', 'Bad Debt', 'Receivables' -> set is_operating_expense=false\n\n"
        f"SHEET DATA:\n{opex_text}"
    )
    try:
        parsed_exp: ParsedExpenses = client.chat.completions.create(
            model="qwen2.5:7b",
            response_model=ParsedExpenses,
            messages=[{"role": "user", "content": opex_prompt}],
        )
        processed = _postprocess_expense_lines([e.model_dump() for e in parsed_exp.expense_lines])
        for d in processed:
            flag = "" if d["is_operating_expense"] else "  [EXCLUDED]"
            conf = d["confidence"]
            conf_label = "HIGH" if conf >= 0.85 else ("MID" if conf >= 0.60 else "LOW")
            print(f"  [{conf_label:4s} {conf:.2f}]{flag}")
            print(f"    {d['original_label']!r:40s} -> {d['mapped_category']!r}")
            print(f"    Annual: ${d['annual_amount']:,.0f}")
        if parsed_exp.warnings:
            print("Warnings:", parsed_exp.warnings)
    except Exception as exc:
        print(f"OpEx parse ERROR: {exc}")


if __name__ == "__main__":
    main()
