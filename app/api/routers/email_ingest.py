"""Email ingest routes: webhook endpoint, inbox list, side-by-side review UI."""

from __future__ import annotations

import base64
import hashlib
import hmac
import json
import time
import uuid
from pathlib import Path
from typing import Any

from fastapi import APIRouter, HTTPException, Request, status
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import select

from app.api.deps import DBSession, CurrentUserId
from app.config import settings

# Reject webhooks with timestamps more than 5 minutes off (Svix default)
_WEBHOOK_TOLERANCE_SECONDS = 300

# Debug-log download is gated to a single hardcoded operator email — feature only
# meaningful for the person triaging AI extraction failures.
_DEBUG_LOG_ALLOWED_EMAIL = "stephenjketch@gmail.com"

# ---------------------------------------------------------------------------
# Template setup (same dir as ui.py)
# ---------------------------------------------------------------------------
_PACKAGE_DIR = Path(__file__).resolve().parent.parent.parent
_TEMPLATES_DIR = _PACKAGE_DIR / "templates"
templates = Jinja2Templates(directory=str(_TEMPLATES_DIR))

# ---------------------------------------------------------------------------
# Two routers: api_router gets /api prefix via ROUTERS; ui_router has no prefix
# ---------------------------------------------------------------------------
router = APIRouter(tags=["email-ingest"], include_in_schema=False)
ui_router = APIRouter(include_in_schema=False)


# ===========================================================================
# API routes (mounted at /api/...)
# ===========================================================================

@router.post("/email-ingest", status_code=status.HTTP_200_OK)
async def receive_inbound_email(
    request: Request,
    session: DBSession,
) -> dict[str, Any]:
    """Resend inbound webhook. Verifies Svix signature, stores email metadata, queues Celery task to fetch full content."""
    raw_body = await request.body()
    _verify_svix_signature(request, raw_body)

    try:
        payload = json.loads(raw_body)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Invalid JSON") from exc

    if payload.get("type") != "email.received":
        return {"status": "ignored", "reason": "event type not email.received"}

    data = payload.get("data") or {}
    resend_email_id = data.get("email_id")
    if not resend_email_id:
        raise HTTPException(status_code=400, detail="Missing email_id")

    from app.models.email_ingest import InboundEmail, InboundEmailStatus
    from app.models.org import Organization

    org_result = await session.execute(select(Organization).limit(1))
    org = org_result.scalar_one_or_none()
    if org is None:
        raise HTTPException(status_code=500, detail="No organization configured")

    email_row = InboundEmail(
        id=uuid.uuid4(),
        org_id=org.id,
        sender_email=str(data.get("from") or ""),
        subject=data.get("subject") or None,
        raw_mime_b64=None,  # Fetched by Celery task via Resend API
        status=InboundEmailStatus.pending.value,
        proforma_task_ids=[],
        attachments_meta=data.get("attachments") or [],
    )
    session.add(email_row)
    await session.commit()
    await session.refresh(email_row)

    from app.tasks.email_ingest import process_inbound_email  # noqa: PLC0415
    process_inbound_email.delay(str(email_row.id), resend_email_id)

    return {"status": "accepted", "id": str(email_row.id)}


def _verify_svix_signature(request: Request, body: bytes) -> None:
    """Verify Resend (Svix) webhook signature using HMAC-SHA256.

    Raises 403 if signature invalid, secret unset, or timestamp out of tolerance.
    """
    secret = settings.resend_webhook_secret
    if not secret:
        raise HTTPException(status_code=403, detail="Webhook secret not configured")

    svix_id = request.headers.get("svix-id", "")
    svix_timestamp = request.headers.get("svix-timestamp", "")
    svix_signature = request.headers.get("svix-signature", "")
    if not (svix_id and svix_timestamp and svix_signature):
        raise HTTPException(status_code=403, detail="Missing Svix headers")

    try:
        ts = int(svix_timestamp)
    except ValueError as exc:
        raise HTTPException(status_code=403, detail="Invalid timestamp") from exc
    if abs(time.time() - ts) > _WEBHOOK_TOLERANCE_SECONDS:
        raise HTTPException(status_code=403, detail="Timestamp out of tolerance")

    # Strip "whsec_" prefix if present, then base64-decode the key
    secret_key = secret.removeprefix("whsec_")
    try:
        key_bytes = base64.b64decode(secret_key)
    except Exception as exc:
        raise HTTPException(status_code=500, detail="Malformed webhook secret") from exc

    signed_payload = f"{svix_id}.{svix_timestamp}.".encode() + body
    expected_sig = base64.b64encode(
        hmac.new(key_bytes, signed_payload, hashlib.sha256).digest()
    ).decode()

    # Header format: "v1,sig1 v1,sig2" — any version-prefixed match wins
    provided_sigs = [s.split(",", 1)[1] for s in svix_signature.split() if "," in s]
    if not any(hmac.compare_digest(expected_sig, sig) for sig in provided_sigs):
        raise HTTPException(status_code=403, detail="Invalid signature")


@router.post("/email-suggestions/{suggestion_id}/accept")
async def accept_suggestion(
    suggestion_id: uuid.UUID,
    current_user_id: CurrentUserId,
    session: DBSession,
) -> HTMLResponse:
    """Mark a suggestion as accepted. Returns updated suggestion row HTML."""
    return await _set_suggestion(suggestion_id, True, current_user_id, session)


@router.post("/email-suggestions/{suggestion_id}/reject")
async def reject_suggestion(
    suggestion_id: uuid.UUID,
    current_user_id: CurrentUserId,
    session: DBSession,
) -> HTMLResponse:
    """Mark a suggestion as rejected. Returns updated suggestion row HTML."""
    return await _set_suggestion(suggestion_id, False, current_user_id, session)


async def _set_suggestion(
    suggestion_id: uuid.UUID,
    accepted: bool,
    current_user_id: uuid.UUID,
    session: Any,
) -> HTMLResponse:
    from app.models.email_ingest import EmailDealSuggestion, InboundEmail

    suggestion = await session.get(EmailDealSuggestion, suggestion_id)
    if suggestion is None:
        raise HTTPException(status_code=404, detail="Suggestion not found")

    email_row = await session.get(InboundEmail, suggestion.inbound_email_id)
    user_org = await _get_user_org(session, current_user_id)
    if email_row is None or email_row.org_id != user_org:
        raise HTTPException(status_code=404, detail="Suggestion not found")

    suggestion.accepted = accepted
    await session.commit()

    accepted_class = "active" if accepted else ""
    return HTMLResponse(
        f'<div class="suggestion-row" id="suggestion-{suggestion_id}" '
        f'style="opacity:{0.6 if not accepted else 1}">'
        f'<div class="suggestion-label">{suggestion.field_path.replace("_", " ").title()}</div>'
        f'<div class="suggestion-value">{suggestion.suggested_value or "—"}</div>'
        f'<div class="suggestion-actions">'
        f'<span class="chip chip-accept {accepted_class}">{"✓ Accepted" if accepted else "✓ Accept"}</span>'
        f'</div></div>'
    )


# ===========================================================================
# UI routes (no /api prefix)
# ===========================================================================

@ui_router.get("/ui/email-inbox", response_class=HTMLResponse)
async def email_inbox(
    request: Request,
    current_user_id: CurrentUserId,
    session: DBSession,
) -> HTMLResponse:
    """Email inbox — lists all inbound emails for the org."""
    from app.api.routers.ui_helpers import _base_ctx, _get_counts  # noqa: PLC0415
    from app.models.email_ingest import InboundEmail
    from app.models.org import User

    user = await session.get(User, current_user_id)
    if user is None:
        raise HTTPException(status_code=401, detail="Not authenticated")

    dedup_count, conflicts_count = await _get_counts(session)

    result = await session.execute(
        select(InboundEmail)
        .where(InboundEmail.org_id == user.org_id)
        .order_by(InboundEmail.received_at.desc())
        .limit(100)
    )
    emails = result.scalars().all()

    ctx = _base_ctx(user, dedup_count, "email_inbox", conflicts_count=conflicts_count)
    ctx.update({
        "emails": emails,
        "can_view_debug_log": (user.email or "").lower() == _DEBUG_LOG_ALLOWED_EMAIL,
    })

    return templates.TemplateResponse(request, "email_inbox.html", ctx)


@ui_router.get("/ui/deals/email/{inbound_email_id}/review", response_class=HTMLResponse)
async def email_deal_review(
    request: Request,
    inbound_email_id: uuid.UUID,
    current_user_id: CurrentUserId,
    session: DBSession,
) -> HTMLResponse:
    """Config table: one row per staged proforma file. User assigns deal names
    and sheet/page ranges, then submits to create deals in bulk."""
    from app.api.routers.ui_helpers import _base_ctx, _get_counts  # noqa: PLC0415
    from app.models.email_ingest import EmailDealSuggestion, InboundEmail
    from app.models.org import User

    user = await session.get(User, current_user_id)
    if user is None:
        raise HTTPException(status_code=401, detail="Not authenticated")

    email_row = await session.get(InboundEmail, inbound_email_id)
    if email_row is None or email_row.org_id != user.org_id:
        raise HTTPException(status_code=404, detail="Email not found")

    dedup_count, conflicts_count = await _get_counts(session)

    suggestions_result = await session.execute(
        select(EmailDealSuggestion)
        .where(EmailDealSuggestion.inbound_email_id == inbound_email_id)
        .order_by(EmailDealSuggestion.field_path)
    )
    suggestions = suggestions_result.scalars().all()

    # Build attachment config list — one entry per staged proforma file.
    # For xlsx: read bytes from Redis, extract sheet names with openpyxl.
    staged_attachments = _build_staged_attachments(email_row)

    ctx = _base_ctx(user, dedup_count, "email_inbox", conflicts_count=conflicts_count)
    ctx.update({
        "email": email_row,
        "suggestions": suggestions,
        "staged_attachments": staged_attachments,
        "can_view_debug_log": (user.email or "").lower() == _DEBUG_LOG_ALLOWED_EMAIL,
    })

    return templates.TemplateResponse(request, "email_deal_review.html", ctx)


def _build_staged_attachments(email_row: Any) -> list[dict]:
    """Return one dict per staged proforma attachment, with sheet names for xlsx files."""
    import io
    import os

    import redis as _redis  # type: ignore
    from app.config import settings as _settings

    _XLSX_EXTS = {"xlsx", "xlsm", "xlsb"}
    r = _redis.from_url(_settings.redis_url, decode_responses=False)

    # Index attachment meta by task_id for fast lookup
    meta_by_task: dict[str, dict] = {}
    for m in (email_row.attachments_meta or []):
        tid = m.get("proforma_task_id")
        if tid:
            meta_by_task[tid] = m

    result = []
    for task_id in (email_row.proforma_task_ids or []):
        meta = meta_by_task.get(task_id, {})
        filename = meta.get("filename") or "attachment"
        ext = os.path.splitext(filename)[1].lower().lstrip(".")
        file_kind = "xlsx" if ext in _XLSX_EXTS else "doc"

        sheet_names: list[str] = []
        if file_kind == "xlsx":
            file_bytes = r.get(f"proforma:{task_id}:file")
            if file_bytes:
                try:
                    from openpyxl import load_workbook  # type: ignore
                    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
                    sheet_names = wb.sheetnames
                    wb.close()
                except Exception:
                    sheet_names = []

        result.append({
            "task_id": task_id,
            "filename": filename,
            "file_kind": file_kind,
            "sheet_names": sheet_names,
            "single_sheet": len(sheet_names) == 1,
            "size_bytes": meta.get("size_bytes"),
        })

    return result


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

async def _get_user_org(session: Any, user_id: uuid.UUID) -> uuid.UUID | None:
    from app.models.org import User
    user = await session.get(User, user_id)
    return user.org_id if user else None


@ui_router.post("/ui/deals/email/{inbound_email_id}/create-deals")
async def email_create_deals(
    request: Request,
    inbound_email_id: uuid.UUID,
    current_user_id: CurrentUserId,
    session: DBSession,
) -> Any:
    """Parse the deal config table submitted from email_deal_review.html.

    Creates one Deal per unique deal name (preserving row order), stores
    sheet/page choices per file in Redis, then redirects to the first
    deal's wizard.
    """
    import json as _json
    from decimal import Decimal, InvalidOperation

    import redis as _redis  # type: ignore
    from fastapi.responses import RedirectResponse
    from sqlalchemy import select as _select

    from app.api.routers.ui_helpers import (
        _auto_assign_opportunity_to_project,
        _seed_milestones,
    )
    from app.models.deal import Deal, ProjectType, UseLine, UseLinePhase
    from app.models.email_ingest import EmailDealSuggestion, InboundEmail, InboundEmailStatus
    from app.models.opportunity import Opportunity, OpportunityStatus
    from app.models.org import User
    from app.services.scenario_factory import create_scenario as _create_scenario
    from app.services.vehicle_preload import preload_equity_modules

    user = await session.get(User, current_user_id)
    if user is None:
        raise HTTPException(status_code=401, detail="Not authenticated")

    email_row = await session.get(InboundEmail, inbound_email_id)
    if email_row is None or email_row.org_id != user.org_id:
        raise HTTPException(status_code=404, detail="Email not found")

    form = await request.form()

    # Collect rows in submission order
    rows: list[dict] = []
    i = 0
    while True:
        task_id = form.get(f"task_id_{i}")
        if task_id is None:
            break
        rows.append({
            "task_id": str(task_id),
            "file_kind": str(form.get(f"file_kind_{i}") or "doc"),
            "deal_name": (str(form.get(f"deal_name_{i}") or "")).strip() or "Unnamed Deal",
            "rev_sheet": str(form.get(f"rev_sheet_{i}") or ""),
            "rev_range": str(form.get(f"rev_range_{i}") or ""),
            "opex_sheet": str(form.get(f"opex_sheet_{i}") or ""),
            "opex_range": str(form.get(f"opex_range_{i}") or ""),
            "rev_pages": str(form.get(f"rev_pages_{i}") or ""),
            "opex_pages": str(form.get(f"opex_pages_{i}") or ""),
        })
        i += 1

    if not rows:
        raise HTTPException(status_code=400, detail="No files submitted")

    # Group by deal name, preserving first-seen order
    deal_groups: dict[str, list[dict]] = {}
    for row in rows:
        deal_groups.setdefault(row["deal_name"], []).append(row)

    # Pull asking price from suggestions (shared across all deals from this email)
    acq_result = await session.execute(
        _select(EmailDealSuggestion)
        .where(EmailDealSuggestion.inbound_email_id == inbound_email_id)
        .where(EmailDealSuggestion.field_path == "acquisition_cost")
        .limit(1)
    )
    acq_suggestion = acq_result.scalar_one_or_none()
    acq_cost = Decimal("0")
    if acq_suggestion and acq_suggestion.suggested_value:
        try:
            acq_cost = Decimal(str(acq_suggestion.suggested_value))
        except InvalidOperation:
            pass

    org_id = user.org_id
    deal_type = ProjectType.value_add
    r = _redis.from_url(settings.redis_url, decode_responses=True)

    first_scenario_id: uuid.UUID | None = None

    for deal_name, deal_rows in deal_groups.items():
        opportunity = Opportunity(
            org_id=org_id,
            name=deal_name,
            opp_status=OpportunityStatus.active.value,
            source="manual",
            source_id=uuid.uuid4().hex,
            source_url="",
            created_by_user_id=user.id,
        )
        session.add(opportunity)
        await session.flush()

        top_deal = Deal(
            org_id=org_id,
            name=deal_name,
            created_by_user_id=user.id,
        )
        session.add(top_deal)
        await session.flush()

        scenario, dev_project, _ = await _create_scenario(
            session=session,
            deal_id=top_deal.id,
            deal_type=deal_type,
            user_id=user.id,
            org_id=org_id,
            opportunity_id=opportunity.id,
        )

        await _auto_assign_opportunity_to_project(opportunity, dev_project, session)
        for milestone in _seed_milestones(dev_project, deal_type):
            session.add(milestone)

        session.add(UseLine(
            project_id=dev_project.id,
            label=f"{deal_name} - Acquisition",
            phase=UseLinePhase.acquisition,
            cost_category="acquisition",
            dev_fee_basis_bucket="acquisition",
            amount=acq_cost,
            timing_type="first_day",
        ))

        await preload_equity_modules(session, scenario.id, org_id, project_id=dev_project.id)

        # Store sheet/page config per file so proforma-from-staged skips the picker
        for row in deal_rows:
            config = {
                "file_kind": row["file_kind"],
                "rev_sheet": row["rev_sheet"],
                "rev_range": row["rev_range"],
                "opex_sheet": row["opex_sheet"],
                "opex_range": row["opex_range"],
                "rev_pages": row["rev_pages"],
                "opex_pages": row["opex_pages"],
                "import_revenue": bool(row["rev_sheet"] or row["rev_pages"]),
                "import_opex": bool(row["opex_sheet"] or row["opex_pages"]),
            }
            r.set(f"proforma:{row['task_id']}:email_config", _json.dumps(config), ex=7 * 86_400)

        # Stash first file's task_id for wizard auto-load
        r.set(
            f"proforma:scenario:{scenario.id}:email_task_id",
            deal_rows[0]["task_id"],
            ex=7 * 86_400,
        )

        if first_scenario_id is None:
            first_scenario_id = scenario.id
            email_row.opportunity_id = opportunity.id

    email_row.status = InboundEmailStatus.opportunity_created.value
    await session.commit()

    return RedirectResponse(
        url=f"/models/{first_scenario_id}/builder?module=timeline&wizard=1",
        status_code=303,
    )


@ui_router.get("/ui/email-inbox/{inbound_email_id}/debug-log.txt")
async def email_debug_log(
    inbound_email_id: uuid.UUID,
    current_user_id: CurrentUserId,
    session: DBSession,
) -> Any:
    """Return AI processing debug log as plain text. Hardcoded to a single operator email."""
    from fastapi.responses import PlainTextResponse  # noqa: PLC0415

    from app.models.email_ingest import InboundEmail
    from app.models.org import User

    user = await session.get(User, current_user_id)
    if user is None or (user.email or "").lower() != _DEBUG_LOG_ALLOWED_EMAIL:
        raise HTTPException(status_code=404, detail="Not found")

    email_row = await session.get(InboundEmail, inbound_email_id)
    if email_row is None or email_row.org_id != user.org_id:
        raise HTTPException(status_code=404, detail="Email not found")

    body = email_row.debug_log or "(no debug log captured for this email)"
    header = (
        f"# Email ingest debug log\n"
        f"# inbound_email_id: {email_row.id}\n"
        f"# received_at: {email_row.received_at}\n"
        f"# status: {email_row.status}\n"
        f"# sender: {email_row.sender_email}\n"
        f"# subject: {email_row.subject}\n"
        f"# error_message: {email_row.error_message or '(none)'}\n"
        f"#" + ("-" * 60) + "\n\n"
    )
    return PlainTextResponse(
        header + body,
        headers={
            "Content-Disposition": f'attachment; filename="email-{email_row.id}-debug.txt"'
        },
    )
