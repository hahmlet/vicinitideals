"""HTML UI routes — Jinja2 templates served directly from FastAPI."""

from __future__ import annotations

import asyncio
import io
import json
import time
import uuid as _uuid_mod
from datetime import UTC, date, datetime
from zoneinfo import ZoneInfo
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from urllib.parse import quote_plus
from uuid import UUID

import httpx
from fastapi import APIRouter, File, Form, HTTPException, Query, Request, Response, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, PlainTextResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from starlette.templating import _TemplateResponse
from sqlalchemy import and_, func, literal, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

import app as _pkg
from app.api.deps import DBSession
from app.config import settings
from app.models.broker import Broker, Brokerage
from app.models.deal import Scenario, STANDARD_OPEX_CATEGORIES, USE_CATEGORY_LABELS, USE_CATEGORY_PRESETS, USE_COST_CATEGORIES, Deal, DealModel, DealOpportunity, DealStatus, IncomeStream, IncomeStreamType, OperatingExpenseLine, OperationalInputs, ProjectType, UnitMix, UseLine, UseLinePhase
from app.models.ingestion import DedupCandidate, DedupStatus, IngestJob, RecordType, SavedSearchCriteria
from app.models.org import Organization, User
from app.models.capital import CapitalModule, DrawSource, WaterfallTier
from app.models.cashflow import OperationalOutputs
from app.models.parcel import Parcel
from app.reconciliation.matcher import normalize_apn
from app.reconciliation.conflict_rules import _FIELD_MAP, auto_resolve_conflict
from app.models.portfolio import Portfolio, PortfolioProject
from app.models.milestone import DEFAULT_DURATIONS, Milestone, MilestoneType, MilestoneType as MT
from app.models.opportunity import Opportunity, OpportunitySource, OpportunityStatus
from app.models.project import Project, ProjectStatus
from app.models.scraped_listing import ScrapedListing
from app.models.realie_usage import RealieUsage
from app.models.settings import UserSetting
from app.scrapers.realie import _current_month
from app.settings.resolver import resolve_dev_fee_config

router = APIRouter(include_in_schema=False)


class _UMRow:
    """Attribute-compatible proxy for unit_mix JSONB dicts.

    Wraps a dict so code that was written against the old UnitMix ORM rows
    (using `.label`, `.unit_count`, etc.) continues to work without change.
    Unknown attributes return None, matching the old ORM nullable behaviour.
    """
    def __init__(self, d: dict) -> None:
        self.__dict__.update(d)

    def __getattr__(self, k: str):
        return None


def _ensure_unit_mix_ids(project) -> bool:
    """Backfill missing 'id' fields on unit_mix JSONB rows.

    Legacy rows uploaded before unit_mix had stable IDs end up missing the
    'id' key, which makes the delete handler unable to target them. Assign
    UUIDs in place and mark the JSONB column dirty so the update persists.
    """
    if project is None or not project.unit_mix:
        return False
    from uuid import uuid4 as _u4
    rows = list(project.unit_mix)
    changed = False
    for r in rows:
        if not r.get("id"):
            r["id"] = str(_u4())
            changed = True
    if changed:
        from sqlalchemy.orm.attributes import flag_modified
        project.unit_mix = rows
        flag_modified(project, "unit_mix")
    return changed


# ---------------------------------------------------------------------------
# Template setup
# ---------------------------------------------------------------------------

_PACKAGE_DIR = Path(_pkg.__file__).parent
_TEMPLATES_DIR = _PACKAGE_DIR / "templates"

templates = Jinja2Templates(directory=str(_TEMPLATES_DIR))


def _fmt_currency(value: object) -> str:
    if value is None:
        return "—"
    try:
        return f"${float(value):,.0f}"
    except (TypeError, ValueError):
        return "—"


def _fmt_currency_m(value: object) -> str:
    """Format as $XM (millions shorthand)."""
    if value is None:
        return "—"
    try:
        m = float(value) / 1_000_000
        return f"${m:.1f}M"
    except (TypeError, ValueError):
        return "—"


def _fmt_pct(value: object) -> str:
    if value is None:
        return "—"
    try:
        v = float(value)
        # Values stored as fractions (0.0–1.0) → multiply to get percentage
        if v <= 1.0:
            v *= 100
        return f"{v:.2f}%"
    except (TypeError, ValueError):
        return "—"


def _fmt_multiple(value: object) -> str:
    if value is None:
        return "—"
    try:
        return f"{float(value):.1f}×"
    except (TypeError, ValueError):
        return "—"


def _fmt_number(value: object) -> str:
    if value is None:
        return "—"
    try:
        return f"{float(value):,.0f}"
    except (TypeError, ValueError):
        return "—"


templates.env.filters["currency"] = _fmt_currency
templates.env.filters["currency_m"] = _fmt_currency_m
templates.env.filters["pct"] = _fmt_pct
templates.env.filters["multiple"] = _fmt_multiple
templates.env.filters["number_fmt"] = _fmt_number
templates.env.filters["urlencode"] = quote_plus

_PACIFIC = ZoneInfo("America/Los_Angeles")
_PROXYON_STATUS_CACHE_TTL_SECONDS = 3600
_proxyon_status_lock = asyncio.Lock()
_proxyon_status_cache: dict[str, Any] = {
    "fetched_monotonic": 0.0,
    "status_label": "Not Configured",
    "connected": False,
    "remaining_gb": None,
    "expected_days_left": None,
    "account_balance_usd": None,
    "active_subscription_id": None,
    "datacenter_count_live": None,   # active proxies from /datacenter/list
    "datacenter_by_country": None,   # dict {"us": 5, ...} from live list
    "checked_at": None,
}

# ---------------------------------------------------------------------------
# Data Cleanup (Dedup + Conflict Resolution) helpers
# ---------------------------------------------------------------------------

# Fields shown in the side-by-side listing comparison.
# Keys are ScrapedListing ORM attribute names (not column names).
_LISTING_COMPARE_FIELDS: list[tuple[str, str]] = [
    ("address_raw",        "Address"),
    ("zip_code",           "ZIP Code"),
    ("asking_price",       "Asking Price"),
    ("units",              "Units"),
    ("gba_sqft",           "Bldg SqFt"),
    ("lot_sqft",           "Lot SqFt"),
    ("year_built",         "Year Built"),
    ("year_renovated",     "Year Renovated"),
    ("cap_rate",           "Cap Rate"),
    ("noi",                "NOI"),
    ("proforma_cap_rate",  "Cap Rate (Pro Forma)"),
    ("proforma_noi",       "NOI (Pro Forma)"),
    ("property_type",      "Property Type"),
    ("zoning",             "Zoning"),
    ("apn",                "APN"),
    ("occupancy_pct",      "Occupancy %"),
    ("price_per_unit",     "Price/Unit"),
    ("price_per_sqft",     "Price/SqFt"),
    ("class_",             "Class"),
    ("stories",            "Stories"),
    ("buildings",          "Buildings"),
    ("status",             "Listing Status"),
    ("source",             "Source"),
]

_ALLOWED_OVERRIDE_FIELDS: frozenset[str] = frozenset(f for f, _ in _LISTING_COMPARE_FIELDS)


def _fmt_cmp(val: Any) -> str:
    """Format a field value for the comparison table."""
    if val is None:
        return "—"
    if isinstance(val, Decimal):
        f = float(val)
        if f >= 10_000:
            return f"${f:,.0f}"
        if f >= 1:
            return f"{f:,.2f}"
        return f"{f:.4f}"
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val)


def _build_listing_compare(
    a: ScrapedListing, b: ScrapedListing
) -> dict[str, list[dict[str, str]]]:
    conflicts: list[dict[str, str]] = []
    matches:   list[dict[str, str]] = []
    for field, label in _LISTING_COMPARE_FIELDS:
        val_a = getattr(a, field, None)
        val_b = getattr(b, field, None)
        fmt_a = _fmt_cmp(val_a)
        fmt_b = _fmt_cmp(val_b)
        entry = {"field": field, "label": label, "val_a": fmt_a, "val_b": fmt_b}
        if fmt_a != fmt_b and not (fmt_a == "—" and fmt_b == "—"):
            conflicts.append(entry)
        else:
            matches.append(entry)
    return {"conflicts": conflicts, "matches": matches}


def _record_type_str(rt: Any) -> str:
    return str(getattr(rt, "value", rt))


async def _load_listings_for_candidates(
    candidates: list[DedupCandidate], session: AsyncSession
) -> dict[_uuid_mod.UUID, ScrapedListing]:
    ids: set[_uuid_mod.UUID] = set()
    for c in candidates:
        if _record_type_str(c.record_a_type) == RecordType.listing.value:
            ids.add(c.record_a_id)
        if _record_type_str(c.record_b_type) == RecordType.listing.value:
            ids.add(c.record_b_id)
    if not ids:
        return {}
    rows = (await session.execute(
        select(ScrapedListing).where(ScrapedListing.id.in_(ids))
    )).scalars()
    return {l.id: l for l in rows}


def _candidate_row(
    c: DedupCandidate,
    listings_by_id: dict[_uuid_mod.UUID, ScrapedListing],
) -> dict[str, Any]:
    def record_label(rt: str, rid: _uuid_mod.UUID) -> tuple[str, str]:
        if rt == RecordType.listing.value:
            l = listings_by_id.get(rid)
            if l:
                addr = l.address_raw or l.full_address or "Unknown address"
                return addr, l.source.title()
        return f"{rt} …{str(rid)[-6:]}", rt.title()

    a_type = _record_type_str(c.record_a_type)
    b_type = _record_type_str(c.record_b_type)
    addr_a, src_a = record_label(a_type, c.record_a_id)
    addr_b, src_b = record_label(b_type, c.record_b_id)
    score = c.confidence_score
    tier = "high" if score >= 0.85 else "mid" if score >= 0.60 else "low"
    return {
        "id": str(c.id),
        "confidence": score,
        "tier": tier,
        "conflict_type": f"{a_type.title()} × {b_type.title()}",
        "record_a_address": addr_a,
        "record_a_source": src_a,
        "record_b_address": addr_b,
        "record_b_source": src_b,
        "match_signals": c.match_signals or {},
        "status": _record_type_str(c.status),
        "resolved_at": c.resolved_at,
    }


# ---------------------------------------------------------------------------
# Display mappings
# ---------------------------------------------------------------------------

_STATUS_DISPLAY: dict[str, tuple[str, str]] = {
    "hypothetical": ("Evaluation", "badge-blue"),
    "active": ("Execution", "badge-green"),
    "archived": ("Archived", "badge-gray"),
    "evaluation": ("Evaluation", "badge-blue"),
    "execution": ("Execution", "badge-green"),
    "under_contract": ("Under Contract", "badge-yellow"),
    "closed": ("Closed", "badge-gray"),
}

_TYPE_DISPLAY: dict[str, str] = {
    "acquisition": "Acquisition",
    "value_add": "Value-Add",
    "conversion": "Conversion",
    "new_construction": "New Construction",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _primary_scenario(deal: Deal) -> DealModel | None:
    """Return the active Scenario (financial plan) for a Deal."""
    active = [s for s in deal.scenarios if s.is_active]
    if active:
        return active[0]
    if deal.scenarios:
        return sorted(deal.scenarios, key=lambda s: s.version, reverse=True)[0]
    return None


def _first_opportunity(deal: Deal) -> Opportunity | None:
    """Return the first Opportunity linked to a Deal via Scenario→Project→Opportunity."""
    for scenario in (deal.scenarios or []):
        for proj in (scenario.projects or []):
            if proj.opportunity is not None:
                return proj.opportunity
    return None


def _deal_address(deal: Deal) -> str | None:
    opp = _first_opportunity(deal)
    if opp is None:
        return None
    return opp.address_normalized or opp.address_raw


def _deal_building_description(deal: Deal) -> str | None:
    """Build a short building description from Opportunity physical attributes."""
    opp = _first_opportunity(deal)
    if opp is None:
        return None
    parts: list[str] = []
    unit_count = opp.units if opp.units is not None else None
    if unit_count:
        parts.append(f"{unit_count} units")
    sqft = opp.gba_sqft if opp.gba_sqft is not None else None
    if sqft:
        parts.append(f"{int(float(sqft)):,} sqft")
    if parts:
        return " · ".join(parts)
    return None


def _build_deal_row(deal: Deal) -> dict:
    scenario = _primary_scenario(deal)
    opp = _first_opportunity(deal)
    outputs = scenario.operational_outputs if scenario else None
    status_key = str(opp.status.value if opp and hasattr(opp.status, "value") else (opp.status if opp else "active"))
    status_display, status_badge = _STATUS_DISPLAY.get(status_key, ("Unknown", "badge-gray"))
    type_key = str(scenario.project_type.value if scenario and hasattr(scenario.project_type, "value") else (scenario.project_type if scenario else ""))
    return {
        "id": str(deal.id),
        "name": deal.name,
        "status": status_key,
        "status_display": status_display,
        "status_badge": status_badge,
        "type_display": _TYPE_DISPLAY.get(type_key, "—") if scenario else "—",
        "primary_model_name": scenario.name if scenario else None,
        "primary_model_id": str(scenario.id) if scenario else None,
        "address": _deal_address(deal),
        "building_description": _deal_building_description(deal),
        "noi": float(outputs.noi_stabilized) if outputs and outputs.noi_stabilized is not None else None,
        "irr": float(outputs.project_irr_levered) if outputs and outputs.project_irr_levered is not None else None,
        "equity_multiple": None,  # TODO: load from SensitivityResult (needs join)
        "last_updated_fmt": deal.created_at.strftime("%b %-d, %Y") if deal.created_at else None,
    }


# Maps UI filter value → DB enum. Statuses not in this map (under_contract, closed)
# don't exist in the DB yet — selecting them returns 0 results intentionally.
_STATUS_DB_MAP = {
    "evaluation": OpportunityStatus.hypothetical,
    "execution": OpportunityStatus.active,
}
_VALID_STATUS_FILTERS = {"evaluation", "execution", "under_contract", "closed"}


async def _load_deals(
    session: DBSession,
    status_filter=None,
    type_filter=None,
    model_filter=None,
    q: str = "",
    include_archived: bool = False,
    hide_test: bool = False,
) -> list[Deal]:
    """Load Deals with their Scenarios and linked Opportunities for the deals page."""
    stmt = (
        select(Deal)
        .options(
            selectinload(Deal.scenarios)
                .selectinload(DealModel.projects)
                .selectinload(Project.opportunity),
            selectinload(Deal.scenarios).selectinload(DealModel.operational_outputs),
        )
        .order_by(Deal.created_at.desc())
    )

    if not include_archived:
        stmt = stmt.where(Deal.status != DealStatus.archived)

    if hide_test:
        stmt = stmt.where(
            ~Deal.name.ilike("%e2e%") &
            ~Deal.name.op("~*")(r"phase\s+\w+\s+test\s+\w+")
        )

    if q:
        stmt = stmt.where(Deal.name.ilike(f"%{q}%"))

    result = await session.execute(stmt)
    deals = list(result.scalars().unique())

    statuses = _as_list(status_filter)
    if statuses:
        known = [s for s in statuses if s in _STATUS_DB_MAP]
        targets = {_STATUS_DB_MAP[s] for s in known}
        if targets:
            # Deals with no linked opportunity cannot be filtered by status — keep them.
            # Deals with a linked opportunity are kept only if their opp_status matches.
            deals = [
                d for d in deals
                if _first_opportunity(d) is None or _first_opportunity(d).opp_status in targets
            ]
        # If none of the selected status values map to DB values, no filtering is applied.

    model_filters = _as_list(model_filter)
    has_primary = "has" in model_filters
    no_primary = "none" in model_filters
    if has_primary and not no_primary:
        deals = [d for d in deals if _primary_scenario(d) is not None]
    elif no_primary and not has_primary:
        deals = [d for d in deals if _primary_scenario(d) is None]

    types = _as_list(type_filter)
    if types:
        deals = [
            d for d in deals
            if _primary_scenario(d) is None or str(getattr(_primary_scenario(d).project_type, "value", _primary_scenario(d).project_type)) in types
        ]

    return deals


# Phase color palette (milestone_type → CSS class)
_PHASE_COLORS: dict[str, str] = {
    "offer_made":            "gantt-phase-offer",
    "under_contract":        "gantt-phase-contract",
    "close":                 "gantt-phase-close",
    "pre_development":       "gantt-phase-predev",
    "construction":          "gantt-phase-construction",
    "operation_lease_up":    "gantt-phase-leaseup",
    "operation_stabilized":  "gantt-phase-stabilized",
    "divestment":            "gantt-phase-exit",
}

_PHASE_LABELS: dict[str, str] = {
    "offer_made":           "Offer",
    "under_contract":       "Under Contract",
    "close":                "Close",
    "pre_development":      "Pre-Dev",
    "construction":         "Construction",
    "operation_lease_up":   "Lease-Up",
    "operation_stabilized": "Stabilized",
    "divestment":           "Divestment",
}


_GANTT_DISPLAY_CAPS: dict[str, int] = {
    "operation_stabilized": 730,   # show max 2 years; actual dates still shown in tooltips
    "operation_lease_up": 365,     # show max 1 year
}


_GANTT_DISPLAY_MINS: dict[str, int] = {
    "divestment": 30,   # single-day event needs visual presence on multi-year Gantt
}


def _apply_display_positions(bars: list[dict]) -> None:
    """Add display_start_day / display_duration_days, capping long hold phases.

    Uses actual calendar positions (not sequential cursor) so concurrent
    phases render at their real dates in the per-row layout.
    Sets is_truncated=True when the bar is shorter than its real duration.
    """
    for bar in bars:
        phase = bar.get("phase_key", "")
        cap = _GANTT_DISPLAY_CAPS.get(phase, bar["duration_days"])
        display_dur = min(bar["duration_days"], cap)
        min_dur = _GANTT_DISPLAY_MINS.get(phase)
        if min_dur and display_dur < min_dur:
            display_dur = min_dur
        bar["display_duration_days"] = display_dur
        bar["display_start_day"] = bar["start_day"]
        bar["is_truncated"] = display_dur < bar["duration_days"]


_NON_STAB_PHASES: frozenset[str] = frozenset({
    "offer_made", "under_contract", "close", "pre_development",
    "construction", "operation_lease_up", "divestment",
})


def _override_stabilized_cap(raw_rows: "list[dict]") -> None:
    """Cap operation_stabilized bars to end ~3 months after the last non-stabilized phase.

    This keeps the stabilized bar short enough to be readable while still
    indicating that operations continue indefinitely (via the truncation fade).
    """
    g_max_non_stab = 0
    for row in raw_rows:
        for bar in row["bars"]:
            if bar.get("phase_key") in _NON_STAB_PHASES:
                end = bar["start_day"] + bar["duration_days"]
                if end > g_max_non_stab:
                    g_max_non_stab = end

    if g_max_non_stab == 0:
        return  # no non-stab phases; keep static cap

    _THREE_MONTHS = 91

    for row in raw_rows:
        for bar in row["bars"]:
            if bar.get("phase_key") == "operation_stabilized":
                cap = max(30, g_max_non_stab + _THREE_MONTHS - bar["start_day"])
                actual_dur = bar["duration_days"]
                bar["display_duration_days"] = min(actual_dur, cap)
                bar["is_truncated"] = bar["display_duration_days"] < actual_dur


def _extract_milestone_bars(
    project: "Project",
    shared_epoch: "date | None" = None,
    milestones: "list | None" = None,
) -> "tuple[list[dict], date | None, bool]":
    """Extract Gantt bars from a project's milestones.

    Returns (bars, epoch_used, has_dates).
    epoch_used is the date origin for start_day values (None if no dates).
    has_dates is True when at least one anchor date was resolved.
    start_day values are relative to shared_epoch when provided.
    """
    from datetime import timedelta as _td

    milestones = sorted(milestones or project.milestones, key=lambda m: m.sequence_order)
    if not milestones:
        return [], shared_epoch, False

    m_map = {m.id: m for m in milestones}
    has_dates = any(m.target_date for m in milestones)
    bars: list[dict] = []
    epoch = shared_epoch

    if has_dates:
        for m in milestones:
            start = m.computed_start(m_map)
            if start is None and m.target_date:
                start = m.target_date
            if start is None:
                continue
            end = m.computed_end(m_map)
            if epoch is None:
                epoch = start
            start_day = (start - epoch).days
            dur = m.duration_days if m.duration_days > 0 else max(1, (end - start).days if end else 1)
            end_day = start_day + dur
            m_type = m.milestone_type.value if hasattr(m.milestone_type, "value") else m.milestone_type
            bars.append({
                "phase_key": m_type,
                "label": _PHASE_LABELS.get(m_type, m_type),
                "color_class": _PHASE_COLORS.get(m_type, "gantt-phase-other"),
                "start_day": start_day,
                "duration_days": dur,
                "end_day": end_day,
                "start_fmt": start.strftime("%b %Y"),
                "end_fmt": (epoch + _td(days=end_day)).strftime("%b %Y") if epoch else "",
            })
    else:
        cursor = 0
        for m in milestones:
            dur = m.duration_days if m.duration_days > 0 else 30
            m_type = m.milestone_type.value if hasattr(m.milestone_type, "value") else m.milestone_type
            bars.append({
                "phase_key": m_type,
                "label": _PHASE_LABELS.get(m_type, m_type),
                "color_class": _PHASE_COLORS.get(m_type, "gantt-phase-other"),
                "start_day": cursor,
                "duration_days": dur,
                "end_day": cursor + dur,
                "start_fmt": "",
                "end_fmt": "",
            })
            cursor += dur

    _apply_display_positions(bars)
    return bars, epoch, has_dates


def _apply_pct_positions(rows: list[dict], global_min: int, global_max: int) -> None:
    """Mutate each bar in rows to add left_pct / width_pct using display positions."""
    total_span = max(global_max - global_min, 1)
    for row in rows:
        for bar in row["bars"]:
            start = bar.get("display_start_day", bar["start_day"])
            dur = bar.get("display_duration_days", bar["duration_days"])
            bar["left_pct"] = round(100 * (start - global_min) / total_span, 2)
            bar["width_pct"] = max(round(100 * dur / total_span, 2), 1.5)


def _compute_gantt_axis(
    epoch: "date | None",
    global_min_day: int,
    global_max_day: int,
    has_dates: bool,
) -> "tuple[list[dict], list[dict]]":
    """Return (month_ticks, year_spans) with left_pct coordinates for the Gantt time axis."""
    import datetime as _dt

    total_span = max(global_max_day - global_min_day, 1)

    def _pct(day_offset: int) -> float:
        return round(100.0 * max(0, day_offset) / total_span, 2)

    if not has_dates or epoch is None:
        # Relative mode: 30-day pseudo-months, 360-day pseudo-years
        month_ticks: list[dict] = []
        day = 30
        while day < total_span:
            is_yr = day > 0 and day % 360 == 0
            month_ticks.append({"left_pct": _pct(day), "label": f"M{day // 30 + 1}", "is_year_start": is_yr})
            day += 30
        year_spans: list[dict] = []
        y, yr = 0, 1
        while y < total_span:
            end = min(y + 360, total_span)
            year_spans.append({"label": f"Year {yr}", "left_pct": _pct(y), "width_pct": round(_pct(end) - _pct(y), 2)})
            y += 360
            yr += 1
        return month_ticks, year_spans

    # Calendar mode
    start_date = epoch + _dt.timedelta(days=global_min_day)
    end_date = epoch + _dt.timedelta(days=global_max_day)

    def _date_pct(d: "_dt.date") -> float:
        return _pct((d - epoch).days - global_min_day)

    # Month ticks: first of each calendar month within the range
    month_ticks = []
    cur = start_date.replace(day=1)
    # Don't skip the starting month — it may start mid-month but we still want the label
    while cur <= end_date:
        lp = _date_pct(cur)
        if 0 <= lp < 100:
            month_ticks.append({"left_pct": lp, "label": cur.strftime("%b").upper(), "is_year_start": cur.month == 1})
        cur = (cur.replace(month=cur.month + 1) if cur.month < 12 else _dt.date(cur.year + 1, 1, 1))

    # Year spans
    year_spans = []
    for yr in range(start_date.year, end_date.year + 1):
        s = max(0.0, _date_pct(_dt.date(yr, 1, 1)))
        e = min(100.0, _date_pct(_dt.date(yr + 1, 1, 1)))
        if e <= 0 or s >= 100:
            continue
        year_spans.append({"label": str(yr), "left_pct": round(s, 2), "width_pct": round(max(0.0, e - s), 2)})

    return month_ticks, year_spans


def _gantt_apply_pct(bars: list[dict], g_min: int, g_max: int) -> None:
    """Mutate bars in-place to add left_pct / width_pct."""
    total_span = max(g_max - g_min, 1)
    for bar in bars:
        start = bar.get("display_start_day", bar["start_day"])
        dur = bar.get("display_duration_days", bar["duration_days"])
        bar["left_pct"] = round(100 * (start - g_min) / total_span, 2)
        bar["width_pct"] = max(round(100 * dur / total_span, 2), 1.5)


def _bars_to_phase_rows(bars: list[dict]) -> list[dict]:
    """Convert a list of bar dicts to phase-type rows for the Gantt v2 template."""
    return [{
        "type": "phase",
        "phase_key": b.get("phase_key", ""),
        "label": b["label"],
        "color_class": b["color_class"],
        "left_pct": b["left_pct"],
        "width_pct": b["width_pct"],
        "start_fmt": b.get("start_fmt", ""),
        "end_fmt": b.get("end_fmt", ""),
        "is_truncated": b.get("is_truncated", False),
        "duration_days": b.get("duration_days", 0),
    } for b in bars]


def _build_gantt_rows(deal: "Deal") -> "dict | None":
    """Build Gantt data for deal_detail.html (Gantt v2).

    Returns a dict with keys: has_dates, month_ticks, year_spans, rows.
    rows is a flat list of project_header and phase items.
    Returns None if no milestone data.
    """
    active_scenario = _primary_scenario(deal)
    if active_scenario is None or not active_scenario.projects:
        return None

    raw_rows: list[dict] = []
    epoch = None
    any_has_dates = False

    for project in active_scenario.projects:
        bars, epoch, hd = _extract_milestone_bars(project, epoch)
        if not bars:
            continue
        any_has_dates = any_has_dates or hd
        raw_rows.append({"project_name": project.name, "bars": bars})

    if not raw_rows:
        return None

    # Override stabilized bar length to 3 months past last non-stab phase
    _override_stabilized_cap(raw_rows)

    # Compute global extent AFTER the override (stabilized cap changes global_max)
    all_bars = [b for row in raw_rows for b in row["bars"]]
    g_min = min(b["display_start_day"] for b in all_bars)
    g_max = max(b["display_start_day"] + b["display_duration_days"] for b in all_bars)

    for row in raw_rows:
        _gantt_apply_pct(row["bars"], g_min, g_max)

    multi = len(raw_rows) > 1
    rows: list[dict] = []
    for raw_row in raw_rows:
        if multi:
            rows.append({"type": "project_header", "name": raw_row["project_name"]})
        rows.extend(_bars_to_phase_rows(raw_row["bars"]))

    month_ticks, year_spans = _compute_gantt_axis(epoch, g_min, g_max, any_has_dates)
    return {"has_dates": any_has_dates, "month_ticks": month_ticks, "year_spans": year_spans, "rows": rows}


def _build_portfolio_gantt(portfolio_entries: "list[tuple[str, str, Deal]]") -> "dict | None":
    """Build multi-deal Gantt data for portfolio_detail.html (Gantt v2).

    portfolio_entries: list of (deal_name, scenario_name, Deal) tuples.
    All deals share one global epoch so they align on the same calendar axis.
    """
    raw: list[tuple[str, list[dict]]] = []
    global_epoch = None
    any_has_dates = False

    for deal_name, scenario_name, deal in portfolio_entries:
        active_scenario = _primary_scenario(deal)
        if active_scenario is None or not active_scenario.projects:
            continue
        for project in active_scenario.projects:
            bars, global_epoch, has_dates = _extract_milestone_bars(project, global_epoch)
            if not bars:
                continue
            any_has_dates = any_has_dates or has_dates
            row_name = deal_name if len(active_scenario.projects) == 1 else f"{deal_name} / {project.name}"
            raw.append((row_name, bars))

    if not raw:
        return None

    # Wrap into raw_rows format for _override_stabilized_cap
    raw_rows = [{"project_name": name, "bars": bars} for name, bars in raw]
    _override_stabilized_cap(raw_rows)

    all_bars = [b for row in raw_rows for b in row["bars"]]
    g_min = min(b["display_start_day"] for b in all_bars)
    g_max = max(b["display_start_day"] + b["display_duration_days"] for b in all_bars)

    rows: list[dict] = []
    for row in raw_rows:
        _gantt_apply_pct(row["bars"], g_min, g_max)
        rows.append({"type": "project_header", "name": row["project_name"]})
        rows.extend(_bars_to_phase_rows(row["bars"]))

    month_ticks, year_spans = _compute_gantt_axis(global_epoch, g_min, g_max, any_has_dates)
    return {"has_dates": any_has_dates, "month_ticks": month_ticks, "year_spans": year_spans, "rows": rows}


async def _get_user(session: DBSession, request: Request) -> User | None:
    """Resolve the current user from the signed session cookie."""
    from app.api.auth import COOKIE_NAME, decode_session_token

    token = request.cookies.get(COOKIE_NAME)
    if token:
        uid = decode_session_token(token)
        if uid is not None:
            return await session.get(User, uid)

    return None


async def _get_dedup_count(session: DBSession) -> int:
    try:
        result = await session.execute(
            select(func.count())
            .select_from(DedupCandidate)
            .where(DedupCandidate.status == DedupStatus.pending)
        )
        return int(result.scalar_one())
    except Exception:
        return 0


async def _get_address_issues_count(session: AsyncSession) -> int:
    try:
        result = await session.execute(
            select(func.count())
            .select_from(ScrapedListing)
            .where(
                ScrapedListing.realie_skip.is_(True),
                ScrapedListing.realie_enriched_at.is_(None),
                ScrapedListing.apn.is_(None),  # listings with a valid APN don't need address resolution
            )
        )
        return int(result.scalar_one())
    except Exception:
        return 0


_CONFLICT_PCT_TOL = 0.05


def _pct_conflict(a, b) -> bool:
    if a is None or b is None:
        return False
    try:
        fa, fb = float(a), float(b)
        if fb == 0:
            return fa != 0
        return abs(fa - fb) / fb > _CONFLICT_PCT_TOL
    except (TypeError, ValueError):
        return False


def _parcel_map_url(parcel: "Parcel") -> str:
    j = (parcel.jurisdiction or "").lower().strip()
    if j == "portland":
        return "https://www.portlandmaps.com/"
    if j == "gresham":
        return "https://gis.greshamoregon.gov/GreshamMap/"
    county = (parcel.county or "").lower().strip()
    if county == "multnomah":
        return "https://www.portlandmaps.com/"
    return "https://www.clackamas.us/cmap"


def _build_conflicts(opps: list) -> list:
    """Return list of (opp, field_key, opp_val, parcel_val, map_url) tuples."""
    conflicts = []
    for opp in opps:
        p = opp.parcel
        if p is None:
            continue
        ack = opp.parcel_conflicts_ack or {}
        url = _parcel_map_url(p)
        if "units" not in ack and opp.units is not None and p.unit_count is not None and opp.units != p.unit_count:
            conflicts.append((opp, "units", opp.units, p.unit_count, url))
        if "gba_sqft" not in ack and _pct_conflict(opp.gba_sqft, p.building_sqft):
            conflicts.append((opp, "gba_sqft", opp.gba_sqft, p.building_sqft, url))
        if "year_built" not in ack and opp.year_built is not None and p.year_built is not None and opp.year_built != p.year_built:
            conflicts.append((opp, "year_built", opp.year_built, p.year_built, url))
        if "lot_sqft" not in ack and _pct_conflict(opp.lot_sqft, p.lot_sqft):
            conflicts.append((opp, "lot_sqft", opp.lot_sqft, p.lot_sqft, url))
    return conflicts


def _apply_auto_resolutions(opps: list) -> int:
    """Apply auto-resolution rules to ORM objects in place.

    Checks every conflict field on every linked Opportunity.  When a rule
    fires the ORM object is mutated (no commit — caller owns the session)
    and the field is acked so it no longer surfaces in the queue.

    Returns the number of field resolutions applied.
    """
    resolved = 0
    _null_fields = {"units", "gba_sqft", "year_built", "lot_sqft"}
    for opp in opps:
        p = opp.parcel
        if p is None:
            continue
        ack = dict(opp.parcel_conflicts_ack or {})
        changed = False
        for field, (opp_attr, parcel_attr) in _FIELD_MAP.items():
            if field in ack:
                continue
            opp_val = getattr(opp, opp_attr, None)
            par_val = getattr(p, parcel_attr, None)
            if opp_val is None or par_val is None:
                continue
            action = auto_resolve_conflict(field, opp_val, par_val)
            if action is None:
                continue
            if action == "use_parcel" and field in _null_fields:
                setattr(opp, opp_attr, None)
            ack[field] = action
            changed = True
            resolved += 1
        if changed:
            opp.parcel_conflicts_ack = ack
    return resolved


async def _get_conflicts_count(session: AsyncSession) -> int:
    """Count Opportunities with at least one unacknowledged parcel field conflict."""
    try:
        stmt = (
            select(Opportunity)
            .options(selectinload(Opportunity.parcel))
            .where(Opportunity.parcel_id.isnot(None), Opportunity.archived.is_(False))
        )
        opps = list((await session.execute(stmt)).scalars().unique())
        return len(set(c[0].id for c in _build_conflicts(opps)))
    except Exception:
        return 0


async def _get_counts(session: AsyncSession) -> tuple[int, int]:
    return await _get_dedup_count(session), await _get_conflicts_count(session)


def _base_ctx(
    user: User | None,
    dedup_count: int,
    active_nav: str,
    address_issues_count: int = 0,
    conflicts_count: int = 0,
) -> dict:
    initials = "??"
    if user:
        parts = user.name.split()
        initials = (parts[0][0] + parts[-1][0]).upper() if len(parts) >= 2 else user.name[:2].upper()
    return {
        "user_name": user.name if user else "Guest",
        "user_initials": initials,
        "user_color": (user.display_color if user else None) or "#2563EB",
        # Soft email-verification gate: templates show a banner when False.
        # None / missing is treated as verified to avoid false positives.
        "user_email_verified": bool(getattr(user, "email_verified", True)) if user else True,
        "is_org_admin": bool(getattr(user, "is_org_admin", False)) if user else False,
        "is_admin": bool(getattr(user, "is_admin", False)) if user else False,
        "active_nav": active_nav,
        "dedup_count": dedup_count,
        "address_issues_count": address_issues_count,
        "conflicts_count": conflicts_count,
    }


def _require_settings_owner(user: User | None) -> None:
    if not (user and user.is_admin):
        raise HTTPException(status_code=404, detail="Not found")


def _stripe_secret_key() -> str:
    return (settings.stripe_secret_key or "").strip()


def _stripe_is_configured() -> bool:
    return bool(_stripe_secret_key())


def _stripe_state_message(state: str | None) -> str | None:
    if state == "success":
        return "Payment method saved and validated by Stripe."
    if state == "cancel":
        return "Stripe checkout was cancelled before saving a payment method."
    if state == "config-missing":
        return "Billing is not configured yet. Add Stripe keys in environment settings first."
    if state == "error":
        return "Stripe returned an error while starting checkout. Try again in a moment."
    return None


async def _stripe_api_request(
    method: str,
    endpoint: str,
    *,
    data: dict[str, Any] | list[tuple[str, str]] | None = None,
    params: dict[str, Any] | None = None,
) -> dict[str, Any]:
    api_key = _stripe_secret_key()
    if not api_key:
        raise HTTPException(status_code=400, detail="Stripe is not configured")

    url = f"https://api.stripe.com{endpoint}"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Accept": "application/json",
    }
    timeout = httpx.Timeout(20.0)

    try:
        async with httpx.AsyncClient(timeout=timeout, trust_env=False) as client:
            resp = await client.request(method, url, headers=headers, data=data, params=params)
    except Exception as exc:  # pragma: no cover - network exception guard
        raise HTTPException(status_code=502, detail=f"Stripe connection failed: {exc}") from exc

    body: dict[str, Any]
    try:
        body = resp.json()
    except Exception as exc:
        raise HTTPException(status_code=502, detail="Stripe returned a non-JSON response") from exc

    if resp.status_code >= 400:
        err = body.get("error") if isinstance(body, dict) else None
        err_msg = (err or {}).get("message") if isinstance(err, dict) else None
        raise HTTPException(status_code=502, detail=f"Stripe API error: {err_msg or 'unknown error'}")

    return body


async def _get_stripe_customer_id(session: AsyncSession, user_id: UUID) -> str | None:
    row = (
        await session.execute(
            select(UserSetting).where(
                UserSetting.user_id == user_id,
                UserSetting.field_key == "stripe_customer_id",
            )
        )
    ).scalar_one_or_none()
    if row is None:
        return None
    value = (row.value or "").strip()
    return value or None


async def _clear_stripe_customer_id(session: AsyncSession, user_id: UUID) -> None:
    row = (
        await session.execute(
            select(UserSetting).where(
                UserSetting.user_id == user_id,
                UserSetting.field_key == "stripe_customer_id",
            )
        )
    ).scalar_one_or_none()
    if row is None:
        return
    await session.delete(row)
    await session.commit()


async def _ensure_stripe_customer_id(session: AsyncSession, user: User) -> str:
    existing_id = await _get_stripe_customer_id(session, user.id)
    if existing_id:
        return existing_id

    payload: list[tuple[str, str]] = [
        ("name", user.name or "Viciniti Deals user"),
        ("metadata[user_id]", str(user.id)),
        ("metadata[org_id]", str(user.org_id)),
    ]
    if user.email:
        payload.append(("email", user.email))

    customer = await _stripe_api_request("POST", "/v1/customers", data=payload)
    customer_id = str(customer.get("id") or "").strip()
    if not customer_id:
        raise HTTPException(status_code=502, detail="Stripe customer creation failed")

    row = UserSetting(
        id=_uuid_mod.uuid4(),
        user_id=user.id,
        org_id=user.org_id,
        field_key="stripe_customer_id",
        value=customer_id,
    )
    session.add(row)
    await session.commit()
    return customer_id


async def _list_stripe_payment_methods(customer_id: str) -> list[dict[str, str]]:
    body = await _stripe_api_request(
        "GET",
        "/v1/payment_methods",
        params={"customer": customer_id, "type": "card", "limit": 5},
    )

    rows: list[dict[str, str]] = []
    for item in body.get("data", []):
        card = item.get("card") or {}
        brand = str(card.get("brand") or "card").title()
        last4 = str(card.get("last4") or "••••")
        exp_month = str(card.get("exp_month") or "").zfill(2)
        exp_year = str(card.get("exp_year") or "")
        rows.append(
            {
                "brand": brand,
                "last4": last4,
                "exp": f"{exp_month}/{exp_year}" if exp_month and exp_year else "",
            }
        )
    return rows


def _fmt_ts(ts: datetime | None) -> str:
    if ts is None:
        return "Never"
    return ts.astimezone(_PACIFIC).strftime("%Y-%m-%d %H:%M PT")


def _freshness_status(ts: datetime | None, stale_after_hours: int) -> str:
    if ts is None:
        return "No activity"
    age_hours = (datetime.now(UTC) - ts.astimezone(UTC)).total_seconds() / 3600
    return "Healthy" if age_hours <= stale_after_hours else "Stale"


async def _direct_live_ping(url: str, timeout_seconds: float = 8.0) -> bool:
    """Ping a URL directly without inheriting system/env proxy settings."""
    headers = {
        "User-Agent": "VicinitiDeals/1.0",
        "Accept": "application/json, text/html;q=0.9,*/*;q=0.8",
    }
    timeout = httpx.Timeout(timeout_seconds)
    try:
        async with httpx.AsyncClient(timeout=timeout, follow_redirects=True, trust_env=False) as client:
            response = await client.get(url, headers=headers)
            return int(response.status_code) < 500
    except Exception:
        return False


async def _proxyon_remaining_gb(timeout_seconds: float = 8.0) -> str | None:
    """Return residential GB remaining from ProxyOn API if credentials are configured."""
    api_key = (settings.proxyon_api_key or "").strip()
    if not api_key:
        return None

    timeout = httpx.Timeout(timeout_seconds)
    try:
        async with httpx.AsyncClient(timeout=timeout, follow_redirects=True, trust_env=False) as client:
            auth = await client.post(
                "https://api.proxyon.io/v1/auth/token",
                data={"apikey": api_key},
                headers={"Accept": "application/json"},
            )
            auth.raise_for_status()
            auth_body = auth.json()
            if not auth_body.get("success"):
                return None
            token = (auth_body.get("result") or {}).get("token")
            if not token:
                return None

            stats = await client.get(
                "https://api.proxyon.io/v1/residential/stats",
                headers={"X-Session-Token": token, "Accept": "application/json"},
            )
            stats.raise_for_status()
            stats_body = stats.json()
            if not stats_body.get("success"):
                return None
            result = stats_body.get("result") or {}

            for key in ("remaining_gb", "gb_remaining", "left_gb", "remaining", "traffic_left_gb"):
                if key in result and result[key] is not None:
                    return f"{float(result[key]):,.2f} GB"

            for key in ("total_gb", "used_gb"):
                if key in result and result[key] is not None:
                    total = float(result.get("total_gb") or 0)
                    used = float(result.get("used_gb") or 0)
                    if total > 0:
                        return f"{max(total - used, 0):,.2f} GB"
    except Exception:
        return None
    return None


async def _proxyon_residential_snapshot(timeout_seconds: float = 8.0) -> dict[str, Any]:
    """Return cached (hourly) ProxyOn residential connection state and remaining GB."""
    now_monotonic = time.monotonic()
    cached_age = now_monotonic - float(_proxyon_status_cache.get("fetched_monotonic") or 0.0)
    if cached_age < _PROXYON_STATUS_CACHE_TTL_SECONDS:
        return dict(_proxyon_status_cache)

    async with _proxyon_status_lock:
        now_monotonic = time.monotonic()
        cached_age = now_monotonic - float(_proxyon_status_cache.get("fetched_monotonic") or 0.0)
        if cached_age < _PROXYON_STATUS_CACHE_TTL_SECONDS:
            return dict(_proxyon_status_cache)

        checked_at = datetime.now(UTC)
        api_key = (settings.proxyon_api_key or "").strip()
        if not api_key:
            _proxyon_status_cache.update(
                {
                    "fetched_monotonic": now_monotonic,
                    "status_label": "Not Configured",
                    "connected": False,
                    "remaining_gb": None,
                    "expected_days_left": None,
                    "account_balance_usd": None,
                    "active_subscription_id": None,
                    "datacenter_count_live": None,
                    "datacenter_by_country": None,
                    "checked_at": None,
                }
            )
            return dict(_proxyon_status_cache)

        account_balance_usd: str | None = None
        data_left_gb: float | None = None
        expected_days_left: float | None = None
        active_sub_id: int | None = None
        dc_count_live: int | None = None
        dc_by_country: dict[str, int] | None = None
        connected = False
        status_label = "API Key Invalid"
        timeout = httpx.Timeout(timeout_seconds)
        try:
            async with httpx.AsyncClient(timeout=timeout, follow_redirects=True, trust_env=False) as client:
                auth = await client.post(
                    "https://api.proxyon.io/v1/auth/token",
                    data={"apikey": api_key},
                    headers={"Accept": "application/json"},
                )
                auth.raise_for_status()
                auth_body = auth.json()
                token: str | None = None
                if auth_body.get("success"):
                    token = ((auth_body.get("result") or {}).get("token")
                             or (auth_body.get("result") or {}).get("sessionToken"))
                if token:
                    hdrs = {"X-Session-Token": token, "Accept": "application/json"}
                    # Account balance (USD) — same endpoint as before
                    acct = await client.get(
                        "https://api.proxyon.io/v1/account/info", headers=hdrs,
                    )
                    acct.raise_for_status()
                    acct_body = acct.json()
                    if acct_body.get("success"):
                        connected = True
                        balance = (acct_body.get("result") or {}).get("balance")
                        if balance is not None:
                            account_balance_usd = f"${float(balance):,.2f}"

                    # Residential subscription list — this is the authoritative
                    # "are credentials provisioned" signal. Creds come from the
                    # API (user+password returned per subscription), so API-key
                    # presence + an active subscription means we're configured.
                    # Note: /list does NOT return dataLeft; we fetch it via
                    # /residential/{id}/info for the active subscription.
                    subs = await client.get(
                        "https://api.proxyon.io/v1/residential/list", headers=hdrs,
                    )
                    subs.raise_for_status()
                    subs_body = subs.json()
                    sub_list: list = []
                    if subs_body.get("success"):
                        result = subs_body.get("result") or {}
                        if isinstance(result, list):
                            sub_list = result
                        elif isinstance(result, dict):
                            sub_list = (result.get("subscriptions")
                                        or result.get("list")
                                        or [])

                    if sub_list:
                        # Pick the main sub if tagged, else first entry
                        active = next(
                            (s for s in sub_list
                             if isinstance(s, dict) and s.get("isMain")),
                            sub_list[0] if isinstance(sub_list[0], dict) else None,
                        )
                        if isinstance(active, dict):
                            active_sub_id = active.get("id")

                        # Second API call: /info endpoint has the usage data
                        if active_sub_id is not None:
                            try:
                                info = await client.get(
                                    f"https://api.proxyon.io/v1/residential/{active_sub_id}/info",
                                    headers=hdrs,
                                )
                                info.raise_for_status()
                                info_body = info.json()
                                if info_body.get("success"):
                                    info_result = info_body.get("result") or {}
                                    # ProxyOn returns dataLeft in megabytes
                                    # (confirmed against live account with
                                    # expectedDaysLeft cross-check). Convert to GB.
                                    _left_mb = info_result.get("dataLeft")
                                    if _left_mb is not None:
                                        try:
                                            data_left_gb = float(_left_mb) / 1024.0
                                        except (TypeError, ValueError):
                                            data_left_gb = None
                                    _days = info_result.get("expectedDaysLeft")
                                    if _days is not None:
                                        try:
                                            expected_days_left = float(_days)
                                        except (TypeError, ValueError):
                                            expected_days_left = None
                            except Exception:
                                pass

                        status_label = (
                            "Active" if data_left_gb and data_left_gb > 0
                            else "Configured (No Data Left)" if data_left_gb == 0
                            else "Configured"  # has sub, data unknown (info call failed)
                        )
                    else:
                        status_label = "Configured (No Subscription)"

                    # Datacenter proxies — live count from /datacenter/list.
                    # Env-var PROXYON_DATACENTER_PROXIES is a comma-separated
                    # list of pre-wired connection strings the scraper uses at
                    # runtime; the live list is the authoritative inventory.
                    try:
                        dc = await client.get(
                            "https://api.proxyon.io/v1/datacenter/list", headers=hdrs,
                        )
                        dc.raise_for_status()
                        dc_body = dc.json()
                        if dc_body.get("success"):
                            dc_result = dc_body.get("result") or {}
                            if isinstance(dc_result, dict):
                                proxies = dc_result.get("proxies") or dc_result.get("list") or []
                            elif isinstance(dc_result, list):
                                proxies = dc_result
                            else:
                                proxies = []
                            active = [p for p in proxies
                                      if isinstance(p, dict)
                                      and (p.get("status") or "").lower() == "active"]
                            dc_count_live = len(active)
                            by_country: dict[str, int] = {}
                            for p in active:
                                cc = (p.get("country") or "??").lower()
                                by_country[cc] = by_country.get(cc, 0) + 1
                            dc_by_country = by_country or None
                    except Exception:
                        pass
        except Exception:
            connected = False
            status_label = "API Error"

        _proxyon_status_cache.update(
            {
                "fetched_monotonic": now_monotonic,
                "status_label": status_label,
                "connected": connected,
                "remaining_gb": (
                    f"{data_left_gb:.2f} GB" if data_left_gb is not None else None
                ),
                "expected_days_left": expected_days_left,
                "account_balance_usd": account_balance_usd,
                "active_subscription_id": active_sub_id,
                "datacenter_count_live": dc_count_live,
                "datacenter_by_country": dc_by_country,
                "checked_at": checked_at,
            }
        )
        return dict(_proxyon_status_cache)


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------


@router.get("/", response_class=HTMLResponse)
async def root() -> RedirectResponse:
    return RedirectResponse(url="/deals")


@router.get("/splash", response_class=HTMLResponse)
async def splash(request: Request, session: DBSession) -> HTMLResponse:
    users = list((await session.execute(select(User).order_by(User.name))).scalars())
    return templates.TemplateResponse(request, "splash.html", {"users": users})



@router.get("/settings/scraping-services", response_class=HTMLResponse)
async def settings_scraping_services(
    request: Request,
    session: DBSession,
) -> HTMLResponse:
    user = await _get_user(session, request)
    _require_settings_owner(user)
    dedup_count, conflicts_count = await _get_counts(session)
    address_issues_count = await _get_address_issues_count(session)

    loopnet_job = (await session.execute(
        select(IngestJob)
        .where(IngestJob.source == "loopnet")
        .order_by(IngestJob.started_at.desc())
        .limit(1)
    )).scalar_one_or_none()
    crexi_job = (await session.execute(
        select(IngestJob)
        .where(IngestJob.source == "crexi")
        .order_by(IngestJob.started_at.desc())
        .limit(1)
    )).scalar_one_or_none()

    residential_username = (settings.proxyon_residential_username or "").strip()
    residential_password = (settings.proxyon_residential_password or "").strip()
    residential_env_creds = bool(residential_username and residential_password)
    datacenter_env_count = len([p for p in (settings.proxyon_datacenter_proxies or "").split(",") if p.strip()])
    proxyon_snapshot = await _proxyon_residential_snapshot()

    # Authoritative status = live API state. Residential creds are provisioned
    # via the ProxyOn API (GET /residential/list returns user+password per
    # subscription), so API-key presence + an active subscription means we
    # have usable credentials regardless of what's in the env vars.
    residential_status = proxyon_snapshot.get("status_label") or "Not Configured"
    residential_gb_remaining = proxyon_snapshot.get("remaining_gb")
    residential_balance = proxyon_snapshot.get("account_balance_usd")
    residential_sub_id = proxyon_snapshot.get("active_subscription_id")
    residential_days_left = proxyon_snapshot.get("expected_days_left")
    datacenter_count_live = proxyon_snapshot.get("datacenter_count_live")
    datacenter_by_country = proxyon_snapshot.get("datacenter_by_country") or {}
    # Prefer live count from API; fall back to env-var count if API unavailable
    datacenter_count = (
        datacenter_count_live
        if datacenter_count_live is not None
        else datacenter_env_count
    )
    # A residential subscription exists (live signal from API, not env-var peek)
    residential_configured = bool(residential_sub_id) or residential_env_creds
    _checked_at = proxyon_snapshot.get("checked_at")
    residential_last_checked = _fmt_ts(_checked_at) if _checked_at else "API key not configured"

    loopnet_lease_job = (await session.execute(
        select(IngestJob)
        .where(IngestJob.source == "loopnet_lease")
        .order_by(IngestJob.started_at.desc())
        .limit(1)
    )).scalar_one_or_none()

    rapidapi_configured = bool((settings.rapidapi_key or "").strip())
    loopnet_status = (
        "Not Configured" if not rapidapi_configured
        else _freshness_status(
            loopnet_job.started_at if loopnet_job else None,
            stale_after_hours=180,  # weekly cadence + 12h grace
        )
    )
    loopnet_lease_status: str
    if not rapidapi_configured:
        loopnet_lease_status = "Not Configured"
    elif loopnet_lease_job is None:
        loopnet_lease_status = "Never Run"
    elif loopnet_lease_job.status == "completed":
        loopnet_lease_status = "Completed"
    else:
        loopnet_lease_status = (loopnet_lease_job.status or "").title() or "Unknown"

    services = [
        {
            "name": "LoopNet Ingest",
            "description": "Weekly LoopNet scrape of sale listings via RapidAPI, with bulk-triage and polygon-tiered categorization (MF + Land + Mixed-Use in target tier, MF-only in comp tier).",
            "status": loopnet_status,
            "schedule": "Weekly Monday at 07:00 UTC via Celery beat",
            "proxy": "Direct (RapidAPI, configured)" if rapidapi_configured else "Direct (RapidAPI, not configured)",
            "last_run": _fmt_ts(loopnet_job.started_at if loopnet_job else None),
            "last_result": loopnet_job.status if loopnet_job else "never",
        },
        {
            "name": "LoopNet Lease Seed",
            "description": "Manual one-off: scrape MF + mixed-use lease listings for income-side comp library.",
            "status": loopnet_lease_status,
            "schedule": "Manual (on-demand)",
            "proxy": "Direct (RapidAPI, configured)" if rapidapi_configured else "Direct (RapidAPI, not configured)",
            "last_run": _fmt_ts(loopnet_lease_job.started_at if loopnet_lease_job else None),
            "last_result": loopnet_lease_job.status if loopnet_lease_job else "never",
        },
        {
            "name": "Crexi Ingest",
            "description": "Daily Crexi crawler run for refreshed multifamily listing coverage.",
            "status": _freshness_status(crexi_job.started_at if crexi_job else None, stale_after_hours=30),
            "schedule": "Daily at 06:00 PT via Celery beat",
            "proxy": "Residential (ProxyOn)" if residential_configured else "Residential (ProxyOn, not configured)",
            "last_run": _fmt_ts(crexi_job.started_at if crexi_job else None),
            "last_result": crexi_job.status if crexi_job else "never",
        },
        {
            "name": "Oregon eLicense",
            "description": "Monthly enrichment of broker license records (license type, status, personal address, affiliated firm, disciplinary actions) from the Oregon Real Estate Agency public lookup.",
            "status": "Pending Implementation",
            "schedule": "Monthly on the 2nd at 05:00 UTC via Celery beat",
            "proxy": "Residential (ProxyOn)" if residential_configured else "Residential (ProxyOn, not configured)",
            "last_run": "—",
            "last_result": "—",
            "action_url": "/scraper/oregon-elicense/run",
            "action_label": "Trigger Sweep",
            "action_method": "post",
        },
    ]

    return templates.TemplateResponse(
        request,
        "settings_scraping_services.html",
        {
            "services": services,
            "residential_status": residential_status,
            "datacenter_count": datacenter_count,
            "datacenter_count_live": datacenter_count_live,
            "datacenter_env_count": datacenter_env_count,
            "datacenter_by_country": datacenter_by_country,
            "residential_gb_remaining": residential_gb_remaining,
            "residential_balance": residential_balance,
            "residential_sub_id": residential_sub_id,
            "residential_days_left": residential_days_left,
            "residential_last_checked": residential_last_checked,
            **_base_ctx(user, dedup_count, "", address_issues_count, conflicts_count=conflicts_count),
        },
    )


@router.get("/settings/data-sources", response_class=HTMLResponse)
async def settings_data_sources(
    request: Request,
    session: DBSession,
) -> HTMLResponse:
    user = await _get_user(session, request)
    _require_settings_owner(user)
    dedup_count, conflicts_count = await _get_counts(session)
    address_issues_count = await _get_address_issues_count(session)

    parcel_count = int((await session.execute(select(func.count()).select_from(Parcel))).scalar_one())

    # cache_file: path relative to /app/data/gis_cache/  (empty = not cached as file)
    # ping_host:  hostname key used to share one liveness check across all layers on that host
    def _layer(name: str, slug: str, type_: str, status: str, notes: str,
               cache_file: str = "", ping_host: str = "") -> dict:
        return {"name": name, "slug": slug, "type": type_, "status": status,
                "notes": notes, "cache_file": cache_file, "ping_host": ping_host}

    groups = [
        {
            "id": "parcel_seeding", "title": "Parcel Seeding", "provider": "",
            "layers": [
                _layer("Metro RLIS Taxlots", "tax_lots_metro_rlis", "FeatureServer", "Active",
                       f"Primary seed — ~430k features (Multnomah + Clackamas). {parcel_count:,} parcels in DB. "
                       "Polygon geometry + assessed values. Owner name stripped in public layer. Monthly refresh by Metro.",
                       cache_file="oregon/tax_lots_metro_rlis.geojson", ping_host="services2.arcgis.com"),
                _layer("Oregon Address Points", "address_points_or", "FeatureServer", "Cached",
                       "462,110 features downloaded. PARCEL_ID 0% populated by Portland/Clackamas 911 agencies — "
                       "cannot be used for parcel seeding. Retained for potential address enrichment.",
                       cache_file="oregon/address_points_or.geojson", ping_host="services8.arcgis.com"),
            ],
        },
        {
            "id": "boundary_routing", "title": "Boundary & Routing", "provider": "",
            "layers": [
                _layer("City Limits (Oregon)", "city_limits_or", "MapServer", "Active",
                       "ODOT source. Point-in-polygon routing for listings without a clean jurisdiction.",
                       cache_file="oregon/city_limits_or.geojson", ping_host="gis.odot.state.or.us"),
                _layer("County Boundaries (Oregon)", "county_boundaries_or", "FeatureServer", "Active",
                       "BLM source. County routing fallback for unincorporated parcels.",
                       cache_file="oregon/county_boundaries_or.geojson", ping_host="services1.arcgis.com"),
                _layer("Urban Growth Boundaries (Oregon)", "urban_growth_boundaries_or", "FeatureServer", "Active",
                       "DLCD source. Out-of-market screening gate — parcels outside UGB are Out of Market.",
                       cache_file="oregon/urban_growth_boundaries_or.geojson", ping_host="services8.arcgis.com"),
            ],
        },
        {
            "id": "incentive_screening", "title": "Incentive Screening", "provider": "",
            "layers": [
                _layer("Enterprise Zones (Oregon)", "enterprise_zones_or", "FeatureServer", "Active",
                       "Oregon Business Development Dept. Statewide enterprise zone polygons.",
                       cache_file="oregon/enterprise_zones_or.geojson", ping_host="services8.arcgis.com"),
                _layer("Opportunity Zones (Oregon)", "opportunity_zones_or", "FeatureServer", "Active",
                       "Filter: STATE='41'. Federal Opportunity Zone census tracts.",
                       cache_file="external/opportunity_zones_or.geojson", ping_host="services.arcgis.com"),
                _layer("NMTC Qualified Tracts", "nmtc_qualified_tracts_or", "FeatureServer", "Active",
                       "Filter: STATE_FIPS='41'. New Markets Tax Credit qualified census tracts.",
                       cache_file="external/nmtc_qualified_tracts_or.geojson", ping_host="services6.arcgis.com"),
            ],
        },
        {
            "id": "environmental", "title": "Environmental", "provider": "Oregon GEO (services8.arcgis.com)",
            "layers": [
                _layer("Wetlands — LWI", "wetlands_lwi_or", "FeatureServer", "Active",
                       "Oregon Local Wetland Inventory. Additive evidence family with NWI + MORE.",
                       cache_file="oregon/wetlands_lwi_or.geojson", ping_host="services8.arcgis.com"),
                _layer("Wetlands — NWI", "wetlands_nwi_or", "FeatureServer", "Active",
                       "USFWS National Wetland Inventory. Combined with LWI + MORE for better coverage.",
                       cache_file="oregon/wetlands_nwi_or.geojson", ping_host="services8.arcgis.com"),
                _layer("Wetlands — MORE Oregon", "wetlands_more_or", "FeatureServer", "Active",
                       "More Oregon Wetlands dataset. Third additive layer in the family.",
                       cache_file="oregon/wetlands_more_or.geojson", ping_host="services8.arcgis.com"),
            ],
        },
        {
            "id": "street_classifications", "title": "Street Classifications", "provider": "",
            "layers": [
                _layer("ODOT State Roads", "street_functional_class_state_or", "MapServer", "Active",
                       "Federal functional class for ODOT-owned roads statewide. NEW_FC_TYP: Interstate / Freeway / Arterial / Collector / Local.",
                       ping_host="gis.odot.state.or.us"),
                _layer("ODOT Non-State Roads", "street_functional_class_nonstate_or", "MapServer", "Active",
                       "Federal functional class for county/city/other roads. Covers all of Multnomah + Clackamas combined with State layer.",
                       ping_host="gis.odot.state.or.us"),
            ],
        },
        {
            "id": "reference", "title": "Reference Layers", "provider": "Oregon GEO (services8.arcgis.com)",
            "layers": [
                _layer("Building Footprints (Oregon)", "building_footprints_or", "FeatureServer", "Active",
                       "Structural screening — confirms existing building presence and approximate footprint area.",
                       cache_file="oregon/building_footprints_or.geojson", ping_host="services8.arcgis.com"),
                _layer("Oregon ZIP Reference", "oregon_zip_reference", "FeatureServer", "Active",
                       "ZIP code polygon reference for address routing.",
                       cache_file="oregon/oregon_zip_reference.geojson", ping_host="services8.arcgis.com"),
                _layer("Census Block Groups 2020", "census_block_groups_2020_or", "FeatureServer", "Active",
                       "Demographic context for NMTC / Opportunity Zone joins.",
                       cache_file="oregon/census_block_groups_2020_or.geojson", ping_host="services8.arcgis.com"),
                _layer("Census Tracts 2020", "census_tracts_2020_or", "FeatureServer", "Active",
                       "Demographic context for NMTC / Opportunity Zone joins.",
                       cache_file="oregon/census_tracts_2020_or.geojson", ping_host="services8.arcgis.com"),
            ],
        },
        {
            "id": "local_fairview", "title": "Local GIS — Fairview",
            "provider": "services5.arcgis.com · Fairview ArcGIS Online (Org: 3DoY8p7EnUTzaIE7)",
            "layers": [
                _layer("Natural Resource Protection Areas", "natural_resources_fairview", "FeatureServer", "Active",
                       "TYPE field: riparian buffers (35'/40'/55'/80'), Fairview Lake 50' buffer, platted protected areas, upland habitat, wetlands.",
                       cache_file="fairview/natural_resources_fairview.geojson", ping_host="services5.arcgis.com"),
                _layer("Fairview Lake 35ft Buffer", "fairview_lake_35ft_buffer", "FeatureServer", "Active",
                       "Additive to natural resource layer.",
                       cache_file="fairview/fairview_lake_35ft_buffer.geojson", ping_host="services5.arcgis.com"),
                _layer("Fairview Lake 50ft Buffer", "fairview_lake_50ft_buffer", "FeatureServer", "Active",
                       "Additive to natural resource layer.",
                       cache_file="fairview/fairview_lake_50ft_buffer.geojson", ping_host="services5.arcgis.com"),
                _layer("Enterprise Zone", "enterprise_zone_fairview", "FeatureServer", "Active",
                       "Columbia Cascade Enterprise Zone, ~34 parcels. Supplement to statewide EZ layer.",
                       cache_file="fairview/enterprise_zone_fairview.geojson", ping_host="services5.arcgis.com"),
                _layer("Overlay Districts", "overlay_districts_fairview", "FeatureServer", "Active",
                       "Airport Overlay, Storefront District (TCC), Four Corners Area (VMU), R/SFLD.",
                       cache_file="fairview/overlay_districts_fairview.geojson", ping_host="services5.arcgis.com"),
                _layer("Street Jurisdiction Routing", "streets_jurisdiction_fairview", "FeatureServer", "Active",
                       "OWNER field: City of Fairview / Gresham / Multnomah County / ODOT / Private.",
                       cache_file="fairview/streets_jurisdiction_fairview.geojson", ping_host="services5.arcgis.com"),
                _layer("Zoning", "—", "PDF / Manual", "Manual",
                       "No queryable GIS layer. zoning_lookup_url set on parcels at seed time. Zone Painter used for manual zoning_code assignment."),
            ],
        },
        {
            "id": "local_gresham", "title": "Local GIS — Gresham",
            "provider": "gis.greshamoregon.gov · Gresham MapServer",
            "layers": [
                _layer("East County Taxlots (RLIS+)", "tax_lots_east_county", "MapServer", "Active",
                       "Full RLIS dataset with ZONE + owner fields intact. Covers Portland, Troutdale, Fairview, Wood Village, unincorporated Multnomah — not Gresham city parcels.",
                       cache_file="gresham/tax_lots_east_county.geojson", ping_host="gis.greshamoregon.gov"),
                _layer("City Limits", "city_limits", "MapServer", "Active",
                       "Gresham jurisdiction boundary.",
                       cache_file="gresham/city_limits.geojson", ping_host="gis.greshamoregon.gov"),
                _layer("Neighborhoods", "neighborhoods", "MapServer", "Active",
                       "Neighborhood routing layer.",
                       cache_file="gresham/neighborhoods.geojson", ping_host="gis.greshamoregon.gov"),
                _layer("Addresses", "addresses_all", "MapServer", "Active",
                       "Address-level routing.",
                       cache_file="gresham/addresses_all.geojson", ping_host="gis.greshamoregon.gov"),
                _layer("Multifamily Housing Inventory", "multifamily_housing", "MapServer", "Active",
                       "Existing MF housing stock — used for comparables context.",
                       cache_file="gresham/multifamily_housing.geojson", ping_host="gis.greshamoregon.gov"),
                _layer("Planning Overlays", "—", "MapServer", "Active",
                       "Pleasant Valley, Kelley Creek Headwaters, Springwater plan areas; Rockwood Plan District; Design Districts.",
                       cache_file="gresham/pleasant_valley_plan_area.geojson", ping_host="gis.greshamoregon.gov"),
                _layer("Street Classifications", "street_classifications", "MapServer", "Active",
                       "Local Gresham Planning dept street designations.",
                       cache_file="gresham/street_classifications.geojson", ping_host="gis.greshamoregon.gov"),
                _layer("Environmental Overlays", "—", "MapServer", "Active",
                       "Streams, other waters, environmental overlay districts.",
                       cache_file="gresham/streams.geojson", ping_host="gis.greshamoregon.gov"),
                _layer("Transit Layers", "—", "MapServer", "Active",
                       "Bike routes, MAX stops, bus stops, bus lines.",
                       cache_file="gresham/bus_stops.geojson", ping_host="gis.greshamoregon.gov"),
                _layer("Incentive Zones (6 layers)", "—", "MapServer", "Active",
                       "Incentive eligibility overlays from Gresham Incentives MapServer.",
                       cache_file="gresham/rockwood_urban_renewal_area.geojson", ping_host="gis.greshamoregon.gov"),
            ],
        },
        {
            "id": "local_wood_village", "title": "Local GIS — Wood Village",
            "provider": "services7.arcgis.com · City of Wood Village ArcGIS Online (Org: 5Loh3xXKWLd2M7xA)",
            "layers": [
                _layer("Zoning", "zoning_wood_village", "FeatureServer", "Active",
                       "Labeling (zone code), Name (description). Supports advanced queries.",
                       cache_file="wood_village/zoning_wood_village.geojson", ping_host="services7.arcgis.com"),
                _layer("Taxlots (RLIS-compatible)", "taxlots_wood_village", "FeatureServer", "Active",
                       "RLIS-compatible fields: TLID, LANDVAL, ASSESSVAL, LANDUSE, STATECLASS, YEARBUILT, BLDGSQFT, SITEADDR.",
                       cache_file="wood_village/taxlots_wood_village.geojson", ping_host="services7.arcgis.com"),
                _layer("City Limits", "city_limits_wood_village", "FeatureServer", "Active",
                       "Jurisdiction boundary.",
                       cache_file="wood_village/city_limits_wood_village.geojson", ping_host="services7.arcgis.com"),
            ],
        },
        {
            "id": "local_troutdale", "title": "Local GIS — Troutdale",
            "provider": "maps.troutdaleoregon.gov · Self-hosted ArcGIS Enterprise",
            "layers": [
                _layer("Zoning (Urban Planning Area)", "zoning_troutdale", "MapServer", "Active",
                       "ZONE field (R10, GI, etc.). supportsAdvancedQueries: true.",
                       ping_host="maps.troutdaleoregon.gov"),
                _layer("Street Centerlines", "streets_troutdale", "MapServer", "Active",
                       "CLASS (designation), OWNER, CONDTN fields.",
                       ping_host="maps.troutdaleoregon.gov"),
            ],
        },
        # ── Clackamas County cities ───────────────────────────────────────
        {
            "id": "local_happy_valley", "title": "Local GIS — Happy Valley",
            "provider": "services5.arcgis.com · ArcGIS Online (Org: fuVQ9NIPGnPhCBXp)",
            "layers": [
                _layer("Zoning", "zoning_happy_valley", "FeatureServer", "Active",
                       "Authoritative 2024 zoning. Fields: ZONE (e.g. R-1, C-1), ZOVER (overlay), ORDINANCE, DATE_.",
                       cache_file="happy_valley/zoning_happy_valley.geojson", ping_host="services5.arcgis.com"),
                _layer("City Limits", "city_limits_happy_valley", "FeatureServer", "Active",
                       "Jurisdiction boundary.",
                       cache_file="happy_valley/city_limits_happy_valley.geojson", ping_host="services5.arcgis.com"),
                _layer("Natural Resource Overlay", "natural_resources_happy_valley", "FeatureServer", "Active",
                       "Natural resource protection overlay zones.",
                       cache_file="happy_valley/natural_resources_happy_valley.geojson", ping_host="services5.arcgis.com"),
                _layer("FEMA Floodplain", "fema_floodplain_happy_valley", "FeatureServer", "Active",
                       "FEMA flood hazard zones.",
                       cache_file="happy_valley/fema_floodplain_happy_valley.geojson", ping_host="services5.arcgis.com"),
            ],
        },
        {
            "id": "local_milwaukie", "title": "Local GIS — Milwaukie",
            "provider": "services6.arcgis.com · ArcGIS Online (Org: 8e6aYcxt8yhvXvO9)",
            "layers": [
                _layer("Zoning", "zoning_milwaukie", "FeatureServer", "Active",
                       "COM_Zoning_SDE layer 11. Field: ZONE (MUTSA, BI, GMU, C-CS, DMU, C-G, NMU, SMU, OS, M, R-MD, R-HD).",
                       cache_file="milwaukie/zoning_milwaukie.geojson", ping_host="services6.arcgis.com"),
                _layer("City Limits", "city_limits_milwaukie", "FeatureServer", "Active",
                       "Jurisdiction boundary.",
                       cache_file="milwaukie/city_limits_milwaukie.geojson", ping_host="services6.arcgis.com"),
                _layer("Wetlands", "wetlands_milwaukie", "FeatureServer", "Active",
                       "Local wetland inventory. Same service also has vegetated corridors (6), habitat conservation areas (7), Willamette Greenway (8).",
                       cache_file="milwaukie/wetlands_milwaukie.geojson", ping_host="services6.arcgis.com"),
                _layer("FEMA Floodplain", "floodplain_milwaukie", "FeatureServer", "Active",
                       "FEMA flood hazard zones (COM_FEMA_Hazards service).",
                       cache_file="milwaukie/floodplain_milwaukie.geojson", ping_host="services6.arcgis.com"),
                _layer("Urban Renewal Area", "urban_renewal_milwaukie", "FeatureServer", "Active",
                       "Urban renewal district boundary (COM_URA service).",
                       cache_file="milwaukie/urban_renewal_milwaukie.geojson", ping_host="services6.arcgis.com"),
            ],
        },
        {
            "id": "local_oregon_city", "title": "Local GIS — Oregon City",
            "provider": "maps.orcity.org · Self-hosted ArcGIS Enterprise (v11.5)",
            "layers": [
                _layer("Zoning", "zoning_oregon_city", "MapServer", "Active",
                       "LandUseAndPlanning_PUBLIC layer 62. Same service: comp plan (57), enterprise zones (3, 85), opportunity zones (73), urban renewal (33), historic districts (31-32).",
                       cache_file="oregon_city/zoning_oregon_city.geojson", ping_host="maps.orcity.org"),
                _layer("City Limits", "city_limits_oregon_city", "MapServer", "Active",
                       "City boundary and annexation history.",
                       cache_file="oregon_city/city_limits_oregon_city.geojson", ping_host="maps.orcity.org"),
                _layer("Taxlots", "taxlots_oregon_city", "MapServer", "Active",
                       "Taxlot polygons. Max 50k records, min scale 1:20,000.",
                       cache_file="oregon_city/taxlots_oregon_city.geojson", ping_host="maps.orcity.org"),
                _layer("Hazards & Flood", "hazards_flood_oregon_city", "MapServer", "Active",
                       "100yr/500yr floodplain, floodway, landslides, geologic hazards, slope categories, riparian buffer zone.",
                       cache_file="oregon_city/hazards_flood_oregon_city.geojson", ping_host="maps.orcity.org"),
                _layer("Urban Renewal District", "urban_renewal_oregon_city", "MapServer", "Active",
                       "Urban renewal district boundary.",
                       cache_file="oregon_city/urban_renewal_oregon_city.geojson", ping_host="maps.orcity.org"),
                _layer("Enterprise Zones", "enterprise_zone_oregon_city", "MapServer", "Active",
                       "Enterprise zone polygons.",
                       cache_file="oregon_city/enterprise_zone_oregon_city.geojson", ping_host="maps.orcity.org"),
            ],
        },
        {
            "id": "local_gladstone", "title": "Local GIS — Gladstone",
            "provider": "maps.orcity.org · Hosted on Oregon City ArcGIS Enterprise",
            "layers": [
                _layer("Zoning", "zoning_gladstone", "MapServer", "Active",
                       "Gladstone_LandUseAndPlanning layer 7. Same service: comp plan (6), urban renewal (5), multifamily housing (3), vacant lands (2).",
                       cache_file="gladstone/zoning_gladstone.geojson", ping_host="maps.orcity.org"),
                _layer("City Limits", "city_limits_gladstone", "MapServer", "Active",
                       "Jurisdiction boundary.",
                       cache_file="gladstone/city_limits_gladstone.geojson", ping_host="maps.orcity.org"),
                _layer("Hazards & Flood", "hazards_flood_gladstone", "MapServer", "Active",
                       "FEMA floodplain, landslide, and geologic hazard layers.",
                       cache_file="gladstone/hazards_flood_gladstone.geojson", ping_host="maps.orcity.org"),
                _layer("Natural Resources", "natural_resources_gladstone", "MapServer", "Active",
                       "Streams and natural resource areas.",
                       cache_file="gladstone/natural_resources_gladstone.geojson", ping_host="maps.orcity.org"),
                _layer("Multifamily Housing", "multifamily_housing_gladstone", "MapServer", "Active",
                       "Existing MF housing stock inventory.",
                       cache_file="gladstone/multifamily_housing_gladstone.geojson", ping_host="maps.orcity.org"),
            ],
        },
        {
            "id": "local_lake_oswego", "title": "Local GIS — Lake Oswego",
            "provider": "maps.ci.oswego.or.us · Self-hosted ArcGIS Enterprise (v12)",
            "layers": [
                _layer("Zoning", "zoning_lake_oswego", "MapServer", "Active",
                       "Layers_Geocortex layer 68. Also: comp plan (69), design districts (58), neighborhood overlays (60), Willamette River Greenway mgmt district (62).",
                       cache_file="lake_oswego/zoning_lake_oswego.geojson", ping_host="maps.ci.oswego.or.us"),
                _layer("City Limits", "city_limits_lake_oswego", "MapServer", "Active",
                       "City boundary (Layers_Geocortex layer 1).",
                       cache_file="lake_oswego/city_limits_lake_oswego.geojson", ping_host="maps.ci.oswego.or.us"),
                _layer("Sensitive Lands", "sensitive_lands_lake_oswego", "MapServer", "Active",
                       "Layer 57 = Sensitive Lands polygons. Also: streams (55), wetland (200), 50ft riparian protection area (308).",
                       cache_file="lake_oswego/sensitive_lands_lake_oswego.geojson", ping_host="maps.ci.oswego.or.us"),
                _layer("FEMA Flood / Hazards", "fema_flood_lake_oswego", "MapServer", "Active",
                       "FEMA (17), 1996 flood level (18), soils (19), fault (20), shallow/deep landslide susceptibility (22-23).",
                       cache_file="lake_oswego/fema_flood_lake_oswego.geojson", ping_host="maps.ci.oswego.or.us"),
                _layer("Urban Renewal Districts", "urban_renewal_lake_oswego", "MapServer", "Active",
                       "East End URA (layer 10) and Lake Grove URA (layer 11).",
                       cache_file="lake_oswego/urban_renewal_lake_oswego.geojson", ping_host="maps.ci.oswego.or.us"),
            ],
        },
        {
            "id": "local_west_linn", "title": "Local GIS — West Linn",
            "provider": "geo.westlinnoregon.gov · Self-hosted ArcGIS Enterprise (v10.9)",
            "layers": [
                _layer("Zoning", "zoning_west_linn", "MapServer", "Active",
                       "ZoningComPlan layer 8 + comp plan (10). Max 2,000 records.",
                       cache_file="west_linn/zoning_west_linn.geojson", ping_host="geo.westlinnoregon.gov"),
                _layer("Wetland Inventory", "wetlands_west_linn", "MapServer", "Active",
                       "WetlandInventory MapServer layers 0-1.",
                       cache_file="west_linn/wetlands_west_linn.geojson", ping_host="geo.westlinnoregon.gov"),
                _layer("FEMA Flood Hazard", "fema_flood_west_linn", "MapServer", "Active",
                       "FEMA Flood Hazard Zones (2020), layer 1.",
                       cache_file="west_linn/fema_flood_west_linn.geojson", ping_host="geo.westlinnoregon.gov"),
                _layer("Habitat Conservation Area", "habitat_conservation_west_linn", "MapServer", "Active",
                       "Verified HCA polygons.",
                       cache_file="west_linn/habitat_conservation_west_linn.geojson", ping_host="geo.westlinnoregon.gov"),
                _layer("Riparian Corridor", "riparian_corridor_west_linn", "MapServer", "Active",
                       "Riparian Corridor Inventory polygons.",
                       cache_file="west_linn/riparian_corridor_west_linn.geojson", ping_host="geo.westlinnoregon.gov"),
                _layer("Regulatory Overlays", "regulatory_zones_west_linn", "MapServer", "Active",
                       "Willamette Falls Drive Commercial Design District, Willamette Historic District (local + National Register).",
                       cache_file="west_linn/regulatory_zones_west_linn.geojson", ping_host="geo.westlinnoregon.gov"),
            ],
        },
        {
            "id": "local_tualatin", "title": "Local GIS — Tualatin",
            "provider": "tualgis.ci.tualatin.or.us · Self-hosted ArcGIS Enterprise (v10.91)",
            "layers": [
                _layer("Zoning / Planning Districts", "zoning_tualatin", "MapServer", "Active",
                       "LandusePlanningExplorer layers 6-7. Zone code field: PLANDIST.CZONE (e.g. CO, RH, IN). Max 1,000 records.",
                       cache_file="tualatin/zoning_tualatin.geojson", ping_host="tualgis.ci.tualatin.or.us"),
                _layer("City Limits", "city_limits_tualatin", "MapServer", "Active",
                       "TualatinBoundaries layer 0.",
                       cache_file="tualatin/city_limits_tualatin.geojson", ping_host="tualgis.ci.tualatin.or.us"),
                _layer("Environmental Overlays", "environmental_tualatin", "MapServer", "Active",
                       "EnvironmentalExplorer: wetlands (24), 100yr floodplain (9), floodway (11), natural resources protection overlay (23), 50ft stream buffer (26), slope ≥25% (3).",
                       cache_file="tualatin/environmental_tualatin.geojson", ping_host="tualgis.ci.tualatin.or.us"),
                _layer("Urban Renewal Areas", "urban_renewal_tualatin", "MapServer", "Active",
                       "Core Opportunity and Reinvestment Area, Leveton TID, SW & Basalt Creek URAs.",
                       cache_file="tualatin/urban_renewal_tualatin.geojson", ping_host="tualgis.ci.tualatin.or.us"),
            ],
        },
        {
            "id": "local_wilsonville", "title": "Local GIS — Wilsonville",
            "provider": "gis.wilsonvillemaps.com · Self-hosted ArcGIS Enterprise (v11.5)",
            "layers": [
                _layer("Zoning", "zoning_wilsonville", "FeatureServer", "Active",
                       "Map___WilsonvilleMaps_MIL1 layer 40. ZONE_CODE field (OTR, PDC, PDI, R, V, Future Development).",
                       cache_file="wilsonville/zoning_wilsonville.geojson", ping_host="gis.wilsonvillemaps.com"),
                _layer("City Limits", "city_limits_wilsonville", "FeatureServer", "Active",
                       "Layer 2. Also: UGB (0), county boundary (1).",
                       cache_file="wilsonville/city_limits_wilsonville.geojson", ping_host="gis.wilsonvillemaps.com"),
                _layer("Taxlots", "taxlots_wilsonville", "FeatureServer", "Active",
                       "County assessor taxlots covering Clackamas + Washington County portions (layer 11).",
                       cache_file="wilsonville/taxlots_wilsonville.geojson", ping_host="gis.wilsonvillemaps.com"),
                _layer("Natural Resources / Environmental", "environmental_wilsonville", "FeatureServer", "Active",
                       "Map___NaturalResources: significant wetlands (1099), non-significant (1090), upland habitat (1080), FEMA 100yr floodplain (1107), 1996 flood inundation (1030), streams (1050-1060).",
                       cache_file="wilsonville/environmental_wilsonville.geojson", ping_host="gis.wilsonvillemaps.com"),
                _layer("SROZ — Significant Resource Overlay Zone", "sroz_wilsonville", "FeatureServer", "Active",
                       "Primary environmental overlay. Layer 60 = SROZ polygon, layer 70 = SROZ Impact Area.",
                       cache_file="wilsonville/sroz_wilsonville.geojson", ping_host="gis.wilsonvillemaps.com"),
                _layer("Urban Renewal Areas", "urban_renewal_wilsonville", "MapServer", "Active",
                       "Map___URA: URA_Coffee, URA_East, URA_TWIST, URA_West, URA_WIN.",
                       cache_file="wilsonville/urban_renewal_wilsonville.geojson", ping_host="gis.wilsonvillemaps.com"),
            ],
        },
        {
            "id": "local_canby_johnson_city", "title": "Local GIS — Canby / Johnson City",
            "provider": "DLCD statewide fallback — no public REST service",
            "layers": [
                _layer("Canby Zoning", "—", "DLCD Fallback", "Planned",
                       "No public ArcGIS REST service found. Zoning via DLCD statewide layer (ownerName='Canby'). Contact Canby Planning (503-266-7001) for direct shapefile/REST access."),
                _layer("Johnson City Zoning", "—", "DLCD Fallback", "Planned",
                       "0.07 sq mi micro-municipality (single mobile home park). No city GIS program. DLCD statewide is the only queryable source (ownerName='Johnson City')."),
            ],
        },
    ]

    # ── Parallel host liveness checks ──────────────────────────────────────
    _PING_HOST_URLS: dict[str, str] = {
        "services2.arcgis.com":     "https://services2.arcgis.com",
        "services8.arcgis.com":     "https://services8.arcgis.com",
        "services1.arcgis.com":     "https://services1.arcgis.com",
        "services.arcgis.com":      "https://services.arcgis.com",
        "services5.arcgis.com":     "https://services5.arcgis.com",
        "services6.arcgis.com":     "https://services6.arcgis.com",
        "services7.arcgis.com":     "https://services7.arcgis.com",
        "gis.odot.state.or.us":        "https://gis.odot.state.or.us",
        "gis.greshamoregon.gov":       "https://gis.greshamoregon.gov",
        "maps.troutdaleoregon.gov":    "https://maps.troutdaleoregon.gov",
        "maps.orcity.org":             "https://maps.orcity.org",
        "maps.ci.oswego.or.us":        "https://maps.ci.oswego.or.us",
        "geo.westlinnoregon.gov":      "https://geo.westlinnoregon.gov",
        "tualgis.ci.tualatin.or.us":   "https://tualgis.ci.tualatin.or.us",
        "gis.wilsonvillemaps.com":     "https://gis.wilsonvillemaps.com",
    }
    unique_hosts = {layer["ping_host"] for g in groups for layer in g["layers"] if layer["ping_host"]}

    async def _ping(host: str) -> tuple[str, bool]:
        url = _PING_HOST_URLS.get(host, f"https://{host}")
        ok = await _direct_live_ping(url, timeout_seconds=5.0)
        return host, ok

    ping_pairs = await asyncio.gather(*[_ping(h) for h in unique_hosts])
    host_ok: dict[str, bool] = dict(ping_pairs)

    # ── File mtime → last pull timestamp ───────────────────────────────────
    _CACHE_ROOT = Path("/app/data/gis_cache")
    heartbeat_ts = datetime.now(_PACIFIC).strftime("%Y-%m-%d %H:%M PT")

    for group in groups:
        for layer in group["layers"]:
            cf = layer.get("cache_file", "")
            if cf:
                fpath = _CACHE_ROOT / cf
                if fpath.exists():
                    mtime = datetime.fromtimestamp(fpath.stat().st_mtime, tz=_PACIFIC)
                    layer["last_pull"] = mtime.strftime("%Y-%m-%d %H:%M PT")
                else:
                    layer["last_pull"] = "Not cached"
            else:
                layer["last_pull"] = "—"

            h = layer.get("ping_host", "")
            if h:
                layer["heartbeat_ok"] = host_ok.get(h, False)
                layer["heartbeat_ts"] = heartbeat_ts
            else:
                layer["heartbeat_ok"] = None  # no check applicable
                layer["heartbeat_ts"] = "—"

    return templates.TemplateResponse(
        request,
        "settings_data_sources.html",
        {
            "groups": groups,
            "heartbeat_ts": heartbeat_ts,
            **_base_ctx(user, dedup_count, "", address_issues_count, conflicts_count=conflicts_count),
        },
    )


# ---------------------------------------------------------------------------
# GET /settings/organization
# ---------------------------------------------------------------------------

@router.get("/settings/organization", response_class=HTMLResponse)
async def settings_organization(
    request: Request,
    session: DBSession,
) -> HTMLResponse:
    user = await _get_user(session, request)
    if user is None:
        return RedirectResponse(url="/login?next=/settings/organization", status_code=303)
    dedup_count, conflicts_count = await _get_counts(session)
    address_issues_count = await _get_address_issues_count(session)

    org = await session.get(Organization, user.org_id)
    org_users = list(
        (
            await session.execute(
                select(User).where(User.org_id == user.org_id).order_by(User.created_at)
            )
        ).scalars()
    )

    from app.models.settings import OrgSetting as _OrgSetting
    from app.settings.defaults import ORG_SET_FIELDS as _ORG_SET_FIELDS
    from app.settings.resolver import resolve_all_defaults as _resolve_all

    resolved = await _resolve_all(user.id, user.org_id, session)
    _org_rows = (
        await session.execute(select(_OrgSetting).where(_OrgSetting.org_id == user.org_id))
    ).scalars().all()
    org_settings_map = {
        r.field_key: {"value": r.value, "user_overridable": r.user_overridable}
        for r in _org_rows
    }

    from app.models.source_vehicle import SourceVehicle as _OSV_org
    org_source_vehicles = (
        await session.execute(
            select(_OSV_org).where(
                _OSV_org.scope == "org", _OSV_org.owner_id == user.org_id
            ).order_by(_OSV_org.label)
        )
    ).scalars().all()

    from app.models.settings import OrgDealTypeDefault as _OrgDTD
    from app.settings.resolver import resolve_timeline_defaults as _resolve_tl_org
    timeline_defaults_map = await _resolve_tl_org(user.id, user.org_id, session)
    _org_tl_rows = (
        await session.execute(select(_OrgDTD).where(_OrgDTD.org_id == user.org_id))
    ).scalars().all()
    org_timeline_map = {
        (r.deal_type, r.milestone_type): {
            "included": r.included,
            "duration_days": int(r.duration_days) if r.duration_days is not None else None,
            "starts_after_type": r.starts_after_type,
            "offset_days": int(r.offset_days),
            "user_overridable": r.user_overridable,
        }
        for r in _org_tl_rows
    }

    return templates.TemplateResponse(
        request,
        "settings_organization.html",
        {
            "org": org,
            "org_users": org_users,
            "user": user,
            "resolved": resolved,
            "org_settings_map": org_settings_map,
            "org_set_fields": _ORG_SET_FIELDS,
            "org_source_vehicles": org_source_vehicles,
            "timeline_defaults_map": timeline_defaults_map,
            "org_timeline_map": org_timeline_map,
            **_base_ctx(user, dedup_count, "", address_issues_count, conflicts_count=conflicts_count),
        },
    )


@router.post("/settings/organization", response_class=HTMLResponse)
async def settings_organization_post(
    request: Request,
    session: DBSession,
    org_name: str = Form(...),
    org_slug: str = Form(None),
) -> HTMLResponse:
    user = await _get_user(session, request)
    if user is None:
        return RedirectResponse(url="/login?next=/settings/organization", status_code=303)
    if not getattr(user, "is_org_admin", False):
        return HTMLResponse("Access denied", status_code=403)

    org = await session.get(Organization, user.org_id)
    if org is None:
        return HTMLResponse("Organization not found", status_code=404)

    org.name = org_name.strip()
    if org_slug and org_slug.strip():
        org.slug = org_slug.strip().lower().replace(" ", "-")
    await session.commit()

    # Redirect back to GET to show updated data
    return RedirectResponse(url="/settings/organization", status_code=303)


# ---------------------------------------------------------------------------
# GET /settings/org — org-wide underwriting defaults (admin only)
# ---------------------------------------------------------------------------


@router.get("/settings/org", response_class=HTMLResponse)
async def settings_org_defaults(request: Request) -> HTMLResponse:
    return RedirectResponse(url="/settings/organization", status_code=301)


# ---------------------------------------------------------------------------
# GET /settings/preferences — per-user underwriting preferences
# ---------------------------------------------------------------------------


@router.get("/settings/preferences", response_class=HTMLResponse)
async def settings_user_preferences(
    request: Request,
    session: DBSession,
) -> HTMLResponse:
    from app.models.settings import OrgSetting as _OrgSetting
    from app.models.settings import UserSetting as _UserSetting
    from app.settings.defaults import ORG_SET_FIELDS as _ORG_SET_FIELDS
    from app.settings.resolver import build_overridable_map as _build_overridable_map
    from app.settings.resolver import resolve_all_defaults as _resolve_all

    user = await _get_user(session, request)
    if user is None:
        return RedirectResponse(url="/login?next=/settings/preferences", status_code=303)

    dedup_count, conflicts_count = await _get_counts(session)
    address_issues_count = await _get_address_issues_count(session)
    resolved = await _resolve_all(user.id, user.org_id, session)

    _org_rows = (
        await session.execute(select(_OrgSetting).where(_OrgSetting.org_id == user.org_id))
    ).scalars().all()
    _user_rows = (
        await session.execute(select(_UserSetting).where(_UserSetting.user_id == user.id))
    ).scalars().all()
    overridable = _build_overridable_map(_org_rows)
    user_values = {r.field_key: r.value for r in _user_rows}

    from app.models.settings import UserDealTypeDefault as _UserDTD
    from app.settings.resolver import resolve_timeline_defaults as _resolve_tl2

    timeline_defaults_map = await _resolve_tl2(user.id, user.org_id, session)
    _user_dtd_rows = (
        await session.execute(select(_UserDTD).where(_UserDTD.user_id == user.id))
    ).scalars().all()
    user_timeline_map = {
        (r.deal_type, r.milestone_type): {
            "included": r.included,
            "duration_days": int(r.duration_days) if r.duration_days is not None else None,
            "starts_after_type": r.starts_after_type,
            "offset_days": int(r.offset_days),
        }
        for r in _user_dtd_rows
    }
    from app.models.settings import OrgDealTypeDefault as _OrgDTD2
    _org_dtd_rows2 = (
        await session.execute(select(_OrgDTD2).where(_OrgDTD2.org_id == user.org_id))
    ).scalars().all()
    timeline_overridable = {
        (r.deal_type, r.milestone_type): r.user_overridable for r in _org_dtd_rows2
    }

    from app.models.source_vehicle import SourceVehicle as _SV_usr
    org_source_vehicles_usr = (
        await session.execute(
            select(_SV_usr).where(
                _SV_usr.scope == "org", _SV_usr.owner_id == user.org_id
            ).order_by(_SV_usr.label)
        )
    ).scalars().all()
    user_source_vehicles = (
        await session.execute(
            select(_SV_usr).where(
                _SV_usr.scope == "user", _SV_usr.owner_id == user.id
            ).order_by(_SV_usr.label)
        )
    ).scalars().all()

    return templates.TemplateResponse(
        request,
        "settings_user.html",
        {
            "user": user,
            "resolved": resolved,
            "overridable": overridable,
            "user_values": user_values,
            "org_set_fields": _ORG_SET_FIELDS,
            "timeline_defaults_map": timeline_defaults_map,
            "user_timeline_map": user_timeline_map,
            "timeline_overridable": timeline_overridable,
            "org_source_vehicles": org_source_vehicles_usr,
            "user_source_vehicles": user_source_vehicles,
            **_base_ctx(user, dedup_count, "", address_issues_count, conflicts_count=conflicts_count),
        },
    )


@router.get("/settings/billing", response_class=HTMLResponse)
async def settings_billing(
    request: Request,
    session: DBSession,
    stripe: str = Query(default=""),
) -> HTMLResponse:
    user = await _get_user(session, request)
    if user is None:
        return RedirectResponse(url="/login?next=/settings/billing", status_code=303)

    dedup_count, conflicts_count = await _get_counts(session)
    address_issues_count = await _get_address_issues_count(session)

    stripe_configured = _stripe_is_configured()
    stripe_test_mode = _stripe_secret_key().startswith("sk_test_") if stripe_configured else False
    customer_id = await _get_stripe_customer_id(session, user.id) if stripe_configured else None

    cards: list[dict[str, str]] = []
    cards_error: str | None = None
    if stripe_configured and customer_id:
        try:
            cards = await _list_stripe_payment_methods(customer_id)
        except HTTPException as exc:
            cards_error = str(exc.detail)

    return templates.TemplateResponse(
        request,
        "settings_billing.html",
        {
            "stripe_configured": stripe_configured,
            "stripe_test_mode": stripe_test_mode,
            "stripe_customer_id": customer_id,
            "cards": cards,
            "cards_error": cards_error,
            "stripe_state": stripe,
            "stripe_state_message": _stripe_state_message(stripe),
            **_base_ctx(user, dedup_count, "", address_issues_count, conflicts_count=conflicts_count),
        },
    )


@router.post("/settings/billing/stripe/setup-session", response_class=HTMLResponse)
async def settings_billing_create_setup_session(
    request: Request,
    session: DBSession,
) -> HTMLResponse:
    user = await _get_user(session, request)
    if user is None:
        return RedirectResponse(url="/login?next=/settings/billing", status_code=303)

    if not _stripe_is_configured():
        return RedirectResponse(url="/settings/billing?stripe=config-missing", status_code=303)

    def _checkout_payload(customer_id: str) -> list[tuple[str, str]]:
        return [
            ("mode", "setup"),
            ("customer", customer_id),
            ("payment_method_types[]", "card"),
            ("billing_address_collection", "auto"),
            ("success_url", f"{settings.app_base_url}/settings/billing?stripe=success"),
            ("cancel_url", f"{settings.app_base_url}/settings/billing?stripe=cancel"),
        ]

    try:
        customer_id = await _ensure_stripe_customer_id(session, user)
        checkout = await _stripe_api_request(
            "POST",
            "/v1/checkout/sessions",
            data=_checkout_payload(customer_id),
        )
    except HTTPException as exc:
        detail = str(exc.detail or "")
        if "No such customer" in detail or "resource_missing" in detail:
            try:
                await _clear_stripe_customer_id(session, user.id)
                customer_id = await _ensure_stripe_customer_id(session, user)
                checkout = await _stripe_api_request(
                    "POST",
                    "/v1/checkout/sessions",
                    data=_checkout_payload(customer_id),
                )
            except HTTPException:
                return RedirectResponse(url="/settings/billing?stripe=error", status_code=303)
        else:
            return RedirectResponse(url="/settings/billing?stripe=error", status_code=303)

    checkout_url = str(checkout.get("url") or "").strip()
    if not checkout_url:
        return RedirectResponse(url="/settings/billing?stripe=error", status_code=303)
    return RedirectResponse(url=checkout_url, status_code=303)


@router.post("/ui/admin/rlis-refresh")
async def admin_rlis_refresh(
    request: Request,
    session: DBSession,
) -> JSONResponse:
    """
    Dispatch rlis_quarterly_refresh_task to the Celery default queue.
    Assumes rlis_delta.py has already been run (cache + sidecar are fresh).
    Returns the Celery task ID.
    """
    from app.tasks.parcel_seed import rlis_quarterly_refresh_task
    user = await _get_user(session, request)
    _require_settings_owner(user)
    result = rlis_quarterly_refresh_task.delay()
    return JSONResponse({"task_id": result.id, "status": "queued"})


@router.post("/ui/admin/seed-rlis")
async def admin_seed_rlis(
    request: Request,
    session: DBSession,
) -> JSONResponse:
    """Dispatch seed_rlis_task — re-seeds parcels from the cached taxlot GeoJSON."""
    from app.tasks.parcel_seed import seed_rlis_task
    user = await _get_user(session, request)
    _require_settings_owner(user)
    result = seed_rlis_task.delay()
    return JSONResponse({"task_id": result.id, "status": "queued"})


@router.post("/ui/admin/backfill-listing-buckets")
async def admin_backfill_listing_buckets(
    session: DBSession,
) -> JSONResponse:
    """Classify all ScrapedListings that have no priority_bucket yet.
    Uses zoning/county/property_type from the linked Parcel if available,
    otherwise falls back to listing fields."""
    from app.utils.priority import classify as _classify
    from app.models.parcel import Parcel

    stmt = (
        select(ScrapedListing)
        .where(ScrapedListing.priority_bucket.is_(None))
        .options(selectinload(ScrapedListing.broker))
    )
    listings = list((await session.execute(stmt)).scalars())
    updated = 0
    for listing in listings:
        parcel: Parcel | None = None
        if listing.parcel_id:
            parcel = await session.get(Parcel, listing.parcel_id)
        elif listing.apn:
            parcel = (await session.execute(
                select(Parcel).where(Parcel.apn == listing.apn.split(",")[0].strip().upper())
            )).scalar_one_or_none()

        bucket = _classify(
            zoning_code=(parcel.zoning_code if parcel else None) or listing.zoning,
            zoning_description=(parcel.zoning_description if parcel else None),
            county=(parcel.county if parcel else None) or listing.county,
            jurisdiction=(parcel.jurisdiction if parcel else None) or listing.city,
            current_use=(parcel.current_use if parcel else None),
            property_type=listing.property_type,
        )
        listing.priority_bucket = bucket.value
        updated += 1

    await session.commit()
    return JSONResponse({"updated": updated})


@router.post("/ui/admin/classify-parcels")
async def admin_classify_parcels(
    request: Request,
    session: DBSession,
) -> JSONResponse:
    """Dispatch classify_parcels_task — classifies parcels with data but no bucket."""
    from app.tasks.parcel_seed import classify_parcels_task
    user = await _get_user(session, request)
    _require_settings_owner(user)
    result = classify_parcels_task.delay()
    return JSONResponse({"task_id": result.id, "status": "queued"})


@router.get("/deals/new", response_class=HTMLResponse)
async def deals_new_page(
    request: Request,
    session: DBSession,
    opp_id: str = Query(default=""),
    from_opp: str = Query(default=""),
    from_listing: str = Query(default=""),
    clone_of: str = Query(default=""),
) -> HTMLResponse:
    """Single landing page for creating a new deal.

    Query params carry context from upstream entry points so every "Create
    Deal" button across the app funnels through this one URL:

    - ``?from_opp=<opportunity_id>`` (alias: ``?opp_id=``) — pre-fill name +
      asking-price from a linked Opportunity. Submit creates a new Scenario
      linked to that Opportunity.
    - ``?from_listing=<scraped_listing_id>`` — pre-fill from a ScrapedListing.
      The listing is promoted into an Opportunity at submit time.
    - ``?clone_of=<scenario_id>`` — pre-fill from an existing Scenario.
      Submit clones the source's Projects, UseLines, IncomeStreams,
      ExpenseLines, CapitalModules, WaterfallTiers, etc. while re-applying
      current org Type 1 defaults.
    """
    user = await _get_user(session, request)
    dedup_count, conflicts_count = await _get_counts(session)
    ctx = _base_ctx(user, dedup_count, "deals", conflicts_count=conflicts_count)

    # Normalize ``opp_id`` (legacy) → ``from_opp``. Both supported during the
    # transition period; from_opp wins if both are set.
    effective_opp_id = (from_opp or opp_id).strip()

    pre_name: str = ""
    pre_acquisition_cost: float | None = None
    pre_deal_type: str = "acquisition"
    banner_text: str = ""
    context_kind: str = "blank"  # one of: blank | from_opp | from_listing | clone_of

    if clone_of:
        try:
            _src = await session.get(DealModel, UUID(clone_of))
        except ValueError:
            _src = None
        if _src is not None:
            context_kind = "clone_of"
            pre_name = f"{_src.name} (Copy)"
            pre_deal_type = getattr(_src.project_type, "value", _src.project_type) or "acquisition"
            banner_text = f"Cloning from: {_src.name}"
    elif from_listing:
        try:
            _listing = await session.get(
                ScrapedListing, UUID(from_listing),
                options=[selectinload(ScrapedListing.broker)],
            )
        except ValueError:
            _listing = None
        if _listing is not None:
            context_kind = "from_listing"
            pre_name = _listing.address_normalized or _listing.address_raw or "Unnamed Listing Deal"
            if _listing.asking_price is not None and float(_listing.asking_price) > 0:
                pre_acquisition_cost = float(_listing.asking_price)
            banner_text = f"From listing: {pre_name}"
    elif effective_opp_id:
        try:
            _opp = await session.get(Opportunity, UUID(effective_opp_id))
        except ValueError:
            _opp = None
        if _opp is not None:
            context_kind = "from_opp"
            pre_name = _opp.name or ""
            if _opp.asking_price is not None and _opp.asking_price > 0:
                pre_acquisition_cost = float(_opp.asking_price)
            banner_text = f"Linked to opportunity: {pre_name}"

    ctx.update({
        "context_kind": context_kind,
        "banner_text": banner_text,
        "from_opp": effective_opp_id if context_kind == "from_opp" else "",
        "from_listing": from_listing if context_kind == "from_listing" else "",
        "clone_of": clone_of if context_kind == "clone_of" else "",
        "pre_name": pre_name,
        "pre_acquisition_cost": pre_acquisition_cost,
        "pre_deal_type": pre_deal_type,
        # Legacy template variable kept for safety; new template prefers pre_acquisition_cost.
        "opp_id": effective_opp_id,
        "opp_name": pre_name,
        "opp_asking_price": pre_acquisition_cost,
    })
    return templates.TemplateResponse(request, "deals_new.html", ctx)


@router.get("/deals", response_class=HTMLResponse)
async def deals_page(
    request: Request,
    session: DBSession,
    q: str = Query(default=""),
    status: list[str] = Query(default=[]),
    type: list[str] = Query(default=[]),
    model: list[str] = Query(default=[]),
    include_archived: str = Query(default=""),
    hide_test: str = Query(default=""),
) -> HTMLResponse:
    user = await _get_user(session, request)
    dedup_count, conflicts_count = await _get_counts(session)

    archived = include_archived == "1"
    # Default hide_test ON for admin users when not explicitly set
    is_admin = user is not None and bool(getattr(user, "is_admin", False))
    effective_hide_test = (hide_test == "1") if hide_test != "" else is_admin
    loaded_deals = await _load_deals(session, status, type, model, q, archived, effective_hide_test)
    deals = [_build_deal_row(d) for d in loaded_deals]

    total_result = await session.execute(
        select(func.count()).select_from(Deal).where(Deal.status != DealStatus.archived)
    )
    total_count = int(total_result.scalar_one())

    archived_result = await session.execute(
        select(func.count()).select_from(Deal).where(Deal.status == DealStatus.archived)
    )
    archived_count = int(archived_result.scalar_one())

    irr_values = [d["irr"] for d in deals if d["irr"] is not None]
    avg_irr = sum(irr_values) / len(irr_values) if irr_values else None
    equity_values = [d["noi"] for d in deals if d.get("noi") is not None]  # use NOI as pipeline proxy
    # pipeline_value = total equity required across deals with outputs
    equity_req_values: list[float] = []
    for loaded_deal in loaded_deals:
        scenario = _primary_scenario(loaded_deal)
        if scenario and scenario.operational_outputs and scenario.operational_outputs.equity_required is not None:
            equity_req_values.append(float(scenario.operational_outputs.equity_required))
    pipeline_value = sum(equity_req_values) if equity_req_values else None

    return templates.TemplateResponse(
        request,
        "deals.html",
        {
            "deals": deals,
            "total_count": total_count,
            "archived_count": archived_count,
            "include_archived": archived,
            "hide_test": effective_hide_test,
            "q": q,
            "status": status,
            "deal_type": type,
            "model_filter": model,
            "stats": {
                "pipeline_count": total_count,
                "avg_irr": avg_irr,
                "pipeline_value": pipeline_value,
                "no_model_count": sum(1 for d in deals if not d["primary_model_name"]),
            },
            **_base_ctx(user, dedup_count, "deals", conflicts_count=conflicts_count),
        },
    )


@router.get("/ui/deals/rows", response_class=HTMLResponse)
async def deals_rows(
    request: Request,
    session: DBSession,
    q: str = Query(default=""),
    status: list[str] = Query(default=[]),
    type: list[str] = Query(default=[]),
    model: list[str] = Query(default=[]),
    include_archived: str = Query(default=""),
    hide_test: str = Query(default=""),
) -> HTMLResponse:
    archived = include_archived == "1"
    loaded_deals = await _load_deals(session, status, type, model, q, archived, hide_test == "1")
    deals = [_build_deal_row(d) for d in loaded_deals]
    return templates.TemplateResponse(request, "partials/deals_rows.html", {"deals": deals})


def _seed_milestones(project: Project, deal_type: ProjectType) -> list[Milestone]:
    """Return unseeded Milestone rows for a new dev Project based on deal_type defaults."""
    durations = DEFAULT_DURATIONS.get(deal_type.value, {})
    milestones = []
    for seq, (type_str, duration) in enumerate(durations.items(), start=1):
        try:
            mtype = MilestoneType(type_str)
        except ValueError:
            continue
        milestones.append(Milestone(
            project_id=project.id,
            milestone_type=mtype,
            duration_days=duration,
            sequence_order=seq,
        ))
    return milestones


@router.post("/ui/deals/create", response_class=HTMLResponse)
async def create_deal(
    request: Request,
    session: DBSession,
    new: str = Query(default=""),
) -> HTMLResponse:
    """Unified deal-creation entry point. All non-clone deal creations land here.

    Form hidden fields carry context from /deals/new:
      - ``opportunity_id`` — link to an existing Opportunity (from-opp path)
      - ``listing_id``    — promote a ScrapedListing into an Opportunity
                            (from-listing path); falls back to acquisition_cost
                            taken from the listing if the form omits it.

    Existing-Opportunity path: links to that Opportunity.
    From-listing path: promotes the listing to org-scoped Opportunity.
    Blank: creates a new manual Opportunity from the form name.

    Clone path uses POST /ui/deals/{deal_id}/variant — separate handler
    because of its deep-copy semantics.
    """
    form = await request.form()
    name = str(form.get("name", "")).strip()
    deal_type_raw = str(form.get("deal_type", "acquisition")).strip()
    org_id_raw = str(form.get("org_id", "")).strip()
    opp_id_raw = str(form.get("opportunity_id", "")).strip()
    listing_id_raw = str(form.get("listing_id", "")).strip()
    acq_cost_raw = str(form.get("acquisition_cost", "")).strip()

    user = await _get_user(session, request)

    # ── Listing-promotion path: resolve the ScrapedListing first so it can
    # feed defaults into name + acq_cost when the form omits them.
    listing: ScrapedListing | None = None
    if listing_id_raw:
        try:
            listing = await session.get(
                ScrapedListing, UUID(listing_id_raw),
                options=[selectinload(ScrapedListing.broker)],
            )
        except ValueError:
            listing = None
        if listing is None:
            return HTMLResponse("<p class='text-muted'>Listing not found.</p>", status_code=404)
        # Enforce listing has a real asking price — without it the seeded
        # Acquisition UseLine lands at $0 and downstream debt sizing produces
        # gaps that can't be reconciled later.
        if not listing.asking_price or float(listing.asking_price) <= 0:
            return HTMLResponse(
                "<p class='text-muted'>This listing has no asking price. "
                "Set a price on the listing first, or create the deal manually "
                "from <a href='/deals/new'>Deals → New</a>.</p>",
                status_code=400,
            )
        # Listing-derived defaults when form fields are blank.
        if not name:
            name = listing.address_normalized or listing.address_raw or "Unnamed Listing Deal"
        if not acq_cost_raw:
            acq_cost_raw = str(listing.asking_price)

    if not name:
        return HTMLResponse("<p class='text-muted'>Deal name is required.</p>", status_code=400)

    # Required: acquisition_cost > 0. Same invariant as the Add-Project flow —
    # the seeded UseLine must have a real value or downstream debt sizing
    # silently produces gaps that can't be reconciled by a later edit.
    try:
        acq_cost = Decimal(acq_cost_raw) if acq_cost_raw else Decimal("0")
    except (InvalidOperation, ValueError):
        return HTMLResponse(
            "<p class='text-muted'>Invalid acquisition cost.</p>", status_code=400,
        )
    if acq_cost <= 0:
        return HTMLResponse(
            "<p class='text-muted'>Acquisition cost must be greater than zero.</p>",
            status_code=400,
        )

    # Resolve org_id: form value → user's org → first org
    org_id = None
    if org_id_raw:
        try:
            org_id = UUID(org_id_raw)
        except ValueError:
            pass
    if org_id is None and user is not None:
        org_id = user.org_id
    if org_id is None:
        from app.models.org import Organization
        first_org = (await session.execute(select(Organization).limit(1))).scalar_one_or_none()
        if first_org is None:
            return HTMLResponse("<p class='text-muted'>No organization found. Create one first.</p>", status_code=400)
        org_id = first_org.id

    try:
        deal_type = ProjectType(deal_type_raw)
    except ValueError:
        deal_type = ProjectType.acquisition

    # Resolve / create Opportunity. Three paths:
    #   1. opportunity_id form field → link existing
    #   2. listing_id form field → promote the ScrapedListing
    #   3. blank → create a fresh manual Opportunity from the form name
    opportunity: Opportunity | None = None
    if opp_id_raw:
        try:
            opportunity = await session.get(Opportunity, UUID(opp_id_raw))
        except ValueError:
            pass

    if opportunity is None and listing is not None:
        # ScrapedListing IS the Opportunity (single-table inheritance) —
        # check for an existing Deal linked via Scenario→Project before
        # creating a duplicate.
        existing_listing_deal = (await session.execute(
            select(Deal)
            .join(DealModel, DealModel.deal_id == Deal.id)
            .join(Project, Project.scenario_id == DealModel.id)
            .where(Project.opportunity_id == listing.id)
            .limit(1)
        )).scalar_one_or_none()
        if existing_listing_deal is not None:
            return RedirectResponse(url=f"/deals/{existing_listing_deal.id}", status_code=303)
        opportunity = listing
        if not opportunity.org_id:
            opportunity.org_id = org_id
            opportunity.opp_status = OpportunityStatus.active.value
            if not opportunity.name:
                opportunity.name = name
            # Enrich parcel link from APN/address if missing
            if opportunity.parcel_id is None and (opportunity.apn or opportunity.address_normalized):
                try:
                    from app.scrapers.parcel_enrichment import enrich_parcel as _enrich
                    _parcel = await _enrich(
                        session,
                        address=opportunity.address_normalized or opportunity.address_raw,
                        apn=opportunity.apn,
                    )
                    if _parcel is not None:
                        opportunity.parcel_id = _parcel.id
                except Exception:
                    pass
            await session.flush()

    _opportunity_is_new = False
    if opportunity is None:
        _opportunity_is_new = True
        opportunity = Opportunity(
            org_id=org_id,
            name=name,
            opp_status=OpportunityStatus.active.value,
            source="manual",
            source_id=_uuid_mod.uuid4().hex,
            source_url="",
            created_by_user_id=user.id if user else None,
        )
        session.add(opportunity)
        await session.flush()
    else:
        # Existing-opp path: ensure parcel is linked on the Opportunity itself.
        if opportunity.parcel_id is None and (opportunity.apn or opportunity.address_normalized):
            try:
                from app.scrapers.parcel_enrichment import enrich_parcel as _enrich
                _p = await _enrich(
                    session,
                    address=opportunity.address_normalized or opportunity.address_raw,
                    apn=opportunity.apn,
                )
                if _p is not None:
                    opportunity.parcel_id = _p.id
            except Exception:
                pass
        await session.flush()

    # Deal → Scenario (financial plan) → Project → Opportunity
    top_deal = Deal(
        org_id=org_id,
        name=name,
        created_by_user_id=user.id if user else None,
    )
    session.add(top_deal)
    await session.flush()

    from app.services.scenario_factory import create_scenario as _create_scenario
    scenario, dev_project, _ = await _create_scenario(
        session=session,
        deal_id=top_deal.id,
        deal_type=deal_type,
        user_id=user.id if user else None,
        org_id=org_id,
        opportunity_id=opportunity.id,
    )

    await _auto_assign_opportunity_to_project(opportunity, dev_project, session)
    for milestone in _seed_milestones(dev_project, deal_type):
        session.add(milestone)

    # Seed the Acquisition UseLine with the user-confirmed cost. Pre-fill
    # in the form pulls from the linked listing's asking_price; user can
    # override with their underwriting price before submit.
    session.add(UseLine(
        project_id=dev_project.id,
        label=f"{opportunity.name or 'Property'} - Acquisition",
        phase=UseLinePhase.acquisition,
        cost_category="acquisition",
        amount=acq_cost,
        timing_type="first_day",
    ))

    from app.services.vehicle_preload import preload_equity_modules
    await preload_equity_modules(session, scenario.id, org_id, project_id=dev_project.id)

    # Auto Developer Fee — engine recomputes $ each pass; user overrides %
    # in the Use drawer. Set pct=0 to effectively disable for this deal.
    if user is not None:
        dev_fee_cfg = await resolve_dev_fee_config(
            user.id, org_id, deal_type, session
        )
        if str(dev_fee_cfg["enabled"]).lower() == "true":
            try:
                _pct = Decimal(dev_fee_cfg["pct"])
            except (InvalidOperation, TypeError):
                _pct = Decimal("0")
            _phase_str = dev_fee_cfg["phase"]
            try:
                _phase_enum = UseLinePhase(_phase_str)
            except ValueError:
                _phase_enum = UseLinePhase.construction
            session.add(UseLine(
                project_id=dev_project.id,
                label="Developer Fee",
                phase=_phase_enum,
                cost_category="soft",
                amount=Decimal("0"),
                timing_type=dev_fee_cfg["timing"],
                is_auto_dev_fee=True,
                dev_fee_pct=_pct,
                dev_fee_basis=dev_fee_cfg["basis"],
            ))

    await session.commit()

    # Single-flow wizard: land directly inside the wizard chrome at the timeline
    # step. The user never sees the full builder UI until they finish the
    # setup wizard. Builder route reads wizard=1 to hide sidebar/topbar and
    # the approve-timeline handler reads _wizard to route into setup.
    redirect_url = f"/models/{scenario.id}/builder?module=timeline&wizard=1"
    if new == "1":
        redirect_url += "&new=1"
    return RedirectResponse(url=redirect_url, status_code=303)


@router.get("/deals/{deal_id}", response_class=HTMLResponse)
async def deal_detail(
    request: Request,
    deal_id: UUID,
    session: DBSession,
    tab: str = Query(default="overview"),
    error: str = Query(default=""),
) -> HTMLResponse:
    user = await _get_user(session, request)
    dedup_count, conflicts_count = await _get_counts(session)

    deal = await session.get(
        Deal,
        deal_id,
        options=[
            selectinload(Deal.scenarios).selectinload(DealModel.operational_outputs),
            selectinload(Deal.scenarios).selectinload(DealModel.projects).selectinload(Project.milestones),
            selectinload(Deal.scenarios).selectinload(DealModel.projects).selectinload(Project.opportunity),
        ],
    )
    if deal is None or (user is not None and deal.org_id != user.org_id):
        return HTMLResponse("<p class='text-muted'>Deal not found.</p>", status_code=404)

    opportunity = _first_opportunity(deal)

    buildings = []  # Building entity removed — physical attrs now on Opportunity

    # Parcel linked directly to the Opportunity
    parcels = []
    if opportunity and opportunity.parcel_id:
        parcel = await session.get(Parcel, opportunity.parcel_id)
        if parcel:
            parcels = [_build_parcel_row(parcel)]

    # Financial models (Scenarios) for this Deal
    models = []
    for scenario in deal.scenarios:
        out = scenario.operational_outputs
        type_key = str(scenario.project_type.value if hasattr(scenario.project_type, "value") else scenario.project_type)
        first_proj = scenario.projects[0] if scenario.projects else None
        models.append({
            "id": str(scenario.id),
            "name": scenario.name,
            "version": scenario.version,
            "is_active": scenario.is_active,
            "type_display": _TYPE_DISPLAY.get(type_key, type_key),
            "project_name": first_proj.name if first_proj else "—",
            "project_id": str(first_proj.id) if first_proj else None,
            "noi": float(out.noi_stabilized) if out and out.noi_stabilized is not None else None,
            "irr": float(out.project_irr_levered) if out and out.project_irr_levered is not None else None,
            "equity_required": float(out.equity_required) if out and out.equity_required is not None else None,
            "created_at_fmt": scenario.created_at.strftime("%b %-d, %Y") if scenario.created_at else None,
        })
    models.sort(key=lambda m: (0 if m["is_active"] else 1, -m["version"]))

    # Build Gantt data from milestones across all scenarios/projects
    gantt_data = _build_gantt_rows(deal)

    # Status comes from the linked Opportunity (pipeline stage)
    status_key = str(opportunity.status.value if opportunity and hasattr(opportunity.status, "value") else (opportunity.status if opportunity else "active"))
    status_display, status_badge = _STATUS_DISPLAY.get(status_key, ("Unknown", "badge-gray"))

    return templates.TemplateResponse(
        request,
        "deal_detail.html",
        {
            "deal": deal,
            "deal_id": str(deal.id),
            "deal_name": deal.name,
            "opp": opportunity,
            "opp_id": str(opportunity.id) if opportunity else "",
            "opp_name": opportunity.name if opportunity else "",
            "status_key": status_key,
            "status_display": status_display,
            "status_badge": status_badge,
            "buildings": buildings,
            "parcels": parcels,
            "models": models,
            "gantt_data": gantt_data,
            "active_tab": tab,
            "primary_model_id": models[0]["id"] if models else None,
            "flash_error": error,
            **_base_ctx(user, dedup_count, "deals", conflicts_count=conflicts_count),
        },
    )


@router.post("/ui/deals/{deal_id}/archive", response_class=HTMLResponse)
async def archive_deal(
    request: Request,
    deal_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    user = await _get_user(session, request)
    deal = await session.get(Deal, deal_id)
    if deal is not None and (user is None or deal.org_id == user.org_id):
        deal.status = DealStatus.archived
        await session.flush()
    loaded_deals = await _load_deals(session)
    deals = [_build_deal_row(d) for d in loaded_deals]
    return templates.TemplateResponse(request, "partials/deals_rows.html", {"deals": deals})


@router.post("/ui/deals/{deal_id}/update", response_class=HTMLResponse)
async def update_deal(
    request: Request,
    deal_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    user = await _get_user(session, request)
    form = await request.form()
    name = str(form.get("name", "")).strip()
    status_raw = str(form.get("status", "hypothetical")).strip()

    deal = await session.get(
        Deal,
        deal_id,
        options=[
            selectinload(Deal.scenarios).selectinload(DealModel.projects).selectinload(Project.opportunity),
        ],
    )
    if deal is None or (user is not None and deal.org_id != user.org_id):
        return HTMLResponse("<p class='text-muted'>Not found.</p>", status_code=404)

    if name:
        deal.name = name
    # Pipeline stage is stored on the linked Opportunity
    opp = _first_opportunity(deal)
    if opp is not None:
        try:
            opp.opp_status = status_raw
        except ValueError:
            pass
    await session.commit()
    return RedirectResponse(url=f"/deals/{deal_id}", status_code=303)


@router.post("/ui/deals/{deal_id}/link-parcel", response_class=HTMLResponse)
async def link_parcel_to_deal(
    request: Request,
    deal_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    form = await request.form()
    apn = str(form.get("apn", "")).strip()
    rel_raw = str(form.get("relationship", "unchanged")).strip()

    # Find parcel by APN
    parcel_result = await session.execute(select(Parcel).where(Parcel.apn == apn))
    parcel = parcel_result.scalar_one_or_none()
    if parcel is None:
        return RedirectResponse(
            url=f"/deals/{deal_id}?error=parcel_not_found",
            status_code=303,
        )

    # Get the primary opportunity and set parcel_id directly
    deal = await session.get(
        Deal, deal_id,
        options=[
            selectinload(Deal.scenarios).selectinload(DealModel.projects).selectinload(Project.opportunity),
        ],
    )
    opp = _first_opportunity(deal) if deal else None
    if opp is None:
        return RedirectResponse(url=f"/deals/{deal_id}?error=no_opportunity", status_code=303)

    opp.parcel_id = parcel.id
    await session.commit()

    return RedirectResponse(url=f"/deals/{deal_id}", status_code=303)


# ---------------------------------------------------------------------------
# Buildings
# ---------------------------------------------------------------------------

def _build_building_row(prop: object) -> dict:
    """Stub — Building entity removed. Physical attrs now on Opportunity."""
    return {}


# ── Opportunities ─────────────────────────────────────────────────────────────

@router.get("/opportunities", response_class=HTMLResponse)
async def opportunities_page(
    request: Request,
    session: DBSession,
) -> HTMLResponse:
    user = await _get_user(session, request)
    dedup_count, conflicts_count = await _get_counts(session)
    jur_rows = await session.execute(
        select(func.lower(Opportunity.jurisdiction), func.lower(Opportunity.county))
        .where(Opportunity.jurisdiction.isnot(None))
        .where(Opportunity.jurisdiction != "")
        .distinct()
        .order_by(func.lower(Opportunity.jurisdiction), func.lower(Opportunity.county))
    )
    seen_jur: set[str] = set()
    jurisdiction_options: list[dict[str, str]] = []
    for jur, county in jur_rows:
        if not jur or not jur.strip():
            continue
        if jur == "unincorporated":
            county_part = (county or "").strip()
            value = f"unincorporated_{county_part}" if county_part else "unincorporated"
            label = f"Unin. {county_part.title()}" if county_part else "Unincorporated"
        else:
            value, label = jur, jur.title()
        if value not in seen_jur:
            seen_jur.add(value)
            jurisdiction_options.append({"value": value, "label": label})
    jurisdiction_options.sort(key=lambda x: x["label"])
    is_admin = user is not None and bool(getattr(user, "is_admin", False))
    return templates.TemplateResponse(request, "opportunities.html", {
        "request": request,
        "jurisdiction_options": jurisdiction_options,
        "hide_test_default": is_admin,
        **_base_ctx(user, dedup_count, "opportunities", conflicts_count=conflicts_count),
    })


# ── Opportunities sub-table HTMX endpoints ────────────────────────────────────

def _filter_opps(opps: list, q: str) -> list:
    if not q:
        return opps
    q_lower = q.lower()
    return [
        o for o in opps
        if q_lower in (o.name or "").lower()
        or q_lower in (o.listing_name or "").lower()
        or q_lower in (o.address_normalized or "").lower()
        or q_lower in (o.street or "").lower()
        or q_lower in (o.apn or "").lower()
    ]


def _apply_opp_filters(
    stmt: object,
    favorited: int,
    jurisdiction: list[str],
    min_units: int | None,
    max_units: int | None,
    property_type: list[str],
    hide_test: bool = False,
) -> object:
    if hide_test:
        stmt = stmt.where(
            ~Opportunity.name.ilike("%e2e%") &
            ~Opportunity.name.op("~*")(r"phase\s+\w+\s+test\s+\w+")
        )
    if favorited:
        stmt = stmt.where(Opportunity.is_favorited.is_(True))
    if jurisdiction:
        plain, compound = [], []
        for j in jurisdiction:
            if j.startswith("unincorporated_"):
                compound.append(j[len("unincorporated_"):])
            else:
                plain.append(j.lower())
        clauses = []
        if plain:
            clauses.append(func.lower(Opportunity.jurisdiction).in_(plain))
        for county_part in compound:
            clauses.append(
                and_(
                    func.lower(Opportunity.jurisdiction) == "unincorporated",
                    func.lower(Opportunity.county) == county_part,
                )
            )
        if clauses:
            stmt = stmt.where(or_(*clauses))
    if min_units is not None:
        stmt = stmt.where(Opportunity.units >= min_units)
    if max_units is not None:
        stmt = stmt.where(Opportunity.units <= max_units)
    if property_type:
        stmt = stmt.where(func.lower(Opportunity.property_type).in_([p.lower() for p in property_type]))
    return stmt


@router.get("/ui/opportunities/rows/deals", response_class=HTMLResponse)
async def opportunities_rows_deals(
    request: Request,
    session: DBSession,
    q: str = Query(default=""),
    favorited: int = Query(default=0),
    jurisdiction: list[str] = Query(default=[]),
    min_units: int | None = Query(default=None),
    max_units: int | None = Query(default=None),
    property_type: list[str] = Query(default=[]),
    hide_test: str = Query(default=""),
) -> HTMLResponse:
    active_oppo_ids = select(Project.opportunity_id).where(
        Project.opportunity_id.isnot(None)
    )
    stmt = (
        select(Opportunity)
        .options(selectinload(Opportunity.parcel))
        .where(
            Opportunity.id.in_(active_oppo_ids),
            Opportunity.archived.is_(False),
        )
        .order_by(Opportunity.last_seen_at.desc())
    )
    stmt = _apply_opp_filters(stmt, favorited, jurisdiction, min_units, max_units, property_type, hide_test == "1")
    opps = _filter_opps(list((await session.execute(stmt)).scalars().unique()), q)
    return templates.TemplateResponse(request, "partials/opportunities_rows.html", {
        "request": request, "opps": opps, "table": "deals",
    })


@router.get("/ui/opportunities/rows/offmarket", response_class=HTMLResponse)
async def opportunities_rows_offmarket(
    request: Request,
    session: DBSession,
    q: str = Query(default=""),
    favorited: int = Query(default=0),
    jurisdiction: list[str] = Query(default=[]),
    min_units: int | None = Query(default=None),
    max_units: int | None = Query(default=None),
    property_type: list[str] = Query(default=[]),
    hide_test: str = Query(default=""),
) -> HTMLResponse:
    active_oppo_ids = select(Project.opportunity_id).where(
        Project.opportunity_id.isnot(None)
    )
    stmt = (
        select(Opportunity)
        .options(selectinload(Opportunity.parcel))
        .where(
            Opportunity.promotion_source == "manual",
            Opportunity.id.notin_(active_oppo_ids),
            Opportunity.archived.is_(False),
        )
        .order_by(Opportunity.last_seen_at.desc())
    )
    stmt = _apply_opp_filters(stmt, favorited, jurisdiction, min_units, max_units, property_type, hide_test == "1")
    opps = _filter_opps(list((await session.execute(stmt)).scalars().unique()), q)
    return templates.TemplateResponse(request, "partials/opportunities_rows.html", {
        "request": request, "opps": opps, "table": "offmarket",
    })


@router.get("/ui/opportunities/rows/onmarket", response_class=HTMLResponse)
async def opportunities_rows_onmarket(
    request: Request,
    session: DBSession,
    q: str = Query(default=""),
    favorited: int = Query(default=0),
    jurisdiction: list[str] = Query(default=[]),
    min_units: int | None = Query(default=None),
    max_units: int | None = Query(default=None),
    property_type: list[str] = Query(default=[]),
    hide_test: str = Query(default=""),
) -> HTMLResponse:
    active_oppo_ids = select(Project.opportunity_id).where(
        Project.opportunity_id.isnot(None)
    )
    stmt = (
        select(Opportunity)
        .options(selectinload(Opportunity.parcel))
        .where(
            Opportunity.promotion_source.in_(["loopnet", "crexi", "scraper"]),
            Opportunity.id.notin_(active_oppo_ids),
            Opportunity.archived.is_(False),
        )
        .order_by(Opportunity.last_seen_at.desc())
    )
    stmt = _apply_opp_filters(stmt, favorited, jurisdiction, min_units, max_units, property_type, hide_test == "1")
    opps = _filter_opps(list((await session.execute(stmt)).scalars().unique()), q)
    return templates.TemplateResponse(request, "partials/opportunities_rows.html", {
        "request": request, "opps": opps, "table": "onmarket",
    })


@router.patch("/ui/opportunities/{opp_id}/favorite", response_class=HTMLResponse)
async def toggle_opportunity_favorite(
    request: Request,
    session: DBSession,
    opp_id: UUID,
) -> HTMLResponse:
    opp = await session.get(Opportunity, opp_id)
    if opp is None:
        return HTMLResponse("Not found", status_code=404)
    opp.is_favorited = not opp.is_favorited
    await session.commit()
    await session.refresh(opp)
    # Return updated star button only
    starred = "★" if opp.is_favorited else "☆"
    title = "Unfavorite" if opp.is_favorited else "Favorite"
    return HTMLResponse(
        f'<button class="star-btn {"starred" if opp.is_favorited else ""}"'
        f' hx-patch="/ui/opportunities/{opp_id}/favorite"'
        f' hx-target="closest .star-cell"'
        f' hx-swap="innerHTML"'
        f' title="{title}">{starred}</button>'
    )


# ── Opportunity conflicts page ────────────────────────────────────────────────

@router.get("/ui/opportunities/conflicts", response_class=HTMLResponse)
async def opportunities_conflicts(
    request: Request,
    session: DBSession,
) -> HTMLResponse:
    """Show field-level conflicts between Opportunity physical attrs and linked Parcel."""
    user = await _get_user(session, request)
    dedup_count, conflicts_count = await _get_counts(session)

    stmt = (
        select(Opportunity)
        .options(selectinload(Opportunity.parcel))
        .where(
            Opportunity.parcel_id.isnot(None),
            Opportunity.archived.is_(False),
        )
        .order_by(Opportunity.last_seen_at.desc())
    )
    opps = list((await session.execute(stmt)).scalars().unique())

    # Auto-resolve any conflicts that match the rules before showing the queue
    auto_count = _apply_auto_resolutions(opps)
    if auto_count:
        await session.commit()

    conflicts = _build_conflicts(opps)

    return templates.TemplateResponse(request, "conflicts.html", {
        "request": request, "conflicts": conflicts,
        **_base_ctx(user, dedup_count, "conflicts", conflicts_count=len(set(c[0].id for c in conflicts))),
    })


@router.post("/ui/opportunities/conflicts/auto-resolve", response_class=HTMLResponse)
async def trigger_conflict_auto_resolve(
    request: Request,
    session: DBSession,
) -> HTMLResponse:
    """Enqueue a background task to auto-resolve all existing conflicts.

    Returns a small status fragment suitable for HTMX swap.
    """
    from app.tasks.conflict_backflow import conflict_backflow_task  # noqa: PLC0415
    task = conflict_backflow_task.delay()
    return HTMLResponse(
        f'<span class="badge badge-gray" title="task id: {task.id}">Auto-resolve queued</span>',
        status_code=202,
    )


@router.post("/ui/opportunities/{opp_id}/conflicts/resolve", response_class=HTMLResponse)
async def resolve_opportunity_conflict(
    request: Request,
    session: DBSession,
    opp_id: UUID,
) -> HTMLResponse:
    """Resolve a single field conflict: use_listing, use_parcel, or dismiss."""
    form = await request.form()
    field = str(form.get("field", ""))
    action = str(form.get("action", ""))

    opp = await session.get(Opportunity, opp_id)
    if opp is None or not field:
        return HTMLResponse("Not found", status_code=404)

    if action == "use_parcel":
        # NULL out opp field — defers to parcel via display_* property
        _null_map = {
            "units": "units", "gba_sqft": "gba_sqft",
            "year_built": "year_built", "lot_sqft": "lot_sqft",
        }
        if field in _null_map:
            setattr(opp, _null_map[field], None)

    # Both use_listing and dismiss: ack the field to suppress from queue
    ack = dict(opp.parcel_conflicts_ack or {})
    ack[field] = action
    opp.parcel_conflicts_ack = ack

    await session.commit()
    return HTMLResponse("", status_code=200)


# ── Opportunity creation wizard ────────────────────────────────────────────────

def _safe_uuid_str(raw: str) -> str:
    """Return raw if it parses as a UUID, else empty string."""
    if not raw:
        return ""
    try:
        UUID(raw)
        return raw
    except ValueError:
        return ""


def _safe_return_path(raw: str) -> str:
    """Only allow same-origin paths starting with `/` and free of CRLF / scheme.

    Caller-provided redirect target — we filter to prevent open-redirect to an
    attacker-controlled URL.
    """
    if not raw or not raw.startswith("/") or raw.startswith("//"):
        return ""
    if any(ch in raw for ch in ("\r", "\n")):
        return ""
    return raw


@router.get("/ui/opportunities/wizard", response_class=HTMLResponse)
async def opportunity_wizard_get(
    request: Request,
    session: DBSession,
    step: int = Query(default=1),
    opp_id: str = Query(default=""),
    link_to_deal: str = Query(default=""),
    return_to: str = Query(default=""),
) -> HTMLResponse:
    user = await _get_user(session, request)
    dedup_count, conflicts_count = await _get_counts(session)
    opp = None
    buildings: list = []  # Building entity removed
    if opp_id:
        try:
            opp = await session.get(Opportunity, UUID(opp_id))
        except (ValueError, Exception):
            pass
    ctx = {
        "request": request, "step": step, "opp": opp,
        "opp_id": opp_id, "buildings": buildings,
        "deal_type": request.query_params.get("deal_type", ""),
        "opp_asking_price": "", "opp_notes": "",
        "deal_type_label": "",
        # Carry-through for the "create-from-deal" flow: when the wizard is
        # opened from the Add-Project drawer's empty state, both params are
        # set, threaded through every step's form, and consumed by /complete
        # to link the new opp to the deal and bounce back to the builder.
        "link_to_deal": _safe_uuid_str(link_to_deal),
        "return_to": _safe_return_path(return_to),
        **_base_ctx(user, dedup_count, "opportunities", conflicts_count=conflicts_count),
    }
    return templates.TemplateResponse(request, "opportunity_wizard.html", ctx)


@router.get("/ui/opportunities/wizard/search", response_class=HTMLResponse)
async def opportunity_wizard_search(
    request: Request,
    session: DBSession,
    q: str = Query(default=""),
    opp_id: str = Query(default=""),
) -> HTMLResponse:
    """Step 2 HTMX search — finds a matching scraped Opportunity or Parcel."""
    if not q or len(q.strip()) < 3:
        return HTMLResponse("")

    q_lower = q.strip().lower()

    # Priority 1: search scraped Opportunities by address or APN
    stmt = (
        select(Opportunity)
        .where(
            Opportunity.promotion_source.in_(["loopnet", "crexi", "scraper"]),
            Opportunity.archived.is_(False),
        )
        .order_by(Opportunity.last_seen_at.desc())
        .limit(200)
    )
    candidates = list((await session.execute(stmt)).scalars().unique())
    matched_opp: Opportunity | None = None
    for c in candidates:
        if (
            q_lower in (c.address_normalized or "").lower()
            or q_lower in (c.street or "").lower()
            or q_lower in (c.apn or "").lower()
            or q_lower in (c.name or "").lower()
            or q_lower in (c.listing_name or "").lower()
        ):
            matched_opp = c
            break

    if matched_opp:
        return templates.TemplateResponse(request, "partials/wizard_match_card.html", {
            "request": request,
            "match_type": "listing",
            "match": matched_opp,
            "opp_id": opp_id,
        })

    # Priority 2: fall back to Parcel search
    parcel_stmt = (
        select(Parcel)
        .where(
            Parcel.address_normalized.ilike(f"%{q.strip()}%")
        )
        .limit(1)
    )
    matched_parcel = (await session.execute(parcel_stmt)).scalar_one_or_none()

    if matched_parcel:
        return templates.TemplateResponse(request, "partials/wizard_match_card.html", {
            "request": request,
            "match_type": "parcel",
            "match": matched_parcel,
            "opp_id": opp_id,
        })

    return HTMLResponse(
        '<div style="color:var(--text-muted);font-size:13px;padding:12px 0">'
        "No match found for that address or APN.</div>"
    )


@router.post("/ui/opportunities/wizard/step", response_class=HTMLResponse)
async def opportunity_wizard_step(
    request: Request,
    session: DBSession,
) -> HTMLResponse:
    form = await request.form()
    step = int(form.get("step", 1))
    opp_id_str = str(form.get("opp_id", "") or "")
    user = await _get_user(session, request)
    dedup_count, conflicts_count = await _get_counts(session)

    # Carry-through params from the create-from-deal flow.
    _link_to_deal = _safe_uuid_str(str(form.get("link_to_deal", "") or ""))
    _return_to = _safe_return_path(str(form.get("return_to", "") or ""))

    _deal_type_labels = {
        "acquisition": "Acquisition",
        "value_add": "Value-Add",
        "conversion": "Conversion",
        "new_construction": "New Construction",
    }

    if step == 1:
        name = str(form.get("name", "")).strip()
        deal_type = str(form.get("deal_type", "value_add"))
        notes = str(form.get("notes", "") or "").strip()

        if opp_id_str:
            try:
                opp = await session.get(Opportunity, UUID(opp_id_str))
            except ValueError:
                opp = None
        else:
            opp = None

        if opp is None:
            from app.models.org import Organization as _Org
            org = (await session.execute(select(_Org).limit(1))).scalar_one_or_none()
            if org is None:
                return HTMLResponse("No organization found", status_code=400)
            opp = Opportunity(
                org_id=org.id,
                name=name,
                notes=notes,
                source="manual",
                source_url="",
                promotion_source="manual",
                created_by_user_id=user.id if user else None,
            )
            session.add(opp)
        else:
            opp.name = name
            opp.notes = notes

        await session.commit()
        await session.refresh(opp)
        opp_id_str = str(opp.id)

        return templates.TemplateResponse(request, "opportunity_wizard.html", {
            "request": request, "step": 2, "opp": opp,
            "opp_id": opp_id_str,
            "deal_type": deal_type,
            "deal_type_label": _deal_type_labels.get(deal_type, deal_type),
            "link_to_deal": _link_to_deal,
            "return_to": _return_to,
            **_base_ctx(user, dedup_count, "opportunities", conflicts_count=conflicts_count),
        })

    elif step == 2:
        # Step 2: optional parcel/listing attachment.
        # Form sends attach_type ('listing'|'parcel'|'') and attach_id (UUID str).
        opp = await session.get(Opportunity, UUID(opp_id_str)) if opp_id_str else None
        if opp is None:
            return HTMLResponse("Opportunity not found", status_code=400)

        attach_type = str(form.get("attach_type", "") or "").strip()
        attach_id_raw = str(form.get("attach_id", "") or "").strip()

        if attach_type == "parcel" and attach_id_raw:
            try:
                opp.parcel_id = UUID(attach_id_raw)
            except ValueError:
                pass
        elif attach_type == "listing" and attach_id_raw:
            # Link via parcel if the matched listing already has one
            try:
                linked_opp = await session.get(Opportunity, UUID(attach_id_raw))
                if linked_opp and linked_opp.parcel_id:
                    opp.parcel_id = linked_opp.parcel_id
            except ValueError:
                pass
        # "skip" or empty = no attachment; proceed to review

        await session.commit()

        # Re-fetch with parcel eager-loaded so step 3 template can render
        # opp.parcel.address_normalized without an async lazy load.
        opp = (await session.execute(
            select(Opportunity)
            .where(Opportunity.id == UUID(opp_id_str))
            .options(selectinload(Opportunity.parcel))
        )).scalar_one()

        deal_type = str(form.get("deal_type", "value_add"))
        return templates.TemplateResponse(request, "opportunity_wizard.html", {
            "request": request, "step": 3, "opp": opp,
            "opp_id": opp_id_str,
            "deal_type": deal_type,
            "deal_type_label": _deal_type_labels.get(deal_type, deal_type),
            "link_to_deal": _link_to_deal,
            "return_to": _return_to,
            **_base_ctx(user, dedup_count, "opportunities", conflicts_count=conflicts_count),
        })

    return HTMLResponse("Invalid step", status_code=400)


@router.post("/ui/opportunities/wizard/complete")
async def opportunity_wizard_complete(
    request: Request,
    session: DBSession,
) -> Response:
    """Finalize opportunity creation — redirect to deal or opportunity detail.

    ``link_to_deal`` is no longer used to create a junction row (DealOpportunity
    was dropped in migration 0067). The opportunity is already linked via
    Scenario→Project→Opportunity. ``return_to`` controls the post-finalize
    landing URL — same-origin paths only; falls back to the opportunity detail.
    """
    form = await request.form()
    opp_id_str = str(form.get("opp_id", "") or "")
    if not opp_id_str:
        return HTMLResponse("Missing opp_id", status_code=400)

    return_to = _safe_return_path(str(form.get("return_to", "") or ""))
    target = return_to or f"/opportunities/{opp_id_str}"
    return RedirectResponse(url=target, status_code=303)


@router.get("/opportunities/{opp_id}", response_class=HTMLResponse)
async def opportunity_detail(
    request: Request,
    opp_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    user = await _get_user(session, request)
    dedup_count, conflicts_count = await _get_counts(session)
    opp = (await session.execute(
        select(Opportunity)
        .where(Opportunity.id == opp_id)
        .options(
            selectinload(Opportunity.parcel),
            selectinload(Opportunity.dev_projects).selectinload(Project.scenario),
        )
    )).scalar_one_or_none()
    if opp is None:
        return HTMLResponse("Not found", status_code=404)
    return templates.TemplateResponse(request, "opportunity_detail.html", {
        "request": request, "opp": opp,
        **_base_ctx(user, dedup_count, "opportunities", conflicts_count=conflicts_count),
    })


@router.post("/ui/opportunities/{opp_id}/archive")
async def archive_opportunity(
    opp_id: UUID,
    session: DBSession,
) -> RedirectResponse:
    """Archive a manually-created opportunity (sets archived=True, keeps all data)."""
    opp = await session.get(Opportunity, opp_id)
    if opp is None:
        return RedirectResponse("/opportunities", status_code=303)
    opp.archived = True
    opp.opp_status = OpportunityStatus.archived.value
    await session.commit()
    return RedirectResponse("/opportunities", status_code=303)



@router.get("/buildings", response_class=HTMLResponse)
async def buildings_page(
    request: Request,
    session: DBSession,
    q: str = Query(default=""),
    source: str = Query(default=""),
) -> HTMLResponse:
    """Building inventory page — removed. Physical attributes now live on Opportunity."""
    user = await _get_user(session, request)
    dedup_count, conflicts_count = await _get_counts(session)
    return RedirectResponse("/opportunities", status_code=302)


@router.get("/ui/buildings/rows", response_class=HTMLResponse)
async def buildings_rows(
    request: Request, session: DBSession,
    q: str = Query(default=""), source: str = Query(default=""),
) -> HTMLResponse:
    return HTMLResponse("<p class='text-muted'>Building inventory removed.</p>")


@router.get("/ui/buildings/{property_id}/detail", response_class=HTMLResponse)
async def building_detail(request: Request, property_id: UUID, session: DBSession) -> HTMLResponse:
    return HTMLResponse("<p class='text-muted'>Building entity removed.</p>", status_code=410)


# ---------------------------------------------------------------------------
# Parcels
# ---------------------------------------------------------------------------

def _extract_city(address_normalized: str | None) -> str | None:
    """Extract and title-case the city portion from a normalized address."""
    if not address_normalized:
        return None
    parts = address_normalized.split(",")
    if len(parts) >= 2:
        city_part = parts[1].strip().split()[0] if parts[1].strip() else None
        return city_part.title() if city_part else None
    return None


_STATE_CLASS_LABELS: dict[str, str] = {
    "101": "Single Family", "100": "Residential (Vacant)", "541": "Manufactured Home",
    "551": "Apartment / MF", "550": "Condo / MF",
    "201": "Commercial", "200": "Commercial (Vacant)",
    "401": "Industrial", "400": "Industrial (Vacant)",
    "701": "Farm", "700": "Farm (Vacant)", "301": "Timber", "300": "Timber (Vacant)", "303": "Timber",
    "801": "Exempt", "800": "Exempt (Vacant)", "641": "Utility", "640": "Utility (Vacant)",
    "601": "Mining", "600": "Mining (Vacant)", "000": "Unknown",
}


def _build_parcel_row(p: Parcel) -> dict:
    city = _extract_city(p.address_normalized)
    return {
        "id": str(p.id),
        "apn": p.apn,
        "address": p.address_normalized or p.address_raw or "",
        "street": (p.address_normalized or "").split(",")[0] if p.address_normalized else "",
        "city_state_zip": ", ".join(
            part.strip() for part in (p.address_normalized or "").split(",")[1:]
        ) if p.address_normalized else "",
        "address_city": city,
        "jurisdiction_mismatch": False,
        "zoning_code": p.zoning_code,
        "zoning_description": p.zoning_description,
        "lot_sqft": float(p.lot_sqft) if p.lot_sqft else None,
        "gis_acres": float(p.gis_acres) if p.gis_acres else None,
        "state_class": p.state_class,
        "state_class_label": _STATE_CLASS_LABELS.get(p.state_class or "", None),
        "total_assessed_value": float(p.total_assessed_value) if p.total_assessed_value else None,
        "assessed_value_land": float(p.assessed_value_land) if p.assessed_value_land else None,
        "assessed_value_improvements": float(p.assessed_value_improvements) if p.assessed_value_improvements else None,
        "year_built": p.year_built,
        "owner_name": p.owner_name,
        "owner_mailing_address": p.owner_mailing_address,
        "current_use": p.current_use,
        "county": p.county,
        "jurisdiction": p.jurisdiction,
        "priority_bucket": p.priority_bucket,
        "overridden_fields": [],
        "scraped_at_fmt": p.scraped_at.strftime("%b %-d, %Y") if p.scraped_at else None,
    }


_PARCEL_PAGE_SIZE = 500

# Oregon DOR state class → display label mapping (grouped for filter UI)
_STATE_CLASS_GROUPS: dict[str, tuple[str, list[str]]] = {
    "residential":  ("Residential (SFR)",     ["101", "100", "541"]),
    "multifamily":  ("Multi-Family / Apt",     ["551", "550"]),
    "commercial":   ("Commercial",             ["201", "200"]),
    "industrial":   ("Industrial",             ["401", "400"]),
    "farm":         ("Farm / Timber",          ["701", "700", "301", "300", "303"]),
    "exempt":       ("Exempt / Gov / Utility", ["801", "800", "641", "640", "601", "600"]),
}


def _parcel_jurisdiction_clause(jurisdictions: list[str]):
    """Build a SQL clause matching Parcel.jurisdiction, with county-scoped unincorporated.

    City values match ``jurisdiction`` directly. The special ``uninc:<County>`` prefix
    matches ``jurisdiction='unincorporated' AND county=<County>``, so the filter can
    distinguish unincorporated Multnomah from unincorporated Clackamas.
    """
    cities: list[str] = []
    uninc_counties: list[str] = []
    for j in jurisdictions:
        if j.startswith("uninc:"):
            uninc_counties.append(j[6:])
        else:
            cities.append(j)
    clauses = []
    if cities:
        clauses.append(Parcel.jurisdiction.in_(cities))
    if uninc_counties:
        clauses.append(
            and_(Parcel.jurisdiction == "unincorporated", Parcel.county.in_(uninc_counties))
        )
    if not clauses:
        return literal(False)
    return or_(*clauses) if len(clauses) > 1 else clauses[0]


async def _get_parcel_jurisdictions(session) -> list[dict]:
    """Return sorted {value, label} entries for the parcels jurisdiction filter.

    Splits the literal ``unincorporated`` jurisdiction into per-county rows
    (``uninc:Clackamas``, ``uninc:Multnomah``) so the UI can filter each
    unincorporated area separately.
    """
    rows = (await session.execute(
        select(Parcel.jurisdiction, Parcel.county, func.count())
        .where(Parcel.jurisdiction.isnot(None))
        .group_by(Parcel.jurisdiction, Parcel.county)
    )).all()

    city_totals: dict[str, int] = {}
    uninc_totals: dict[str, int] = {}
    for jurisdiction_name, county, cnt in rows:
        if jurisdiction_name == "unincorporated":
            key = (county or "Unknown").strip()
            uninc_totals[key] = uninc_totals.get(key, 0) + cnt
        else:
            city_totals[jurisdiction_name] = city_totals.get(jurisdiction_name, 0) + cnt

    out: list[dict] = [
        {"value": name, "label": f"{name.title()} ({cnt})"}
        for name, cnt in sorted(city_totals.items())
    ]
    for county, cnt in sorted(uninc_totals.items()):
        out.append({"value": f"uninc:{county}", "label": f"Unin. {county} ({cnt})"})
    return out


def _parcel_base_stmt(
    q: str, zoning: list[str], jurisdiction,
    use_group, min_acres: str, max_acres: str,
    min_year: str, max_year: str,
):
    stmt = select(Parcel).order_by(Parcel.apn)
    if q:
        # Also match apn_normalized with the punctuation-stripped query so searches
        # like "1S3E10AD -05800" / "1S3E10AD 05800" / "1S3E10AD05800" all resolve
        # to the same parcel regardless of how the stored APN is formatted.
        q_compact = normalize_apn(q)
        clauses = [Parcel.apn.ilike(f"%{q}%"), Parcel.address_normalized.ilike(f"%{q}%")]
        if q_compact:
            clauses.append(Parcel.apn_normalized.ilike(f"%{q_compact}%"))
        stmt = stmt.where(or_(*clauses))
    if zoning:
        stmt = stmt.where(Parcel.zoning_code.in_(zoning))
    jurs = _as_list(jurisdiction)
    if jurs:
        stmt = stmt.where(_parcel_jurisdiction_clause(jurs))
    use_groups = [g for g in _as_list(use_group) if g in _STATE_CLASS_GROUPS]
    if use_groups:
        codes: list[str] = []
        for g in use_groups:
            codes.extend(_STATE_CLASS_GROUPS[g][1])
        stmt = stmt.where(Parcel.state_class.in_(codes))
    if min_acres:
        try:
            stmt = stmt.where(Parcel.gis_acres >= float(min_acres))
        except ValueError:
            pass
    if max_acres:
        try:
            stmt = stmt.where(Parcel.gis_acres <= float(max_acres))
        except ValueError:
            pass
    if min_year:
        try:
            stmt = stmt.where(Parcel.year_built >= int(min_year))
        except ValueError:
            pass
    if max_year:
        try:
            stmt = stmt.where(Parcel.year_built <= int(max_year))
        except ValueError:
            pass
    return stmt


def _parcel_filter_ctx(
    q: str, zoning: list[str], jurisdiction,
    use_group, min_acres: str, max_acres: str,
    min_year: str, max_year: str,
) -> dict:
    return {
        "q": q, "zoning": zoning,
        "jurisdiction": _as_list(jurisdiction),
        "use_group": _as_list(use_group),
        "min_acres": min_acres, "max_acres": max_acres,
        "min_year": min_year, "max_year": max_year,
        "use_group_options": [(k, v[0]) for k, v in _STATE_CLASS_GROUPS.items()],
    }


@router.get("/parcels", response_class=HTMLResponse)
async def parcels_page(
    request: Request, session: DBSession,
    q: str = Query(default=""),
    zoning: list[str] = Query(default=[]),
    jurisdiction: list[str] = Query(default=[]),
    use_group: list[str] = Query(default=[]),
    min_acres: str = Query(default=""),
    max_acres: str = Query(default=""),
    min_year: str = Query(default=""),
    max_year: str = Query(default=""),
    offset: int = Query(default=0, ge=0),
) -> HTMLResponse:
    user = await _get_user(session, request)
    dedup_count, conflicts_count = await _get_counts(session)
    base = _parcel_base_stmt(q, zoning, jurisdiction, use_group, min_acres, max_acres, min_year, max_year)
    filtered_count, total = await asyncio.gather(
        session.execute(select(func.count()).select_from(base.subquery())),
        session.execute(select(func.count()).select_from(Parcel)),
    )
    filtered_count = int(filtered_count.scalar_one())
    total = int(total.scalar_one())
    parcels_list = list((await session.execute(base.offset(offset).limit(_PARCEL_PAGE_SIZE))).scalars())
    zoning_codes_stmt = select(Parcel.zoning_code).where(Parcel.zoning_code.isnot(None)).distinct().order_by(Parcel.zoning_code)
    if jurisdiction:
        zoning_codes_stmt = zoning_codes_stmt.where(_parcel_jurisdiction_clause(jurisdiction))
    zoning_codes_result = (await session.execute(zoning_codes_stmt)).all()
    zoning_codes = [r[0] for r in zoning_codes_result]
    jurisdictions = await _get_parcel_jurisdictions(session)
    parcels_data = [_build_parcel_row(p) for p in parcels_list]
    return templates.TemplateResponse(request, "parcels.html", {
        "parcels": parcels_data,
        "total_count": total,
        "filtered_count": filtered_count,
        "page_size": _PARCEL_PAGE_SIZE,
        "offset": offset,
        "zoning_codes": zoning_codes,
        "jurisdictions": jurisdictions,
        **_parcel_filter_ctx(q, zoning, jurisdiction, use_group, min_acres, max_acres, min_year, max_year),
        **_base_ctx(user, dedup_count, "parcels", conflicts_count=conflicts_count),
    })


@router.get("/ui/parcels/rows", response_class=HTMLResponse)
async def parcels_rows(
    request: Request, session: DBSession,
    q: str = Query(default=""),
    zoning: list[str] = Query(default=[]),
    jurisdiction: list[str] = Query(default=[]),
    use_group: list[str] = Query(default=[]),
    min_acres: str = Query(default=""),
    max_acres: str = Query(default=""),
    min_year: str = Query(default=""),
    max_year: str = Query(default=""),
    offset: int = Query(default=0, ge=0),
) -> HTMLResponse:
    base = _parcel_base_stmt(q, zoning, jurisdiction, use_group, min_acres, max_acres, min_year, max_year)
    filtered_count, total = await asyncio.gather(
        session.execute(select(func.count()).select_from(base.subquery())),
        session.execute(select(func.count()).select_from(Parcel)),
    )
    filtered_count = int(filtered_count.scalar_one())
    total = int(total.scalar_one())
    parcels_list = list((await session.execute(base.offset(offset).limit(_PARCEL_PAGE_SIZE))).scalars())
    parcels_data = [_build_parcel_row(p) for p in parcels_list]
    return templates.TemplateResponse(request, "partials/parcels_rows.html", {
        "parcels": parcels_data,
        "total_count": total,
        "filtered_count": filtered_count,
        "page_size": _PARCEL_PAGE_SIZE,
        "offset": offset,
        **_parcel_filter_ctx(q, zoning, jurisdiction, use_group, min_acres, max_acres, min_year, max_year),
    })


@router.get("/ui/parcels/{parcel_id}/detail", response_class=HTMLResponse)
async def parcel_detail(request: Request, parcel_id: UUID, session: DBSession) -> HTMLResponse:
    parcel = await session.get(Parcel, parcel_id)
    if parcel is None:
        return HTMLResponse("<p class='text-muted'>Not found.</p>")
    return templates.TemplateResponse(request, "partials/parcel_detail.html", {"p": _build_parcel_row(parcel)})


@router.post("/ui/parcels/{parcel_id}/gis-refresh", response_class=HTMLResponse)
async def parcel_gis_refresh(request: Request, parcel_id: UUID, session: DBSession) -> HTMLResponse:
    """Pull fresh GIS data for a parcel from the appropriate county source and re-render the detail panel."""
    from app.scrapers.parcel_enrichment import enrich_parcel

    parcel = await session.get(Parcel, parcel_id)
    if parcel is None:
        return HTMLResponse("<p class='text-muted'>Parcel not found.</p>", status_code=404)

    address = parcel.address_normalized or parcel.address_raw
    updated = await enrich_parcel(session, address=address, apn=parcel.apn)
    await session.commit()

    result = updated or parcel
    return templates.TemplateResponse(request, "partials/parcel_detail.html", {"p": _build_parcel_row(result)})


# ---------------------------------------------------------------------------
# Listings
# ---------------------------------------------------------------------------

def _build_listing_row(listing: ScrapedListing) -> dict:
    prop = getattr(listing, "_property", None)
    broker = listing.broker
    brokerage = broker.brokerage if broker else None

    # Jurisdiction display: prefer parcel-reconciled jurisdiction over city,
    # collapse the literal "unincorporated" / county-name-as-jurisdiction
    # cases into a friendly bucket label.
    _ej = (listing.jurisdiction or listing.city or "").strip()
    _county = (listing.county or "").strip()
    _bucket = _classify_listing_uninc_bucket(_ej.lower(), _county.lower())
    if _bucket == "uninc:Clackamas":
        jurisdiction_label = "Unin. Clackamas"
    elif _bucket == "uninc:Multnomah":
        jurisdiction_label = "Unin. Multnomah"
    elif _bucket == "uninc:other":
        jurisdiction_label = "Unincorporated"
    else:
        jurisdiction_label = _ej.title() if _ej else None

    return {
        "id": str(listing.id),
        "address": listing.address_normalized or listing.address_raw or "Undisclosed",
        "jurisdiction_label": jurisdiction_label,
        "is_new": listing.is_new,
        "source": listing.source,
        "source_label": listing.source.title(),
        "source_url": listing.source_url,
        "source_id": listing.source_id,
        "asking_price": float(listing.asking_price) if listing.asking_price else None,
        "price_per_unit": float(listing.price_per_unit) if listing.price_per_unit else None,
        "units": listing.units,
        "cap_rate": float(listing.cap_rate) if listing.cap_rate else None,
        "proforma_cap_rate": float(listing.proforma_cap_rate) if listing.proforma_cap_rate else None,
        "noi": float(listing.noi) if listing.noi else None,
        "proforma_noi": float(listing.proforma_noi) if listing.proforma_noi else None,
        "building_sqft": float(listing.gba_sqft) if listing.gba_sqft else None,
        "net_rentable_sqft": float(listing.net_rentable_sqft) if listing.net_rentable_sqft else None,
        "lot_sqft": float(listing.lot_sqft) if listing.lot_sqft else None,
        "year_built": listing.year_built,
        "property_type": listing.property_type,
        "status": listing.status,
        "description": listing.description,
        "buildings": listing.buildings,
        "stories": listing.stories,
        "parking_spaces": listing.parking_spaces,
        "class_": listing.class_,
        "zoning": listing.zoning,
        "apn": listing.apn,
        "occupancy_pct": float(listing.occupancy_pct) if listing.occupancy_pct else None,
        "year_renovated": listing.year_renovated,
        "price_per_sqft": float(listing.price_per_sqft) if listing.price_per_sqft else None,
        "broker_co_op": listing.broker_co_op,
        "broker_name": f"{broker.first_name or ''} {broker.last_name or ''}".strip() if broker else None,
        "brokerage_name": brokerage.name if brokerage else None,
        "broker_phone": broker.phone if broker else None,
        "broker_email": broker.email if broker else None,
        "property_id": str(prop.id) if prop else None,
        "first_seen_fmt": listing.first_seen_at.strftime("%b %-d, %Y") if listing.first_seen_at else None,
        "last_updated_fmt": listing.updated_at_source.strftime("%b %-d, %Y") if listing.updated_at_source else None,
        "last_checked_fmt": listing.last_seen_at.strftime("%b %-d, %Y") if listing.last_seen_at else None,
        "updated_highlight": listing.updated_at_source is not None,
        "raw_json": listing.raw_json,
        "archived": listing.archived,
        "linked_opportunity_id": str(listing.id) if listing.org_id else None,
        "linked_opportunity_name": listing.name or None,
        "linked_deal_id": None,  # Resolved separately when needed (avoid N+1 on list page)
        "priority_bucket": listing.priority_bucket,
    }


def _as_list(v) -> list[str]:
    """Normalize a filter value to a list of non-empty strings.

    Accepts either a single string (legacy single-select) or a list (multi-select).
    """
    if v is None:
        return []
    if isinstance(v, str):
        return [v] if v else []
    return [s for s in v if s]


def _listings_base_stmt(
    q: str,
    source,
    is_new: str,
    property_type=None,
    min_units: str = "",
    max_units: str = "",
    priority_bucket=None,
    cities: list[str] | None = None,
):
    stmt = (
        select(ScrapedListing)
        .options(
            selectinload(ScrapedListing.broker).selectinload(Broker.brokerage),
        )
        .order_by(ScrapedListing.last_seen_at.desc())
    )
    if q:
        stmt = stmt.where(or_(
            ScrapedListing.address_normalized.ilike(f"%{q}%"),
            ScrapedListing.address_raw.ilike(f"%{q}%"),
        ))
    sources = _as_list(source)
    if sources:
        stmt = stmt.where(ScrapedListing.source.in_(sources))
    if is_new == "1":
        stmt = stmt.where(ScrapedListing.is_new.is_(True))
    ptypes = _as_list(property_type)
    if ptypes:
        stmt = stmt.where(ScrapedListing.property_type.in_(ptypes))
    if min_units:
        try:
            n = int(min_units)
            if n > 0:
                stmt = stmt.where(ScrapedListing.units >= n)
        except ValueError:
            pass
    if max_units:
        try:
            stmt = stmt.where(ScrapedListing.units <= int(max_units))
        except ValueError:
            pass
    buckets = _as_list(priority_bucket)
    if buckets:
        stmt = stmt.where(ScrapedListing.priority_bucket.in_(buckets))
    if cities is not None:
        stmt = _apply_jurisdiction_filter(stmt, cities)
    return stmt


_LISTING_UNINC_TOKENS = {"uninc:Clackamas", "uninc:Multnomah", "uninc:other"}


def _classify_listing_uninc_bucket(ej_norm: str | None, county_norm: str | None) -> str | None:
    """Classify a listing into uninc:Clackamas / uninc:Multnomah / uninc:other or None.

    ej_norm: lowercased COALESCE(jurisdiction, city) string
    county_norm: lowercased county string

    Returns the bucket token, or None if the listing isn't unincorporated.
    """
    ej = (ej_norm or "").strip()
    cty = (county_norm or "").strip()
    is_unincorp_label = ej.startswith("unincorp")
    is_county_as_jur = ej in {"clackamas", "clackamas county", "multnomah", "multnomah county"}

    if not (is_unincorp_label or is_county_as_jur):
        return None

    # Pick a county hint from either ej or county.
    hints = " ".join([ej, cty])
    if "clackamas" in hints:
        return "uninc:Clackamas"
    if "multnomah" in hints:
        return "uninc:Multnomah"
    return "uninc:other"


def _apply_jurisdiction_filter(stmt, jurisdictions: list[str]):
    """Apply jurisdiction filter — cities and 'uninc:<bucket>' entries.

    Uses COALESCE(jurisdiction, city) so that parcel-reconciled listings
    filter by the authoritative GIS jurisdiction, while unreconciled
    listings fall back to the broker-provided city.

    Three unincorporated buckets are recognized:
      uninc:Clackamas, uninc:Multnomah — explicit unincorporated label or
        the county name standing in as the jurisdiction, with the county
        column matching.
      uninc:other — every other unincorporated row (other counties or
        rows with no county info).
    """
    effective_jurisdiction = func.coalesce(ScrapedListing.jurisdiction, ScrapedListing.city)
    ej_lower = func.lower(effective_jurisdiction)
    county_lower = func.lower(ScrapedListing.county)

    def _county_match(county: str):
        return county_lower.like(f"{county}%")

    def _is_uninc_label():
        return ej_lower.like("unincorp%")

    def _is_county_as_jur(county: str):
        return ej_lower.in_([county, f"{county} county"])

    def _is_clackamas_bucket():
        return or_(
            _is_county_as_jur("clackamas"),
            _is_uninc_label() & _county_match("clackamas"),
        )

    def _is_multnomah_bucket():
        return or_(
            _is_county_as_jur("multnomah"),
            _is_uninc_label() & _county_match("multnomah"),
        )

    def _is_other_uninc_bucket():
        return or_(
            _is_uninc_label() & ~(_county_match("clackamas") | _county_match("multnomah") | county_lower.is_(None)),
            _is_uninc_label() & county_lower.is_(None),
        )

    city_names = []
    selected_uninc: set[str] = set()
    for j in jurisdictions:
        if j in _LISTING_UNINC_TOKENS:
            selected_uninc.add(j)
        elif j.startswith("uninc:"):
            # Legacy uninc:<other-county> tokens — bucket as 'other'.
            selected_uninc.add("uninc:other")
        else:
            city_names.append(j)

    clauses = []
    if city_names:
        clauses.append(ej_lower.in_([c.lower() for c in city_names]))
    if "uninc:Clackamas" in selected_uninc:
        clauses.append(_is_clackamas_bucket())
    if "uninc:Multnomah" in selected_uninc:
        clauses.append(_is_multnomah_bucket())
    if "uninc:other" in selected_uninc:
        clauses.append(_is_other_uninc_bucket())
    if clauses:
        stmt = stmt.where(or_(*clauses))
    else:
        stmt = stmt.where(ScrapedListing.id.is_(None))
    return stmt


async def _get_jurisdictions(session) -> list[dict]:
    """Return sorted list of {value, label, type} for the listings jurisdiction filter.

    Cities are emitted as discrete rows. Anything that classifies as
    unincorporated (literal "unincorporated" jurisdiction, or a county name
    standing in as the jurisdiction, or a row with no jurisdiction but a
    county hint) is rolled up into one of three buckets:

      uninc:Clackamas, uninc:Multnomah, uninc:other

    so the user picks "Unin. Clackamas" without seeing a confusing raw
    "Clackamas" entry alongside actual cities.
    """
    effective_jurisdiction = func.coalesce(ScrapedListing.jurisdiction, ScrapedListing.city)

    rows = (await session.execute(
        select(effective_jurisdiction.label("ej"), ScrapedListing.county, func.count())
        .group_by(effective_jurisdiction, ScrapedListing.county)
    )).all()

    seen_cities: dict[str, tuple[str, int]] = {}
    uninc_totals: dict[str, int] = {"uninc:Clackamas": 0, "uninc:Multnomah": 0, "uninc:other": 0}

    for ej, county, cnt in rows:
        ej_norm = (ej or "").strip().lower()
        county_norm = (county or "").strip().lower()
        bucket = _classify_listing_uninc_bucket(ej_norm, county_norm)
        if bucket is not None:
            uninc_totals[bucket] += cnt
            continue
        if not ej:
            # No jurisdiction *or* city, and not classifiable as unincorp →
            # quietly bucket as 'other' so the row isn't lost.
            uninc_totals["uninc:other"] += cnt
            continue
        # Treat as a city. Dedup case variants (KLAMATH FALLS vs Klamath Falls).
        key = ej_norm
        if key in seen_cities:
            existing_label, existing_cnt = seen_cities[key]
            seen_cities[key] = (existing_label if existing_label[0].isupper() else ej.strip(), existing_cnt + cnt)
        else:
            seen_cities[key] = (ej.strip(), cnt)

    jurisdictions: list[dict] = [
        {"value": label, "label": f"{label} ({cnt})", "type": "city"}
        for _key, (label, cnt) in sorted(seen_cities.items())
    ]
    for bucket_label, bucket_value in [
        ("Unin. Clackamas", "uninc:Clackamas"),
        ("Unin. Multnomah", "uninc:Multnomah"),
        ("Unincorporated (other)", "uninc:other"),
    ]:
        if uninc_totals[bucket_value] > 0:
            jurisdictions.append({
                "value": bucket_value,
                "label": f"{bucket_label} ({uninc_totals[bucket_value]})",
                "type": "unincorporated",
            })
    return jurisdictions


def _split_listings(all_listings: list) -> tuple[list, list, list]:
    """Split into (new, promoted, archived) buckets. Promoted = org_id set."""
    new, promoted, archived = [], [], []
    for l in all_listings:
        if l.org_id:
            promoted.append(_build_listing_row(l))
        elif l.archived:
            archived.append(_build_listing_row(l))
        else:
            new.append(_build_listing_row(l))
    return new, promoted, archived


@router.get("/listings", response_class=HTMLResponse)
async def listings_page(request: Request) -> Response:
    """Listings page merged into Opportunities — redirect permanently."""
    return RedirectResponse(url="/opportunities", status_code=302)


@router.get("/ui/listings/rows", response_class=HTMLResponse)
async def listings_rows(
    request: Request, session: DBSession,
    q: str = Query(default=""),
    source: list[str] = Query(default=[]),
    property_type: list[str] = Query(default=[]),
    min_units: str = Query(default=""),
    max_units: str = Query(default=""),
    priority_bucket: list[str] = Query(default=[]),
    jurisdiction: list[str] = Query(default=[]),
) -> HTMLResponse:
    cities = jurisdiction if jurisdiction else None
    all_listings = list((await session.execute(
        _listings_base_stmt(q, source, "", property_type, min_units, max_units, priority_bucket, cities=cities)
    )).scalars())
    new_listings, promoted, archived = _split_listings(all_listings)
    return templates.TemplateResponse(request, "partials/listings_rows.html", {
        "new_listings": new_listings,
        "promoted_listings": promoted,
        "archived_listings": archived,
        "oob": True,
    })


@router.get("/ui/listings/export.csv")
async def listings_export_csv(
    session: DBSession,
    q: str = Query(default=""),
    source: list[str] = Query(default=[]),
    property_type: list[str] = Query(default=[]),
    min_units: str = Query(default=""),
    max_units: str = Query(default=""),
    priority_bucket: list[str] = Query(default=[]),
    jurisdiction: list[str] = Query(default=[]),
) -> StreamingResponse:
    """Export filtered listings as CSV (address, units, asking price, city, county, property type)."""
    cities = jurisdiction if jurisdiction else None
    all_listings = list((await session.execute(
        _listings_base_stmt(q, source, "", property_type, min_units, max_units, priority_bucket, cities=cities)
    )).scalars())

    import csv as _csv

    buf = io.StringIO()
    writer = _csv.writer(buf)
    writer.writerow(["Address", "City", "County", "Units", "Asking Price", "Property Type", "Cap Rate", "Year Built", "Source"])
    for l in all_listings:
        addr = l.address_normalized or l.address_raw or "Undisclosed"
        price = float(l.asking_price) if l.asking_price else ""
        cap = f"{float(l.cap_rate):.2f}%" if l.cap_rate else ""
        writer.writerow([addr, l.city or "", l.county or "", l.units or "", price, l.property_type or "", cap, l.year_built or "", l.source or ""])

    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=listings_export.csv"},
    )


@router.get("/ui/listings/promoted/rows", response_class=HTMLResponse)
async def listings_promoted_rows(
    request: Request, session: DBSession,
    q_promoted: str = Query(default=""),
    promoted_source: str = Query(default=""),
    promoted_property_type: str = Query(default=""),
    promoted_min_units: str = Query(default=""),
    promoted_max_units: str = Query(default=""),
) -> HTMLResponse:
    stmt = (
        select(ScrapedListing)
        .options(
            selectinload(ScrapedListing.broker).selectinload(Broker.brokerage),
        )
        .where(ScrapedListing.org_id.isnot(None))
        .order_by(ScrapedListing.last_seen_at.desc())
    )
    if q_promoted:
        stmt = stmt.where(or_(
            ScrapedListing.address_normalized.ilike(f"%{q_promoted}%"),
            ScrapedListing.address_raw.ilike(f"%{q_promoted}%"),
        ))
    if promoted_source:
        stmt = stmt.where(ScrapedListing.source == promoted_source)
    if promoted_property_type:
        stmt = stmt.where(ScrapedListing.property_type == promoted_property_type)
    if promoted_min_units:
        try:
            n = int(promoted_min_units)
            if n > 0:
                stmt = stmt.where(ScrapedListing.units >= n)
        except ValueError:
            pass
    if promoted_max_units:
        try:
            stmt = stmt.where(ScrapedListing.units <= int(promoted_max_units))
        except ValueError:
            pass
    promoted = [_build_listing_row(l) for l in (await session.execute(stmt)).scalars()]
    return templates.TemplateResponse(request, "partials/listings_promoted_rows.html", {
        "promoted_listings": promoted,
    })


@router.get("/ui/listings/{listing_id}/raw", response_class=PlainTextResponse)
async def listing_raw_json(listing_id: UUID, session: DBSession) -> PlainTextResponse:
    listing = await session.get(ScrapedListing, listing_id)
    if listing is None:
        return PlainTextResponse('{"error": "not found"}')
    data = listing.raw_json or {
        "id": str(listing.id),
        "source": listing.source,
        "source_id": listing.source_id,
        "address": listing.address_normalized,
        "asking_price": float(listing.asking_price) if listing.asking_price else None,
        "units": listing.units,
        "year_built": listing.year_built,
        "cap_rate": float(listing.cap_rate) if listing.cap_rate else None,
        "status": listing.status,
        "scraped_at": listing.last_seen_at.isoformat() if listing.last_seen_at else None,
    }
    return PlainTextResponse(json.dumps(data, indent=2, default=str))


@router.get("/ui/map/context")
async def map_context(
    session: DBSession,
    listing_id: UUID | None = Query(default=None),
    opportunity_id: UUID | None = Query(default=None),
    project_id: UUID | None = Query(default=None),
    parcel_id: UUID | None = Query(default=None),
    parcel_ids: str | None = Query(default=None),
) -> dict:
    """
    Resolve parcels + overlay GeoJSON for the map modal.
    Accepts one context param: listing_id, opportunity_id, project_id, or comma-sep parcel_ids.
    """
    from app.utils.gis import (
        combined_envelope, bbox_to_leaflet, envelope_str,
        esri_to_geojson, is_wgs84, detect_jurisdiction,
        fetch_overlay_features, compute_overlap_assessment,
        OVERLAY_REGISTRY,
    )

    parcels: list[Parcel] = []
    context_label = "Parcels"

    if listing_id:
        listing = await session.get(ScrapedListing, listing_id)
        if listing:
            context_label = listing.address_normalized or listing.address_raw or str(listing_id)
            # Resolve parcels for the listing. Try in order:
            #   1) exact apn match
            #   2) any apn_normalized token match (handles cross-source format
            #      drift like "1N1E03BD-09700" vs "1N1E03BD09700")
            #   3) reconciled parcel via ProjectParcel (when promoted to opp)
            apn_candidates: list[str] = []
            if listing.apn:
                apn_candidates.extend(
                    p.strip().upper() for p in listing.apn.replace(";", ",").split(",")
                    if p.strip()
                )
            if apn_candidates:
                hits = list((await session.execute(
                    select(Parcel).where(Parcel.apn.in_(apn_candidates))
                )).scalars())
                if hits:
                    parcels = hits
            if not parcels and listing.apn_normalized:
                tokens = [t for t in (listing.apn_normalized or []) if t]
                if tokens:
                    hits = list((await session.execute(
                        select(Parcel).where(Parcel.apn_normalized.in_(tokens))
                    )).scalars())
                    if hits:
                        parcels = hits
            if not parcels and listing.parcel_id:
                parcel = await session.get(Parcel, listing.parcel_id)
                if parcel:
                    parcels = [parcel]
            if not parcels:
                # Address-match fallback. APN coding differs across sources,
                # but many listings share an exact address with a parcel row.
                addr = (listing.address_normalized or "").strip()
                if addr:
                    hits = list((await session.execute(
                        select(Parcel).where(
                            func.lower(Parcel.address_normalized) == addr.lower()
                        )
                    )).scalars())
                    if hits:
                        parcels = hits
            # Fallback: listing lat/lng only — handled below via centroid

    elif opportunity_id:
        opp = await session.get(Opportunity, opportunity_id)
        if opp:
            context_label = opp.name or str(opportunity_id)
            if opp.parcel_id:
                parcel = await session.get(Parcel, opp.parcel_id)
                if parcel:
                    parcels = [parcel]

    elif project_id:
        proj = await session.get(Project, project_id)
        if proj:
            context_label = proj.name or str(project_id)
            if proj.opportunity_id:
                opp = await session.get(Opportunity, proj.opportunity_id)
                if opp:
                    context_label = f"{opp.name or ''} — {proj.name or ''}".strip(" —")
            # Parcel via Project.parcel_id or Opportunity.parcel_id
            if proj.parcel_id:
                parcel = await session.get(Parcel, proj.parcel_id)
                if parcel:
                    parcels = [parcel]
            elif proj.opportunity_id:
                opp = opp if proj.opportunity_id == (opp.id if opp else None) else await session.get(Opportunity, proj.opportunity_id)
                if opp and opp.parcel_id:
                    parcel = await session.get(Parcel, opp.parcel_id)
                    if parcel:
                        parcels = [parcel]

    elif parcel_id:
        result = await session.get(Parcel, parcel_id)
        if result:
            parcels = [result]

    elif parcel_ids:
        ids = [s.strip() for s in parcel_ids.replace(";", ",").split(",") if s.strip()]
        try:
            uuid_list = [UUID(i) for i in ids]
            parcels = list((await session.execute(
                select(Parcel).where(Parcel.id.in_(uuid_list))
            )).scalars())
        except ValueError:
            pass

    # --- Build parcel data ---
    parcel_data: list[dict] = []
    geometries: list[dict] = []  # raw ESRI geometries for combined envelope
    jurisdiction: str | None = None

    for parcel in parcels:
        geom_raw = parcel.geometry  # stored as ESRI rings dict
        geojson = esri_to_geojson(geom_raw) if geom_raw else None

        # Use stored jurisdiction first; fall back to address detection
        if not jurisdiction:
            jurisdiction = parcel.jurisdiction or detect_jurisdiction(
                parcel.address_normalized or parcel.address_raw,
                parcel.owner_city,
            )

        if geom_raw and is_wgs84(geom_raw):
            geometries.append(geom_raw)
        elif geojson:
            geometries.append(geom_raw)

        parcel_data.append({
            "id": str(parcel.id),
            "apn": parcel.apn,
            "address": parcel.address_normalized or parcel.address_raw or parcel.apn,
            "geojson": geojson,
            "lot_sqft": float(parcel.lot_sqft) if parcel.lot_sqft else None,
            "zoning": parcel.zoning_code,
            "has_geometry": geojson is not None,
        })

    # Lat/lng fallback from listing (no parcel geometry)
    centroid: list[float] | None = None
    if not geometries and listing_id:
        listing = await session.get(ScrapedListing, listing_id)
        if listing and listing.lat and listing.lng:
            centroid = [float(listing.lat), float(listing.lng)]

    # --- Compute envelope ---
    bbox: list[list[float]] | None = None
    overlay_data: dict = {}
    assessments: dict = {}

    if geometries:
        env_tuple = combined_envelope(geometries)
        env_str = envelope_str(env_tuple)
        bbox = bbox_to_leaflet(env_tuple)

        overlay_data = await fetch_overlay_features(env_str, jurisdiction=jurisdiction)

        # Overlap assessment per parcel × overlay
        for p_dict, parcel in zip(parcel_data, parcels):
            if not parcel.geometry:
                continue
            parcel_assessments: dict = {}
            for layer_key, layer_info in overlay_data.items():
                assessment = compute_overlap_assessment(
                    parcel.geometry,
                    layer_info.get("features") or [],
                    parcel_sqft=float(parcel.lot_sqft) if parcel.lot_sqft else None,
                )
                if assessment:
                    parcel_assessments[layer_key] = assessment
            if parcel_assessments:
                assessments[p_dict["id"]] = parcel_assessments

    elif centroid:
        # No geometry but we have a lat/lng — tiny envelope just to run overlays
        lat, lng = centroid[0], centroid[1]
        env_str = f"{lng-0.001},{lat-0.001},{lng+0.001},{lat+0.001}"
        bbox = [[lat - 0.002, lng - 0.002], [lat + 0.002, lng + 0.002]]
        overlay_data = await fetch_overlay_features(env_str, jurisdiction=jurisdiction)

    return {
        "parcels": parcel_data,
        "overlays": overlay_data,
        "assessments": assessments,
        "bbox": bbox,
        "centroid": centroid,
        "context_label": context_label,
        "jurisdiction": jurisdiction,
    }


@router.get("/ui/listings/{listing_id}/detail", response_class=HTMLResponse)
async def listing_detail(request: Request, listing_id: UUID, session: DBSession) -> HTMLResponse:
    listing = await session.get(
        ScrapedListing, listing_id,
        options=[
            selectinload(ScrapedListing.broker).selectinload(Broker.brokerage),
        ]
    )
    if listing is None:
        return HTMLResponse("<p class='text-muted'>Not found.</p>")
    l = _build_listing_row(listing)
    # Resolve linked deal — listing IS the opportunity; find deal via Scenario→Project path
    if listing.org_id:
        deal_row = (await session.execute(
            select(Deal.id)
            .join(Scenario, Scenario.deal_id == Deal.id)
            .join(Project, Project.scenario_id == Scenario.id)
            .where(Project.opportunity_id == listing.id)
            .limit(1)
        )).scalar_one_or_none()
        if deal_row:
            l["linked_deal_id"] = str(deal_row)
    return templates.TemplateResponse(request, "partials/listing_detail.html", {"l": l})


@router.post("/ui/listings/{listing_id}/promote", response_class=HTMLResponse)
async def promote_listing(
    request: Request,
    listing_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Manually promote a listing to an Opportunity (set org_id on listing row)."""
    from app.tasks.scraper import _promote_listing as _do_promote, _get_default_org_id  # local import avoids circular

    listing = await session.get(
        ScrapedListing, listing_id,
        options=[selectinload(ScrapedListing.broker).selectinload(Broker.brokerage)]
    )
    if listing is None:
        return HTMLResponse("<span class='text-muted text-small'>Not found</span>")

    if listing.org_id:
        # Already promoted
        l = _build_listing_row(listing)
        return templates.TemplateResponse(request, "partials/listings_promoted_row.html", {"l": l})

    org_id = await _get_default_org_id(session)
    opp = await _do_promote(
        listing, session,
        promotion_source="manual",
        ruleset_id=None,
        org_id=org_id,
    )
    await session.commit()

    if opp is None:
        return HTMLResponse("<span class='text-muted text-small'>Promotion failed</span>")

    await session.refresh(listing)
    l = _build_listing_row(listing)
    return templates.TemplateResponse(request, "partials/listings_promoted_row.html", {"l": l})


@router.post("/ui/listings/{listing_id}/promote-redirect")
async def promote_listing_redirect(
    listing_id: UUID,
    session: DBSession,
) -> RedirectResponse:
    """Promote listing to Opportunity (or reuse existing), then redirect to opportunity detail."""
    from app.tasks.scraper import _promote_listing as _do_promote, _get_default_org_id

    listing = await session.get(ScrapedListing, listing_id)
    if listing is None:
        return RedirectResponse("/listings", status_code=303)

    if listing.org_id:
        return RedirectResponse(f"/opportunities/{listing.id}", status_code=303)

    org_id = await _get_default_org_id(session)
    opp = await _do_promote(listing, session, promotion_source="manual", ruleset_id=None, org_id=org_id)
    await session.commit()

    if opp is None:
        return RedirectResponse("/listings", status_code=303)

    return RedirectResponse(f"/opportunities/{opp.id}", status_code=303)


@router.post("/ui/listings/{listing_id}/revert", response_class=HTMLResponse)
async def revert_listing(
    request: Request,
    listing_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Revert a promoted listing back to unpromoted: archives the Opportunity and clears the link."""
    listing = await session.get(
        ScrapedListing, listing_id,
        options=[selectinload(ScrapedListing.broker).selectinload(Broker.brokerage)]
    )
    if listing is None:
        return HTMLResponse("<span class='text-muted text-small'>Not found</span>")

    if listing.org_id:
        # Demote: clear org_id and set opp_status to archived
        listing.org_id = None
        listing.opp_status = None
        await session.commit()

    # Reload and return as a New row (revert = back to New, not archived)
    listing_reloaded = await session.get(
        ScrapedListing, listing_id,
        options=[selectinload(ScrapedListing.broker).selectinload(Broker.brokerage)]
    )
    if listing_reloaded:
        listing_reloaded.is_new = True
        listing_reloaded.archived = False
        await session.commit()
        await session.refresh(listing_reloaded)
    l = _build_listing_row(listing_reloaded)
    return templates.TemplateResponse(request, "partials/listings_new_row.html", {"l": l})


@router.post("/ui/listings/{listing_id}/archive", response_class=HTMLResponse)
async def archive_listing(
    request: Request,
    listing_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Move a listing from New to Archived."""
    listing = await session.get(
        ScrapedListing, listing_id,
        options=[selectinload(ScrapedListing.broker).selectinload(Broker.brokerage)]
    )
    if listing is None:
        return HTMLResponse("")
    listing.archived = True
    listing.is_new = False
    await session.commit()
    await session.refresh(listing)
    l = _build_listing_row(listing)
    return templates.TemplateResponse(request, "partials/listings_archived_row.html", {"l": l})


@router.post("/ui/listings/{listing_id}/unarchive", response_class=HTMLResponse)
async def unarchive_listing(
    request: Request,
    listing_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Restore an archived listing back to New."""
    listing = await session.get(
        ScrapedListing, listing_id,
        options=[selectinload(ScrapedListing.broker).selectinload(Broker.brokerage)]
    )
    if listing is None:
        return HTMLResponse("")
    listing.archived = False
    listing.is_new = True
    await session.commit()
    await session.refresh(listing)
    l = _build_listing_row(listing)
    return templates.TemplateResponse(request, "partials/listings_new_row.html", {"l": l})


# ---------------------------------------------------------------------------
# Brokers
# ---------------------------------------------------------------------------

def _build_broker_row(broker: Broker, listing_count: int) -> dict:
    bg = broker.brokerage
    return {
        "id": str(broker.id),
        "full_name": f"{broker.first_name or ''} {broker.last_name or ''}".strip() or "Unknown",
        "brokerage_name": bg.name if bg else None,
        "brokerage_status": (bg.firm_scrape_status if bg else None) or "unknown",
        "email": broker.email,
        "phone": broker.phone,
        "license_number": broker.license_number,
        "license_state": broker.license_state,
        "is_platinum": broker.is_platinum,
        "number_of_assets": broker.number_of_assets,
        "listing_count": listing_count,
    }


def _join_address(*parts: str | None) -> str | None:
    """Join non-empty address parts with separators suitable for inline display."""
    cleaned = [str(p).strip() for p in parts if p]
    if not cleaned:
        return None
    # street, street2, city, state, zip → "street, street2, city, state zip"
    if len(cleaned) >= 4:
        head = ", ".join(cleaned[:-2])
        tail = " ".join(cleaned[-2:])
        return f"{head}, {tail}"
    return ", ".join(cleaned)


def _build_broker_detail(broker: Broker, listings: list[ScrapedListing]) -> dict:
    row = _build_broker_row(broker, len(listings))
    bg = broker.brokerage
    row.update(
        {
            "license_number_locked": bool(broker.license_number_locked),
            "license_personal_address": _join_address(
                broker.license_personal_street,
                broker.license_personal_street2,
                broker.license_personal_city,
                broker.license_personal_state,
                broker.license_personal_zip,
            ),
            "license_type": broker.license_type,
            "license_status": broker.license_status or "unknown",
            "oregon_last_pulled_at": broker.oregon_last_pulled_at.isoformat()
            if broker.oregon_last_pulled_at else None,
            "oregon_lookup_status": broker.oregon_lookup_status,
            "oregon_failure_count": int(broker.oregon_failure_count or 0),
            "oregon_detail_url": broker.oregon_detail_url,
            "brokerage_id": str(bg.id) if bg else None,
            "firm_scrape_status": (bg.firm_scrape_status if bg else None) or "unknown",
            "firm_scrape_domain": bg.firm_scrape_domain if bg else None,
            "oregon_company_name": bg.oregon_company_name if bg else None,
            "oregon_company_address": _join_address(
                bg.oregon_company_street,
                bg.oregon_company_street2,
                bg.oregon_company_city,
                bg.oregon_company_state,
                bg.oregon_company_zip,
            ) if bg else None,
            "disciplinary_actions": [
                {
                    "case_number": d.case_number,
                    "order_signed_date": d.order_signed_date.isoformat()
                    if d.order_signed_date else None,
                    "resolution": d.resolution,
                    "found_issues": d.found_issues,
                }
                for d in (broker.disciplinary_actions or [])
            ],
        }
    )
    row["listings"] = [
        {
            "address": l.address_normalized or l.address_raw or "Unknown",
            "source": l.source,
            "asking_price": float(l.asking_price) if l.asking_price else None,
        }
        for l in listings
    ]
    return row


def _broker_stmt(q: str = "", company: str = "", listings_op: str = "", listings_val: str = ""):
    stmt = (
        select(Broker)
        .options(selectinload(Broker.brokerage), selectinload(Broker.scraped_listings))
        .order_by(Broker.last_name, Broker.first_name)
    )
    needs_brokerage_join = bool(q or company)
    if needs_brokerage_join:
        stmt = stmt.outerjoin(Broker.brokerage)
    if q:
        stmt = stmt.where(or_(
            Broker.first_name.ilike(f"%{q}%"),
            Broker.last_name.ilike(f"%{q}%"),
            (Broker.first_name + " " + Broker.last_name).ilike(f"%{q}%"),
            Brokerage.name.ilike(f"%{q}%"),
        ))
    if company:
        stmt = stmt.where(Brokerage.name.ilike(f"%{company}%"))
    return stmt


def _apply_listings_filter(brokers_list: list, listings_op: str, listings_val: str) -> list:
    if not listings_op or not listings_val:
        return brokers_list
    try:
        val = int(listings_val)
    except (ValueError, TypeError):
        return brokers_list
    if listings_op == "gte":
        return [b for b in brokers_list if len(b.scraped_listings) >= val]
    if listings_op == "lte":
        return [b for b in brokers_list if len(b.scraped_listings) <= val]
    if listings_op == "eq":
        return [b for b in brokers_list if len(b.scraped_listings) == val]
    return brokers_list


@router.get("/brokers", response_class=HTMLResponse)
async def brokers_page(
    request: Request, session: DBSession,
    q: str = Query(default=""),
    company: str = Query(default=""),
    listings_op: str = Query(default=""),
    listings_val: str = Query(default=""),
) -> HTMLResponse:
    user = await _get_user(session, request)
    dedup_count, conflicts_count = await _get_counts(session)
    stmt = _broker_stmt(q, company, listings_op, listings_val)
    brokers_list = list((await session.execute(stmt)).scalars().unique())
    brokers_list = _apply_listings_filter(brokers_list, listings_op, listings_val)
    total = int((await session.execute(select(func.count()).select_from(Broker))).scalar_one())
    brokers_data = [_build_broker_row(b, len(b.scraped_listings)) for b in brokers_list]
    return templates.TemplateResponse(request, "brokers.html", {
        "brokers": brokers_data, "total_count": total,
        "q": q, "company": company, "listings_op": listings_op, "listings_val": listings_val,
        **_base_ctx(user, dedup_count, "brokers", conflicts_count=conflicts_count),
    })


@router.get("/ui/brokers/rows", response_class=HTMLResponse)
async def brokers_rows(
    request: Request, session: DBSession,
    q: str = Query(default=""),
    company: str = Query(default=""),
    listings_op: str = Query(default=""),
    listings_val: str = Query(default=""),
) -> HTMLResponse:
    stmt = _broker_stmt(q, company, listings_op, listings_val)
    brokers_list = list((await session.execute(stmt)).scalars().unique())
    brokers_list = _apply_listings_filter(brokers_list, listings_op, listings_val)
    brokers_data = [_build_broker_row(b, len(b.scraped_listings)) for b in brokers_list]
    return templates.TemplateResponse(request, "partials/brokers_rows.html", {"brokers": brokers_data})


@router.get("/ui/brokers/{broker_id}/detail", response_class=HTMLResponse)
async def broker_detail(request: Request, broker_id: UUID, session: DBSession) -> HTMLResponse:
    broker = await session.get(
        Broker, broker_id,
        options=[
            selectinload(Broker.brokerage),
            selectinload(Broker.scraped_listings),
            selectinload(Broker.disciplinary_actions),
        ],
    )
    if broker is None:
        return HTMLResponse("<p class='text-muted'>Not found.</p>")
    b = _build_broker_detail(broker, broker.scraped_listings)
    return templates.TemplateResponse(request, "partials/broker_detail.html", {"b": b})


@router.post("/ui/brokers/{broker_id}/license", response_class=HTMLResponse)
async def broker_license_update(
    request: Request,
    broker_id: UUID,
    session: DBSession,
    license_number: str = Form(default=""),
    license_state: str = Form(default=""),
) -> HTMLResponse:
    """Manually set a broker's license number. Sets license_number_locked=True
    so listing scrapers won't overwrite it. The user does this specifically to
    align the license with the Oregon database; subsequent Oregon enrichment
    runs against this value."""
    broker = await session.get(
        Broker, broker_id,
        options=[
            selectinload(Broker.brokerage),
            selectinload(Broker.scraped_listings),
            selectinload(Broker.disciplinary_actions),
        ],
    )
    if broker is None:
        return HTMLResponse("<p class='text-muted'>Not found.</p>")
    cleaned_number = (license_number or "").strip() or None
    cleaned_state = (license_state or "").strip().upper() or None
    broker.license_number = cleaned_number
    broker.license_state = cleaned_state
    broker.license_number_locked = cleaned_number is not None
    await session.commit()
    await session.refresh(broker)
    b = _build_broker_detail(broker, broker.scraped_listings)
    resp = templates.TemplateResponse(request, "partials/broker_detail.html", {"b": b})
    resp.headers["HX-Trigger"] = "brokerSaved"
    return resp


@router.post("/ui/brokers/{broker_id}/oregon-update", response_class=HTMLResponse)
async def broker_oregon_update(
    request: Request, broker_id: UUID, session: DBSession,
) -> HTMLResponse:
    """Queue a one-shot Oregon eLicense enrichment for a single broker. Does
    not affect license_number_locked — manual lock and Oregon enrichment are
    independent (lock blocks listing-source scrapers, not Oregon)."""
    broker = await session.get(
        Broker, broker_id,
        options=[
            selectinload(Broker.brokerage),
            selectinload(Broker.scraped_listings),
            selectinload(Broker.disciplinary_actions),
        ],
    )
    if broker is None:
        return HTMLResponse("<p class='text-muted'>Not found.</p>")
    broker.oregon_lookup_status = "pending"
    await session.commit()
    # Local import keeps Celery out of the router import path in unit tests
    # that don't load celery_app.
    from app.tasks.oregon_elicense import enrich_broker_oregon  # noqa: PLC0415

    enrich_broker_oregon.delay(str(broker_id))
    await session.refresh(broker)
    b = _build_broker_detail(broker, broker.scraped_listings)
    return templates.TemplateResponse(request, "partials/broker_detail.html", {"b": b})


# ---------------------------------------------------------------------------
# Model Builder
# ---------------------------------------------------------------------------

def _sum_amount(rows: list) -> float | None:
    if not rows:
        return None
    total = sum(float(r.amount) for r in rows)
    return total if total else None


def _sum_annual(rows: list, field: str = "annual_amount") -> float | None:
    if not rows:
        return None
    total = sum(float(getattr(r, field, 0) or 0) for r in rows)
    return total if total else None


def _income_annual(streams: list) -> float | None:
    """Effective gross annual revenue at stabilization — applies stabilized_occupancy_pct."""
    if not streams:
        return None
    total = 0.0
    for s in streams:
        occupancy = float(s.stabilized_occupancy_pct or 100) / 100.0
        if s.amount_per_unit_monthly and s.unit_count:
            total += float(s.amount_per_unit_monthly) * int(s.unit_count) * occupancy * 12
        elif s.amount_fixed_monthly:
            total += float(s.amount_fixed_monthly) * occupancy * 12
    return total if total else None


def _capital_total(modules: list, junction_amts: dict[str, float] | None = None) -> float | None:
    """Sum module principals. If junction_amts is provided (project-scoped
    view), use the per-project junction amount instead of scenario-level
    source.amount so multi-project Sources totals reflect this project's
    share only."""
    total = 0.0
    for m in modules:
        if m.source and isinstance(m.source, dict):
            if m.source.get("is_bridge"):
                continue
            if junction_amts is not None:
                amt = junction_amts.get(str(m.id), 0.0)
            else:
                amt = m.source.get("amount")
            if amt:
                total += float(amt)
    return total if total else None


# ---------------------------------------------------------------------------
# Builder form helpers
# ---------------------------------------------------------------------------

from decimal import Decimal

_ITEM_TYPE_TO_MODULE: dict[str, str] = {
    "use-lines": "sources_uses",
    "income-streams": "revenue",
    "expense-lines": "opex",
    "capital-modules": "sources_uses",
    "waterfall-tiers": "owners_profit",
    "milestones": "timeline",
    "unit-mix": "property",
}


def _parse_vehicle_carry_schedule(form) -> dict | None:
    """Parse carry schedule arrays from a vehicle settings form into carry_config dict."""
    labels = form.getlist("v_carry_phase_label[]")
    types = form.getlist("v_carry_phase_type[]")
    dur_types = form.getlist("v_carry_phase_duration_type[]")
    months_vals = form.getlist("v_carry_phase_months[]")
    milestones = form.getlist("v_carry_phase_milestone_key[]")
    rates = form.getlist("v_carry_phase_rate_pct[]")
    amorts = form.getlist("v_carry_phase_amort_years[]")
    if not types:
        return None
    phases = []
    for i, ct in enumerate(types):
        ct = ct.strip()
        if not ct or ct == "none":
            continue
        dur_type = (dur_types[i] if i < len(dur_types) else "remainder").strip()
        if dur_type == "months":
            try:
                n = int(months_vals[i]) if i < len(months_vals) and months_vals[i].strip() else 0
            except ValueError:
                n = 0
            dur: dict = {"type": "months", "months": n}
        elif dur_type == "milestone":
            mk = (milestones[i] if i < len(milestones) else "").strip()
            dur = {"type": "milestone", "milestone_key": mk}
        else:
            dur = {"type": "remainder"}
        p: dict = {
            "label": (labels[i] if i < len(labels) else "").strip() or ct,
            "carry_type": ct,
            "duration": dur,
        }
        try:
            if i < len(rates) and rates[i].strip():
                p["rate_pct"] = float(rates[i].strip())
        except ValueError:
            pass
        try:
            if i < len(amorts) and amorts[i].strip():
                p["amort_term_years"] = int(amorts[i].strip())
        except ValueError:
            pass
        phases.append(p)
    return {"schedule": phases} if phases else None


def _fd(v: str | None) -> Decimal | None:
    """Parse an optional Decimal from a form field. Strips commas tolerantly."""
    if not v or not v.strip():
        return None
    try:
        return Decimal(v.strip().replace(",", ""))
    except Exception:
        return None


def _fi(v: str | None, default: int = 0) -> int:
    """Parse an optional int from a form field."""
    if not v or not v.strip():
        return default
    try:
        return int(v.strip())
    except Exception:
        return default


def _fp(v: str | None, default: list[str] | None = None) -> list[str]:
    """Parse phases from a comma-separated or JSON-array form field."""
    if not v or not v.strip():
        return default or []
    v = v.strip()
    if v.startswith("["):
        try:
            return json.loads(v)
        except Exception:
            pass
    return [p.strip() for p in v.split(",") if p.strip()]


def _builder_gantt_from_milestones(project: "Project | None", milestones: list) -> "dict | None":
    """Build Gantt v2 data from pre-loaded milestones for the model builder timeline panel."""
    if not project or not milestones:
        return None
    bars, epoch, has_dates = _extract_milestone_bars(project, milestones=milestones)
    if not bars:
        return None
    # Apply the same stabilized cap used by the full gantt
    raw_rows = [{"project_name": project.name, "bars": bars}]
    _override_stabilized_cap(raw_rows)
    g_min = min(b["display_start_day"] for b in bars)
    g_max = max(b["display_start_day"] + b["display_duration_days"] for b in bars)
    _gantt_apply_pct(bars, g_min, g_max)
    bars.sort(key=lambda b: b["display_start_day"])  # chronological row order
    month_ticks, year_spans = _compute_gantt_axis(epoch, g_min, g_max, has_dates)
    return {
        "has_dates": has_dates,
        "epoch": epoch,           # exposed for source bar positioning
        "g_min": g_min,
        "g_max": g_max,
        "month_ticks": month_ticks,
        "year_spans": year_spans,
        "rows": _bars_to_phase_rows(bars),
    }


async def _per_project_capital_modules_ui(
    session: AsyncSession, scenario_id: UUID, project_id: UUID
) -> list:
    """UI-side equivalent of the engine's ``_per_project_capital_modules``
    — loads CapitalModule rows attached to this project via the junction.
    Duplicated here to avoid cross-package import between UI and engine.
    """
    from app.models.capital import CapitalModuleProject
    result = await session.execute(
        select(CapitalModule)
        .join(
            CapitalModuleProject,
            CapitalModuleProject.capital_module_id == CapitalModule.id,
        )
        .where(
            CapitalModule.scenario_id == scenario_id,
            CapitalModuleProject.project_id == project_id,
        )
        .order_by(CapitalModule.stack_position)
    )
    return list(result.scalars())


async def _lightweight_project_status_data(
    session: AsyncSession, scenario_id: UUID, project_id: UUID
) -> dict:
    """Minimal per-project data for ``_compute_calc_status``.

    Cheaper than re-running the full ``_load_builder_data`` per project.
    Fetches only the five keys ``_compute_calc_status`` reads: outputs,
    inputs, capital_modules, capital_total, uses_total.
    """
    outputs = (
        await session.execute(
            select(OperationalOutputs).where(
                OperationalOutputs.scenario_id == scenario_id,
                OperationalOutputs.project_id == project_id,
            )
        )
    ).scalar_one_or_none()
    inputs = (
        await session.execute(
            select(OperationalInputs).where(
                OperationalInputs.project_id == project_id
            )
        )
    ).scalar_one_or_none()
    capital_modules = await _per_project_capital_modules_ui(
        session, scenario_id, project_id
    )
    # capital_total must use this project's junction-scoped amount, not the
    # module's source.amount (which holds the scenario-wide last-sized value
    # for shared modules — e.g. P2's $4.3M would show on P1's pill,
    # producing a phantom $1.45M Sources/Uses surplus on P1).
    from app.models.capital import CapitalModuleProject as _CMP_lite
    _junction_by_module = {
        str(mid): float(amt or 0)
        for mid, amt in (
            await session.execute(
                select(_CMP_lite.capital_module_id, _CMP_lite.amount).where(
                    _CMP_lite.project_id == project_id
                )
            )
        ).all()
    }
    capital_total = 0.0
    for cm in capital_modules:
        _jam = _junction_by_module.get(str(cm.id))
        if _jam is not None:
            capital_total += _jam
            continue
        src = cm.source or {}
        amt = src.get("amount")
        if amt is None:
            continue
        try:
            capital_total += float(amt)
        except (TypeError, ValueError):
            continue
    use_lines = list(
        (
            await session.execute(
                select(UseLine).where(UseLine.project_id == project_id)
            )
        ).scalars()
    )
    uses_total = 0.0
    for ul in use_lines:
        if ul.is_deferred:
            continue
        if ul.amount is None:
            continue
        try:
            uses_total += float(ul.amount)
        except (TypeError, ValueError):
            continue
    return {
        "outputs": outputs,
        "inputs": inputs,
        "capital_modules": capital_modules,
        "capital_total": capital_total,
        "uses_total": uses_total,
    }


# Severity ranking for rolling worst-status up from per-project to Underwriting.
_STATUS_RANK = {"ok": 0, "na": 0, "warn": 1, "fail": 2}


async def _compute_scenario_statuses(
    session: AsyncSession, scenario_id: UUID
) -> dict:
    """Per-project status dicts + Underwriting aggregate for tab-chip pills.

    Each per-project status matches the shape ``_compute_calc_status``
    returns (sources_uses / dscr / ltv + overall + failing_count). The
    Underwriting aggregate picks the worst severity across projects and
    sums their failing counts. No cross-project soft rules yet — combined
    DSCR / LTV portfolio checks are informational-only per the Phase 2f
    product decision; a future Phase 3b1 can add dedicated Underwriting-
    only rule rows if needed.
    """
    project_ids = [
        row[0]
        for row in (
            await session.execute(
                select(Project.id)
                .where(Project.scenario_id == scenario_id)
                .order_by(Project.created_at.asc())
            )
        )
    ]

    per_project: dict = {}
    worst_overall = "ok"
    worst_project_id = None
    total_failing = 0
    for pid in project_ids:
        data = await _lightweight_project_status_data(session, scenario_id, pid)
        status = _compute_calc_status(data)
        per_project[pid] = status
        total_failing += int(status.get("failing_count", 0) or 0)
        if _STATUS_RANK.get(status["overall"], 0) > _STATUS_RANK.get(
            worst_overall, 0
        ):
            worst_overall = status["overall"]
            worst_project_id = pid

    return {
        "per_project": per_project,
        "underwriting": {
            "overall": worst_overall,
            "failing_count": total_failing,
            "worst_project_id": worst_project_id,
        },
    }


async def _staleness_map(session: AsyncSession, scenario_id: UUID) -> dict:
    """Compute per-project + scenario-level staleness for tab chips.

    A project is stale iff any of its inputs (project-scoped or scenario-
    scoped) was updated after its ``OperationalOutputs.computed_at``. A
    never-computed project is stale by definition.

    Scenario-level input updates (edits to CapitalModule, WaterfallTier,
    CapitalModuleProject, ProjectAnchor) mark EVERY project on the scenario
    stale, since those tables are shared.

    Returns::

        {
            "per_project": {project_id: bool, ...},
            "underwriting": bool,   # True iff any project is stale
        }

    Migration 0052 added ``updated_at`` to the input tables; for fresh
    installs the initial value is server-default ``now()``, so existing
    deals won't light up as stale until a real edit happens.
    """
    from app.models.capital import CapitalModuleProject as _CMP
    from app.models.deal import Scenario, IncomeStream as _IS, OperatingExpenseLine as _OEL, OperationalInputs as _OI, UseLine as _UL
    from app.models.project import ProjectAnchor as _PA

    # ── Scenario-scoped max(updated_at) — applies to every project on this scenario.
    async def _scalar_max(stmt):
        return (await session.execute(stmt)).scalar_one_or_none()

    scenario_maxes = [
        await _scalar_max(
            select(func.max(CapitalModule.updated_at)).where(
                CapitalModule.scenario_id == scenario_id
            )
        ),
        await _scalar_max(
            select(func.max(WaterfallTier.updated_at)).where(
                WaterfallTier.scenario_id == scenario_id
            )
        ),
        await _scalar_max(
            select(func.max(_CMP.updated_at))
            .join(CapitalModule, CapitalModule.id == _CMP.capital_module_id)
            .where(CapitalModule.scenario_id == scenario_id)
        ),
        await _scalar_max(
            select(func.max(_PA.updated_at))
            .join(Project, Project.id == _PA.project_id)
            .where(Project.scenario_id == scenario_id)
        ),
    ]
    scenario_max = max((t for t in scenario_maxes if t is not None), default=None)

    # Per-project max(updated_at) across project-scoped input tables.
    project_ids = [
        row[0]
        for row in (
            await session.execute(
                select(Project.id).where(Project.scenario_id == scenario_id)
            )
        )
    ]

    per_project: dict = {}
    for pid in project_ids:
        proj_maxes = [
            await _scalar_max(
                select(func.max(_UL.updated_at)).where(_UL.project_id == pid)
            ),
            await _scalar_max(
                select(func.max(_IS.updated_at)).where(_IS.project_id == pid)
            ),
            await _scalar_max(
                select(func.max(_OEL.updated_at)).where(_OEL.project_id == pid)
            ),
            await _scalar_max(
                select(func.max(_OI.updated_at)).where(_OI.project_id == pid)
            ),
                await _scalar_max(
                select(func.max(Milestone.updated_at)).where(Milestone.project_id == pid)
            ),
        ]
        proj_max = max((t for t in proj_maxes if t is not None), default=None)

        computed_at = (
            await session.execute(
                select(OperationalOutputs.computed_at).where(
                    OperationalOutputs.scenario_id == scenario_id,
                    OperationalOutputs.project_id == pid,
                )
            )
        ).scalar_one_or_none()

        if computed_at is None:
            per_project[pid] = True
            continue

        max_input = max(
            (t for t in (proj_max, scenario_max) if t is not None), default=None
        )
        per_project[pid] = bool(max_input is not None and max_input > computed_at)

    return {
        "per_project": per_project,
        "underwriting": any(per_project.values()) if per_project else False,
    }


def _wizard_mode_from_request(request: Request) -> bool:
    """Read the single-flow deal-creation wizard flag off the parent page.

    HTMX form posts in the wizard chrome don't carry the `wizard=1` query
    param themselves, but the browser's current URL (sent in HX-Current-URL)
    does. Used by panel-render endpoints so the timeline approve button text
    and sidebar visibility stay consistent across in-wizard mutations.
    """
    _hx_url = request.headers.get("HX-Current-URL", "")
    if not _hx_url:
        return False
    from urllib.parse import urlparse, parse_qs
    return parse_qs(urlparse(_hx_url).query).get("wizard", [""])[0] == "1"


async def _active_project_from_request(
    request: Request, session: AsyncSession, model_id: UUID,
) -> UUID | None:
    """Extract the active project_id from HX-Current-URL's `?project=` query param.

    Multi-project deals: form submits don't carry the active project, so the
    route must look at the user's current URL (sent by HTMX in the
    HX-Current-URL header) to know which Project tab is active. Returns None
    if the header is missing, the param is invalid, or the project doesn't
    belong to this scenario.
    """
    _hx_url = request.headers.get("HX-Current-URL", "")
    if not _hx_url:
        return None
    from urllib.parse import urlparse, parse_qs
    _qs_proj = parse_qs(urlparse(_hx_url).query).get("project", [""])[0]
    if not _qs_proj:
        return None
    try:
        _candidate = UUID(_qs_proj)
    except ValueError:
        return None
    _candidate_proj = await session.get(Project, _candidate)
    if _candidate_proj and _candidate_proj.scenario_id == model_id:
        return _candidate_proj.id
    return None


async def _load_builder_data(session: AsyncSession, model_id: UUID, project_id: UUID | None = None) -> dict:
    """Load all line-item data for the model builder page/panel.

    model_id = Deal.id.  Line items (use_lines, income_streams, expense_lines,
    operational_inputs) belong to the active Project for this Deal.
    Capital modules and waterfall tiers belong to the Deal directly.

    project_id: if provided, load data for that specific Project; else default to first.
    """
    # Load the scenario (DealModel) to access income_mode and deal_id
    _scenario = await session.get(DealModel, model_id)

    # Resolve active Project for this Scenario
    default_project = None
    if project_id is not None:
        candidate = await session.get(Project, project_id)
        if candidate and candidate.scenario_id == model_id:
            default_project = candidate
    if default_project is None:
        default_project = (await session.execute(
            select(Project).where(Project.scenario_id == model_id).order_by(Project.created_at.asc()).limit(1)
        )).scalar_one_or_none()
    project_id = default_project.id if default_project else None
    timeline_approved = default_project.timeline_approved if default_project else False

    inputs = None
    use_lines: list = []
    income_streams: list = []
    expense_lines: list = []
    unit_mix_rows: list = []

    if project_id is not None:
        inputs = (await session.execute(
            select(OperationalInputs).where(OperationalInputs.project_id == project_id)
        )).scalar_one_or_none()

        use_lines = list((await session.execute(
            select(UseLine).where(UseLine.project_id == project_id).order_by(UseLine.phase, UseLine.label)
        )).scalars())

        income_streams = list((await session.execute(
            select(IncomeStream).where(IncomeStream.project_id == project_id).order_by(IncomeStream.label)
        )).scalars())

        expense_lines = list((await session.execute(
            select(OperatingExpenseLine).where(OperatingExpenseLine.project_id == project_id).order_by(OperatingExpenseLine.label)
        )).scalars())

        # unit_mix is JSONB on Project; wrap dicts for attribute-compatible access
        if default_project is not None and _ensure_unit_mix_ids(default_project):
            await session.flush()
        unit_mix_rows = [_UMRow(r) for r in (default_project.unit_mix or [])] if default_project else []

    # Scope outputs to the active project. Phase 2b's migration 0051
    # swapped UNIQUE(scenario_id) for UNIQUE(scenario_id, project_id), so
    # a scenario may now carry N rows (one per project). scalar_one_or_none
    # without project_id raised MultipleResultsFound on multi-project deals.
    _outputs_q = select(OperationalOutputs).where(
        OperationalOutputs.scenario_id == model_id
    )
    if project_id is not None:
        _outputs_q = _outputs_q.where(OperationalOutputs.project_id == project_id)
    outputs = (await session.execute(_outputs_q.limit(1))).scalar_one_or_none()

    # Carrying annual = avg monthly debt service in stabilized/operations phase × 12.
    # None = never computed; 0.0 = computed but no debt service.
    # Multi-project: filter CashFlow by active project_id so the bottom
    # "Est. Annual Debt Service" reflects this project's service only — not
    # a scenario-wide average that mixes projects together.
    from app.models.cashflow import CashFlow as _CashFlow, PeriodType as _PT

    def _cf_scope(q):
        q = q.where(_CashFlow.scenario_id == model_id)
        if project_id is not None:
            q = q.where(_CashFlow.project_id == project_id)
        return q

    _cf_count = (await session.execute(
        _cf_scope(select(func.count()).select_from(_CashFlow))
    )).scalar_one()
    if _cf_count:
        # Prefer stabilized phase; fall back to any operation phase; last resort = any period
        _ops_avg = (await session.execute(
            _cf_scope(select(func.avg(_CashFlow.debt_service)))
            .where(_CashFlow.period_type == _PT.stabilized)
        )).scalar_one_or_none()
        if _ops_avg is None:
            _ops_avg = (await session.execute(
                _cf_scope(select(func.avg(_CashFlow.debt_service)))
                .where(_CashFlow.period_type.in_([_PT.lease_up, _PT.stabilized]))
            )).scalar_one_or_none()
        if _ops_avg is None:
            _ops_avg = (await session.execute(
                _cf_scope(select(func.avg(_CashFlow.debt_service)))
            )).scalar_one_or_none()
        carrying_annual_computed: float | None = float(_ops_avg) * 12 if _ops_avg else 0.0

        # First stabilized period NCF → first-month and first-year profit metrics
        _stab_rows = list((await session.execute(
            _cf_scope(select(_CashFlow.net_cash_flow))
            .where(_CashFlow.period_type == _PT.stabilized)
            .order_by(_CashFlow.period)
        )).scalars())
        stabilized_month1_ncf: float | None = float(_stab_rows[0]) if _stab_rows else None
        stabilized_year1_ncf: float | None = float(sum(_stab_rows[:12])) if _stab_rows else None
    else:
        carrying_annual_computed = None  # not yet computed
        stabilized_month1_ncf = None
        stabilized_year1_ncf = None

    capital_modules = list((await session.execute(
        select(CapitalModule).where(CapitalModule.scenario_id == model_id).order_by(CapitalModule.stack_position)
    )).scalars())

    # Per-module-per-phase annual debt service for the carrying costs table rows.
    # carrying_detail[module_id_str][phase_name] = annual_amount (float)

    def _annual_carry_amt(source: dict, carry_type: str) -> float:
        amount = source.get("amount")
        rate_pct = source.get("interest_rate_pct")
        if not amount or not rate_pct:
            return 0.0
        principal = float(amount)
        rate = float(rate_pct)
        if carry_type in ("io_only", "interest_reserve"):
            # True IO or pre-funded IR: annual interest cost = principal × rate
            # (IR: reserve pays it; io_only: borrower pays it — same carrying display)
            return principal * rate / 100.0
        elif carry_type == "capitalized_interest":
            # No periodic cash outflow — interest accrues to balance, paid at payoff
            return 0.0
        elif carry_type == "pi":
            r = rate / 100.0 / 12.0
            n = int(source.get("amort_term_years") or 30) * 12
            if r == 0:
                return principal / (n / 12)
            factor = (1 + r) ** n
            return (principal * r * factor / (factor - 1)) * 12
        return 0.0

    # Multi-project: per-row estimate should reflect THIS project's share of
    # the module principal (from the CapitalModuleProject junction), not the
    # scenario-level source.amount (which after auto-size holds the last
    # project's sized value). Fall back to source.amount only when no project
    # is selected (e.g. underwriting rollup view).
    _junction_amts: dict[str, float] = {}
    if project_id is not None:
        from app.models.capital import CapitalModuleProject as _CMP
        _jrows = list((await session.execute(
            select(_CMP.capital_module_id, _CMP.amount).where(_CMP.project_id == project_id)
        )).all())
        _junction_amts = {str(mid): float(amt or 0) for mid, amt in _jrows}

    carrying_detail: dict[str, dict[str, float]] = {}
    for _cm in capital_modules:
        if str(getattr(_cm, "vehicle_type", "") or "").replace("VehicleType.", "") != "debt":
            continue
        _src = dict(_cm.source or {})
        _carry = _cm.carry or {}
        _mid = str(_cm.id)
        if project_id is not None:
            # Override amount with this project's junction share
            _src["amount"] = _junction_amts.get(_mid, 0.0)
        if _carry.get("phases"):
            carrying_detail[_mid] = {
                p.get("name", ""): _annual_carry_amt(_src, p.get("carry_type", "none"))
                for p in _carry["phases"]
            }
        else:
            _ct = _carry.get("carry_type", "none")
            _amt = _annual_carry_amt(_src, _ct)
            carrying_detail[_mid] = {"construction": _amt, "operation": _amt}

    waterfall_tiers = list((await session.execute(
        select(WaterfallTier).where(WaterfallTier.scenario_id == model_id).order_by(WaterfallTier.priority)
    )).scalars())

    # Milestones for the default dev Project
    milestones: list = []
    if project_id is not None:
        milestones = list((await session.execute(
            select(Milestone)
            .where(Milestone.project_id == project_id)
        )).scalars())

    # Build milestone map for trigger-chain resolution
    _PHASE_ORDER = [
        "offer_made", "under_contract", "close", "pre_development",
        "construction", "operation_lease_up", "operation_stabilized", "divestment",
    ]
    ms_map = {m.id: m for m in milestones}

    # Auto-cap operation_stabilized at 30 years when no divestment milestone exists
    _STABILIZED_AUTO_DAYS = 10950
    _has_divestment = any(
        str(m.milestone_type).replace("MilestoneType.", "") == "divestment"
        for m in milestones
    )
    if not _has_divestment:
        for _m in milestones:
            if (
                str(_m.milestone_type).replace("MilestoneType.", "") == "operation_stabilized"
                and (_m.duration_days or 0) == 0
            ):
                _m.duration_days = _STABILIZED_AUTO_DAYS
                session.add(_m)

    def _phase_idx(m):
        raw = str(m.milestone_type).replace("MilestoneType.", "")
        return next((i for i, v in enumerate(_PHASE_ORDER) if v == raw), 99)

    def _ms_sort_key(m):
        start = m.computed_start(ms_map)
        return (start is None, start or 0, _phase_idx(m))

    milestones = sorted(milestones, key=_ms_sort_key)

    exit_lines = [u for u in use_lines if getattr(u.phase, "value", str(u.phase)) == "exit"]
    deferred_uses = [u for u in use_lines if getattr(u, "is_deferred", False)]
    deferred_total = sum(float(u.amount or 0) for u in deferred_uses)
    revenue_annual = _income_annual(income_streams)
    opex_annual = _sum_annual(expense_lines, "annual_amount")
    _capex_per_unit = float(inputs.capex_reserve_per_unit_annual or 0) if inputs else 0.0
    _total_units_for_reserve = sum((u.unit_count or 0) for u in unit_mix_rows)
    capex_reserve_annual = _capex_per_unit * _total_units_for_reserve
    opex_total_annual = (opex_annual or 0) + capex_reserve_annual
    # Multi-project Sources totals use per-project junction amounts so the
    # panel total doesn't show the scenario-wide (last-sized) principal.
    _cap_junction_amts: dict[str, float] = {}
    if project_id is not None:
        from app.models.capital import CapitalModuleProject as _CMP_sum
        _cap_junction_amts = {
            str(mid): float(amt or 0)
            for mid, amt in (await session.execute(
                select(_CMP_sum.capital_module_id, _CMP_sum.amount).where(
                    _CMP_sum.project_id == project_id
                )
            )).all()
        }
    capital_total = _capital_total(
        capital_modules, junction_amts=_cap_junction_amts if project_id is not None else None
    )
    uses_total_val = sum(float(u.amount or 0) for u in use_lines)

    # Equity ownership — computed from equity-type capital modules
    equity_modules = [
        m for m in capital_modules
        if str(getattr(m, "vehicle_type", "") or "").replace("VehicleType.", "") == "equity"
    ]
    _total_equity = sum(
        float((m.source or {}).get("amount", 0) or 0) for m in equity_modules
    )
    equity_ownership = [
        {
            "module": m,
            "amount": float((m.source or {}).get("amount", 0) or 0),
            "pct": (float((m.source or {}).get("amount", 0) or 0) / _total_equity * 100)
                   if _total_equity > 0 else 0.0,
        }
        for m in equity_modules
    ]
    # If no equity partners defined, synthesize a 100% org-owner row
    org_owner_fallback = not equity_ownership

    # Load org name for fallback display
    org_name = "Sponsor"
    try:
        from app.models.org import Organization as _Org
        _scenario_for_org = await session.get(DealModel, model_id)
        if _scenario_for_org:
            _deal_for_org = await session.get(Deal, _scenario_for_org.deal_id)
            if _deal_for_org and _deal_for_org.org_id:
                _org = await session.get(_Org, _deal_for_org.org_id)
                if _org:
                    org_name = _org.name
    except Exception:
        pass

    # ── Phase summaries ──────────────────────────────────────────────────────
    # Four logical phases built from milestone types.
    from app.models.milestone import MilestoneType as MT
    _PRE_DEV   = {MT.offer_made, MT.under_contract, MT.close, MT.pre_development}
    _CONSTRUCT = {MT.construction}
    _OPERATION = {MT.operation_lease_up, MT.operation_stabilized}
    _DIVEST    = {MT.divestment}

    def _phase_bucket(types: set) -> list:
        return [m for m in milestones if MT(m.milestone_type) in types]

    def _bucket_summary(bucket: list) -> dict:
        starts = [m.computed_start(ms_map) for m in bucket]
        starts = [s for s in starts if s]
        ends = [m.computed_end(ms_map) for m in bucket]
        ends = [e for e in ends if e]
        if not starts:
            return {"start": None, "end": None, "duration_days": None}
        start = min(starts)
        end = max(ends) if ends else None
        duration_days = (end - start).days if end and start else None
        return {"start": start, "end": end, "duration_days": duration_days}

    pre_dev_bucket  = _phase_bucket(_PRE_DEV)
    construct_bucket = _phase_bucket(_CONSTRUCT)
    operation_bucket = _phase_bucket(_OPERATION)
    divest_bucket    = _phase_bucket(_DIVEST)

    phase_summaries = {
        "pre_dev":      _bucket_summary(pre_dev_bucket),
        "construction": _bucket_summary(construct_bucket),
        "operation":    _bucket_summary(operation_bucket),
        "divestment":   _bucket_summary(divest_bucket),
        "has_divestment": bool(divest_bucket),
    }

    # Total timeline: earliest computed start → latest computed end
    _all_starts = [m.computed_start(ms_map) for m in milestones]
    _all_ends   = [m.computed_end(ms_map)   for m in milestones]
    _all_starts = [s for s in _all_starts if s]
    _all_ends   = [e for e in _all_ends   if e]
    if _all_starts and _all_ends:
        total_timeline_days = (max(_all_ends) - min(_all_starts)).days
    else:
        total_timeline_days = sum(m.duration_days for m in milestones)

    # ── Capital module source bars for Sources & Uses Gantt ─────────────────
    # Each source bar spans active_phase_start → active_phase_end. If no end
    # phase is set the bar extends to the right edge with a fade-out (perpetuity
    # convention for equity / permanent debt). Zero-amount modules are hidden
    # unless explicitly auto-sized (placeholder dashed bar).
    #
    # Phase → (milestone_key, side) mapping:
    # - side="end": the phase begins when that milestone *completes* (e.g.
    #   "acquisition" phase starts when the Close milestone ends — money
    #   changes hands at the end of the closing process).
    # - side="start": the phase begins when that milestone *starts* (e.g.
    #   "construction" phase starts at Construction milestone start).
    # Phase → ordered list of candidate (milestone_key, side) tuples.  The
    # first candidate whose milestone is actually on this deal's timeline is
    # used.  This lets a single phase like "construction" resolve against
    # whichever work-phase milestone the project uses: `construction` for new
    # construction, `minor_renovation` / `major_renovation` for acq-reno
    # deals, `conversion` for conversions, etc.
    # Each candidate list walks from the ideal match to progressively looser
    # fallbacks, so a bar still resolves when the exact milestone isn't on
    # this deal's timeline.  Example: an acq-minor-reno deal may have no
    # `construction` or `minor_renovation` milestone — "construction" phase
    # then falls back to `close.end` (construction starts right after close).
    _CM_PHASE_TO_MS: dict[str, list[tuple[str, str]]] = {
        "acquisition":          [("close", "end"), ("under_contract", "end"), ("offer_made", "end")],
        "pre_development":      [("pre_development", "start"), ("close", "end")],
        "pre_construction":     [("pre_development", "start"), ("close", "end")],
        "construction":         [
            ("construction", "start"),
            ("minor_renovation", "start"),
            ("major_renovation", "start"),
            ("renovation", "start"),
            ("conversion", "start"),
            ("close", "end"),  # fallback: construction begins right after close
        ],
        "lease_up":             [
            ("operation_lease_up", "start"),
            ("construction", "end"),
            ("minor_renovation", "end"),
            ("major_renovation", "end"),
            ("close", "end"),
        ],
        "operation_lease_up":   [("operation_lease_up", "start")],
        "stabilized":           [
            ("operation_stabilized", "start"),
            ("operation_lease_up", "end"),
        ],
        "operation_stabilized": [("operation_stabilized", "start")],
        "exit":                 [("divestment", "start"), ("operation_stabilized", "end")],
        "divestment":           [("divestment", "start"), ("operation_stabilized", "end")],
        # Legacy value written by the pre-cleanup wizard — treat as "runs
        # through divestment" so the bar renders to the far right.
        "perpetuity":           [("divestment", "start"), ("operation_stabilized", "end")],
    }
    _cm_gantt_rows: list[dict] = []
    _bgd_cm = _builder_gantt_from_milestones(default_project, milestones)
    if _bgd_cm and capital_modules:
        _epoch_cm = _bgd_cm.get("epoch")
        _g_min_cm = _bgd_cm.get("g_min", 0)
        _g_max_cm = _bgd_cm.get("g_max", 1)
        _span_cm = max(_g_max_cm - _g_min_cm, 1)
        _ms_start_map: dict = {
            str(m.milestone_type).replace("MilestoneType.", ""): m.computed_start(ms_map)
            for m in milestones
            if m.computed_start(ms_map)
        }
        _ms_end_map: dict = {
            str(m.milestone_type).replace("MilestoneType.", ""): m.computed_end(ms_map)
            for m in milestones
            if m.computed_end(ms_map)
        }

        def _phase_to_date(phase: str | None) -> date | None:
            if not phase:
                return None
            candidates = _CM_PHASE_TO_MS.get(phase)
            if not candidates:
                return None
            for ms_key, side in candidates:
                source_map = _ms_end_map if side == "end" else _ms_start_map
                found = source_map.get(ms_key)
                if found:
                    return found
            return None

        # Derive end-phase string from Exit Vehicle when possible — matches
        # the engine's `_resolve_active_end_rank` logic.  Falls back to the
        # legacy `active_phase_end` field when vehicle is unset.
        def _gantt_end_phase(_cm: object) -> str | None:
            vt = str(getattr(_cm, "vehicle_type", "") or "").replace("VehicleType.", "")
            if vt != "debt":
                # Non-debt: runs through the end of the hold (display-wise)
                return "divestment"
            et = getattr(_cm, "exit_terms", None) or {}
            saved = (et.get("vehicle") or "").strip() if isinstance(et, dict) else ""
            if saved == "sale":
                return "divestment"
            if saved == "maturity":
                return "divestment"  # display convention: through end of deal
            if saved:
                for r in capital_modules:
                    if r is _cm:
                        continue
                    if str(getattr(r, "id", "")) == saved:
                        return str(r.active_phase_start or "")
            # Legacy fallback.
            legacy = str(getattr(_cm, "active_phase_end", "") or "")
            return legacy or None

        for _cm in capital_modules:
            _src = _cm.source or {}
            _src_amount = float(_src.get("amount") or 0)
            _auto_size = bool(_src.get("auto_size"))
            # Hide $0 sources that aren't marked for auto-sizing (zero means
            # the user hasn't committed this source, so it shouldn't take up
            # a Gantt row).
            if _src_amount <= 0 and not _auto_size:
                continue
            if not _cm.active_phase_start or not _epoch_cm:
                continue
            _from_date = _phase_to_date(_cm.active_phase_start)
            if not _from_date:
                continue
            _from_day = (_from_date - _epoch_cm).days
            _left = max(0.0, round(100.0 * (_from_day - _g_min_cm) / _span_cm, 2))
            _end_phase = _gantt_end_phase(_cm)
            _to_date = _phase_to_date(_end_phase) if _end_phase else None
            if _to_date:
                _to_day = (_to_date - _epoch_cm).days
                _right = min(100.0, round(100.0 * (_to_day - _g_min_cm) / _span_cm, 2))
                _width = max(_right - _left, 1.5)
                _fade = False
            else:
                _width = max(100.0 - _left, 1.5)
                _fade = True
            _vt = str(getattr(_cm, "vehicle_type", "") or "").replace("VehicleType.", "")
            _label = (_cm.label or _vt).replace(" (auto)", "").strip()
            _cm_gantt_rows.append({
                "label": _label,
                "source_type": _vt if _vt in ("equity", "debt", "grant", "forgivable_loan") else "debt",
                "vehicle_type": _vt,
                "left_pct": _left,
                "width_pct": _width,
                "fade_right": _fade,
                "unsized": _src_amount <= 0,
            })

    # AMI rent tier lookup — only computed when affordable_housing_project is enabled.
    ami_unit_data: dict = {}
    if inputs and getattr(inputs, "affordable_housing_project", False):
        from app.data.ami_portland_2025 import get_ami_tier as _get_ami_tier
        _stream_by_label = {s.label: s for s in income_streams}
        for _u in unit_mix_rows:
            _sl = (
                f"{_u.label} Rent (Renovated)"
                if getattr(_u, "unit_strategy", None) == "value_add_renovation"
                else f"{_u.label} Rent"
            )
            _s = _stream_by_label.get(_sl)
            _proposed: float | None = None
            if _s and _s.amount_per_unit_monthly:
                _proposed = float(_s.amount_per_unit_monthly)
            elif _u.market_rent_per_unit:
                _proposed = float(_u.market_rent_per_unit)
            if _proposed is not None:
                _beds = int(_u.beds or 1)
                ami_unit_data[str(_u.id)] = _get_ami_tier(_beds, _proposed)

    return {
        "inputs": inputs,
        "outputs": outputs,
        "use_lines": use_lines,
        "income_streams": income_streams,
        "expense_lines": expense_lines,
        "unit_mix_rows": unit_mix_rows,
        "capital_modules": capital_modules,
        # Per-project junction-scoped principal by module id (str). Template
        # uses this in the Sources table so Project N's tab shows Project N's
        # share, not the scenario-wide last-sized amount.
        "capital_junction_amts": _cap_junction_amts,
        "waterfall_tiers": waterfall_tiers,
        "milestones": milestones,
        "milestone_rows": [
            {
                "ms": m,
                "start": m.computed_start(ms_map),
                "end": m.computed_end(ms_map),
            }
            for m in milestones
        ],
        "use_line_count": len(use_lines),
        "income_stream_count": len(income_streams),
        "expense_line_count": len(expense_lines),
        "unit_mix_count": len(unit_mix_rows),
        "total_units": sum((u.unit_count or 0) for u in unit_mix_rows),
        "capital_module_count": len(capital_modules),
        "waterfall_tier_count": len(waterfall_tiers),
        "capital_total": capital_total,
        "uses_total": uses_total_val,
        "revenue_annual": revenue_annual,
        "opex_annual": opex_annual,
        "capex_reserve_annual": capex_reserve_annual,
        "opex_total_annual": opex_total_annual,
        "carrying_annual": carrying_annual_computed,
        "carrying_detail": carrying_detail,
        "stabilized_month1_ncf": stabilized_month1_ncf,
        "stabilized_year1_ncf": stabilized_year1_ncf,
        "divestment_total": _sum_amount(exit_lines),
        "profit_total": float(outputs.noi_stabilized) if outputs and outputs.noi_stabilized else None,
        "equity_ownership": equity_ownership,
        "org_owner_fallback": org_owner_fallback,
        "org_name": org_name,
        "deferred_uses": deferred_uses,
        "deferred_total": deferred_total,
        "total_timeline_days": total_timeline_days,
        "total_timeline_months": round(total_timeline_days / 30.4) if total_timeline_days else 0,
        "phase_summaries": phase_summaries,
        "timeline_approved": timeline_approved,
        "deal_setup_complete": bool(getattr(inputs, "deal_setup_complete", False)) if inputs else False,
        "default_project_id": project_id,
        # Staleness map for top tab chips (Phase 3a): each project + the
        # Underwriting rollup get a dot when inputs have moved past the
        # last compute. Computed once per _load_builder_data call.
        "staleness": await _staleness_map(session, model_id),
        # Phase 3b: per-project calc-status + Underwriting aggregate for the
        # tab-chip pills. Each status dict follows the _compute_calc_status
        # shape; the Underwriting entry is the worst severity across projects.
        "scenario_statuses": await _compute_scenario_statuses(session, model_id),
        # Approval gate: every milestone needs a position AND a non-zero duration
        "timeline_approvable": len(milestones) > 0 and all(
            (m.target_date or m.trigger_milestone_id) and (m.duration_days or 0) > 0
            for m in milestones
        ),
        "timeline_missing_position": [
            m for m in milestones
            if not m.target_date and not m.trigger_milestone_id
        ],
        "timeline_missing_duration": [
            m for m in milestones
            if not (m.duration_days or 0) > 0
        ],
        # Gantt data for the timeline module panel (Gantt v2)
        "builder_gantt_data": _builder_gantt_from_milestones(default_project, milestones),
        "capital_module_gantt_rows": _cm_gantt_rows,
        # Wizard: show when unapproved and no milestone has a start date yet
        "wizard_needed": (not timeline_approved) and (not any(m.target_date for m in milestones)),
        "wizard_default_types": list(DEFAULT_DURATIONS.get(
            _scenario.project_type if _scenario else "", {}
        ).keys()),
        "wizard_deal_type": _scenario.project_type if _scenario else "",
        "wizard_deal_type_label": {
            "acquisition": "Acquisition",
            "value_add": "Value-Add",
            "conversion": "Conversion",
            "new_construction": "New Construction",
        }.get(_scenario.project_type if _scenario else "", "Project"),
        "income_mode": (_scenario.income_mode if _scenario else "revenue_opex") or "revenue_opex",
        "noi_annual": float(inputs.noi_stabilized_input) if inputs and inputs.noi_stabilized_input is not None else None,
        "ami_unit_data": ami_unit_data,
    }


@router.post("/ui/forms/{model_id}/{item_type}", response_class=HTMLResponse)
@router.put("/ui/forms/{model_id}/{item_type}/{item_id}", response_class=HTMLResponse)
async def handle_form_create_or_update(
    request: Request,
    model_id: UUID,
    item_type: str,
    session: DBSession,
    item_id: str = "",
) -> HTMLResponse:
    """Accept form-encoded data, persist the mutation, return refreshed panel HTML."""
    model = await session.get(DealModel, model_id)
    if model is None:
        return HTMLResponse("<p class='text-muted'>Model not found.</p>", status_code=404)

    # Resolve active Project for line items that belong to Project level.
    # Multi-project deals: prefer the project from the form's active URL
    # (HX-Current-URL header carries it), fall back to the oldest project.
    default_project = (await session.execute(
        select(Project).where(Project.scenario_id == model_id).order_by(Project.created_at.asc()).limit(1)
    )).scalar_one_or_none()
    project_id = await _active_project_from_request(request, session, model_id)
    if project_id is None:
        project_id = default_project.id if default_project else None

    form = await request.form()
    module = _ITEM_TYPE_TO_MODULE.get(item_type, "uses")

    if item_type == "use-lines":
        _ms_key = form.get("milestone_key") or None
        _ms_key_to = form.get("milestone_key_to") or None
        # If "to" == "from" or blank, treat as single-point (no range)
        if _ms_key_to == _ms_key:
            _ms_key_to = None
        # Derive phase from milestone_key for backward compat with anything that reads phase
        _milestone_to_phase = {
            "close": "acquisition", "pre_development": "pre_construction",
            "construction": "construction", "renovation": "construction",
            "conversion": "construction", "operation_lease_up": "operation",
            "operation_stabilized": "operation", "divestment": "exit",
            "maturity": "other",
        }
        _phase = _milestone_to_phase.get(_ms_key or "", "") or form.get("phase", "acquisition")
        data: dict = {
            "label": form.get("label", ""),
            "phase": _phase,
            "amount": _fd(form.get("amount")) or Decimal("0"),
            "timing_type": form.get("timing_type") or "first_day",
            "is_deferred": form.get("is_deferred") == "true",
            "cost_category": form.get("cost_category") or "soft",
            "notes": form.get("notes") or None,
        }
        if item_id:
            row = await session.get(UseLine, UUID(item_id))
            if row:
                if row.is_auto_dev_fee:
                    # Auto Dev Fee row: user edits % and basis only.
                    # Label, phase, amount ($) are managed by the engine.
                    pct_raw = _fd(form.get("dev_fee_pct"))
                    if pct_raw is not None:
                        row.dev_fee_pct = pct_raw
                    basis_raw = form.get("dev_fee_basis")
                    if basis_raw in ("purchase_price", "tpc_excl_self"):
                        row.dev_fee_basis = basis_raw
                    if form.get("notes") is not None:
                        row.notes = form.get("notes") or None
                else:
                    # User edit on an auto Total Finance Costs row turns off
                    # the auto flag so engine stops recomputing.  User can
                    # delete the row to reset; next compute regenerates it.
                    if getattr(row, "is_auto_finance_cost", False):
                        row.is_auto_finance_cost = False
                    for k, v in data.items():
                        setattr(row, k, v)
        elif project_id:
            session.add(UseLine(project_id=project_id, **data))

    elif item_type == "income-streams":
        _amount_type = str(form.get("amount_type", "")).strip()
        _per_unit_val = _fd(form.get("amount_per_unit_monthly"))
        _fixed_val = _fd(form.get("amount_fixed_monthly"))
        # Clear the unused field so engine logic is unambiguous.
        if _amount_type == "flat":
            _per_unit_val = None
        elif _amount_type == "per_unit":
            _fixed_val = None
        data = {
            "label": form.get("label", ""),
            "stream_type": form.get("stream_type", "residential_rent"),
            "unit_count": _fi(form.get("unit_count")) or None,
            "amount_per_unit_monthly": _per_unit_val,
            "amount_fixed_monthly": _fixed_val,
            "stabilized_occupancy_pct": _fd(form.get("stabilized_occupancy_pct")) or Decimal("95"),
            "bad_debt_pct": _fd(form.get("bad_debt_pct")) or Decimal("0"),
            "concessions_pct": _fd(form.get("concessions_pct")) or Decimal("0"),
            "catchup_target_rent": _fd(form.get("catchup_target_rent")),
            "renovation_absorption_rate": _fd(form.get("renovation_absorption_rate")),
            "escalation_rate_pct_annual": _fd(form.get("escalation_rate_pct_annual")) or Decimal("0"),
            "active_in_phases": form.getlist("active_in_phases") or _fp(form.get("active_in_phases"), ["stabilized"]),
            "notes": form.get("notes") or None,
        }
        if item_id:
            row = await session.get(IncomeStream, UUID(item_id))
            if row:
                for k, v in data.items():
                    setattr(row, k, v)
        elif project_id:
            session.add(IncomeStream(project_id=project_id, **data))

    elif item_type == "expense-lines":
        _aip_list = form.getlist("active_in_phases")
        active_phases = _aip_list if _aip_list else _fp(form.get("active_in_phases"), ["stabilized"])
        per_type_val = form.get("per_type") or None
        per_value_val = _fd(form.get("per_value"))
        # For flat type, annual_amount mirrors per_value for backward-compat display
        # For per_unit/sqft types, annual_amount stays 0 until compute engine scales it
        if per_value_val and per_type_val in (None, "flat"):
            annual_amt = per_value_val
        else:
            annual_amt = _fd(form.get("annual_amount")) or Decimal("0")
        data = {
            "label": form.get("label", ""),
            "annual_amount": annual_amt,
            "per_value": per_value_val,
            "per_type": per_type_val,
            "scale_with_lease_up": form.get("scale_with_lease_up") == "on",
            "lease_up_floor_pct": _fd(form.get("lease_up_floor_pct")),
            "escalation_rate_pct_annual": _fd(form.get("escalation_rate_pct_annual")) or Decimal("3"),
            "active_in_phases": active_phases,
            "notes": form.get("notes") or None,
        }
        if item_id:
            row = await session.get(OperatingExpenseLine, UUID(item_id))
            if row:
                for k, v in data.items():
                    setattr(row, k, v)
        elif project_id:
            session.add(OperatingExpenseLine(project_id=project_id, **data))

    elif item_type == "capital-modules":
        source_d: dict = {}
        if src_amt := _fd(form.get("source_amount")):
            source_d["amount"] = float(src_amt)
        # Grant cap (per-Use eligibility). When set, source.amount is engine-
        # computed each compute pass and ignored from form input above.
        if src_max := _fd(form.get("source_maximum")):
            source_d["maximum"] = float(src_max)
        if src_pct := _fd(form.get("source_pct")):
            source_d["pct_of_total_cost"] = float(src_pct)
        if src_rate := _fd(form.get("source_interest_rate")):
            source_d["interest_rate_pct"] = float(src_rate)
        if cp := form.get("compounding_period"):
            source_d["compounding_period"] = cp
        if amort := _fi(form.get("amort_term_years"), None):
            source_d["amort_term_years"] = amort
        if hold_term := _fi(form.get("hold_term_years"), None):
            source_d["hold_term_years"] = hold_term
        if ppct := _fd(form.get("prepay_penalty_pct")):
            source_d["prepay_penalty_pct"] = float(ppct)
        if ltv := _fd(form.get("ltv_pct")):
            source_d["ltv_pct"] = float(ltv)
        constr_carry_type = form.get("construction_carry_type", "none")
        # Carry rate: use source rate so the engine finds it in both places
        _carry_rate = _fd(form.get("source_interest_rate"))

        _carry_schedule_mode = (form.get("carry_schedule_mode") or "simple").strip()
        if _carry_schedule_mode == "schedule":
            # Parse N-phase carry schedule from form arrays.
            _sched_labels = form.getlist("carry_phase_label[]")
            _sched_types = form.getlist("carry_phase_type[]")
            _sched_dur_types = form.getlist("carry_phase_duration_type[]")
            _sched_months = form.getlist("carry_phase_months[]")
            _sched_milestones = form.getlist("carry_phase_milestone_key[]")
            _sched_rates = form.getlist("carry_phase_rate_pct[]")
            _sched_amorts = form.getlist("carry_phase_amort_years[]")
            _schedule_phases: list[dict] = []
            for _si in range(len(_sched_types)):
                _p_ct = (_sched_types[_si] if _si < len(_sched_types) else "none").strip()
                if not _p_ct or _p_ct == "none":
                    continue
                _p_dur_type = (_sched_dur_types[_si] if _si < len(_sched_dur_types) else "remainder").strip()
                if _p_dur_type == "months":
                    _p_dur: dict = {"type": "months", "months": int((_sched_months[_si] if _si < len(_sched_months) else "") or 0)}
                elif _p_dur_type == "milestone":
                    _p_dur = {"type": "milestone", "milestone_key": (_sched_milestones[_si] if _si < len(_sched_milestones) else "").strip()}
                else:
                    _p_dur = {"type": "remainder"}
                _p: dict = {
                    "label": (_sched_labels[_si] if _si < len(_sched_labels) else "").strip() or _p_ct,
                    "carry_type": _p_ct,
                    "duration": _p_dur,
                }
                _p_rate = _fd(_sched_rates[_si] if _si < len(_sched_rates) else None)
                if _p_rate is not None:
                    _p["rate_pct"] = float(_p_rate)
                _p_amort = _fi(_sched_amorts[_si] if _si < len(_sched_amorts) else None, None)
                if _p_amort:
                    _p["amort_term_years"] = _p_amort
                _schedule_phases.append(_p)
            carry_d: dict = {"schedule": _schedule_phases}
        else:
            constr_phase: dict = {
                "name": "construction",
                "carry_type": constr_carry_type,
                "payment_frequency": form.get("construction_payment_frequency", "monthly"),
            }
            if _carry_rate is not None:
                constr_phase["io_rate_pct"] = float(_carry_rate)
            if constr_carry_type == "converts_to_permanent":
                if perm_rate := _fd(form.get("perm_rate_pct")):
                    constr_phase["perm_rate_pct"] = float(perm_rate)
                if perm_term := _fi(form.get("perm_term_years"), None):
                    constr_phase["perm_term_years"] = perm_term
                if perm_trig := form.get("perm_conversion_trigger"):
                    constr_phase["perm_conversion_trigger"] = perm_trig
            _op_phase: dict = {
                "name": "operation",
                "carry_type": form.get("operation_carry_type", "none"),
                "payment_frequency": form.get("operation_payment_frequency", "monthly"),
            }
            if _carry_rate is not None:
                _op_phase["io_rate_pct"] = float(_carry_rate)
            if amort:
                _op_phase["amort_term_years"] = amort
            carry_d = {"phases": [constr_phase, _op_phase]}
        # Exit Vehicle: "maturity" | "sale" | "<module_uuid>" (retiring source).
        # Validate: must be one of the literals OR a UUID of another module on
        # the same scenario. Fall back to "maturity" if invalid.
        #
        # Non-debt vehicle types (equity, grants, etc.) are forced to
        # "maturity" as a no-op sentinel — their UI hides Exit Vehicle entirely.
        _vehicle_type = form.get("vehicle_type", "debt") or "debt"
        _equity_role = (form.get("equity_role") or "").strip() or None
        if _vehicle_type != "debt":
            _vehicle_value = "maturity"
        else:
            _vehicle_raw = (form.get("exit_vehicle") or "").strip()
            _vehicle_value = "maturity"
            if _vehicle_raw in {"maturity", "sale"}:
                _vehicle_value = _vehicle_raw
            elif _vehicle_raw:
                try:
                    _vehicle_uuid = UUID(_vehicle_raw)
                    _sibling = (await session.execute(
                        select(CapitalModule.id).where(
                            CapitalModule.scenario_id == model_id,
                            CapitalModule.id == _vehicle_uuid,
                        )
                    )).scalar_one_or_none()
                    if _sibling is not None and (not item_id or str(_sibling) != item_id):
                        _vehicle_value = str(_sibling)
                except (ValueError, AttributeError):
                    pass
        exit_d = {
            "exit_type": form.get("exit_type", "full_payoff"),
            "vehicle": _vehicle_value,
        }

        # Derive `active_phase_end` + `active_to_milestone` from the vehicle.
        # `active_phase_end` is deprecated as a user input but still written
        # server-side from the vehicle so legacy read-paths (Gantt, reports)
        # keep working without a DB migration.
        _APS_TO_MS = {
            "acquisition": "close", "close": "close",
            "pre_construction": "pre_development",
            "construction": "construction",
            "lease_up": "operation_lease_up", "operation_lease_up": "operation_lease_up",
            "stabilized": "operation_stabilized", "operation_stabilized": "operation_stabilized",
            "exit": "divestment", "divestment": "divestment",
        }
        _retirer_aps: str | None = None
        if _vehicle_value not in {"maturity", "sale"} and _vehicle_value:
            try:
                _retirer = await session.get(CapitalModule, UUID(_vehicle_value))
                if _retirer is not None:
                    _retirer_aps = _retirer.active_phase_start
            except (ValueError, AttributeError):
                _retirer_aps = None
        if _vehicle_value == "sale":
            _derived_end_phase = "exit"
            _derived_to_ms = "divestment"
        elif _retirer_aps:
            _derived_end_phase = _retirer_aps
            _derived_to_ms = _APS_TO_MS.get(_retirer_aps, _retirer_aps)
        else:
            # "maturity" or unresolved → perpetuity sentinel; Gantt shows through exit.
            _derived_end_phase = "exit"
            _derived_to_ms = "divestment"
        explicit_pos = _fi(form.get("stack_position"), None)
        if not item_id and (not explicit_pos or explicit_pos == 0):
            # Auto-assign: place at end of current stack
            max_pos_result = await session.execute(
                select(func.max(CapitalModule.stack_position)).where(CapitalModule.scenario_id == model_id)
            )
            max_pos = max_pos_result.scalar_one_or_none() or 0
            explicit_pos = max_pos + 1
        final_pos = explicit_pos or 1
        # Uniqueness: if another module already holds this position, resolve the conflict.
        conflict_stmt = (
            select(CapitalModule)
            .where(CapitalModule.scenario_id == model_id, CapitalModule.stack_position == final_pos)
        )
        if item_id:
            conflict_stmt = conflict_stmt.where(CapitalModule.id != UUID(item_id))
        _conflict_mod = (await session.execute(conflict_stmt)).scalars().first()
        if _conflict_mod is not None:
            if item_id:
                # Edit: honor the user's chosen position — swap with the conflicting
                # module so it takes the edited module's previous slot.
                _editing_row = await session.get(CapitalModule, UUID(item_id))
                _old_pos = (_editing_row.stack_position if _editing_row else None) or final_pos
                _conflict_mod.stack_position = _old_pos
                session.add(_conflict_mod)
            else:
                # Create: bump the new module to the end of the stack.
                max_pos_result = await session.execute(
                    select(func.max(CapitalModule.stack_position)).where(CapitalModule.scenario_id == model_id)
                )
                final_pos = (max_pos_result.scalar_one_or_none() or 0) + 1
        # Record which vehicle pre-filled this module (nullable; SET NULL on vehicle delete)
        _sv_id_raw = (form.get("source_vehicle_id") or "").strip()
        _sv_uuid: UUID | None = None
        if _sv_id_raw:
            try:
                _sv_uuid = UUID(_sv_id_raw)
            except (ValueError, AttributeError):
                pass

        data = {
            "label": form.get("label", ""),
            "vehicle_type": _vehicle_type,
            "equity_role": _equity_role,
            "stack_position": final_pos,
            "source": source_d,
            "carry": carry_d,
            "exit_terms": exit_d,
            "active_phase_end": _derived_end_phase,
            "source_vehicle_id": _sv_uuid,
        }
        # Draw schedule fields from form.  `ds_active_to_milestone` is ignored —
        # the repayment milestone is derived from Exit Vehicle (above).
        _ds_from_ms = form.get("ds_active_from_milestone") or ""
        _ds_to_ms = _derived_to_ms
        _ds_from_offset = _fi(form.get("ds_active_from_offset_days"), 0) or 0
        _ds_to_offset = _fi(form.get("ds_active_to_offset_days"), 0) or 0
        _ds_frequency = _fi(form.get("ds_draw_every_n_months"), 1) or 1
        _ds_rate = source_d.get("interest_rate_pct", 0.0)

        if item_id:
            row = await session.get(CapitalModule, UUID(item_id))
            if row:
                # Preserve internal-only source keys (auto_size) not exposed in the UI form
                if (row.source or {}).get("auto_size"):
                    source_d["auto_size"] = True
                data["source"] = source_d
                for k, v in data.items():
                    setattr(row, k, v)
                # Fixed-amount sources (grant/forgivable_loan/tax_credit/equity): mirror the
                # new source.amount onto the active-project junction row so the engine
                # actually picks up the change. Without this the junction's stale amount
                # overlays source.amount in memory at engine load and the new value is
                # silently dropped.
                if data["vehicle_type"] in ("grant", "forgivable_loan", "tax_credit", "equity"):
                    from app.models.capital import CapitalModuleProject as _CMP_edit
                    _new_amt = Decimal(str(source_d.get("amount") or 0))
                    _active_pid = project_id if project_id is not None else (
                        default_project.id if default_project else None
                    )
                    if _active_pid is not None:
                        _j_row = (await session.execute(
                            select(_CMP_edit).where(
                                _CMP_edit.capital_module_id == row.id,
                                _CMP_edit.project_id == _active_pid,
                            )
                        )).scalar_one_or_none()
                        if _j_row is None:
                            session.add(_CMP_edit(
                                capital_module_id=row.id,
                                project_id=_active_pid,
                                amount=_new_amt,
                                active_from=None,
                                active_to=row.active_phase_end,
                                auto_size=False,
                            ))
                        else:
                            _j_row.amount = _new_amt
                # Mirror perm-debt hold_term_years / dscr_min to wizard staging
                # so re-opening Deal Setup wizard shows the latest value.
                if data["vehicle_type"] == "debt" and (
                    "hold_term_years" in source_d or "dscr_min" in source_d
                ):
                    _default_proj = (await session.execute(
                        select(Project).where(Project.scenario_id == model_id)
                        .order_by(Project.created_at.asc()).limit(1)
                    )).scalar_one_or_none()
                    if _default_proj is not None:
                        _oi = (await session.execute(
                            select(OperationalInputs).where(
                                OperationalInputs.project_id == _default_proj.id
                            )
                        )).scalar_one_or_none()
                        if _oi is not None:
                            _dt = dict(_oi.debt_terms or {})
                            _pd_entry = dict(_dt.get("permanent_debt", {}))
                            if "hold_term_years" in source_d:
                                _pd_entry["hold_term_years"] = source_d["hold_term_years"]
                            if "dscr_min" in source_d:
                                _pd_entry["dscr_min"] = source_d["dscr_min"]
                            _dt["permanent_debt"] = _pd_entry
                            _oi.debt_terms = _dt
                            session.add(_oi)
            # Update matching DrawSource (active window, offsets, frequency)
            _ds_id_raw = str(form.get("ds_id") or "").strip()
            if _ds_id_raw:
                try:
                    _ds_row = await session.get(DrawSource, UUID(_ds_id_raw))
                    if _ds_row and _ds_row.scenario_id == model_id:
                        if _ds_from_ms:
                            _ds_row.active_from_milestone = _ds_from_ms
                        if _ds_to_ms:
                            _ds_row.active_to_milestone = _ds_to_ms
                        _ds_row.active_from_offset_days = _ds_from_offset
                        _ds_row.active_to_offset_days = _ds_to_offset
                        _ds_row.draw_every_n_months = _ds_frequency
                        _ds_row.annual_interest_rate = Decimal(str(_ds_rate))
                        _ds_row.label = data["label"]
                except (ValueError, TypeError):
                    pass
        else:
            _cm_id = _uuid_mod.uuid4()
            cm = CapitalModule(id=_cm_id, scenario_id=model_id, **data)
            session.add(cm)
            # Create CapitalModuleProject junction row(s) — the engine's per-project
            # query INNER JOINs on this table, so a module with no junction row is
            # invisible to sizing/cashflow. Put the full source.amount on the active
            # project; zero rows on other projects so the source appears in their
            # coverage lists without double-counting.
            from app.models.capital import CapitalModuleProject as _CMP_create
            _all_projects = (await session.execute(
                select(Project).where(Project.scenario_id == model_id).order_by(Project.created_at.asc())
            )).scalars().all()
            _src_amt_dec = Decimal(str(source_d.get("amount") or 0))
            _is_auto = bool(source_d.get("auto_size"))
            _primary_pid = project_id if project_id is not None else (default_project.id if default_project else None)
            for _p in _all_projects:
                _amt = _src_amt_dec if (_p.id == _primary_pid) else Decimal("0")
                session.add(_CMP_create(
                    capital_module_id=_cm_id,
                    project_id=_p.id,
                    amount=_amt,
                    active_from=None,
                    active_to=_derived_end_phase,
                    auto_size=_is_auto,
                ))
            # Auto-create linked DrawSource.  Non-debt vehicle types map to source_type="equity".
            _src_type = "debt" if data["vehicle_type"] == "debt" else "equity"
            # Determine sort order for new DrawSource
            _max_sort = (await session.execute(
                select(func.max(DrawSource.sort_order)).where(DrawSource.scenario_id == model_id)
            )).scalar_one_or_none() or 0
            ds = DrawSource(
                scenario_id=model_id,
                label=data["label"],
                source_type=_src_type,
                sort_order=_max_sort + 1,
                draw_every_n_months=_ds_frequency,
                annual_interest_rate=Decimal(str(_ds_rate)),
                active_from_milestone=_ds_from_ms or "construction",
                active_to_milestone=_ds_to_ms or "maturity",
                active_from_offset_days=_ds_from_offset,
                active_to_offset_days=_ds_to_offset,
                capital_module_id=_cm_id,
            )
            session.add(ds)

        # ── Per-Use eligibility sync (source-side editor writes use-side column) ──
        # Form field `eligible_use_ids[]` lists Use UUIDs this grant funds.
        # Maintains `use_lines.eligible_module_ids` bidirectional consistency:
        #   - For each ticked Use → append this module's ID
        #   - For each previously-linked Use no longer ticked → remove this module's ID
        # Also validates the cap/eligibility combination: maximum requires
        # at least one Use checked, and any ticked Use requires a maximum.
        _module_uuid_for_sync: UUID | None = None
        if item_id:
            try:
                _module_uuid_for_sync = UUID(item_id)
            except (ValueError, AttributeError):
                _module_uuid_for_sync = None
        else:
            _module_uuid_for_sync = _cm_id if "_cm_id" in locals() else None

        if _module_uuid_for_sync is not None:
            _raw_ids = form.getlist("eligible_use_ids")
            _new_eligible: set[UUID] = set()
            for _rid in _raw_ids:
                _rid = (_rid or "").strip()
                if not _rid:
                    continue
                try:
                    _new_eligible.add(UUID(_rid))
                except (ValueError, AttributeError):
                    continue

            _has_maximum = source_d.get("maximum") is not None
            # Validation: maximum + eligibility must agree
            if _has_maximum and not _new_eligible:
                from fastapi import HTTPException as _HTTPExc
                raise _HTTPExc(
                    status_code=422,
                    detail="Maximum requires at least one eligible Use to be selected.",
                )
            if _new_eligible and not _has_maximum:
                from fastapi import HTTPException as _HTTPExc
                raise _HTTPExc(
                    status_code=422,
                    detail="Eligible Uses selected — Maximum must be entered.",
                )

            # Apply bi-directional sync against all Uses on this scenario
            _all_uses = (await session.execute(
                select(UseLine)
                .join(Project, UseLine.project_id == Project.id)
                .where(Project.scenario_id == model_id)
            )).scalars().all()
            _mod_id_str = str(_module_uuid_for_sync)
            for _ul in _all_uses:
                _cur = [x for x in (_ul.eligible_module_ids or [])]
                _cur_strs = [str(x) for x in _cur]
                _should_include = _ul.id in _new_eligible
                _is_present = _mod_id_str in _cur_strs
                if _should_include and not _is_present:
                    _ul.eligible_module_ids = _cur + [_module_uuid_for_sync]
                elif (not _should_include) and _is_present:
                    _ul.eligible_module_ids = [
                        x for x in _cur if str(x) != _mod_id_str
                    ]

    elif item_type == "waterfall-tiers":
        data = {
            "priority": _fi(form.get("priority"), 1),
            "tier_type": form.get("tier_type", "residual"),
            "description": form.get("description") or None,
            "lp_split_pct": _fd(form.get("lp_split_pct")) or Decimal("0"),
            "gp_split_pct": _fd(form.get("gp_split_pct")) or Decimal("0"),
            "irr_hurdle_pct": _fd(form.get("irr_hurdle_pct")),
            "max_pct_of_distributable": _fd(form.get("max_pct_of_distributable")),
            "interest_rate_pct": _fd(form.get("interest_rate_pct")),
        }
        if item_id:
            row = await session.get(WaterfallTier, UUID(item_id))
            if row:
                for k, v in data.items():
                    setattr(row, k, v)
        else:
            session.add(WaterfallTier(scenario_id=model_id, **data))

    elif item_type == "milestones":
        from datetime import date as _date
        def _parse_date(v: str | None) -> _date | None:
            if not v or not v.strip():
                return None
            try:
                return _date.fromisoformat(v.strip()[:10])
            except ValueError:
                return None

        mtype_raw = form.get("milestone_type", "construction")
        try:
            mtype = MilestoneType(mtype_raw)
        except ValueError:
            mtype = MilestoneType.construction

        trigger_raw = str(form.get("trigger_milestone_id") or "").strip()
        try:
            trigger_id = UUID(trigger_raw) if trigger_raw else None
        except ValueError:
            trigger_id = None

        # Guard: reject a trigger that belongs to a different project.
        # This prevents the cross-project trigger corruption that caused
        # milestones on project 2 to point at milestones on project 1.
        if trigger_id is not None:
            _trigger_ms = await session.get(Milestone, trigger_id)
            _ms_project_id: UUID | None = project_id
            if item_id:
                try:
                    _existing_for_check = await session.get(Milestone, UUID(item_id))
                    if _existing_for_check:
                        _ms_project_id = _existing_for_check.project_id
                except (ValueError, AttributeError):
                    pass
            if _trigger_ms is None or (_ms_project_id is not None and _trigger_ms.project_id != _ms_project_id):
                trigger_id = None

        data = {
            "duration_days": _fi(form.get("duration_days"), 0),
            "milestone_type": mtype,
            "trigger_milestone_id": trigger_id,
            "trigger_offset_days": _fi(form.get("trigger_offset_days"), 0),
            # anchor: keep target_date only when no trigger; clear it when trigger set
            "target_date": _parse_date(str(form.get("target_date") or "")) if not trigger_id else None,
        }
        if item_id:
            row = await session.get(Milestone, UUID(item_id))
            if row:
                for k, v in data.items():
                    setattr(row, k, v)
        elif project_id:
            session.add(Milestone(
                project_id=project_id,
                sequence_order=0,
                **data,
            ))

    elif item_type == "unit-mix":
        from uuid import uuid4 as _uuid4
        def _fj(v): return float(v) if v is not None else None
        data = {
            "label": form.get("label", "").strip() or "Units",
            "unit_count": _fi(form.get("unit_count"), 1) or 1,
            "avg_sqft": _fj(_fd(form.get("avg_sqft"))),
            "beds": _fj(_fd(form.get("beds"))),
            "baths": _fj(_fd(form.get("baths"))),
            "market_rent_per_unit": _fj(_fd(form.get("market_rent_per_unit"))),
            "in_place_rent_per_unit": _fj(_fd(form.get("in_place_rent_per_unit"))),
            "unit_strategy": form.get("unit_strategy") or None,
            "post_reno_rent_per_unit": _fj(_fd(form.get("post_reno_rent_per_unit"))),
            "notes": form.get("notes") or None,
        }
        if project_id:
            _um_proj = await session.get(Project, project_id)
            if _um_proj is not None:
                rows = list(_um_proj.unit_mix or [])
                if item_id:
                    _uid_str = str(UUID(item_id))
                    idx = next((i for i, d in enumerate(rows) if d.get("id") == _uid_str), None)
                    if idx is not None:
                        rows[idx] = {**data, "id": _uid_str}
                    else:
                        rows.append({**data, "id": _uid_str})
                else:
                    rows.append({**data, "id": str(_uuid4())})
                _um_proj.unit_mix = rows
                session.add(_um_proj)

    # Re-sync CapitalModule milestone FKs whenever debt sources OR milestones
    # were touched in this request. Lookups are scenario-scoped and cheap; one
    # round-trip per module. Engine reads the FK first, so the next compute
    # will pick up any renamed milestone target without further user action.
    if item_type in ("capital-modules", "milestones"):
        from app.services.capital_module_milestones import (
            sync_milestone_fks_for_scenario,
        )
        await sync_milestone_fks_for_scenario(session, model_id)

    await session.flush()
    panel_data = await _load_builder_data(session, model_id, project_id=project_id)
    ctx = {"model": model, "active_module": module, "wizard_mode": _wizard_mode_from_request(request), **panel_data}
    return templates.TemplateResponse(request, "partials/model_builder_panel.html", ctx)


@router.delete("/ui/forms/{model_id}/{item_type}/{item_id}", response_class=HTMLResponse)
async def handle_form_delete(
    request: Request,
    model_id: UUID,
    item_type: str,
    item_id: str,
    session: DBSession,
) -> HTMLResponse:
    """Delete a line item and return the refreshed panel HTML."""
    model = await session.get(DealModel, model_id)
    if model is None:
        return HTMLResponse("<p class='text-muted'>Model not found.</p>", status_code=404)

    module = _ITEM_TYPE_TO_MODULE.get(item_type, "uses")
    try:
        uid = UUID(item_id)
    except (ValueError, AttributeError):
        # Legacy unit-mix rows may have rendered without an id; surface a panel
        # refresh rather than 500 so the user can retry after backfill runs.
        if item_type == "unit-mix":
            _active_proj_id = await _active_project_from_request(request, session, model_id)
            panel_data = await _load_builder_data(session, model_id, project_id=_active_proj_id)
            ctx = {"model": model, "active_module": module, "wizard_mode": _wizard_mode_from_request(request), **panel_data}
            return templates.TemplateResponse(request, "partials/model_builder_panel.html", ctx)
        return HTMLResponse("<p class='text-muted'>Invalid item id.</p>", status_code=400)

    row = None
    if item_type == "use-lines":
        row = await session.get(UseLine, uid)
    elif item_type == "income-streams":
        row = await session.get(IncomeStream, uid)
    elif item_type == "expense-lines":
        row = await session.get(OperatingExpenseLine, uid)
    elif item_type == "capital-modules":
        row = await session.get(CapitalModule, uid)
    elif item_type == "waterfall-tiers":
        row = await session.get(WaterfallTier, uid)
    elif item_type == "milestones":
        row = await session.get(Milestone, uid)
    elif item_type == "unit-mix":
        # JSONB — find the project and filter out the row by id
        _del_proj_id = await _active_project_from_request(request, session, model_id)
        if _del_proj_id:
            _del_proj = await session.get(Project, _del_proj_id)
            if _del_proj is not None:
                _del_proj.unit_mix = [d for d in (_del_proj.unit_mix or []) if d.get("id") != str(uid)]
                session.add(_del_proj)
        await session.flush()
        _active_proj_id2 = await _active_project_from_request(request, session, model_id)
        panel_data = await _load_builder_data(session, model_id, project_id=_active_proj_id2)
        ctx = {"model": model, "active_module": module, "wizard_mode": _wizard_mode_from_request(request), **panel_data}
        return templates.TemplateResponse(request, "partials/model_builder_panel.html", ctx)

    if row is not None:
        if item_type == "use-lines" and getattr(row, "is_auto_dev_fee", False):
            return HTMLResponse(
                "<p class='text-muted'>The auto Developer Fee Use Line cannot be "
                "deleted. Set its % to 0 to disable.</p>",
                status_code=403,
            )
        await session.delete(row)
        await session.flush()

    _active_proj_id = await _active_project_from_request(request, session, model_id)
    panel_data = await _load_builder_data(session, model_id, project_id=_active_proj_id)
    ctx = {"model": model, "active_module": module, "wizard_mode": _wizard_mode_from_request(request), **panel_data}
    return templates.TemplateResponse(request, "partials/model_builder_panel.html", ctx)


@router.post("/ui/models/{model_id}/unit-mix/apply-to-revenue", response_class=HTMLResponse)
async def apply_unit_mix_to_revenue(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Add IncomeStream rows for UnitMix rows that don't already have one.
    **Additive only** — never deletes or overwrites existing streams.

    Strategy → generated label mapping:
      - base_escalation:       "{unit_label} Rent"
      - ltl_catchup:           "{unit_label} Rent"
      - value_add_renovation:  "{unit_label} Rent (Renovated)"

    If a stream already exists at the generated label this handler skips
    it (preserving rent / occupancy / escalation edits). Stub streams like
    "All Units" stay put — the user deletes them manually once the
    per-unit-type streams look right.
    """
    model = await session.get(DealModel, model_id)
    if model is None:
        return HTMLResponse("<p class='text-muted'>Model not found.</p>", status_code=404)

    # Resolve active project from HX-Current-URL (multi-project tabs) and
    # fall back to the default project. All inserts and the panel re-render
    # must use this project_id, not default_project.
    active_proj_id = await _active_project_from_request(request, session, model_id)
    if active_proj_id is None:
        default_project = (await session.execute(
            select(Project).where(Project.scenario_id == model_id).order_by(Project.created_at.asc()).limit(1)
        )).scalar_one_or_none()
        if default_project is None:
            return HTMLResponse("<p class='text-muted'>No project.</p>", status_code=400)
        active_proj_id = default_project.id

    _active_proj = await session.get(Project, active_proj_id)
    if _active_proj is not None and _ensure_unit_mix_ids(_active_proj):
        await session.flush()
    unit_mix_rows = [_UMRow(r) for r in (_active_proj.unit_mix or [])] if _active_proj else []
    if not unit_mix_rows:
        panel_data = await _load_builder_data(session, model_id, project_id=active_proj_id)
        ctx = {"model": model, "active_module": "property", **panel_data}
        return templates.TemplateResponse(request, "partials/model_builder_panel.html", ctx)

    # Build the candidate stream list from each UnitMix row's strategy.
    # Additive sync filters this list against existing labels below.
    to_generate: list[dict] = []
    for u in unit_mix_rows:
        strategy = (u.unit_strategy or "base_escalation")
        label = f"{u.label} Rent"
        count = int(u.unit_count or 0)
        if count <= 0:
            continue

        ip_rent = _to_decimal_or_none(u.in_place_rent_per_unit)
        mkt_rent = _to_decimal_or_none(u.market_rent_per_unit)
        post_reno = _to_decimal_or_none(u.post_reno_rent_per_unit)

        if strategy == "value_add_renovation":
            # Post-reno rent is the target; fall back to market if unset
            base_rent = post_reno or mkt_rent or ip_rent
            stream = dict(
                label=f"{u.label} Rent (Renovated)",
                stream_type=IncomeStreamType.residential_rent,
                unit_count=count,
                amount_per_unit_monthly=base_rent,
                stabilized_occupancy_pct=Decimal("95"),
                escalation_rate_pct_annual=Decimal("3"),
                renovation_absorption_rate=Decimal("1"),
                active_in_phases=["lease_up", "stabilized"],
            )
        elif strategy == "ltl_catchup":
            stream = dict(
                label=label,
                stream_type=IncomeStreamType.residential_rent,
                unit_count=count,
                amount_per_unit_monthly=(ip_rent or mkt_rent or Decimal("0")),
                stabilized_occupancy_pct=Decimal("95"),
                catchup_target_rent=mkt_rent,
                escalation_rate_pct_annual=Decimal("3"),
                active_in_phases=["lease_up", "stabilized"],
            )
        else:  # base_escalation
            stream = dict(
                label=label,
                stream_type=IncomeStreamType.residential_rent,
                unit_count=count,
                amount_per_unit_monthly=(ip_rent or mkt_rent or Decimal("0")),
                stabilized_occupancy_pct=Decimal("95"),
                escalation_rate_pct_annual=Decimal("3"),
                active_in_phases=["lease_up", "stabilized"],
            )
        to_generate.append(stream)

    # Additive-only: skip any generated label whose stream already exists.
    # Preserves manual rent / occupancy / escalation edits the user made.
    generated_labels = {s["label"] for s in to_generate}
    existing_labels: set[str] = set()
    if generated_labels:
        existing_labels = set((await session.execute(
            select(IncomeStream.label).where(
                IncomeStream.project_id == active_proj_id,
                IncomeStream.label.in_(generated_labels),
            )
        )).scalars())

    for data in to_generate:
        if data["label"] in existing_labels:
            continue
        session.add(IncomeStream(project_id=active_proj_id, **data))
    await session.flush()

    # Return the refreshed Revenue panel — the sync banner triggers from
    # there, so stay oriented on Revenue rather than bouncing to Property.
    panel_data = await _load_builder_data(session, model_id, project_id=active_proj_id)
    ctx = {"model": model, "active_module": "revenue", **panel_data}
    return templates.TemplateResponse(request, "partials/model_builder_panel.html", ctx)


def _to_decimal_or_none(v) -> Decimal | None:
    """Coerce a numeric value to Decimal, or None if zero/missing."""
    if v is None:
        return None
    try:
        d = Decimal(str(v))
        return d if d != 0 else None
    except Exception:
        return None


@router.post("/ui/models/{model_id}/sensitivity/run", response_class=HTMLResponse)
async def run_sensitivity_analysis(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Run a 5x5 sensitivity compute and persist the result as JSON on
    OperationalOutputs.sensitivity_matrix. Returns the refreshed Sensitivity
    panel so the user sees results inline."""
    from app.engines.sensitivity_matrix import compute_sensitivity_matrix

    model = await session.get(DealModel, model_id)
    if model is None:
        return HTMLResponse("<p class='text-muted'>Model not found.</p>", status_code=404)

    form = await request.form()
    axis_x = form.get("axis_x") or "noi_escalation_rate_pct"
    axis_y = form.get("axis_y") or "exit_cap_rate_pct"
    metric = form.get("metric") or "project_irr_levered"

    try:
        matrix = await compute_sensitivity_matrix(
            deal_model_id=model_id,
            session=session,
            axis_x=axis_x,
            axis_y=axis_y,
            metric=metric,
        )
    except ValueError as e:
        # Bad axis/metric combo — surface the error in the panel
        panel_data = await _load_builder_data(session, model_id)
        ctx = {"model": model, "active_module": "sensitivity", **panel_data,
               "sensitivity_error": str(e)}
        return templates.TemplateResponse(request, "partials/model_builder_panel.html", ctx)

    # Persist on OperationalOutputs.sensitivity_matrix (JSON column).
    # compute_sensitivity_matrix runs a final compute_cash_flows so a fresh
    # OperationalOutputs row now exists. Scope to default project because
    # sensitivity_matrix is a scenario-wide artifact today (Phase 3d/e will
    # revisit per-project sensitivity).
    outputs = (await session.execute(
        select(OperationalOutputs)
        .where(OperationalOutputs.scenario_id == model_id)
        .join(Project, Project.id == OperationalOutputs.project_id)
        .order_by(Project.created_at.asc())
        .limit(1)
    )).scalar_one_or_none()
    if outputs is not None:
        outputs.sensitivity_matrix = matrix
        session.add(outputs)
        await session.flush()

    panel_data = await _load_builder_data(session, model_id)
    ctx = {"model": model, "active_module": "sensitivity", **panel_data}
    return templates.TemplateResponse(request, "partials/model_builder_panel.html", ctx)


async def _sync_opportunity_buildings_to_projects(
    opportunity: Opportunity,
    buildings: object,
    session: AsyncSession,
) -> None:
    """No-op stub — Building entity removed. Physical attrs live on Opportunity."""
    pass


async def _auto_assign_opportunity_to_project(
    opportunity: Opportunity,
    project: Project,
    session: AsyncSession,
) -> None:
    """Set project.parcel_id from opportunity.parcel_id when available."""
    if project.parcel_id is None and opportunity.parcel_id is not None:
        project.parcel_id = opportunity.parcel_id


async def _copy_project_data(
    src_proj: Project,
    dst_proj: Project,
    session: AsyncSession,
    *,
    user_id: UUID | None = None,
    org_id: UUID | None = None,
) -> None:
    """Copy milestones (with trigger remapping), use lines, income streams,
    expense lines, and operational inputs from src_proj to dst_proj.
    Caller is responsible for deleting dst_proj's existing data first.

    When ``user_id`` and ``org_id`` are provided, Type 1 (Org-Set) defaults
    are re-applied to the copied OperationalInputs row so the clone picks up
    current org policy instead of inheriting a stale baseline."""
    # Copy milestones (preserve trigger chain with remapped IDs)
    src_milestones = list((await session.execute(
        select(Milestone).where(Milestone.project_id == src_proj.id)
    )).scalars())
    ms_id_map: dict = {}
    for ms in src_milestones:
        new_ms = Milestone(
            project_id=dst_proj.id,
            milestone_type=ms.milestone_type,
            label=ms.label,
            target_date=ms.target_date,
            duration_days=ms.duration_days,
            sequence_order=ms.sequence_order,
        )
        session.add(new_ms)
        await session.flush()
        ms_id_map[ms.id] = new_ms.id

    # Resolve trigger_milestone_id after all are created
    for ms in src_milestones:
        if ms.trigger_milestone_id and ms.trigger_milestone_id in ms_id_map:
            new_ms_obj = await session.get(Milestone, ms_id_map[ms.id])
            if new_ms_obj:
                new_ms_obj.trigger_milestone_id = ms_id_map[ms.trigger_milestone_id]
                new_ms_obj.trigger_offset_days = ms.trigger_offset_days

    # Copy Use lines
    for u in (await session.execute(
        select(UseLine).where(UseLine.project_id == src_proj.id)
    )).scalars():
        session.add(UseLine(
            project_id=dst_proj.id,
            label=u.label, phase=u.phase,
            amount=u.amount, is_deferred=u.is_deferred, notes=u.notes,
        ))

    # Copy Income streams
    for s in (await session.execute(
        select(IncomeStream).where(IncomeStream.project_id == src_proj.id)
    )).scalars():
        session.add(IncomeStream(
            project_id=dst_proj.id,
            stream_type=s.stream_type, label=s.label,
            unit_count=s.unit_count,
            amount_per_unit_monthly=s.amount_per_unit_monthly,
            amount_fixed_monthly=s.amount_fixed_monthly,
            stabilized_occupancy_pct=s.stabilized_occupancy_pct,
            escalation_rate_pct_annual=s.escalation_rate_pct_annual,
            active_in_phases=s.active_in_phases, notes=s.notes,
        ))

    # Copy Expense lines
    for e in (await session.execute(
        select(OperatingExpenseLine).where(OperatingExpenseLine.project_id == src_proj.id)
    )).scalars():
        session.add(OperatingExpenseLine(
            project_id=dst_proj.id,
            label=e.label, annual_amount=e.annual_amount,
            escalation_rate_pct_annual=e.escalation_rate_pct_annual,
            active_in_phases=e.active_in_phases, notes=e.notes,
        ))

    # Copy unit_mix JSONB
    if src_proj.unit_mix:
        dst_proj.unit_mix = list(src_proj.unit_mix)
        session.add(dst_proj)

    # Copy OperationalInputs if any
    src_inputs = (await session.execute(
        select(OperationalInputs).where(OperationalInputs.project_id == src_proj.id)
    )).scalar_one_or_none()
    if src_inputs:
        new_inputs = OperationalInputs(project_id=dst_proj.id)
        skip = {"id", "project_id"}
        for col in OperationalInputs.__table__.columns:
            if col.name not in skip:
                setattr(new_inputs, col.name, getattr(src_inputs, col.name, None))
        session.add(new_inputs)
        # Clone path: re-apply Type 1 (Org-Set) defaults from current org so a
        # cloned deal picks up the latest policy instead of stale source values.
        # Type 2 inherits from source via the verbatim column copy above.
        if user_id is not None and org_id is not None:
            from app.services.scenario_factory import force_type1_on_existing
            await session.flush()
            await force_type1_on_existing(
                session=session,
                scenario=None,
                inputs=new_inputs,
                user_id=user_id,
                org_id=org_id,
            )


@router.post("/ui/deals/{deal_id}/variant", response_class=HTMLResponse)
async def create_deal_copy(
    request: Request,
    deal_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Deep-copy a Scenario into a new Scenario with the same Projects, milestones, and line items."""
    from decimal import Decimal as _Dec
    source = await session.get(DealModel, deal_id)
    if source is None:
        return HTMLResponse("<p class='text-muted'>Deal not found.</p>", status_code=404)

    user = await _get_user(session, request)
    form = await request.form()
    variant_name = str(form.get("name", "")).strip() or f"{source.name} (Copy)"
    selected_project_ids = set(form.getlist("project_ids"))

    # New Scenario under same top-level Deal. Use the factory in Scenario-only
    # mode so org/user defaults (Type 1) get re-applied and the source's Type 2
    # values overlay them, while leaving Project + OperationalInputs creation
    # to the per-source-project clone logic below.
    from app.services.scenario_factory import create_scenario as _create_scenario
    # Resolve org_id from the source Deal so defaults resolve against the
    # right organization even when source's creator is no longer the caller.
    _src_deal = await session.get(Deal, source.deal_id)
    _org_id = _src_deal.org_id if _src_deal else (user.org_id if user else None)
    new_deal, _, _ = await _create_scenario(
        session=session,
        deal_id=source.deal_id,
        deal_type=source.project_type,
        user_id=user.id if user else None,
        org_id=_org_id,
        name=variant_name,
        version=source.version + 1,
        is_active=False,
        project_name=None,  # Scenario-only mode — clone loop creates Projects.
        source_scenario=source,
    )

    # Copy Projects (all if none selected, otherwise only checked ones)
    source_projects = list((await session.execute(
        select(Project).where(Project.scenario_id == deal_id).order_by(Project.created_at.asc())
    )).scalars())
    if selected_project_ids:
        source_projects = [p for p in source_projects if str(p.id) in selected_project_ids]

    # Track src_project_id → new_project_id so we can remap per-project FKs
    # (CapitalModuleProject junction, ProjectAnchor, waterfall_tiers) onto
    # the new projects below.
    project_id_map: dict = {}

    for src_proj in source_projects:
        new_proj = Project(
            scenario_id=new_deal.id,
            opportunity_id=src_proj.opportunity_id,
            name=src_proj.name,
            timeline_approved=src_proj.timeline_approved,
        )
        session.add(new_proj)
        await session.flush()
        project_id_map[src_proj.id] = new_proj.id

        await _copy_project_data(
            src_proj, new_proj, session,
            user_id=user.id if user else None,
            org_id=_org_id,
        )

    # Copy Scenario-level Capital modules + rebuild their project junction rows.
    from app.models.capital import CapitalModuleProject as _CMP
    from app.models.project import ProjectAnchor as _PA

    src_modules = list(
        (
            await session.execute(
                select(CapitalModule).where(
                    CapitalModule.scenario_id == deal_id
                )
            )
        ).scalars()
    )
    # Map old module_id → new module instance so we can attach junction rows
    # pointing at the copied Source.
    module_id_map: dict = {}
    for cm in src_modules:
        new_cm = CapitalModule(
            scenario_id=new_deal.id,
            label=cm.label,
            vehicle_type=cm.vehicle_type,
            equity_role=cm.equity_role,
            stack_position=cm.stack_position,
            source=cm.source,
            carry=cm.carry,
            exit_terms=cm.exit_terms,
            active_phase_start=cm.active_phase_start,
            active_phase_end=cm.active_phase_end,
        )
        session.add(new_cm)
        await session.flush()
        module_id_map[cm.id] = new_cm.id

    # Phase 3e: copy capital_module_projects junction rows. Without this the
    # copied Sources on the new Scenario are orphaned (no project links), and
    # the per-project loader in cashflow.py returns an empty list → the
    # engine silently skips sizing on every module.
    src_junctions = list(
        (
            await session.execute(
                select(_CMP).where(
                    _CMP.capital_module_id.in_(list(module_id_map.keys()))
                )
            )
        ).scalars()
    )
    for j in src_junctions:
        new_pid = project_id_map.get(j.project_id)
        if new_pid is None:
            continue  # project was excluded from copy; skip its junction
        new_mid = module_id_map.get(j.capital_module_id)
        if new_mid is None:
            continue
        session.add(
            _CMP(
                capital_module_id=new_mid,
                project_id=new_pid,
                amount=j.amount,
                active_from=j.active_from,
                active_to=j.active_to,
                active_from_offset_days=j.active_from_offset_days,
                active_to_offset_days=j.active_to_offset_days,
                auto_size=j.auto_size,
            )
        )

    # Copy ProjectAnchor rows — dormant in prod today (no rows exist), but
    # defensive for the future when cross-project timeline coupling lands.
    src_anchors = list(
        (
            await session.execute(
                select(_PA).where(
                    _PA.project_id.in_(list(project_id_map.keys()))
                )
            )
        ).scalars()
    )
    for a in src_anchors:
        new_anchor_pid = project_id_map.get(a.project_id)
        new_parent_pid = project_id_map.get(a.anchor_project_id)
        if new_anchor_pid is None or new_parent_pid is None:
            continue  # drop dangling anchors when parent project wasn't copied
        session.add(
            _PA(
                project_id=new_anchor_pid,
                anchor_project_id=new_parent_pid,
                anchor_milestone_id=None,  # milestone-id remap would need cross-walk; skip for v1
                offset_months=a.offset_months,
                offset_days=a.offset_days,
            )
        )

    # Copy Scenario-level Waterfall tiers (project_id remapped to new projects)
    for t in (await session.execute(
        select(WaterfallTier).where(WaterfallTier.scenario_id == deal_id)
    )).scalars():
        remapped_pid = project_id_map.get(t.project_id) if t.project_id else None
        # If the tier's capital_module_id points at an old module, remap to
        # the new one via module_id_map. Tiers with capital_module_id=None
        # (e.g. a sponsor-level promote tier) are copied as-is.
        remapped_mid = (
            module_id_map.get(t.capital_module_id)
            if t.capital_module_id
            else None
        )
        session.add(WaterfallTier(
            scenario_id=new_deal.id,
            project_id=remapped_pid,
            capital_module_id=remapped_mid,
            priority=t.priority, tier_type=t.tier_type,
            irr_hurdle_pct=t.irr_hurdle_pct,
            lp_split_pct=t.lp_split_pct, gp_split_pct=t.gp_split_pct,
            description=t.description,
            max_pct_of_distributable=t.max_pct_of_distributable,
            interest_rate_pct=t.interest_rate_pct,
        ))

    await session.commit()
    # Phase 3e: land on the Underwriting tab of the new variant so the user
    # sees the rolled-up view (with every project's staleness dot lit) and
    # knows to click Compute before any per-project edits.
    return RedirectResponse(
        url=f"/models/{new_deal.id}/builder?view=underwriting",
        status_code=303,
    )


@router.post("/ui/deals/{deal_id}/new-project", response_class=HTMLResponse)
async def create_deal_project(
    request: Request,
    deal_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Add a new Project to an existing Scenario (max 5). Redirects to builder with timeline wizard."""
    deal = await session.get(DealModel, deal_id)
    if deal is None:
        return HTMLResponse("<p class='text-muted'>Deal not found.</p>", status_code=404)

    project_count = int((await session.execute(
        select(func.count()).select_from(Project).where(Project.scenario_id == deal_id)
    )).scalar_one())
    if project_count >= 5:
        return HTMLResponse("<p class='text-muted'>Maximum 5 projects per deal.</p>", status_code=400)

    form = await request.form()
    project_name = str(form.get("name", "")).strip() or f"Project {project_count + 1}"

    try:
        pt = ProjectType(str(form.get("deal_type", "")))
    except ValueError:
        try:
            pt = ProjectType(str(deal.project_type))
        except ValueError:
            pt = ProjectType.acquisition

    # Required: opportunity_id. Tying every project to an opportunity is the
    # invariant that lets us seed a non-zero Acquisition UseLine — without it
    # the model has $0 of Uses and downstream debt sizing produces gaps that
    # can't be reconciled by a later edit.
    _opp_id_raw = str(form.get("opportunity_id", "")).strip()
    if not _opp_id_raw:
        return HTMLResponse(
            "<p class='text-muted'>Pick an opportunity to convert into this project.</p>",
            status_code=400,
        )
    try:
        _opp_id = UUID(_opp_id_raw)
    except ValueError:
        return HTMLResponse(
            "<p class='text-muted'>Invalid opportunity id.</p>", status_code=400,
        )

    opp = await session.get(Opportunity, _opp_id)
    if opp is None:
        return HTMLResponse(
            "<p class='text-muted'>Opportunity not found.</p>", status_code=404,
        )

    # Required: acquisition_cost > 0. Pre-filled from the opportunity's
    # listing price client-side, but always editable so the user can
    # override with their underwriting price. Without this we end up with
    # a $0 Acquisition UseLine and the same recompute-gap bug that motivated
    # this whole flow.
    _acq_raw = str(form.get("acquisition_cost", "")).strip()
    try:
        _acq_amount = Decimal(_acq_raw) if _acq_raw else Decimal("0")
    except (InvalidOperation, ValueError):
        return HTMLResponse(
            "<p class='text-muted'>Invalid acquisition cost.</p>", status_code=400,
        )
    if _acq_amount <= 0:
        return HTMLResponse(
            "<p class='text-muted'>Acquisition cost must be greater than zero.</p>",
            status_code=400,
        )

    new_proj = Project(
        scenario_id=deal_id,
        opportunity_id=_opp_id,
        name=project_name,
    )
    session.add(new_proj)
    await session.flush()

    await _auto_assign_opportunity_to_project(opp, new_proj, session)

    for milestone in _seed_milestones(new_proj, pt):
        session.add(milestone)
    await session.flush()

    # Seed the Acquisition UseLine using the user-confirmed cost. Mirrors the
    # pattern used at deal creation (project 1) so multi-project deals are
    # symmetric — every project lands with a populated Acquisition row.
    session.add(UseLine(
        project_id=new_proj.id,
        label=f"{opp.name or 'Property'} - Acquisition",
        phase=UseLinePhase.acquisition,
        cost_category="acquisition",
        amount=_acq_amount,
        timing_type="first_day",
    ))

    # ── Seed OperationalInputs with scenario-level sizing config ──────────
    # If the deal already has Deal Setup completed, the default project's
    # OperationalInputs holds the scenario-level debt + sizing config the
    # user filled in. Copy those onto the new project's inputs so the
    # engine respects the same DSCR cap, sizing mode, reserve months, etc.
    # Without this the new project silently falls back to gap-fill, leaving
    # debt over-leveraged and the Sources/Uses panel showing zero gap when
    # there should be one.
    _default_inputs = (await session.execute(
        select(OperationalInputs)
        .join(Project, Project.id == OperationalInputs.project_id)
        .where(
            Project.scenario_id == deal_id,
            Project.id != new_proj.id,
        )
        .order_by(Project.created_at.asc())
        .limit(1)
    )).scalar_one_or_none()
    if _default_inputs is not None:
        session.add(OperationalInputs(
            project_id=new_proj.id,
            debt_types=_default_inputs.debt_types,
            debt_structure=_default_inputs.debt_structure,
            debt_milestone_config=_default_inputs.debt_milestone_config,
            debt_sizing_mode=_default_inputs.debt_sizing_mode,
            construction_floor_pct=_default_inputs.construction_floor_pct,
            operation_reserve_months=_default_inputs.operation_reserve_months,
            deal_setup_complete=_default_inputs.deal_setup_complete,
        ))
        await session.flush()

    # ── Source share / clone decisions ────────────────────────────────────
    # Default: each project gets its OWN copy of every existing CapitalModule
    # (cloned with its own junction). User opts in to sharing a Source by
    # checking the box in the drawer — checked Sources add the new project
    # to the existing module's junction (one principal underwriting both).
    from app.models.capital import CapitalModule as _CM_create, CapitalModuleProject as _CMP_create
    _shared_ids: set[str] = {
        s.strip() for s in form.getlist("share_source_ids") if s and s.strip()
    }
    _existing_modules = list((await session.execute(
        select(_CM_create).where(_CM_create.scenario_id == deal_id)
    )).scalars())
    _max_pos = max((m.stack_position or 0) for m in _existing_modules) if _existing_modules else 0
    for _m in _existing_modules:
        if str(_m.id) in _shared_ids:
            # Share: add new project to existing junction (idempotent).
            _existing_j = (await session.execute(
                select(_CMP_create).where(
                    _CMP_create.capital_module_id == _m.id,
                    _CMP_create.project_id == new_proj.id,
                )
            )).scalar_one_or_none()
            if _existing_j is None:
                session.add(_CMP_create(
                    capital_module_id=_m.id,
                    project_id=new_proj.id,
                    amount=Decimal("0"),
                    active_from=_m.active_phase_start,
                    active_to=_m.active_phase_end,
                    auto_size=bool((_m.source or {}).get("auto_size")),
                ))
        else:
            # Clone: new CapitalModule + junction tied to new project only.
            # Reset auto-sized principal so the engine sizes for the new
            # project's own Uses; user-set amounts are preserved as a
            # starting point the user can edit afterwards.
            _src_copy = dict(_m.source or {})
            if _src_copy.get("auto_size"):
                _src_copy["amount"] = 0
            _max_pos += 1
            _new_mod = _CM_create(
                scenario_id=deal_id,
                label=_m.label,
                vehicle_type=_m.vehicle_type,
                equity_role=_m.equity_role,
                stack_position=_max_pos,
                source=_src_copy,
                carry=dict(_m.carry or {}),
                exit_terms=dict(_m.exit_terms or {}),
                active_phase_start=_m.active_phase_start,
                active_phase_end=_m.active_phase_end,
            )
            session.add(_new_mod)
            await session.flush()
            session.add(_CMP_create(
                capital_module_id=_new_mod.id,
                project_id=new_proj.id,
                amount=Decimal("0"),
                active_from=_new_mod.active_phase_start,
                active_to=_new_mod.active_phase_end,
                auto_size=bool((_new_mod.source or {}).get("auto_size")),
            ))
    await session.flush()

    # Optional Timeline Anchor — set when the user picks a parent milestone
    # in the Add Project drawer. anchor_project_id is derived from the
    # milestone's project_id so the form only needs one dropdown.
    _anchor_ms_raw = str(form.get("anchor_milestone_id", "")).strip()
    if _anchor_ms_raw:
        from app.models.milestone import Milestone as _AnchorMS
        from app.models.project import ProjectAnchor as _PA_create
        try:
            _anchor_ms_id = UUID(_anchor_ms_raw)
        except ValueError:
            _anchor_ms_id = None
        if _anchor_ms_id is not None:
            _pivot = await session.get(_AnchorMS, _anchor_ms_id)
            if _pivot is not None and _pivot.project_id != new_proj.id:
                # Confirm parent project belongs to this scenario
                _parent_proj = await session.get(Project, _pivot.project_id)
                if _parent_proj and _parent_proj.scenario_id == deal_id:
                    try:
                        _off_m = int(form.get("offset_months") or 0)
                    except (TypeError, ValueError):
                        _off_m = 0
                    try:
                        _off_d = int(form.get("offset_days") or 0)
                    except (TypeError, ValueError):
                        _off_d = 0
                    session.add(_PA_create(
                        project_id=new_proj.id,
                        anchor_project_id=_parent_proj.id,
                        anchor_milestone_id=_anchor_ms_id,
                        offset_months=_off_m,
                        offset_days=_off_d,
                    ))
                    await session.flush()

    # HTMX-driven submit (drawer form uses hx-post) — return HX-Redirect so
    # the browser does a full navigation to the new project. Plain
    # RedirectResponse retained as fallback for direct (non-HTMX) calls.
    if request.headers.get("HX-Request") == "true":
        from starlette.responses import Response as _StarletteResp
        _resp = _StarletteResp(status_code=204)
        _resp.headers["HX-Redirect"] = (
            f"/models/{deal_id}/builder?project={new_proj.id}"
        )
        return _resp
    return RedirectResponse(
        url=f"/models/{deal_id}/builder?project={new_proj.id}", status_code=303
    )


@router.post("/ui/projects/{project_id}/rename", response_class=HTMLResponse)
async def rename_project(
    request: Request,
    project_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Rename a Project. Returns the rendered breadcrumb span for HTMX swap."""
    project = await session.get(Project, project_id)
    if project is None:
        return HTMLResponse("<span>Project not found</span>", status_code=404)

    form = await request.form()
    new_name = str(form.get("name", "")).strip()
    if not new_name:
        new_name = project.name  # ignore empty submits, keep existing
    project.name = new_name[:255]
    await session.commit()

    safe = (
        new_name.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")
    )
    return HTMLResponse(
        f'<span class="proj-tab-name" '
        f'data-project-id="{project_id}" '
        f'onclick="event.preventDefault(); event.stopPropagation(); _startProjectRename(event)" '
        f'style="cursor:pointer" '
        f'title="Click to rename">{safe}</span>'
    )


@router.post("/ui/deals/{deal_id}/project/{project_id}/clone-from", response_class=HTMLResponse)
async def clone_project_from(
    deal_id: UUID,
    project_id: UUID,
    request: Request,
    session: DBSession,
) -> HTMLResponse:
    """Replace target project's financial data with a copy from another project in the same scenario."""
    from sqlalchemy import delete as sa_delete

    form = await request.form()
    source_project_id_raw = str(form.get("source_project_id", "")).strip()
    if not source_project_id_raw:
        return HTMLResponse("<p class='text-muted'>No source project selected.</p>", status_code=400)
    try:
        source_project_id = UUID(source_project_id_raw)
    except ValueError:
        return HTMLResponse("<p class='text-muted'>Invalid source project ID.</p>", status_code=400)

    target_proj = await session.get(Project, project_id)
    source_proj = await session.get(Project, source_project_id)

    if target_proj is None or source_proj is None:
        return HTMLResponse("<p class='text-muted'>Project not found.</p>", status_code=404)
    if target_proj.scenario_id != deal_id or source_proj.scenario_id != deal_id:
        return HTMLResponse("<p class='text-muted'>Projects must belong to the same scenario.</p>", status_code=400)
    if target_proj.id == source_proj.id:
        return HTMLResponse("<p class='text-muted'>Cannot clone a project onto itself.</p>", status_code=400)

    # Clear existing data on target
    await session.execute(sa_delete(Milestone).where(Milestone.project_id == project_id))
    await session.execute(sa_delete(UseLine).where(UseLine.project_id == project_id))
    await session.execute(sa_delete(IncomeStream).where(IncomeStream.project_id == project_id))
    await session.execute(sa_delete(OperatingExpenseLine).where(OperatingExpenseLine.project_id == project_id))
    if target_proj:
        target_proj.unit_mix = []
        session.add(target_proj)
    await session.execute(sa_delete(OperationalInputs).where(OperationalInputs.project_id == project_id))
    await session.flush()

    # Copy from source
    await _copy_project_data(source_proj, target_proj, session)
    await session.flush()

    return RedirectResponse(url=f"/models/{deal_id}/builder?project={project_id}", status_code=303)


@router.post("/ui/deals/{deal_id}/project/{project_id}/delete")
async def delete_deal_project(
    deal_id: UUID,
    project_id: UUID,
    session: DBSession,
) -> RedirectResponse:
    """Delete a Project and all its child rows from a Scenario.

    `deal_id` here is the Scenario id (URL kept consistent with the rest
    of the model-builder routes). The deal must keep at least one project,
    so deletion is rejected when only one remains.
    """
    from sqlalchemy import delete as sa_delete
    from app.models.portfolio import GanttEntry
    from app.models.parcel import ParcelTransformation
    from app.models.org import ProjectVisibility
    from app.models.project import PermitStub

    proj = await session.get(Project, project_id)
    if proj is None or proj.scenario_id != deal_id:
        return RedirectResponse(url=f"/models/{deal_id}/builder", status_code=303)

    project_count = int((await session.execute(
        select(func.count()).select_from(Project).where(Project.scenario_id == deal_id)
    )).scalar_one())
    if project_count <= 1:
        return RedirectResponse(
            url=f"/models/{deal_id}/builder?project={project_id}", status_code=303
        )

    # Non-CASCADE child tables: delete explicitly. Most other children
    # (capital_modules, draw_sources, waterfall_tiers, cash_flows,
    # cash_flow_line_items, operational_outputs) cascade from projects.id
    # ondelete=CASCADE.
    await session.execute(sa_delete(UseLine).where(UseLine.project_id == project_id))
    await session.execute(sa_delete(IncomeStream).where(IncomeStream.project_id == project_id))
    await session.execute(sa_delete(OperatingExpenseLine).where(OperatingExpenseLine.project_id == project_id))
    await session.execute(sa_delete(OperationalInputs).where(OperationalInputs.project_id == project_id))
    # Defensive deletes for tables whose FK isn't ondelete=CASCADE everywhere.
    await session.execute(sa_delete(PortfolioProject).where(PortfolioProject.project_id == project_id))
    await session.execute(sa_delete(GanttEntry).where(GanttEntry.project_id == project_id))
    await session.execute(sa_delete(ParcelTransformation).where(ParcelTransformation.project_id == project_id))
    await session.execute(sa_delete(ProjectVisibility).where(ProjectVisibility.project_id == project_id))
    await session.execute(sa_delete(PermitStub).where(PermitStub.project_id == project_id))
    # Milestones: must delete via raw SQL BEFORE session.delete(proj) so the
    # ORM doesn't try to null out project_id on cached milestone rows (which
    # would violate ck_milestones_single_parent — both parent FKs null).
    # Trigger-chain refs from sibling projects' milestones are auto-nulled by
    # the DB (trigger_milestone_id FK is ondelete=SET NULL).
    await session.execute(sa_delete(Milestone).where(Milestone.project_id == project_id))
    await session.flush()
    # Expire any cached state on `proj` so the ORM doesn't replay a stale
    # milestones collection during the delete cascade.
    await session.refresh(proj)

    await session.delete(proj)
    await session.commit()

    return RedirectResponse(url=f"/models/{deal_id}/builder", status_code=303)


@router.get("/ui/deals/{deal_id}/add-project/search", response_class=HTMLResponse)
async def add_project_search(
    deal_id: UUID,
    request: Request,
    session: DBSession,
    q: str = Query(default=""),
) -> HTMLResponse:
    """HTMX search for the Add-Project drawer. Mirrors the wizard step 2
    parcel/opp search UX: returns one best-match opportunity card.

    Eligibility: any opp the user can see (their org OR scraped/null-org),
    excluding archived rows. Flags rows already bound to a project on this
    scenario so the user gets a clear 'already used' message instead of a
    silent re-add attempt.
    """
    if not q or len(q.strip()) < 3:
        return HTMLResponse("")

    user = await _get_user(session, request)
    if user is None:
        return HTMLResponse("")

    q_clean = q.strip()
    q_lower = q_clean.lower()

    bound_rows = (await session.execute(
        select(Project.opportunity_id).where(Project.scenario_id == deal_id)
    )).all()
    bound_ids = {r[0] for r in bound_rows if r[0] is not None}

    stmt = (
        select(Opportunity)
        .where(
            ((Opportunity.org_id == user.org_id) | (Opportunity.org_id.is_(None))),
            Opportunity.archived.is_(False),
            (
                Opportunity.address_normalized.ilike(f"%{q_clean}%")
                | Opportunity.street.ilike(f"%{q_clean}%")
                | Opportunity.name.ilike(f"%{q_clean}%")
                | Opportunity.listing_name.ilike(f"%{q_clean}%")
                | Opportunity.apn.ilike(f"%{q_clean}%")
            ),
        )
        .order_by(
            Opportunity.org_id.desc().nullslast(),
            Opportunity.last_seen_at.desc().nullslast(),
        )
        .limit(25)
    )
    candidates = list((await session.execute(stmt)).scalars().unique())

    matched: Opportunity | None = None
    for c in candidates:
        haystack = " ".join([
            (c.address_normalized or ""),
            (c.street or ""),
            (c.name or ""),
            (c.listing_name or ""),
            (c.apn or ""),
        ]).lower()
        if q_lower in haystack:
            matched = c
            break
    if matched is None and candidates:
        matched = candidates[0]

    return templates.TemplateResponse(request, "partials/add_project_search_match.html", {
        "request": request,
        "match": matched,
        "already_bound": (matched is not None and matched.id in bound_ids),
    })


@router.post("/ui/deals/{deal_id}/split-projects", response_class=HTMLResponse)
async def split_multiparcel_projects(
    deal_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Split a multi-APN listing into one Project per parcel in the same Scenario."""
    import re as _re

    scenario = await session.get(DealModel, deal_id)
    if scenario is None:
        return HTMLResponse("")

    # Get existing single project
    existing_proj = (await session.execute(
        select(Project).where(Project.scenario_id == deal_id).order_by(Project.created_at.asc()).limit(1)
    )).scalar_one_or_none()
    if existing_proj is None:
        return HTMLResponse("")

    # Opportunity IS the listing — APN is directly on the opportunity row
    opp = await session.get(Opportunity, existing_proj.opportunity_id) if existing_proj.opportunity_id else None
    if opp is None:
        return HTMLResponse("")

    if not opp.apn or not _re.search(r"[,;]", opp.apn):
        return HTMLResponse("")

    apns = [a.strip() for a in _re.split(r"[,;]", opp.apn) if a.strip()]
    if len(apns) < 2:
        return HTMLResponse("")

    try:
        pt = ProjectType(scenario.project_type) if scenario and scenario.project_type else ProjectType.acquisition
    except ValueError:
        pt = ProjectType.acquisition

    # Rename existing project to include first APN and seed its parcel assignment
    existing_proj.name = f"Project 1 — {apns[0]}"
    await _auto_assign_opportunity_to_project(opp, existing_proj, session)

    # Parcel lookup helper: find parcel by APN for per-project scoping
    async def _parcel_for_apn(apn: str) -> "Parcel | None":
        return (await session.execute(
            select(Parcel).where(Parcel.apn == apn).limit(1)
        )).scalar_one_or_none()

    # Create one new project per remaining APN
    for i, apn in enumerate(apns[1:], start=2):
        proj_count = await session.scalar(
            select(func.count()).select_from(Project).where(Project.scenario_id == deal_id)
        ) or 0
        if proj_count >= 5:
            break
        new_proj = Project(
            scenario_id=deal_id,
            opportunity_id=existing_proj.opportunity_id,
            name=f"Project {i} — {apn}",
        )
        session.add(new_proj)
        await session.flush()
        await _auto_assign_opportunity_to_project(opp, new_proj, session)
        parcel = await _parcel_for_apn(apn)
        if parcel:
            new_proj.parcel_id = parcel.id
        for milestone in _seed_milestones(new_proj, pt):
            session.add(milestone)

    await session.flush()

    return RedirectResponse(url=f"/models/{deal_id}/builder", status_code=303)


@router.post("/ui/deals/{deal_id}/dismiss-multiparcel", response_class=HTMLResponse)
async def dismiss_multiparcel_banner(
    deal_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Suppress the multi-parcel banner for this opportunity."""
    # multi_parcel_dismissed was on the old Opportunity entity — no longer supported.
    return HTMLResponse("")  # replaces the banner with nothing


@router.post("/ui/models/{model_id}/stack-order", response_class=HTMLResponse)
async def save_stack_order(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Batch-update stack_position for all capital modules from a confirm-order form submission."""
    form = await request.form()
    modules = list((await session.execute(
        select(CapitalModule).where(CapitalModule.scenario_id == model_id)
    )).scalars())
    for m in modules:
        key = f"pos_{m.id}"
        if val := _fi(form.get(key), None):
            m.stack_position = val
    await session.flush()
    _active_proj_id = await _active_project_from_request(request, session, model_id)
    ctx = await _load_builder_data(session, model_id, project_id=_active_proj_id)
    ctx["request"] = request
    ctx["active_module"] = "sources"
    return templates.TemplateResponse(request, "partials/model_builder_panel.html", ctx)


@router.post("/ui/models/{model_id}/capital-modules/reorder", response_class=HTMLResponse)
async def reorder_capital_modules(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Drag-reorder capital modules — receives ordered list of IDs, assigns stack_position 1..N."""
    form = await request.form()
    ordered_ids = form.getlist("order")
    for i, id_str in enumerate(ordered_ids, start=1):
        try:
            mod = await session.get(CapitalModule, UUID(id_str))
            if mod and mod.scenario_id == model_id:
                mod.stack_position = i
        except (ValueError, Exception):
            pass
    await session.flush()
    _active_proj_id = await _active_project_from_request(request, session, model_id)
    ctx = await _load_builder_data(session, model_id, project_id=_active_proj_id)
    ctx["request"] = request
    ctx["active_module"] = "sources"
    model = await session.get(DealModel, model_id)
    return templates.TemplateResponse(
        request, "partials/model_builder_panel.html",
        {"model": model, "active_module": "sources", **ctx}
    )


@router.post("/ui/models/{model_id}/settings", response_class=HTMLResponse)
async def save_model_settings(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Save model name, project type, and scalar operational inputs from the Settings drawer."""
    form = await request.form()
    name = str(form.get("name", "")).strip()
    deal_type_raw = str(form.get("deal_type", "")).strip()
    expense_growth = form.get("expense_growth_rate_pct_annual")
    exit_cap = form.get("exit_cap_rate_pct")
    going_in_cap = form.get("going_in_cap_rate_pct")
    capex_reserve = form.get("capex_reserve_per_unit_annual")
    risk_free_rate = form.get("risk_free_rate_pct")
    discount_rate = form.get("discount_rate_pct")
    hold_period = form.get("hold_period_years")
    debt_structure = str(form.get("debt_structure") or "").strip() or None
    debt_sizing_mode = str(form.get("debt_sizing_mode") or "").strip() or None
    dscr_minimum = form.get("dscr_minimum")
    operation_reserve_months = form.get("operation_reserve_months")
    perm_rate_pct = form.get("perm_rate_pct")
    construction_rate_pct = form.get("construction_rate_pct")
    perm_amort_years = form.get("perm_amort_years")

    deal = await session.get(DealModel, model_id)
    if deal is None:
        return HTMLResponse("<p class='text-muted'>Not found.</p>", status_code=404)

    if name:
        deal.name = name
    if deal_type_raw:
        try:
            deal.project_type = ProjectType(deal_type_raw)
        except ValueError:
            pass
    if risk_free_rate is not None:
        try:
            deal.risk_free_rate_pct = float(risk_free_rate)
        except (ValueError, TypeError):
            pass
    if discount_rate is not None:
        try:
            deal.discount_rate_pct = float(discount_rate)
        except (ValueError, TypeError):
            pass

    await session.flush()

    # Update OperationalInputs for the default project
    default_project = (await session.execute(
        select(Project).where(Project.scenario_id == model_id).order_by(Project.created_at.asc()).limit(1)
    )).scalar_one_or_none()
    if default_project:
        inputs = (await session.execute(
            select(OperationalInputs).where(OperationalInputs.project_id == default_project.id)
        )).scalar_one_or_none()
        if inputs is None:
            inputs = OperationalInputs(project_id=default_project.id)
            session.add(inputs)

        if expense_growth is not None:
            try:
                inputs.expense_growth_rate_pct_annual = float(expense_growth)
            except (ValueError, TypeError):
                pass
        if exit_cap is not None:
            try:
                inputs.exit_cap_rate_pct = float(exit_cap)
            except (ValueError, TypeError):
                pass
        if going_in_cap is not None:
            try:
                inputs.going_in_cap_rate_pct = float(going_in_cap)
            except (ValueError, TypeError):
                pass
        if capex_reserve is not None:
            try:
                inputs.capex_reserve_per_unit_annual = float(capex_reserve)
            except (ValueError, TypeError):
                pass
        if hold_period is not None:
            try:
                hold_years = float(hold_period)
                # Hold period now sourced from operation_stabilized milestone
                # duration AND per-perm-debt CapitalModule.source.hold_term_years.
                from app.models.milestone import Milestone as _Milestone
                stabilized_ms = (await session.execute(
                    select(_Milestone).where(
                        _Milestone.project_id == default_project.id,
                        _Milestone.milestone_type == "operation_stabilized",
                    )
                )).scalar_one_or_none()
                if stabilized_ms is not None:
                    stabilized_ms.duration_days = round(hold_years * 365)
                # Mirror to all perm-debt modules so engine resolver picks it up.
                hold_int = max(1, int(round(hold_years)))
                _perm_mods = list((await session.execute(
                    select(CapitalModule).where(CapitalModule.scenario_id == model_id)
                )).scalars())
                for _cm in _perm_mods:
                    if str(getattr(_cm, "vehicle_type", "") or "").replace("VehicleType.", "") != "debt":
                        continue
                    _src = dict(_cm.source or {})
                    _src["hold_term_years"] = hold_int
                    _cm.source = _src
                    session.add(_cm)
                # Mirror to wizard staging (inputs.debt_terms) so re-opening
                # the Deal Setup wizard reflects the latest value instead of
                # the stale staging dict from initial setup.
                _dt = dict(inputs.debt_terms or {})
                _pd_entry = dict(_dt.get("permanent_debt", {}))
                _pd_entry["hold_term_years"] = hold_int
                _dt["permanent_debt"] = _pd_entry
                inputs.debt_terms = _dt
            except (ValueError, TypeError):
                pass

        if debt_structure:
            inputs.debt_structure = debt_structure
        if debt_sizing_mode:
            inputs.debt_sizing_mode = debt_sizing_mode
        if dscr_minimum:
            try:
                _dscr_val = float(dscr_minimum)
                # Per-perm-debt dscr_min replaces deal-level dscr_minimum.
                _dscr_mods = list((await session.execute(
                    select(CapitalModule).where(CapitalModule.scenario_id == model_id)
                )).scalars())
                for _cm in _dscr_mods:
                    if str(getattr(_cm, "vehicle_type", "") or "").replace("VehicleType.", "") != "debt":
                        continue
                    _src = dict(_cm.source or {})
                    _src["dscr_min"] = _dscr_val
                    _cm.source = _src
                    session.add(_cm)
                # Mirror to wizard staging.
                _dt = dict(inputs.debt_terms or {})
                _pd_entry = dict(_dt.get("permanent_debt", {}))
                _pd_entry["dscr_min"] = _dscr_val
                _dt["permanent_debt"] = _pd_entry
                inputs.debt_terms = _dt
            except (ValueError, TypeError):
                pass
        if operation_reserve_months:
            try:
                inputs.operation_reserve_months = int(operation_reserve_months)
            except Exception:
                pass
        _ahp_val = (form.get("affordable_housing_project") == "1")
        inputs.affordable_housing_project = _ahp_val
        # Scenario-level toggle — propagate to every project's OperationalInputs
        # so the flag stays consistent regardless of which project tab is
        # active when the Settings drawer is reopened.
        _sibling_oi = list((await session.execute(
            select(OperationalInputs)
            .join(Project, Project.id == OperationalInputs.project_id)
            .where(Project.scenario_id == model_id)
            .where(OperationalInputs.project_id != default_project.id)
        )).scalars())
        for _oi in _sibling_oi:
            _oi.affordable_housing_project = _ahp_val
        # Sync auto-sized CapitalModules with rate / amort form fields
        # (deal-level OperationalInputs.debt_terms is no longer authoritative).
        if any([perm_rate_pct, construction_rate_pct, perm_amort_years, debt_structure]):
            auto_mods = list((await session.execute(
                select(CapitalModule).where(CapitalModule.scenario_id == model_id)
            )).scalars())
            for cm in auto_mods:
                src = cm.source or {}
                if not src.get("auto_size"):
                    continue
                src = dict(src)
                carry = dict(cm.carry or {})
                from app.engines.cashflow import _loan_subtype_from_module as _ls_fn
                ft = _ls_fn(cm)
                if ft in ("bond",) and perm_rate_pct:
                    src["interest_rate_pct"] = float(perm_rate_pct)
                    if "phases" in carry:
                        for ph in carry["phases"]:
                            if ph.get("name") == "operation":
                                if perm_rate_pct:
                                    ph["io_rate_pct"] = float(perm_rate_pct)
                                if perm_amort_years:
                                    ph["amort_term_years"] = int(perm_amort_years)
                            elif ph.get("name") == "construction" and construction_rate_pct:
                                ph["io_rate_pct"] = float(construction_rate_pct)
                elif ft in ("permanent_debt",) and perm_rate_pct:
                    src["interest_rate_pct"] = float(perm_rate_pct)
                    if perm_amort_years:
                        carry["amort_term_years"] = int(perm_amort_years)
                elif ft in ("construction_loan",) and construction_rate_pct:
                    src["interest_rate_pct"] = float(construction_rate_pct)
                cm.source = src
                cm.carry = carry
                session.add(cm)

    _ht: dict[str, float] = {}
    for _key in ("ht_occ_green", "ht_oer_green", "ht_dscr_green", "ht_margin_green"):
        _raw = form.get(_key)
        if _raw is not None:
            try:
                _ht[_key.removeprefix("ht_")] = float(_raw)
            except (ValueError, TypeError):
                pass
    if _ht:
        deal.health_thresholds = {**(deal.health_thresholds or {}), **_ht}

    await session.commit()

    return RedirectResponse(url=f"/models/{model_id}/builder", status_code=303)


@router.post("/ui/projects/{project_id}/timeline-wizard", response_class=HTMLResponse)
async def timeline_wizard_submit(
    request: Request,
    project_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Process wizard form: clear seeded milestones, create anchor + selected milestones."""
    from sqlalchemy import delete as sa_delete
    from datetime import date as _date

    proj = await session.get(Project, project_id)
    if proj is None:
        return HTMLResponse("<p class='text-muted'>Project not found.</p>", status_code=404)

    form = await request.form()
    anchor_type_raw = str(form.get("anchor_type", ""))
    anchor_date_raw = str(form.get("anchor_date", ""))
    anchor_duration_raw = str(form.get("anchor_duration_days", "0"))
    selected_types = form.getlist("milestone_types")  # includes anchor
    new_name = str(form.get("new_name", "")).strip()
    new_deal_type_raw = str(form.get("new_deal_type", "")).strip()

    # If name/type provided (new deal wizard step 0), update deal records
    if new_name or new_deal_type_raw:
        scenario = await session.get(DealModel, proj.scenario_id)
        if scenario:
            if new_deal_type_raw:
                try:
                    new_dt = ProjectType(new_deal_type_raw)
                    scenario.project_type = new_dt
                except ValueError:
                    pass
            if new_name:
                if scenario.deal_id:
                    deal_obj = await session.get(Deal, scenario.deal_id)
                    if deal_obj:
                        deal_obj.name = new_name
                if proj.opportunity_id:
                    opp_obj = await session.get(Opportunity, proj.opportunity_id)
                    if opp_obj:
                        opp_obj.name = new_name

    try:
        anchor_mt = MilestoneType(anchor_type_raw)
    except ValueError:
        return HTMLResponse("<p class='text-muted'>Invalid anchor type.</p>", status_code=400)

    try:
        anchor_date = _date.fromisoformat(anchor_date_raw.strip()[:10])
    except (ValueError, AttributeError):
        return HTMLResponse("<p class='text-muted'>Invalid start date.</p>", status_code=400)

    try:
        anchor_duration = max(0, int(anchor_duration_raw))
    except (ValueError, TypeError):
        anchor_duration = 0

    _STABILIZED_AUTO_DAYS = 10950  # 30 years — applied when no divestment milestone
    # Hardcoded default durations for acquisition-phase milestones when the
    # user hasn't supplied an override.  Trigger chain (Pass 2 below) wires
    # these in submitted order, so "Close Starts After Under Contract" etc.
    # falls out automatically whenever the predecessor is present.
    _ACQUISITION_DEFAULT_DAYS: dict[str, int] = {
        "offer_made":     7,
        "under_contract": 30,
        "close":          30,
    }

    # Clear existing milestones for this project
    await session.execute(sa_delete(Milestone).where(Milestone.project_id == project_id))
    await session.flush()

    has_divestment = "divestment" in selected_types
    valid_types = {mt.value for mt in MilestoneType}

    # Filter + de-dupe while preserving submitted order (the canonical CRE
    # timeline order the user picked in the UI).  Unknown types are skipped.
    ordered_types: list[str] = []
    seen: set[str] = set()
    for mt_str in selected_types:
        if mt_str in valid_types and mt_str not in seen:
            ordered_types.append(mt_str)
            seen.add(mt_str)

    # Two-pass creation so we can build a trigger chain.
    # Pass 1: instantiate every milestone with its duration + target_date
    # on the anchor.  Pass 2: assign trigger_milestone_id so each non-anchor
    # milestone starts at the end of the previous one in submitted order.
    # Without the trigger chain, computed_start() returns None for non-
    # anchor milestones, _milestone_dates_from_orm skips them, and the
    # cashflow engine falls back to the legacy OperationalInputs scalar
    # fields (which are NULL on wizard-created deals) → every phase
    # defaults to 1 month and the carry-type math collapses.
    created: list[Milestone] = []
    for seq, mt_str in enumerate(ordered_types):
        mt = MilestoneType(mt_str)
        is_anchor = mt == anchor_mt

        # Per-milestone duration override via ``duration_{type}=N`` form field.
        override_raw = form.get(f"duration_{mt_str}")
        if override_raw is not None and str(override_raw).strip() != "":
            try:
                dur = max(0, int(override_raw))
            except (ValueError, TypeError):
                dur = 0
        elif mt == MilestoneType.operation_stabilized and not has_divestment:
            # Auto-cap stabilized at 30 years when no divestment
            dur = _STABILIZED_AUTO_DAYS
        elif mt == MilestoneType.divestment:
            # Divestment is a single-day event (sale closing date)
            dur = 1
        elif is_anchor:
            dur = anchor_duration
        elif mt_str in _ACQUISITION_DEFAULT_DAYS:
            dur = _ACQUISITION_DEFAULT_DAYS[mt_str]
        else:
            dur = 0

        row = Milestone(
            project_id=project_id,
            milestone_type=mt,
            target_date=anchor_date if is_anchor else None,
            duration_days=dur,
            sequence_order=seq,
        )
        session.add(row)
        created.append(row)

    # Flush so every Milestone gets a primary key before we wire trigger refs.
    await session.flush()

    # Pass 2: build the trigger chain in submitted order.  Each non-anchor
    # milestone triggers off the previous one with offset=0 so its start date
    # equals the prior milestone's end date (prev.start + prev.duration_days).
    prev: Milestone | None = None
    for row in created:
        if row.milestone_type == anchor_mt:
            prev = row
            continue
        if prev is not None:
            row.trigger_milestone_id = prev.id
            row.trigger_offset_days = 0
        prev = row

    await session.commit()
    _tw_url = f"/models/{proj.scenario_id}/builder?project={project_id}&module=timeline"
    if str(form.get("_wizard", "")).strip() == "1":
        _tw_url += "&wizard=1"
    return RedirectResponse(url=_tw_url, status_code=303)


@router.post("/ui/projects/{project_id}/approve-timeline", response_class=HTMLResponse)
async def approve_timeline(
    request: Request,
    project_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Toggle timeline approval on the dev Project.

    Normal POST → approve (redirect to sources).
    POST with _unapprove=1 → re-open (redirect back to timeline).
    POST with _wizard=1 → single-flow deal-creation wizard mode; redirect into
    the setup wizard with wizard chrome still active.
    """
    proj = await session.get(Project, project_id)
    if proj is None:
        return HTMLResponse("<p class='text-muted'>Project not found.</p>", status_code=404)
    form = await request.form()
    unapprove = str(form.get("_unapprove", "")).strip() == "1"
    wizard_mode = str(form.get("_wizard", "")).strip() == "1"
    proj.timeline_approved = not unapprove
    await session.commit()
    if unapprove:
        _unapp_url = f"/models/{proj.scenario_id}/builder?project={project_id}&module=timeline"
        if wizard_mode:
            _unapp_url += "&wizard=1"
        return RedirectResponse(url=_unapp_url, status_code=303)
    if wizard_mode:
        return RedirectResponse(
            url=f"/models/{proj.scenario_id}/builder?project={project_id}&module=deal_setup&wizard=1",
            status_code=303,
        )
    return RedirectResponse(url=f"/models/{proj.scenario_id}/builder?project={project_id}&module=sources", status_code=303)


async def _wizard_active_from_options(
    session: "AsyncSession",
    default_project: "Project | None",
) -> list[tuple[str, str]]:
    """Build the Active From dropdown options for wizard step 3 from the
    default project's seeded milestones. Each option is (key, label) where
    `key` is the milestone_type the wizard submits (resolved to a CapitalModule
    milestone FK during finalize). Labels fall back to a humanized form of the
    milestone_type when no override label is set.
    """
    if default_project is None:
        return []
    rows = list((await session.execute(
        select(Milestone)
        .where(Milestone.project_id == default_project.id)
        .order_by(Milestone.sequence_order, Milestone.id)
    )).scalars())
    if not rows:
        return []
    out: list[tuple[str, str]] = []
    for m in rows:
        mt = str(getattr(m, "milestone_type", "") or "").replace("MilestoneType.", "")
        label = m.label or _milestone_label(mt)
        out.append((mt, label))
    return out


async def _wizard_phases_present(
    session: "AsyncSession",
    default_project: "Project | None",
) -> set[str]:
    """Return the set of MilestoneType keys present on the default project.

    Used by Step 2 of the setup wizard to filter the debt-card list:
      - construction_loan / construction_to_perm require a Construction milestone
      - pre_development_loan requires a Pre-Development milestone

    Cards whose required phase is absent are hidden (not disabled) so the picker
    stays uncluttered. Re-derived on every wizard GET/POST — milestones are the
    source of truth, no persisted allow-list.
    """
    if default_project is None:
        return set()
    rows = list((await session.execute(
        select(Milestone.milestone_type).where(Milestone.project_id == default_project.id)
    )).scalars())
    out: set[str] = set()
    for mt in rows:
        out.add(str(mt or "").replace("MilestoneType.", ""))
    return out


async def _seed_wizard_perm_defaults(inputs: "OperationalInputs", session: "DBSession", request: "Request") -> None:
    """Seed permanent-debt staging from user/org resolved defaults when not already set."""
    _wiz_user = await _get_user(session, request)
    if _wiz_user is None:
        return
    from app.settings.resolver import resolve_all_defaults as _resolve_all
    _wiz_defs = await _resolve_all(_wiz_user.id, _wiz_user.org_id, session)
    _wiz_dt = dict(inputs.debt_terms or {})
    _wiz_pd = dict(_wiz_dt.get("permanent_debt", {}))
    _wiz_dirty = False
    for _def_key, _staging_key, _cast in (
        ("hold_term_years", "hold_term_years", int),
        ("amort_term_years", "amort_years", int),
        ("ltv_pct", "ltv_pct", float),
        ("dscr_min", "dscr_min", float),
    ):
        if _staging_key not in _wiz_pd:
            _raw = _wiz_defs.get(_def_key)
            if _raw:
                try:
                    _wiz_pd[_staging_key] = _cast(_raw)
                    _wiz_dirty = True
                except (ValueError, TypeError):
                    pass
    if _wiz_dirty:
        _wiz_dt["permanent_debt"] = _wiz_pd
        inputs.debt_terms = _wiz_dt
        session.add(inputs)
        await session.flush()
    # Org-Set debt_sizing_mode reflects org policy when no value yet on inputs.
    if not inputs.debt_sizing_mode:
        _dsm_default = _wiz_defs.get("debt_sizing_mode")
        if _dsm_default:
            inputs.debt_sizing_mode = _dsm_default
            session.add(inputs)
            await session.flush()


async def _wizard_debt_vehicles(session: "AsyncSession", user) -> list[dict]:
    """Debt+forgivable_loan Source Vehicles visible to this user (for Step 2 picker)."""
    from app.models.source_vehicle import SourceVehicle as _SV_wiz
    rows = (await session.execute(
        select(_SV_wiz).where(
            _SV_wiz.vehicle_type.in_(["debt", "forgivable_loan"]),
            (
                ((_SV_wiz.scope == "org") & (_SV_wiz.owner_id == user.org_id)) |
                ((_SV_wiz.scope == "user") & (_SV_wiz.owner_id == user.id))
            ),
        ).order_by(_SV_wiz.label)
    )).scalars().all()
    return [{"id": str(v.id), "name": v.label, "vehicle_type": v.vehicle_type} for v in rows]


@router.get("/ui/models/{model_id}/setup", response_class=HTMLResponse)
async def deal_setup_wizard_get(
    request: Request,
    model_id: UUID,
    session: DBSession,
    step: int = Query(default=-1),
) -> HTMLResponse:
    """Render a single wizard step fragment (used by Back buttons and direct links)."""
    model = await session.get(DealModel, model_id)
    if model is None:
        return HTMLResponse("Not found", status_code=404)
    default_project = (await session.execute(
        select(Project).where(Project.scenario_id == model_id).order_by(Project.created_at).limit(1)
    )).scalar_one_or_none()
    inputs = (await session.execute(
        select(OperationalInputs).where(OperationalInputs.project_id == default_project.id)
    )).scalar_one_or_none() if default_project else None

    # If step not explicitly requested, start at step 1.
    if step == -1:
        step = 1

    # Sync wizard staging (inputs.debt_terms) from live CapitalModule.source so
    # re-opening the wizard reflects edits made via Settings drawer or Edit
    # Source drawer between setups. Read-only mirror — engine still uses
    # CapitalModule.source as source of truth.
    if inputs is not None:
        _perm = (await session.execute(
            select(CapitalModule).where(
                CapitalModule.scenario_id == model_id,
                CapitalModule.vehicle_type == "debt",
            ).limit(1)
        )).scalar_one_or_none()
        if _perm is not None and isinstance(_perm.source, dict):
            _dt = dict(inputs.debt_terms or {})
            _pd_entry = dict(_dt.get("permanent_debt", {}))
            if "hold_term_years" in _perm.source:
                _pd_entry["hold_term_years"] = _perm.source["hold_term_years"]
            if "dscr_min" in _perm.source:
                _pd_entry["dscr_min"] = _perm.source["dscr_min"]
            _dt["permanent_debt"] = _pd_entry
            inputs.debt_terms = _dt
            session.add(inputs)
            await session.flush()

    # Seed perm-debt staging from user/org resolved defaults when not already set.
    if inputs is not None:
        await _seed_wizard_perm_defaults(inputs, session, request)
    _wiz_user = await _get_user(session, request)
    _svd = await _wizard_debt_vehicles(session, _wiz_user) if _wiz_user else []
    _wiz_active_from_opts = await _wizard_active_from_options(session, default_project)

    # Review (Step 6) Back button: jump to Step 2 when every selected debt has
    # a vehicle (Steps 3-5 were skipped). Otherwise return to Step 5.
    _review_back_step = 5
    if step == 6 and inputs is not None:
        _selected = inputs.debt_types or []
        _dt_now = inputs.debt_terms or {}
        if _selected and all((_dt_now.get(_ft) or {}).get("vehicle_id") for _ft in _selected):
            _review_back_step = 2

    _wiz_phases_present = await _wizard_phases_present(session, default_project)
    return templates.TemplateResponse(request, "partials/deal_setup_wizard.html", {
        "request": request, "model": model, "inputs": inputs, "step": step,
        "source_vehicles_debt": _svd,
        "wizard_active_from_opts": _wiz_active_from_opts,
        "review_back_step": _review_back_step,
        "phases_present": _wiz_phases_present,
    })


async def _prefill_noi_from_listing(
    model: "DealModel",
    default_project: "Project",
    inputs: "OperationalInputs",
    session: "AsyncSession",
) -> None:
    """If the project's linked opportunity has NOI data, pre-fill
    OperationalInputs.noi_stabilized_input. Does nothing if already set."""
    if inputs.noi_stabilized_input is not None:
        return  # already set — don't overwrite a previous entry
    # Opportunity IS the listing — read NOI directly from Project.opportunity_id
    if default_project.opportunity_id is None:
        return
    listing = await session.get(ScrapedListing, default_project.opportunity_id)
    if listing is None:
        return
    noi_value = listing.proforma_noi if listing.proforma_noi is not None else listing.noi
    if noi_value is not None:
        inputs.noi_stabilized_input = noi_value
        inputs.noi_auto_seeded = True
        session.add(inputs)


@router.post("/ui/models/{model_id}/setup/step", response_class=HTMLResponse)
async def deal_setup_wizard_step(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Save a wizard step's data and return the next step fragment."""
    form = await request.form()
    step = int(form.get("step", 1))
    # Field-level validation errors collected during the step handler.
    # Keyed by vehicle_type (or "_form" for cross-cutting errors).  When
    # non-empty, the same step is re-rendered instead of advancing.
    wizard_errors: dict[str, str] = {}

    model = await session.get(DealModel, model_id)
    if model is None:
        return HTMLResponse("Not found", status_code=404)
    default_project = (await session.execute(
        select(Project).where(Project.scenario_id == model_id).order_by(Project.created_at).limit(1)
    )).scalar_one_or_none()
    if default_project is None:
        return HTMLResponse("No project found", status_code=400)

    inputs = (await session.execute(
        select(OperationalInputs).where(OperationalInputs.project_id == default_project.id)
    )).scalar_one_or_none()
    if inputs is None:
        # Legacy fallback: post-refactor every Scenario is created with an
        # OperationalInputs row via app.services.scenario_factory, so this
        # branch only fires for pre-factory deals. Create an empty row and
        # apply current org/user defaults — same logic the factory uses on
        # fresh deals — so the engine never sees a NULL Type 1 field.
        from app.services.scenario_factory import apply_defaults_to_existing
        inputs = OperationalInputs(project_id=default_project.id)
        session.add(inputs)
        await session.flush()
        _wizard_user = await _get_user(session, request)
        if _wizard_user is not None:
            await apply_defaults_to_existing(
                session=session,
                scenario=model,
                inputs=inputs,
                user_id=_wizard_user.id,
                org_id=_wizard_user.org_id,
            )
        # Discount Rate / Hurdle lives on Scenario and tracks the org/user
        # IRR Hurdle Tier 1 (required return for NPV + waterfall). The factory
        # doesn't write it because it isn't in DEFAULT_REGISTRY (no scalar
        # equivalent), so the legacy fallback handles it explicitly.
        if _wizard_user is not None and model.discount_rate_pct is None:
            from decimal import Decimal as _D
            from app.settings.resolver import resolve_default as _resolve_one
            _irr1 = await _resolve_one(
                "irr_hurdle_pct_tier1", _wizard_user.id, _wizard_user.org_id, session,
            )
            if _irr1 is not None:
                try:
                    model.discount_rate_pct = _D(_irr1)
                    session.add(model)
                except Exception:
                    pass

    # Save current step's data
    if step == 1:
        # Step 1 (new): income mode + permanent-debt sizing mode + optional
        # pro forma upload. Three knobs on one screen; sizing-mode was moved
        # off Step 5 in the May 2026 wizard refactor.
        income_mode = str(form.get("income_mode") or "revenue_opex")
        if income_mode not in ("revenue_opex", "noi"):
            income_mode = "revenue_opex"
        model.income_mode = income_mode

        # Permanent-debt sizing mode (gap_fill / dscr_capped / dual_constraint)
        sizing_mode = str(form.get("debt_sizing_mode") or "").strip()
        if sizing_mode in ("gap_fill", "dscr_capped", "dual_constraint"):
            inputs.debt_sizing_mode = sizing_mode

        session.add(model)
        session.add(inputs)

        # Pre-fill NOI from linked opportunity's scraped listing
        if income_mode == "noi":
            await _prefill_noi_from_listing(model, default_project, inputs, session)
    elif step == 2:
        # Debt type checkboxes → debt_types list
        _valid_types = {
            "pre_development_loan", "acquisition_loan", "construction_loan",
            "bridge", "permanent_debt", "construction_to_perm",
        }
        selected = [t for t in form.getlist("debt_types") if t in _valid_types]
        if selected:
            inputs.debt_types = selected

        # Per-debt-type Source Vehicle selection → pre-populate debt_terms so
        # Step 4 shows vehicle's rate/carry without the user having to re-enter.
        from app.models.source_vehicle import SourceVehicle as _SV_s2
        _dt2 = dict(inputs.debt_terms or {})
        for _ft2 in (selected or []):
            _vid_raw = str(form.get(f"vehicle_id_{_ft2}", "") or "").strip()
            _entry2 = dict(_dt2.get(_ft2, {}))
            if not _vid_raw:
                _entry2.pop("vehicle_id", None)
                _dt2[_ft2] = _entry2
                continue
            try:
                _vid = UUID(_vid_raw)
            except (ValueError, AttributeError):
                continue
            _sv2 = (await session.execute(
                select(_SV_s2).where(_SV_s2.id == _vid)
            )).scalar_one_or_none()
            if _sv2 is None:
                continue
            _entry2["vehicle_id"] = str(_vid)
            if _sv2.interest_rate_pct is not None:
                _entry2["rate_pct"] = float(_sv2.interest_rate_pct)
            if _sv2.carry_type:
                _entry2["loan_type"] = _sv2.carry_type
            if _sv2.amort_term_years:
                _entry2["amort_years"] = int(_sv2.amort_term_years)
            _dt2[_ft2] = _entry2
        inputs.debt_terms = _dt2 if _dt2 else inputs.debt_terms
    elif step == 3:
        # Per-debt milestone & Exit Vehicle config.  The old "Active To"
        # column is gone — end of loan is derived from Exit Vehicle in the
        # finalize step (and in the engine at compute time).
        _valid_debt_types = set(inputs.debt_types or [])
        dmc: dict = {}
        for ft in _valid_debt_types:
            active_from  = form.get(f"{ft}_active_from") or ""
            # New field name; accept the old "retired_by" for form re-submits
            # that round-tripped through the old POST shape.
            exit_vehicle = (
                form.get(f"{ft}_exit_vehicle")
                or form.get(f"{ft}_retired_by")
                or ""
            )
            # Normalise: legacy 'perpetuity' → 'maturity'.
            if exit_vehicle == "perpetuity":
                exit_vehicle = "maturity"
            # Validate vehicle: must be 'maturity' | 'sale' | another picked debt.
            if exit_vehicle and exit_vehicle not in {"maturity", "sale"} and exit_vehicle not in _valid_debt_types:
                wizard_errors[ft] = (
                    f"Exit Vehicle {exit_vehicle!r} is not one of the selected debt types."
                )
                continue
            if exit_vehicle == ft:
                wizard_errors[ft] = "A loan cannot retire itself. Pick a different Exit Vehicle."
                continue
            if active_from or exit_vehicle:
                dmc[ft] = {
                    "active_from":  active_from,
                    "exit_vehicle": exit_vehicle or "maturity",
                }
        if wizard_errors:
            # Re-render step 3 with errors; do not advance
            pass
        elif dmc:
            inputs.debt_milestone_config = dmc
    elif step == 4:
        # Per-debt terms: loan type, rate, amort years
        # Validate with explicit try/except and range checks so bad input
        # surfaces as a field error, not a 500.
        _VALID_LOAN_TYPES = {
            "io_only", "interest_reserve", "capitalized_interest",
            "pi", "io_then_pi",
        }
        dt_terms = dict(inputs.debt_terms or {})
        for ft in (inputs.debt_types or []):
            loan_type   = form.get(f"{ft}_loan_type")
            rate_raw    = form.get(f"{ft}_rate_pct")
            amort_raw   = form.get(f"{ft}_amort_years")
            hold_raw    = form.get(f"{ft}_hold_term_years")
            entry = dict(dt_terms.get(ft, {}))

            if loan_type:
                if loan_type not in _VALID_LOAN_TYPES:
                    wizard_errors[ft] = f"Unknown loan type: {loan_type!r}"
                    continue
                entry["loan_type"] = loan_type

            if rate_raw:
                try:
                    rate_val = float(rate_raw)
                except (TypeError, ValueError):
                    wizard_errors[ft] = f"Interest rate must be a number (got {rate_raw!r})"
                    continue
                if rate_val < 0 or rate_val > 30:
                    wizard_errors[ft] = (
                        f"Interest rate {rate_val}% is outside 0–30%. Enter a realistic rate."
                    )
                    continue
                entry["rate_pct"] = rate_val

            if amort_raw:
                try:
                    amort_val = int(amort_raw)
                except (TypeError, ValueError):
                    wizard_errors[ft] = f"Amortization must be a whole number of years (got {amort_raw!r})"
                    continue
                if amort_val < 1 or amort_val > 40:
                    wizard_errors[ft] = f"Amortization {amort_val} years is outside 1–40 years."
                    continue
                entry["amort_years"] = amort_val

            if hold_raw:
                try:
                    hold_val = int(hold_raw)
                except (TypeError, ValueError):
                    wizard_errors[ft] = f"Hold term must be a whole number of years (got {hold_raw!r})"
                    continue
                if hold_val < 1 or hold_val > 40:
                    wizard_errors[ft] = f"Hold term {hold_val} years is outside 1–40 years."
                    continue
                entry["hold_term_years"] = hold_val

            if entry:
                dt_terms[ft] = entry
        if not wizard_errors:
            inputs.debt_terms = dt_terms
    elif step == 5:
        # Per-debt sizing — LTV / fixed amount / minimum DSCR. The permanent-
        # debt sizing-mode picker moved to Step 1 in the May 2026 refactor;
        # this step is no longer responsible for it. Reserves & Floors (old
        # Step 6) is gone entirely — those fields come from org defaults.
        dt_terms = dict(inputs.debt_terms or {})
        for ft in (inputs.debt_types or []):
            sizing_approach = form.get(f"{ft}_sizing_approach")
            ltv_pct         = form.get(f"{ft}_ltv_pct")
            fixed_amount    = form.get(f"{ft}_fixed_amount")
            if sizing_approach or ltv_pct or fixed_amount:
                entry = dict(dt_terms.get(ft, {}))
                if sizing_approach: entry["sizing_approach"] = sizing_approach
                if ltv_pct:         entry["ltv_pct"]        = float(ltv_pct)
                if fixed_amount:    entry["fixed_amount"]   = float(fixed_amount)
                dt_terms[ft] = entry
        dscr_val = form.get("dscr_minimum")
        if dscr_val:
            try:
                _perm = dict(dt_terms.get("permanent_debt", {}))
                _perm["dscr_min"] = float(dscr_val)
                dt_terms["permanent_debt"] = _perm
            except (TypeError, ValueError):
                pass
        inputs.debt_terms = dt_terms

    _post_user = await _get_user(session, request)
    _post_svd = await _wizard_debt_vehicles(session, _post_user) if _post_user else []

    # If validation failed, don't persist and re-render the same step with errors
    if wizard_errors:
        _post_active_from_opts = await _wizard_active_from_options(session, default_project)
        _post_phases_present = await _wizard_phases_present(session, default_project)
        return templates.TemplateResponse(request, "partials/deal_setup_wizard.html", {
            "request": request, "model": model, "inputs": inputs, "step": step,
            "wizard_errors": wizard_errors, "source_vehicles_debt": _post_svd,
            "wizard_active_from_opts": _post_active_from_opts,
            "phases_present": _post_phases_present,
        })

    session.add(inputs)
    await session.commit()
    await session.refresh(inputs)
    await session.refresh(model)

    # ── Step 1 → optional pro forma upload ────────────────────────────────
    # New (May 2026): the upload zone lives directly on Step 1 instead of a
    # separate page. When a file is attached we dispatch into the preflight
    # logic that used to live behind /proforma-preflight. With no file the
    # user clicked "Skip Import →" (or NOI mode) and we advance to Step 2.
    if step == 1:
        _proforma_file = form.get("file")
        # FormData returns a starlette UploadFile when one was attached, or a
        # bare string ("") when the input was empty. Filter by truthy filename.
        from starlette.datastructures import UploadFile as _StarletteUploadFile
        if isinstance(_proforma_file, _StarletteUploadFile) and (_proforma_file.filename or ""):
            return await _dispatch_proforma_preflight(
                request=request, model_id=model_id, upload=_proforma_file,
            )

    # ── Step 2 → vehicle-skip routing ─────────────────────────────────────
    # If every selected debt type was assigned a Source Vehicle, Steps 3-5
    # (milestones, terms, sizing) are redundant — the vehicle carries that
    # data already. Jump straight to Review. If ANY debt is on "Use defaults"
    # we drop into 3/4/5 for all of them (per-debt skipping is future work).
    next_step = step + 1
    review_back_step = step  # default: Review's back goes to the prior step
    if step == 2:
        _selected = inputs.debt_types or []
        _dt_now = inputs.debt_terms or {}
        _all_have_vehicle = bool(_selected) and all(
            (_dt_now.get(_ft) or {}).get("vehicle_id") for _ft in _selected
        )
        if _all_have_vehicle:
            next_step = 6
            review_back_step = 2

    # Seed perm-debt defaults so Step 4 shows user/org preference, not template fallback.
    await _seed_wizard_perm_defaults(inputs, session, request)
    _next_active_from_opts = await _wizard_active_from_options(session, default_project)
    _next_phases_present = await _wizard_phases_present(session, default_project)
    return templates.TemplateResponse(request, "partials/deal_setup_wizard.html", {
        "request": request, "model": model, "inputs": inputs, "step": next_step,
        "source_vehicles_debt": _post_svd,
        "wizard_active_from_opts": _next_active_from_opts,
        "review_back_step": review_back_step,
        "phases_present": _next_phases_present,
    })


@router.post("/ui/models/{model_id}/setup/complete")
async def deal_setup_wizard_complete(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> Response:
    """Finalize setup: mark complete and auto-create the primary debt CapitalModule(s)."""
    from app.models.capital import CapitalModule, CapitalModuleProject

    model = await session.get(DealModel, model_id)
    if model is None:
        return HTMLResponse("Not found", status_code=404)
    default_project = (await session.execute(
        select(Project).where(Project.scenario_id == model_id).order_by(Project.created_at).limit(1)
    )).scalar_one_or_none()
    if default_project is None:
        return HTMLResponse("No project", status_code=400)

    inputs = (await session.execute(
        select(OperationalInputs).where(OperationalInputs.project_id == default_project.id)
    )).scalar_one_or_none()
    if inputs is None:
        return HTMLResponse("No inputs", status_code=400)

    form = await request.form()
    dt = inputs.debt_terms or {}
    debt_types = inputs.debt_types  # None for pre-migration deals
    debt_structure = inputs.debt_structure or "perm_only"

    # Remove any existing auto-created debt modules (clean re-run).
    # Cascade dependent rows manually first — WaterfallResult and
    # WaterfallTier both reference capital_module_id with NOT NULL +
    # default-NO-ACTION FKs, so SQLAlchemy's default "set FK to NULL on
    # parent delete" behavior throws a NotNullViolation. Delete the
    # dependents explicitly to keep the wizard re-runnable on scenarios
    # that have been computed before.
    from sqlalchemy import delete as _sa_delete
    from app.models.capital import WaterfallResult as _WR
    existing_auto = list((await session.execute(
        select(CapitalModule).where(
            CapitalModule.scenario_id == model_id,
            CapitalModule.label.like("%(auto)%"),
        )
    )).scalars())
    # Snapshot share state so wizard re-runs don't silently strip junctions
    # added by Add Project. Map vehicle_type → set of non-default project IDs
    # that had a junction on the OLD auto module. After recreation we restore
    # those junctions on the matching new module so shared auto-debt persists.
    _prior_auto_shares: dict[str, set[UUID]] = {}
    if existing_auto:
        _prior_junctions = list((await session.execute(
            select(CapitalModuleProject).where(
                CapitalModuleProject.capital_module_id.in_([cm.id for cm in existing_auto])
            )
        )).scalars())
        _ft_by_mod = {cm.id: str(getattr(cm, "vehicle_type", "") or "").replace("VehicleType.", "") for cm in existing_auto}
        for _j in _prior_junctions:
            if _j.project_id == default_project.id:
                continue
            _ft = _ft_by_mod.get(_j.capital_module_id)
            if _ft:
                _prior_auto_shares.setdefault(_ft, set()).add(_j.project_id)
    if existing_auto:
        _auto_ids = [cm.id for cm in existing_auto]
        await session.execute(
            _sa_delete(_WR).where(_WR.capital_module_id.in_(_auto_ids))
        )
        await session.execute(
            _sa_delete(WaterfallTier).where(WaterfallTier.capital_module_id.in_(_auto_ids))
        )
        await session.flush()
    for cm in existing_auto:
        await session.delete(cm)

    # Closing-cost pre-load data collected during Phase B module creation.
    # Populated in the loop below; used after both branches to write $0 Use line stubs.
    _cc_preload_modules: list[dict] = []

    # If a Source Vehicle was chosen at deal creation, use it for the auto module
    # instead of the wizard's per-debt-type config.
    _sv_module_created = False
    if model.source_vehicle_id is not None:
        from app.models.source_vehicle import SourceVehicle as _SV2
        _sv = (await session.execute(
            select(_SV2).where(_SV2.id == model.source_vehicle_id)
        )).scalar_one_or_none()
        if _sv is not None:
            _sv_module_created = True
            _sv_label = f"{_sv.label} (auto)"
            _sv_ft_str = _sv.vehicle_type or "debt"  # for _cc_preload_modules key
            session.add(CapitalModule(
                scenario_id=model_id,
                label=_sv_label,
                vehicle_type=_sv.vehicle_type,
                equity_role=_sv.equity_role,
                stack_position=1,
                source={**(_sv.source_config or {}), "auto_size": True},
                carry=_sv.carry_config or {},
                exit_terms=_sv.exit_config or {"exit_type": "full_payoff", "trigger": "end of hold period"},
                active_phase_start=_sv.active_phase_start or "acquisition",
                active_phase_end=_sv.active_phase_end or "stabilized",
                source_vehicle_id=model.source_vehicle_id,
            ))
            _cc_preload_modules.append({
                "loan_subtype": _sv_ft_str,
                "label": _sv_label,
                "active_phase_start": _sv.active_phase_start or "acquisition",
            })

    if debt_types and not _sv_module_created:
        # ── New multi-debt path ───────────────────────────────────────────────
        # Build one CapitalModule per selected debt type using debt_milestone_config
        # and debt_terms (per-debt dicts).  Falls back to sensible defaults if the
        # wizard steps were skipped (e.g. re-running on a backfilled deal).
        dmc = inputs.debt_milestone_config or {}

        # Valid debt sub-type keys (used to validate/skip unknown entries).
        _FT_MAP: set[str] = {
            "pre_development_loan", "acquisition_loan", "construction_loan",
            "bridge", "permanent_debt", "construction_to_perm",
        }
        _LABEL: dict[str, str] = {
            "pre_development_loan": "Pre-Development Loan",
            "acquisition_loan":     "Acquisition Loan",
            "construction_loan":    "Construction Loan",
            "bridge":               "Bridge Loan",
            "permanent_debt":       "Permanent Debt",
            "construction_to_perm": "Construction-to-Perm",
        }
        # Defaults must match deal_setup_wizard.html Step 4 (_dt_default_*).
        # Any mismatch means wizard re-runs show ghost field changes.
        _DEFAULT_RATE: dict[str, float] = {
            "pre_development_loan": 8.0,
            "acquisition_loan":     6.5,
            "construction_loan":    6.0,
            "bridge":               7.5,
            "permanent_debt":       5.0,
            "construction_to_perm": 5.0,
        }
        _DEFAULT_LOAN_TYPE: dict[str, str] = {
            "pre_development_loan": "interest_reserve",
            "acquisition_loan":     "interest_reserve",
            "construction_loan":    "interest_reserve",
            "bridge":               "interest_reserve",
            "permanent_debt":       "pi",
            "construction_to_perm": "io_then_pi",
        }
        _DEFAULT_FROM: dict[str, str] = {
            "pre_development_loan": "pre_construction",
            "acquisition_loan":     "acquisition",
            "construction_loan":    "acquisition",
            "bridge":               "lease_up",
            "permanent_debt":       "lease_up",
            "construction_to_perm": "acquisition",
        }
        # Exit Vehicle default per debt type — first available retirer in the
        # preference chain wins, else fall back to the trailing sentinel.
        # Mirrors the Jinja `_vehicle_pref_chain` in deal_setup_wizard.html
        # (§3 Milestones & Retirement).
        _VEHICLE_PREF_CHAIN: dict[str, list[str]] = {
            "pre_development_loan": ["acquisition_loan","construction_loan","construction_to_perm","bridge","maturity"],
            "acquisition_loan":     ["construction_loan","construction_to_perm","permanent_debt","bridge","maturity"],
            "construction_loan":    ["permanent_debt","construction_to_perm","sale","maturity"],
            "bridge":               ["permanent_debt","sale","maturity"],
            "permanent_debt":       ["maturity"],
            "construction_to_perm": ["sale","maturity"],
        }

        def _default_vehicle(ft_str: str, picked: set[str]) -> str:
            for cand in _VEHICLE_PREF_CHAIN.get(ft_str, ["maturity"]):
                if cand in {"maturity", "sale"}:
                    return cand
                if cand in picked and cand != ft_str:
                    return cand
            return "maturity"

        # Phase-rank lookup used to derive active_phase_end from the resolved
        # vehicle.  Mirrors cashflow._APS_TO_RANK; duplicated here to avoid a
        # circular import between the router and the engine.
        _APS_TO_PHASE: dict[int, str] = {
            0: "acquisition", 2: "pre_construction", 3: "construction",
            4: "lease_up", 5: "stabilized", 6: "exit",
        }
        _PHASE_TO_RANK: dict[str, int] = {
            "acquisition": 0, "close": 0,
            "pre_construction": 2,
            "construction": 3,
            "lease_up": 4, "operation_lease_up": 4,
            "stabilized": 5, "operation_stabilized": 5,
            "exit": 6, "divestment": 6,
        }

        # ── Pass 1: pre-generate UUIDs so Exit Vehicle can reference siblings
        # by id.  Collect per-debt config so Pass 2 can resolve vehicle and
        # derive active_phase_end without rescanning the form.
        _picked = {ft_str for ft_str in debt_types if ft_str in _FT_MAP}
        _module_plan: list[dict] = []
        _uuid_by_debt_type: dict[str, "_uuid_mod.UUID"] = {}
        for pos, ft_str in enumerate(debt_types, start=1):
            if ft_str not in _FT_MAP:
                continue
            cfg   = dmc.get(ft_str, {})
            terms = dt.get(ft_str, {})
            rate        = float(terms.get("rate_pct") or _DEFAULT_RATE.get(ft_str, 6.0))
            loan_type   = terms.get("loan_type") or _DEFAULT_LOAN_TYPE.get(ft_str, "io_only")
            amort_years = int(terms.get("amort_years") or 30)
            ltv_pct     = float(terms.get("ltv_pct")) if terms.get("ltv_pct") is not None else None
            active_from = cfg.get("active_from") or _DEFAULT_FROM.get(ft_str, "acquisition")
            # Stabilized-acquisition deals: the perm loan IS the acquisition
            # financing — no construction/lease-up to refinance out of. Default
            # active_from = "acquisition" so closing costs flow at close instead
            # of dumping into month 1 of stabilized (where they'd otherwise
            # show as a phantom -$25K NCF spike). Value-add / new-construction
            # keep the standard "lease_up" default — perm refinances out the
            # construction loan at stabilization, costs hit then.
            _scn_proj_type = str(getattr(model, "project_type", "") or "").replace("ProjectType.", "")
            if (
                ft_str == "permanent_debt"
                and _scn_proj_type == "acquisition"
                and not cfg.get("active_from")
            ):
                active_from = "acquisition"

            # Resolve Exit Vehicle — legacy 'retired_by' + 'active_to' are
            # honoured as fallbacks so in-flight wizard re-submits still work
            # while old rows propagate through the finalize path.
            vehicle_raw = (cfg.get("exit_vehicle") or cfg.get("retired_by") or "").strip()
            if vehicle_raw == "perpetuity" or not vehicle_raw:
                vehicle_raw = _default_vehicle(ft_str, _picked)
            _module_plan.append({
                "pos": pos,
                "ft_str": ft_str,
                "rate": rate,
                "loan_type": loan_type,
                "amort_years": amort_years,
                "active_from": active_from,
                "vehicle_raw": vehicle_raw,
            })
            _uuid_by_debt_type[ft_str] = _uuid_mod.uuid4()

        # ── Pass 2: create CapitalModule rows with fully-resolved exit_terms
        # and derived active_phase_end.
        for plan in _module_plan:
            ft_str      = plan["ft_str"]
            rate        = plan["rate"]
            loan_type   = plan["loan_type"]
            amort_years = plan["amort_years"]
            active_from = plan["active_from"]
            vehicle_raw = plan["vehicle_raw"]
            pos         = plan["pos"]
            cm_id       = _uuid_by_debt_type[ft_str]

            if loan_type == "interest_reserve":
                carry: dict = {"carry_type": "interest_reserve", "io_rate_pct": rate}
            elif loan_type == "capitalized_interest":
                carry = {"carry_type": "capitalized_interest", "io_rate_pct": rate}
            elif loan_type == "io_only":
                carry = {"carry_type": "io_only", "io_rate_pct": rate}
            elif loan_type == "pi":
                carry = {"carry_type": "pi", "amort_term_years": amort_years, "io_rate_pct": rate}
            else:  # io_then_pi
                carry = {
                    "phases": [
                        {"name": "construction", "carry_type": "interest_reserve", "io_rate_pct": rate},
                        {"name": "operation", "carry_type": "pi", "amort_term_years": amort_years, "io_rate_pct": rate},
                    ]
                }

            # Resolve vehicle to the persisted form:
            #   "maturity" / "sale" → literal
            #   <picked debt-type>   → retirer module's UUID
            if vehicle_raw in {"maturity", "sale"}:
                vehicle_value = vehicle_raw
                derived_end = "exit" if vehicle_raw == "sale" else "exit"  # maturity = perpetuity, through exit
                exit_trigger = "end of hold period" if vehicle_raw == "maturity" else "Sale"
            elif vehicle_raw in _uuid_by_debt_type:
                retirer_uuid = _uuid_by_debt_type[vehicle_raw]
                vehicle_value = str(retirer_uuid)
                # Derived end = retirer's active_from (handoff point)
                retirer_plan = next(p for p in _module_plan if p["ft_str"] == vehicle_raw)
                derived_end = retirer_plan["active_from"]
                exit_trigger = _LABEL.get(vehicle_raw, vehicle_raw.replace("_", " "))
            else:
                # Shouldn't happen after validation, but stay defensive.
                vehicle_value = "maturity"
                derived_end = "exit"
                exit_trigger = "end of hold period"

            exit_terms_dict: dict = {
                "exit_type": "full_payoff",
                "trigger":   exit_trigger,
                "vehicle":   vehicle_value,
            }

            _cm_label_for_cc = f"{_LABEL.get(ft_str, ft_str)} (auto)"
            _source_dict: dict = {"auto_size": True, "interest_rate_pct": rate}
            if ltv_pct is not None:
                _source_dict["ltv_pct"] = ltv_pct
            # Perm-debt requires hold_term_years (validator-enforced) and
            # optionally dscr_min. Read from wizard staging (debt_terms[ft]),
            # default hold_term_years to amort_years if user didn't pick one.
            if ft_str == "permanent_debt":
                _terms_for_pd = dt.get(ft_str, {}) if isinstance(dt, dict) else {}
                _hold_raw = _terms_for_pd.get("hold_term_years") if isinstance(_terms_for_pd, dict) else None
                _source_dict["hold_term_years"] = (
                    int(_hold_raw) if _hold_raw else amort_years
                )
                _dscr_raw = _terms_for_pd.get("dscr_min") if isinstance(_terms_for_pd, dict) else None
                if _dscr_raw is not None:
                    try:
                        _source_dict["dscr_min"] = float(_dscr_raw)
                    except (TypeError, ValueError):
                        pass
            _cm_vid_raw = dt.get(ft_str, {}).get("vehicle_id") if isinstance(dt, dict) else None
            _cm_vid = UUID(_cm_vid_raw) if _cm_vid_raw else None

            # If a Source Vehicle was picked for this debt type, inherit
            # rate/term/sizing/carry/exit from the vehicle's JSONB presets.
            # Vehicle wins on overlap; wizard form values fill in gaps.
            _cm_vehicle_type = "debt"
            if _cm_vid is not None:
                from app.models.source_vehicle import SourceVehicle as _SVDebt
                _sv_pick = (await session.execute(
                    select(_SVDebt).where(_SVDebt.id == _cm_vid)
                )).scalar_one_or_none()
                if _sv_pick is not None:
                    _cm_label_for_cc = f"{_sv_pick.label} (auto)"
                    _cm_vehicle_type = _sv_pick.vehicle_type or "debt"
                    _source_dict = {
                        **(_sv_pick.source_config or {}),
                        **{k: v for k, v in _source_dict.items() if k == "auto_size"},
                    }
                    if _sv_pick.carry_config:
                        carry = dict(_sv_pick.carry_config)
                    if _sv_pick.exit_config:
                        exit_terms_dict = dict(_sv_pick.exit_config)
                    if _sv_pick.active_phase_start:
                        active_from = _sv_pick.active_phase_start
                    if _sv_pick.active_phase_end:
                        derived_end = _sv_pick.active_phase_end

            session.add(CapitalModule(
                id=cm_id,
                scenario_id=model_id,
                label=_cm_label_for_cc,
                vehicle_type=_cm_vehicle_type,
                stack_position=pos,
                source=_source_dict,
                carry=carry,
                exit_terms=exit_terms_dict,
                active_phase_start=active_from,
                active_phase_end=derived_end,
                source_vehicle_id=_cm_vid,
            ))
            _cc_preload_modules.append({
                "loan_subtype": ft_str,
                "label": _cm_label_for_cc,
                "active_phase_start": active_from,
            })

    elif not _sv_module_created:
        # ── Legacy 3-path (backward compat for pre-migration deals) ──────────
        if debt_structure == "construction_to_perm":
            construction_rate = dt.get("construction_rate_pct") or dt.get("perm_rate_pct") or 4.5
            perm_rate = dt.get("perm_rate_pct") or construction_rate
            amort_years = int(dt.get("perm_amort_years") or 30)
            session.add(CapitalModule(
                scenario_id=model_id,
                label="Bond / Construction-to-Perm (auto)",
                vehicle_type="debt",
                stack_position=1,
                source={"auto_size": True, "interest_rate_pct": perm_rate},
                carry={
                    "phases": [
                        {"name": "construction", "carry_type": "io_only", "io_rate_pct": construction_rate},
                        {"name": "operation", "carry_type": "pi", "amort_term_years": amort_years, "io_rate_pct": perm_rate},
                    ]
                },
                exit_terms={"exit_type": "full_payoff", "trigger": "end of hold period"},
                active_phase_start="pre_construction",
                active_phase_end="stabilized",
            ))

        elif debt_structure == "construction_and_perm":
            construction_rate = dt.get("construction_rate_pct") or 6.0
            perm_rate = dt.get("perm_rate_pct") or 5.0
            amort_years = int(dt.get("perm_amort_years") or 30)
            session.add(CapitalModule(
                scenario_id=model_id,
                label="Construction Loan (auto)",
                vehicle_type="debt",
                stack_position=1,
                source={"auto_size": True, "interest_rate_pct": construction_rate},
                carry={"carry_type": "io_only", "io_rate_pct": construction_rate},
                exit_terms={"exit_type": "full_payoff", "trigger": "permanent_financing_close"},
                active_phase_start="pre_construction",
                active_phase_end="lease_up",
            ))
            session.add(CapitalModule(
                scenario_id=model_id,
                label="Permanent Debt (auto)",
                vehicle_type="debt",
                stack_position=2,
                source={"auto_size": True, "interest_rate_pct": perm_rate},
                carry={"carry_type": "pi", "amort_term_years": amort_years, "io_rate_pct": perm_rate},
                exit_terms={"exit_type": "full_payoff", "trigger": "end of hold period"},
                active_phase_start="lease_up",
                active_phase_end="stabilized",
            ))

        else:  # perm_only
            perm_rate = dt.get("perm_rate_pct") or 5.0
            amort_years = int(dt.get("perm_amort_years") or 30)
            session.add(CapitalModule(
                scenario_id=model_id,
                label="Permanent Debt (auto)",
                vehicle_type="debt",
                stack_position=1,
                source={"auto_size": True, "interest_rate_pct": perm_rate},
                carry={"carry_type": "pi", "amort_term_years": amort_years, "io_rate_pct": perm_rate},
                exit_terms={"exit_type": "full_payoff", "trigger": "end of hold period"},
                active_phase_start="acquisition",
                active_phase_end="stabilized",
            ))

    # Sync debt_structure from debt_types for engine backward compat.
    # Phase B will generalise the engine to use debt_types directly; until then
    # the sizing function gates on debt_structure to detect construction+perm bridges.
    if debt_types:
        if "construction_to_perm" in debt_types:
            inputs.debt_structure = "construction_to_perm"
        elif "construction_loan" in debt_types and "permanent_debt" in debt_types:
            inputs.debt_structure = "construction_and_perm"
        elif debt_types == ["permanent_debt"]:
            inputs.debt_structure = "perm_only"
        # Other combinations (pre_development, acquisition, bridge) left as-is until Phase B

    # All projects in the scenario — used below to seed Uses, Reserves,
    # OpEx, and CapitalModuleProject junction rows on every project so
    # multi-project deals don't leave Project 2+ empty after Deal Setup.
    all_scenario_projects = list((await session.execute(
        select(Project).where(Project.scenario_id == model_id).order_by(Project.created_at.asc())
    )).scalars())



    inputs.deal_setup_complete = True
    session.add(inputs)

    # Deal Setup is scenario-level (income mode, debt stack). Propagate
    # the completion flag to every other Project's OperationalInputs so
    # their tabs don't keep gating users back to the wizard. Create rows
    # for projects that lack OperationalInputs (e.g. freshly added
    # Projects 2+ that haven't hit the wizard individually).
    other_projects = list((await session.execute(
        select(Project).where(
            Project.scenario_id == model_id,
            Project.id != default_project.id,
        )
    )).scalars())
    for other_proj in other_projects:
        other_inputs = (await session.execute(
            select(OperationalInputs).where(OperationalInputs.project_id == other_proj.id)
        )).scalar_one_or_none()
        if other_inputs is None:
            other_inputs = OperationalInputs(project_id=other_proj.id)
            session.add(other_inputs)
        other_inputs.deal_setup_complete = True
        # Propagate scenario-level debt config so the engine's closing-cost
        # auto-sizing runs for every project. Without these, the per-project
        # compute skips the CC write-back block (it checks debt_types_list)
        # and leaves Appraisal / Origination / Legal / Title at $0.
        other_inputs.debt_types = inputs.debt_types
        other_inputs.debt_structure = inputs.debt_structure
        other_inputs.debt_milestone_config = inputs.debt_milestone_config
        # Sizing-policy fields are scenario-level too. Without them every
        # non-default project silently falls back to gap-fill on compute,
        # so DSCR caps don't bind and debt over-leverages without a gap.
        other_inputs.debt_sizing_mode = inputs.debt_sizing_mode
        other_inputs.construction_floor_pct = inputs.construction_floor_pct
        other_inputs.operation_reserve_months = inputs.operation_reserve_months

    # Create $0 Operating Reserve placeholder in Uses for every project
    # (populated at compute time). Multi-project deals each need their own.
    for _proj in all_scenario_projects:
        _existing_reserve = (await session.execute(
            select(UseLine).where(
                UseLine.project_id == _proj.id,
                UseLine.label == "Operating Reserve",
            )
        )).scalar_one_or_none()
        if _existing_reserve is None:
            session.add(UseLine(
                project_id=_proj.id,
                label="Operating Reserve",
                phase="operation",
                amount=Decimal("0"),
                timing_type="first_day",
                notes="Sized at compute time: max(OpEx, Debt Service) × reserve months",
            ))

    # Attach every newly-created CapitalModule to the DEFAULT project only.
    # Other projects start without coverage on these auto modules; when the
    # user runs Add Project (drawer), the share-or-clone checkboxes decide
    # whether each Source is shared (junction extended) or cloned (new module
    # owned solely by the new project). Default is per-project, not shared.
    #
    # Explicit flush before the SELECT — autoflush has historically dropped
    # the just-added rows in this handler (likely an interaction with the
    # earlier cascade-delete batch in this same transaction). Without the
    # flush, the SELECT comes back empty and no junctions get attached, so
    # subsequent computes return empty module lists for default_project and
    # the deal silently has no debt coverage.
    await session.flush()
    _auto_modules = list((await session.execute(
        select(CapitalModule).where(
            CapitalModule.scenario_id == model_id,
            CapitalModule.label.like("%(auto)%"),
        )
    )).scalars())
    for _mod in _auto_modules:
        _existing_j = (await session.execute(
            select(CapitalModuleProject).where(
                CapitalModuleProject.capital_module_id == _mod.id,
                CapitalModuleProject.project_id == default_project.id,
            )
        )).scalar_one_or_none()
        if _existing_j is None:
            session.add(CapitalModuleProject(
                capital_module_id=_mod.id,
                project_id=default_project.id,
                amount=Decimal(str((_mod.source or {}).get("amount") or 0)),
                active_from=_mod.active_phase_start,
                active_to=_mod.active_phase_end,
                auto_size=bool((_mod.source or {}).get("auto_size")),
            ))
        # Restore prior shared-with-other-projects junctions for this funder
        # type so wizard re-runs preserve coverage added via Add Project.
        _mod_vt = str(getattr(_mod, "vehicle_type", "") or "").replace("VehicleType.", "")
        for _pid in _prior_auto_shares.get(_mod_vt, set()):
            session.add(CapitalModuleProject(
                capital_module_id=_mod.id,
                project_id=_pid,
                amount=Decimal("0"),
                active_from=_mod.active_phase_start,
                active_to=_mod.active_phase_end,
                auto_size=bool((_mod.source or {}).get("auto_size")),
            ))

    # ── Pre-load $0 closing cost Use line stubs for Phase B modules ──────────
    # Cost names match _DEFAULT_LOAN_COSTS in cashflow.py (keep in sync).
    # amount=0 → engine computes at run time; amount>0 → user override, engine skips.
    # Users see and edit these in the S&U table before running Compute.
    _CC_PRELOAD_COSTS: dict[str, list[str]] = {
        "construction_loan":    ["Origination Fee", "Lender Legal", "Title / Survey", "Environmental Phase I"],
        "permanent_debt":       ["Origination Fee", "Lender Legal", "Appraisal", "Title"],
        "pre_development_loan": ["Origination Fee", "Lender Legal"],
        "acquisition_loan":     ["Origination Fee", "Lender Legal", "Title / Survey"],
        "bridge":               ["Origination Fee", "Lender Legal"],
        "bond":                 ["Bond Issuance Fee", "Bond Counsel Legal"],
    }
    _APS_TO_PHASE: dict[str, str] = {
        "acquisition": "acquisition",      "close": "acquisition",
        "pre_construction": "pre_construction",
        "construction": "construction",
        "lease_up": "operation",           "operation_lease_up": "operation",
        "stabilized": "operation",         "operation_stabilized": "operation",
        "exit": "exit",                    "divestment": "exit",
    }
    for _cc_mod in _cc_preload_modules:
        _cc_ft_str = _cc_mod["loan_subtype"]
        # Map construction_to_perm → bond for cost lookup
        _cc_ft_key = "bond" if _cc_ft_str == "construction_to_perm" else _cc_ft_str
        _cost_names = _CC_PRELOAD_COSTS.get(_cc_ft_key)
        if not _cost_names:
            continue
        _cc_lbl  = _cc_mod["label"]
        _cc_phase = _APS_TO_PHASE.get(_cc_mod["active_phase_start"] or "", "pre_construction")
        for _proj in all_scenario_projects:
            for _cost_name in _cost_names:
                _full_cc_lbl = f"{_cc_lbl} — {_cost_name}"
                _existing_cc = (await session.execute(
                    select(UseLine).where(
                        UseLine.project_id == _proj.id,
                        UseLine.label == _full_cc_lbl,
                    )
                )).scalar_one_or_none()
                if _existing_cc is None:
                    session.add(UseLine(
                        project_id=_proj.id,
                        label=_full_cc_lbl,
                        phase=_cc_phase,
                        amount=Decimal("0"),
                        timing_type="first_day",
                        notes="Auto-computed — edit to override",
                    ))

    # ── UnitMix: seed from opportunity if none exist ─────────────────────────
    # unit_mix is JSONB on Project; shape: [{label, unit_count, beds, baths, sqft, rent_monthly, notes}]
    existing_unit_mix: list = default_project.unit_mix or []

    # Physical attributes from the linked opportunity
    _opp_for_units: Opportunity | None = None
    if default_project.opportunity_id:
        _opp_for_units = await session.get(Opportunity, default_project.opportunity_id)

    building_unit_count: int = int(inputs.unit_count_new or 0)
    if _opp_for_units and _opp_for_units.units and not building_unit_count:
        building_unit_count = int(_opp_for_units.units)

    if not existing_unit_mix and building_unit_count:
        # Keep OperationalInputs in sync
        if not inputs.unit_count_new:
            inputs.unit_count_new = building_unit_count
            session.add(inputs)
        default_project.unit_mix = [{
            "label": "All Units",
            "unit_count": building_unit_count,
            "notes": "Seeded from opportunity — break into unit types as needed",
        }]
        session.add(default_project)
        existing_unit_mix = default_project.unit_mix

    # ── Market recommendation: KNN query for revenue/expense prefill ────────
    from app.engines.market import SubjectProperty, get_market_recommendation

    _market_rec = None
    _market_occupancy = Decimal("95")
    if _opp_for_units:
        _subj_units = building_unit_count or int(_opp_for_units.units or 0)
        _subj_year = _opp_for_units.year_built
        _subj_sqft = float(_opp_for_units.gba_sqft) if _opp_for_units.gba_sqft else None
        _subj_sqft_per_unit = _subj_sqft / _subj_units if _subj_sqft and _subj_units > 0 else None
        _subj_juris = _opp_for_units.jurisdiction or _opp_for_units.city
        _exclude_listing_id = str(_opp_for_units.id)
        if _subj_units > 0 and _subj_year:
            try:
                _market_rec = await get_market_recommendation(
                    session,
                    SubjectProperty(
                        units=_subj_units,
                        year_built=_subj_year,
                        sqft_per_unit=_subj_sqft_per_unit,
                        jurisdiction=_subj_juris,
                    ),
                    exclude_listing_id=_exclude_listing_id,
                )
                if _market_rec and not _market_rec.low_confidence:
                    if _market_rec.occupancy_pct is not None:
                        _market_occupancy = Decimal(str(round(_market_rec.occupancy_pct * 100, 1)))
            except Exception:
                pass  # market recommendation failed; fall back to defaults

    # If NOI mode and no NOI prefilled from listing, use market recommendation.
    # Mark as auto-seeded so the builder can show a confirm-or-override banner
    # — a silent KNN-based number shouldn't be accepted as the user's input.
    if model.income_mode == "noi" and inputs.noi_stabilized_input is None and _market_rec and not _market_rec.low_confidence:
        _market_noi = Decimal(str(round(_market_rec.noi_per_unit * building_unit_count, 2)))
        inputs.noi_stabilized_input = _market_noi
        inputs.noi_auto_seeded = True
        session.add(inputs)

    # ── Revenue: seed one IncomeStream per UnitMix row ──────────────────────
    # Skip entirely in NOI mode — Revenue module is not used
    # Only seed if no income streams exist yet
    existing_income = (await session.execute(
        select(IncomeStream).where(IncomeStream.project_id == default_project.id).limit(1)
    )).scalar_one_or_none()

    if existing_income is None and model.income_mode != "noi":
        # Reload unit_mix in case it was just seeded above
        await session.flush()
        await session.refresh(default_project)
        unit_mix_rows: list[dict] = default_project.unit_mix or []

        total_units = sum(int(row.get("unit_count", 0)) for row in unit_mix_rows)
        for row in unit_mix_rows:
            _uc = int(row.get("unit_count", 0))
            _rent = Decimal(str(
                row.get("in_place_rent_per_unit") or row.get("market_rent_per_unit")
                or row.get("rent_monthly") or 0
            ))
            session.add(IncomeStream(
                project_id=default_project.id,
                stream_type=IncomeStreamType.residential_rent,
                label=row.get("label", "Unit Mix"),
                unit_count=_uc,
                amount_per_unit_monthly=_rent,
                stabilized_occupancy_pct=_market_occupancy,
                escalation_rate_pct_annual=Decimal("3"),
                active_in_phases=["lease_up", "stabilized"],
            ))

        if total_units > 0:
            session.add(IncomeStream(
                project_id=default_project.id,
                stream_type=IncomeStreamType.deposit_forfeit,
                label="Turnover on Deposit",
                unit_count=total_units,
                amount_per_unit_monthly=Decimal("0"),
                stabilized_occupancy_pct=Decimal("100"),
                escalation_rate_pct_annual=Decimal("3"),
                active_in_phases=["stabilized"],
                notes="$/unit/mo to configure — typically 4.5% annual turnover × avg rent × recovery rate / 12",
            ))

    # ── OpEx: seed canonical standard lines for every project ───────────────
    # Skip individual labels that already exist (idempotent re-run).
    # Seeded in all income modes — user may switch from NOI to revenue/opex.
    # (label, per_type, scale_with_lease_up, lease_up_floor_pct, active_phases)
    _OPEX_SEEDS = [
        ("Real Estate Taxes",          "flat",     False, None,   ["lease_up", "stabilized"]),
        ("Insurance",                  "flat",     False, None,   ["lease_up", "stabilized"]),
        ("Property Management",        "per_unit", True,  25.0,   ["lease_up", "stabilized"]),
        ("Utilities — Water/Sewer",    "per_unit", True,  50.0,   ["lease_up", "stabilized"]),
        ("Utilities — Electric",       "flat",     True,  100.0,  ["lease_up", "stabilized"]),
        ("Utilities — Gas",            "per_unit", True,  50.0,   ["lease_up", "stabilized"]),
        ("Utilities — Trash",          "flat",     True,  75.0,   ["lease_up", "stabilized"]),
        ("Repairs & Maintenance",      "per_unit", True,  25.0,   ["stabilized"]),
        ("Marketing & Leasing",        "per_unit", True,  100.0,  ["lease_up", "stabilized"]),
        ("Administrative",             "flat",     False, None,   ["lease_up", "stabilized"]),
        ("Payroll",                    "flat",     False, None,   ["lease_up", "stabilized"]),
        ("Landscaping & Snow Removal", "flat",     False, None,   ["lease_up", "stabilized"]),
        ("Pest Control",               "flat",     False, None,   ["lease_up", "stabilized"]),
        ("Cleaning & Janitorial",      "flat",     False, None,   ["lease_up", "stabilized"]),
        ("Security",                   "flat",     False, None,   ["lease_up", "stabilized"]),
        ("Resident Services",          "flat",     True,  25.0,   ["stabilized"]),
        ("Jurisdiction Fees",           "flat",     False, None,   ["lease_up", "stabilized"]),
        ("Legal",                       "flat",     False, None,   ["lease_up", "stabilized"]),
        ("Bank/Software Fees",         "flat",     False, None,   ["lease_up", "stabilized"]),
        ("Unit Turnover",              "per_unit", False, None,   ["stabilized"]),
    ]

    for _opex_proj in all_scenario_projects:
        _existing_opex_labels = set((await session.execute(
            select(OperatingExpenseLine.label).where(
                OperatingExpenseLine.project_id == _opex_proj.id
            )
        )).scalars())
        for label, per_type, scale, floor_pct, phases in _OPEX_SEEDS:
            if label in _existing_opex_labels:
                continue
            session.add(OperatingExpenseLine(
                project_id=_opex_proj.id,
                label=label,
                annual_amount=Decimal("0"),
                per_type=per_type,
                scale_with_lease_up=scale,
                lease_up_floor_pct=Decimal(str(floor_pct)) if floor_pct is not None else None,
                escalation_rate_pct_annual=Decimal("3"),
                active_in_phases=phases,
            ))

    # Save Deal Health thresholds from wizard step 7 form.
    _ht: dict[str, float] = {}
    for _key, _default in (
        ("ht_occ_green", None), ("ht_oer_green", None),
        ("ht_dscr_green", None), ("ht_margin_green", None),
    ):
        _raw = form.get(_key)
        if _raw is not None:
            try:
                _ht[_key.removeprefix("ht_")] = float(_raw)
            except (ValueError, TypeError):
                pass
    if _ht:
        model.health_thresholds = {**(model.health_thresholds or {}), **_ht}

    # Resolve CapitalModule.active_from_milestone_id / active_to_milestone_id
    # FKs from the active_phase_start / active_phase_end strings. Milestones
    # are seeded earlier in this handler (per project), so the FK lookup is
    # safe here. Engine prefers FK over string when both exist, so future
    # milestone renames / date drags carry through to debt activation timing
    # without touching the legacy string field.
    from app.services.capital_module_milestones import (
        sync_milestone_fks_for_scenario,
    )
    await sync_milestone_fks_for_scenario(session, model_id)

    await session.commit()

    # Redirect to builder — NOI mode lands on the NOI module first, else
    # Property (the natural starting point for filling in unit mix /
    # building data before moving on to S&U).
    # Preserve the active Project from HX-Current-URL so multi-project deals
    # don't snap the user back to Project 1 after finishing Deal Setup.
    _first_module = "noi" if model.income_mode == "noi" else "property"
    _active_proj_q = ""
    _hx_url = request.headers.get("HX-Current-URL", "")
    if _hx_url:
        from urllib.parse import urlparse, parse_qs
        _qs = parse_qs(urlparse(_hx_url).query)
        _p = _qs.get("project", [""])[0]
        if _p:
            _active_proj_q = f"&project={_p}"
    from starlette.responses import Response as StarletteResponse
    response = StarletteResponse(status_code=204)
    response.headers["HX-Redirect"] = f"/models/{model_id}/builder?module={_first_module}{_active_proj_q}"
    return response


@router.get("/models/{model_id}/builder", response_class=HTMLResponse)
async def model_builder(
    request: Request,
    model_id: UUID,
    session: DBSession,
    module: str = Query(default=""),
    project: str = Query(default=""),  # optional Project.id to view a specific project
    view: str = Query(default=""),  # "underwriting" for the scenario-level rollup; default = per-project
    new: str = Query(default=""),  # set to "1" when redirected from new deal creation
    wizard: str = Query(default=""),  # "1" = single-flow deal-creation wizard mode (hide chrome)
) -> HTMLResponse:
    model = await session.get(DealModel, model_id)
    if model is None:
        return HTMLResponse("<p class='text-muted'>Model not found.</p>", status_code=404)

    # `project` context var = the Opportunity (purchase target), for display in topbar
    # Find Opportunity via the first Project linked to this Scenario
    _first_proj = (await session.execute(
        select(Project).where(Project.scenario_id == model_id).limit(1)
    )).scalar_one_or_none()
    opportunity = (
        await session.get(Opportunity, _first_proj.opportunity_id)
        if _first_proj and _first_proj.opportunity_id else None
    )
    user = await _get_user(session, request)
    dedup_count, conflicts_count = await _get_counts(session)
    address_issues_count = await _get_address_issues_count(session)

    # All Projects in this Scenario (tab row)
    deal_projects = list((await session.execute(
        select(Project).where(Project.scenario_id == model_id).order_by(Project.created_at.asc())
    )).scalars())

    # Resolve which project to view
    active_project_id: UUID | None = None
    if project:
        try:
            candidate_id = UUID(project)
            if any(p.id == candidate_id for p in deal_projects):
                active_project_id = candidate_id
        except ValueError:
            pass
    if active_project_id is None and deal_projects:
        active_project_id = deal_projects[0].id

    # Phase 3a: ``view=underwriting`` routes to the scenario-level rollup.
    # active_project_id stays populated (for tab-chip rendering); the
    # template branches on active_view to show rollup panels instead of
    # the per-project sidebar / module editor.
    active_view = "underwriting" if view == "underwriting" else "project"

    data = await _load_builder_data(session, model_id, project_id=active_project_id)

    # Load rollup data only when the Underwriting tab is active — cheap DB
    # aggregations; safe to run scenario-wide.
    underwriting_rollup_data: dict = {}
    if active_view == "underwriting":
        from app.engines.underwriting_rollup import (
            rollup_cashflow,
            rollup_draws,
            rollup_irr,
            rollup_sources,
            rollup_summary,
            rollup_waterfall,
        )
        underwriting_rollup_data = {
            "cashflow": await rollup_cashflow(model_id, session),
            "sources": await rollup_sources(model_id, session),
            "waterfall": await rollup_waterfall(model_id, session),
            "draws": await rollup_draws(model_id, session),
            "combined_irr_pct": await rollup_irr(model_id, session),
            "summary": await rollup_summary(model_id, session),
        }
        # Phase 2d1: Timeline Anchors panel data.
        # Load current ProjectAnchor rows + every project's milestones so the
        # form can render a per-parent <optgroup> milestone dropdown.
        from app.models.milestone import Milestone as _MS
        from app.models.project import ProjectAnchor as _PA
        _anchors = list(
            (
                await session.execute(
                    select(_PA)
                    .join(Project, Project.id == _PA.project_id)
                    .where(Project.scenario_id == model_id)
                )
            ).scalars()
        )
        _ms_rows = list(
            (
                await session.execute(
                    select(_MS)
                    .join(Project, Project.id == _MS.project_id)
                    .where(Project.scenario_id == model_id)
                    .order_by(_MS.sequence_order.asc())
                )
            ).scalars()
        )
        _ms_by_project: dict = {}
        for _m in _ms_rows:
            _ms_by_project.setdefault(_m.project_id, []).append(_m)
        underwriting_rollup_data["anchors"] = {
            "by_project": {a.project_id: a for a in _anchors},
            "milestones_by_project": _ms_by_project,
        }

    # Sibling scenarios for the variant tab row: other Scenarios sharing the same Opportunity (via Projects)
    deal_variants: list = []
    if opportunity:
        _dv_result = await session.execute(
            select(DealModel)
            .join(Project, Project.scenario_id == DealModel.id)
            .where(Project.opportunity_id == opportunity.id)
            .order_by(DealModel.created_at)
        )
        deal_variants = list(_dv_result.scalars().unique())
    if not deal_variants:
        deal_variants = [model]

    # Resolve Deal.id for the breadcrumb — model.deal_id IS the parent Deal
    parent_deal_id: UUID | None = model.deal_id

    # Determine active module — deal_setup gates modules after timeline approval
    _inputs = data.get("inputs")
    _deal_setup_complete = bool(getattr(_inputs, "deal_setup_complete", False)) if _inputs else False
    _timeline_approved = data.get("timeline_approved", False)
    if not _timeline_approved:
        active_module = module or "timeline"
    elif not _deal_setup_complete and module not in ("timeline", "deal_setup", ""):
        # Redirect so the URL reflects where the user actually lands. Preserve
        # the active project so multi-project deals don't snap back to P1.
        _proj_q = f"&project={active_project_id}" if active_project_id else ""
        return RedirectResponse(url=f"/models/{model_id}/builder?module=deal_setup{_proj_q}", status_code=302)
    else:
        active_module = module or ("sources_uses" if _deal_setup_complete else "deal_setup")

    # Canonicalize URL so it always carries ?project= for per-project views.
    # Without this, HX-Current-URL on form posts/deletes lacks the project
    # param and `_active_project_from_request` returns None, silently no-op'ing
    # JSONB writes (e.g. unit-mix delete). Skip for scenario-level rollup view.
    if active_view != "underwriting" and not project and active_project_id is not None:
        _view_q = f"&view={view}" if view else ""
        _new_q = f"&new={new}" if new else ""
        # Preserve wizard=1 across the canonicalization redirect; otherwise the
        # single-flow deal-creation chrome drops on the very first hop after
        # /ui/deals/create.
        _wiz_q = f"&wizard={wizard}" if wizard else ""
        return RedirectResponse(
            url=f"/models/{model_id}/builder?module={active_module}&project={active_project_id}{_view_q}{_new_q}{_wiz_q}",
            status_code=302,
        )

    # Cash flow periods — only loaded when the cashflow module is active.
    # Multi-project: filter by active project_id so the per-project tab
    # doesn't interleave both projects' rows (which previously showed
    # Project 2's $5M acquisition on Project 1's cashflow page).
    cash_flow_rows: list = []
    if active_module == "cashflow":
        from app.models.cashflow import CashFlow
        _cf_q = select(CashFlow).where(CashFlow.scenario_id == model_id)
        if active_project_id is not None:
            _cf_q = _cf_q.where(CashFlow.project_id == active_project_id)
        cash_flow_rows = list((await session.execute(
            _cf_q.order_by(CashFlow.period)
        )).scalars())

    # Multi-parcel detection — opportunity IS the listing; APN is on opportunity directly
    import re as _re
    multi_parcel_apns: list[str] = []
    if opportunity and opportunity.apn and len(deal_projects) <= 1:
        if _re.search(r"[,;]", opportunity.apn):
            multi_parcel_apns = [a.strip() for a in _re.split(r"[,;]", opportunity.apn) if a.strip()]

    # Lot-size mismatch detection — flag from parcel reconciliation
    lot_size_mismatch_info: dict | None = None
    if opportunity and opportunity.lot_size_mismatch and opportunity.parcel_id:
        _p = await session.get(Parcel, opportunity.parcel_id)
        if _p:
            _parcel_lot = float(_p.lot_sqft) if _p.lot_sqft else (float(_p.gis_acres) * 43560 if _p.gis_acres else None)
            _listing_lot = float(opportunity.lot_sqft) if opportunity.lot_sqft else None
            if _parcel_lot and _listing_lot:
                lot_size_mismatch_info = {
                    "listing_sqft": f"{_listing_lot:,.0f}",
                    "parcel_sqft": f"{_parcel_lot:,.0f}",
                }

    # Active project object (for Clone From drawer label)
    active_project = next((p for p in deal_projects if p.id == active_project_id), None)

    # Resolved anchor start date for the active project. When the project is
    # anchored to another project's milestone, the timeline wizard's Start
    # Date input is replaced by a read-only display of this computed date.
    active_project_anchor_date = None
    active_project_anchor_parent = None
    if active_project_id is not None:
        from app.models.project import ProjectAnchor as _PA_active
        _active_anchor = (await session.execute(
            select(_PA_active).where(_PA_active.project_id == active_project_id)
        )).scalar_one_or_none()
        if _active_anchor is not None:
            from app.engines.anchor_resolver import resolve_project_start_dates
            from app.models.deal import Scenario, Scenario as _Scn
            _scn = await session.get(_Scn, model_id)
            await session.refresh(_scn, ["projects"])
            _resolved = await resolve_project_start_dates(_scn, session)
            active_project_anchor_date = _resolved.get(active_project_id)
            active_project_anchor_parent = next(
                (p for p in deal_projects if p.id == _active_anchor.anchor_project_id),
                None,
            )

    # Milestones for every project on the scenario — feeds the Add Project
    # drawer's "Anchor Start Date To" optgroup dropdown so the user can wire
    # the new project to a parent milestone at creation time.
    anchor_milestones_by_project: dict = {}
    if deal_projects:
        from app.models.milestone import Milestone as _MS_anchor
        _anchor_ms_rows = list(
            (
                await session.execute(
                    select(_MS_anchor)
                    .join(Project, Project.id == _MS_anchor.project_id)
                    .where(Project.scenario_id == model_id)
                    .order_by(_MS_anchor.sequence_order.asc())
                )
            ).scalars()
        )
        for _ms in _anchor_ms_rows:
            anchor_milestones_by_project.setdefault(_ms.project_id, []).append(_ms)

    # Add-Project drawer is search-driven (see GET /ui/deals/{deal_id}/add-project/search),
    # so no eager opportunity list is needed here.

    # When deal_setup is the active module, resolve wizard step and missing building data
    # so the included partials/deal_setup_wizard.html has everything it needs.
    # Deal Setup is a scenario-level wizard (income mode, debt stack), so the
    # template always reads the default project's OperationalInputs — not the
    # active project's. Otherwise Project 2's tab would render the wizard with
    # inputs=None and Step 7 (Review) crashes on `inputs.debt_sizing_mode`.
    wizard_step: int = 1
    deal_setup_inputs = data.get("inputs")
    # Map debt_type → list of project names sharing the existing auto module
    # (excluding the currently-active project). Step 7 of the wizard renders
    # a "shared with X" chip when this list is non-empty.
    wizard_share_info: dict[str, list[str]] = {}
    if active_module == "deal_setup":
        _default_proj = (await session.execute(
            select(Project).where(Project.scenario_id == model_id).order_by(Project.created_at.asc()).limit(1)
        )).scalar_one_or_none()
        if _default_proj is not None:
            deal_setup_inputs = (await session.execute(
                select(OperationalInputs).where(OperationalInputs.project_id == _default_proj.id)
            )).scalar_one_or_none()
            wizard_step = 1
            # Find existing auto modules + their junction-shared projects so
            # the Step 7 review can flag "shared with {project}" entries.
            from app.models.capital import CapitalModuleProject as _CMP_ws
            _auto_mods = list((await session.execute(
                select(CapitalModule).where(
                    CapitalModule.scenario_id == model_id,
                    CapitalModule.label.like("%(auto)%"),
                )
            )).scalars())
            _proj_name_by_id = {p.id: p.name for p in deal_projects}
            for _mod in _auto_mods:
                _junctions = list((await session.execute(
                    select(_CMP_ws.project_id).where(_CMP_ws.capital_module_id == _mod.id)
                )).scalars())
                _other_names = [
                    _proj_name_by_id[pid] for pid in _junctions
                    if pid != active_project_id and pid in _proj_name_by_id
                ]
                if _other_names:
                    _vt_str = str(getattr(_mod, "vehicle_type", "") or "").replace("VehicleType.", "")
                    wizard_share_info[_vt_str] = _other_names

    # Draw schedule data — loaded for draw_schedule and cashflow modules
    draw_schedule_data: dict = {}
    if active_module in ("draw_schedule", "cashflow"):
        draw_schedule_data = await _load_draw_schedule_ctx(session, model_id)
        if active_module == "cashflow":
            _sched = await _run_draw_schedule(session, model_id, writeback=False)
            if _sched:
                draw_schedule_data["schedule"] = _sched

    # Pre-render the calc-status pill so it's visible on initial page load
    # without depending on an HTMX hx-trigger="load" round-trip (which was
    # silently failing for some states, leaving the topbar empty).
    # On the Underwriting view, the pill should reflect the scenario's
    # worst-status aggregate (so it doesn't just mirror Project 1's pill);
    # on per-project tabs, it remains the active project's status.
    if active_view == "underwriting":
        _scen_st = data.get("scenario_statuses") or {}
        _uw_st = (_scen_st.get("underwriting") or {})
        _calc_status = {
            "overall": _uw_st.get("overall", "na"),
            "failing_count": int(_uw_st.get("failing_count") or 0),
            # The detail modal uses sources_uses/dscr/ltv keys; for the
            # Underwriting pill we rely on the per-project chips for drilldown
            # and just surface the aggregate severity here.
            "sources_uses": {"status": "na", "label": "See per-project chips", "detail": "", "meta": {}},
            "dscr": {"status": "na", "label": "See per-project chips", "detail": "", "meta": {}},
            "ltv": {"status": "na", "label": "See per-project chips", "detail": "", "meta": {}},
        }
    else:
        _calc_status = _compute_calc_status(data)
    _has_adj = False
    if active_view != "underwriting" and active_project_id is not None:
        _has_adj = await _has_any_gap_adjustment(session, active_project_id)
    calc_status_pill_html = _render_calc_status_pill_html(
        _calc_status, model_id, has_any_adjustment=_has_adj
    )

    # Override wizard_default_types from org/user resolved timeline template.
    org_discount_rate_default: str | None = None
    if user is not None:
        from app.settings.resolver import resolve_timeline_defaults as _resolve_tl_bld
        from app.settings.resolver import resolve_default as _resolve_one_bld
        _tl_template = await _resolve_tl_bld(user.id, user.org_id, session)
        _bld_deal_type = data.get("wizard_deal_type", "")
        _tl_for_dt = _tl_template.get(_bld_deal_type, {})
        data["wizard_default_types"] = [mt for mt, cfg in _tl_for_dt.items() if cfg.get("included")]
        data["wizard_timeline_template"] = _tl_template
        # Discount Rate / Hurdle placeholder — mirror IRR Hurdle Tier 1 org/user
        # default so the empty-field hint matches what the engine will use.
        try:
            org_discount_rate_default = await _resolve_one_bld(
                "irr_hurdle_pct_tier1", user.id, user.org_id, session
            )
        except Exception:
            org_discount_rate_default = None

    ctx = {
        "model": model,
        "project": active_project or opportunity,  # template uses `project.name` for the topbar breadcrumb
        "parent_deal_id": str(parent_deal_id) if parent_deal_id else None,
        "deal_variants": deal_variants,
        "deal_projects": deal_projects,
        "anchor_milestones_by_project": anchor_milestones_by_project,
        "active_project_id": str(active_project_id) if active_project_id else None,
        "active_project": active_project,
        "active_project_anchor_date": active_project_anchor_date,
        "active_project_anchor_parent": active_project_anchor_parent,
        "active_view": active_view,
        "underwriting_rollup": underwriting_rollup_data,
        "active_module": active_module,
        "deal_setup_complete": _deal_setup_complete,
        "new_deal": new == "1",
        "wizard_mode": wizard == "1",
        "cash_flow_rows": cash_flow_rows,
        "multi_parcel_apns": multi_parcel_apns,
        "lot_size_mismatch": lot_size_mismatch_info,
        "step": wizard_step,
        "wizard_share_info": wizard_share_info,
        "calc_status_pill_html": calc_status_pill_html,
        **data,
        # Override inputs when the wizard is active so it always reads the
        # default project's OperationalInputs (Deal Setup is scenario-level).
        # Spread AFTER `**data` so this wins.
        **({"inputs": deal_setup_inputs} if active_module == "deal_setup" else {}),
        **draw_schedule_data,
        **_base_ctx(user, dedup_count, "deals", address_issues_count, conflicts_count=conflicts_count),
        "org_set_fields": frozenset({"debt_sizing_mode", "operation_reserve_months", "capex_reserve_per_unit_annual", "risk_free_rate_pct"}),
        "org_discount_rate_default": org_discount_rate_default,
    }
    return templates.TemplateResponse(request, "model_builder.html", ctx)


@router.get("/ui/panel/{model_id}", response_class=HTMLResponse)
async def builder_panel(
    request: Request,
    model_id: UUID,
    session: DBSession,
    module: str = Query(default="sources"),
) -> HTMLResponse:
    """HTMX endpoint — returns the panel partial after a mutation."""
    from app.models.cashflow import CashFlow

    model = await session.get(DealModel, model_id)
    if model is None:
        return HTMLResponse("<p class='text-muted'>Model not found.</p>", status_code=404)

    _active_proj_id = await _active_project_from_request(request, session, model_id)
    data = await _load_builder_data(session, model_id, project_id=_active_proj_id)
    ctx: dict = {
        "model": model,
        "active_module": module,
        "active_project_id": str(_active_proj_id) if _active_proj_id else None,
        **data,
    }

    # Cash flow periods — only loaded when the cashflow module is active.
    # Filter by active project so per-project tab doesn't interleave rows.
    if module == "cashflow":
        _cf_q2 = select(CashFlow).where(CashFlow.scenario_id == model_id)
        if _active_proj_id is not None:
            _cf_q2 = _cf_q2.where(CashFlow.project_id == _active_proj_id)
        cf_rows = list((await session.execute(
            _cf_q2.order_by(CashFlow.period)
        )).scalars())
        ctx["cash_flow_rows"] = cf_rows

    # Draw schedule data — loaded for draw_schedule and cashflow modules
    if module in ("draw_schedule", "cashflow"):
        ds_ctx = await _load_draw_schedule_ctx(session, model_id)
        ctx.update(ds_ctx)
        if module == "cashflow":
            _sched = await _run_draw_schedule(session, model_id, writeback=False)
            if _sched:
                ctx["schedule"] = _sched

    return templates.TemplateResponse(request, "partials/model_builder_panel.html", ctx)


# ─────────────────────────────────────────────────────────────────────────────
# Phase 2d1: Timeline Anchors — upsert / remove ProjectAnchor rows.
# ─────────────────────────────────────────────────────────────────────────────


@router.post(
    "/ui/models/{model_id}/anchors", response_class=HTMLResponse
)
async def upsert_project_anchor(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Upsert a ProjectAnchor row from the Timeline Anchors form.

    Form fields:
      project_id          — the anchored (child) project
      anchor_milestone_id — the pivot milestone on the parent project
      offset_months       — int
      offset_days         — int

    ``anchor_project_id`` is derived server-side from the picked milestone's
    ``project_id`` so the UI only needs one milestone dropdown. Returns
    HX-Redirect to the Underwriting tab so pill, staleness dots, and
    timeline re-render.
    """
    from app.models.milestone import Milestone as _MS
    from app.models.project import ProjectAnchor as _PA
    from app.engines.anchor_resolver import AnchorCycleError, _check_no_cycles
    form = await request.form()
    try:
        child_id = UUID(str(form.get("project_id", "")))
        ms_id = UUID(str(form.get("anchor_milestone_id", "")))
    except (ValueError, TypeError):
        return HTMLResponse(
            "<p class='text-muted'>Invalid project_id or anchor_milestone_id.</p>",
            status_code=400,
        )
    offset_months = int(form.get("offset_months") or 0)
    offset_days = int(form.get("offset_days") or 0)

    # Resolve parent project from the picked milestone.
    pivot = await session.get(_MS, ms_id)
    if pivot is None:
        return HTMLResponse(
            "<p class='text-muted'>Anchor milestone not found.</p>", status_code=404
        )
    parent_id = pivot.project_id
    if parent_id == child_id:
        return HTMLResponse(
            "<p class='text-muted'>A project cannot anchor to its own milestone.</p>",
            status_code=400,
        )

    # Gather every project on the scenario for the cycle check. Include the
    # proposed edge ``child ← parent`` plus all existing anchor edges.
    project_ids = [
        row[0]
        for row in (
            await session.execute(
                select(Project.id).where(Project.scenario_id == model_id)
            )
        )
    ]
    if child_id not in project_ids or parent_id not in project_ids:
        return HTMLResponse(
            "<p class='text-muted'>Projects must belong to this scenario.</p>",
            status_code=400,
        )
    existing_anchors = list(
        (
            await session.execute(
                select(_PA).where(_PA.project_id.in_(project_ids))
            )
        ).scalars()
    )
    parent_of: dict = {a.project_id: a.anchor_project_id for a in existing_anchors}
    parent_of[child_id] = parent_id  # proposed edge
    try:
        _check_no_cycles(parent_of, project_ids)
    except AnchorCycleError as exc:
        return HTMLResponse(
            f"<p class='text-muted'>Cycle rejected: {exc}</p>", status_code=400
        )

    existing = (
        await session.execute(
            select(_PA).where(_PA.project_id == child_id)
        )
    ).scalar_one_or_none()
    if existing is not None:
        existing.anchor_project_id = parent_id
        existing.anchor_milestone_id = ms_id
        existing.offset_months = offset_months
        existing.offset_days = offset_days
        session.add(existing)
    else:
        session.add(
            _PA(
                project_id=child_id,
                anchor_project_id=parent_id,
                anchor_milestone_id=ms_id,
                offset_months=offset_months,
                offset_days=offset_days,
            )
        )
    await session.commit()

    response = HTMLResponse("")
    response.headers["HX-Redirect"] = (
        f"/models/{model_id}/builder?view=underwriting"
    )
    return response


@router.delete(
    "/ui/models/{model_id}/anchors/{project_id}", response_class=HTMLResponse
)
async def delete_project_anchor(
    model_id: UUID,
    project_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Remove a ProjectAnchor row — project reverts to its own start date."""
    from app.models.project import ProjectAnchor as _PA
    existing = (
        await session.execute(
            select(_PA).where(_PA.project_id == project_id)
        )
    ).scalar_one_or_none()
    if existing is not None:
        await session.delete(existing)
        await session.commit()
    response = HTMLResponse("")
    response.headers["HX-Redirect"] = (
        f"/models/{model_id}/builder?view=underwriting"
    )
    return response


# ─────────────────────────────────────────────────────────────────────────────
# Phase 3c: Source coverage modal — edit which Projects a CapitalModule is
# attached to (capital_module_projects junction rows).
# ─────────────────────────────────────────────────────────────────────────────


@router.get(
    "/ui/models/{model_id}/sources/{source_id}/coverage",
    response_class=HTMLResponse,
)
async def source_coverage_modal(
    request: Request,
    model_id: UUID,
    source_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Render the Source Coverage modal.

    Lists every Project on the Scenario with an Included checkbox + per-
    project amount / active_from / active_to / auto_size fields. Existing
    ``capital_module_projects`` rows pre-fill their projects' inputs; not-
    yet-attached projects render with empty values + include-checkbox off.
    """
    from app.models.capital import CapitalModuleProject
    module = await session.get(CapitalModule, source_id)
    if module is None or module.scenario_id != model_id:
        return HTMLResponse(
            "<p class='text-muted'>Source not found.</p>", status_code=404
        )
    projects = list(
        (
            await session.execute(
                select(Project)
                .where(Project.scenario_id == model_id)
                .order_by(Project.created_at.asc())
            )
        ).scalars()
    )
    junction_rows = list(
        (
            await session.execute(
                select(CapitalModuleProject).where(
                    CapitalModuleProject.capital_module_id == source_id
                )
            )
        ).scalars()
    )
    by_project = {j.project_id: j for j in junction_rows}
    rows = []
    for p in projects:
        j = by_project.get(p.id)
        rows.append(
            {
                "project": p,
                "included": j is not None,
                "amount": j.amount if j else (module.source or {}).get("amount"),
                "active_from": j.active_from
                if j
                else module.active_phase_start,
                "active_to": j.active_to if j else module.active_phase_end,
                "auto_size": j.auto_size if j else False,
            }
        )
    return templates.TemplateResponse(
        request,
        "partials/underwriting/coverage_modal.html",
        {
            "model_id": str(model_id),
            "module": module,
            "rows": rows,
        },
    )


@router.post(
    "/ui/models/{model_id}/sources/{source_id}/coverage",
    response_class=HTMLResponse,
)
async def source_coverage_write(
    request: Request,
    model_id: UUID,
    source_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Write junction changes submitted by the coverage modal.

    Form fields per project (keyed by project_id):

        included[<pid>]         — "1" to attach (checkbox); absent = detach
        amount[<pid>]           — Decimal-ish string
        active_from[<pid>]      — milestone key
        active_to[<pid>]        — milestone key
        auto_size[<pid>]        — "1" / absent

    Invariant: a CapitalModule must retain at least one junction row. If the
    submitted form would leave zero, the last row stays untouched with a
    soft warning (TODO: surface warning to user — for now the handler
    silently keeps the row).
    """
    from app.models.capital import CapitalModuleProject
    module = await session.get(CapitalModule, source_id)
    if module is None or module.scenario_id != model_id:
        return HTMLResponse(
            "<p class='text-muted'>Source not found.</p>", status_code=404
        )
    form = await request.form()

    projects = list(
        (
            await session.execute(
                select(Project)
                .where(Project.scenario_id == model_id)
                .order_by(Project.created_at.asc())
            )
        ).scalars()
    )
    junction_rows = list(
        (
            await session.execute(
                select(CapitalModuleProject).where(
                    CapitalModuleProject.capital_module_id == source_id
                )
            )
        ).scalars()
    )
    by_project = {j.project_id: j for j in junction_rows}

    included_after: list[UUID] = []
    for p in projects:
        key = str(p.id)
        included = form.get(f"included[{key}]") == "1"
        if not included:
            continue
        included_after.append(p.id)

    # Orphan guard — never leave a module with zero junction rows. If the
    # submission would, keep the first existing row.
    if not included_after and junction_rows:
        fallback = junction_rows[0]
        included_after = [fallback.project_id]

    for p in projects:
        key = str(p.id)
        included = p.id in included_after
        amt_raw = (form.get(f"amount[{key}]") or "").strip()
        af = (form.get(f"active_from[{key}]") or "").strip() or None
        at = (form.get(f"active_to[{key}]") or "").strip() or None
        auto = form.get(f"auto_size[{key}]") == "1"
        amt = _fd(amt_raw)
        existing = by_project.get(p.id)
        if included:
            if existing is not None:
                if amt is not None:
                    existing.amount = amt
                existing.active_from = af
                existing.active_to = at
                existing.auto_size = auto
                session.add(existing)
            else:
                session.add(
                    CapitalModuleProject(
                        capital_module_id=source_id,
                        project_id=p.id,
                        amount=amt if amt is not None else Decimal("0"),
                        active_from=af,
                        active_to=at,
                        auto_size=auto,
                    )
                )
        elif existing is not None:
            await session.delete(existing)

    await session.commit()
    # Signal the client to reload the Underwriting page so every panel
    # (pill, staleness dots, source package, rollup CF) picks up the junction
    # change. HX-Redirect is the cleanest way from an HTMX form submission.
    response = HTMLResponse("")
    response.headers["HX-Redirect"] = (
        f"/models/{model_id}/builder?view=underwriting"
    )
    return response


# ─────────────────────────────────────────────────────────────────────────────
# Adopt Source — project-perspective recovery path. Surfaces scenario
# CapitalModules not yet junctioned to a given project, then writes the
# junction when one is chosen. Recovery for Add Project's share box being
# missed or wiped by a wizard re-run.
# ─────────────────────────────────────────────────────────────────────────────


@router.get(
    "/ui/models/{model_id}/projects/{project_id}/adopt-source",
    response_class=HTMLResponse,
)
async def adopt_source_modal(
    request: Request,
    model_id: UUID,
    project_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Render the Adopt Source picker for a project."""
    from app.models.capital import CapitalModuleProject
    project = await session.get(Project, project_id)
    if project is None or project.scenario_id != model_id:
        return HTMLResponse(
            "<p class='text-muted'>Project not found.</p>", status_code=404
        )
    all_modules = list((await session.execute(
        select(CapitalModule)
        .where(CapitalModule.scenario_id == model_id)
        .order_by(CapitalModule.stack_position.asc())
    )).scalars())
    attached_ids = set((await session.execute(
        select(CapitalModuleProject.capital_module_id).where(
            CapitalModuleProject.project_id == project_id
        )
    )).scalars())
    candidates = [m for m in all_modules if m.id not in attached_ids]
    return templates.TemplateResponse(
        request,
        "partials/adopt_source_modal.html",
        {
            "model_id": str(model_id),
            "project_id": str(project_id),
            "project_name": project.name,
            "candidates": candidates,
        },
    )


@router.post(
    "/ui/models/{model_id}/projects/{project_id}/adopt-source",
    response_class=HTMLResponse,
)
async def adopt_source_write(
    request: Request,
    model_id: UUID,
    project_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Attach the chosen scenario CapitalModule to this project."""
    from app.models.capital import CapitalModuleProject
    project = await session.get(Project, project_id)
    if project is None or project.scenario_id != model_id:
        return HTMLResponse(
            "<p class='text-muted'>Project not found.</p>", status_code=404
        )
    form = await request.form()
    src_raw = str(form.get("source_id", "")).strip()
    if not src_raw:
        return HTMLResponse(
            "<p class='text-muted'>Pick a Source.</p>", status_code=400
        )
    try:
        source_id = UUID(src_raw)
    except ValueError:
        return HTMLResponse(
            "<p class='text-muted'>Invalid Source.</p>", status_code=400
        )
    module = await session.get(CapitalModule, source_id)
    if module is None or module.scenario_id != model_id:
        return HTMLResponse(
            "<p class='text-muted'>Source not on this Deal.</p>", status_code=404
        )
    existing = (await session.execute(
        select(CapitalModuleProject).where(
            CapitalModuleProject.capital_module_id == source_id,
            CapitalModuleProject.project_id == project_id,
        )
    )).scalar_one_or_none()
    if existing is None:
        session.add(CapitalModuleProject(
            capital_module_id=source_id,
            project_id=project_id,
            amount=Decimal("0"),
            active_from=module.active_phase_start,
            active_to=module.active_phase_end,
            auto_size=bool((module.source or {}).get("auto_size")),
        ))
        await session.commit()
    response = HTMLResponse("")
    response.headers["HX-Redirect"] = (
        f"/models/{model_id}/builder?module=sources_uses&project={project_id}"
    )
    return response


def _compute_calc_status(data: dict) -> dict:
    """Produce the 3-factor calculation status: Sources=Uses, DSCR, LTV.

    Each factor returns a dict with:
      status: "ok" | "warn" | "fail" | "na"
      label: short human summary
      detail: longer explanation for the modal
      meta: structured data (numbers) for display

    Overall status rolls up to "ok" (green) only if ALL factors are ok/na.
    """
    capital_total = float(data.get("capital_total") or 0.0)
    uses_total = float(data.get("uses_total") or 0.0)
    outputs = data.get("outputs")
    inputs = data.get("inputs")
    capital_modules = data.get("capital_modules") or []

    # ── Factor 1: Sources vs Uses ──
    gap = capital_total - uses_total
    if not capital_total and not uses_total:
        su_status = {
            "status": "na",
            "label": "No sources/uses yet",
            "detail": "Add Sources and Uses to see balance check.",
            "meta": {"capital_total": 0, "uses_total": 0, "gap": 0},
        }
    elif abs(gap) < 1.0:
        su_status = {
            "status": "ok",
            "label": "Sources = Uses",
            "detail": f"Balanced at {_fmt_currency(uses_total)}.",
            "meta": {"capital_total": capital_total, "uses_total": uses_total, "gap": 0},
        }
    elif gap > 0:
        su_status = {
            "status": "warn",
            "label": f"Surplus {_fmt_currency(gap)}",
            "detail": f"Sources ({_fmt_currency(capital_total)}) exceed Uses ({_fmt_currency(uses_total)}) by {_fmt_currency(gap)}. The extra capital isn't needed — consider reducing a debt amount or equity.",
            "meta": {"capital_total": capital_total, "uses_total": uses_total, "gap": gap},
        }
    else:
        su_status = {
            "status": "fail",
            "label": f"Gap {_fmt_currency(-gap)}",
            "detail": f"Uses ({_fmt_currency(uses_total)}) exceed Sources ({_fmt_currency(capital_total)}) by {_fmt_currency(-gap)}. Either increase debt sizing (raise LTV or DSCR), reduce Uses, or add equity.",
            "meta": {"capital_total": capital_total, "uses_total": uses_total, "gap": gap},
        }

    # ── Factor 2: DSCR ──
    dscr_val = None
    dscr_min = None
    if outputs is not None and getattr(outputs, "dscr", None):
        try:
            dscr_val = float(outputs.dscr)
        except (TypeError, ValueError):
            dscr_val = None
    # DSCR floor: same precedence as the cashflow engine —
    #   source.dscr_min → operational_inputs.debt_terms.permanent_debt.dscr_min → 1.20
    _dt_perm_dscr: float | None = None
    if inputs is not None:
        _dt = getattr(inputs, "debt_terms", None) or {}
        if isinstance(_dt, dict):
            _dt_perm = _dt.get("permanent_debt") or {}
            if isinstance(_dt_perm, dict):
                try:
                    _v_dt = _dt_perm.get("dscr_min")
                    if _v_dt is not None:
                        _dt_perm_dscr = float(_v_dt)
                except (TypeError, ValueError):
                    pass
    for _cm in capital_modules:
        if str(getattr(_cm, "vehicle_type", "") or "").replace("VehicleType.", "") != "debt":
            continue
        _src = getattr(_cm, "source", None) or {}
        try:
            _v = _src.get("dscr_min") if isinstance(_src, dict) else None
            if _v is not None:
                dscr_min = float(_v)
            elif _dt_perm_dscr is not None:
                dscr_min = _dt_perm_dscr
            else:
                dscr_min = 1.20
        except (TypeError, ValueError):
            dscr_min = _dt_perm_dscr if _dt_perm_dscr is not None else 1.20
        break

    if dscr_val is None or dscr_min is None:
        dscr_status = {
            "status": "na",
            "label": "DSCR not computed",
            "detail": "Run Compute to calculate DSCR. Requires debt modules and an operational NOI.",
            "meta": {"dscr": None, "dscr_min": dscr_min},
        }
    elif dscr_val >= dscr_min:
        headroom = dscr_val - dscr_min
        dscr_status = {
            "status": "ok",
            "label": f"DSCR {dscr_val:.2f}× (min {dscr_min:.2f}×)",
            "detail": f"DSCR is {headroom:.2f}× above the minimum. The deal comfortably covers its debt service.",
            "meta": {"dscr": dscr_val, "dscr_min": dscr_min, "headroom": headroom},
        }
    else:
        shortfall = dscr_min - dscr_val
        dscr_status = {
            "status": "fail",
            "label": f"DSCR {dscr_val:.2f}× < min {dscr_min:.2f}×",
            "detail": f"DSCR is {shortfall:.2f}× below the minimum. The deal isn't producing enough NOI to cover debt service at lender requirements. Reduce debt, increase NOI, or lower the DSCR minimum if your lender allows.",
            "meta": {"dscr": dscr_val, "dscr_min": dscr_min, "shortfall": shortfall},
        }

    # ── Factor 3: LTV ──
    # Compute actual LTV = total non-bridge debt / property value (NOI / exit cap).
    # Always shown to the user as informational. Green/red treatment ONLY when
    # sizing mode is dual_constraint (because otherwise LTV isn't a constraint
    # the engine actively targets — it's just a derived number).
    sizing_mode = (getattr(inputs, "debt_sizing_mode", None) or "") if inputs else ""
    is_dual_constraint = (sizing_mode == "dual_constraint")

    # Actual LTV calculation. Multi-project: prefer the per-project junction
    # amount over the scenario-level source.amount. source.amount holds the
    # last-sized value (P2 in last-write-wins), so on P1's tab the LTV would
    # otherwise compute against P2's much larger principal — easily pushing
    # the displayed LTV above 100% for the smaller project.
    _DEC_ZERO = Decimal("0")
    _junction_amts = data.get("capital_junction_amts") or {}
    total_non_bridge_debt = _DEC_ZERO
    for m in capital_modules:
        src = m.source or {}
        if src.get("is_bridge"):
            continue
        if str(getattr(m, "vehicle_type", "") or "").replace("VehicleType.", "") == "debt":
            _jam = _junction_amts.get(str(m.id))
            amt = _jam if _jam is not None else src.get("amount")
            if amt:
                try:
                    total_non_bridge_debt += Decimal(str(amt))
                except Exception:
                    pass

    noi_dec = Decimal(str(outputs.noi_stabilized)) if (outputs and getattr(outputs, "noi_stabilized", None)) else _DEC_ZERO
    exit_cap = Decimal(str(inputs.exit_cap_rate_pct)) if (inputs and getattr(inputs, "exit_cap_rate_pct", None)) else _DEC_ZERO
    property_value = (noi_dec / (exit_cap / Decimal("100"))) if (noi_dec > 0 and exit_cap > 0) else _DEC_ZERO
    actual_ltv_pct = float(total_non_bridge_debt / property_value * Decimal("100")) if property_value > 0 and total_non_bridge_debt > 0 else None

    # Collect per-module binding info (for dual_constraint diagnostics)
    ltv_binding_modules: list[dict] = []
    any_has_ltv = False
    for m in capital_modules:
        src = m.source or {}
        if src.get("is_bridge"):
            continue
        binding = src.get("binding_constraint")
        ltv_pct_cfg = src.get("ltv_pct")
        if ltv_pct_cfg:
            any_has_ltv = True
        if binding == "ltv":
            ltv_binding_modules.append({
                "label": m.label,
                "ltv_pct": float(ltv_pct_cfg) if ltv_pct_cfg else None,
                "amount": float(src.get("amount") or 0),
            })

    # Headline LTV cap for the "max debt at X% LTV" display. Prefer the LTV
    # setting on the first non-bridge debt module with ltv_pct; fall back to
    # the engine's 65% default when dual_constraint is on; else None.
    _headline_ltv_pct: float | None = None
    for m in capital_modules:
        src = m.source or {}
        if src.get("is_bridge"):
            continue
        _cfg = src.get("ltv_pct")
        if _cfg:
            try:
                _headline_ltv_pct = float(_cfg)
                break
            except (TypeError, ValueError):
                pass
    if _headline_ltv_pct is None and is_dual_constraint:
        _headline_ltv_pct = 65.0

    max_debt_at_ltv: float | None = None
    if _headline_ltv_pct and property_value > 0:
        max_debt_at_ltv = float(property_value * Decimal(str(_headline_ltv_pct)) / Decimal("100"))

    ltv_meta = {
        "actual_ltv_pct": actual_ltv_pct,
        "total_debt": float(total_non_bridge_debt) if total_non_bridge_debt else 0,
        "property_value": float(property_value) if property_value else 0,
        "binding_modules": ltv_binding_modules,
        "headline_ltv_pct": _headline_ltv_pct,
        "max_debt_at_ltv": max_debt_at_ltv,
    }

    if actual_ltv_pct is None:
        # Diagnose exactly which input is missing so the user knows what to fix.
        _missing: list[str] = []
        if noi_dec <= 0:
            _missing.append("stabilized NOI (run Compute)")
        if exit_cap <= 0:
            _missing.append("exit cap rate (set in Settings or the Divestment module)")
        if total_non_bridge_debt <= 0:
            _missing.append("non-bridge debt (add a permanent loan source)")
        if _missing:
            _detail = "Missing: " + "; ".join(_missing) + "."
        else:
            _detail = "Derived property value came out to zero — check inputs."
        ltv_status = {
            "status": "na",
            "label": "LTV not computable",
            "detail": _detail,
            "meta": ltv_meta,
        }
    elif not is_dual_constraint:
        # gap_fill / dscr_capped: LTV is a derived outcome, not an active
        # constraint. Treat it as healthy (ok) when it lands inside the
        # configured LTV cap, warn when it exceeds the cap. Detail still
        # notes the user can switch to Dual-Constraint to make it actively bind.
        _cap = _headline_ltv_pct
        _within_cap = (_cap is None) or (actual_ltv_pct <= float(_cap) + 0.05)
        _detail = (
            f"Debt ${float(total_non_bridge_debt):,.0f} / property value "
            f"${float(property_value):,.0f} = {actual_ltv_pct:.1f}%. Sizing "
            f"mode is '{sizing_mode or 'gap_fill'}', so LTV is a derived "
            f"outcome — not an active constraint. Switch to Dual-Constraint "
            f"in Deal Setup to size debt by MIN(LTV, DSCR, gap-fill)."
        )
        if _within_cap:
            ltv_status = {
                "status": "ok",
                "label": f"LTV {actual_ltv_pct:.1f}%",
                "detail": _detail,
                "meta": ltv_meta,
            }
        else:
            ltv_status = {
                "status": "warn",
                "label": f"LTV {actual_ltv_pct:.1f}% — exceeds {_cap}% cap",
                "detail": _detail,
                "meta": ltv_meta,
            }
    elif ltv_binding_modules and gap < -1.0:
        first = ltv_binding_modules[0]
        pct = first.get("ltv_pct")
        ltv_status = {
            "status": "fail",
            "label": f"LTV {pct}% — binding with gap" if pct else "LTV binding with gap",
            "detail": (
                f"{first['label']} is sized at its LTV cap "
                f"({pct}% of $NOI/exit_cap property value)"
                if pct else
                f"{first['label']} is sized at its LTV cap"
            ) + (
                ". DSCR may have headroom, but dual-constraint sizing uses "
                "MIN(LTV, DSCR, gap-fill) — the lowest cap wins. To close "
                "the Sources gap, raise the LTV on this source (or switch "
                "to a different sizing mode)."
            ),
            "meta": ltv_meta,
        }
    elif ltv_binding_modules:
        first = ltv_binding_modules[0]
        pct = first.get("ltv_pct")
        ltv_status = {
            "status": "ok",
            "label": f"LTV {actual_ltv_pct:.1f}% (cap {pct}%)" if pct else f"LTV {actual_ltv_pct:.1f}%",
            "detail": f"{first['label']} is sized exactly at the LTV cap. Sources = Uses, so this is fine.",
            "meta": ltv_meta,
        }
    else:
        ltv_status = {
            "status": "ok",
            "label": f"LTV {actual_ltv_pct:.1f}% (slack)",
            "detail": "LTV is not the binding constraint — DSCR or gap-fill is sizing your debt. Plenty of LTV headroom.",
            "meta": ltv_meta,
        }

    # ── Overall rollup ──
    factors = [su_status, dscr_status, ltv_status]
    failing_count = sum(1 for f in factors if f["status"] in ("fail", "warn"))
    if failing_count == 0:
        overall = "ok"
    else:
        overall = "warn"

    return {
        "overall": overall,
        "failing_count": failing_count,
        "sources_uses": su_status,
        "dscr": dscr_status,
        "ltv": ltv_status,
    }


async def _get_gap_adjustment_amounts(
    session: AsyncSession, project_id: UUID
) -> dict[str, float]:
    """Return current Gap Adjustment phantom amounts as a dict.

    Keys: ``revenue_monthly`` (IncomeStream.amount_fixed_monthly),
    ``opex_annual`` (OperatingExpenseLine.annual_amount),
    ``pp`` (UseLine.amount).  Missing rows resolve to 0.0. Used by the
    calc-status modal to drive per-section yellow override + adjustment
    notes (Sources=Uses ← PP; DSCR/LTV ← Revenue + OpEx).
    """
    from app.schemas.gap_adjustment_names import (
        OPEX_ADJUSTMENT_LABEL,
        PURCHASE_PRICE_ADJUSTMENT_LABEL,
        REVENUE_ADJUSTMENT_LABEL,
    )

    out = {"revenue_monthly": 0.0, "opex_annual": 0.0, "pp": 0.0}
    rev = (await session.execute(
        select(IncomeStream).where(
            IncomeStream.project_id == project_id,
            IncomeStream.label == REVENUE_ADJUSTMENT_LABEL,
        )
    )).scalars().first()
    if rev and rev.amount_fixed_monthly is not None:
        try:
            out["revenue_monthly"] = float(rev.amount_fixed_monthly)
        except (TypeError, ValueError):
            pass
    opex = (await session.execute(
        select(OperatingExpenseLine).where(
            OperatingExpenseLine.project_id == project_id,
            OperatingExpenseLine.label == OPEX_ADJUSTMENT_LABEL,
        )
    )).scalars().first()
    if opex and opex.annual_amount is not None:
        try:
            out["opex_annual"] = float(opex.annual_amount)
        except (TypeError, ValueError):
            pass
    pp = (await session.execute(
        select(UseLine).where(
            UseLine.project_id == project_id,
            UseLine.label == PURCHASE_PRICE_ADJUSTMENT_LABEL,
        )
    )).scalars().first()
    if pp and pp.amount is not None:
        try:
            out["pp"] = float(pp.amount)
        except (TypeError, ValueError):
            pass
    return out


async def _has_any_gap_adjustment(session: AsyncSession, project_id: UUID) -> bool:
    """True iff at least one Gap Adjustment phantom row has a nonzero amount.

    Gap Adjustment phantom rows materialize the slider deltas; the
    calc-status pill stays yellow as long as any of them are non-zero,
    even if Sources=Uses / DSCR / LTV all individually pass — the user
    needs visible feedback that the model balances "with adjustments"
    rather than "for real."
    """
    from app.schemas.gap_adjustment_names import (
        OPEX_ADJUSTMENT_LABEL,
        PURCHASE_PRICE_ADJUSTMENT_LABEL,
        REVENUE_ADJUSTMENT_LABEL,
    )

    rev = (await session.execute(
        select(IncomeStream).where(
            IncomeStream.project_id == project_id,
            IncomeStream.label == REVENUE_ADJUSTMENT_LABEL,
        )
    )).scalars().first()
    if rev and rev.amount_fixed_monthly and float(rev.amount_fixed_monthly) != 0:
        return True
    opex = (await session.execute(
        select(OperatingExpenseLine).where(
            OperatingExpenseLine.project_id == project_id,
            OperatingExpenseLine.label == OPEX_ADJUSTMENT_LABEL,
        )
    )).scalars().first()
    if opex and opex.annual_amount and float(opex.annual_amount) != 0:
        return True
    pp = (await session.execute(
        select(UseLine).where(
            UseLine.project_id == project_id,
            UseLine.label == PURCHASE_PRICE_ADJUSTMENT_LABEL,
        )
    )).scalars().first()
    if pp and pp.amount and float(pp.amount) != 0:
        return True
    return False


def _render_calc_status_pill_html(
    status: dict,
    model_id: UUID,
    has_any_adjustment: bool = False,
) -> str:
    """Render the calc-status pill button HTML from a computed status dict.

    ``has_any_adjustment`` overrides "ok" → "warn" (yellow) so a model
    that pencils only via Gap Adjustment phantom rows is visibly distinct
    from a model that pencils unaided. Real failures still surface as
    "warn" with their existing label — the override only applies to the
    otherwise-green case.
    """
    if status["overall"] == "ok" and has_any_adjustment:
        label = "⚠ Balanced w/ adjustments"
        cls = "warn"
    elif status["overall"] == "ok":
        label = "✓ Calculation Valid"
        cls = "ok"
    else:
        n = status["failing_count"]
        _specific: list[str] = []
        su = status.get("sources_uses", {})
        if su.get("status") in ("fail", "warn"):
            _gap = (su.get("meta") or {}).get("gap") or 0
            if _gap < 0:
                _specific.append(f"-${abs(float(_gap)):,.0f} Sources Gap")
            elif _gap > 0:
                _specific.append(f"+${float(_gap):,.0f} Sources Surplus")
        _dscr_f = status.get("dscr", {})
        if _dscr_f.get("status") in ("fail", "warn"):
            _dscr_val = (_dscr_f.get("meta") or {}).get("dscr")
            if _dscr_val is not None:
                _specific.append(f"{float(_dscr_val):.2f}× DSCR — Too Low")
            else:
                _specific.append("DSCR — Issue")
        _ltv_f = status.get("ltv", {})
        if _ltv_f.get("status") in ("fail", "warn"):
            _actual = (_ltv_f.get("meta") or {}).get("actual_ltv_pct")
            if _actual is not None:
                _specific.append(f"{float(_actual):.1f}% LTV — Too High")
            else:
                _specific.append("LTV — Too High")
        if n == 1 and _specific:
            label = f"⚠ {_specific[0]}"
        else:
            label = f"⚠ {n} issue{'s' if n != 1 else ''}"
        cls = "warn"
    return (
        f'<button type="button" class="calc-status-pill {cls}" '
        f'hx-get="/ui/models/{model_id}/calc-status/modal" '
        f'hx-target="#calc-status-modal-body" '
        f'hx-swap="innerHTML" '
        f'onclick="document.getElementById(\'calc-status-modal\').style.display=\'flex\'">'
        f'{label}</button>'
    )


def _is_underwriting_view_request(request: Request) -> bool:
    """True when the user is currently on the Underwriting (rollup) view.

    HTMX sends HX-Current-URL on every refresh; the pill + modal routes
    use it to decide whether to render scenario-aggregate state (when the
    parent page is on ?view=underwriting) or per-project state.
    """
    _hx_url = request.headers.get("HX-Current-URL", "")
    if not _hx_url:
        return False
    from urllib.parse import urlparse, parse_qs
    return parse_qs(urlparse(_hx_url).query).get("view", [""])[0] == "underwriting"


async def _aggregate_status_for_underwriting(
    session: AsyncSession, model_id: UUID
) -> dict:
    """Build a calc_status-shaped dict from _compute_scenario_statuses so
    the pill on Underwriting reflects the worst-project severity instead
    of mirroring the default project's pill."""
    _scen_st = await _compute_scenario_statuses(session, model_id)
    _uw_st = (_scen_st.get("underwriting") or {})
    return {
        "overall": _uw_st.get("overall", "na"),
        "failing_count": int(_uw_st.get("failing_count") or 0),
        "sources_uses": {"status": "na", "label": "See per-project chips", "detail": "", "meta": {}},
        "dscr": {"status": "na", "label": "See per-project chips", "detail": "", "meta": {}},
        "ltv": {"status": "na", "label": "See per-project chips", "detail": "", "meta": {}},
    }


@router.get("/ui/models/{model_id}/calc-status", response_class=HTMLResponse)
async def model_calc_status_pill(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Returns the center-top calculation status pill HTML.

    On the per-project view: 3-factor (Sources=Uses, DSCR, LTV) for the
    active project. On the Underwriting view: scenario aggregate via
    _compute_scenario_statuses so the pill stays consistent with the tab
    chips and doesn't snap to Project 1's pill after modal interactions.
    """
    has_adj = False
    if _is_underwriting_view_request(request):
        status = await _aggregate_status_for_underwriting(session, model_id)
    else:
        _active_proj_id = await _active_project_from_request(request, session, model_id)
        data = await _load_builder_data(session, model_id, project_id=_active_proj_id)
        status = _compute_calc_status(data)
        if _active_proj_id is not None:
            has_adj = await _has_any_gap_adjustment(session, _active_proj_id)
    return HTMLResponse(
        _render_calc_status_pill_html(status, model_id, has_any_adjustment=has_adj)
    )


@router.get("/ui/models/{model_id}/calc-status/modal", response_class=HTMLResponse)
async def model_calc_status_modal(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Returns the modal body HTML with the 3-factor diagnostic.

    Emits an HX-Trigger response header so the topbar pill re-fetches its
    state whenever the modal opens — keeps pill and modal in lockstep.
    On Underwriting view the modal shows the aggregate (no specific
    factor drill-downs); per-project drilldowns are reachable via the
    project tab chips.
    """
    if _is_underwriting_view_request(request):
        status = await _aggregate_status_for_underwriting(session, model_id)
        # Underwriting modal shows per-project breakdown, not the 3-factor
        # diagnostic. Load full scenario_statuses + project list for the
        # template's severity-sorted card list.
        _scen_st_full = await _compute_scenario_statuses(session, model_id)
        _deal_projects = list(
            (
                await session.execute(
                    select(Project)
                    .where(Project.scenario_id == model_id)
                    .order_by(Project.created_at.asc())
                )
            ).scalars()
        )
        response = templates.TemplateResponse(
            request,
            "partials/calc_status_modal_underwriting.html",
            {
                "status": status,
                "model_id": str(model_id),
                "scenario_statuses": _scen_st_full,
                "deal_projects": _deal_projects,
            },
        )
    else:
        _active_proj_id = await _active_project_from_request(request, session, model_id)
        data = await _load_builder_data(session, model_id, project_id=_active_proj_id)
        status = _compute_calc_status(data)
        gap_adjustments = (
            await _get_gap_adjustment_amounts(session, _active_proj_id)
            if _active_proj_id is not None
            else {"revenue_monthly": 0.0, "opex_annual": 0.0, "pp": 0.0}
        )
        response = templates.TemplateResponse(
            request,
            "partials/calc_status_modal.html",
            {
                "status": status,
                "model_id": str(model_id),
                "gap_adjustments": gap_adjustments,
            },
        )
    response.headers["HX-Trigger"] = "calcStatusChanged"
    return response


@router.get("/ui/models/{model_id}/balance-bar", response_class=HTMLResponse)
async def model_balance_bar(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """DEPRECATED — kept for back-compat. Redirects to calc-status pill.
    Sidebar balance bar replaced by center-top status pill.
    """
    return await model_calc_status_pill(request, model_id, session)


@router.get("/ui/models/{model_id}/module-nav")
async def model_module_nav(
    request: Request,
    model_id: UUID,
    session: DBSession,
    module: str = "",
) -> _TemplateResponse:
    """Returns the module nav cards partial for sidebar live-refresh after mutations."""
    _active_proj_id = await _active_project_from_request(request, session, model_id)
    data = await _load_builder_data(session, model_id, project_id=_active_proj_id)
    ctx: dict[str, Any] = {
        "request": request,
        "active_module": module,
        "locked": not data.get("timeline_approved", False),
        "deal_setup_complete": data.get("deal_setup_complete", False),
        "nav_base_path": f"/models/{model_id}/builder",
        **{k: data.get(k) for k in (
            "capital_module_count", "capital_total",
            "use_line_count", "uses_total",
            "income_stream_count", "revenue_annual",
            "expense_line_count", "opex_annual",
            "capex_reserve_annual", "opex_total_annual",
            "carrying_annual",
            "equity_ownership", "org_owner_fallback",
            "deferred_uses", "deferred_total", "profit_total",
            "divestment_total", "phase_summaries", "outputs",
            "income_mode", "noi_annual",
            "unit_mix_count", "total_units",
            # Keep active project on every nav link — without this, _proj_qs
            # falls back to empty and OpEx/Sources/etc. links drop the project.
            "default_project_id",
        )},
    }
    return templates.TemplateResponse(request, "partials/model_builder_nav_cards.html", ctx)


@router.get("/ui/models/{model_id}/export.xlsx")
async def download_model_export(
    model_id: UUID,
    session: DBSession,
) -> StreamingResponse:
    """Download a round-trip-capable Excel workbook for this deal model.

    Deprecated path; superseded by ``/investor-export.xlsx``. Kept available
    while the investor export bakes — see plan §10 in
    ``docs/feature-plans/investor-excel-export-v2.md``.
    """
    from app.exporters.excel_export import export_deal_model_workbook, make_export_filename
    model = await session.get(DealModel, model_id)
    if model is None:
        return HTMLResponse("Not found", status_code=404)
    workbook_bytes = await export_deal_model_workbook(model_id, session)
    filename = make_export_filename(model)
    return StreamingResponse(
        iter([workbook_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/ui/models/{model_id}/investor-export.xlsx")
async def download_investor_export(
    model_id: UUID,
    session: DBSession,
    profile: str = Query(default="internal"),
) -> StreamingResponse:
    """Download the LP-facing investor Excel workbook for this Scenario."""
    from app.exporters.investor_export import export_investor_workbook, make_investor_filename
    scenario = await session.get(DealModel, model_id)
    if scenario is None:
        return HTMLResponse("Not found", status_code=404)
    deal = await session.get(Deal, scenario.deal_id) if scenario.deal_id else None
    workbook_bytes = await export_investor_workbook(model_id, session, profile=profile)
    filename = make_investor_filename(scenario, deal)
    return StreamingResponse(
        iter([workbook_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Async investor export ────────────────────────────────────────────────────
#
# The synchronous endpoint above blows past the NGINX 60s proxy timeout once
# the live Sensitivity matrix lands (25 cashflow cycles per export). The async
# path below kicks a Celery task that builds the workbook off the request
# path and emails the .xlsx as an attachment when finished. Job state
# (``queued → calculating → sending → sent`` or ``failed``) is persisted on
# ``ExportJob`` rows so the UI can poll for hover-modal updates and the user
# can resend a cached build when the scenario hasn't been recomputed since
# the last successful export.


@router.get("/ui/models/{model_id}/investor-export/preflight")
async def preflight_investor_export(
    model_id: UUID,
    session: DBSession,
    request: Request,
    profile: str = Query(default="internal"),
) -> JSONResponse:
    """Cheap idempotent check: is a cached resend eligible for this scenario?

    Returns ``{"resend_eligible": bool, "resend_job_id": <uuid|null>}``.
    UI uses this to decide whether to prompt the user with a "Resend last
    export?" modal before enqueueing a fresh build.

    Eligibility = last ``sent`` job's ``created_at`` > every
    ``OperationalOutputs.computed_at`` for this scenario AND that job
    still has ``xlsx_bytes`` cached.
    """
    from app.models.cashflow import OperationalOutputs
    from app.models.export_job import ExportJob, ExportJobStatus

    user = await _get_user(session, request)
    if user is None:
        return JSONResponse({"error": "not authenticated"}, status_code=401)
    scenario = await session.get(DealModel, model_id)
    if scenario is None:
        return JSONResponse({"error": "scenario not found"}, status_code=404)

    latest_outputs_computed_at = (
        await session.execute(
            select(func.max(OperationalOutputs.computed_at))
            .where(OperationalOutputs.scenario_id == model_id)
        )
    ).scalar_one_or_none()

    last_sent_job = (
        await session.execute(
            select(ExportJob)
            .where(ExportJob.scenario_id == model_id)
            .where(ExportJob.status == ExportJobStatus.sent)
            .where(ExportJob.xlsx_bytes.isnot(None))
            .where(ExportJob.export_profile == profile)
            .order_by(ExportJob.created_at.desc())
            .limit(1)
        )
    ).scalar_one_or_none()

    resend_eligible = bool(
        last_sent_job is not None
        and (
            latest_outputs_computed_at is None
            or last_sent_job.created_at > latest_outputs_computed_at
        )
    )
    return JSONResponse(
        {
            "resend_eligible": resend_eligible,
            "resend_job_id": (
                str(last_sent_job.id) if (resend_eligible and last_sent_job) else None
            ),
        }
    )


@router.post("/ui/models/{model_id}/investor-export/async")
async def start_investor_export_async(
    model_id: UUID,
    session: DBSession,
    request: Request,
) -> JSONResponse:
    """Enqueue a fresh-build investor-export job and return its id.

    Caller (UI) is expected to have hit ``/preflight`` first to decide
    whether to prompt for "Resend last export?" — this endpoint always
    builds fresh.
    """
    from app.models.export_job import ExportJob, ExportJobStatus
    from app.tasks.export import RUN_EXPORT_TASK
    from app.tasks.celery_app import celery_app as _celery

    user = await _get_user(session, request)
    if user is None:
        return JSONResponse({"error": "not authenticated"}, status_code=401)
    scenario = await session.get(DealModel, model_id)
    if scenario is None:
        return JSONResponse({"error": "scenario not found"}, status_code=404)

    try:
        body = await request.json()
    except Exception:  # noqa: BLE001
        body = {}
    _raw_profile = body.get("profile", "internal") if isinstance(body, dict) else "internal"
    _profile = _raw_profile if _raw_profile in {"internal", "lp", "lender", "proforma"} else "internal"

    job = ExportJob(
        scenario_id=model_id,
        user_id=user.id,
        recipient_email=user.email or "",
        status=ExportJobStatus.queued,
        export_profile=_profile,
    )
    session.add(job)
    await session.commit()
    await session.refresh(job)

    _celery.send_task(RUN_EXPORT_TASK, args=[str(job.id)])

    return JSONResponse(
        {
            "job_id": str(job.id),
            "status": job.status.value,
        }
    )


@router.get("/ui/exports/{job_id}/status")
async def get_export_job_status(
    job_id: UUID,
    session: DBSession,
    request: Request,
) -> JSONResponse:
    """Return current status of an export job for poll/hover-modal use."""
    from app.models.export_job import ExportJob

    user = await _get_user(session, request)
    if user is None:
        return JSONResponse({"error": "not authenticated"}, status_code=401)
    job = await session.get(ExportJob, job_id)
    if job is None or job.user_id != user.id:
        return JSONResponse({"error": "job not found"}, status_code=404)

    return JSONResponse(
        {
            "job_id": str(job.id),
            "status": job.status.value,
            "error_message": job.error_message,
            "created_at": job.created_at.isoformat() if job.created_at else None,
            "completed_at": job.completed_at.isoformat() if job.completed_at else None,
        }
    )


@router.post("/ui/exports/{job_id}/resend")
async def resend_investor_export_endpoint(
    job_id: UUID,
    session: DBSession,
    request: Request,
) -> JSONResponse:
    """Re-send a previously-completed export from cached xlsx_bytes.

    Spawns a fresh ``ExportJob`` row pointing at the same scenario; the
    resend task copies bytes from the source job before sending so the
    "last sent" lookup keeps walking forward in time.
    """
    from app.models.export_job import ExportJob, ExportJobStatus
    from app.tasks.export import RESEND_EXPORT_TASK
    from app.tasks.celery_app import celery_app as _celery

    user = await _get_user(session, request)
    if user is None:
        return JSONResponse({"error": "not authenticated"}, status_code=401)
    src = await session.get(ExportJob, job_id)
    if src is None or src.user_id != user.id:
        return JSONResponse({"error": "job not found"}, status_code=404)
    if not src.xlsx_bytes:
        return JSONResponse({"error": "no cached export bytes"}, status_code=409)

    new_job = ExportJob(
        scenario_id=src.scenario_id,
        user_id=user.id,
        recipient_email=user.email or src.recipient_email,
        status=ExportJobStatus.queued,
        xlsx_bytes=src.xlsx_bytes,
        filename=src.filename,
    )
    session.add(new_job)
    await session.commit()
    await session.refresh(new_job)

    _celery.send_task(RESEND_EXPORT_TASK, args=[str(new_job.id)])
    return JSONResponse(
        {
            "job_id": str(new_job.id),
            "status": new_job.status.value,
        }
    )


@router.get("/ui/models/{model_id}/import-template.xlsx")
async def download_import_template(model_id: UUID) -> StreamingResponse:
    """Download a pre-formatted Excel template for bulk import of Uses and OpEx line items."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()

    # ── Shared styles ──────────────────────────────────────────────────────────
    hdr_font = Font(bold=True, size=10, color="FFFFFF")
    hdr_fill_uses = PatternFill("solid", fgColor="2563EB")   # blue
    hdr_fill_opex = PatternFill("solid", fgColor="059669")   # green
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hint_font = Font(italic=True, size=9, color="6B7280")
    hint_fill = PatternFill("solid", fgColor="F9FAFB")

    def _set_col_width(ws, col_letter, width):
        ws.column_dimensions[col_letter].width = width

    def _header_row(ws, headers, fill):
        ws.append(headers)
        for i, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i)
            cell.font = hdr_font
            cell.fill = fill
            cell.alignment = hdr_align
        ws.row_dimensions[1].height = 28

    def _hint_row(ws, hints):
        ws.append(hints)
        for i, _ in enumerate(hints, 1):
            cell = ws.cell(row=2, column=i)
            cell.font = hint_font
            cell.fill = hint_fill
            cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[2].height = 36

    # ── Uses sheet ─────────────────────────────────────────────────────────────
    ws_uses = wb.active
    ws_uses.title = "Uses"
    _header_row(ws_uses, ["Label", "Phase", "Amount ($)", "Deferred Dev Fee?", "Notes"], hdr_fill_uses)
    _hint_row(ws_uses, [
        "e.g. Hard Costs, Soft Costs, Contingency",
        "acquisition | pre_development | construction | exit",
        "Dollar amount (no commas)",
        "yes / no — deferred developer fee?",
        "Optional notes",
    ])
    # Phase validation
    phase_dv = DataValidation(
        type="list",
        formula1='"acquisition,pre_development,construction,exit"',
        allow_blank=True,
    )
    ws_uses.add_data_validation(phase_dv)
    phase_dv.sqref = "B3:B500"
    # Deferred dv
    bool_dv = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
    ws_uses.add_data_validation(bool_dv)
    bool_dv.sqref = "D3:D500"
    # Widths
    for col, w in zip("ABCDE", [32, 22, 16, 18, 30]):
        _set_col_width(ws_uses, col, w)
    # 3 sample rows
    for label, phase, amt in [
        ("Hard Costs", "construction", 480000),
        ("Soft Costs", "construction", 72000),
        ("Contingency (10%)", "construction", 55200),
    ]:
        ws_uses.append([label, phase, amt, "no", ""])

    # ── OpEx sheet ─────────────────────────────────────────────────────────────
    ws_opex = wb.create_sheet("OpEx")
    _header_row(ws_opex, [
        "Label", "Amount", "Per", "Escalation (%/yr)",
        "Scale w/ Lease-Up?", "Lease-Up Floor (%)", "Active Phases", "Notes",
    ], hdr_fill_opex)
    _hint_row(ws_opex, [
        "e.g. Property Tax, Insurance",
        "Dollar value",
        "flat | per_unit | per_sqft_residential | per_sqft_commercial",
        "e.g. 3.0",
        "yes / no",
        "0–100 (% of stabilized when vacant)",
        "construction, lease_up, stabilized (comma-separated)",
        "Optional",
    ])
    per_dv = DataValidation(
        type="list",
        formula1='"flat,per_unit,per_sqft_residential,per_sqft_commercial"',
        allow_blank=True,
    )
    ws_opex.add_data_validation(per_dv)
    per_dv.sqref = "C3:C500"
    ws_opex.add_data_validation(bool_dv)
    bool_dv.sqref = "E3:E500"
    for col, w in zip("ABCDEFGH", [28, 14, 22, 16, 18, 18, 30, 24]):
        _set_col_width(ws_opex, col, w)
    # 3 sample rows
    for label, amt, per, esc, scale, floor, phases in [
        ("Property Tax", 18000, "flat", 3.0, "no", "", "stabilized"),
        ("Insurance", 9600, "flat", 3.0, "no", "", "stabilized"),
        ("Property Management", 8, "per_unit", 3.0, "yes", 25, "lease_up, stabilized"),
    ]:
        ws_opex.append([label, amt, per, esc, scale, floor, phases, ""])

    # ── Stream to response ─────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=import-template.xlsx"},
    )


# ---------------------------------------------------------------------------
# Pro forma import — preflight, upload, status, confirm, skip
# ---------------------------------------------------------------------------

def _render_proforma_sheet_picker(
    request: Request,
    model_id: UUID,
    task_id: str,
    content: bytes,
) -> HTMLResponse:
    """Read sheet names + first-row columns from xlsx bytes and return the
    sheet-picker fragment. Shared by the preflight dispatch and the
    reanalyze flow (which reuses bytes already in Redis)."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        sheet_names = wb.sheetnames

        sheet_columns: dict[str, list[str]] = {}
        for name in sheet_names:
            ws = wb[name]
            for row in ws.iter_rows(max_row=10, values_only=True):
                non_empty = [str(c).strip() for c in row if c is not None]
                if non_empty:
                    sheet_columns[name] = non_empty
                    break
            else:
                sheet_columns[name] = []
        wb.close()
    except Exception as exc:
        return HTMLResponse(f"<p class='text-red-500'>Could not read file: {exc}</p>", status_code=400)

    return templates.TemplateResponse(
        request,
        "partials/proforma_preflight.html",
        {
            "model_id": model_id,
            "task_id": task_id,
            "sheet_names": sheet_names,
            "sheet_columns": sheet_columns,
            "STANDARD_OPEX_CATEGORIES": STANDARD_OPEX_CATEGORIES,
        },
    )


async def _dispatch_proforma_preflight(
    *,
    request: Request,
    model_id: UUID,
    upload: UploadFile,
) -> HTMLResponse:
    """Stash the uploaded pro forma in redis, then return either the sheet
    picker (for .xlsx) or the parse-progress poller (for PDF/DOCX/HTML).

    Extracted helper so both the dedicated POST route and the Step 1 wizard
    handler can dispatch the same flow when a file rides along with the
    income-mode form.

    Computes SHA-256 of the file bytes. If a parse result already lives in
    Redis under ``proforma:filehash:{hash}:result`` (7-day TTL), returns the
    cache-hit fragment so the user can skip the LLM call.
    """
    import hashlib
    import os as _os

    content = await upload.read()
    if not content:
        return HTMLResponse("<p class='text-red-500'>Empty file uploaded.</p>", status_code=400)

    filename = upload.filename or ""
    ext = _os.path.splitext(filename)[1].lower().lstrip(".")
    file_kind = "xlsx" if ext in {"xlsx", "xlsm", "xlsb"} else "doc"
    file_hash = hashlib.sha256(content).hexdigest()

    task_id = str(_uuid_mod.uuid4())
    import redis as _redis  # type: ignore
    _RESUME_TTL = 7 * 86_400
    r = _redis.from_url(settings.redis_url, decode_responses=False)
    r.set(f"proforma:{task_id}:file", content, ex=86_400)
    r.set(f"proforma:{task_id}:filename", filename, ex=86_400)
    r.set(f"proforma:{task_id}:kind", file_kind, ex=86_400)
    r.set(f"proforma:{task_id}:file_hash", file_hash, ex=86_400)
    # Track the most-recent proforma hash + filename per scenario so the
    # wizard's Step 2 "Back" can resume on the review page instead of
    # forcing a re-upload. 7d TTL matches the cache TTL.
    r.set(f"scenario:{model_id}:last_proforma_hash", file_hash.encode(), ex=_RESUME_TTL)
    r.set(f"scenario:{model_id}:last_proforma_filename", filename.encode(), ex=_RESUME_TTL)

    # ── Content-hash cache check ───────────────────────────────────────────
    # Cache hit: skip the LLM call and render the review page directly with
    # cached data. A banner on the review page surfaces the cache origin and
    # re-analyze/purge actions.
    r_str = _redis.from_url(settings.redis_url, decode_responses=True)
    cached_raw = r_str.get(f"proforma:filehash:{file_hash}:result")
    if cached_raw:
        try:
            cached_result = json.loads(cached_raw)
        except Exception:
            cached_result = None
        if cached_result is not None:
            parsed_at = r_str.get(f"proforma:filehash:{file_hash}:parsed_at") or ""
            # Mirror cached result to task-keyed key so /proforma-confirm
            # (reads by task_id) finds it.
            r_str.set(f"proforma:{task_id}:result", cached_raw, ex=86_400)
            return templates.TemplateResponse(
                request,
                "partials/proforma_review.html",
                {
                    "model_id": model_id,
                    "task_id": task_id,
                    "unit_types": cached_result.get("unit_types", []),
                    "expense_lines": cached_result.get("expense_lines", []),
                    "warnings": cached_result.get("warnings", []),
                    "STANDARD_OPEX_CATEGORIES": STANDARD_OPEX_CATEGORIES,
                    "from_cache": True,
                    "file_hash": file_hash,
                    "filename": filename,
                    "parsed_at": parsed_at,
                },
            )

    if file_kind == "doc":
        # No sheets to pick — queue the task immediately and return the
        # progress poller. MarkitDown will convert the whole document.
        from app.tasks.proforma_parse import PARSE_PROFORMA_TASK
        from app.tasks.celery_app import celery_app as _celery
        _celery.send_task(
            PARSE_PROFORMA_TASK,
            kwargs={
                "task_id": task_id,
                "model_id": str(model_id),
                "revenue_sheet": "",
                "opex_sheet": "",
                "property_column": None,
                "file_kind": "doc",
            },
        )
        return templates.TemplateResponse(
            request,
            "partials/proforma_progress.html",
            {"model_id": model_id, "task_id": task_id},
        )

    return _render_proforma_sheet_picker(request, model_id, task_id, content)


@router.post("/ui/models/{model_id}/proforma-preflight", response_class=HTMLResponse)
async def proforma_preflight(
    request: Request,
    model_id: UUID,
    file: UploadFile = File(...),
) -> HTMLResponse:
    """Receive uploaded file at the dedicated endpoint. Thin wrapper over
    ``_dispatch_proforma_preflight`` — kept for direct re-upload (proforma-
    restart) and for any external callers that still POST a file directly."""
    return await _dispatch_proforma_preflight(
        request=request, model_id=model_id, upload=file,
    )


@router.get("/ui/models/{model_id}/proforma-restart", response_class=HTMLResponse)
async def proforma_restart(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Return the wizard at Step 1 so the user can upload a different file
    (or pick a different income mode). Re-uses the GET /setup handler so the
    full context (inputs, vehicles, phases) is populated."""
    return await deal_setup_wizard_get(
        request=request, model_id=model_id, session=session, step=1,
    )


@router.get("/ui/models/{model_id}/proforma-resume", response_class=HTMLResponse)
async def proforma_resume(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Wizard Step-2 "Back" target. If the scenario has a recent proforma
    import whose parse result is still cached, re-render the review page
    so the user can adjust line items without re-uploading. Falls back to
    Step 1 (upload UI) when nothing is cached."""
    import redis as _redis  # type: ignore

    r = _redis.from_url(settings.redis_url, decode_responses=True)
    file_hash = r.get(f"scenario:{model_id}:last_proforma_hash")
    if file_hash:
        cached_raw = r.get(f"proforma:filehash:{file_hash}:result")
        if cached_raw:
            try:
                result = json.loads(cached_raw)
            except Exception:
                result = None
            if result is not None:
                filename = r.get(f"scenario:{model_id}:last_proforma_filename") or ""
                parsed_at = r.get(f"proforma:filehash:{file_hash}:parsed_at") or ""
                # Synthetic task_id — re-analyze/purge routes will gracefully
                # handle missing file bytes (24h task TTL vs 7d hash TTL).
                resume_task_id = str(_uuid_mod.uuid4())
                r.set(f"proforma:{resume_task_id}:result", cached_raw, ex=86_400)
                return templates.TemplateResponse(
                    request,
                    "partials/proforma_review.html",
                    {
                        "model_id": model_id,
                        "task_id": resume_task_id,
                        "unit_types": result.get("unit_types", []),
                        "expense_lines": result.get("expense_lines", []),
                        "warnings": result.get("warnings", []),
                        "STANDARD_OPEX_CATEGORIES": STANDARD_OPEX_CATEGORIES,
                        "from_cache": True,
                        "file_hash": file_hash,
                        "filename": filename,
                        "parsed_at": parsed_at,
                    },
                )

    return await deal_setup_wizard_get(
        request=request, model_id=model_id, session=session, step=1,
    )


def _render_proforma_reanalyze(
    request: Request,
    model_id: UUID,
    task_id: str,
) -> HTMLResponse:
    """Run a fresh parse for an already-uploaded file (bytes still in Redis).

    For xlsx files, returns the sheet picker so the user can pick sheets again.
    For doc files (PDF/DOCX/etc.), queues the Celery task immediately and
    returns the progress poller.
    """
    import redis as _redis  # type: ignore

    r = _redis.from_url(settings.redis_url, decode_responses=False)
    file_bytes = r.get(f"proforma:{task_id}:file")
    if not file_bytes:
        return HTMLResponse(
            "<p class='text-red-500'>Upload expired. Please re-upload the file.</p>",
            status_code=410,
        )

    kind_raw = r.get(f"proforma:{task_id}:kind") or b"xlsx"
    file_kind = kind_raw.decode() if isinstance(kind_raw, bytes) else str(kind_raw)

    if file_kind == "doc":
        from app.tasks.proforma_parse import PARSE_PROFORMA_TASK
        from app.tasks.celery_app import celery_app as _celery
        _celery.send_task(
            PARSE_PROFORMA_TASK,
            kwargs={
                "task_id": task_id,
                "model_id": str(model_id),
                "revenue_sheet": "",
                "opex_sheet": "",
                "property_column": None,
                "file_kind": "doc",
            },
        )
        return templates.TemplateResponse(
            request,
            "partials/proforma_progress.html",
            {"model_id": model_id, "task_id": task_id},
        )

    return _render_proforma_sheet_picker(request, model_id, task_id, file_bytes)


@router.post("/ui/models/{model_id}/proforma-reanalyze", response_class=HTMLResponse)
async def proforma_reanalyze(
    request: Request,
    model_id: UUID,
    task_id: str = Form(...),
) -> HTMLResponse:
    """Skip the cache and run a fresh parse. Cache is left intact (use
    /proforma-purge-cache to delete the cached result)."""
    return _render_proforma_reanalyze(request, model_id, task_id)


@router.post("/ui/models/{model_id}/proforma-purge-cache", response_class=HTMLResponse)
async def proforma_purge_cache(
    request: Request,
    model_id: UUID,
    task_id: str = Form(...),
    file_hash: str = Form(...),
) -> HTMLResponse:
    """Delete the content-hash cache entry, then trigger a fresh parse."""
    import redis as _redis  # type: ignore

    if file_hash and len(file_hash) == 64:
        r = _redis.from_url(settings.redis_url, decode_responses=True)
        r.delete(f"proforma:filehash:{file_hash}:result")
        r.delete(f"proforma:filehash:{file_hash}:parsed_at")

    return _render_proforma_reanalyze(request, model_id, task_id)


@router.post("/ui/models/{model_id}/upload-proforma", response_class=HTMLResponse)
async def upload_proforma(
    request: Request,
    model_id: UUID,
    task_id: str = Form(...),
    revenue_sheet: str = Form(...),
    opex_sheet: str = Form(...),
    property_column: str = Form(""),
) -> HTMLResponse:
    """Queue the Celery parse task with the user-selected sheet/column coordinates,
    then return the progress-polling fragment."""
    from app.tasks.proforma_parse import PARSE_PROFORMA_TASK
    from app.tasks.celery_app import celery_app as _celery

    _celery.send_task(
        PARSE_PROFORMA_TASK,
        kwargs={
            "task_id": task_id,
            "model_id": str(model_id),
            "revenue_sheet": revenue_sheet,
            "opex_sheet": opex_sheet,
            "property_column": property_column or None,
            "file_kind": "xlsx",
        },
    )

    return templates.TemplateResponse(
        request,
        "partials/proforma_progress.html",
        {"model_id": model_id, "task_id": task_id},
    )


@router.get("/ui/models/{model_id}/proforma-status/{task_id}", response_class=HTMLResponse)
async def proforma_status(
    request: Request,
    model_id: UUID,
    task_id: str,
    session: DBSession,
) -> HTMLResponse:
    """HTMX poll endpoint. Returns progress fragment while running; switches to
    the review fragment when the task completes or errors."""
    import redis as _redis  # type: ignore

    r = _redis.from_url(settings.redis_url, decode_responses=True)
    raw = r.get(f"proforma:{task_id}:progress")

    if not raw:
        return templates.TemplateResponse(
            request,
            "partials/proforma_progress.html",
            {"model_id": model_id, "task_id": task_id, "step": 0, "total": 3, "message": "Queued…"},
        )

    progress = json.loads(raw)
    status = progress.get("status", "running")

    if status == "error":
        return templates.TemplateResponse(
            request,
            "partials/proforma_progress.html",
            {
                "model_id": model_id,
                "task_id": task_id,
                "error": progress.get("message", "Unknown error"),
            },
        )

    if status != "done":
        return templates.TemplateResponse(
            request,
            "partials/proforma_progress.html",
            {
                "model_id": model_id,
                "task_id": task_id,
                "step": progress.get("step", 0),
                "total": progress.get("total", 3),
                "message": progress.get("message", ""),
            },
        )

    # Done — load result and render review UI
    raw_result = r.get(f"proforma:{task_id}:result")
    result = json.loads(raw_result) if raw_result else {"unit_types": [], "expense_lines": [], "warnings": []}

    return templates.TemplateResponse(
        request,
        "partials/proforma_review.html",
        {
            "model_id": model_id,
            "task_id": task_id,
            "unit_types": result.get("unit_types", []),
            "expense_lines": result.get("expense_lines", []),
            "warnings": result.get("warnings", []),
            "STANDARD_OPEX_CATEGORIES": STANDARD_OPEX_CATEGORIES,
        },
    )


@router.post("/ui/models/{model_id}/proforma-confirm", response_class=HTMLResponse)
async def proforma_confirm(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Commit user-approved revenue and expense lines to the database.

    Accepts multipart form data built by the review template:
    - ``unit_type_name[]``, ``unit_type_count[]``, ``unit_type_sqft[]``,
      ``unit_type_rent[]`` — parallel arrays for each confirmed unit type
    - ``expense_label[]``, ``expense_amount[]``, ``expense_category[]``,
      ``expense_include[]`` — parallel arrays for each expense line
      (``expense_include`` contains the indices of rows the user kept checked)
    """
    from sqlalchemy import delete

    form = await request.form()

    deal_model = await session.get(DealModel, model_id)
    if not deal_model:
        raise HTTPException(status_code=404, detail="Deal model not found")

    # Multi-project deals: route import to the project the user is viewing
    # (from HX-Current-URL ?project=...). Fall back to oldest only if missing.
    project_id = await _active_project_from_request(request, session, model_id)
    if project_id is None:
        default_project = (await session.execute(
            select(Project).where(Project.scenario_id == model_id).order_by(Project.created_at).limit(1)
        )).scalar_one_or_none()
        if not default_project:
            raise HTTPException(status_code=400, detail="No project found")
        project_id = default_project.id

    inputs = (await session.execute(
        select(OperationalInputs).where(OperationalInputs.project_id == project_id)
    )).scalar_one_or_none()

    # ---------- Revenue / unit mix (JSONB on Project) ----------
    names = form.getlist("unit_type_name[]")
    counts = form.getlist("unit_type_count[]")
    sqfts = form.getlist("unit_type_sqft[]")
    rents = form.getlist("unit_type_rent[]")
    market_rents = form.getlist("unit_type_market_rent[]")
    # Indices of unit-mix rows the user kept checked. When the include list
    # is absent (older clients), accept all rows for backward compat.
    unit_included_raw = form.getlist("unit_type_include[]")
    unit_included = {int(i) for i in unit_included_raw if str(i).strip().isdigit()}
    unit_filter_enabled = bool(unit_included_raw)

    rent_type = (form.get("rent_type") or "in_place").strip().lower()
    rent_field = "market_rent_per_unit" if rent_type == "market" else "in_place_rent_per_unit"

    if names:
        proj_result = await session.execute(
            select(Project).where(Project.id == project_id)
        )
        project = proj_result.scalar_one()
        unit_mix_rows = []
        for idx, (name, count_s, sqft_s, rent_s) in enumerate(zip(names, counts, sqfts, rents)):
            if unit_filter_enabled and idx not in unit_included:
                continue
            name = name.strip()
            if not name:
                continue
            try:
                row = {
                    "label": name,
                    "unit_count": int((count_s or "0").replace(",", "")),
                    "avg_sqft": float((sqft_s or "0").replace(",", "")),
                    "beds": None,
                    "baths": None,
                    "unit_strategy": "base_escalation",
                    "notes": None,
                }
                row[rent_field] = float((rent_s or "0").replace(",", ""))
                # If user supplied a Market Rent alongside in-place, capture it
                # so LTL catchup can be modeled later without re-entry.
                if rent_type == "in_place" and idx < len(market_rents):
                    mkt_s = (market_rents[idx] or "").replace(",", "").strip()
                    if mkt_s:
                        try:
                            mkt_v = float(mkt_s)
                            if mkt_v > 0:
                                row["market_rent_per_unit"] = mkt_v
                        except ValueError:
                            pass
                unit_mix_rows.append(row)
            except Exception:
                pass
        if unit_mix_rows:
            from sqlalchemy.orm.attributes import flag_modified
            project.unit_mix = unit_mix_rows
            flag_modified(project, "unit_mix")
            session.add(project)

            # Seed IncomeStream rows so the Revenue tab is populated.
            # Overwrite semantics: drop existing streams for this project, then
            # insert one per unit mix row using rent_monthly as the base.
            await session.execute(
                delete(IncomeStream).where(IncomeStream.project_id == project_id)
            )
            for row in unit_mix_rows:
                rent = Decimal(str(row.get(rent_field) or 0))
                count = int(row.get("unit_count") or 0)
                if count <= 0:
                    continue
                mkt_rent_raw = row.get("market_rent_per_unit")
                mkt_rent = Decimal(str(mkt_rent_raw)) if mkt_rent_raw else None
                session.add(IncomeStream(
                    project_id=project_id,
                    label=f"{row['label']} Rent",
                    stream_type=IncomeStreamType.residential_rent,
                    unit_count=count,
                    amount_per_unit_monthly=rent,
                    catchup_target_rent=mkt_rent,
                    stabilized_occupancy_pct=Decimal("95"),
                    escalation_rate_pct_annual=Decimal("3"),
                    active_in_phases=["lease_up", "stabilized"],
                ))

    # ---------- OpEx lines ----------
    # Label field holds the mapped category (investor export groups by label).
    # Original source label is preserved in notes.
    orig_labels = form.getlist("expense_orig_label[]")
    labels = form.getlist("expense_label[]")
    amounts = form.getlist("expense_amount[]")
    included_indices = {int(i) for i in form.getlist("expense_include[]")}

    if labels:
        await session.execute(
            delete(OperatingExpenseLine).where(OperatingExpenseLine.project_id == project_id)
        )
        for idx, (orig_label, label, amount_s) in enumerate(zip(orig_labels, labels, amounts)):
            if idx not in included_indices:
                continue
            label = label.strip()
            if not label:
                continue
            try:
                session.add(OperatingExpenseLine(
                    project_id=project_id,
                    label=label,
                    annual_amount=Decimal((amount_s or "0").replace(",", "")),
                    escalation_rate_pct_annual=Decimal("3"),
                    active_in_phases=["lease_up", "stabilized"],
                    notes=orig_label.strip() if orig_label.strip() != label else None,
                ))
            except Exception:
                pass

    await session.commit()

    # Re-enter the wizard at Step 2 via the canonical GET handler so the full
    # context (source_vehicles_debt, phases_present, review_back_step, etc.)
    # is populated — without it the Source Vehicle dropdowns never render.
    return await deal_setup_wizard_get(
        request=request, model_id=model_id, session=session, step=2,
    )


@router.get("/ui/models/{model_id}/proforma-skip", response_class=HTMLResponse)
async def proforma_skip(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Skip pro forma import — advance wizard to Step 2 (debt types)."""
    # Delegate to the canonical GET so source_vehicles_debt et al. are
    # populated (needed for the Source Vehicle dropdown on each debt card).
    return await deal_setup_wizard_get(
        request=request, model_id=model_id, session=session, step=2,
    )


@router.post("/ui/models/{model_id}/noi-inputs", response_class=HTMLResponse)
async def save_noi_inputs(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Save NOI mode inputs (stabilized NOI + escalation rate) and return refreshed form."""
    model = await session.get(DealModel, model_id)
    if model is None:
        return HTMLResponse("Not found", status_code=404)
    default_project = (await session.execute(
        select(Project).where(Project.scenario_id == model_id).order_by(Project.created_at).limit(1)
    )).scalar_one_or_none()
    if default_project is None:
        return HTMLResponse("No project", status_code=400)
    inputs = (await session.execute(
        select(OperationalInputs).where(OperationalInputs.project_id == default_project.id)
    )).scalar_one_or_none()
    if inputs is None:
        return HTMLResponse("No inputs", status_code=400)

    form = await request.form()
    noi_raw = str(form.get("noi_stabilized_input", "")).strip()
    esc_raw = form.get("noi_escalation_rate_pct", "3")
    # Strip any display formatting ($ and commas) before parsing.
    noi_clean = noi_raw.replace("$", "").replace(",", "").strip()
    try:
        inputs.noi_stabilized_input = Decimal(noi_clean) if noi_clean else None
    except Exception:
        inputs.noi_stabilized_input = None
    try:
        inputs.noi_escalation_rate_pct = Decimal(str(esc_raw)) if esc_raw else Decimal("3")
    except Exception:
        inputs.noi_escalation_rate_pct = Decimal("3")
    # User explicitly submitted — clear the auto-seeded flag whether they
    # accepted the suggested value or overrode it. Banner disappears.
    inputs.noi_auto_seeded = False
    session.add(inputs)
    await session.commit()
    await session.refresh(inputs)

    _noi_val = float(inputs.noi_stabilized_input) if inputs.noi_stabilized_input else ""
    _esc_val = float(inputs.noi_escalation_rate_pct) if inputs.noi_escalation_rate_pct else 3.0
    html = f"""<form hx-post="/ui/models/{model_id}/noi-inputs"
        hx-target="this"
        hx-swap="outerHTML"
        style="max-width:480px">
    <div style="background:var(--success-faint,#f0fdf4);border:1px solid var(--success,#22c55e);border-radius:6px;padding:8px 12px;margin-bottom:16px;font-size:12px;color:var(--success,#16a34a)">
      ✓ NOI inputs saved.
    </div>
    <div class="field-group" style="margin-bottom:20px">
      <label style="display:block;font-size:12px;font-weight:600;text-transform:uppercase;letter-spacing:.04em;color:var(--text-secondary);margin-bottom:4px">Stabilized NOI (Annual)</label>
      <input type="number" name="noi_stabilized_input" step="1000" min="0"
             value="{_noi_val}" placeholder="e.g. 500000"
             style="width:100%;box-sizing:border-box;padding:8px 10px;border:1px solid var(--border);border-radius:6px;font-size:14px;background:var(--bg);color:var(--text)">
      <div style="font-size:11px;color:var(--text-muted);margin-top:3px">Net Operating Income at stabilization — pre-debt service, post-OpEx (even though OpEx is not modeled separately).</div>
    </div>
    <div class="field-group" style="margin-bottom:20px">
      <label style="display:block;font-size:12px;font-weight:600;text-transform:uppercase;letter-spacing:.04em;color:var(--text-secondary);margin-bottom:4px">Annual NOI Escalation Rate (%)</label>
      <input type="number" name="noi_escalation_rate_pct" step="0.25" min="0" max="20"
             value="{_esc_val}" placeholder="3.0"
             style="width:140px;box-sizing:border-box;padding:8px 10px;border:1px solid var(--border);border-radius:6px;font-size:14px;background:var(--bg);color:var(--text)">
      <div style="font-size:11px;color:var(--text-muted);margin-top:3px">Compound annual growth applied to NOI each year. Typical: 2–4%.</div>
    </div>
    <div>
      <button type="submit" class="btn btn-primary">Save NOI Inputs</button>
    </div>
  </form>"""
    return HTMLResponse(html)


@router.get("/ui/models/{model_id}/line-form", response_class=HTMLResponse)
async def model_builder_line_form(
    request: Request,
    model_id: UUID,
    session: DBSession,
    type: str = Query(default="uses"),
    id: str = Query(default=""),
    phase: str = Query(default=""),
    category: str = Query(default="soft"),
) -> HTMLResponse:
    """Serves the add/edit form inside the line-item drawer."""
    model = await session.get(DealModel, model_id)
    if model is None:
        return HTMLResponse("<p class='text-muted'>Model not found.</p>", status_code=404)

    existing = None
    if id:
        try:
            eid = UUID(id)
            if type in ("use_lines", "uses"):
                existing = await session.get(UseLine, eid)
            elif type in ("income_streams", "revenue"):
                existing = await session.get(IncomeStream, eid)
            elif type in ("expense_lines", "opex"):
                existing = await session.get(OperatingExpenseLine, eid)
            elif type in ("capital_modules", "sources"):
                existing = await session.get(CapitalModule, eid)
            elif type in ("waterfall_tiers", "waterfall"):
                existing = await session.get(WaterfallTier, eid)
            elif type in ("milestones", "timeline"):
                existing = await session.get(Milestone, eid)
            elif type == "unit_mix":
                _lf_proj = (await session.execute(
                    select(Project).where(Project.scenario_id == model_id).order_by(Project.created_at.asc()).limit(1)
                )).scalar_one_or_none()
                if _lf_proj:
                    _um_dict = next((d for d in (_lf_proj.unit_mix or []) if d.get("id") == str(eid)), None)
                    if _um_dict:
                        existing = _UMRow(_um_dict)
        except ValueError:
            pass

    # For milestone forms: load siblings + compute which would be circular triggers
    sibling_milestones = []
    circular_ids: set = set()
    trigger_end_date = None  # ISO string passed to JS for end-date preview
    default_trigger_id: str | None = None
    if type in ("milestones", "timeline"):
        # Determine which project's milestones to load for the trigger dropdown:
        # 1. If editing an existing milestone, use its own project_id.
        # 2. If adding, use the ?project= query param passed by the caller.
        # 3. Fall back to the first (oldest) project only as a last resort.
        ms_project_id: UUID | None = None
        if existing is not None and hasattr(existing, "project_id"):
            ms_project_id = existing.project_id
        if ms_project_id is None:
            _proj_param = request.query_params.get("project", "")
            if _proj_param:
                try:
                    ms_project_id = UUID(_proj_param)
                except ValueError:
                    pass
        if ms_project_id is None:
            _fp = (await session.execute(
                select(Project).where(Project.scenario_id == model_id).order_by(Project.created_at.asc()).limit(1)
            )).scalar_one_or_none()
            ms_project_id = _fp.id if _fp else None
        if ms_project_id is not None:
            all_ms = list((await session.execute(
                select(Milestone).where(Milestone.project_id == ms_project_id)
            )).scalars())
            _SPHASE_ORDER = [
                "offer_made", "under_contract", "close", "pre_development",
                "construction", "operation_lease_up", "operation_stabilized", "divestment",
            ]
            def _sphase_idx(m):
                raw = str(m.milestone_type).replace("MilestoneType.", "")
                return next((i for i, v in enumerate(_SPHASE_ORDER) if v == raw), 99)
            editing_id = existing.id if existing else None
            sibling_milestones = sorted(
                [m for m in all_ms if m.id != editing_id],
                key=_sphase_idx
            )
            ms_map_local = {m.id: m for m in all_ms}
            # Detect circular: candidate Y is circular if following Y's trigger chain hits editing_id
            if editing_id:
                for candidate in sibling_milestones:
                    visited: set = set()
                    cur = candidate
                    while cur and cur.trigger_milestone_id:
                        if cur.trigger_milestone_id == editing_id:
                            circular_ids.add(candidate.id)
                            break
                        if cur.id in visited:
                            break
                        visited.add(cur.id)
                        cur = ms_map_local.get(cur.trigger_milestone_id)
            # default_trigger_id is no longer used by the template — predecessor
            # auto-selection is handled client-side by _msAutoTrigger() JS in the form.
            # Kept as a no-op so the template context key still exists.
            # Resolve trigger's end date so JS can preview end date on the form
            if existing and existing.trigger_milestone_id:
                trigger = ms_map_local.get(existing.trigger_milestone_id)
                if trigger:
                    t_end = trigger.computed_end(ms_map_local)
                    if t_end:
                        trigger_end_date = t_end.isoformat()

    # Lock duration for operation_stabilized when no divestment milestone exists
    lock_duration = False
    _STABILIZED_AUTO_DAYS = 10950
    if (
        existing
        and hasattr(existing, "milestone_type")
        and str(existing.milestone_type).replace("MilestoneType.", "") == "operation_stabilized"
    ):
        has_div = any(
            str(m.milestone_type).replace("MilestoneType.", "") == "divestment"
            for m in sibling_milestones
        )
        if not has_div:
            lock_duration = True

    _PHASE_LABELS = {
        "offer_made": "Offer Made", "under_contract": "Under Contract",
        "close": "Close / Acquisition", "pre_development": "Pre-Development",
        "construction": "Construction", "operation_lease_up": "Lease-Up",
        "operation_stabilized": "Stabilized Operations", "divestment": "Divestment / Exit",
    }

    # Phase options scoped to this deal type — prevents assigning costs to phases that don't exist
    _project_type_str = str(getattr(model, "project_type", "") or "").replace("ProjectType.", "")
    _USE_PHASES_BY_TYPE: dict[str, list[tuple[str, str]]] = {
        "acquisition": [
            ("acquisition", "Acquisition"),
            ("operation", "Operations"),
            ("exit", "Exit / Sale"),
            ("other", "Other"),
        ],
        "value_add": [
            ("acquisition", "Acquisition"),
            ("pre_construction", "Pre-Development"),
            ("construction", "Construction / Renovation"),
            ("operation", "Operations"),
            ("exit", "Exit / Sale"),
            ("other", "Other"),
        ],
        "conversion": [
            ("acquisition", "Acquisition"),
            ("conversion", "Conversion"),
            ("operation", "Operations"),
            ("exit", "Exit / Sale"),
            ("other", "Other"),
        ],
        "new_construction": [
            ("acquisition", "Acquisition"),
            ("pre_construction", "Pre-Construction"),
            ("construction", "Construction"),
            ("operation", "Operations"),
            ("exit", "Exit / Sale"),
            ("other", "Other"),
        ],
    }
    _default_phases = [
        ("acquisition", "Acquisition"), ("pre_construction", "Pre-Construction"),
        ("construction", "Construction"), ("renovation", "Renovation"),
        ("conversion", "Conversion"), ("operation", "Operations"),
        ("exit", "Exit / Sale"), ("other", "Other"),
    ]
    valid_use_phases = _USE_PHASES_BY_TYPE.get(_project_type_str, _default_phases)

    # For capital module and use line forms: load milestones for pickers
    milestones_dated_ds: list[dict] = []
    draw_source_window = None
    if type in ("capital_modules", "sources", "use_lines", "uses"):
        from app.models.project import Project as _LFProject
        from app.models.milestone import Milestone as _LFMilestone
        _lf_proj = (await session.execute(
            select(_LFProject).where(_LFProject.scenario_id == model_id).order_by(_LFProject.created_at.asc()).limit(1)
        )).scalar_one_or_none()
        if _lf_proj:
            _lf_opp_ms = list((await session.execute(
                select(_LFMilestone).where(_LFMilestone.opportunity_id == _lf_proj.opportunity_id)
            )).scalars()) if _lf_proj.opportunity_id else []
            _lf_proj_ms = list((await session.execute(
                select(_LFMilestone).where(_LFMilestone.project_id == _lf_proj.id)
            )).scalars())
            _lf_all_ms = _lf_opp_ms + _lf_proj_ms
            _lf_ms_map = {m.id: m for m in _lf_all_ms}
            for m in _lf_all_ms:
                _start = m.computed_start(_lf_ms_map)
                if _start:
                    _key = m.milestone_type.value if hasattr(m.milestone_type, "value") else str(m.milestone_type)
                    milestones_dated_ds.append({"key": _key, "label": _milestone_label(_key), "date": _start})
            milestones_dated_ds.sort(key=lambda x: x["date"])
            # Append "maturity" pseudo-milestone for the Active To dropdown only
            milestones_dated_ds.append({"key": "maturity", "label": "Maturity", "date": None})
        if existing and type in ("capital_modules", "sources"):
            # Prefer lookup by capital_module_id (reliable for wizard-created sources);
            # fall back to label match for legacy sources created before the FK existed.
            _ds_q = select(DrawSource).where(
                DrawSource.scenario_id == model_id,
                DrawSource.capital_module_id == existing.id,
            ).limit(1)
            draw_source_window = (await session.execute(_ds_q)).scalar_one_or_none()
            if draw_source_window is None:
                _ds_q = select(DrawSource).where(
                    DrawSource.scenario_id == model_id,
                    DrawSource.label == existing.label,
                ).limit(1)
                draw_source_window = (await session.execute(_ds_q)).scalar_one_or_none()

    # Exit Vehicle dropdown options (capital modules only). Dynamic from the
    # current module's active_phase_end + siblings' active windows.
    exit_vehicle_options: list[dict] = []
    show_exit_vehicle = False
    show_active_window = False
    # Exit Vehicle applies to all debt modules (vehicle_type == "debt").
    _EXIT_VEHICLE_APPLIES_UI: set[str] = set()  # unused — replaced by vehicle_type check below
    if type in ("capital_modules", "sources"):
        from app.engines.cashflow import (
            _APS_TO_RANK as _EXIT_APS_RANK,
            _resolve_vehicle as _exit_resolve,
        )

        siblings = list((await session.execute(
            select(CapitalModule).where(CapitalModule.scenario_id == model_id)
        )).scalars())
        others = [m for m in siblings if not existing or m.id != existing.id]
        # Build a "candidate" module stand-in for the resolve call — for new
        # modules we have no saved active_phase_end yet; default to "perpetuity"
        # (→ Maturity as the only option) to match the form's initial blank
        # state.  For existing modules we use their actual saved values.
        # New-source wizards haven't picked active_phase_end yet — so
        # eligible-by-rank gives zero results. Fall back to "all other
        # modules" so the user can pre-select a takeout target. The engine
        # re-validates at compute time and falls back to maturity if the
        # eventual active_phase_end doesn't actually overlap.
        is_new = existing is None
        if not is_new:
            candidate = existing
        else:
            class _Stub:  # minimal shim
                id = None
                active_phase_start = "acquisition"
                active_phase_end = ""
                exit_terms: dict = {}
            candidate = _Stub()

        _vehicle_now, _retirer_now = _exit_resolve(candidate, [candidate] + others)
        saved_val = ""
        if existing is not None and isinstance(existing.exit_terms, dict):
            saved_val = (existing.exit_terms.get("vehicle") or "").strip()

        # Compute eligible source retirers via same rank logic used by engine
        e_rank = _EXIT_APS_RANK.get(
            str(getattr(candidate, "active_phase_end", "") or ""), 99
        )

        def _rank(m: object, side: str) -> int:
            raw = str(getattr(m, f"active_phase_{side}", "") or "")
            if side == "end":
                return _EXIT_APS_RANK.get(raw, 99)
            return _EXIT_APS_RANK.get(raw, 0)

        # List all other sources as candidates. Overlap is too brittle a
        # filter — adjacent-vs-overlapping distinctions flip on rank
        # mapping (a new loan often starts the day the old closes). The
        # engine honours the user's explicit pick at compute time.
        eligible_sources = list(others)

        def _opt(value: str, label: str) -> dict:
            # If saved vehicle is present, honour it; else default to what
            # _resolve_vehicle picked.
            if saved_val:
                selected = (value == saved_val)
            elif _vehicle_now == "source" and _retirer_now is not None:
                selected = (value == str(getattr(_retirer_now, "id", "")))
            else:
                selected = (value == _vehicle_now)
            return {"value": value, "label": label, "selected": selected}

        exit_vehicle_options.append(_opt("maturity", "Maturity"))
        # Sale is always a valid exit for any debt instrument — the asset can
        # be sold at any point, retiring outstanding balances.
        exit_vehicle_options.append(_opt("sale", "Sale (divestment)"))
        for m in sorted(
            eligible_sources,
            key=lambda r: (int(getattr(r, "stack_position", 0) or 0), str(r.label or "")),
        ):
            exit_vehicle_options.append(_opt(str(m.id), m.label or "(unlabeled)"))

        # Gate Exit Vehicle + draw cadence UI on vehicle type.  Non-debt vehicle types
        # (equity, grants, forgivable loans) don't have a repayment concept — form hides them.
        _existing_vt = ""
        if existing is not None:
            _existing_vt = str(getattr(existing, "vehicle_type", "") or "").replace("VehicleType.", "")
        # New modules default to debt (see line form template default).
        _effective_vt = _existing_vt or "debt"
        show_exit_vehicle = _effective_vt == "debt"
        show_active_window = show_exit_vehicle

    # Source vehicle presets for the wizard dropdown (only shown on add, not edit)
    _sv_list: list[dict] = []
    if not id:
        _lf_user = await _get_user(session, request)
        if _lf_user is not None:
            from app.models.source_vehicle import SourceVehicle as _SV_lf
            _all_svs_lf = (await session.execute(
                select(_SV_lf).where(
                    ((_SV_lf.scope == "org") & (_SV_lf.owner_id == _lf_user.org_id)) |
                    ((_SV_lf.scope == "user") & (_SV_lf.owner_id == _lf_user.id))
                ).order_by(_SV_lf.label)
            )).scalars().all()
            _sv_list = [
                {
                    "id": str(v.id),
                    "name": v.label,
                    "vehicle_type": v.vehicle_type,
                    "equity_role": v.equity_role or "",
                    "owner": v.scope,
                }
                for v in _all_svs_lf
            ]

    # Per-Use eligibility checklist (grant cap UI). Pull all Use lines in this
    # scenario across projects so the source-side edit form can render
    # checkboxes for each Use; pre-tick those already referencing this module.
    _eligibility_uses: list[dict] = []
    if type in ("capital_modules", "sources", "capital-modules"):
        from app.schemas.gap_adjustment_names import is_reserved_label
        _ul_rows = (await session.execute(
            select(UseLine)
            .join(Project, UseLine.project_id == Project.id)
            .where(Project.scenario_id == model_id)
            .order_by(UseLine.label.asc())
        )).scalars().all()
        _existing_id_str = str(existing.id) if existing is not None else ""
        for _ul in _ul_rows:
            if is_reserved_label(_ul.label or ""):
                continue
            if not _ul.amount or float(_ul.amount) <= 0:
                continue
            _eligible_ids = _ul.eligible_module_ids or []
            _is_ticked = any(str(x) == _existing_id_str for x in _eligible_ids) if _existing_id_str else False
            _eligibility_uses.append({
                "id": str(_ul.id),
                "label": _ul.label or "(unlabeled)",
                "amount": float(_ul.amount or 0),
                "phase": str(getattr(_ul.phase, "value", _ul.phase) or ""),
                "ticked": _is_ticked,
            })

    return templates.TemplateResponse(request, "partials/model_builder_line_form.html", {
        "model": model,
        "form_type": type,
        "existing": existing,
        "default_phase": phase or "acquisition",
        "sibling_milestones": sibling_milestones,
        "circular_ids": circular_ids,
        "trigger_end_date": trigger_end_date,
        "default_trigger_id": default_trigger_id,
        "lock_duration": lock_duration,
        "phase_labels": _PHASE_LABELS,
        "valid_use_phases": valid_use_phases,
        "milestones_dated_ds": milestones_dated_ds,
        "draw_source_window": draw_source_window,
        "exit_vehicle_options": exit_vehicle_options,
        "show_exit_vehicle": show_exit_vehicle,
        "show_active_window": show_active_window,
        "exit_vehicle_applies": [],  # deprecated — templates now use vehicle_type == "debt" check
        "opex_categories": STANDARD_OPEX_CATEGORIES,
        "default_category": (getattr(existing, "cost_category", None) if existing else None) or category,
        "use_cost_categories": USE_COST_CATEGORIES,
        "use_category_labels": USE_CATEGORY_LABELS,
        "use_category_presets": USE_CATEGORY_PRESETS,
        "source_vehicles": _sv_list,
        "eligibility_uses": _eligibility_uses,
    })


# ---------------------------------------------------------------------------
# Source Vehicle prefill endpoint
# ---------------------------------------------------------------------------


@router.get("/ui/source-vehicles/{vehicle_id}/prefill")
async def source_vehicle_prefill(
    request: Request,
    vehicle_id: UUID,
    session: DBSession,
) -> JSONResponse:
    """Return flat form-field values for a source vehicle (used by wizard dropdown JS)."""
    from app.models.source_vehicle import SourceVehicle as _SV_pf

    user = await _get_user(session, request)
    if user is None:
        raise HTTPException(status_code=401, detail="Not authenticated")

    vehicle = (await session.execute(
        select(_SV_pf).where(
            _SV_pf.id == vehicle_id,
            (
                ((_SV_pf.scope == "org") & (_SV_pf.owner_id == user.org_id)) |
                ((_SV_pf.scope == "user") & (_SV_pf.owner_id == user.id))
            ),
        )
    )).scalar_one_or_none()

    if vehicle is None:
        raise HTTPException(status_code=404, detail="Vehicle not found")

    owner = vehicle.scope

    source = vehicle.source_config or {}
    carry = vehicle.carry_config or {}
    exit_cfg = vehicle.exit_config or {}

    phases = carry.get("phases", [])
    constr = next((p for p in phases if p.get("name") == "construction"), {})
    oper = next((p for p in phases if p.get("name") == "operation"), {})

    exit_vehicle_raw = exit_cfg.get("vehicle") or exit_cfg.get("exit_vehicle")
    # Only pre-fill exit vehicle if it's a named sentinel (not a stale UUID from another deal)
    safe_exit_vehicle = exit_vehicle_raw if exit_vehicle_raw in ("maturity", "sale") else None

    # Prefer ORM columns (set by vehicle_form.html) over JSONB source_config (set by
    # settings inline forms) so both creation paths return consistent prefill data.
    _rate = (float(vehicle.interest_rate_pct) if vehicle.interest_rate_pct is not None
             else source.get("interest_rate_pct"))
    _amort = vehicle.amort_term_years or source.get("amort_term_years")
    _fallback_ct = vehicle.carry_type  # set by vehicle_form.html; absent in settings-created vehicles
    _constr_ct = constr.get("carry_type") or _fallback_ct
    _oper_ct = oper.get("carry_type") or _fallback_ct

    return JSONResponse({
        "vehicle_name": vehicle.name,
        "owner": owner,
        "vehicle_type": vehicle.vehicle_type,
        "equity_role": vehicle.equity_role,
        "source_interest_rate": _rate,
        "ltv_pct": source.get("ltv_pct"),
        "amort_term_years": _amort,
        "hold_term_years": source.get("hold_term_years"),
        "dscr_min": source.get("dscr_min"),
        "construction_carry_type": _constr_ct,
        "operation_carry_type": _oper_ct,
        "perm_rate_pct": oper.get("perm_rate_pct"),
        "perm_term_years": oper.get("perm_term_years"),
        "perm_conversion_trigger": oper.get("perm_conversion_trigger"),
        "exit_type": exit_cfg.get("exit_type"),
        "exit_vehicle": safe_exit_vehicle,
        "draw_every_n_months": source.get("draw_every_n_months"),
        "draw_active_from_milestone": source.get("draw_active_from_milestone"),
        "draw_active_from_offset_days": source.get("draw_active_from_offset_days"),
        "carry_schedule": carry.get("schedule"),
    })


# ---------------------------------------------------------------------------
# Source Vehicle management (Phase G)
# ---------------------------------------------------------------------------

_VEHICLE_TYPE_LABELS = {
    "equity": "Equity",
    "debt": "Debt",
    "forgivable_loan": "Forgivable Loan",
    "grant": "Grant",
}


@router.get("/settings/vehicles", response_class=HTMLResponse)
async def vehicle_settings_page(
    request: Request,
    session: DBSession,
) -> HTMLResponse:
    user = await _get_user(session, request)
    if user is None:
        return RedirectResponse(url="/login?next=/settings/vehicles", status_code=303)

    from app.models.source_vehicle import SourceVehicle as _SV_list
    dedup_count, conflicts_count = await _get_counts(session)
    address_issues_count = await _get_address_issues_count(session, user)

    org_vehicles = (
        await session.execute(
            select(_SV_list).where(
                _SV_list.scope == "org", _SV_list.owner_id == user.org_id
            ).order_by(_SV_list.vehicle_type, _SV_list.label)
        )
    ).scalars().all()

    user_vehicles = (
        await session.execute(
            select(_SV_list).where(
                _SV_list.scope == "user", _SV_list.owner_id == user.id
            ).order_by(_SV_list.vehicle_type, _SV_list.label)
        )
    ).scalars().all()

    return templates.TemplateResponse(
        request,
        "settings_vehicles.html",
        {
            "org_vehicles": org_vehicles,
            "user_vehicles": user_vehicles,
            "vehicle_type_labels": _VEHICLE_TYPE_LABELS,
            **_base_ctx(user, dedup_count, "", address_issues_count, conflicts_count=conflicts_count),
        },
    )


@router.get("/settings/vehicles/{vehicle_id}/form", response_class=HTMLResponse)
async def vehicle_edit_form(
    request: Request,
    vehicle_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    user = await _get_user(session, request)
    if user is None:
        return HTMLResponse("Unauthorized", status_code=401)

    from app.models.source_vehicle import SourceVehicle as _SV_ef
    vehicle = (await session.execute(
        select(_SV_ef).where(
            _SV_ef.id == vehicle_id,
            (
                ((_SV_ef.scope == "org") & (_SV_ef.owner_id == user.org_id)) |
                ((_SV_ef.scope == "user") & (_SV_ef.owner_id == user.id))
            ),
        )
    )).scalar_one_or_none()
    if vehicle is None:
        return HTMLResponse("Vehicle not found", status_code=404)

    return templates.TemplateResponse(
        request,
        "partials/vehicle_form.html",
        {"vehicle": vehicle, "vehicle_type_labels": _VEHICLE_TYPE_LABELS},
    )


@router.post("/settings/vehicles", response_class=HTMLResponse)
async def vehicle_create(
    request: Request,
    session: DBSession,
) -> HTMLResponse:
    user = await _get_user(session, request)
    if user is None:
        return HTMLResponse("Unauthorized", status_code=401)

    form = await request.form()
    scope = str(form.get("scope", "org")).strip()
    label = str(form.get("label", "")).strip()
    vehicle_type = str(form.get("vehicle_type", "equity")).strip()
    equity_role = str(form.get("equity_role", "")).strip() or None

    if not label:
        return HTMLResponse("<p class='text-muted'>Label is required.</p>", status_code=400)
    if vehicle_type not in _VEHICLE_TYPE_LABELS:
        return HTMLResponse("<p class='text-muted'>Invalid vehicle type.</p>", status_code=400)
    if vehicle_type == "equity" and equity_role not in ("gp", "lp"):
        return HTMLResponse("<p class='text-muted'>Equity vehicles must have GP or LP role.</p>", status_code=400)
    if vehicle_type != "equity":
        equity_role = None

    owner_id = user.org_id if scope == "org" else user.id

    from app.models.source_vehicle import SourceVehicle as _SV_cr
    _v_carry_config = _parse_vehicle_carry_schedule(form)
    vehicle = _SV_cr(
        scope=scope,
        owner_id=owner_id,
        label=label,
        vehicle_type=vehicle_type,
        equity_role=equity_role,
        default_waterfall_position=int(form.get("default_waterfall_position") or 0),
        draw_cadence=str(form.get("draw_cadence", "monthly") or "monthly"),
        interest_rate_pct=form.get("interest_rate_pct") or None,
        carry_type=form.get("carry_type") or None,
        day_count_convention=str(form.get("day_count_convention", "actual_360") or "actual_360"),
        io_period_months=int(form.get("io_period_months")) if form.get("io_period_months") else None,
        amort_term_years=int(form.get("amort_term_years")) if form.get("amort_term_years") else None,
        pref_rate_pct=form.get("pref_rate_pct") or None,
        carry_config=_v_carry_config if _v_carry_config else None,
        created_by=user.id,
        updated_by=user.id,
    )
    session.add(vehicle)
    await session.commit()

    return RedirectResponse(url="/settings/vehicles", status_code=303)


@router.post("/settings/vehicles/{vehicle_id}", response_class=HTMLResponse)
async def vehicle_update(
    request: Request,
    vehicle_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    user = await _get_user(session, request)
    if user is None:
        return HTMLResponse("Unauthorized", status_code=401)

    from app.models.source_vehicle import SourceVehicle as _SV_up
    vehicle = (await session.execute(
        select(_SV_up).where(
            _SV_up.id == vehicle_id,
            (
                ((_SV_up.scope == "org") & (_SV_up.owner_id == user.org_id)) |
                ((_SV_up.scope == "user") & (_SV_up.owner_id == user.id))
            ),
        )
    )).scalar_one_or_none()
    if vehicle is None:
        return HTMLResponse("Vehicle not found", status_code=404)

    form = await request.form()
    vehicle.label = str(form.get("label", vehicle.label)).strip()
    new_vt = str(form.get("vehicle_type", vehicle.vehicle_type)).strip()
    if new_vt in _VEHICLE_TYPE_LABELS:
        vehicle.vehicle_type = new_vt
    er = str(form.get("equity_role", "")).strip() or None
    vehicle.equity_role = er if vehicle.vehicle_type == "equity" else None
    vehicle.draw_cadence = str(form.get("draw_cadence", vehicle.draw_cadence) or vehicle.draw_cadence)
    vehicle.day_count_convention = str(form.get("day_count_convention", vehicle.day_count_convention) or "actual_360")
    vehicle.interest_rate_pct = form.get("interest_rate_pct") or None
    vehicle.carry_type = form.get("carry_type") or None
    vehicle.io_period_months = int(form.get("io_period_months")) if form.get("io_period_months") else None
    vehicle.amort_term_years = int(form.get("amort_term_years")) if form.get("amort_term_years") else None
    vehicle.pref_rate_pct = form.get("pref_rate_pct") or None
    vehicle.default_waterfall_position = int(form.get("default_waterfall_position") or vehicle.default_waterfall_position)
    _new_carry_config = _parse_vehicle_carry_schedule(form)
    vehicle.carry_config = _new_carry_config if _new_carry_config else vehicle.carry_config
    vehicle.updated_by = user.id
    await session.commit()

    # Propagate carry schedule to all CapitalModules using this vehicle.
    if _new_carry_config and _new_carry_config.get("schedule"):
        _linked_modules = (await session.execute(
            select(CapitalModule).where(CapitalModule.source_vehicle_id == vehicle_id)
        )).scalars().all()
        for _lm in _linked_modules:
            _lm_carry = dict(_lm.carry or {})
            if not _lm_carry.get("_schedule_override"):
                _lm_carry["schedule"] = _new_carry_config["schedule"]
                _lm.carry = _lm_carry
                session.add(_lm)
        if _linked_modules:
            await session.commit()

    return RedirectResponse(url="/settings/vehicles", status_code=303)


@router.delete("/settings/vehicles/{vehicle_id}", response_class=HTMLResponse)
async def vehicle_delete(
    request: Request,
    vehicle_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    user = await _get_user(session, request)
    if user is None:
        return HTMLResponse("Unauthorized", status_code=401)

    from app.models.source_vehicle import SourceVehicle as _SV_del
    vehicle = (await session.execute(
        select(_SV_del).where(
            _SV_del.id == vehicle_id,
            (
                ((_SV_del.scope == "org") & (_SV_del.owner_id == user.org_id)) |
                ((_SV_del.scope == "user") & (_SV_del.owner_id == user.id))
            ),
        )
    )).scalar_one_or_none()
    if vehicle is None:
        return HTMLResponse("", status_code=404)

    await session.delete(vehicle)
    await session.commit()
    return HTMLResponse("")


# ---------------------------------------------------------------------------
# Portfolios
# ---------------------------------------------------------------------------


@router.get("/portfolios", response_class=HTMLResponse)
async def portfolios_page(
    request: Request,
    session: DBSession,
) -> HTMLResponse:
    user = await _get_user(session, request)
    dedup_count, conflicts_count = await _get_counts(session)

    portfolios_result = await session.execute(
        select(Portfolio)
        .options(
            selectinload(Portfolio.portfolio_projects)
                .selectinload(PortfolioProject.opportunity),
            selectinload(Portfolio.portfolio_projects)
                .selectinload(PortfolioProject.scenario)
                .selectinload(DealModel.operational_outputs),
        )
        .order_by(Portfolio.created_at.desc())
    )
    portfolios = list(portfolios_result.scalars().unique())

    # Build summary row per portfolio
    portfolio_rows = []
    for p in portfolios:
        deal_count = len(p.portfolio_projects)
        irr_values = [
            float(pp.scenario.operational_outputs.project_irr_levered)
            for pp in p.portfolio_projects
            if pp.scenario and pp.scenario.operational_outputs
            and pp.scenario.operational_outputs.project_irr_levered is not None
        ]
        avg_irr = sum(irr_values) / len(irr_values) if irr_values else None
        portfolio_rows.append({
            "id": str(p.id),
            "name": p.name,
            "deal_count": deal_count,
            "avg_irr": avg_irr,
            "created_at_fmt": p.created_at.strftime("%b %-d, %Y") if p.created_at else None,
        })

    return templates.TemplateResponse(
        request, "portfolios.html",
        {
            "portfolios": portfolio_rows,
            **_base_ctx(user, dedup_count, "portfolios", conflicts_count=conflicts_count),
        },
    )


@router.get("/ui/deals/search", response_class=HTMLResponse)
async def deals_search(
    request: Request,
    session: DBSession,
    q: str = Query(default=""),
) -> HTMLResponse:
    """HTMX deal search — returns an <ul> of results for portfolio add-deal picker."""
    if not q or len(q) < 2:
        return HTMLResponse("")
    stmt = (
        select(Deal)
        .where(Deal.name.ilike(f"%{q}%"), Deal.status != DealStatus.archived)
        .order_by(Deal.name)
        .limit(8)
    )
    results = list((await session.execute(stmt)).scalars())
    if not results:
        return HTMLResponse('<li style="padding:8px 12px;color:var(--text-muted);font-size:13px">No deals found</li>')
    items = "".join(
        f'<li style="padding:8px 12px;cursor:pointer;font-size:13px;border-bottom:1px solid var(--border)" '
        f'onclick="document.getElementById(\'deal-id-input\').value=\'{deal.id}\'; '
        f'document.getElementById(\'deal-search-display\').value=\'{deal.name.replace(chr(39), chr(39)+chr(39))}\'; '
        f'document.getElementById(\'deal-search-results\').innerHTML=\'\'">'
        f'{deal.name}</li>'
        for deal in results
    )
    return HTMLResponse(items)


@router.post("/ui/portfolios/create", response_class=HTMLResponse)
async def create_portfolio(
    request: Request,
    session: DBSession,
) -> HTMLResponse:
    form = await request.form()
    name = str(form.get("name", "")).strip()
    if not name:
        return HTMLResponse("<p class='text-muted'>Portfolio name is required.</p>", status_code=400)

    user = await _get_user(session, request)
    org_id = user.org_id if user else None
    if org_id is None:
        from app.models.org import Organization
        first_org = (await session.execute(select(Organization).limit(1))).scalar_one_or_none()
        if first_org is None:
            return HTMLResponse("<p class='text-muted'>No organization found.</p>", status_code=400)
        org_id = first_org.id

    p = Portfolio(org_id=org_id, name=name)
    session.add(p)
    await session.commit()
    return RedirectResponse(url=f"/portfolios/{p.id}", status_code=303)


@router.get("/portfolios/{portfolio_id}", response_class=HTMLResponse)
async def portfolio_detail(
    request: Request,
    portfolio_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    user = await _get_user(session, request)
    dedup_count, conflicts_count = await _get_counts(session)

    portfolio = await session.get(
        Portfolio,
        portfolio_id,
        options=[
            selectinload(Portfolio.portfolio_projects)
                .selectinload(PortfolioProject.opportunity),
            selectinload(Portfolio.portfolio_projects)
                .selectinload(PortfolioProject.scenario)
                .selectinload(DealModel.operational_outputs),
        ],
    )
    if portfolio is None:
        return HTMLResponse("<p class='text-muted'>Portfolio not found.</p>", status_code=404)

    # Build deal summary rows
    deal_rows = []
    for pp in portfolio.portfolio_projects:
        out = pp.scenario.operational_outputs if pp.scenario else None
        deal_rows.append({
            "opportunity_id": str(pp.project_id),
            "opportunity_name": pp.opportunity.name if pp.opportunity else "—",
            "scenario_id": str(pp.scenario_id) if pp.scenario_id else None,
            "scenario_name": pp.scenario.name if pp.scenario else None,
            "noi": float(out.noi_stabilized) if out and out.noi_stabilized is not None else None,
            "irr": float(out.project_irr_levered) if out and out.project_irr_levered is not None else None,
            "equity_required": float(out.equity_required) if out and out.equity_required is not None else None,
        })

    # Build Gantt — find Deals whose scenarios/projects reference these opportunity IDs
    opp_ids = [pp.project_id for pp in portfolio.portfolio_projects if pp.project_id]
    gantt_rows: list[dict] = []
    if opp_ids:
        deals_stmt = (
            select(Deal)
            .join(DealModel, DealModel.deal_id == Deal.id)
            .join(Project, Project.scenario_id == DealModel.id)
            .where(Project.opportunity_id.in_(opp_ids))
            .options(
                selectinload(Deal.scenarios).selectinload(DealModel.projects).selectinload(Project.milestones),
                selectinload(Deal.scenarios).selectinload(DealModel.projects).selectinload(Project.opportunity),
            )
            .distinct()
        )
        deals_for_gantt = list((await session.execute(deals_stmt)).scalars().unique())

        # Match each pp opportunity → Deal, build entries list
        entries = []
        for deal in deals_for_gantt:
            opp = _first_opportunity(deal)
            if opp is None or opp.id not in opp_ids:
                continue
            scenario = _primary_scenario(deal)
            entries.append((deal.name, scenario.name if scenario else "", deal))

        gantt_data = _build_portfolio_gantt(entries)

    # Aggregate stats
    irr_values = [r["irr"] for r in deal_rows if r["irr"] is not None]
    equity_values = [r["equity_required"] for r in deal_rows if r["equity_required"] is not None]
    noi_values = [r["noi"] for r in deal_rows if r["noi"] is not None]

    return templates.TemplateResponse(
        request, "portfolio_detail.html",
        {
            "portfolio": portfolio,
            "portfolio_id": str(portfolio.id),
            "portfolio_name": portfolio.name,
            "deal_rows": deal_rows,
            "gantt_data": gantt_data,
            "stats": {
                "deal_count": len(deal_rows),
                "avg_irr": sum(irr_values) / len(irr_values) if irr_values else None,
                "total_equity": sum(equity_values) if equity_values else None,
                "total_noi": sum(noi_values) if noi_values else None,
            },
            **_base_ctx(user, dedup_count, "portfolios", conflicts_count=conflicts_count),
        },
    )


@router.post("/ui/portfolios/{portfolio_id}/add-deal", response_class=HTMLResponse)
async def portfolio_add_deal(
    request: Request,
    portfolio_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Add a deal (by Deal.id) to a portfolio."""
    form = await request.form()
    deal_id_raw = str(form.get("deal_id", "")).strip()
    try:
        deal_id = UUID(deal_id_raw)
    except ValueError:
        return HTMLResponse("<p class='text-muted'>Invalid deal ID.</p>", status_code=400)

    # Resolve opportunity + active scenario from the Deal
    deal = await session.get(
        Deal, deal_id,
        options=[
            selectinload(Deal.scenarios).selectinload(DealModel.projects),
        ],
    )
    if deal is None:
        return HTMLResponse("<p class='text-muted'>Deal not found.</p>", status_code=404)

    active_scenario = _primary_scenario(deal)
    _first_proj = active_scenario.projects[0] if active_scenario and active_scenario.projects else None
    if _first_proj is None or _first_proj.opportunity_id is None:
        return HTMLResponse("<p class='text-muted'>Deal has no linked opportunity.</p>", status_code=400)

    # Upsert — skip if opportunity already in portfolio
    existing = (await session.execute(
        select(PortfolioProject).where(
            PortfolioProject.portfolio_id == portfolio_id,
            PortfolioProject.project_id == _first_proj.opportunity_id,
        )
    )).scalar_one_or_none()

    if existing is None:
        pp = PortfolioProject(
            portfolio_id=portfolio_id,
            project_id=_first_proj.opportunity_id,
            scenario_id=active_scenario.id if active_scenario else None,
        )
        session.add(pp)
    await session.commit()

    return RedirectResponse(url=f"/portfolios/{portfolio_id}", status_code=303)


@router.post("/ui/portfolios/{portfolio_id}/remove-deal", response_class=HTMLResponse)
async def portfolio_remove_deal(
    request: Request,
    portfolio_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    from sqlalchemy import delete as sa_delete
    form = await request.form()
    opp_id_raw = str(form.get("opportunity_id", "")).strip()
    try:
        opp_id = UUID(opp_id_raw)
    except ValueError:
        return HTMLResponse("<p class='text-muted'>Invalid opportunity ID.</p>", status_code=400)

    await session.execute(
        sa_delete(PortfolioProject).where(
            PortfolioProject.portfolio_id == portfolio_id,
            PortfolioProject.project_id == opp_id,
        )
    )
    await session.commit()
    return RedirectResponse(url=f"/portfolios/{portfolio_id}", status_code=303)


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Data Cleanup
# ---------------------------------------------------------------------------

@router.get("/dedup", response_class=HTMLResponse)
async def dedup_page(
    request: Request, session: DBSession,
    tab: str = Query(default="pending"),
) -> HTMLResponse:
    user = await _get_user(session, request)
    dedup_count, conflicts_count = await _get_counts(session)
    address_issues_count = await _get_address_issues_count(session)

    address_issues: list[ScrapedListing] = []
    candidates: list[Any] = []

    if tab == "address_issues":
        ai_result = await session.execute(
            select(ScrapedListing)
            .where(
                ScrapedListing.realie_skip.is_(True),
                ScrapedListing.realie_enriched_at.is_(None),
                ScrapedListing.apn.is_(None),
            )
            .order_by(ScrapedListing.city.asc(), ScrapedListing.street.asc())
        )
        address_issues = list(ai_result.scalars())
    elif tab == "resolved":
        result = await session.execute(
            select(DedupCandidate)
            .where(DedupCandidate.status != DedupStatus.pending)
            .order_by(DedupCandidate.resolved_at.desc())
            .limit(200)
        )
        candidates = list(result.scalars())
    else:
        result = await session.execute(
            select(DedupCandidate)
            .where(DedupCandidate.status == DedupStatus.pending)
            .order_by(DedupCandidate.confidence_score.desc())
        )
        candidates = list(result.scalars())

    listings_map = await _load_listings_for_candidates(candidates, session)
    rows = [_candidate_row(c, listings_map) for c in candidates]

    return templates.TemplateResponse(request, "dedup.html", {
        "request": request,
        "tab": tab,
        "candidates": rows,
        "address_issues": address_issues,
        **_base_ctx(user, dedup_count, "dedup", address_issues_count, conflicts_count=conflicts_count),
    })


@router.get("/ui/dedup/{candidate_id}/compare", response_class=HTMLResponse)
async def dedup_compare(
    request: Request, candidate_id: UUID, session: DBSession,
) -> HTMLResponse:
    candidate = await session.get(DedupCandidate, candidate_id)
    if candidate is None:
        return HTMLResponse("<div class='text-muted text-small'>Candidate not found.</div>")

    a_type = _record_type_str(candidate.record_a_type)
    b_type = _record_type_str(candidate.record_b_type)
    compare: dict[str, Any] = {"conflicts": [], "matches": []}
    record_a: ScrapedListing | None = None
    record_b: ScrapedListing | None = None

    if a_type == RecordType.listing.value and b_type == RecordType.listing.value:
        record_a = await session.get(ScrapedListing, candidate.record_a_id)
        record_b = await session.get(ScrapedListing, candidate.record_b_id)
        if record_a and record_b:
            compare = _build_listing_compare(record_a, record_b)

    src_a = (record_a.source.title() if record_a else a_type.title())
    src_b = (record_b.source.title() if record_b else b_type.title())
    addr_a = (record_a.address_raw or record_a.full_address or "—") if record_a else "—"
    addr_b = (record_b.address_raw or record_b.full_address or "—") if record_b else "—"
    url_a = getattr(record_a, "source_url", None) if record_a else None
    url_b = getattr(record_b, "source_url", None) if record_b else None

    return templates.TemplateResponse(request, "partials/dedup_compare.html", {
        "request": request,
        "candidate_id": str(candidate_id),
        "src_a": src_a,
        "src_b": src_b,
        "addr_a": addr_a,
        "addr_b": addr_b,
        "url_a": url_a,
        "url_b": url_b,
        "conflicts": compare["conflicts"],
        "matches": compare["matches"],
    })


@router.post("/ui/dedup/{candidate_id}/keep-separate", response_class=HTMLResponse)
async def ui_dedup_keep_separate(
    request: Request, candidate_id: UUID, session: DBSession,
) -> HTMLResponse:
    candidate = await session.get(DedupCandidate, candidate_id)
    if candidate is None:
        return HTMLResponse("")
    user = await _get_user(session, request)
    candidate.status = DedupStatus.kept_separate
    candidate.resolved_by_user_id = user.id if user else None
    candidate.resolved_at = datetime.now(UTC)
    await session.flush()
    return HTMLResponse(
        f'<tr id="dedup-row-{candidate_id}" class="text-muted" style="opacity:.4">'
        f'<td colspan="6" style="padding:10px 12px;font-size:12px">✓ Marked as separate records</td>'
        f'</tr>'
    )


@router.post("/ui/dedup/{candidate_id}/resolve", response_class=HTMLResponse)
async def ui_dedup_resolve(
    request: Request, candidate_id: UUID, session: DBSession,
) -> HTMLResponse:
    candidate = await session.get(DedupCandidate, candidate_id)
    if candidate is None:
        return HTMLResponse("")

    form = await request.form()
    action = str(form.get("action", "keep_separate"))
    winner = str(form.get("winner", "a"))
    user = await _get_user(session, request)

    if action == "keep_separate":
        candidate.status = DedupStatus.kept_separate
        candidate.resolved_by_user_id = user.id if user else None
        candidate.resolved_at = datetime.now(UTC)
        await session.flush()
        return HTMLResponse(
            f'<tr id="dedup-row-{candidate_id}" class="text-muted" style="opacity:.4">'
            f'<td colspan="6" style="padding:10px 12px;font-size:12px">✓ Kept as separate records</td>'
            f'</tr>'
        )

    # merge: apply field choices, mark loser as duplicate of winner
    a_type = _record_type_str(candidate.record_a_type)
    b_type = _record_type_str(candidate.record_b_type)

    if a_type == RecordType.listing.value and b_type == RecordType.listing.value:
        rec_a = await session.get(ScrapedListing, candidate.record_a_id)
        rec_b = await session.get(ScrapedListing, candidate.record_b_id)
        if rec_a and rec_b:
            winner_rec = rec_a if winner == "a" else rec_b
            loser_rec  = rec_b if winner == "a" else rec_a
            loser_source_key = "b" if winner == "a" else "a"

            # Apply per-field choices: if user picked the loser's source for a field,
            # copy that value onto the winner record
            for key, val in form.items():
                if not key.startswith("field_"):
                    continue
                field_name = key[6:]
                if field_name not in _ALLOWED_OVERRIDE_FIELDS:
                    continue
                if str(val) == loser_source_key:
                    setattr(winner_rec, field_name, getattr(loser_rec, field_name, None))

            loser_rec.canonical_id = winner_rec.id
            loser_rec.is_new = False
            loser_rec.archived = True

    candidate.status = DedupStatus.merged if winner == "a" else DedupStatus.swapped
    candidate.resolved_by_user_id = user.id if user else None
    candidate.resolved_at = datetime.now(UTC)
    await session.flush()

    label = "merged into primary" if winner == "a" else "merged (B preferred)"
    return HTMLResponse(
        f'<tr id="dedup-row-{candidate_id}" class="text-muted" style="opacity:.4">'
        f'<td colspan="6" style="padding:10px 12px;font-size:12px">✓ Records {label}</td>'
        f'</tr>'
    )


@router.post("/ui/listings/{listing_id}/realie-skip", response_class=HTMLResponse)
async def ui_toggle_realie_skip(
    listing_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Toggle realie_skip on a listing. Returns updated toggle button HTML."""
    listing = await session.get(ScrapedListing, listing_id)
    if listing is None:
        return HTMLResponse("")
    listing.realie_skip = not listing.realie_skip
    await session.flush()
    label = "Enable Realie" if listing.realie_skip else "Skip Realie"
    style = "color:var(--warning,#f59e0b)" if listing.realie_skip else ""
    return HTMLResponse(
        f'<button id="skip-btn-{listing_id}" style="{style}"'
        f' hx-post="/ui/listings/{listing_id}/realie-skip"'
        f' hx-swap="outerHTML" hx-target="#skip-btn-{listing_id}"'
        f' class="btn btn-sm btn-secondary">{label}</button>'
    )


# ---------------------------------------------------------------------------
# Draw Schedule module
# ---------------------------------------------------------------------------

def _milestone_label(key: str) -> str:
    labels = {
        "offer_made": "Offer Made",
        "under_contract": "Under Contract",
        "close": "Close",
        "pre_development": "Pre-Development",
        "construction": "Construction",
        "operation_lease_up": "Lease-Up",
        "operation_stabilized": "Stabilized",
        "divestment": "Divestment",
    }
    return labels.get(key, key.replace("_", " ").title())


async def _run_draw_schedule(
    session: AsyncSession,
    model_id: UUID,
    *,
    writeback: bool = False,
) -> "Any | None":
    """Run the draw schedule engine; optionally write computed amounts back to DB.

    All sources are auto-sized (total_commitment=None) so the engine determines
    each source's commitment from Uses + carry. Returns the DrawSchedule, or None
    if the engine cannot run (missing milestones / sources).
    """
    from app.engines.draw_schedule import (
        DealMilestone,
        DrawScheduleCalculator,
        DrawScheduleConfig,
        DrawScheduleInputs,
        SourceDef,
        UseLineItem,
    )
    from datetime import datetime as _dt_cls

    ctx = await _load_draw_schedule_ctx(session, model_id)
    if not ctx:
        return None

    milestones_dated = ctx["milestones_dated"]
    draw_sources_db   = ctx["draw_sources"]
    use_lines_db      = ctx["use_lines_db"]

    if not milestones_dated or not draw_sources_db:
        return None

    # ── Milestones ──────────────────────────────────────────────────────────
    engine_milestones = [
        DealMilestone(
            key=m["key"],
            label=m["label"],
            date=_dt_cls.combine(m["date"], _dt_cls.min.time()),
        )
        for m in milestones_dated
    ]

    # ── Use lines ───────────────────────────────────────────────────────────
    _phase_to_ms = {
        "acquisition": "close", "pre_construction": "pre_development",
        "construction": "construction", "renovation": "construction",
        "conversion": "construction", "operation": "operation_stabilized",
        "exit": "divestment", "other": "close",
    }
    _phase_to_cat = {
        "acquisition": "land", "pre_construction": "soft_costs",
        "construction": "hard_costs", "renovation": "hard_costs",
        "conversion": "hard_costs", "operation": "reserves",
        "exit": "fees", "other": "other",
    }
    _ms_keys_set = {m["key"] for m in milestones_dated}
    _ms_date_idx  = {m["key"]: m["date"] for m in milestones_dated}
    engine_uses: list[UseLineItem] = []
    for ul in use_lines_db:
        raw_phase = str(ul.phase or "").replace("UseLinePhase.", "")
        ms_key = _phase_to_ms.get(raw_phase, "close")
        if ms_key not in _ms_keys_set and _ms_keys_set:
            ms_key = next(iter(_ms_keys_set))
        raw_timing   = str(ul.timing_type).replace("UseLineTiming.", "")
        spread_months = 1
        spread_to_date = None
        if raw_timing in ("spread", "spread_across_range"):
            for i, m in enumerate(milestones_dated):
                if m["key"] == ms_key and i + 1 < len(milestones_dated):
                    nxt = milestones_dated[i + 1]["date"]
                    cur = m["date"]
                    diff_months = (nxt.year - cur.year) * 12 + (nxt.month - cur.month)
                    spread_months = max(1, diff_months)
                    break
        engine_uses.append(UseLineItem(
            key=str(ul.id), label=ul.label,
            category=_phase_to_cat.get(raw_phase, "other"),
            total_amount=Decimal(str(ul.amount)),
            milestone_key=ms_key, spread_months=spread_months, spread_to_date=spread_to_date,
        ))

    # ── Sources — always auto-size (total_commitment=None) ─────────────────
    _last_real_ms  = milestones_dated[-1]["key"] if milestones_dated else "operation_stabilized"
    _real_ms_keys  = {m["key"] for m in milestones_dated}
    engine_sources: list[SourceDef] = []
    for ds in draw_sources_db:
        _to  = ds.active_to_milestone   if ds.active_to_milestone   in _real_ms_keys else _last_real_ms
        _frm = ds.active_from_milestone if ds.active_from_milestone in _real_ms_keys else (
            milestones_dated[0]["key"] if milestones_dated else _to
        )
        engine_sources.append(SourceDef(
            key=str(ds.id), label=ds.label,
            source_type=ds.source_type,
            draw_every_n_months=ds.draw_every_n_months,
            annual_interest_rate=Decimal(str(ds.annual_interest_rate)),
            active_from_milestone=_frm, active_to_milestone=_to,
            active_from_offset_days=getattr(ds, "active_from_offset_days", 0) or 0,
            active_to_offset_days=getattr(ds, "active_to_offset_days", 0) or 0,
            total_commitment=None,  # auto-size always
            # Non-exit-vehicle sources fund as a single lump-sum draw; the
            # engine ignores the active_to milestone for these.
            single_draw=(ds.source_type != "debt"),
        ))
    engine_sources.sort(key=lambda s: _ms_date_idx.get(s.active_from_milestone, _dt_cls.max))

    if not engine_sources:
        return None

    config = DrawScheduleConfig(
        min_reserve_construction=ctx["reserve_construction"],
        min_reserve_operational=ctx["reserve_operational"],
        operational_start_milestone="operation_lease_up",
    )
    try:
        schedule = DrawScheduleCalculator(DrawScheduleInputs(
            milestones=engine_milestones, uses=engine_uses,
            sources=engine_sources, config=config,
        )).calculate()
    except Exception:
        return None

    if writeback:
        _drawn_by_key = {ss.source_key: ss.total_drawn for ss in schedule.source_summaries}
        for ds in draw_sources_db:
            _drawn = _drawn_by_key.get(str(ds.id))
            if _drawn is None:
                continue
            ds.total_commitment = Decimal(str(_drawn))
            # Equity draw sources: update draw_sources.total_commitment for display
            # only — never overwrite capital module source["amount"] for equity
            # because the sequential payoff model produces inflated figures.
            if ds.source_type == "equity":
                continue
            if ds.capital_module_id:
                _cm = await session.get(CapitalModule, ds.capital_module_id)
                if _cm:
                    _src = dict(_cm.source or {})
                    _src["amount"] = float(_drawn)
                    _cm.source = _src
            else:
                _cm_q = select(CapitalModule).where(
                    CapitalModule.scenario_id == model_id,
                    CapitalModule.label == ds.label,
                ).limit(1)
                _cm = (await session.execute(_cm_q)).scalar_one_or_none()
                if _cm:
                    _src = dict(_cm.source or {})
                    _src["amount"] = float(_drawn)
                    _cm.source = _src
        await session.flush()

    return schedule


async def _load_draw_schedule_ctx(
    session: AsyncSession,
    model_id: UUID,
) -> dict[str, Any]:
    """Shared context for draw schedule panel and calculate endpoint."""
    from app.models.project import Project
    from app.models.milestone import Milestone

    model = await session.get(DealModel, model_id)
    if model is None:
        return {}

    # Load draw sources ordered by sort_order
    draw_sources = list((await session.execute(
        select(DrawSource)
        .where(DrawSource.scenario_id == model_id)
        .order_by(DrawSource.sort_order)
    )).scalars())

    # Load use lines (via Project) so we can pass them to the engine
    first_proj = (await session.execute(
        select(Project).where(Project.scenario_id == model_id).limit(1)
    )).scalar_one_or_none()

    use_lines_db: list = []
    project_milestones: list = []
    if first_proj:
        use_lines_db = list((await session.execute(
            select(UseLine).where(UseLine.project_id == first_proj.id)
        )).scalars())
        # Load milestones from both opportunity and project
        opp_ms = list((await session.execute(
            select(Milestone)
            .where(Milestone.opportunity_id == first_proj.opportunity_id)
            .order_by(Milestone.sequence_order)
        )).scalars()) if first_proj.opportunity_id else []
        proj_ms = list((await session.execute(
            select(Milestone)
            .where(Milestone.project_id == first_proj.id)
            .order_by(Milestone.sequence_order)
        )).scalars())
        project_milestones = opp_ms + proj_ms

    # Build milestone map for date resolution
    ms_map = {m.id: m for m in project_milestones}
    milestones_dated = []
    for m in project_milestones:
        start = m.computed_start(ms_map)
        if start:
            milestones_dated.append({
                "key": m.milestone_type.value if hasattr(m.milestone_type, "value") else str(m.milestone_type),
                "label": m.label or _milestone_label(str(m.milestone_type.value if hasattr(m.milestone_type, "value") else m.milestone_type)),
                "date": start,
            })

    # Sort milestones by date (opp + proj may interleave in unusual order)
    milestones_dated.sort(key=lambda m: m["date"])
    milestone_keys = [m["key"] for m in milestones_dated]

    # ---------------------------------------------------------------------------
    # Auto-seed draw_sources from capital_modules when none exist yet
    # ---------------------------------------------------------------------------
    if not draw_sources:
        capital_modules = list((await session.execute(
            select(CapitalModule)
            .where(CapitalModule.scenario_id == model_id)
            .order_by(CapitalModule.stack_position)
        )).scalars())

        # Map capital module phase strings → milestone keys (best-effort)
        _phase_to_ms = {
            "offer_made": "offer_made",
            "under_contract": "under_contract",
            "acquisition": "close",
            "pre_construction": "pre_development",
            "pre_development": "pre_development",
            "construction": "construction",
            "renovation": "construction",
            "lease_up": "operation_lease_up",
            "operation_lease_up": "operation_lease_up",
            "stabilized": "operation_stabilized",
            "operation_stabilized": "operation_stabilized",
            "divestment": "divestment",
        }
        for i, cm in enumerate(capital_modules):
            raw_from = cm.active_phase_start or "close"
            raw_to = cm.active_phase_end or "operation_stabilized"
            ms_from = _phase_to_ms.get(raw_from, raw_from)
            ms_to = _phase_to_ms.get(raw_to, raw_to)
            # Fall back to first/last milestone if mapped key not in timeline
            if milestone_keys:
                if ms_from not in milestone_keys:
                    ms_from = milestone_keys[0]
                if ms_to not in milestone_keys:
                    ms_to = milestone_keys[-1]

            src = cm.source or {}
            rate_pct = src.get("interest_rate_pct") or 0.0
            annual_rate = Decimal(str(rate_pct)) / Decimal("100")

            source_type = "debt" if str(getattr(cm, "vehicle_type", "") or "").replace("VehicleType.", "") == "debt" else "equity"
            draw_freq = 2 if source_type == "debt" else 1

            ds = DrawSource(
                id=_uuid_mod.uuid4(),
                scenario_id=model_id,
                capital_module_id=cm.id,
                sort_order=i + 1,
                label=cm.label,
                source_type=source_type,
                draw_every_n_months=draw_freq,
                annual_interest_rate=annual_rate,
                active_from_milestone=ms_from,
                active_to_milestone=ms_to,
                total_commitment=Decimal(str(src["amount"])) if src.get("amount") else None,
            )
            session.add(ds)

        if capital_modules:
            await session.flush()
            draw_sources = list((await session.execute(
                select(DrawSource)
                .where(DrawSource.scenario_id == model_id)
                .order_by(DrawSource.sort_order)
            )).scalars())

    # ---------------------------------------------------------------------------
    # Auto-populate reserve floors from computed use lines when still unset
    # ---------------------------------------------------------------------------
    reserve_construction = Decimal(str(model.min_reserve_construction or 0))
    reserve_operational = Decimal(str(model.min_reserve_operational or 0))

    if (reserve_construction == 0 or reserve_operational == 0) and use_lines_db:
        for ul in use_lines_db:
            lbl = (ul.label or "").strip()
            amt = Decimal(str(ul.amount or 0))
            if reserve_construction == 0 and lbl == "Capitalized Construction Interest":
                reserve_construction = amt
            elif reserve_operational == 0 and lbl == "Operating Reserve":
                reserve_operational = amt

    # ---------------------------------------------------------------------------
    # Build source Gantt rows using the same g2- coordinate system as the
    # timeline Gantt.  builder_gantt_data has epoch/g_min/g_max exposed.
    # ---------------------------------------------------------------------------
    import datetime as _dt
    builder_gantt_data_ds = _builder_gantt_from_milestones(first_proj, project_milestones)
    source_gantt_rows: list[dict] = []
    if builder_gantt_data_ds and draw_sources:
        epoch_d = builder_gantt_data_ds.get("epoch")
        g_min_d = builder_gantt_data_ds.get("g_min", 0)
        g_max_d = builder_gantt_data_ds.get("g_max", 1)
        total_span = max(g_max_d - g_min_d, 1)

        def _day_pct(day_offset: int) -> float:
            return round(100.0 * (day_offset - g_min_d) / total_span, 2)

        ms_date_map = {m["key"]: m["date"] for m in milestones_dated}
        for ds in draw_sources:
            from_date = ms_date_map.get(ds.active_from_milestone)
            to_date = ms_date_map.get(ds.active_to_milestone)
            fade_right = False
            if from_date and epoch_d:
                from_day = (from_date - epoch_d).days
                left = max(0.0, _day_pct(from_day))
                if ds.active_to_milestone not in ms_date_map:
                    # pseudo-milestone (e.g. "maturity"): extend to Gantt right edge, fade out
                    right = 100.0
                    fade_right = True
                elif to_date:
                    to_day = (to_date - epoch_d).days
                    right = min(100.0, _day_pct(to_day))
                else:
                    continue
                source_gantt_rows.append({
                    "label": ds.label,
                    "source_type": ds.source_type,
                    "left_pct": left,
                    "width_pct": max(right - left, 1.5),
                    "fade_right": fade_right,
                })

    # Label map used by the panel to display current active window as text (not editable here)
    milestone_label_map = {m["key"]: m["label"] for m in milestones_dated}
    milestone_label_map["maturity"] = "Maturity"

    # Phase E draw events — populated by compute_cash_flows(); empty until then.
    from app.models.capital_draw_event import CapitalDrawEvent as _CDE_ds
    from app.models.project import Project as _Proj_ds
    _cde_rows = list((await session.execute(
        select(_CDE_ds)
        .where(_CDE_ds.scenario_id == model_id)
        .order_by(_CDE_ds.period, _CDE_ds.allocation_reason)
    )).scalars())
    # Collect project names for multi-project labelling
    _proj_ids = {str(r.project_id) for r in _cde_rows if r.project_id}
    _proj_name_map: dict[str, str] = {}
    if _proj_ids:
        _projs_ds = list((await session.execute(
            select(_Proj_ds).where(_Proj_ds.id.in_([_uuid_mod.UUID(p) for p in _proj_ids]))
        )).scalars())
        _proj_name_map = {str(p.id): p.name for p in _projs_ds}
    capital_draw_events = [
        {
            "period": r.period,
            "period_type": r.period_type or "",
            "allocation_reason": (r.allocation_reason.value if hasattr(r.allocation_reason, "value") else str(r.allocation_reason or "")),
            "amount": float(r.amount or 0),
            "project_id": str(r.project_id) if r.project_id else None,
            "project_name": _proj_name_map.get(str(r.project_id), "—") if r.project_id else "—",
            "use_line_label": r.use_line_label or "",
        }
        for r in _cde_rows
    ]

    return {
        "model": model,
        "draw_sources": draw_sources,
        "milestones_dated": milestones_dated,
        "milestone_label_map": milestone_label_map,
        "use_lines_db": use_lines_db,
        "reserve_construction": reserve_construction,
        "reserve_operational": reserve_operational,
        "milestone_keys": milestone_keys,
        "builder_gantt_data": builder_gantt_data_ds,
        "source_gantt_rows": source_gantt_rows,
        "capital_draw_events": capital_draw_events,
    }


@router.get("/ui/models/{model_id}/draw-schedule", response_class=HTMLResponse)
async def draw_schedule_panel(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Returns the draw schedule panel partial for HTMX swap."""
    ctx = await _load_draw_schedule_ctx(session, model_id)
    if not ctx:
        return HTMLResponse("<p class='text-muted'>Model not found.</p>", status_code=404)
    ctx["request"] = request
    ctx["active_module"] = "draw_schedule"
    return templates.TemplateResponse(request, "partials/draw_schedule_panel.html", ctx)


@router.post("/ui/models/{model_id}/draw-schedule/sources", response_class=HTMLResponse)
async def add_draw_source(
    request: Request,
    model_id: UUID,
    session: DBSession,
    label: str = Form(...),
    source_type: str = Form("equity"),
    draw_every_n_months: int = Form(1),
    annual_interest_rate: str = Form("0"),
    active_from_milestone: str = Form(...),
    active_to_milestone: str = Form(...),
    total_commitment: str = Form(""),
) -> HTMLResponse:
    """Add a draw source row."""
    model = await session.get(DealModel, model_id)
    if model is None:
        return HTMLResponse("Model not found", status_code=404)

    # Determine next sort_order
    max_order_row = (await session.execute(
        select(DrawSource.sort_order)
        .where(DrawSource.scenario_id == model_id)
        .order_by(DrawSource.sort_order.desc())
        .limit(1)
    )).scalar_one_or_none()
    next_order = (max_order_row or 0) + 1

    commitment = None
    if total_commitment.strip():
        try:
            commitment = Decimal(total_commitment.strip().replace(",", ""))
        except Exception:
            commitment = None

    ds = DrawSource(
        id=_uuid_mod.uuid4(),
        scenario_id=model_id,
        sort_order=next_order,
        label=label.strip(),
        source_type=source_type,
        draw_every_n_months=max(1, draw_every_n_months),
        annual_interest_rate=Decimal(annual_interest_rate.strip() or "0"),
        active_from_milestone=active_from_milestone,
        active_to_milestone=active_to_milestone,
        total_commitment=commitment,
    )
    session.add(ds)
    await session.flush()

    ctx = await _load_draw_schedule_ctx(session, model_id)
    ctx["request"] = request
    ctx["active_module"] = "draw_schedule"
    return templates.TemplateResponse(request, "partials/draw_schedule_panel.html", ctx)


@router.delete("/ui/models/{model_id}/draw-schedule/sources/{source_id}", response_class=HTMLResponse)
async def delete_draw_source(
    request: Request,
    model_id: UUID,
    source_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Delete a draw source row."""
    ds = await session.get(DrawSource, source_id)
    if ds and ds.scenario_id == model_id:
        await session.delete(ds)
        await session.flush()
    ctx = await _load_draw_schedule_ctx(session, model_id)
    ctx["request"] = request
    ctx["active_module"] = "draw_schedule"
    return templates.TemplateResponse(request, "partials/draw_schedule_panel.html", ctx)


@router.post("/ui/models/{model_id}/draw-schedule/settings", response_class=HTMLResponse)
async def update_draw_schedule_settings(
    request: Request,
    model_id: UUID,
    session: DBSession,
    min_reserve_construction: str = Form("0"),
    min_reserve_operational: str = Form("0"),
) -> HTMLResponse:
    """Update reserve floor settings on the scenario."""
    model = await session.get(DealModel, model_id)
    if model is None:
        return HTMLResponse("Model not found", status_code=404)

    def _parse_dec(val: str) -> Decimal:
        try:
            return Decimal(val.strip().replace(",", "") or "0")
        except Exception:
            return Decimal("0")

    model.min_reserve_construction = _parse_dec(min_reserve_construction)
    model.min_reserve_operational = _parse_dec(min_reserve_operational)
    await session.flush()

    ctx = await _load_draw_schedule_ctx(session, model_id)
    ctx["request"] = request
    ctx["active_module"] = "draw_schedule"
    return templates.TemplateResponse(request, "partials/draw_schedule_panel.html", ctx)


@router.post("/ui/models/{model_id}/draw-schedule/calculate", response_class=HTMLResponse)
async def calculate_draw_schedule(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Run the draw schedule engine and return the results HTML fragment."""
    ctx = await _load_draw_schedule_ctx(session, model_id)
    if not ctx:
        return HTMLResponse("Model not found", status_code=404)

    milestones_dated = ctx["milestones_dated"]
    draw_sources_db  = ctx["draw_sources"]

    if not milestones_dated:
        return HTMLResponse(
            "<div class='module-empty'><div class='module-empty-icon'>📅</div>"
            "<div class='module-empty-title'>No timeline yet</div>"
            "<div class='module-empty-desc'>Set up milestones in the Timeline module first.</div></div>"
        )
    if not draw_sources_db:
        return HTMLResponse(
            "<div class='module-empty'><div class='module-empty-icon'>💰</div>"
            "<div class='module-empty-title'>No sources defined</div>"
            "<div class='module-empty-desc'>Add at least one funding source above.</div></div>"
        )

    schedule = await _run_draw_schedule(session, model_id, writeback=True)
    if schedule is None:
        return HTMLResponse(
            "<div class='alert alert-danger' style='padding:12px;border-radius:6px;"
            "background:#fef2f2;border:1px solid #fca5a5;color:#dc2626;font-size:13px'>"
            "⚠ Engine error: check milestones and sources are configured.</div>"
        )

    # ── Detect unfunded uses ─────────────────────────────────────────────────
    from app.engines.draw_schedule import UseLineItem
    from datetime import datetime
    _ms_date_index   = {m["key"]: m["date"] for m in milestones_dated}
    _milestone_label_map = {m["key"]: m["label"] for m in milestones_dated}
    _milestone_label_map["maturity"] = "Maturity"
    _covered_ms_keys: set[str] = set()
    for ss in schedule.source_summaries:
        # Find the source's active window from ctx draw_sources_db
        _ds = next((d for d in draw_sources_db if str(d.id) == ss.source_key), None)
        if _ds:
            from_idx = next((i for i, m in enumerate(milestones_dated) if m["key"] == _ds.active_from_milestone), None)
            to_idx   = next((i for i, m in enumerate(milestones_dated) if m["key"] == _ds.active_to_milestone), None)
            if from_idx is not None and to_idx is not None:
                for i in range(from_idx, to_idx + 1):
                    _covered_ms_keys.add(milestones_dated[i]["key"])
    # Build use items for unfunded check
    _phase_to_ms = {
        "acquisition": "close", "pre_construction": "pre_development",
        "construction": "construction", "renovation": "construction",
        "conversion": "construction", "operation": "operation_stabilized",
        "exit": "divestment", "other": "close",
    }
    _ms_keys_set = {m["key"] for m in milestones_dated}
    unfunded_uses: list[dict] = []
    for ul in ctx["use_lines_db"]:
        raw_phase = str(ul.phase or "").replace("UseLinePhase.", "")
        ms_key = _phase_to_ms.get(raw_phase, "close")
        if ms_key not in _ms_keys_set and _ms_keys_set:
            ms_key = next(iter(_ms_keys_set))
        if ms_key not in _covered_ms_keys and (ul.amount or 0) > 0:
            unfunded_uses.append({
                "label": ul.label,
                "amount": ul.amount,
                "milestone_key": ms_key,
                "milestone_label": _milestone_label_map.get(ms_key, ms_key),
            })

    # Filter display: hide sources with no draws and $0-commitment sources from Gantt/table
    active_labels = {ss.source_label for ss in schedule.source_summaries if ss.total_drawn > 0}
    committed_labels = {
        ds.label for ds in draw_sources_db
        if ds.total_commitment and float(ds.total_commitment) > 0
    }
    show_labels = active_labels & committed_labels
    ctx["source_gantt_rows"] = [r for r in ctx.get("source_gantt_rows", []) if r["label"] in show_labels]
    ctx["draw_sources"] = [ds for ds in ctx["draw_sources"] if ds.label in show_labels]

    ctx["schedule"] = schedule
    ctx["unfunded_uses"] = unfunded_uses
    ctx["request"] = request
    ctx["active_module"] = "draw_schedule"
    # Return the full panel (not just results) so sources table always reflects current DB state
    return templates.TemplateResponse(request, "partials/draw_schedule_panel.html", ctx)



@router.patch("/ui/models/{model_id}/draw-schedule/sources/{source_id}", response_class=HTMLResponse)
async def update_draw_source_window(
    request: Request,
    model_id: UUID,
    source_id: UUID,
    session: DBSession,
    active_from_milestone: str = Form(...),
    active_to_milestone: str = Form(...),
) -> HTMLResponse:
    """Update the active window (from/to milestone) of a draw source."""
    ds = await session.get(DrawSource, source_id)
    if ds and ds.scenario_id == model_id:
        ds.active_from_milestone = active_from_milestone
        ds.active_to_milestone = active_to_milestone
        await session.flush()
    ctx = await _load_draw_schedule_ctx(session, model_id)
    ctx["request"] = request
    ctx["active_module"] = "draw_schedule"
    return templates.TemplateResponse(request, "partials/draw_schedule_panel.html", ctx)


# ── Saved Filters ────────────────────────────────────────────────────────────
# Per-user, per-page named filter snapshots used by the Listings/Parcels/
# Opportunities/Deals filter bars. Stored as the URL query string the page
# already speaks, so loading a saved filter == redirect to /<page>?<query>
# and the URL itself is shareable.

_SAVED_FILTER_PAGES = {"listings", "parcels", "opportunities", "deals"}


def _saved_filter_landing(page: str) -> str:
    """Map a filter-form page key back to the URL the saved filter loads against."""
    return {
        "listings": "/listings",
        "parcels": "/parcels",
        "opportunities": "/opportunities",
        "deals": "/deals",
    }.get(page, "/")


@router.get("/api/saved-filters")
async def list_saved_filters(
    request: Request,
    session: DBSession,
    page: str = Query(...),
) -> dict:
    """List the current user's saved filters for one page."""
    from app.models.saved_filter import SavedFilter
    user = await _get_user(session, request)
    if user is None or page not in _SAVED_FILTER_PAGES:
        return {"items": []}
    rows = list((await session.execute(
        select(SavedFilter)
        .where(SavedFilter.user_id == user.id, SavedFilter.page == page)
        .order_by(SavedFilter.name)
    )).scalars())
    base = _saved_filter_landing(page)
    return {
        "items": [
            {
                "id": str(r.id),
                "name": r.name,
                "url": f"{base}?{r.query_string}" if r.query_string else base,
                "query_string": r.query_string,
            }
            for r in rows
        ]
    }


@router.post("/api/saved-filters")
async def create_saved_filter(
    request: Request,
    session: DBSession,
) -> dict:
    """Create or rename-overwrite a saved filter for the current user."""
    from app.models.saved_filter import SavedFilter
    user = await _get_user(session, request)
    if user is None:
        return JSONResponse({"detail": "Not authenticated"}, status_code=401)
    form = await request.form()
    page = str(form.get("page", "")).strip()
    name = str(form.get("name", "")).strip()[:120]
    query_string = str(form.get("query_string", "")).strip()
    if page not in _SAVED_FILTER_PAGES or not name:
        return JSONResponse({"detail": "Missing page or name"}, status_code=400)

    existing = (await session.execute(
        select(SavedFilter).where(
            SavedFilter.user_id == user.id,
            SavedFilter.page == page,
            SavedFilter.name == name,
        )
    )).scalar_one_or_none()
    if existing:
        existing.query_string = query_string
        existing.updated_at = datetime.now(UTC)
        row = existing
    else:
        row = SavedFilter(
            user_id=user.id,
            page=page,
            name=name,
            query_string=query_string,
        )
        session.add(row)
    await session.commit()
    base = _saved_filter_landing(page)
    return {
        "id": str(row.id),
        "name": row.name,
        "url": f"{base}?{row.query_string}" if row.query_string else base,
    }


@router.delete("/api/saved-filters/{filter_id}")
async def delete_saved_filter(
    request: Request,
    filter_id: UUID,
    session: DBSession,
) -> dict:
    from app.models.saved_filter import SavedFilter
    user = await _get_user(session, request)
    if user is None:
        return JSONResponse({"detail": "Not authenticated"}, status_code=401)
    row = await session.get(SavedFilter, filter_id)
    if row is None or row.user_id != user.id:
        return JSONResponse({"detail": "Not found"}, status_code=404)
    await session.delete(row)
    await session.commit()
    return {"ok": True}


# ── Deal change-log (history drawer) ─────────────────────────────────────────

@router.get("/ui/models/{model_id}/history", response_class=HTMLResponse)
async def history_drawer(
    model_id: UUID, request: Request, session: DBSession
) -> HTMLResponse:
    """Render the history drawer partial (list of compute snapshots)."""
    from app.exporters.snapshot import diff_snapshots, list_snapshots

    user = await _get_user(session, request)
    if user is None:
        return HTMLResponse('<p style="color:var(--text-muted)">Not authenticated</p>', status_code=401)

    model = await session.get(DealModel, model_id, options=[selectinload(DealModel.deal)])
    if model is None:
        return HTMLResponse('<p style="color:var(--text-muted)">Model not found</p>', status_code=404)
    if model.deal is None or model.deal.org_id != user.org_id:
        return HTMLResponse('<p style="color:var(--text-muted)">Forbidden</p>', status_code=403)

    snaps = await list_snapshots(session, model_id)

    # Build diff summaries between consecutive snapshots for rendering
    entries = []
    for i, snap in enumerate(snaps):
        diff = diff_snapshots(snaps[i - 1], snap) if i > 0 else None
        entries.append({"snap": snap, "diff": diff})

    return templates.TemplateResponse(
        request,
        "partials/history_drawer.html",
        {"request": request, "model": model, "entries": entries, "model_id": model_id},
    )


@router.post("/ui/models/{model_id}/history/{snapshot_id}/revert", response_class=HTMLResponse)
async def revert_snapshot(
    model_id: UUID,
    snapshot_id: UUID,
    request: Request,
    session: DBSession,
) -> HTMLResponse:
    """Revert scenario inputs to a prior snapshot state and return a status banner."""
    from app.exporters.snapshot import revert_to_snapshot

    user = await _get_user(session, request)
    if user is None:
        return HTMLResponse('<p style="color:var(--color-error)">Not authenticated</p>', status_code=401)

    model = await session.get(DealModel, model_id, options=[selectinload(DealModel.deal)])
    if model is None:
        return HTMLResponse('<p style="color:var(--color-error)">Model not found</p>', status_code=404)
    if model.deal is None or model.deal.org_id != user.org_id:
        return HTMLResponse('<p style="color:var(--color-error)">Forbidden</p>', status_code=403)

    try:
        await revert_to_snapshot(session, model_id, snapshot_id)
        await session.commit()
    except ValueError as exc:
        return HTMLResponse(
            f'<div class="alert alert-error" role="alert">{exc}</div>',
            status_code=404,
        )

    from starlette.responses import RedirectResponse
    return RedirectResponse(url=f"/models/{model_id}/builder", status_code=303)


@router.get("/ui/models/{model_id}/history/export.json")
async def export_history_json_endpoint(
    model_id: UUID, request: Request, session: DBSession
) -> JSONResponse:
    """Return the full change-log as a structured JSON (AI-readable)."""
    from app.exporters.snapshot import export_history_json
    import json as _json

    user = await _get_user(session, request)
    if user is None:
        return JSONResponse({"error": "not authenticated"}, status_code=401)

    model = await session.get(DealModel, model_id, options=[selectinload(DealModel.deal)])
    if model is None:
        return JSONResponse({"error": "model not found"}, status_code=404)
    if model.deal is None or model.deal.org_id != user.org_id:
        return JSONResponse({"error": "forbidden"}, status_code=403)

    try:
        payload = await export_history_json(session, model_id)
    except ValueError as exc:
        return JSONResponse({"error": str(exc)}, status_code=404)

    return JSONResponse(
        content=_json.loads(_json.dumps(payload, default=str)),
        headers={
            "Content-Disposition": f'attachment; filename="history-{model_id}.json"',
        },
    )


