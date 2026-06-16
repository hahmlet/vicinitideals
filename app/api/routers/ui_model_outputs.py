"""Excel/investor exports, proforma import, NOI inputs, line form, source vehicles, draw schedule, history.

Extracted from ui.py as part of the Phase 2a sub-router split. Route handlers
for model output operations: file downloads, bulk imports, draw schedule,
change history. Imports from ui_helpers; never from ui.py at module level.
"""
from __future__ import annotations

import io
import json
import uuid as _uuid_mod
from decimal import Decimal
from typing import Any
from uuid import UUID

from fastapi import APIRouter, File, Form, HTTPException, Query, Request, Response, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from sqlalchemy import and_, delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.deps import DBSession
from app.config import settings
from app.models.capital import CapitalModule, DrawSource, WaterfallTier
from app.models.deal import (
    Deal,
    Scenario,
    IncomeStream,
    IncomeStreamType,
    OperatingExpenseLine,
    OperationalInputs,
    STANDARD_OPEX_CATEGORIES,
    USE_CATEGORY_LABELS,
    USE_CATEGORY_PRESETS,
    USE_COST_CATEGORIES,
    UseLine,
    UseLinePhase,
)
from app.models.milestone import Milestone, MilestoneType
from app.models.org import User
from app.models.project import Project
from app.engines.dev_fee import BASIS_BUCKETS
from app.api.routers.ui_helpers import (
    _UMRow,
    _active_project_from_request,
    _builder_gantt_from_milestones,
    _get_user,
    templates,
)

router = APIRouter(include_in_schema=False)

@router.get("/ui/models/{model_id}/export.xlsx")
async def download_model_export(
    model_id: UUID,
    session: DBSession,
) -> StreamingResponse:
    """Download a round-trip-capable Excel workbook for this deal model.

    Deprecated path; superseded by ``/investor-export.xlsx``. Kept available
    while the investor export bakes — see plan §10 in
    ``docs/feature-plans/investor-excel-export-v2.md``.
    """
    from app.exporters.excel_export import export_deal_model_workbook, make_export_filename
    model = await session.get(Scenario, model_id)
    if model is None:
        return HTMLResponse("Not found", status_code=404)
    workbook_bytes = await export_deal_model_workbook(model_id, session)
    filename = make_export_filename(model)
    return StreamingResponse(
        iter([workbook_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/ui/models/{model_id}/investor-export.xlsx")
async def download_investor_export(
    model_id: UUID,
    session: DBSession,
    profile: str = Query(default="internal"),
) -> StreamingResponse:
    """Download the LP-facing investor Excel workbook for this Scenario."""
    from app.exporters.investor_export import export_investor_workbook, make_investor_filename
    scenario = await session.get(Scenario, model_id)
    if scenario is None:
        return HTMLResponse("Not found", status_code=404)
    deal = await session.get(Deal, scenario.deal_id) if scenario.deal_id else None
    filename = make_investor_filename(scenario, deal)
    workbook_bytes = await export_investor_workbook(model_id, session, profile=profile)
    return StreamingResponse(
        iter([workbook_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Async investor export ────────────────────────────────────────────────────
#
# The synchronous endpoint above blows past the NGINX 60s proxy timeout once
# the live Sensitivity matrix lands (25 cashflow cycles per export). The async
# path below kicks a Celery task that builds the workbook off the request
# path and emails the .xlsx as an attachment when finished. Job state
# (``queued → calculating → sending → sent`` or ``failed``) is persisted on
# ``ExportJob`` rows so the UI can poll for hover-modal updates and the user
# can resend a cached build when the scenario hasn't been recomputed since
# the last successful export.


@router.get("/ui/models/{model_id}/investor-export/preflight")
async def preflight_investor_export(
    model_id: UUID,
    session: DBSession,
    request: Request,
    profile: str = Query(default="internal"),
) -> JSONResponse:
    """Cheap idempotent check: is a cached resend eligible for this scenario?

    Returns ``{"resend_eligible": bool, "resend_job_id": <uuid|null>}``.
    UI uses this to decide whether to prompt the user with a "Resend last
    export?" modal before enqueueing a fresh build.

    Eligibility = last ``sent`` job's ``created_at`` > every
    ``OperationalOutputs.computed_at`` for this scenario AND that job
    still has ``xlsx_bytes`` cached.
    """
    from app.models.cashflow import OperationalOutputs
    from app.models.export_job import ExportJob, ExportJobStatus

    user = await _get_user(session, request)
    if user is None:
        return JSONResponse({"error": "not authenticated"}, status_code=401)
    scenario = await session.get(Scenario, model_id)
    if scenario is None:
        return JSONResponse({"error": "scenario not found"}, status_code=404)

    latest_outputs_computed_at = (
        await session.execute(
            select(func.max(OperationalOutputs.computed_at))
            .where(OperationalOutputs.scenario_id == model_id)
        )
    ).scalar_one_or_none()

    last_sent_job = (
        await session.execute(
            select(ExportJob)
            .where(ExportJob.scenario_id == model_id)
            .where(ExportJob.status == ExportJobStatus.sent)
            .where(ExportJob.xlsx_bytes.isnot(None))
            .where(ExportJob.export_profile == profile)
            .order_by(ExportJob.created_at.desc())
            .limit(1)
        )
    ).scalar_one_or_none()

    resend_eligible = bool(
        last_sent_job is not None
        and (
            latest_outputs_computed_at is None
            or last_sent_job.created_at > latest_outputs_computed_at
        )
    )
    return JSONResponse(
        {
            "resend_eligible": resend_eligible,
            "resend_job_id": (
                str(last_sent_job.id) if (resend_eligible and last_sent_job) else None
            ),
        }
    )


@router.post("/ui/models/{model_id}/investor-export/async")
async def start_investor_export_async(
    model_id: UUID,
    session: DBSession,
    request: Request,
) -> JSONResponse:
    """Enqueue a fresh-build investor-export job and return its id.

    Caller (UI) is expected to have hit ``/preflight`` first to decide
    whether to prompt for "Resend last export?" — this endpoint always
    builds fresh.
    """
    from app.models.export_job import ExportJob, ExportJobStatus
    from app.tasks.export import RUN_EXPORT_TASK
    from app.tasks.celery_app import celery_app as _celery

    user = await _get_user(session, request)
    if user is None:
        return JSONResponse({"error": "not authenticated"}, status_code=401)
    scenario = await session.get(Scenario, model_id)
    if scenario is None:
        return JSONResponse({"error": "scenario not found"}, status_code=404)

    try:
        body = await request.json()
    except Exception:  # noqa: BLE001
        body = {}
    _raw_profile = body.get("profile", "internal") if isinstance(body, dict) else "internal"
    _profile = _raw_profile if _raw_profile in {"internal", "lp", "lender", "proforma"} else "internal"

    job = ExportJob(
        scenario_id=model_id,
        user_id=user.id,
        recipient_email=user.email or "",
        status=ExportJobStatus.queued,
        export_profile=_profile,
    )
    session.add(job)
    await session.commit()
    await session.refresh(job)

    _celery.send_task(RUN_EXPORT_TASK, args=[str(job.id)])

    return JSONResponse(
        {
            "job_id": str(job.id),
            "status": job.status.value,
        }
    )


@router.get("/ui/exports/{job_id}/status")
async def get_export_job_status(
    job_id: UUID,
    session: DBSession,
    request: Request,
) -> JSONResponse:
    """Return current status of an export job for poll/hover-modal use."""
    from app.models.export_job import ExportJob

    user = await _get_user(session, request)
    if user is None:
        return JSONResponse({"error": "not authenticated"}, status_code=401)
    job = await session.get(ExportJob, job_id)
    if job is None or job.user_id != user.id:
        return JSONResponse({"error": "job not found"}, status_code=404)

    return JSONResponse(
        {
            "job_id": str(job.id),
            "status": job.status.value,
            "error_message": job.error_message,
            "created_at": job.created_at.isoformat() if job.created_at else None,
            "completed_at": job.completed_at.isoformat() if job.completed_at else None,
        }
    )


@router.post("/ui/exports/{job_id}/resend")
async def resend_investor_export_endpoint(
    job_id: UUID,
    session: DBSession,
    request: Request,
) -> JSONResponse:
    """Re-send a previously-completed export from cached xlsx_bytes.

    Spawns a fresh ``ExportJob`` row pointing at the same scenario; the
    resend task copies bytes from the source job before sending so the
    "last sent" lookup keeps walking forward in time.
    """
    from app.models.export_job import ExportJob, ExportJobStatus
    from app.tasks.export import RESEND_EXPORT_TASK
    from app.tasks.celery_app import celery_app as _celery

    user = await _get_user(session, request)
    if user is None:
        return JSONResponse({"error": "not authenticated"}, status_code=401)
    src = await session.get(ExportJob, job_id)
    if src is None or src.user_id != user.id:
        return JSONResponse({"error": "job not found"}, status_code=404)
    if not src.xlsx_bytes:
        return JSONResponse({"error": "no cached export bytes"}, status_code=409)

    new_job = ExportJob(
        scenario_id=src.scenario_id,
        user_id=user.id,
        recipient_email=user.email or src.recipient_email,
        status=ExportJobStatus.queued,
        xlsx_bytes=src.xlsx_bytes,
        filename=src.filename,
    )
    session.add(new_job)
    await session.commit()
    await session.refresh(new_job)

    _celery.send_task(RESEND_EXPORT_TASK, args=[str(new_job.id)])
    return JSONResponse(
        {
            "job_id": str(new_job.id),
            "status": new_job.status.value,
        }
    )


@router.get("/ui/models/{model_id}/import-template.xlsx")
async def download_import_template(model_id: UUID) -> StreamingResponse:
    """Download a pre-formatted Excel template for bulk import of Uses and OpEx line items."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()

    # ── Shared styles ──────────────────────────────────────────────────────────
    hdr_font = Font(bold=True, size=10, color="FFFFFF")
    hdr_fill_uses = PatternFill("solid", fgColor="2563EB")   # blue
    hdr_fill_opex = PatternFill("solid", fgColor="059669")   # green
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hint_font = Font(italic=True, size=9, color="6B7280")
    hint_fill = PatternFill("solid", fgColor="F9FAFB")

    def _set_col_width(ws, col_letter, width):
        ws.column_dimensions[col_letter].width = width

    def _header_row(ws, headers, fill):
        ws.append(headers)
        for i, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i)
            cell.font = hdr_font
            cell.fill = fill
            cell.alignment = hdr_align
        ws.row_dimensions[1].height = 28

    def _hint_row(ws, hints):
        ws.append(hints)
        for i, _ in enumerate(hints, 1):
            cell = ws.cell(row=2, column=i)
            cell.font = hint_font
            cell.fill = hint_fill
            cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[2].height = 36

    # ── Uses sheet ─────────────────────────────────────────────────────────────
    ws_uses = wb.active
    ws_uses.title = "Uses"
    _header_row(ws_uses, ["Label", "Phase", "Amount ($)", "Deferred Dev Fee?", "Notes"], hdr_fill_uses)
    _hint_row(ws_uses, [
        "e.g. Hard Costs, Soft Costs, Contingency",
        "acquisition | pre_development | construction | exit",
        "Dollar amount (no commas)",
        "yes / no — deferred developer fee?",
        "Optional notes",
    ])
    # Phase validation
    phase_dv = DataValidation(
        type="list",
        formula1='"acquisition,pre_development,construction,exit"',
        allow_blank=True,
    )
    ws_uses.add_data_validation(phase_dv)
    phase_dv.sqref = "B3:B500"
    # Deferred dv
    bool_dv = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
    ws_uses.add_data_validation(bool_dv)
    bool_dv.sqref = "D3:D500"
    # Widths
    for col, w in zip("ABCDE", [32, 22, 16, 18, 30]):
        _set_col_width(ws_uses, col, w)
    # 3 sample rows
    for label, phase, amt in [
        ("Hard Costs", "construction", 480000),
        ("Soft Costs", "construction", 72000),
        ("Contingency (10%)", "construction", 55200),
    ]:
        ws_uses.append([label, phase, amt, "no", ""])

    # ── OpEx sheet ─────────────────────────────────────────────────────────────
    ws_opex = wb.create_sheet("OpEx")
    _header_row(ws_opex, [
        "Label", "Amount", "Per", "Escalation (%/yr)",
        "Scale w/ Lease-Up?", "Lease-Up Floor (%)", "Active Phases", "Notes",
    ], hdr_fill_opex)
    _hint_row(ws_opex, [
        "e.g. Property Tax, Insurance",
        "Dollar value",
        "flat | per_unit | per_sqft_residential | per_sqft_commercial",
        "e.g. 3.0",
        "yes / no",
        "0–100 (% of stabilized when vacant)",
        "construction, lease_up, stabilized (comma-separated)",
        "Optional",
    ])
    per_dv = DataValidation(
        type="list",
        formula1='"flat,per_unit,per_sqft_residential,per_sqft_commercial"',
        allow_blank=True,
    )
    ws_opex.add_data_validation(per_dv)
    per_dv.sqref = "C3:C500"
    ws_opex.add_data_validation(bool_dv)
    bool_dv.sqref = "E3:E500"
    for col, w in zip("ABCDEFGH", [28, 14, 22, 16, 18, 18, 30, 24]):
        _set_col_width(ws_opex, col, w)
    # 3 sample rows
    for label, amt, per, esc, scale, floor, phases in [
        ("Property Tax", 18000, "flat", 3.0, "no", "", "stabilized"),
        ("Insurance", 9600, "flat", 3.0, "no", "", "stabilized"),
        ("Property Management", 8, "per_unit", 3.0, "yes", 25, "lease_up, stabilized"),
    ]:
        ws_opex.append([label, amt, per, esc, scale, floor, phases, ""])

    # ── Stream to response ─────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=import-template.xlsx"},
    )


# ---------------------------------------------------------------------------
# Pro forma import — preflight, upload, status, confirm, skip
# ---------------------------------------------------------------------------

def _render_proforma_sheet_picker(
    request: Request,
    model_id: UUID,
    task_id: str,
    content: bytes,
) -> HTMLResponse:
    """Read sheet names + first-row columns from xlsx bytes and return the
    sheet-picker fragment. Shared by the preflight dispatch and the
    reanalyze flow (which reuses bytes already in Redis)."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        sheet_names = wb.sheetnames

        sheet_columns: dict[str, list[str]] = {}
        for name in sheet_names:
            ws = wb[name]
            for row in ws.iter_rows(max_row=10, values_only=True):
                non_empty = [str(c).strip() for c in row if c is not None]
                if non_empty:
                    sheet_columns[name] = non_empty
                    break
            else:
                sheet_columns[name] = []
        wb.close()
    except Exception as exc:
        return HTMLResponse(f"<p class='text-red-500'>Could not read file: {exc}</p>", status_code=400)

    return templates.TemplateResponse(
        request,
        "partials/proforma_preflight.html",
        {
            "model_id": model_id,
            "task_id": task_id,
            "sheet_names": sheet_names,
            "sheet_columns": sheet_columns,
            "STANDARD_OPEX_CATEGORIES": STANDARD_OPEX_CATEGORIES,
        },
    )


async def _dispatch_proforma_preflight(
    *,
    request: Request,
    model_id: UUID,
    upload: UploadFile,
) -> HTMLResponse:
    """Stash the uploaded pro forma in redis, then return either the sheet
    picker (for .xlsx) or the parse-progress poller (for PDF/DOCX/HTML).

    Extracted helper so both the dedicated POST route and the Step 1 wizard
    handler can dispatch the same flow when a file rides along with the
    income-mode form.

    Computes SHA-256 of the file bytes. If a parse result already lives in
    Redis under ``proforma:filehash:{hash}:result`` (7-day TTL), returns the
    cache-hit fragment so the user can skip the LLM call.
    """
    import hashlib
    import os as _os

    content = await upload.read()
    if not content:
        return HTMLResponse("<p class='text-red-500'>Empty file uploaded.</p>", status_code=400)

    filename = upload.filename or ""
    ext = _os.path.splitext(filename)[1].lower().lstrip(".")
    file_kind = "xlsx" if ext in {"xlsx", "xlsm", "xlsb"} else "doc"
    file_hash = hashlib.sha256(content).hexdigest()

    task_id = str(_uuid_mod.uuid4())
    import redis as _redis  # type: ignore
    _RESUME_TTL = 7 * 86_400
    r = _redis.from_url(settings.redis_url, decode_responses=False)
    r.set(f"proforma:{task_id}:file", content, ex=86_400)
    r.set(f"proforma:{task_id}:filename", filename, ex=86_400)
    r.set(f"proforma:{task_id}:kind", file_kind, ex=86_400)
    r.set(f"proforma:{task_id}:file_hash", file_hash, ex=86_400)
    # Track the most-recent proforma hash + filename per scenario so the
    # wizard's Step 2 "Back" can resume on the review page instead of
    # forcing a re-upload. 7d TTL matches the cache TTL.
    r.set(f"scenario:{model_id}:last_proforma_hash", file_hash.encode(), ex=_RESUME_TTL)
    r.set(f"scenario:{model_id}:last_proforma_filename", filename.encode(), ex=_RESUME_TTL)

    # ── Content-hash cache check ───────────────────────────────────────────
    # Cache hit: skip the LLM call and render the review page directly with
    # cached data. A banner on the review page surfaces the cache origin and
    # re-analyze/purge actions.
    r_str = _redis.from_url(settings.redis_url, decode_responses=True)
    cached_raw = r_str.get(f"proforma:filehash:{file_hash}:result")
    if cached_raw:
        try:
            cached_result = json.loads(cached_raw)
        except Exception:
            cached_result = None
        if cached_result is not None:
            parsed_at = r_str.get(f"proforma:filehash:{file_hash}:parsed_at") or ""
            # Mirror cached result to task-keyed key so /proforma-confirm
            # (reads by task_id) finds it.
            r_str.set(f"proforma:{task_id}:result", cached_raw, ex=86_400)
            return templates.TemplateResponse(
                request,
                "partials/proforma_review.html",
                {
                    "model_id": model_id,
                    "task_id": task_id,
                    "unit_types": cached_result.get("unit_types", []),
                    "expense_lines": cached_result.get("expense_lines", []),
                    "warnings": cached_result.get("warnings", []),
                    "STANDARD_OPEX_CATEGORIES": STANDARD_OPEX_CATEGORIES,
                    "from_cache": True,
                    "file_hash": file_hash,
                    "filename": filename,
                    "parsed_at": parsed_at,
                },
            )

    if file_kind == "doc":
        if ext == "pdf":
            # PDF: show page-picker preflight so user can scope large OMs
            page_count: int | None = None
            try:
                import pdfplumber as _pdfplumber  # type: ignore
                with _pdfplumber.open(io.BytesIO(content)) as _pdf:
                    page_count = len(_pdf.pages)
            except Exception:
                pass
            return templates.TemplateResponse(
                request,
                "partials/proforma_preflight_doc.html",
                {
                    "model_id": model_id,
                    "task_id": task_id,
                    "page_count": page_count,
                    "filename": filename,
                },
            )
        # Non-PDF doc (DOCX, HTML, etc.): queue immediately, MarkitDown converts whole doc
        from app.tasks.proforma_parse import PARSE_PROFORMA_TASK
        from app.tasks.celery_app import celery_app as _celery
        _celery.send_task(
            PARSE_PROFORMA_TASK,
            kwargs={
                "task_id": task_id,
                "model_id": str(model_id),
                "revenue_sheet": "",
                "opex_sheet": "",
                "property_column": None,
                "file_kind": "doc",
            },
        )
        return templates.TemplateResponse(
            request,
            "partials/proforma_progress.html",
            {"model_id": model_id, "task_id": task_id},
        )

    return _render_proforma_sheet_picker(request, model_id, task_id, content)


@router.post("/ui/models/{model_id}/proforma-preflight", response_class=HTMLResponse)
async def proforma_preflight(
    request: Request,
    model_id: UUID,
    file: UploadFile = File(...),
) -> HTMLResponse:
    """Receive uploaded file at the dedicated endpoint. Thin wrapper over
    ``_dispatch_proforma_preflight`` — kept for direct re-upload (proforma-
    restart) and for any external callers that still POST a file directly."""
    return await _dispatch_proforma_preflight(
        request=request, model_id=model_id, upload=file,
    )


@router.get("/ui/models/{model_id}/proforma-from-staged", response_class=HTMLResponse)
async def proforma_from_staged(
    request: Request,
    model_id: UUID,
    session: DBSession,
    task_id: str = Query(...),
) -> HTMLResponse:
    """Load a pre-staged proforma from Redis (email attachment path) and run
    the same preflight flow as a manual upload — no file upload needed.

    Called by the deal setup wizard when the URL carries ``proforma_task_id``
    from an email attachment that was stashed in Redis during email ingest.
    """
    # Auth + ownership — same guard pattern as model_builder.
    user = await _get_user(session, request)
    if user is None:
        return HTMLResponse("Not authenticated", status_code=401)
    model = await session.get(Scenario, model_id)
    if model is None:
        return HTMLResponse("Not found", status_code=404)
    if settings.org_isolation_enabled:
        owning_deal = await session.get(Deal, model.deal_id) if model.deal_id else None
        if owning_deal is None or owning_deal.org_id != user.org_id:
            return HTMLResponse("Not found", status_code=404)

    import redis as _redis  # type: ignore

    r = _redis.from_url(settings.redis_url, decode_responses=False)

    # Cross-tenant guard: verify task_id was staged for this org.
    stored_org = r.get(f"proforma:{task_id}:org_id")
    if stored_org is not None and stored_org.decode() != str(user.org_id):
        return HTMLResponse("Not found", status_code=404)

    file_bytes = r.get(f"proforma:{task_id}:file")
    if not file_bytes:
        return HTMLResponse(
            "<div class='wizard-shell' id='deal-setup-wizard'>"
            "<div class='wizard-card'><p style='padding:20px;color:var(--text-muted)'>"
            "Pre-staged file expired or not found. Please upload the file manually.</p>"
            "</div></div>",
            status_code=200,
        )

    # Fast-path: when the user pre-configured sheets/pages in the email review
    # table, skip the picker and go straight to the Celery task.
    r_str = _redis.from_url(settings.redis_url, decode_responses=True)
    email_config_raw = r_str.get(f"proforma:{task_id}:email_config")
    if email_config_raw:
        cfg = json.loads(email_config_raw)
        rev_pages_str: str = cfg.get("rev_pages") or ""
        opex_pages_str: str = cfg.get("opex_pages") or ""

        def _parse_page_str(s: str) -> list[int] | None:
            if not s.strip():
                return None
            pages: list[int] = []
            for part in s.replace(" ", "").split(","):
                if "-" in part:
                    lo, _, hi = part.partition("-")
                    if lo.isdigit() and hi.isdigit():
                        pages.extend(range(int(lo), int(hi) + 1))
                elif part.isdigit():
                    pages.append(int(part))
            return [p - 1 for p in sorted(set(pages)) if p >= 1] or None

        from app.tasks.celery_app import celery_app as _celery
        from app.tasks.proforma_parse import PARSE_PROFORMA_TASK
        _celery.send_task(
            PARSE_PROFORMA_TASK,
            kwargs={
                "task_id": task_id,
                "model_id": str(model_id),
                "revenue_sheet": cfg.get("rev_sheet") or "",
                "opex_sheet": cfg.get("opex_sheet") or "",
                "property_column": None,
                "file_kind": cfg.get("file_kind") or "xlsx",
                "import_revenue": cfg.get("import_revenue", True),
                "import_opex": cfg.get("import_opex", True),
                "revenue_range": cfg.get("rev_range") or None,
                "opex_range": cfg.get("opex_range") or None,
                "revenue_pages": _parse_page_str(rev_pages_str),
                "opex_pages": _parse_page_str(opex_pages_str),
            },
        )
        return templates.TemplateResponse(
            request,
            "partials/proforma_progress.html",
            {"model_id": model_id, "task_id": task_id},
        )

    filename_bytes = r.get(f"proforma:{task_id}:filename")
    filename = filename_bytes.decode() if filename_bytes else "attachment.xlsx"

    class _FakeUpload:
        async def read(self) -> bytes:
            return file_bytes  # type: ignore[return-value]

    _FakeUpload.filename = filename  # type: ignore[attr-defined]

    return await _dispatch_proforma_preflight(
        request=request, model_id=model_id, upload=_FakeUpload(),  # type: ignore[arg-type]
    )


@router.get("/ui/models/{model_id}/proforma-restart", response_class=HTMLResponse)
async def proforma_restart(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Return the wizard at Step 1 so the user can upload a different file
    (or pick a different income mode). Re-uses the GET /setup handler so the
    full context (inputs, vehicles, phases) is populated."""
    from app.api.routers.ui import deal_setup_wizard_get
    return await deal_setup_wizard_get(
        request=request, model_id=model_id, session=session, step=1,
    )


@router.get("/ui/models/{model_id}/proforma-resume", response_class=HTMLResponse)
async def proforma_resume(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Wizard Step-2 "Back" target. If the scenario has a recent proforma
    import whose parse result is still cached, re-render the review page
    so the user can adjust line items without re-uploading. Falls back to
    Step 1 (upload UI) when nothing is cached."""
    import redis as _redis  # type: ignore

    r = _redis.from_url(settings.redis_url, decode_responses=True)
    file_hash = r.get(f"scenario:{model_id}:last_proforma_hash")
    if file_hash:
        cached_raw = r.get(f"proforma:filehash:{file_hash}:result")
        if cached_raw:
            try:
                result = json.loads(cached_raw)
            except Exception:
                result = None
            if result is not None:
                filename = r.get(f"scenario:{model_id}:last_proforma_filename") or ""
                parsed_at = r.get(f"proforma:filehash:{file_hash}:parsed_at") or ""
                # Synthetic task_id — re-analyze/purge routes will gracefully
                # handle missing file bytes (24h task TTL vs 7d hash TTL).
                resume_task_id = str(_uuid_mod.uuid4())
                r.set(f"proforma:{resume_task_id}:result", cached_raw, ex=86_400)
                return templates.TemplateResponse(
                    request,
                    "partials/proforma_review.html",
                    {
                        "model_id": model_id,
                        "task_id": resume_task_id,
                        "unit_types": result.get("unit_types", []),
                        "expense_lines": result.get("expense_lines", []),
                        "warnings": result.get("warnings", []),
                        "STANDARD_OPEX_CATEGORIES": STANDARD_OPEX_CATEGORIES,
                        "from_cache": True,
                        "file_hash": file_hash,
                        "filename": filename,
                        "parsed_at": parsed_at,
                    },
                )

    from app.api.routers.ui import deal_setup_wizard_get
    return await deal_setup_wizard_get(
        request=request, model_id=model_id, session=session, step=1,
    )


def _render_proforma_reanalyze(
    request: Request,
    model_id: UUID,
    task_id: str,
) -> HTMLResponse:
    """Run a fresh parse for an already-uploaded file (bytes still in Redis).

    For xlsx files, returns the sheet picker so the user can pick sheets again.
    For doc files (PDF/DOCX/etc.), queues the Celery task immediately and
    returns the progress poller.
    """
    import redis as _redis  # type: ignore

    r = _redis.from_url(settings.redis_url, decode_responses=False)
    file_bytes = r.get(f"proforma:{task_id}:file")
    if not file_bytes:
        return HTMLResponse(
            "<p class='text-red-500'>Upload expired. Please re-upload the file.</p>",
            status_code=410,
        )

    kind_raw = r.get(f"proforma:{task_id}:kind") or b"xlsx"
    file_kind = kind_raw.decode() if isinstance(kind_raw, bytes) else str(kind_raw)

    if file_kind == "doc":
        from app.tasks.proforma_parse import PARSE_PROFORMA_TASK
        from app.tasks.celery_app import celery_app as _celery
        _celery.send_task(
            PARSE_PROFORMA_TASK,
            kwargs={
                "task_id": task_id,
                "model_id": str(model_id),
                "revenue_sheet": "",
                "opex_sheet": "",
                "property_column": None,
                "file_kind": "doc",
            },
        )
        return templates.TemplateResponse(
            request,
            "partials/proforma_progress.html",
            {"model_id": model_id, "task_id": task_id},
        )

    return _render_proforma_sheet_picker(request, model_id, task_id, file_bytes)


@router.post("/ui/models/{model_id}/proforma-reanalyze", response_class=HTMLResponse)
async def proforma_reanalyze(
    request: Request,
    model_id: UUID,
    task_id: str = Form(...),
) -> HTMLResponse:
    """Skip the cache and run a fresh parse. Cache is left intact (use
    /proforma-purge-cache to delete the cached result)."""
    return _render_proforma_reanalyze(request, model_id, task_id)


@router.post("/ui/models/{model_id}/proforma-purge-cache", response_class=HTMLResponse)
async def proforma_purge_cache(
    request: Request,
    model_id: UUID,
    task_id: str = Form(...),
    file_hash: str = Form(...),
) -> HTMLResponse:
    """Delete the content-hash cache entry, then trigger a fresh parse."""
    import redis as _redis  # type: ignore

    if file_hash and len(file_hash) == 64:
        r = _redis.from_url(settings.redis_url, decode_responses=True)
        r.delete(f"proforma:filehash:{file_hash}:result")
        r.delete(f"proforma:filehash:{file_hash}:parsed_at")

    return _render_proforma_reanalyze(request, model_id, task_id)


@router.post("/ui/models/{model_id}/upload-proforma", response_class=HTMLResponse)
async def upload_proforma(
    request: Request,
    model_id: UUID,
    task_id: str = Form(...),
    revenue_sheet: str = Form(""),
    opex_sheet: str = Form(""),
    property_column: str = Form(""),
    revenue_enabled: str = Form(""),
    opex_enabled: str = Form(""),
    revenue_range: str = Form(""),
    opex_range: str = Form(""),
) -> HTMLResponse:
    """Queue the Celery parse task with the user-selected sheet/column/range
    coordinates, then return the progress-polling fragment."""
    from app.tasks.proforma_parse import PARSE_PROFORMA_TASK
    from app.tasks.celery_app import celery_app as _celery

    import_revenue = revenue_enabled.strip() == "on"
    import_opex = opex_enabled.strip() == "on"

    _celery.send_task(
        PARSE_PROFORMA_TASK,
        kwargs={
            "task_id": task_id,
            "model_id": str(model_id),
            "revenue_sheet": revenue_sheet,
            "opex_sheet": opex_sheet,
            "property_column": property_column or None,
            "file_kind": "xlsx",
            "import_revenue": import_revenue,
            "import_opex": import_opex,
            "revenue_range": revenue_range.strip() or None,
            "opex_range": opex_range.strip() or None,
        },
    )

    return templates.TemplateResponse(
        request,
        "partials/proforma_progress.html",
        {"model_id": model_id, "task_id": task_id},
    )


@router.post("/ui/models/{model_id}/upload-proforma-doc", response_class=HTMLResponse)
async def upload_proforma_doc(
    request: Request,
    model_id: UUID,
    task_id: str = Form(...),
    revenue_enabled: str = Form(""),
    opex_enabled: str = Form(""),
    revenue_pages: str = Form(""),
    opex_pages: str = Form(""),
) -> HTMLResponse:
    """Queue the Celery parse task for a PDF with user-selected page ranges,
    then return the progress-polling fragment."""
    from app.tasks.proforma_parse import PARSE_PROFORMA_TASK, _parse_pages
    from app.tasks.celery_app import celery_app as _celery

    import_revenue = revenue_enabled.strip() == "on"
    import_opex = opex_enabled.strip() == "on"
    rev_pages = _parse_pages(revenue_pages) if revenue_pages.strip() else None
    opx_pages = _parse_pages(opex_pages) if opex_pages.strip() else None

    _celery.send_task(
        PARSE_PROFORMA_TASK,
        kwargs={
            "task_id": task_id,
            "model_id": str(model_id),
            "revenue_sheet": "",
            "opex_sheet": "",
            "property_column": None,
            "file_kind": "doc",
            "import_revenue": import_revenue,
            "import_opex": import_opex,
            "revenue_pages": rev_pages,
            "opex_pages": opx_pages,
        },
    )

    return templates.TemplateResponse(
        request,
        "partials/proforma_progress.html",
        {"model_id": model_id, "task_id": task_id},
    )


@router.post("/ui/models/{model_id}/upload-proforma-multi", response_class=HTMLResponse)
async def upload_proforma_multi(
    request: Request,
    model_id: UUID,
) -> HTMLResponse:
    """Receive the multi-file config table, store email_config per file in Redis,
    dispatch a Celery parse task per file, then return progress for the first file.
    Subsequent files are processed in parallel; their results appear under the
    same scenario when confirmed."""
    import json as _json_multi
    import redis as _redis_multi
    from app.tasks.proforma_parse import PARSE_PROFORMA_TASK
    from app.tasks.celery_app import celery_app as _celery_multi

    form = await request.form()
    r = _redis_multi.from_url(settings.redis_url, decode_responses=True)

    rows: list[dict] = []
    i = 0
    while True:
        task_id = form.get(f"task_id_{i}")
        if task_id is None:
            break
        rows.append({
            "task_id": str(task_id),
            "file_kind": str(form.get(f"file_kind_{i}") or "doc"),
            "rev_sheet": str(form.get(f"rev_sheet_{i}") or ""),
            "rev_range": str(form.get(f"rev_range_{i}") or ""),
            "opex_sheet": str(form.get(f"opex_sheet_{i}") or ""),
            "opex_range": str(form.get(f"opex_range_{i}") or ""),
            "rev_pages": str(form.get(f"rev_pages_{i}") or ""),
            "opex_pages": str(form.get(f"opex_pages_{i}") or ""),
        })
        i += 1

    if not rows:
        return HTMLResponse("<p class='text-red-500'>No files submitted.</p>", status_code=400)

    first_task_id: str | None = None
    for row in rows:
        tid = row["task_id"]
        fkind = row["file_kind"]
        import_revenue = bool(row["rev_sheet"] or row["rev_pages"])
        import_opex = bool(row["opex_sheet"] or row["opex_pages"])

        # If neither enabled, default both on so the file still gets processed
        if not import_revenue and not import_opex:
            import_revenue = True
            import_opex = True

        cfg = {
            "file_kind": fkind,
            "rev_sheet": row["rev_sheet"],
            "rev_range": row["rev_range"],
            "opex_sheet": row["opex_sheet"],
            "opex_range": row["opex_range"],
            "rev_pages": row["rev_pages"],
            "opex_pages": row["opex_pages"],
            "import_revenue": import_revenue,
            "import_opex": import_opex,
        }
        r.set(f"proforma:{tid}:email_config", _json_multi.dumps(cfg), ex=7 * 86400)

        # Parse page strings to 0-based index lists
        def _pps(s: str) -> list[int] | None:
            if not s.strip():
                return None
            pages: list[int] = []
            for part in s.replace(" ", "").split(","):
                if "-" in part:
                    lo, _, hi = part.partition("-")
                    if lo.isdigit() and hi.isdigit():
                        pages.extend(range(int(lo), int(hi) + 1))
                elif part.isdigit():
                    pages.append(int(part))
            return [p - 1 for p in sorted(set(pages)) if p >= 1] or None

        _celery_multi.send_task(
            PARSE_PROFORMA_TASK,
            kwargs={
                "task_id": tid,
                "model_id": str(model_id),
                "revenue_sheet": row["rev_sheet"],
                "opex_sheet": row["opex_sheet"],
                "property_column": None,
                "file_kind": fkind,
                "import_revenue": import_revenue,
                "import_opex": import_opex,
                "revenue_range": row["rev_range"] or None,
                "opex_range": row["opex_range"] or None,
                "revenue_pages": _pps(row["rev_pages"]),
                "opex_pages": _pps(row["opex_pages"]),
            },
        )
        if first_task_id is None:
            first_task_id = tid

    return templates.TemplateResponse(
        request,
        "partials/proforma_progress.html",
        {"model_id": model_id, "task_id": first_task_id},
    )


@router.get("/ui/models/{model_id}/proforma-status/{task_id}", response_class=HTMLResponse)
async def proforma_status(
    request: Request,
    model_id: UUID,
    task_id: str,
    session: DBSession,
) -> HTMLResponse:
    """HTMX poll endpoint. Returns progress fragment while running; switches to
    the review fragment when the task completes or errors."""
    import redis as _redis  # type: ignore

    r = _redis.from_url(settings.redis_url, decode_responses=True)
    raw = r.get(f"proforma:{task_id}:progress")

    if not raw:
        return templates.TemplateResponse(
            request,
            "partials/proforma_progress.html",
            {"model_id": model_id, "task_id": task_id, "step": 0, "total": 3, "message": "Queued…"},
        )

    progress = json.loads(raw)
    status = progress.get("status", "running")

    if status == "error":
        return templates.TemplateResponse(
            request,
            "partials/proforma_progress.html",
            {
                "model_id": model_id,
                "task_id": task_id,
                "error": progress.get("message", "Unknown error"),
            },
        )

    if status != "done":
        import time as _time
        updated_at = progress.get("updated_at")
        if updated_at and (_time.time() - float(updated_at)) > 180:
            return templates.TemplateResponse(
                request,
                "partials/proforma_progress.html",
                {
                    "model_id": model_id,
                    "task_id": task_id,
                    "error": "Analysis timed out — the AI took too long to respond. Please try again.",
                },
            )
        return templates.TemplateResponse(
            request,
            "partials/proforma_progress.html",
            {
                "model_id": model_id,
                "task_id": task_id,
                "step": progress.get("step", 0),
                "total": progress.get("total", 3),
                "message": progress.get("message", ""),
            },
        )

    # Done — load result and render review UI
    raw_result = r.get(f"proforma:{task_id}:result")
    result = json.loads(raw_result) if raw_result else {"unit_types": [], "expense_lines": [], "warnings": []}
    filename = r.get(f"proforma:{task_id}:filename") or ""

    return templates.TemplateResponse(
        request,
        "partials/proforma_review.html",
        {
            "model_id": model_id,
            "task_id": task_id,
            "unit_types": result.get("unit_types", []),
            "expense_lines": result.get("expense_lines", []),
            "warnings": result.get("warnings", []),
            "STANDARD_OPEX_CATEGORIES": STANDARD_OPEX_CATEGORIES,
            "filename": filename,
        },
    )


@router.post("/ui/models/{model_id}/proforma-confirm", response_class=HTMLResponse)
async def proforma_confirm(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Commit user-approved revenue and expense lines to the database.

    Accepts multipart form data built by the review template:
    - ``unit_type_name[]``, ``unit_type_count[]``, ``unit_type_sqft[]``,
      ``unit_type_rent[]`` — parallel arrays for each confirmed unit type
    - ``expense_label[]``, ``expense_amount[]``, ``expense_category[]``,
      ``expense_include[]`` — parallel arrays for each expense line
      (``expense_include`` contains the indices of rows the user kept checked)
    """
    from sqlalchemy import delete

    form = await request.form()

    deal_model = await session.get(Scenario, model_id)
    if not deal_model:
        raise HTTPException(status_code=404, detail="Deal model not found")

    # Multi-project deals: route import to the project the user is viewing
    # (from HX-Current-URL ?project=...). Fall back to oldest only if missing.
    project_id = await _active_project_from_request(request, session, model_id)
    if project_id is None:
        default_project = (await session.execute(
            select(Project).where(Project.scenario_id == model_id).order_by(Project.created_at).limit(1)
        )).scalar_one_or_none()
        if not default_project:
            raise HTTPException(status_code=400, detail="No project found")
        project_id = default_project.id

    inputs = (await session.execute(
        select(OperationalInputs).where(OperationalInputs.project_id == project_id)
    )).scalar_one_or_none()

    # ---------- Revenue / unit mix (JSONB on Project) ----------
    names = form.getlist("unit_type_name[]")
    counts = form.getlist("unit_type_count[]")
    sqfts = form.getlist("unit_type_sqft[]")
    rents = form.getlist("unit_type_rent[]")
    market_rents = form.getlist("unit_type_market_rent[]")
    modes = form.getlist("unit_type_mode[]")          # "unit" or "flat" per row
    stream_types = form.getlist("unit_type_stream_type[]")  # flat-rate type per row
    # Indices of unit-mix rows the user kept checked. When the include list
    # is absent (older clients), accept all rows for backward compat.
    unit_included_raw = form.getlist("unit_type_include[]")
    unit_included = {int(i) for i in unit_included_raw if str(i).strip().isdigit()}
    unit_filter_enabled = bool(unit_included_raw)

    rent_type = (form.get("rent_type") or "in_place").strip().lower()
    rent_field = "market_rent_per_unit" if rent_type == "market" else "in_place_rent_per_unit"

    if names:
        proj_result = await session.execute(
            select(Project).where(Project.id == project_id)
        )
        project = proj_result.scalar_one()
        _modes = modes if modes else ["unit"] * len(names)
        _stypes = stream_types if stream_types else ["residential_rent"] * len(names)
        unit_mix_rows = []
        flat_income_rows: list[dict] = []
        for idx, (name, count_s, sqft_s, rent_s, row_mode, row_stype) in enumerate(
            zip(names, counts, sqfts, rents, _modes, _stypes)
        ):
            if unit_filter_enabled and idx not in unit_included:
                continue
            name = name.strip()
            if not name:
                continue
            if row_mode == "flat":
                try:
                    flat_amt = Decimal((rent_s or "0").replace(",", ""))
                except Exception:
                    flat_amt = Decimal("0")
                if flat_amt > 0:
                    flat_income_rows.append({"label": name, "stream_type": row_stype, "flat_monthly": flat_amt})
                continue
            try:
                row = {
                    "label": name,
                    "unit_count": int((count_s or "0").replace(",", "")),
                    "avg_sqft": float((sqft_s or "0").replace(",", "")),
                    "beds": None,
                    "baths": None,
                    "unit_strategy": "base_escalation",
                    "notes": None,
                }
                row[rent_field] = float((rent_s or "0").replace(",", ""))
                # If user supplied a Market Rent alongside in-place, capture it
                # so LTL catchup can be modeled later without re-entry.
                if rent_type == "in_place" and idx < len(market_rents):
                    mkt_s = (market_rents[idx] or "").replace(",", "").strip()
                    if mkt_s:
                        try:
                            mkt_v = float(mkt_s)
                            if mkt_v > 0:
                                row["market_rent_per_unit"] = mkt_v
                        except ValueError:
                            pass
                unit_mix_rows.append(row)
            except Exception:
                pass
        if unit_mix_rows:
            from sqlalchemy.orm.attributes import flag_modified
            project.unit_mix = unit_mix_rows
            flag_modified(project, "unit_mix")
            session.add(project)

        if unit_mix_rows or flat_income_rows:
            # Seed IncomeStream rows so the Revenue tab is populated.
            # Overwrite semantics: drop existing streams for this project, then
            # insert one per row (unit or flat) using the submitted amounts.
            await session.execute(
                delete(IncomeStream).where(IncomeStream.project_id == project_id)
            )
            for row in unit_mix_rows:
                rent = Decimal(str(row.get(rent_field) or 0))
                count = int(row.get("unit_count") or 0)
                if count <= 0:
                    continue
                mkt_rent_raw = row.get("market_rent_per_unit")
                mkt_rent = Decimal(str(mkt_rent_raw)) if mkt_rent_raw else None
                session.add(IncomeStream(
                    project_id=project_id,
                    label=f"{row['label']} Rent",
                    stream_type=IncomeStreamType.residential_rent,
                    unit_count=count,
                    amount_per_unit_monthly=rent,
                    catchup_target_rent=mkt_rent,
                    stabilized_occupancy_pct=Decimal("95"),
                    escalation_rate_pct_annual=Decimal("3"),
                    active_in_phases=["lease_up", "stabilized"],
                ))
            for flat in flat_income_rows:
                try:
                    stype = IncomeStreamType(flat["stream_type"])
                except ValueError:
                    stype = IncomeStreamType.other
                session.add(IncomeStream(
                    project_id=project_id,
                    label=flat["label"],
                    stream_type=stype,
                    amount_fixed_monthly=flat["flat_monthly"],
                    unit_count=None,
                    amount_per_unit_monthly=None,
                    stabilized_occupancy_pct=Decimal("100"),
                    escalation_rate_pct_annual=Decimal("3"),
                    active_in_phases=["lease_up", "stabilized"],
                ))

    # ---------- OpEx lines ----------
    # Label field holds the mapped category (investor export groups by label).
    # Original source label is preserved in notes.
    orig_labels = form.getlist("expense_orig_label[]")
    labels = form.getlist("expense_label[]")
    amounts = form.getlist("expense_amount[]")
    included_indices = {int(i) for i in form.getlist("expense_include[]")}

    if labels:
        await session.execute(
            delete(OperatingExpenseLine).where(OperatingExpenseLine.project_id == project_id)
        )
        for idx, (orig_label, label, amount_s) in enumerate(zip(orig_labels, labels, amounts)):
            if idx not in included_indices:
                continue
            label = label.strip()
            if not label:
                continue
            try:
                session.add(OperatingExpenseLine(
                    project_id=project_id,
                    label=label,
                    annual_amount=Decimal((amount_s or "0").replace(",", "")),
                    escalation_rate_pct_annual=Decimal("3"),
                    active_in_phases=["lease_up", "stabilized"],
                    notes=orig_label.strip() if orig_label.strip() != label else None,
                ))
            except Exception:
                pass

    await session.commit()

    # Re-enter the wizard at Step 2 via the canonical GET handler so the full
    # context (source_vehicles_debt, phases_present, review_back_step, etc.)
    # is populated — without it the Source Vehicle dropdowns never render.
    from app.api.routers.ui import deal_setup_wizard_get
    return await deal_setup_wizard_get(
        request=request, model_id=model_id, session=session, step=2,
    )


@router.get("/ui/models/{model_id}/proforma-skip", response_class=HTMLResponse)
async def proforma_skip(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Skip pro forma import — advance wizard to Step 2 (debt types)."""
    # Delegate to the canonical GET so source_vehicles_debt et al. are
    # populated (needed for the Source Vehicle dropdown on each debt card).
    from app.api.routers.ui import deal_setup_wizard_get
    return await deal_setup_wizard_get(
        request=request, model_id=model_id, session=session, step=2,
    )


@router.post("/ui/models/{model_id}/noi-inputs", response_class=HTMLResponse)
async def save_noi_inputs(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Save NOI mode inputs (stabilized NOI + escalation rate) and return refreshed form."""
    model = await session.get(Scenario, model_id)
    if model is None:
        return HTMLResponse("Not found", status_code=404)
    default_project = (await session.execute(
        select(Project).where(Project.scenario_id == model_id).order_by(Project.created_at).limit(1)
    )).scalar_one_or_none()
    if default_project is None:
        return HTMLResponse("No project", status_code=400)
    inputs = (await session.execute(
        select(OperationalInputs).where(OperationalInputs.project_id == default_project.id)
    )).scalar_one_or_none()
    if inputs is None:
        return HTMLResponse("No inputs", status_code=400)

    form = await request.form()
    noi_raw = str(form.get("noi_stabilized_input", "")).strip()
    esc_raw = form.get("noi_escalation_rate_pct", "3")
    # Strip any display formatting ($ and commas) before parsing.
    noi_clean = noi_raw.replace("$", "").replace(",", "").strip()
    try:
        inputs.noi_stabilized_input = Decimal(noi_clean) if noi_clean else None
    except Exception:
        inputs.noi_stabilized_input = None
    try:
        inputs.noi_escalation_rate_pct = Decimal(str(esc_raw)) if esc_raw else Decimal("3")
    except Exception:
        inputs.noi_escalation_rate_pct = Decimal("3")
    # User explicitly submitted — clear the auto-seeded flag whether they
    # accepted the suggested value or overrode it. Banner disappears.
    inputs.noi_auto_seeded = False
    session.add(inputs)
    await session.commit()
    await session.refresh(inputs)

    _noi_val = float(inputs.noi_stabilized_input) if inputs.noi_stabilized_input else ""
    _esc_val = float(inputs.noi_escalation_rate_pct) if inputs.noi_escalation_rate_pct else 3.0
    html = f"""<form hx-post="/ui/models/{model_id}/noi-inputs"
        hx-target="this"
        hx-swap="outerHTML"
        style="max-width:480px">
    <div style="background:var(--success-faint,#f0fdf4);border:1px solid var(--success,#22c55e);border-radius:6px;padding:8px 12px;margin-bottom:16px;font-size:12px;color:var(--success,#16a34a)">
      ✓ NOI inputs saved.
    </div>
    <div class="field-group" style="margin-bottom:20px">
      <label style="display:block;font-size:12px;font-weight:600;text-transform:uppercase;letter-spacing:.04em;color:var(--text-secondary);margin-bottom:4px">Stabilized NOI (Annual)</label>
      <input type="number" name="noi_stabilized_input" step="1000" min="0"
             value="{_noi_val}" placeholder="e.g. 500000"
             style="width:100%;box-sizing:border-box;padding:8px 10px;border:1px solid var(--border);border-radius:6px;font-size:14px;background:var(--bg);color:var(--text)">
      <div style="font-size:11px;color:var(--text-muted);margin-top:3px">Net Operating Income at stabilization — pre-debt service, post-OpEx (even though OpEx is not modeled separately).</div>
    </div>
    <div class="field-group" style="margin-bottom:20px">
      <label style="display:block;font-size:12px;font-weight:600;text-transform:uppercase;letter-spacing:.04em;color:var(--text-secondary);margin-bottom:4px">Annual NOI Escalation Rate (%)</label>
      <input type="number" name="noi_escalation_rate_pct" step="0.25" min="0" max="20"
             value="{_esc_val}" placeholder="3.0"
             style="width:140px;box-sizing:border-box;padding:8px 10px;border:1px solid var(--border);border-radius:6px;font-size:14px;background:var(--bg);color:var(--text)">
      <div style="font-size:11px;color:var(--text-muted);margin-top:3px">Compound annual growth applied to NOI each year. Typical: 2–4%.</div>
    </div>
    <div>
      <button type="submit" class="btn btn-primary">Save NOI Inputs</button>
    </div>
  </form>"""
    return HTMLResponse(html)


@router.get("/ui/models/{model_id}/line-form", response_class=HTMLResponse)
async def model_builder_line_form(
    request: Request,
    model_id: UUID,
    session: DBSession,
    type: str = Query(default="uses"),
    id: str = Query(default=""),
    phase: str = Query(default=""),
    category: str = Query(default="soft"),
) -> HTMLResponse:
    """Serves the add/edit form inside the line-item drawer."""
    model = await session.get(Scenario, model_id)
    if model is None:
        return HTMLResponse("<p class='text-muted'>Model not found.</p>", status_code=404)

    existing = None
    if id:
        try:
            eid = UUID(id)
            if type in ("use_lines", "uses"):
                existing = await session.get(UseLine, eid)
            elif type in ("income_streams", "revenue"):
                existing = await session.get(IncomeStream, eid)
            elif type in ("expense_lines", "opex"):
                existing = await session.get(OperatingExpenseLine, eid)
            elif type in ("capital_modules", "sources"):
                existing = await session.get(CapitalModule, eid)
            elif type in ("waterfall_tiers", "waterfall"):
                existing = await session.get(WaterfallTier, eid)
            elif type in ("milestones", "timeline"):
                existing = await session.get(Milestone, eid)
            elif type == "unit_mix":
                _lf_proj = (await session.execute(
                    select(Project).where(Project.scenario_id == model_id).order_by(Project.created_at.asc()).limit(1)
                )).scalar_one_or_none()
                if _lf_proj:
                    _um_dict = next((d for d in (_lf_proj.unit_mix or []) if d.get("id") == str(eid)), None)
                    if _um_dict:
                        existing = _UMRow(_um_dict)
        except ValueError:
            pass

    # For milestone forms: load siblings + compute which would be circular triggers
    sibling_milestones = []
    circular_ids: set = set()
    trigger_end_date = None  # ISO string passed to JS for end-date preview
    default_trigger_id: str | None = None
    if type in ("milestones", "timeline"):
        # Determine which project's milestones to load for the trigger dropdown:
        # 1. If editing an existing milestone, use its own project_id.
        # 2. If adding, use the ?project= query param passed by the caller.
        # 3. Fall back to the first (oldest) project only as a last resort.
        ms_project_id: UUID | None = None
        if existing is not None and hasattr(existing, "project_id"):
            ms_project_id = existing.project_id
        if ms_project_id is None:
            _proj_param = request.query_params.get("project", "")
            if _proj_param:
                try:
                    ms_project_id = UUID(_proj_param)
                except ValueError:
                    pass
        if ms_project_id is None:
            _fp = (await session.execute(
                select(Project).where(Project.scenario_id == model_id).order_by(Project.created_at.asc()).limit(1)
            )).scalar_one_or_none()
            ms_project_id = _fp.id if _fp else None
        if ms_project_id is not None:
            all_ms = list((await session.execute(
                select(Milestone).where(Milestone.project_id == ms_project_id)
            )).scalars())
            _SPHASE_ORDER = [
                "offer_made", "under_contract", "close", "pre_development",
                "construction", "operation_lease_up", "operation_stabilized", "divestment",
            ]
            def _sphase_idx(m):
                raw = str(m.milestone_type).replace("MilestoneType.", "")
                return next((i for i, v in enumerate(_SPHASE_ORDER) if v == raw), 99)
            editing_id = existing.id if existing else None
            sibling_milestones = sorted(
                [m for m in all_ms if m.id != editing_id],
                key=_sphase_idx
            )
            ms_map_local = {m.id: m for m in all_ms}
            # Detect circular: candidate Y is circular if following Y's trigger chain hits editing_id
            if editing_id:
                for candidate in sibling_milestones:
                    visited: set = set()
                    cur = candidate
                    while cur and cur.trigger_milestone_id:
                        if cur.trigger_milestone_id == editing_id:
                            circular_ids.add(candidate.id)
                            break
                        if cur.id in visited:
                            break
                        visited.add(cur.id)
                        cur = ms_map_local.get(cur.trigger_milestone_id)
            # default_trigger_id is no longer used by the template — predecessor
            # auto-selection is handled client-side by _msAutoTrigger() JS in the form.
            # Kept as a no-op so the template context key still exists.
            # Resolve trigger's end date so JS can preview end date on the form
            if existing and existing.trigger_milestone_id:
                trigger = ms_map_local.get(existing.trigger_milestone_id)
                if trigger:
                    t_end = trigger.computed_end(ms_map_local)
                    if t_end:
                        trigger_end_date = t_end.isoformat()

    # Lock duration for operation_stabilized when no divestment milestone exists
    lock_duration = False
    _STABILIZED_AUTO_DAYS = 10950
    if (
        existing
        and hasattr(existing, "milestone_type")
        and str(existing.milestone_type).replace("MilestoneType.", "") == "operation_stabilized"
    ):
        has_div = any(
            str(m.milestone_type).replace("MilestoneType.", "") == "divestment"
            for m in sibling_milestones
        )
        if not has_div:
            lock_duration = True

    _PHASE_LABELS = {
        "offer_made": "Offer Made", "under_contract": "Under Contract",
        "close": "Close / Acquisition", "pre_development": "Pre-Development",
        "construction": "Construction", "operation_lease_up": "Lease-Up",
        "operation_stabilized": "Stabilized Operations", "divestment": "Divestment / Exit",
    }

    # Phase options scoped to this deal type — prevents assigning costs to phases that don't exist
    _project_type_str = str(getattr(model, "project_type", "") or "").replace("ProjectType.", "")
    _USE_PHASES_BY_TYPE: dict[str, list[tuple[str, str]]] = {
        "acquisition": [
            ("acquisition", "Acquisition"),
            ("operation", "Operations"),
            ("exit", "Exit / Sale"),
            ("other", "Other"),
        ],
        "value_add": [
            ("acquisition", "Acquisition"),
            ("pre_construction", "Pre-Development"),
            ("construction", "Construction / Renovation"),
            ("operation", "Operations"),
            ("exit", "Exit / Sale"),
            ("other", "Other"),
        ],
        "conversion": [
            ("acquisition", "Acquisition"),
            ("conversion", "Conversion"),
            ("operation", "Operations"),
            ("exit", "Exit / Sale"),
            ("other", "Other"),
        ],
        "new_construction": [
            ("acquisition", "Acquisition"),
            ("pre_construction", "Pre-Construction"),
            ("construction", "Construction"),
            ("operation", "Operations"),
            ("exit", "Exit / Sale"),
            ("other", "Other"),
        ],
    }
    _default_phases = [
        ("acquisition", "Acquisition"), ("pre_construction", "Pre-Construction"),
        ("construction", "Construction"), ("renovation", "Renovation"),
        ("conversion", "Conversion"), ("operation", "Operations"),
        ("exit", "Exit / Sale"), ("other", "Other"),
    ]
    valid_use_phases = _USE_PHASES_BY_TYPE.get(_project_type_str, _default_phases)

    # For capital module and use line forms: load milestones for pickers
    milestones_dated_ds: list[dict] = []
    draw_source_window = None
    if type in ("capital_modules", "sources", "use_lines", "uses"):
        from app.models.project import Project as _LFProject
        from app.models.milestone import Milestone as _LFMilestone
        _lf_proj = (await session.execute(
            select(_LFProject).where(_LFProject.scenario_id == model_id).order_by(_LFProject.created_at.asc()).limit(1)
        )).scalar_one_or_none()
        if _lf_proj:
            _lf_opp_ms = list((await session.execute(
                select(_LFMilestone).where(_LFMilestone.opportunity_id == _lf_proj.opportunity_id)
            )).scalars()) if _lf_proj.opportunity_id else []
            _lf_proj_ms = list((await session.execute(
                select(_LFMilestone).where(_LFMilestone.project_id == _lf_proj.id)
            )).scalars())
            _lf_all_ms = _lf_opp_ms + _lf_proj_ms
            _lf_ms_map = {m.id: m for m in _lf_all_ms}
            for m in _lf_all_ms:
                _start = m.computed_start(_lf_ms_map)
                if _start:
                    _key = m.milestone_type.value if hasattr(m.milestone_type, "value") else str(m.milestone_type)
                    milestones_dated_ds.append({"key": _key, "label": _milestone_label(_key), "date": _start})
            milestones_dated_ds.sort(key=lambda x: x["date"])
            # Append "maturity" pseudo-milestone for the Active To dropdown only
            milestones_dated_ds.append({"key": "maturity", "label": "Maturity", "date": None})
        if existing and type in ("capital_modules", "sources"):
            # Prefer lookup by capital_module_id (reliable for wizard-created sources);
            # fall back to label match for legacy sources created before the FK existed.
            _ds_q = select(DrawSource).where(
                DrawSource.scenario_id == model_id,
                DrawSource.capital_module_id == existing.id,
            ).limit(1)
            draw_source_window = (await session.execute(_ds_q)).scalar_one_or_none()
            if draw_source_window is None:
                _ds_q = select(DrawSource).where(
                    DrawSource.scenario_id == model_id,
                    DrawSource.label == existing.label,
                ).limit(1)
                draw_source_window = (await session.execute(_ds_q)).scalar_one_or_none()

    # Exit Vehicle dropdown options (capital modules only). Dynamic from the
    # current module's active_phase_end + siblings' active windows.
    exit_vehicle_options: list[dict] = []
    show_exit_vehicle = False
    show_active_window = False
    # Exit Vehicle applies to all debt modules (vehicle_type == "debt").
    _EXIT_VEHICLE_APPLIES_UI: set[str] = set()  # unused — replaced by vehicle_type check below
    if type in ("capital_modules", "sources"):
        from app.engines.cashflow import (
            _APS_TO_RANK as _EXIT_APS_RANK,
            _resolve_vehicle as _exit_resolve,
        )

        siblings = list((await session.execute(
            select(CapitalModule).where(CapitalModule.scenario_id == model_id)
        )).scalars())
        others = [m for m in siblings if not existing or m.id != existing.id]
        # Build a "candidate" module stand-in for the resolve call — for new
        # modules we have no saved active_phase_end yet; default to "perpetuity"
        # (→ Maturity as the only option) to match the form's initial blank
        # state.  For existing modules we use their actual saved values.
        # New-source wizards haven't picked active_phase_end yet — so
        # eligible-by-rank gives zero results. Fall back to "all other
        # modules" so the user can pre-select a takeout target. The engine
        # re-validates at compute time and falls back to maturity if the
        # eventual active_phase_end doesn't actually overlap.
        is_new = existing is None
        if not is_new:
            candidate = existing
        else:
            class _Stub:  # minimal shim
                id = None
                active_phase_start = "acquisition"
                active_phase_end = ""
                exit_terms: dict = {}
            candidate = _Stub()

        _vehicle_now, _retirer_now = _exit_resolve(candidate, [candidate] + others)
        saved_val = ""
        if existing is not None and isinstance(existing.exit_terms, dict):
            saved_val = (existing.exit_terms.get("vehicle") or "").strip()

        # Compute eligible source retirers via same rank logic used by engine
        e_rank = _EXIT_APS_RANK.get(
            str(getattr(candidate, "active_phase_end", "") or ""), 99
        )

        def _rank(m: object, side: str) -> int:
            raw = str(getattr(m, f"active_phase_{side}", "") or "")
            if side == "end":
                return _EXIT_APS_RANK.get(raw, 99)
            return _EXIT_APS_RANK.get(raw, 0)

        # List all other sources as candidates. Overlap is too brittle a
        # filter — adjacent-vs-overlapping distinctions flip on rank
        # mapping (a new loan often starts the day the old closes). The
        # engine honours the user's explicit pick at compute time.
        eligible_sources = list(others)

        def _opt(value: str, label: str) -> dict:
            # If saved vehicle is present, honour it; else default to what
            # _resolve_vehicle picked.
            if saved_val:
                selected = (value == saved_val)
            elif _vehicle_now == "source" and _retirer_now is not None:
                selected = (value == str(getattr(_retirer_now, "id", "")))
            else:
                selected = (value == _vehicle_now)
            return {"value": value, "label": label, "selected": selected}

        exit_vehicle_options.append(_opt("maturity", "Maturity"))
        # Sale is always a valid exit for any debt instrument — the asset can
        # be sold at any point, retiring outstanding balances.
        exit_vehicle_options.append(_opt("sale", "Sale (divestment)"))
        for m in sorted(
            eligible_sources,
            key=lambda r: (int(getattr(r, "stack_position", 0) or 0), str(r.label or "")),
        ):
            exit_vehicle_options.append(_opt(str(m.id), m.label or "(unlabeled)"))

        # Gate Exit Vehicle + draw cadence UI on vehicle type.  Non-debt vehicle types
        # (equity, grants, forgivable loans) don't have a repayment concept — form hides them.
        _existing_vt = ""
        if existing is not None:
            _existing_vt = str(getattr(existing, "vehicle_type", "") or "").replace("VehicleType.", "")
        # New modules default to debt (see line form template default).
        _effective_vt = _existing_vt or "debt"
        show_exit_vehicle = _effective_vt == "debt"
        show_active_window = show_exit_vehicle

    # Source vehicle presets for the wizard dropdown (only shown on add, not edit)
    _sv_list: list[dict] = []
    if not id:
        _lf_user = await _get_user(session, request)
        if _lf_user is not None:
            from app.models.source_vehicle import SourceVehicle as _SV_lf
            _all_svs_lf = (await session.execute(
                select(_SV_lf).where(
                    ((_SV_lf.scope == "org") & (_SV_lf.owner_id == _lf_user.org_id)) |
                    ((_SV_lf.scope == "user") & (_SV_lf.owner_id == _lf_user.id))
                ).order_by(_SV_lf.label)
            )).scalars().all()
            _sv_list = [
                {
                    "id": str(v.id),
                    "name": v.label,
                    "vehicle_type": v.vehicle_type,
                    "equity_role": v.equity_role or "",
                    "owner": v.scope,
                }
                for v in _all_svs_lf
            ]

    # Per-Use eligibility checklist (grant cap UI). Pull all Use lines in this
    # scenario across projects so the source-side edit form can render
    # checkboxes for each Use; pre-tick those already referencing this module.
    _eligibility_uses: list[dict] = []
    if type in ("capital_modules", "sources", "capital-modules"):
        from app.schemas.gap_adjustment_names import is_reserved_label
        _ul_rows = (await session.execute(
            select(UseLine)
            .join(Project, UseLine.project_id == Project.id)
            .where(Project.scenario_id == model_id)
            .order_by(UseLine.label.asc())
        )).scalars().all()
        _existing_id_str = str(existing.id) if existing is not None else ""
        for _ul in _ul_rows:
            if is_reserved_label(_ul.label or ""):
                continue
            if not _ul.amount or float(_ul.amount) <= 0:
                continue
            _eligible_ids = _ul.eligible_module_ids or []
            _is_ticked = any(str(x) == _existing_id_str for x in _eligible_ids) if _existing_id_str else False
            _eligibility_uses.append({
                "id": str(_ul.id),
                "label": _ul.label or "(unlabeled)",
                "amount": float(_ul.amount or 0),
                "phase": str(getattr(_ul.phase, "value", _ul.phase) or ""),
                "ticked": _is_ticked,
            })

    # Sibling CapitalModules + scenario milestones — used by the float-earnings
    # child-source form to populate the parent / paydown-debt / paydown-milestone
    # dropdowns. Loaded only on the capital-module form to keep other form types
    # cheap. Editing module excludes itself from its own parent dropdown.
    _sibling_capital_modules: list[dict] = []
    _scenario_milestones: list[dict] = []
    if type in ("capital_modules", "sources", "capital-modules"):
        _editing_cm_id_str = str(existing.id) if existing is not None else ""
        _cm_rows = (await session.execute(
            select(CapitalModule)
            .where(CapitalModule.scenario_id == model_id)
            .order_by(CapitalModule.stack_position.asc(), CapitalModule.label.asc())
        )).scalars().all()
        for _cm in _cm_rows:
            if str(_cm.id) == _editing_cm_id_str:
                continue
            _src = _cm.source or {}
            _sibling_capital_modules.append({
                "id": str(_cm.id),
                "label": _cm.label or "(unlabeled)",
                "vehicle_type": (str(_cm.vehicle_type or "")).replace("VehicleType.", ""),
                "draw_type": _src.get("draw_type") or "",
                "balance_earns_interest": bool(_src.get("balance_earns_interest")),
                "balance_earns_interest": bool(_src.get("balance_earns_interest")),
            })

        _ms_rows = (await session.execute(
            select(Milestone)
            .join(Project, Milestone.project_id == Project.id)
            .where(Project.scenario_id == model_id)
            .order_by(Milestone.project_id, Milestone.created_at.asc())
        )).scalars().all()
        for _ms in _ms_rows:
            _ms_key = str(_ms.milestone_type or "").replace("MilestoneType.", "")
            _scenario_milestones.append({
                "id": str(_ms.id),
                "label": _ms.label or _milestone_label(_ms_key),
                "milestone_type": _ms_key,
            })

    has_acquisition_costs = False
    if type in ("use_lines", "uses") and existing is not None:
        _acq_count = (await session.execute(
            select(func.count()).select_from(UseLine)
            .join(Project, UseLine.project_id == Project.id)
            .where(
                Project.scenario_id == model_id,
                UseLine.phase == "acquisition",
                UseLine.is_auto_dev_fee == False,  # noqa: E712
                UseLine.is_auto_acquisition_fee == False,  # noqa: E712
            )
        )).scalar_one()
        has_acquisition_costs = _acq_count > 0

    acq_fee_pct_prefill = None
    if existing is not None and getattr(existing, "is_auto_dev_fee", False):
        _acq_fee_row = (await session.execute(
            select(UseLine).join(Project, UseLine.project_id == Project.id)
            .where(
                Project.scenario_id == model_id,
                UseLine.is_auto_acquisition_fee == True,  # noqa: E712
            )
        )).scalars().first()
        if _acq_fee_row is not None:
            acq_fee_pct_prefill = _acq_fee_row.acquisition_fee_pct

    # Auto-FC rows: resolve the parent Source's Active From milestone for the
    # locked, read-only "Active From" display in the line form.
    auto_fc_source_label = None
    auto_fc_milestone_key = None
    auto_fc_milestone_label = None
    if existing is not None and getattr(existing, "is_auto_finance_cost", False):
        _src_cm_id = getattr(existing, "source_capital_module_id", None)
        if _src_cm_id is not None:
            from app.models.capital import CapitalModule as _AFCModule
            from app.models.milestone import Milestone as _AFCMilestone
            _src_cm = await session.get(_AFCModule, _src_cm_id)
            if _src_cm is not None:
                auto_fc_source_label = getattr(_src_cm, "label", None) or "Source"
                _src_ms_id = getattr(_src_cm, "active_from_milestone_id", None)
                if _src_ms_id is not None:
                    _src_ms = await session.get(_AFCMilestone, _src_ms_id)
                    if _src_ms is not None:
                        _mt = _src_ms.milestone_type
                        auto_fc_milestone_key = _mt.value if hasattr(_mt, "value") else str(_mt)
                        auto_fc_milestone_label = _src_ms.label or _milestone_label(auto_fc_milestone_key)

    # Pre-resolve the existing Use Line's Active From / To milestone keys so
    # the drawer dropdown can pre-select the user's saved choice. Migration
    # 0086 dropped the legacy milestone_key string column, so the template
    # cannot read it directly off the row anymore.
    cur_active_from_key = ""
    cur_active_to_key = ""
    if type in ("use_lines", "uses") and existing is not None:
        from app.models.milestone import Milestone as _ULMilestone
        _from_id = getattr(existing, "active_from_milestone_id", None)
        if _from_id is not None:
            _from_ms = await session.get(_ULMilestone, _from_id)
            if _from_ms is not None:
                _mtf = _from_ms.milestone_type
                cur_active_from_key = _mtf.value if hasattr(_mtf, "value") else str(_mtf)
        _to_id = getattr(existing, "spread_to_milestone_id", None)
        if _to_id is not None:
            _to_ms = await session.get(_ULMilestone, _to_id)
            if _to_ms is not None:
                _mtt = _to_ms.milestone_type
                cur_active_to_key = _mtt.value if hasattr(_mtt, "value") else str(_mtt)

    return templates.TemplateResponse(request, "partials/model_builder_line_form.html", {
        "model": model,
        "form_type": type,
        "existing": existing,
        "default_phase": phase or "acquisition",
        "sibling_milestones": sibling_milestones,
        "circular_ids": circular_ids,
        "trigger_end_date": trigger_end_date,
        "default_trigger_id": default_trigger_id,
        "lock_duration": lock_duration,
        "phase_labels": _PHASE_LABELS,
        "valid_use_phases": valid_use_phases,
        "milestones_dated_ds": milestones_dated_ds,
        "draw_source_window": draw_source_window,
        "exit_vehicle_options": exit_vehicle_options,
        "show_exit_vehicle": show_exit_vehicle,
        "show_active_window": show_active_window,
        "exit_vehicle_applies": [],  # deprecated — templates now use vehicle_type == "debt" check
        "opex_categories": STANDARD_OPEX_CATEGORIES,
        "default_category": (getattr(existing, "cost_category", None) if existing else None) or category,
        "use_cost_categories": USE_COST_CATEGORIES,
        "use_category_labels": USE_CATEGORY_LABELS,
        "use_category_presets": USE_CATEGORY_PRESETS,
        "source_vehicles": _sv_list,
        "eligibility_uses": _eligibility_uses,
        "sibling_capital_modules": _sibling_capital_modules,
        "scenario_milestones": _scenario_milestones,
        "basis_buckets": BASIS_BUCKETS,
        "has_acquisition_costs": has_acquisition_costs,
        "acq_fee_pct_prefill": acq_fee_pct_prefill,
        "auto_fc_source_label": auto_fc_source_label,
        "auto_fc_milestone_key": auto_fc_milestone_key,
        "auto_fc_milestone_label": auto_fc_milestone_label,
        "cur_active_from_key": cur_active_from_key,
        "cur_active_to_key": cur_active_to_key,
    })


# ---------------------------------------------------------------------------
# Source Vehicle prefill endpoint
# ---------------------------------------------------------------------------


@router.get("/ui/source-vehicles/{vehicle_id}/prefill")
async def source_vehicle_prefill(
    request: Request,
    vehicle_id: UUID,
    session: DBSession,
) -> JSONResponse:
    """Return flat form-field values for a source vehicle (used by wizard dropdown JS)."""
    from app.models.source_vehicle import SourceVehicle as _SV_pf

    user = await _get_user(session, request)
    if user is None:
        raise HTTPException(status_code=401, detail="Not authenticated")

    vehicle = (await session.execute(
        select(_SV_pf).where(
            _SV_pf.id == vehicle_id,
            (
                ((_SV_pf.scope == "org") & (_SV_pf.owner_id == user.org_id)) |
                ((_SV_pf.scope == "user") & (_SV_pf.owner_id == user.id))
            ),
        )
    )).scalar_one_or_none()

    if vehicle is None:
        raise HTTPException(status_code=404, detail="Vehicle not found")

    owner = vehicle.scope

    source = vehicle.source_config or {}
    carry = vehicle.carry_config or {}
    exit_cfg = vehicle.exit_config or {}

    phases = carry.get("phases", [])
    constr = next((p for p in phases if p.get("name") == "construction"), {})
    oper = next((p for p in phases if p.get("name") == "operation"), {})

    exit_vehicle_raw = exit_cfg.get("vehicle") or exit_cfg.get("exit_vehicle")
    # Only pre-fill exit vehicle if it's a named sentinel (not a stale UUID from another deal)
    safe_exit_vehicle = exit_vehicle_raw if exit_vehicle_raw in ("maturity", "sale") else None

    # Prefer ORM columns (set by vehicle_form.html) over JSONB source_config (set by
    # settings inline forms) so both creation paths return consistent prefill data.
    _rate = (float(vehicle.interest_rate_pct) if vehicle.interest_rate_pct is not None
             else source.get("interest_rate_pct"))
    _amort = vehicle.amort_term_years or source.get("amort_term_years")
    _fallback_ct = vehicle.carry_type  # set by vehicle_form.html; absent in settings-created vehicles
    _constr_ct = constr.get("carry_type") or _fallback_ct
    _oper_ct = oper.get("carry_type") or _fallback_ct

    return JSONResponse({
        "vehicle_name": vehicle.name,
        "owner": owner,
        "vehicle_type": vehicle.vehicle_type,
        "equity_role": vehicle.equity_role,
        "source_interest_rate": _rate,
        "ltv_pct": source.get("ltv_pct"),
        "amort_term_years": _amort,
        "hold_term_years": source.get("hold_term_years"),
        "dscr_min": source.get("dscr_min"),
        "construction_carry_type": _constr_ct,
        "operation_carry_type": _oper_ct,
        "perm_rate_pct": oper.get("perm_rate_pct"),
        "perm_term_years": oper.get("perm_term_years"),
        "perm_conversion_trigger": oper.get("perm_conversion_trigger"),
        "exit_type": exit_cfg.get("exit_type"),
        "exit_vehicle": safe_exit_vehicle,
        "draw_every_n_months": source.get("draw_every_n_months"),
        "draw_active_from_milestone": source.get("draw_active_from_milestone"),
        "draw_active_from_offset_days": source.get("draw_active_from_offset_days"),
        "carry_schedule": carry.get("schedule"),
    })


# ---------------------------------------------------------------------------
# Draw Schedule module
# ---------------------------------------------------------------------------

def _milestone_label(key: str) -> str:
    labels = {
        "offer_made": "Offer Made",
        "under_contract": "Under Contract",
        "close": "Close",
        "pre_development": "Pre-Development",
        "construction": "Construction",
        "operation_lease_up": "Lease-Up",
        "operation_stabilized": "Stabilized",
        "divestment": "Divestment",
    }
    return labels.get(key, key.replace("_", " ").title())


async def _run_draw_schedule(
    session: AsyncSession,
    model_id: UUID,
    *,
    writeback: bool = False,
) -> "Any | None":
    """Run the draw schedule engine; optionally write computed amounts back to DB.

    All sources are auto-sized (total_commitment=None) so the engine determines
    each source's commitment from Uses + carry. Returns the DrawSchedule, or None
    if the engine cannot run (missing milestones / sources).
    """
    from app.engines.draw_schedule import (
        DealMilestone,
        DrawScheduleCalculator,
        DrawScheduleConfig,
        DrawScheduleInputs,
        SourceDef,
        UseLineItem,
    )
    from datetime import datetime as _dt_cls

    ctx = await _load_draw_schedule_ctx(session, model_id)
    if not ctx:
        return None

    milestones_dated = ctx["milestones_dated"]
    draw_sources_db   = ctx["draw_sources"]
    use_lines_db      = ctx["use_lines_db"]

    if not milestones_dated or not draw_sources_db:
        return None

    # ── Milestones ──────────────────────────────────────────────────────────
    engine_milestones = [
        DealMilestone(
            key=m["key"],
            label=m["label"],
            date=_dt_cls.combine(m["date"], _dt_cls.min.time()),
        )
        for m in milestones_dated
    ]

    # ── Use lines ───────────────────────────────────────────────────────────
    # NOTE: This map must cover both UseLinePhase enum values AND off-enum
    # strings that have been written to the DB by legacy or wizard paths
    # (notably ``operation_lease_up`` for the Lease-Up Reserve UseLine).
    # Anything missed here silently falls back to "close" and lands at
    # day 0 of the deal, breaking lease-up reserve timing.
    _phase_to_ms = {
        "acquisition": "close", "pre_construction": "pre_development",
        "construction": "construction", "renovation": "construction",
        "conversion": "construction", "operation": "operation_stabilized",
        "exit": "divestment", "other": "close",
        # Off-enum strings observed in production data:
        "pre_development": "pre_development",
        "lease_up": "operation_lease_up",
        "operation_lease_up": "operation_lease_up",
        "stabilized": "operation_stabilized",
        "operation_stabilized": "operation_stabilized",
        "divestment": "divestment",
    }
    _phase_to_cat = {
        "acquisition": "land", "pre_construction": "soft_costs",
        "construction": "hard_costs", "renovation": "hard_costs",
        "conversion": "hard_costs", "operation": "reserves",
        "exit": "fees", "other": "other",
        "pre_development": "soft_costs",
        "lease_up": "reserves", "operation_lease_up": "reserves",
        "stabilized": "reserves", "operation_stabilized": "reserves",
        "divestment": "fees",
    }
    _ms_keys_set = {m["key"] for m in milestones_dated}
    _ms_date_idx  = {m["key"]: m["date"] for m in milestones_dated}
    engine_uses: list[UseLineItem] = []
    for ul in use_lines_db:
        raw_phase = str(ul.phase or "").replace("UseLinePhase.", "")
        ms_key = _phase_to_ms.get(raw_phase, "close")
        if ms_key not in _ms_keys_set and _ms_keys_set:
            ms_key = next(iter(_ms_keys_set))
        raw_timing   = str(ul.timing_type).replace("UseLineTiming.", "")
        spread_months = 1
        spread_to_date = None
        if raw_timing in ("spread", "spread_across_range"):
            for i, m in enumerate(milestones_dated):
                if m["key"] == ms_key and i + 1 < len(milestones_dated):
                    nxt = milestones_dated[i + 1]["date"]
                    cur = m["date"]
                    diff_months = (nxt.year - cur.year) * 12 + (nxt.month - cur.month)
                    spread_months = max(1, diff_months)
                    break
        engine_uses.append(UseLineItem(
            key=str(ul.id), label=ul.label,
            category=_phase_to_cat.get(raw_phase, "other"),
            total_amount=Decimal(str(ul.amount)),
            milestone_key=ms_key, spread_months=spread_months, spread_to_date=spread_to_date,
        ))

    # ── Sources ────────────────────────────────────────────────────────────
    # Include all funded sources (amount > 0 on the underlying CapitalModule).
    # Debt auto-sizes (total_commitment=None). Grants/equity draw lump-sum
    # capped at the module's configured source.amount and route against
    # eligible_use_tags. Stack position drives routing priority.
    _last_real_ms  = milestones_dated[-1]["key"] if milestones_dated else "operation_stabilized"
    _real_ms_keys  = {m["key"] for m in milestones_dated}

    _cm_ids = [ds.capital_module_id for ds in draw_sources_db if ds.capital_module_id]
    _cm_by_id: dict = {}
    if _cm_ids:
        _cm_rows = list((await session.execute(
            select(CapitalModule).where(CapitalModule.id.in_(_cm_ids))
        )).scalars())
        _cm_by_id = {cm.id: cm for cm in _cm_rows}

    engine_sources: list[SourceDef] = []
    for ds in draw_sources_db:
        cm = _cm_by_id.get(ds.capital_module_id) if ds.capital_module_id else None
        cm_src = (cm.source or {}) if cm else {}
        cm_amount = Decimal(str(cm_src.get("amount") or 0)) if cm else Decimal("0")
        is_debt = ds.source_type == "debt"
        # Funded check: debt is permitted with no preset amount (auto-sizes);
        # grants/equity require a configured CapitalModule.source.amount > 0.
        if not is_debt and cm_amount <= 0:
            continue

        _to  = ds.active_to_milestone   if ds.active_to_milestone   in _real_ms_keys else _last_real_ms
        _frm = ds.active_from_milestone if ds.active_from_milestone in _real_ms_keys else (
            milestones_dated[0]["key"] if milestones_dated else _to
        )
        # Stack position + eligibility routing live on the CapitalModule.
        stack_pos = int(getattr(cm, "stack_position", 0) or 0) if cm else 0
        eligible_tags = list(getattr(cm, "eligible_use_tags", None) or [])
        # When the construction-phase carry_type is "interest_reserve",
        # the cashflow auto-sizer creates a "Capitalized Construction
        # Interest" UseLine that pre-funds the IR pool. The draw schedule
        # already funds that UseLine via the debt draw — capitalizing
        # additional carry into the loan principal would double-count
        # interest. Flag the SourceDef so _calc_source_draws skips the
        # self-referential carry term for these loans.
        _funded_carry = False
        if cm and is_debt:
            _carry_dict = cm.carry or {}
            for _ph in (_carry_dict.get("schedule") or []):
                if _ph.get("carry_type") == "interest_reserve":
                    _funded_carry = True
                    break
        # Map UseLine cost_category → engine UseLineItem category tag
        # (engine uses the same _phase_to_cat values defined above).
        engine_sources.append(SourceDef(
            key=str(ds.id), label=ds.label,
            source_type=ds.source_type,
            draw_every_n_months=ds.draw_every_n_months,
            annual_interest_rate=Decimal(str(ds.annual_interest_rate)),
            active_from_milestone=_frm, active_to_milestone=_to,
            active_from_offset_days=getattr(ds, "active_from_offset_days", 0) or 0,
            active_to_offset_days=getattr(ds, "active_to_offset_days", 0) or 0,
            # Debt auto-sizes; grants/equity cap at the configured amount.
            total_commitment=None if is_debt else cm_amount,
            # Non-exit-vehicle sources fund as a single lump-sum draw.
            single_draw=not is_debt,
            stack_position=stack_pos,
            eligible_use_categories=eligible_tags,
            funded_carry=_funded_carry,
        ))
    engine_sources.sort(key=lambda s: _ms_date_idx.get(s.active_from_milestone, _dt_cls.max))

    if not engine_sources:
        return None

    config = DrawScheduleConfig(
        min_reserve_construction=ctx["reserve_construction"],
        min_reserve_operational=ctx["reserve_operational"],
        operational_start_milestone="operation_lease_up",
    )
    try:
        schedule = DrawScheduleCalculator(DrawScheduleInputs(
            milestones=engine_milestones, uses=engine_uses,
            sources=engine_sources, config=config,
        )).calculate()
    except Exception:
        return None

    if writeback:
        _drawn_by_key = {ss.source_key: ss.total_drawn for ss in schedule.source_summaries}
        for ds in draw_sources_db:
            _drawn = _drawn_by_key.get(str(ds.id))
            if _drawn is None:
                continue
            ds.total_commitment = Decimal(str(_drawn))
            # Only debt sources have engine-computed amounts that should flow
            # back into CapitalModule.source["amount"]. Grants/equity are
            # user-configured caps — preserve them.
            if ds.source_type != "debt":
                continue
            if ds.capital_module_id:
                _cm = await session.get(CapitalModule, ds.capital_module_id)
                if _cm:
                    _src = dict(_cm.source or {})
                    _src["amount"] = float(_drawn)
                    _cm.source = _src
            else:
                _cm_q = select(CapitalModule).where(
                    CapitalModule.scenario_id == model_id,
                    CapitalModule.label == ds.label,
                ).limit(1)
                _cm = (await session.execute(_cm_q)).scalar_one_or_none()
                if _cm:
                    _src = dict(_cm.source or {})
                    _src["amount"] = float(_drawn)
                    _cm.source = _src
        await session.flush()

    return schedule


async def _load_draw_schedule_ctx(
    session: AsyncSession,
    model_id: UUID,
) -> dict[str, Any]:
    """Shared context for draw schedule panel and calculate endpoint."""
    from app.models.project import Project
    from app.models.milestone import Milestone

    model = await session.get(Scenario, model_id)
    if model is None:
        return {}

    # Load draw sources ordered by sort_order
    draw_sources = list((await session.execute(
        select(DrawSource)
        .where(DrawSource.scenario_id == model_id)
        .order_by(DrawSource.sort_order)
    )).scalars())

    # Load use lines (via Project) so we can pass them to the engine
    first_proj = (await session.execute(
        select(Project).where(Project.scenario_id == model_id).limit(1)
    )).scalar_one_or_none()

    use_lines_db: list = []
    project_milestones: list = []
    if first_proj:
        use_lines_db = list((await session.execute(
            select(UseLine).where(UseLine.project_id == first_proj.id)
        )).scalars())
        # Load milestones from both opportunity and project
        opp_ms = list((await session.execute(
            select(Milestone)
            .where(Milestone.opportunity_id == first_proj.opportunity_id)
            .order_by(Milestone.sequence_order)
        )).scalars()) if first_proj.opportunity_id else []
        proj_ms = list((await session.execute(
            select(Milestone)
            .where(Milestone.project_id == first_proj.id)
            .order_by(Milestone.sequence_order)
        )).scalars())
        project_milestones = opp_ms + proj_ms

    # Build milestone map for date resolution
    ms_map = {m.id: m for m in project_milestones}
    milestones_dated = []
    for m in project_milestones:
        start = m.computed_start(ms_map)
        if start:
            milestones_dated.append({
                "key": m.milestone_type.value if hasattr(m.milestone_type, "value") else str(m.milestone_type),
                "label": m.label or _milestone_label(str(m.milestone_type.value if hasattr(m.milestone_type, "value") else m.milestone_type)),
                "date": start,
            })

    # Sort milestones by date (opp + proj may interleave in unusual order)
    milestones_dated.sort(key=lambda m: m["date"])
    milestone_keys = [m["key"] for m in milestones_dated]

    # ---------------------------------------------------------------------------
    # Reconcile draw_sources against capital_modules.
    #
    # CapitalModule is the source-of-truth for capital stack configuration;
    # DrawSource is what the draw schedule engine actually iterates. The two
    # tables drift apart when modules are added/removed after the initial
    # auto-seed:
    #   1. Orphans: DrawSources with ``capital_module_id=NULL`` (or pointing
    #      to a deleted CapitalModule) leak into the engine output as
    #      phantom rows in the Sources Summary KPI tile.
    #   2. Missing: CapitalModules added after the first compute (e.g. a
    #      grant or equity source added via "+ Source") never get a matching
    #      DrawSource and silently never draw.
    #
    # Solution: every load reconciles. Cheap (one DB roundtrip for modules,
    # one delete batch, one insert batch). Idempotent — no-op on healthy
    # data. Mirrors the existing add/edit flows' DrawSource construction so
    # rows look identical regardless of code path.
    # ---------------------------------------------------------------------------
    capital_modules = list((await session.execute(
        select(CapitalModule)
        .where(CapitalModule.scenario_id == model_id)
        .order_by(CapitalModule.stack_position)
    )).scalars())
    _cm_id_set = {cm.id for cm in capital_modules}

    # Delete orphan DrawSources (no CapitalModule or pointing to a removed one).
    _orphans = [
        ds for ds in draw_sources
        if (ds.capital_module_id is None or ds.capital_module_id not in _cm_id_set)
    ]
    if _orphans:
        for ds in _orphans:
            await session.delete(ds)
        _orphan_ids = {ds.id for ds in _orphans}
        draw_sources = [ds for ds in draw_sources if ds.id not in _orphan_ids]

    # Identify CapitalModules with no matching DrawSource — these are the
    # "added after initial seed" rows that need backfilling.
    _cm_with_ds = {ds.capital_module_id for ds in draw_sources if ds.capital_module_id}
    _missing_cms = [cm for cm in capital_modules if cm.id not in _cm_with_ds]

    if _missing_cms or _orphans:
        # Map capital module phase strings → milestone keys (best-effort)
        _phase_to_ms = {
            "offer_made": "offer_made",
            "under_contract": "under_contract",
            "acquisition": "close",
            "pre_construction": "pre_development",
            "pre_development": "pre_development",
            "construction": "construction",
            "renovation": "construction",
            "lease_up": "operation_lease_up",
            "operation_lease_up": "operation_lease_up",
            "stabilized": "operation_stabilized",
            "operation_stabilized": "operation_stabilized",
            "divestment": "divestment",
        }
        _next_sort = max(
            (ds.sort_order or 0 for ds in draw_sources), default=0
        ) + 1
        for i, cm in enumerate(_missing_cms):
            raw_from = cm.active_phase_start or "close"
            raw_to = cm.active_phase_end or "operation_stabilized"
            ms_from = _phase_to_ms.get(raw_from, raw_from)
            ms_to = _phase_to_ms.get(raw_to, raw_to)
            # Fall back to first/last milestone if mapped key not in timeline
            if milestone_keys:
                if ms_from not in milestone_keys:
                    ms_from = milestone_keys[0]
                if ms_to not in milestone_keys:
                    ms_to = milestone_keys[-1]

            src = cm.source or {}
            rate_pct = src.get("interest_rate_pct") or 0.0
            annual_rate = Decimal(str(rate_pct)) / Decimal("100")

            source_type = "debt" if str(getattr(cm, "vehicle_type", "") or "").replace("VehicleType.", "") == "debt" else "equity"
            # Skip zero-amount stubs (unfunded equity placeholders) — only
            # seed sources that have a configured amount or auto-size as debt.
            if source_type != "debt" and not src.get("amount"):
                continue
            draw_freq = 2 if source_type == "debt" else 1

            ds = DrawSource(
                id=_uuid_mod.uuid4(),
                scenario_id=model_id,
                capital_module_id=cm.id,
                sort_order=_next_sort + i,
                label=cm.label,
                source_type=source_type,
                draw_every_n_months=draw_freq,
                annual_interest_rate=annual_rate,
                active_from_milestone=ms_from,
                active_to_milestone=ms_to,
                total_commitment=Decimal(str(src["amount"])) if src.get("amount") else None,
            )
            session.add(ds)

        # Re-load draw_sources after orphan delete + missing insert so
        # downstream code sees the canonical post-reconciliation set.
        await session.flush()
        draw_sources = list((await session.execute(
            select(DrawSource)
            .where(DrawSource.scenario_id == model_id)
            .order_by(DrawSource.sort_order)
        )).scalars())

    # ---------------------------------------------------------------------------
    # Auto-populate reserve floors from computed use lines when still unset
    # ---------------------------------------------------------------------------
    reserve_construction = Decimal(str(model.min_reserve_construction or 0))
    reserve_operational = Decimal(str(model.min_reserve_operational or 0))

    if (reserve_construction == 0 or reserve_operational == 0) and use_lines_db:
        for ul in use_lines_db:
            lbl = (ul.label or "").strip()
            amt = Decimal(str(ul.amount or 0))
            if reserve_construction == 0 and lbl == "Capitalized Construction Interest":
                reserve_construction = amt
            elif reserve_operational == 0 and lbl == "Operating Reserve":
                reserve_operational = amt

    # ---------------------------------------------------------------------------
    # Build source Gantt rows using the same g2- coordinate system as the
    # timeline Gantt.  builder_gantt_data has epoch/g_min/g_max exposed.
    # ---------------------------------------------------------------------------
    import datetime as _dt
    builder_gantt_data_ds = _builder_gantt_from_milestones(first_proj, project_milestones)
    source_gantt_rows: list[dict] = []
    if builder_gantt_data_ds and draw_sources:
        epoch_d = builder_gantt_data_ds.get("epoch")
        g_min_d = builder_gantt_data_ds.get("g_min", 0)
        g_max_d = builder_gantt_data_ds.get("g_max", 1)
        total_span = max(g_max_d - g_min_d, 1)

        def _day_pct(day_offset: int) -> float:
            return round(100.0 * (day_offset - g_min_d) / total_span, 2)

        ms_date_map = {m["key"]: m["date"] for m in milestones_dated}
        for ds in draw_sources:
            from_date = ms_date_map.get(ds.active_from_milestone)
            to_date = ms_date_map.get(ds.active_to_milestone)
            fade_right = False
            if from_date and epoch_d:
                from_day = (from_date - epoch_d).days
                left = max(0.0, _day_pct(from_day))
                if ds.active_to_milestone not in ms_date_map:
                    # pseudo-milestone (e.g. "maturity"): extend to Gantt right edge, fade out
                    right = 100.0
                    fade_right = True
                elif to_date:
                    to_day = (to_date - epoch_d).days
                    right = min(100.0, _day_pct(to_day))
                else:
                    continue
                source_gantt_rows.append({
                    "label": ds.label,
                    "source_type": ds.source_type,
                    "left_pct": left,
                    "width_pct": max(right - left, 1.5),
                    "fade_right": fade_right,
                })

    # Label map used by the panel to display current active window as text (not editable here)
    milestone_label_map = {m["key"]: m["label"] for m in milestones_dated}
    milestone_label_map["maturity"] = "Maturity"

    # Phase E draw events — populated by compute_cash_flows(); empty until then.
    from app.models.capital_draw_event import CapitalDrawEvent as _CDE_ds
    from app.models.project import Project as _Proj_ds
    _cde_rows = list((await session.execute(
        select(_CDE_ds)
        .where(_CDE_ds.scenario_id == model_id)
        .order_by(_CDE_ds.period, _CDE_ds.allocation_reason)
    )).scalars())
    # Collect project names for multi-project labelling
    _proj_ids = {str(r.project_id) for r in _cde_rows if r.project_id}
    _proj_name_map: dict[str, str] = {}
    if _proj_ids:
        _projs_ds = list((await session.execute(
            select(_Proj_ds).where(_Proj_ds.id.in_([_uuid_mod.UUID(p) for p in _proj_ids]))
        )).scalars())
        _proj_name_map = {str(p.id): p.name for p in _projs_ds}
    capital_draw_events = [
        {
            "period": r.period,
            "period_type": r.period_type or "",
            "allocation_reason": (r.allocation_reason.value if hasattr(r.allocation_reason, "value") else str(r.allocation_reason or "")),
            "amount": float(r.amount or 0),
            "project_id": str(r.project_id) if r.project_id else None,
            "project_name": _proj_name_map.get(str(r.project_id), "—") if r.project_id else "—",
            "use_line_label": r.use_line_label or "",
        }
        for r in _cde_rows
    ]

    return {
        "model": model,
        "draw_sources": draw_sources,
        "milestones_dated": milestones_dated,
        "milestone_label_map": milestone_label_map,
        "use_lines_db": use_lines_db,
        "reserve_construction": reserve_construction,
        "reserve_operational": reserve_operational,
        "milestone_keys": milestone_keys,
        "builder_gantt_data": builder_gantt_data_ds,
        "source_gantt_rows": source_gantt_rows,
        "capital_draw_events": capital_draw_events,
    }


@router.get("/ui/models/{model_id}/draw-schedule", response_class=HTMLResponse)
async def draw_schedule_panel(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Returns the draw schedule panel partial for HTMX swap."""
    ctx = await _load_draw_schedule_ctx(session, model_id)
    if not ctx:
        return HTMLResponse("<p class='text-muted'>Model not found.</p>", status_code=404)
    ctx["request"] = request
    ctx["active_module"] = "draw_schedule"
    return templates.TemplateResponse(request, "partials/draw_schedule_panel.html", ctx)


@router.post("/ui/models/{model_id}/draw-schedule/sources", response_class=HTMLResponse)
async def add_draw_source(
    request: Request,
    model_id: UUID,
    session: DBSession,
    label: str = Form(...),
    source_type: str = Form("equity"),
    draw_every_n_months: int = Form(1),
    annual_interest_rate: str = Form("0"),
    active_from_milestone: str = Form(...),
    active_to_milestone: str = Form(...),
    total_commitment: str = Form(""),
) -> HTMLResponse:
    """Add a draw source row."""
    model = await session.get(Scenario, model_id)
    if model is None:
        return HTMLResponse("Model not found", status_code=404)

    # Determine next sort_order
    max_order_row = (await session.execute(
        select(DrawSource.sort_order)
        .where(DrawSource.scenario_id == model_id)
        .order_by(DrawSource.sort_order.desc())
        .limit(1)
    )).scalar_one_or_none()
    next_order = (max_order_row or 0) + 1

    commitment = None
    if total_commitment.strip():
        try:
            commitment = Decimal(total_commitment.strip().replace(",", ""))
        except Exception:
            commitment = None

    ds = DrawSource(
        id=_uuid_mod.uuid4(),
        scenario_id=model_id,
        sort_order=next_order,
        label=label.strip(),
        source_type=source_type,
        draw_every_n_months=max(1, draw_every_n_months),
        annual_interest_rate=Decimal(annual_interest_rate.strip() or "0"),
        active_from_milestone=active_from_milestone,
        active_to_milestone=active_to_milestone,
        total_commitment=commitment,
    )
    session.add(ds)
    await session.flush()

    ctx = await _load_draw_schedule_ctx(session, model_id)
    ctx["request"] = request
    ctx["active_module"] = "draw_schedule"
    return templates.TemplateResponse(request, "partials/draw_schedule_panel.html", ctx)


@router.delete("/ui/models/{model_id}/draw-schedule/sources/{source_id}", response_class=HTMLResponse)
async def delete_draw_source(
    request: Request,
    model_id: UUID,
    source_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Delete a draw source row."""
    ds = await session.get(DrawSource, source_id)
    if ds and ds.scenario_id == model_id:
        await session.delete(ds)
        await session.flush()
    ctx = await _load_draw_schedule_ctx(session, model_id)
    ctx["request"] = request
    ctx["active_module"] = "draw_schedule"
    return templates.TemplateResponse(request, "partials/draw_schedule_panel.html", ctx)


@router.post("/ui/models/{model_id}/draw-schedule/settings", response_class=HTMLResponse)
async def update_draw_schedule_settings(
    request: Request,
    model_id: UUID,
    session: DBSession,
    min_reserve_construction: str = Form("0"),
    min_reserve_operational: str = Form("0"),
) -> HTMLResponse:
    """Update reserve floor settings on the scenario."""
    model = await session.get(Scenario, model_id)
    if model is None:
        return HTMLResponse("Model not found", status_code=404)

    def _parse_dec(val: str) -> Decimal:
        try:
            return Decimal(val.strip().replace(",", "") or "0")
        except Exception:
            return Decimal("0")

    model.min_reserve_construction = _parse_dec(min_reserve_construction)
    model.min_reserve_operational = _parse_dec(min_reserve_operational)
    await session.flush()

    ctx = await _load_draw_schedule_ctx(session, model_id)
    ctx["request"] = request
    ctx["active_module"] = "draw_schedule"
    return templates.TemplateResponse(request, "partials/draw_schedule_panel.html", ctx)


@router.post("/ui/models/{model_id}/draw-schedule/calculate", response_class=HTMLResponse)
async def calculate_draw_schedule(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Run the draw schedule engine and return the results HTML fragment."""
    ctx = await _load_draw_schedule_ctx(session, model_id)
    if not ctx:
        return HTMLResponse("Model not found", status_code=404)

    milestones_dated = ctx["milestones_dated"]
    draw_sources_db  = ctx["draw_sources"]

    if not milestones_dated:
        return HTMLResponse(
            "<div class='module-empty'><div class='module-empty-icon'>📅</div>"
            "<div class='module-empty-title'>No timeline yet</div>"
            "<div class='module-empty-desc'>Set up milestones in the Timeline module first.</div></div>"
        )
    if not draw_sources_db:
        return HTMLResponse(
            "<div class='module-empty'><div class='module-empty-icon'>💰</div>"
            "<div class='module-empty-title'>No sources defined</div>"
            "<div class='module-empty-desc'>Add at least one funding source above.</div></div>"
        )

    schedule = await _run_draw_schedule(session, model_id, writeback=True)
    if schedule is None:
        return HTMLResponse(
            "<div class='alert alert-danger' style='padding:12px;border-radius:6px;"
            "background:#fef2f2;border:1px solid #fca5a5;color:#dc2626;font-size:13px'>"
            "⚠ Engine error: check milestones and sources are configured.</div>"
        )

    # ── Detect unfunded uses ─────────────────────────────────────────────────
    from app.engines.draw_schedule import UseLineItem
    from datetime import datetime
    _ms_date_index   = {m["key"]: m["date"] for m in milestones_dated}
    _milestone_label_map = {m["key"]: m["label"] for m in milestones_dated}
    _milestone_label_map["maturity"] = "Maturity"
    _covered_ms_keys: set[str] = set()
    for ss in schedule.source_summaries:
        # Find the source's active window from ctx draw_sources_db
        _ds = next((d for d in draw_sources_db if str(d.id) == ss.source_key), None)
        if _ds:
            from_idx = next((i for i, m in enumerate(milestones_dated) if m["key"] == _ds.active_from_milestone), None)
            to_idx   = next((i for i, m in enumerate(milestones_dated) if m["key"] == _ds.active_to_milestone), None)
            if from_idx is not None and to_idx is not None:
                for i in range(from_idx, to_idx + 1):
                    _covered_ms_keys.add(milestones_dated[i]["key"])
    # Build use items for unfunded check
    # Keep in sync with the same-named map in `_run_draw_schedule` (around
    # line 14286). Both need to cover the off-enum strings written by
    # legacy paths (notably ``operation_lease_up``).
    _phase_to_ms = {
        "acquisition": "close", "pre_construction": "pre_development",
        "construction": "construction", "renovation": "construction",
        "conversion": "construction", "operation": "operation_stabilized",
        "exit": "divestment", "other": "close",
        "pre_development": "pre_development",
        "lease_up": "operation_lease_up",
        "operation_lease_up": "operation_lease_up",
        "stabilized": "operation_stabilized",
        "operation_stabilized": "operation_stabilized",
        "divestment": "divestment",
    }
    _ms_keys_set = {m["key"] for m in milestones_dated}
    unfunded_uses: list[dict] = []
    for ul in ctx["use_lines_db"]:
        raw_phase = str(ul.phase or "").replace("UseLinePhase.", "")
        ms_key = _phase_to_ms.get(raw_phase, "close")
        if ms_key not in _ms_keys_set and _ms_keys_set:
            ms_key = next(iter(_ms_keys_set))
        if ms_key not in _covered_ms_keys and (ul.amount or 0) > 0:
            unfunded_uses.append({
                "label": ul.label,
                "amount": ul.amount,
                "milestone_key": ms_key,
                "milestone_label": _milestone_label_map.get(ms_key, ms_key),
            })

    # Filter display: hide sources with no draws and $0-commitment sources from Gantt/table
    active_labels = {ss.source_label for ss in schedule.source_summaries if ss.total_drawn > 0}
    committed_labels = {
        ds.label for ds in draw_sources_db
        if ds.total_commitment and float(ds.total_commitment) > 0
    }
    show_labels = active_labels & committed_labels
    ctx["source_gantt_rows"] = [r for r in ctx.get("source_gantt_rows", []) if r["label"] in show_labels]
    ctx["draw_sources"] = [ds for ds in ctx["draw_sources"] if ds.label in show_labels]

    ctx["schedule"] = schedule
    ctx["unfunded_uses"] = unfunded_uses
    ctx["request"] = request
    ctx["active_module"] = "draw_schedule"
    # Return the full panel (not just results) so sources table always reflects current DB state
    return templates.TemplateResponse(request, "partials/draw_schedule_panel.html", ctx)



@router.patch("/ui/models/{model_id}/draw-schedule/sources/{source_id}", response_class=HTMLResponse)
async def update_draw_source_window(
    request: Request,
    model_id: UUID,
    source_id: UUID,
    session: DBSession,
    active_from_milestone: str = Form(...),
    active_to_milestone: str = Form(...),
) -> HTMLResponse:
    """Update the active window (from/to milestone) of a draw source."""
    ds = await session.get(DrawSource, source_id)
    if ds and ds.scenario_id == model_id:
        ds.active_from_milestone = active_from_milestone
        ds.active_to_milestone = active_to_milestone
        await session.flush()
    ctx = await _load_draw_schedule_ctx(session, model_id)
    ctx["request"] = request
    ctx["active_module"] = "draw_schedule"
    return templates.TemplateResponse(request, "partials/draw_schedule_panel.html", ctx)



# ── Deal change-log (history drawer) ─────────────────────────────────────────

@router.get("/ui/models/{model_id}/history", response_class=HTMLResponse)
async def history_drawer(
    model_id: UUID, request: Request, session: DBSession
) -> HTMLResponse:
    """Render the history drawer partial (list of compute snapshots)."""
    from app.exporters.snapshot import diff_snapshots, list_snapshots

    user = await _get_user(session, request)
    if user is None:
        return HTMLResponse('<p style="color:var(--text-muted)">Not authenticated</p>', status_code=401)

    model = await session.get(Scenario, model_id, options=[selectinload(Scenario.deal)])
    if model is None:
        return HTMLResponse('<p style="color:var(--text-muted)">Model not found</p>', status_code=404)
    if model.deal is None or model.deal.org_id != user.org_id:
        return HTMLResponse('<p style="color:var(--text-muted)">Forbidden</p>', status_code=403)

    snaps = await list_snapshots(session, model_id)

    # Build diff summaries between consecutive snapshots for rendering
    entries = []
    for i, snap in enumerate(snaps):
        diff = diff_snapshots(snaps[i - 1], snap) if i > 0 else None
        entries.append({"snap": snap, "diff": diff})

    return templates.TemplateResponse(
        request,
        "partials/history_drawer.html",
        {"request": request, "model": model, "entries": entries, "model_id": model_id},
    )


@router.post("/ui/models/{model_id}/history/{snapshot_id}/revert", response_class=HTMLResponse)
async def revert_snapshot(
    model_id: UUID,
    snapshot_id: UUID,
    request: Request,
    session: DBSession,
) -> HTMLResponse:
    """Revert scenario inputs to a prior snapshot state and return a status banner."""
    from app.exporters.snapshot import revert_to_snapshot

    user = await _get_user(session, request)
    if user is None:
        return HTMLResponse('<p style="color:var(--color-error)">Not authenticated</p>', status_code=401)

    model = await session.get(Scenario, model_id, options=[selectinload(Scenario.deal)])
    if model is None:
        return HTMLResponse('<p style="color:var(--color-error)">Model not found</p>', status_code=404)
    if model.deal is None or model.deal.org_id != user.org_id:
        return HTMLResponse('<p style="color:var(--color-error)">Forbidden</p>', status_code=403)

    try:
        await revert_to_snapshot(session, model_id, snapshot_id)
        await session.commit()
    except ValueError as exc:
        return HTMLResponse(
            f'<div class="alert alert-error" role="alert">{exc}</div>',
            status_code=404,
        )

    from starlette.responses import RedirectResponse
    return RedirectResponse(url=f"/models/{model_id}/builder", status_code=303)


@router.get("/ui/models/{model_id}/history/export.json")
async def export_history_json_endpoint(
    model_id: UUID, request: Request, session: DBSession
) -> JSONResponse:
    """Return the full change-log as a structured JSON (AI-readable)."""
    from app.exporters.snapshot import export_history_json
    import json as _json

    user = await _get_user(session, request)
    if user is None:
        return JSONResponse({"error": "not authenticated"}, status_code=401)

    model = await session.get(Scenario, model_id, options=[selectinload(Scenario.deal)])
    if model is None:
        return JSONResponse({"error": "model not found"}, status_code=404)
    if model.deal is None or model.deal.org_id != user.org_id:
        return JSONResponse({"error": "forbidden"}, status_code=403)

    try:
        payload = await export_history_json(session, model_id)
    except ValueError as exc:
        return JSONResponse({"error": str(exc)}, status_code=404)

    return JSONResponse(
        content=_json.loads(_json.dumps(payload, default=str)),
        headers={
            "Content-Disposition": f'attachment; filename="history-{model_id}.json"',
        },
    )


# ── Save as Template ──────────────────────────────────────────────────────────

@router.post("/ui/models/{model_id}/save-as-template", response_class=HTMLResponse)
async def save_as_template(
    request: Request,
    model_id: UUID,
    session: DBSession,
) -> HTMLResponse:
    """Extract a scenario's structure as a reusable scenario template."""
    from datetime import UTC, datetime
    from app.exporters.template_export import extract_template_json
    from app.models.deal import Deal, Scenario
    from app.models.scenario_template import ScenarioTemplate

    user = await _get_user(session, request)
    if user is None or user.org_id is None:
        return HTMLResponse("<p class='text-muted'>Not authenticated.</p>", status_code=401)

    scenario = await session.get(Scenario, model_id)
    if scenario is None:
        return HTMLResponse("<p class='text-muted'>Scenario not found.</p>", status_code=404)

    deal = await session.get(Deal, scenario.deal_id) if scenario.deal_id else None
    if settings.org_isolation_enabled:
        if deal is None or deal.org_id != user.org_id:
            return HTMLResponse("<p class='text-muted'>Scenario not found.</p>", status_code=404)

    form = await request.form()
    name = str(form.get("name", "")).strip()[:200]
    description = str(form.get("description", "")).strip()[:500] or None
    if not name:
        return HTMLResponse("<p class='text-muted'>Template name is required.</p>", status_code=400)

    template_json = await extract_template_json(session, model_id)

    tmpl = ScenarioTemplate(
        org_id=user.org_id,
        created_by_user_id=user.id,
        source_scenario_id=model_id,
        name=name,
        description=description,
        project_type=template_json.get("project_type"),
        template_json=template_json,
        created_at=datetime.now(UTC),
    )
    session.add(tmpl)
    await session.commit()

    return HTMLResponse(
        f'<div class="alert alert-success" style="margin:0">'
        f'Template "<strong>{name}</strong>" saved. '
        f'<a href="/settings/organization#scenario-templates">View in Settings →</a>'
        f'</div>'
    )


