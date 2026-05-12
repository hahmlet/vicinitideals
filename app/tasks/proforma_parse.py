"""Celery task: parse an uploaded pro forma / OM spreadsheet into
standardised revenue and OpEx line items via a locally-hosted Ollama LLM.

Flow
----
1. Caller (FastAPI route) stores raw file bytes in Redis under
   ``proforma:{task_id}:file`` (TTL 24 h), then queues this task.
2. Task reads the file, extracts the user-specified sheets, formats each
   as a plain-text table, and calls the local Ollama model twice (once for
   revenue, once for OpEx) via *instructor* for structured JSON output.
3. Progress is written to Redis under ``proforma:{task_id}:progress`` so
   the HTMX poller can display live step feedback.
4. Final result JSON is written to ``proforma:{task_id}:result`` (TTL 24 h).

Redis key schema
----------------
``proforma:{task_id}:file``      — raw xlsx bytes (bytes), TTL 24 h
``proforma:{task_id}:progress``  — JSON progress object, TTL 24 h
``proforma:{task_id}:result``    — JSON parse result, TTL 24 h
"""

from __future__ import annotations

import io
import json
import logging
from typing import Any

import openpyxl
from celery.utils.log import get_task_logger
from pydantic import BaseModel, Field

from app.config import settings
from app.models.deal import STANDARD_OPEX_CATEGORIES
from app.tasks.celery_app import celery_app

logger = get_task_logger(__name__)

PARSE_PROFORMA_TASK = "app.tasks.proforma_parse.parse_proforma"
_REDIS_TTL = 86_400  # 24 hours

# ---------------------------------------------------------------------------
# Structured output schemas (returned by the LLM via instructor)
# ---------------------------------------------------------------------------

class UnitTypeResult(BaseModel):
    name: str = Field(description="Descriptive unit type name, e.g. '1BR', 'Studio', 'Tower Small'")
    count: int = Field(description="Number of units of this type")
    avg_sqft: float = Field(description="Average square footage per unit")
    avg_monthly_rent: float = Field(description="Average gross monthly rent per unit in dollars")
    confidence: float = Field(ge=0.0, le=1.0, description="Confidence score 0–1")


class ParsedRevenue(BaseModel):
    unit_types: list[UnitTypeResult]


class ExpenseLineResult(BaseModel):
    original_label: str = Field(description="Exact label as it appeared in the spreadsheet")
    annual_amount: float = Field(description="Annual dollar amount (sum of monthly columns if source is monthly)")
    mapped_category: str | None = Field(
        default=None,
        description="One of the STANDARD_OPEX_CATEGORIES, or null if unable to map",
    )
    confidence: float = Field(ge=0.0, le=1.0, description="Confidence score 0–1 for the category mapping")
    is_operating_expense: bool = Field(
        description=(
            "True for real operating expenses. False for below-the-line items that should be "
            "excluded: debt service, interest, depreciation, loan fees, amortization of financing costs."
        )
    )


class ParsedExpenses(BaseModel):
    expense_lines: list[ExpenseLineResult]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _redis_client() -> Any:
    import redis  # type: ignore
    return redis.from_url(settings.redis_url, decode_responses=False)


def _set_progress(r: Any, task_id: str, step: int, total: int, message: str) -> None:
    payload = json.dumps({"step": step, "total": total, "message": message, "status": "running"})
    r.set(f"proforma:{task_id}:progress", payload, ex=_REDIS_TTL)


def _set_error(r: Any, task_id: str, message: str) -> None:
    payload = json.dumps({"status": "error", "message": message})
    r.set(f"proforma:{task_id}:progress", payload, ex=_REDIS_TTL)


def _set_done(r: Any, task_id: str) -> None:
    payload = json.dumps({"step": 3, "total": 3, "message": "Done", "status": "done"})
    r.set(f"proforma:{task_id}:progress", payload, ex=_REDIS_TTL)


_DEBT_KEYWORDS: frozenset[str] = frozenset({
    "depreciation", "interest", "amortization", "principal",
    "debt service", "mortgage", "loan fee", "income tax",
    "total income", "total expenses", "total other", "net income",
    "net operating", "subtotal", "total revenue", "utilities total",
    "total utilities", "total opex",
    "reserves", "capital reserve", "replacement reserve",
    "ar writeoff", "bad debt",
})

_EXCLUDE_ROW_PREFIXES: tuple[str, ...] = (
    "total", "subtotal", "net ", "grand total",
)


def _is_debt_or_total(label: str) -> bool:
    """Return True if the label looks like a below-the-line or subtotal row."""
    lower = label.lower().strip()
    for kw in _DEBT_KEYWORDS:
        if kw in lower:
            return True
    for prefix in _EXCLUDE_ROW_PREFIXES:
        if lower.startswith(prefix):
            return True
    return False


# Keyword → canonical category overrides.  Applied before fuzzy match so we
# catch verbose LLM outputs (e.g. "Maintenance and Repair", "Waste Management",
# "Advertising and Marketing") that share only one token with the canonical
# category name and so fall under the fuzzy threshold.  Each keyword is
# substring-matched against the lowercase raw label; first hit wins.
_CATEGORY_KEYWORDS: tuple[tuple[str, str], ...] = (
    ("electric",     "Utilities — Electric"),
    ("gas",          "Utilities — Gas"),
    ("water",        "Utilities — Water/Sewer"),
    ("sewer",        "Utilities — Water/Sewer"),
    ("trash",        "Utilities — Trash"),
    ("garbage",      "Utilities — Trash"),
    ("waste",        "Utilities — Trash"),
    ("turnover",     "Unit Turnover"),
    ("make ready",   "Unit Turnover"),
    ("make-ready",   "Unit Turnover"),
    ("repair",       "Repairs & Maintenance"),
    ("maint",        "Repairs & Maintenance"),
    ("landscap",     "Landscaping & Snow Removal"),
    ("snow",         "Landscaping & Snow Removal"),
    ("ground",       "Landscaping & Snow Removal"),
    ("advertis",     "Marketing & Leasing"),
    ("market",       "Marketing & Leasing"),
    ("leasing",      "Marketing & Leasing"),
    ("payroll",      "Payroll"),
    ("salary",       "Payroll"),
    ("wage",         "Payroll"),
    ("pest",         "Pest Control"),
    ("insurance",    "Insurance"),
    ("real estate tax", "Real Estate Taxes"),
    ("property tax", "Real Estate Taxes"),
    ("prop mgmt",    "Property Management"),
    ("property management", "Property Management"),
    ("professional management", "Property Management"),
    ("on-site mgmt", "Property Management"),
    ("off-site mgmt", "Property Management"),
    ("office",       "Administrative"),
    ("admin",        "Administrative"),
    ("phone",        "Bank/Software Fees"),
    ("internet",     "Bank/Software Fees"),
    ("software",     "Bank/Software Fees"),
    ("computer",     "Bank/Software Fees"),
    ("bank",         "Bank/Software Fees"),
    ("financing charge", "Bank/Software Fees"),
    ("tenant event", "Resident Services"),
    ("resident",     "Resident Services"),
    ("security",     "Security"),
    ("fire monitor", "Security"),
    ("compliance",   "Source Compliance"),
    ("lift monit",   "Source Compliance"),
    ("ohcs",         "Source Compliance"),
    ("hud monit",    "Source Compliance"),
    ("legal",        "Legal"),
    ("accounting",   "Legal"),
    ("audit",        "Legal"),
    ("cpa",          "Legal"),
    ("license",      "Legal"),
    ("professional fee", "Legal"),
    ("police",       "Jurisdiction Fees"),
    ("parks levy",   "Jurisdiction Fees"),
    ("municipal",    "Jurisdiction Fees"),
    ("jurisdiction", "Jurisdiction Fees"),
)


def _snap_category(raw: str | None) -> str | None:
    """Snap a raw LLM category string to the nearest STANDARD_OPEX_CATEGORIES entry.

    Tries three strategies in order:
    1. Exact match against the canonical list (already correct, no-op snap).
    2. Substring keyword override (handles verbose LLM phrasings like
       "Maintenance and Repair" → Repairs & Maintenance).
    3. Token-overlap fuzzy match with a 0.30 threshold (last resort).

    Returns None if nothing scores; the caller treats null as "needs human review".
    """
    if raw is None:
        return None

    raw_stripped = raw.strip()
    if not raw_stripped:
        return None

    # 1. Exact canonical match
    if raw_stripped in STANDARD_OPEX_CATEGORIES:
        return raw_stripped

    # 2. Keyword override (substring on lowercased raw)
    lower = raw_stripped.lower()
    for kw, cat in _CATEGORY_KEYWORDS:
        if kw in lower:
            return cat

    # 3. Token-overlap fuzzy fallback
    import re

    def _norm(s: str) -> set[str]:
        s = s.lower()
        s = re.sub(r"[^a-z0-9 ]", " ", s)
        return set(s.split())

    raw_tokens = _norm(raw_stripped)
    if not raw_tokens:
        return None

    best_score = 0.0
    best_cat: str | None = None
    for cat in STANDARD_OPEX_CATEGORIES:
        cat_tokens = _norm(cat)
        if not cat_tokens:
            continue
        overlap = len(raw_tokens & cat_tokens)
        score = overlap / max(len(raw_tokens | cat_tokens), 1)
        if score > best_score:
            best_score = score
            best_cat = cat

    return best_cat if best_score >= 0.30 else None


def _postprocess_expense_lines(lines: list[dict]) -> list[dict]:
    """Apply deterministic rules on top of the LLM output.

    - Snap the original spreadsheet label first (ground truth from the file);
      fall back to the LLM's mapped_category only when the label has no
      recognisable keyword.  This guards against LLM hallucinations like
      "Landscaping" → "Maintenance and Repair".
    - Force is_operating_expense=False for debt/total/reserve rows regardless of LLM
    - Round annual_amount to nearest whole dollar
    - Keep all rows (capture is essential even when unmappable)
    """
    out = []
    for line in lines:
        label = line.get("original_label", "")
        if _is_debt_or_total(label):
            line["is_operating_expense"] = False
        snapped = _snap_category(label) if label else None
        if snapped is None:
            snapped = _snap_category(line.get("mapped_category"))
        line["mapped_category"] = snapped
        try:
            line["annual_amount"] = round(float(line.get("annual_amount", 0)))
        except (TypeError, ValueError):
            line["annual_amount"] = 0
        out.append(line)
    return out


def _sheet_to_text(ws: Any, property_column: str | None = None, max_rows: int = 200) -> str:
    """Convert an openpyxl worksheet to a markdown table for the LLM.

    Column filtering strategy (applied in order):
    1. If *property_column* is set, keep label col + that property's column only.
    2. If the header row contains a TOTAL/Annual column, keep label col + that
       column only — strips out monthly columns that bloat the prompt and cause
       the LLM to drop rows.
    3. Otherwise keep all columns.

    Output is a GitHub-flavored markdown table so the LLM can parse structure
    reliably without counting tabs.
    """
    import re

    _TOTAL_HEADERS = {"total", "annual", "annual total", "ytd", "full year"}

    raw_rows: list[list[str]] = []
    for row in ws.iter_rows(max_row=min(ws.max_row, max_rows), values_only=True):
        if not any(c is not None for c in row):
            continue
        raw_rows.append([str(c).strip() if c is not None else "" for c in row])

    if not raw_rows:
        return ""

    header = raw_rows[0]

    # Determine which columns to keep
    keep: list[int] = []
    if property_column:
        keep = [0] + [
            idx for idx, h in enumerate(header)
            if property_column.lower() in h.lower()
        ]
    else:
        total_col = next(
            (idx for idx, h in enumerate(header)
             if re.sub(r"[^a-z]", "", h.lower()) in {re.sub(r"[^a-z]", "", t) for t in _TOTAL_HEADERS}),
            None,
        )
        if total_col is not None and total_col != 0:
            keep = [0, total_col]

    def _filter(cells: list[str]) -> list[str]:
        if keep:
            return [cells[i] for i in keep if i < len(cells)]
        return cells

    filtered = [_filter(r) for r in raw_rows]

    # Build markdown table
    col_count = max(len(r) for r in filtered)
    header_row = filtered[0] + [""] * (col_count - len(filtered[0]))
    sep = ["---"] * col_count
    lines = [
        "| " + " | ".join(header_row) + " |",
        "| " + " | ".join(sep) + " |",
    ]
    for row in filtered[1:]:
        padded = row + [""] * (col_count - len(row))
        lines.append("| " + " | ".join(padded) + " |")

    return "\n".join(lines)


def _markitdown_to_text(content: bytes, filename: str) -> str:
    """Convert any MarkitDown-supported file (PDF, DOCX, HTML, PPTX, etc.)
    to markdown text suitable for the LLM prompts.

    Writes bytes to a temp file because MarkitDown's public API is path-based.
    Returns empty string on failure so the caller can surface a clean error.
    """
    import os
    import tempfile
    try:
        from markitdown import MarkItDown  # type: ignore
    except ImportError:
        logger.error("markitdown not installed — cannot parse non-xlsx documents")
        return ""

    ext = os.path.splitext(filename)[1] or ".bin"
    tmp_path: str | None = None
    try:
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        md = MarkItDown()
        result = md.convert(tmp_path)
        return (getattr(result, "text_content", "") or "").strip()
    except Exception as exc:
        logger.warning("MarkitDown conversion failed: %s", exc)
        return ""
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass


def build_revenue_prompt(text: str) -> str:
    """Construct the LLM prompt for revenue / unit-mix extraction.

    Shared with scripts/dry_run_proforma_doc.py so dry runs and the
    production Celery task stay in sync.
    """
    return (
        "You are a real estate financial analyst. The following is tabular data from a "
        "spreadsheet's revenue or rent-roll sheet. Group the units into distinct unit types "
        "(e.g. Studio, 1BR, 2BR, or developer-named types like 'Tower Small'). "
        "For each type, return the count, average square footage, and average gross monthly rent. "
        "If the sheet is a unit-by-unit roll (individual unit numbers), group similar-sized units. "
        "Only return actual residential or commercial units — exclude parking, storage, laundry, "
        "and ancillary income rows.\n\n"
        f"SHEET DATA:\n{text}"
    )


def build_opex_prompt(text: str) -> str:
    """Construct the LLM prompt for OpEx extraction + category mapping.

    Shared with scripts/dry_run_proforma_doc.py.
    """
    categories_str = "\n".join(f"- {c}" for c in STANDARD_OPEX_CATEGORIES)
    return (
        "You are a real estate financial analyst. The following is tabular data from a "
        "spreadsheet's operating expense sheet.\n\n"
        "RULES — follow exactly:\n"
        "1. Return EVERY labeled expense row in the sheet, even if you cannot map it. "
        "Do not skip rows or merge rows together.\n"
        "2. For each row, extract the ANNUAL dollar amount for the WHOLE property. "
        "If monthly columns are shown, sum them to get the annual total. "
        "If an ANNUAL or TOTAL column is present, use it directly. "
        "CRITICAL: never use a Per-Unit, Per-Door, /Unit, $/Unit, or PER-UNIT column — those "
        "are already divided by the unit count and will be off by 10–100×. "
        "Per-Unit columns are typically the smaller numbers next to the annual total "
        "(e.g. Annual=$33,519, Per-Unit=$1,381 → use $33,519). "
        "If Current and Pro-Forma annual columns both exist, prefer the Pro-Forma annual.\n"
        "3. Map each label to EXACTLY one category from the STANDARD CATEGORIES list below. "
        "Copy the category name character-for-character including dashes and special characters. "
        "Set mapped_category=null ONLY if none of the standard categories fit at all.\n"
        "4. Set is_operating_expense=false for: debt service, mortgage interest, "
        "depreciation, loan fee amortization, principal payments, income tax — these are NOT OpEx. "
        "Also exclude subtotal/total rows and income rows.\n"
        "5. Set confidence 0.85–1.0 for obvious matches, 0.60–0.84 for reasonable guesses, "
        "below 0.60 when unsure.\n\n"
        f"STANDARD CATEGORIES (use exact spelling):\n{categories_str}\n\n"
        "Mapping hints:\n"
        "- 'LIFT Monitoring', 'OHCS', 'bond compliance', 'HUD monitoring' -> Source Compliance\n"
        "- 'Prop Mgmt', 'On-Site Mgmt', 'Off-Site Mgmt', 'Professional Management' -> Property Management\n"
        "- 'RE Taxes', 'Real Estate Tax', 'Property Tax' -> Real Estate Taxes\n"
        "- 'Gresham Police Fire Parks', 'Municipal assessment', 'Police', 'Fire', 'Parks levy' -> Jurisdiction Fees\n"
        "- 'Accounting', 'CPA', 'Audit', 'Professional Fees', 'Legal', 'Licenses' -> Legal\n"
        "- 'Bank', 'NSF', 'Financing charges', 'Computer', 'Software', 'Internet', 'Phone/Internet' -> Bank/Software Fees\n"
        "- 'Office Supplies', 'Office/Admin', 'Administrative' -> Administrative\n"
        "- 'Tenant Events', 'Resident Activities' -> Resident Services\n"
        "- 'Fire Monitoring', 'Security System' -> Security\n"
        "- 'Garbage', 'Trash', 'Waste' -> Utilities — Trash\n"
        "- 'Electricity', 'Electric' -> Utilities — Electric\n"
        "- 'Gas', 'Natural Gas' -> Utilities — Gas\n"
        "- 'Water', 'Sewer', 'Water/Sewer' -> Utilities — Water/Sewer\n"
        "- 'Advertising', 'Marketing', 'Promotion' -> Marketing & Leasing\n"
        "- 'Landscaping', 'Snow Removal', 'Grounds' -> Landscaping & Snow Removal\n"
        "- 'Maint/Repair', 'Maintenance', 'Repairs' -> Repairs & Maintenance\n"
        "- 'Payroll', 'Wages', 'Salaries' -> Payroll\n"
        "- 'Turnover', 'Make-ready', 'Unit Turn' -> Unit Turnover\n"
        "- 'Reserves', 'Capital Reserves', 'Replacement Reserves' -> set is_operating_expense=false\n"
        "- 'AR Writeoffs', 'Bad Debt', 'Receivables' -> set is_operating_expense=false\n\n"
        f"SHEET DATA:\n{text}"
    )


def _llm_client() -> Any:
    """Return an instructor-patched OpenAI client pointed at local Ollama."""
    import instructor  # type: ignore
    from openai import OpenAI  # type: ignore

    raw = OpenAI(base_url=settings.ollama_base_url, api_key="ollama")
    return instructor.from_openai(raw, mode=instructor.Mode.JSON)


# ---------------------------------------------------------------------------
# Celery task
# ---------------------------------------------------------------------------

@celery_app.task(bind=True, name=PARSE_PROFORMA_TASK)
def parse_proforma(
    self,
    task_id: str,
    model_id: str,  # noqa: ARG001  — reserved for future per-model context
    revenue_sheet: str,
    opex_sheet: str,
    property_column: str | None = None,
    file_kind: str = "xlsx",
) -> None:
    """Parse a pro forma file and write structured results to Redis.

    file_kind="xlsx": user has picked specific sheets via the preflight UI.
    file_kind="doc": PDF / DOCX / HTML / etc. — MarkitDown converts the
    whole document and the resulting markdown is fed to both LLM passes.
    """
    r = _redis_client()
    warnings: list[str] = []

    try:
        # ------------------------------------------------------------------
        # Step 1 — Read file from Redis
        # ------------------------------------------------------------------
        _set_progress(r, task_id, 1, 3, "Reading document…")
        file_bytes: bytes | None = r.get(f"proforma:{task_id}:file")
        if not file_bytes:
            _set_error(r, task_id, "Upload expired or not found. Please upload the file again.")
            return

        if file_kind == "doc":
            filename_raw = r.get(f"proforma:{task_id}:filename") or b"document"
            filename = filename_raw.decode() if isinstance(filename_raw, bytes) else str(filename_raw)
            doc_markdown = _markitdown_to_text(file_bytes, filename)
            if not doc_markdown:
                _set_error(r, task_id, "Could not extract text from this file. Try a PDF, DOCX, or XLSX.")
                return
            revenue_text = doc_markdown
            opex_text = doc_markdown
        else:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

            if revenue_sheet not in wb.sheetnames:
                warnings.append(f"Revenue sheet '{revenue_sheet}' not found — skipping revenue parse.")
                revenue_text = ""
            else:
                revenue_text = _sheet_to_text(wb[revenue_sheet])

            if opex_sheet not in wb.sheetnames:
                warnings.append(f"OpEx sheet '{opex_sheet}' not found — skipping expense parse.")
                opex_text = ""
            else:
                opex_text = _sheet_to_text(wb[opex_sheet], property_column=property_column)

        client = _llm_client()

        # ------------------------------------------------------------------
        # Step 2 — Parse revenue
        # ------------------------------------------------------------------
        _set_progress(r, task_id, 2, 3, "Parsing revenue / unit mix…")
        unit_types: list[dict] = []

        if revenue_text:
            try:
                rev_prompt = build_revenue_prompt(revenue_text)
                parsed_rev: ParsedRevenue = client.chat.completions.create(
                    model=settings.ollama_model,
                    response_model=ParsedRevenue,
                    messages=[{"role": "user", "content": rev_prompt}],
                )
                unit_types = [u.model_dump() for u in parsed_rev.unit_types]
            except Exception as exc:
                logger.warning("Revenue parse failed: %s", exc)
                warnings.append(f"Revenue parsing failed: {exc}")

        # ------------------------------------------------------------------
        # Step 3 — Parse OpEx
        # ------------------------------------------------------------------
        _set_progress(r, task_id, 3, 3, "Mapping expense categories…")
        expense_lines: list[dict] = []

        if opex_text:
            try:
                opex_prompt = build_opex_prompt(opex_text)
                parsed_exp: ParsedExpenses = client.chat.completions.create(
                    model=settings.ollama_model,
                    response_model=ParsedExpenses,
                    messages=[{"role": "user", "content": opex_prompt}],
                )
                expense_lines = _postprocess_expense_lines(
                    [e.model_dump() for e in parsed_exp.expense_lines]
                )
            except Exception as exc:
                logger.warning("OpEx parse failed: %s", exc)
                warnings.append(f"OpEx parsing failed: {exc}")

        # ------------------------------------------------------------------
        # Write result
        # ------------------------------------------------------------------
        result = {
            "unit_types": unit_types,
            "expense_lines": expense_lines,
            "warnings": warnings,
        }
        r.set(f"proforma:{task_id}:result", json.dumps(result), ex=_REDIS_TTL)
        _set_done(r, task_id)

    except Exception as exc:
        logger.exception("proforma_parse task failed: %s", exc)
        _set_error(r, task_id, f"Unexpected error: {exc}")
