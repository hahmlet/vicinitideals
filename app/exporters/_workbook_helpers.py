"""Shared workbook helpers for the investor Excel export.

Provides:

- ``CellRegistry`` — tracks every investor-meaningful cell so the workbook
  emits workbook-scoped, absolute defined names. Phase 2 formulas read by
  name (``=s_cap_rate * s_year_10_noi``) without needing sheet prefixes.
- ``BRAND`` — palette + typography constants the investor sheets share.
- Format constants (``ACCOUNTING``, ``PCT``, ``MULTIPLE``, etc.) + style
  helpers (``header_row``, ``section_label``, ``kv_row``, ``freeze_top``,
  ``set_widths``).

Naming convention enforced by ``CellRegistry``: see plan §4 in
``docs/feature-plans/investor-excel-export-v2.md``.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from enum import Enum
from typing import Any
from uuid import UUID

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

# ── Palette + typography ──────────────────────────────────────────────────────

BRAND: dict[str, str] = {
    "navy": "0D1B2A",
    "slate": "415A77",
    "mist": "778DA9",
    "fog": "E0E1DD",
    "gold": "C9A96E",
    "ink": "1B1B1B",
    "paper": "FFFFFF",
    "rule": "D0D4DA",
    # Input/output color convention — industry standard from
    # `docs/Best Practice Synthesis/03_modeling_conventions.md` / `07_ux_*.md`:
    # blue = user-editable input, black = calc, green = cross-sheet link.
    # Picked these specific shades to be legible in print + e-ink readers
    # without crossing over into the gold accent.
    "input_blue": "0B5394",
    "link_green": "1E7E34",
}

FONT_TITLE = Font(name="Calibri", size=18, bold=True, color=BRAND["navy"])
FONT_SUBTITLE = Font(name="Calibri", size=11, italic=True, color=BRAND["slate"])
FONT_SECTION = Font(name="Calibri", size=12, bold=True, color=BRAND["paper"])
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color=BRAND["paper"])
FONT_LABEL = Font(name="Calibri", size=10, bold=True, color=BRAND["ink"])
FONT_VALUE = Font(name="Calibri", size=10, color=BRAND["ink"])
FONT_HERO_VALUE = Font(name="Calibri", size=12, bold=True, color=BRAND["gold"])
FONT_HINT = Font(name="Calibri", size=9, italic=True, color=BRAND["mist"])
# Input cell — blue. Surfaces on every Assumption row so the LP can tell at
# a glance which numbers are user inputs vs derived calculations.
FONT_INPUT = Font(name="Calibri", size=10, color=BRAND["input_blue"])
# Cross-sheet link — green. Used for HYPERLINK formulas (Glossary refs,
# Per-Project Mini-Summary navigation, per-project sheet back/forward links).
FONT_LINK = Font(name="Calibri", size=10, color=BRAND["link_green"], underline="single")

FILL_SECTION = PatternFill("solid", fgColor=BRAND["navy"])
FILL_HEADER = PatternFill("solid", fgColor=BRAND["slate"])
FILL_HERO = PatternFill("solid", fgColor=BRAND["fog"])
FILL_RAG_GREEN  = PatternFill("solid", fgColor="C8E6C9")
FILL_RAG_YELLOW = PatternFill("solid", fgColor="FFF9C4")
FILL_RAG_RED    = PatternFill("solid", fgColor="FFCDD2")

ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

_RULE = Side(style="thin", color=BRAND["rule"])
THIN_BORDER = Border(left=_RULE, right=_RULE, top=_RULE, bottom=_RULE)
BOTTOM_RULE = Border(bottom=_RULE)

# ── Number formats ────────────────────────────────────────────────────────────

ACCOUNTING = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
ACCOUNTING_2 = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
INT_COMMA = "#,##0"
PCT = "0.00%"
PCT_1 = "0.0%"
MULTIPLE = "0.00\\x"
DATE_FMT = "yyyy-mm-dd"


# ── CellRegistry ──────────────────────────────────────────────────────────────


@dataclass(frozen=True)
class CellAddress:
    """Resolved address for a defined name. Single-cell when end_* is None."""

    sheet: str
    row: int  # 1-based
    col: int  # 1-based
    end_row: int | None = None
    end_col: int | None = None

    def to_ref(self) -> str:
        col_a = get_column_letter(self.col)
        col_b = get_column_letter(self.end_col or self.col)
        if self.end_row is None and self.end_col is None:
            return f"'{self.sheet}'!${col_a}${self.row}"
        return f"'{self.sheet}'!${col_a}${self.row}:${col_b}${self.end_row or self.row}"


@dataclass
class CellRegistry:
    """Workbook-scoped defined-name registry.

    Discipline: every investor-meaningful cell registers a name. Pure
    presentation cells (section dividers, headers) don't. ``emit(wb)``
    writes them all as ``DefinedName`` rows on the workbook so Phase 2
    formulas can reference them by name.
    """

    _names: dict[str, CellAddress] = field(default_factory=dict)

    def __contains__(self, name: str) -> bool:
        return name in self._names

    def __len__(self) -> int:
        return len(self._names)

    def names(self) -> set[str]:
        return set(self._names)

    def get(self, name: str) -> CellAddress | None:
        return self._names.get(name)

    def register(self, name: str, sheet: str, row: int, col: int) -> None:
        if name in self._names:
            raise ValueError(
                f"defined name {name!r} already registered to {self._names[name]}"
            )
        self._names[name] = CellAddress(sheet=sheet, row=row, col=col)

    def register_range(
        self, name: str, sheet: str, top_row: int, bottom_row: int, col: int,
        *, end_col: int | None = None,
    ) -> None:
        if name in self._names:
            raise ValueError(
                f"defined name {name!r} already registered to {self._names[name]}"
            )
        self._names[name] = CellAddress(
            sheet=sheet, row=top_row, col=col,
            end_row=bottom_row, end_col=end_col or col,
        )

    def write(
        self,
        ws: Worksheet,
        row: int,
        col: int,
        value: Any,
        *,
        name: str | None = None,
        fmt: str | None = None,
        font: Font | None = None,
        fill: PatternFill | None = None,
        align: Alignment | None = None,
        border: Border | None = None,
    ) -> None:
        """Write a coerced value at (row, col) and optionally register a name."""
        cell = ws.cell(row=row, column=col, value=to_excel_value(value))
        if fmt is not None:
            cell.number_format = fmt
        if font is not None:
            cell.font = font
        if fill is not None:
            cell.fill = fill
        if align is not None:
            cell.alignment = align
        if border is not None:
            cell.border = border
        if name is not None:
            self.register(name, ws.title, row, col)

    def emit(self, wb: Workbook) -> None:
        """Write every registered name to the workbook's DefinedNameDict."""
        for name, addr in self._names.items():
            wb.defined_names[name] = DefinedName(name=name, attr_text=addr.to_ref())


# ── Value coercion ────────────────────────────────────────────────────────────


def to_excel_value(value: Any) -> Any:
    """Coerce a Python value into something openpyxl will accept cleanly.

    Mirrors the round-trip exporter's ``_to_v`` (excel_export.py:588) so
    behaviour is consistent across the two exporters during the rollout.
    Returns "" for None so empty cells render as blanks rather than the
    literal string "None".
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, Decimal):
        f = float(value)
        return int(f) if f == int(f) else round(f, 6)
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Enum):
        return value.value
    if isinstance(value, UUID):
        return str(value)
    return value


# ── Sheet builder helpers ─────────────────────────────────────────────────────


def section_label(ws: Worksheet, row: int, text: str, *, span_cols: int = 6) -> None:
    """Bold navy banner row spanning the given columns."""
    ws.cell(row=row, column=1, value=text)
    cell = ws.cell(row=row, column=1)
    cell.font = FONT_SECTION
    cell.fill = FILL_SECTION
    cell.alignment = ALIGN_LEFT
    if span_cols > 1:
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=span_cols
        )
        for c in range(2, span_cols + 1):
            ws.cell(row=row, column=c).fill = FILL_SECTION
    ws.row_dimensions[row].height = 22


def header_row(ws: Worksheet, row: int, columns: list[str]) -> None:
    """Slate-fill header strip with white bold text for table headers."""
    for idx, label in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=idx, value=label)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 18


def kv_row(
    ws: Worksheet,
    row: int,
    key: str,
    value: Any,
    *,
    name: str | None = None,
    registry: CellRegistry | None = None,
    fmt: str | None = None,
    hero: bool = False,
    style: str = "calc",
) -> None:
    """Two-column key/value row at (row, 1) and (row, 2).

    ``style`` selects the value-cell font per the input/output color
    convention from ``docs/Best Practice Synthesis/03_modeling_conventions.md``:
    ``"calc"`` = black (default; derived values), ``"input"`` = blue
    (user-editable assumptions), ``"link"`` = green (cross-sheet links;
    typically paired with a ``=HYPERLINK(...)`` value). ``hero=True``
    overrides ``style`` for the gold-accent KPI rows on Underwriting
    Summary's Primary KPIs block, since those are visually-loaded and
    don't need additional color signaling.
    """
    ws.cell(row=row, column=1, value=key).font = FONT_LABEL
    ws.cell(row=row, column=1).alignment = ALIGN_LEFT
    value_font = (
        FONT_HERO_VALUE if hero
        else FONT_INPUT if style == "input"
        else FONT_LINK if style == "link"
        else FONT_VALUE
    )
    if registry is not None:
        registry.write(
            ws,
            row,
            2,
            value,
            name=name,
            fmt=fmt,
            font=value_font,
            fill=FILL_HERO if hero else None,
            align=ALIGN_RIGHT,
        )
    else:
        cell = ws.cell(row=row, column=2, value=to_excel_value(value))
        cell.font = value_font
        cell.alignment = ALIGN_RIGHT
        if fmt:
            cell.number_format = fmt
        if hero:
            cell.fill = FILL_HERO


def freeze_top(ws: Worksheet, row: int = 2) -> None:
    """Freeze the first ``row - 1`` rows (default 1)."""
    ws.freeze_panes = ws.cell(row=row, column=1)


def set_widths(ws: Worksheet, widths: list[int]) -> None:
    """Set column widths positionally (1-indexed via list order)."""
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def print_landscape(ws: Worksheet) -> None:
    """Configure landscape orientation + fit-to-1-page-wide for summary sheets."""
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def hyperlink_to_sheet(target_sheet: str, label: str, cell: str = "A1") -> str:
    """Build an in-workbook hyperlink formula string.

    Use as a cell ``value`` (cell.hyperlink + cell.value would also work but
    HYPERLINK formula keeps the link logic visible to anyone inspecting the
    workbook).
    """
    return f'=HYPERLINK("#\'{target_sheet}\'!{cell}", "{label}")'
