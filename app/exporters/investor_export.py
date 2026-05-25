"""Investor-ready Excel export for a Scenario.

Generates a single ``.xlsx`` workbook per Scenario formatted for an LP /
lender / sponsor audience. See ``docs/feature-plans/investor-excel-export-v2.md``
for the full design (sheet order, named-range convention, doc-driven
glossary, build sequencing).

**Build status.** Commit 1 of the build sequence ships the audit-spine
sheets (Cover, Assumptions, Glossary). Commits 2/3 add the
underwriting-rollup sheets (Underwriting Summary, Pro Forma, Cash Flow,
Investor Returns) and the per-project sheets respectively. Sheet order
on disk grows toward the §2 final order as those commits land.

**Why this exists alongside ``excel_export.py``.** The round-trip exporter
is deprecated (see its docstring); it served the importer round-trip use
case. This module is the LP-facing artifact and is not intended to be
re-imported.
"""
from __future__ import annotations

import re
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.exporters._doc_validator import MetricEntry, parse_doc
from app.exporters._workbook_helpers import (
    ACCOUNTING,
    ALIGN_LEFT,
    ALIGN_RIGHT,
    ALIGN_WRAP,
    BRAND,
    DATE_FMT,
    FILL_HERO,
    FILL_RAG_GREEN,
    FILL_RAG_RED,
    FILL_RAG_YELLOW,
    FONT_HERO_VALUE,
    FONT_HINT,
    FONT_INPUT,
    FONT_LABEL,
    FONT_LINK,
    FONT_SUBTITLE,
    FONT_TITLE,
    FONT_VALUE,
    INT_COMMA,
    PCT,
    PCT_1,
    THIN_BORDER,
    CellRegistry,
    freeze_top,
    header_row,
    kv_row,
    print_landscape,
    section_label,
    set_widths,
)
from app.engines.phase_plan import (
    build_project_phase_windows,
    perm_origination_month,
    total_horizon_months,
)
from app.engines.sensitivity_matrix import compute_sensitivity_matrix
from app.engines.underwriting_rollup import (
    rollup_summary,
    rollup_waterfall,
)
from app.models.capital import (
    CapitalModule,
    CapitalModuleProject,
    WaterfallResult,
    WaterfallTier,
)
from app.models.cashflow import CashFlow, CashFlowLineItem, OperationalOutputs
from app.models.milestone import Milestone
from app.models.deal import (
    ALWAYS_SHOWN_OPEX_CATEGORIES,
    USE_CATEGORY_LABELS,
    USE_COST_CATEGORIES,
    Deal,
    DealModel,
    IncomeStream,
    OperatingExpenseLine,
    OperationalInputs,
    UseLine,
    normalize_opex_label,
)


class _UMProxy:
    """Attribute-compatible proxy for unit_mix JSONB dicts.

    Replaces the old UnitMix ORM rows; unknown attributes return None.
    """
    def __init__(self, d: dict) -> None:
        self.__dict__.update(d)

    def __getattr__(self, k: str):
        return None
from app.models.org import Organization
from app.models.project import Project
from app.config import settings as _app_settings

# Hard cap from the plan; enforced upstream so we never need 3-digit ordinals.
MAX_PROJECTS_PER_SCENARIO = 5

# Sheet-name prefix is exactly 4 chars (`P` + 1- or 2-digit ordinal + space),
# leaving 27 chars for the project name within Excel's 31-char ceiling.
PROJECT_SHEET_NAME_BUDGET = 27


# ── Public entry point ────────────────────────────────────────────────────────


async def export_investor_workbook(
    deal_model_id: UUID,
    session: AsyncSession,
    profile: str = "internal",
) -> bytes:
    """Build the investor workbook for a Scenario and return the bytes.

    ``profile`` controls which sheets are rendered:
    - ``"internal"``  — all sheets (full underwriting working model)
    - ``"lp"``        — LP / investor package (IRR, waterfall, sensitivity)
    - ``"lender"``    — lender package (DSCR, debt schedule, NOI)
    - ``"proforma"``  — pro forma only (NOI build, S&U, per-project)

    Raises ``ValueError`` if the Scenario doesn't exist.
    """
    _profile = profile if profile in {"internal", "lp", "lender", "proforma"} else "internal"
    ctx = await _load_all(session, deal_model_id)
    if ctx is None:
        raise ValueError(f"Scenario {deal_model_id} was not found")
    ctx["_profile"] = _profile

    wb = Workbook()
    registry = CellRegistry()

    # ── Profile membership helpers ─────────────────────────────────────────────
    # Each set lists which profiles render that sheet.  Adding a new profile
    # only requires updating the relevant set(s) below.
    _HAS_UW        = {"internal", "lp", "lender"}   # UW Summary + Pro Forma + Cash Flow
    _HAS_SU        = {"internal", "lp", "lender", "proforma"}  # Sources & Uses dedicated sheet
    _HAS_RETURNS   = {"internal", "lp"}              # Investor Returns + Waterfall
    _HAS_UNIT_MIX  = {"internal", "lp", "lender"}
    _HAS_DEBT      = {"internal", "lender"}          # Debt Schedule
    # Every profile that renders S&U needs Assumptions too: S&U Sources
    # rows emit ``=s_<slug>_principal`` formulas and the new Operating
    # Reserve UseLine emits ``=s_operating_reserve_months*...``. Those
    # defined names only get registered inside _build_assumptions
    # (Block C + Block A). Without the sheet the formulas become
    # dangling references that show #NAME? in Excel.
    _HAS_ASSUMPT   = {"internal", "lp", "lender", "proforma"}  # Assumptions
    _HAS_SENS      = {"internal", "lp"}              # Sensitivity (slow; skip for lender/proforma)
    _HAS_PF        = {"proforma"}                    # New formula-driven Pro Forma sheets

    cover = wb.active
    cover.title = "Cover"
    _build_cover(cover, registry, ctx)

    if _profile in _HAS_UW:
        uw_summary = wb.create_sheet("Underwriting Summary")
        _build_uw_summary(uw_summary, registry, ctx)

        uw_proforma = wb.create_sheet("Underwriting Pro Forma")
        _build_uw_proforma(uw_proforma, registry, ctx)

        uw_cashflow = wb.create_sheet("Underwriting Cash Flow")
        _build_uw_cashflow(uw_cashflow, registry, ctx)

    if _profile in _HAS_SU:
        su_sheet = wb.create_sheet("Sources & Uses")
        _build_su_sheet(su_sheet, registry, ctx)

    if _profile in _HAS_RETURNS:
        investor_returns = wb.create_sheet("Investor Returns")
        _build_investor_returns(investor_returns, registry, ctx)

        waterfall_ws = wb.create_sheet("Waterfall")
        _build_waterfall_sheet(waterfall_ws, registry, ctx)

    if _profile in _HAS_UNIT_MIX:
        unit_mix_ws = wb.create_sheet("Unit Mix")
        _build_unit_mix_sheet(unit_mix_ws, registry, ctx)

    if _profile in _HAS_DEBT:
        debt_schedule = wb.create_sheet("Debt Schedule")
        _build_debt_schedule(debt_schedule, registry, ctx)

    if _profile in _HAS_ASSUMPT:
        assumptions = wb.create_sheet("Assumptions")
        _build_assumptions(assumptions, registry, ctx)

    # Pro Forma profile gets the formula-driven combined + per-project sheets.
    if _profile in _HAS_PF:
        pf_combined = wb.create_sheet("Pro Forma")
        _build_proforma_combined(pf_combined, registry, ctx)

    # Per-project sheets: all profiles. Sheet names `P{n} {Name}` truncated
    # to Excel's 31-char ceiling. UW Summary hyperlinks resolve once these
    # exist.
    #
    # Formula-conversion plan §5 (commit 8): single-project consolidation.
    # When a scenario has exactly one project, the per-project sheet
    # duplicates content already on Underwriting Pro Forma / Cash Flow
    # (combined == per-project when there's one project). Skip the P1
    # sheet to reduce noise; UW Summary's mini-table HYPERLINK then
    # points at Underwriting Pro Forma (or the formula-driven Pro Forma
    # for the PF profile) instead of a P1 sheet that no longer exists.
    projects: list[Project] = ctx["projects"]
    _single_project = len(projects) == 1
    for idx, project in enumerate(projects, start=1):
        if _single_project:
            continue
        if _profile in _HAS_PF:
            sheet_name = _project_sheet_name(idx, project.name)
            ws_proj = wb.create_sheet(sheet_name)
            _build_proforma_project_sheet(ws_proj, registry, ctx, idx, project)
        else:
            sheet_name = _project_sheet_name(idx, project.name)
            ws_proj = wb.create_sheet(sheet_name)
            _build_project_sheet(ws_proj, registry, ctx, idx, project)

    # Collect every string written to the workbook so far so _build_glossary
    # can filter to only terms that actually appear in this export.
    _written_strings: set[str] = {
        cell.value
        for sheet in wb.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value
    }
    glossary = wb.create_sheet("Glossary & Methodology")
    _build_glossary(glossary, registry, ctx, written_strings=_written_strings)

    # Sensitivity is built LAST — compute_sensitivity_matrix calls
    # session.expire_all() each cycle, invalidating ORM rows in ctx.
    # Splice it into position 4 after building so tab order is correct.
    # Skip for lender/proforma profiles (not needed, saves ~30s build time).
    if _profile in _HAS_SENS:
        sensitivity = await compute_sensitivity_matrix(
            deal_model_id=deal_model_id,
            session=session,
            axis_x="noi_escalation_rate_pct",
            axis_y="exit_cap_rate_pct",
            metric="project_irr_levered",
            secondary_metric="equity_multiple",
            mode="combined",
            step_overrides={
                "noi_escalation_rate_pct": Decimal("1.0"),
                "exit_cap_rate_pct": Decimal("0.5"),
            },
        )
        sensitivity_sheet = wb.create_sheet("Sensitivity")
        _build_sensitivity(sensitivity_sheet, registry, ctx, sensitivity)
        # Splice from end → position 5 (after Cover, UW Summary, Pro Forma,
        # Cash Flow, Sources & Uses). openpyxl's wb._sheets is a plain list.
        wb._sheets.remove(sensitivity_sheet)
        wb._sheets.insert(5, sensitivity_sheet)

    registry.emit(wb)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


_PROFILE_SUFFIX: dict[str, str] = {
    "internal": "underwriting",
    "lp": "investor",
    "lender": "lender",
    "proforma": "proforma",
}


def make_investor_filename(
    scenario: DealModel,
    deal: Deal | None,
    profile: str = "internal",
) -> str:
    """Export filename: ``<deal>-<scenario>-<profile-suffix>.xlsx`` (slugged)."""
    deal_part = (deal.name if deal else None) or "deal"
    scen_part = scenario.name or "scenario"
    deal_slug = re.sub(r"[^a-z0-9]+", "-", deal_part.lower()).strip("-") or "deal"
    scen_slug = re.sub(r"[^a-z0-9]+", "-", scen_part.lower()).strip("-") or "scenario"
    suffix = _PROFILE_SUFFIX.get(profile, "export")
    return f"{deal_slug}-{scen_slug}-{suffix}.xlsx"


# ── Data loader ───────────────────────────────────────────────────────────────


async def _load_all(session: AsyncSession, scenario_id: UUID) -> dict | None:
    """Load the context dict the sheet builders read from.

    Shape mirrors plan §5.4 but only the fields commit 1 needs are populated
    today. Commits 2/3 will extend with cash flows, line items, waterfall
    rows, and rollup outputs.
    """
    scenario = (
        await session.execute(select(DealModel).where(DealModel.id == scenario_id))
    ).scalar_one_or_none()
    if scenario is None:
        return None

    deal = (
        await session.execute(select(Deal).where(Deal.id == scenario.deal_id))
    ).scalar_one_or_none()

    org: Organization | None = None
    if deal is not None and deal.org_id is not None:
        org = (
            await session.execute(
                select(Organization).where(Organization.id == deal.org_id)
            )
        ).scalar_one_or_none()

    projects = list(
        (
            await session.execute(
                select(Project)
                .where(Project.scenario_id == scenario_id)
                .order_by(Project.created_at.asc())
            )
        ).scalars()
    )
    project_ids = [p.id for p in projects]

    inputs_by_project: dict[UUID, OperationalInputs] = {}
    use_lines_by_project: dict[UUID, list[UseLine]] = {pid: [] for pid in project_ids}
    # unit_mix is JSONB on Project; wrap dicts in _UMProxy for attribute-compatible access
    unit_mix_by_project: dict[UUID, list] = {
        p.id: [_UMProxy(r) for r in (p.unit_mix or [])] for p in projects
    }

    if project_ids:
        for inp in (
            await session.execute(
                select(OperationalInputs).where(
                    OperationalInputs.project_id.in_(project_ids)
                )
            )
        ).scalars():
            inputs_by_project[inp.project_id] = inp
        for ul in (
            await session.execute(
                select(UseLine)
                .where(UseLine.project_id.in_(project_ids))
                .order_by(UseLine.phase, UseLine.label)
            )
        ).scalars():
            use_lines_by_project.setdefault(ul.project_id, []).append(ul)

    income_streams_by_project: dict[UUID, list[IncomeStream]] = {
        pid: [] for pid in project_ids
    }
    if project_ids:
        for stream in (
            await session.execute(
                select(IncomeStream)
                .where(IncomeStream.project_id.in_(project_ids))
                .order_by(IncomeStream.project_id, IncomeStream.label)
            )
        ).scalars():
            income_streams_by_project.setdefault(stream.project_id, []).append(stream)

    expense_lines_by_project: dict[UUID, list[OperatingExpenseLine]] = {
        pid: [] for pid in project_ids
    }
    if project_ids:
        for line in (
            await session.execute(
                select(OperatingExpenseLine)
                .where(OperatingExpenseLine.project_id.in_(project_ids))
                .order_by(OperatingExpenseLine.project_id, OperatingExpenseLine.label)
            )
        ).scalars():
            expense_lines_by_project.setdefault(line.project_id, []).append(line)

    capital_modules = list(
        (
            await session.execute(
                select(CapitalModule)
                .where(CapitalModule.scenario_id == scenario_id)
                .order_by(CapitalModule.stack_position)
            )
        ).scalars()
    )

    junctions: list[CapitalModuleProject] = []
    if capital_modules:
        module_ids = [m.id for m in capital_modules]
        junctions = list(
            (
                await session.execute(
                    select(CapitalModuleProject).where(
                        CapitalModuleProject.capital_module_id.in_(module_ids)
                    )
                )
            ).scalars()
        )

    # ── Cashflow / waterfall / rollup data (commit 2) ──────────────────────
    # Per-project cashflow + line items so the UW Pro Forma / Cash Flow
    # sheets can aggregate to annual buckets and the Investor Returns sheet
    # can read waterfall tier distributions.
    cash_flows_by_project: dict[UUID, list[CashFlow]] = {pid: [] for pid in project_ids}
    cash_flow_items_by_project: dict[UUID, list[CashFlowLineItem]] = {
        pid: [] for pid in project_ids
    }
    outputs_by_project: dict[UUID, OperationalOutputs] = {}
    if project_ids:
        for cf in (
            await session.execute(
                select(CashFlow)
                .where(CashFlow.scenario_id == scenario_id)
                .order_by(CashFlow.project_id, CashFlow.period)
            )
        ).scalars():
            if cf.project_id is not None:
                cash_flows_by_project.setdefault(cf.project_id, []).append(cf)
        for li in (
            await session.execute(
                select(CashFlowLineItem)
                .where(CashFlowLineItem.scenario_id == scenario_id)
                .order_by(CashFlowLineItem.project_id, CashFlowLineItem.period)
            )
        ).scalars():
            if li.project_id is not None:
                cash_flow_items_by_project.setdefault(li.project_id, []).append(li)
        for o in (
            await session.execute(
                select(OperationalOutputs).where(
                    OperationalOutputs.scenario_id == scenario_id
                )
            )
        ).scalars():
            if o.project_id is not None:
                outputs_by_project[o.project_id] = o

    waterfall_tiers = list(
        (
            await session.execute(
                select(WaterfallTier)
                .where(WaterfallTier.scenario_id == scenario_id)
                .order_by(WaterfallTier.priority)
            )
        ).scalars()
    )
    waterfall_results = list(
        (
            await session.execute(
                select(WaterfallResult)
                .where(WaterfallResult.scenario_id == scenario_id)
                .order_by(WaterfallResult.period)
            )
        ).scalars()
    )

    # Rollup helpers do their own DB roundtrips — call once and stash so
    # every sheet builder reads the same snapshot. ``rollup_summary``
    # returns ``{"per_project": [...], "totals": {...}}``;
    # ``rollup_waterfall`` returns the joined tier table.
    summary = await rollup_summary(scenario_id, session)
    waterfall_rollup = await rollup_waterfall(scenario_id, session)

    milestones_by_project: dict = {}
    if project_ids:
        _all_ms = list(
            (
                await session.execute(
                    select(Milestone)
                    .where(Milestone.project_id.in_(project_ids))
                    .order_by(Milestone.sequence_order.asc())
                )
            ).scalars()
        )
        for _m in _all_ms:
            milestones_by_project.setdefault(_m.project_id, []).append(_m)

    return {
        "scenario": scenario,
        "deal": deal,
        "org": org,
        "projects": projects,
        "operational_inputs": inputs_by_project,
        "use_lines": use_lines_by_project,
        "income_streams": income_streams_by_project,
        "expense_lines": expense_lines_by_project,
        "unit_mix": unit_mix_by_project,
        "capital_modules": capital_modules,
        "module_slugs": _compute_module_slugs(capital_modules),
        "junctions": junctions,
        "cash_flows": cash_flows_by_project,
        "cash_flow_items": cash_flow_items_by_project,
        "outputs": outputs_by_project,
        "waterfall_tiers": waterfall_tiers,
        "waterfall_results": waterfall_results,
        "rollup_summary": summary,
        "rollup_waterfall": waterfall_rollup,
        "snapshot_at": datetime.now(),
        "risk_free_rate_pct": (
            Decimal(str(scenario.risk_free_rate_pct))
            if scenario.risk_free_rate_pct is not None
            else Decimal(str(_app_settings.default_risk_free_rate_pct))
        ),
        "discount_rate_pct": (
            Decimal(str(scenario.discount_rate_pct))
            if scenario.discount_rate_pct is not None
            else await _resolve_discount_rate_default(scenario, deal, session)
        ),
        "milestones": milestones_by_project,
    }


# ── Sheet builders ────────────────────────────────────────────────────────────


_NOI_BASIS_LABELS: dict[str, str] = {
    "revenue_opex": "Revenue/OpEx",
    "noi": "Simplified NOI",
}

# Repo URL for in-workbook hyperlinks back to the FINANCIAL_MODEL.md headings.
# If the repo moves, update here. Anchor format follows GitHub's markdown
# heading convention (see _github_anchor_for).
_FINANCIAL_MODEL_URL = (
    "https://github.com/hahmlet/vicinitideals/blob/main/docs/FINANCIAL_MODEL.md"
)


def _github_anchor_for(metric) -> str:
    """Derive GitHub's auto-generated anchor for a tagged metric heading.

    GitHub renders ``### Total Project Cost (TPC) [investor, lender, app]``
    with the anchor ``#total-project-cost-tpc-investor-lender-app``: lowercase,
    drop everything that isn't alphanumeric / space / hyphen / underscore,
    replace runs of whitespace with single hyphens.
    """
    audiences = sorted(metric.audiences)
    heading = f"{metric.name} [{', '.join(audiences)}]"
    cleaned = re.sub(r"[^a-z0-9\s_-]", "", heading.lower())
    return re.sub(r"\s+", "-", cleaned).strip("-")


def _noi_basis_label(income_mode: str | None) -> str:
    """Translate the engine's `income_mode` enum to the LP-facing NOI Basis label.

    Engine stores `revenue_opex` (default) or `noi`; the LP cares about the
    semantic distinction between full P&L roll-up vs. direct-NOI input.
    """
    return _NOI_BASIS_LABELS.get(str(income_mode or "").lower(), str(income_mode or "—"))


def _build_version_tab(ws, ctx: dict) -> None:
    """Version/Audit tab — workbook metadata for reviewer audit trail.

    Placed first in the workbook per CRE best-practice convention: professional
    models lead with a version tab so any reviewer can immediately identify
    the model version, export timestamp, and change history.
    """
    set_widths(ws, [22, 42, 60])
    scenario = ctx["scenario"]
    ws.cell(row=1, column=1, value="Viciniti Investor Export").font = FONT_TITLE
    ws.cell(row=2, column=1, value="Version & Audit Trail").font = FONT_SUBTITLE
    row = 4
    snap = ctx.get("snapshot_at") or datetime.now()
    fields = [
        ("Export Generated", snap.strftime("%Y-%m-%d %H:%M UTC")),
        ("Export Version", "2.0"),
        ("Compute Version", f"v{scenario.version}" if getattr(scenario, 'version', None) is not None else "—"),
        ("Scenario Name", scenario.name or "—"),
        ("Scenario ID", str(scenario.id)),
        ("Deal Type", str(getattr(scenario, "project_type", "") or "—").replace("_", " ").title()),
    ]
    for label, value in fields:
        ws.cell(row=row, column=1, value=label).font = FONT_LABEL
        ws.cell(row=row, column=2, value=value).font = FONT_VALUE
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="Change Log").font = FONT_LABEL
    row += 1
    ws.cell(row=row, column=1, value="Version").font = FONT_HINT
    ws.cell(row=row, column=2, value="Date").font = FONT_HINT
    ws.cell(row=row, column=3, value="Notes").font = FONT_HINT
    row += 1
    changelog = [
        ("2.0", "2026-05-03", "Added Weighted EM, DCF NPV (configurable hurdle rate), Day Count per loan"),
        ("1.0", "2025-01-01", "Initial release: Cover, UW Summary, Pro Forma, Cash Flow, Returns, Waterfall, Debt Schedule, Sensitivity"),
    ]
    for ver, date_str, notes in changelog:
        ws.cell(row=row, column=1, value=ver).font = FONT_VALUE
        ws.cell(row=row, column=2, value=date_str).font = FONT_VALUE
        ws.cell(row=row, column=3, value=notes).font = FONT_VALUE
        row += 1


def _npv_levered(
    rollup_waterfall: list[dict],
    capital_modules: list,
    equity_required: Decimal,
    discount_rate_pct: Decimal,
) -> Decimal | None:
    """NPV of levered equity CF series discounted at investor's hurdle rate.

    Uses per-period waterfall cash_distributed as the equity CF series —
    this correctly captures operating distributions AND the large terminal
    exit payout that net_cash_flow (NOI−DS) omits.  Initial equity outlay
    is equity_required (the funded gap between Uses and Sources).

    Returns None when no distributable equity CFs found or denominator zero.
    """
    if discount_rate_pct <= 0 or equity_required <= 0:
        return None
    r = discount_rate_pct / Decimal(100)

    # Identify equity capital module IDs so we skip debt tiers.
    # If none have Equity class (auto-funded deals), fall through to sum all tiers.
    equity_ids = {
        str(m.id) for m in capital_modules
        if _funder_class(m) == "Equity"
    }

    per_period: dict[int, Decimal] = {}
    for row in rollup_waterfall:
        mid = row.get("capital_module_id")
        if equity_ids and mid not in equity_ids:
            continue
        t = row.get("period") or 0
        if t <= 0:
            continue
        dist = _coerce_decimal(row.get("cash_distributed") or 0)
        if dist:
            per_period[t] = per_period.get(t, Decimal(0)) + dist

    if not per_period:
        return None

    pv_dist = sum(
        v / (1 + r) ** (Decimal(t) / Decimal(12))
        for t, v in per_period.items()
    )
    return pv_dist - equity_required


def _weighted_em_calc(
    rollup_waterfall: list[dict],
    capital_modules: list,
    equity_required: Decimal,
    discount_rate_pct: Decimal,
) -> Decimal | None:
    """Weighted Equity Multiple = (equity + NPV) / equity.

    Adjusts raw EM for TVM using the investor's hurdle rate. A 2.0× WEM
    means PV of distributions equals 2× equity invested — a better quality
    signal than raw EM when comparing deals with different hold periods.
    """
    npv = _npv_levered(rollup_waterfall, capital_modules, equity_required, discount_rate_pct)
    if npv is None:
        return None
    if equity_required < Decimal(1):
        return None
    return (equity_required + npv) / equity_required


def _write_cover_timeline(
    ws,
    start_row: int,
    projects: list,
    milestones_by_project: dict,
) -> None:
    """Render a phase-timeline table on the Cover sheet."""
    if not any(milestones_by_project.get(p.id) for p in projects):
        return

    section_label(ws, start_row, "Project Timeline", span_cols=5)
    hdr = start_row + 1
    for col, txt in enumerate(["Phase", "Start", "End", "Days", "Months"], start=1):
        ws.cell(row=hdr, column=col, value=txt).font = FONT_LABEL

    row = hdr + 1
    for idx, project in enumerate(projects, start=1):
        ms_list = milestones_by_project.get(project.id, [])
        if not ms_list:
            continue
        if len(projects) > 1:
            ws.cell(row=row, column=1, value=f"P{idx} — {project.name}").font = FONT_LABEL
            row += 1
        ms_map = {m.id: m for m in ms_list}
        for m in ms_list:
            start_dt = m.computed_start(ms_map)
            end_dt = m.computed_end(ms_map)
            days = int(m.duration_days or 0)
            months = round(days / 30.4, 1) if days else 0.0
            raw_type = str(getattr(m.milestone_type, "value", m.milestone_type) or "")
            label = m.label or raw_type.replace("_", " ").title()
            ws.cell(row=row, column=1, value=label).font = FONT_VALUE
            ws.cell(row=row, column=2, value=start_dt.isoformat() if start_dt else "—").font = FONT_VALUE
            ws.cell(row=row, column=3, value=end_dt.isoformat() if end_dt else "—").font = FONT_VALUE
            ws.cell(row=row, column=4, value=days).font = FONT_VALUE
            c = ws.cell(row=row, column=5, value=months)
            c.font = FONT_VALUE
            c.number_format = "0.0"
            row += 1
        row += 1  # blank between projects


def _build_cover(ws, registry: CellRegistry, ctx: dict) -> None:
    """Cover sheet: key metrics summary, deal/scenario metadata, project list."""
    set_widths(ws, [30, 16, 14, 10, 12])
    scenario: DealModel = ctx["scenario"]
    deal: Deal | None = ctx["deal"]
    org: Organization | None = ctx["org"]
    projects: list[Project] = ctx["projects"]
    summary_data = ctx.get("rollup_summary") or {}
    totals = summary_data.get("totals") or {}
    per_project = summary_data.get("per_project") or []
    inputs_by_proj: dict = ctx.get("operational_inputs") or {}
    unit_mix_by_proj: dict = ctx.get("unit_mix") or {}

    row = 1

    # Title block
    ws.cell(row=row, column=1, value=f"{(deal.name if deal else '—')} — {scenario.name}")
    ws.cell(row=row, column=1).font = FONT_TITLE
    ws.cell(row=row, column=1).alignment = ALIGN_LEFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.row_dimensions[row].height = 28
    row += 2

    # Key Metrics summary — high-level figures for all export audiences.
    section_label(ws, row, "Key Metrics", span_cols=2)
    row += 1

    uses_total, sources_total, gap = _compute_sources_gap(ctx)
    combined_noi = _sum_per_project_field(per_project, "noi_stabilized")
    _tpc = _coerce_decimal(totals.get("total_project_cost") or 0)
    cap_rate_pct = (_coerce_pct(combined_noi / _tpc * Decimal(100)) if _tpc > 0 else None)
    combined_dscr = _combined_dscr(per_project)
    combined_irr = _coerce_pct(totals.get("combined_irr_pct") or 0)

    cash_flow_items: dict = ctx.get("cash_flow_items") or {}
    carrying_costs = sum(
        sum(_coerce_decimal(li.net_amount or 0) for li in items if li.label == "Carrying Cost")
        for items in cash_flow_items.values()
    )

    # Formula-conversion plan §4.7: when UW Summary is rendered for this
    # profile, point the Cover hero NOI / IRR cells at the named ranges
    # registered there so LP edits to revenue/OpEx assumptions flow up. For
    # proforma (no UW Summary), fall back to engine-computed scalars.
    _has_uw_summary = ctx.get("_profile") in {"internal", "lp", "lender"}
    kv_row(
        ws, row, "Stabilized NOI",
        "=s_combined_noi" if _has_uw_summary else combined_noi,
        name="s_cover_noi", registry=registry, fmt=ACCOUNTING,
    ); row += 1
    # Cap Rate on Cost = Stabilized NOI / Total Uses. Both operands are
    # available as named ranges when UW Summary + S&U render (always paired
    # for the profiles that get UW Summary). IFERROR guards uses=0 edge.
    kv_row(
        ws, row, "Cap Rate on Cost",
        ("=IFERROR(s_combined_noi/s_su_uses_total,\"\")" if _has_uw_summary else cap_rate_pct),
        name="s_cover_cap_rate", registry=registry, fmt=PCT,
    ); row += 1
    kv_row(ws, row, "DSCR (combined)", combined_dscr,
           name="s_cover_dscr", registry=registry, fmt="0.000")
    row += 1
    kv_row(
        ws, row, "Levered IRR",
        "=s_combined_irr" if _has_uw_summary else combined_irr,
        name="s_cover_irr", registry=registry, fmt=PCT,
    ); row += 1
    # Formula-conversion plan §4.7 (commit 2): Cover Total Uses + Total
    # Sources are cross-sheet references to the Sources & Uses sheet so
    # edits to the per-project Use lines (or to Block C principals) ripple
    # straight to the Cover hero block.
    kv_row(ws, row, "Total Uses", "=s_su_uses_total",
           name="s_cover_uses", registry=registry, fmt=ACCOUNTING)
    row += 1
    kv_row(ws, row, "Total Sources", "=s_su_sources_total",
           name="s_cover_sources", registry=registry, fmt=ACCOUNTING)
    row += 1
    if carrying_costs > Decimal(0):
        kv_row(ws, row, "Carrying Costs", carrying_costs,
               name="s_cover_carrying_costs", registry=registry, fmt=ACCOUNTING)
        row += 1
    row += 1  # spacer

    # Deal metadata block
    section_label(ws, row, "Deal", span_cols=2)
    row += 1
    kv_row(ws, row, "Sponsor / Organization", org.name if org else "—",
           name="s_sponsor_name", registry=registry, style="input")
    row += 1
    kv_row(ws, row, "Deal Name", deal.name if deal else "—",
           name="s_deal_name", registry=registry, style="input")
    row += 1
    kv_row(ws, row, "Scenario Name", scenario.name,
           name="s_scenario_name", registry=registry, style="input")
    row += 1
    snapshot_at: datetime = ctx["snapshot_at"]
    kv_row(ws, row, "Snapshot Date", snapshot_at.date().isoformat(),
           name="s_snapshot_date", registry=registry)
    row += 1
    kv_row(ws, row, "Project Count", len(projects),
           name="s_project_count", registry=registry, fmt=INT_COMMA)
    row += 1
    kv_row(ws, row, "NOI Basis", _noi_basis_label(scenario.income_mode),
           name="s_noi_basis", registry=registry, style="input")
    row += 2

    # Project list — subheader per project + unit count + rentable sqft
    section_label(ws, row, "Projects", span_cols=2)
    cur_row = row + 1
    for idx, proj in enumerate(projects, start=1):
        mix = unit_mix_by_proj.get(proj.id) or []
        inp = inputs_by_proj.get(proj.id)

        # Derive unit count: prefer unit_mix sum, fall back to OperationalInputs
        if mix:
            unit_total = sum(int(r.unit_count or 0) for r in mix)
        elif inp is not None:
            unit_total = int(inp.unit_count_new or 0) + int(inp.unit_count_existing or 0)
        else:
            unit_total = None

        # Rentable sqft: sum(sqft_per_unit * unit_count) across unit mix rows
        if mix and any(r.sqft for r in mix):
            rentable_sqft = sum(
                float(r.sqft or 0) * int(r.unit_count or 0) for r in mix
            )
        elif inp is not None and inp.building_sqft is not None:
            rentable_sqft = float(inp.building_sqft)
        else:
            rentable_sqft = None

        # Project subheader
        cell = ws.cell(row=cur_row, column=1, value=proj.name or f"Project {idx}")
        cell.font = FONT_SUBTITLE
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=2)
        cur_row += 1

        ws.cell(row=cur_row, column=1, value="Units").font = FONT_LABEL
        ws.cell(row=cur_row, column=2,
                value=unit_total if unit_total is not None else "—").font = FONT_VALUE
        cur_row += 1

        ws.cell(row=cur_row, column=1, value="Rentable Sq Ft").font = FONT_LABEL
        v_cell = ws.cell(row=cur_row, column=2,
                         value=round(rentable_sqft) if rentable_sqft is not None else "—")
        v_cell.font = FONT_VALUE
        if rentable_sqft is not None:
            v_cell.number_format = "#,##0"
        cur_row += 1

        cur_row += 1  # blank row between projects

    next_row = cur_row + 1

    # Sources-Gap banner — fires when Uses exceed funded Sources by > $1.
    if gap > Decimal(1):
        section_label(ws, next_row, "⚠ Equity Gap", span_cols=2)
        ws.cell(row=next_row + 1, column=1,
                value="Owner equity not formally committed").font = FONT_LABEL
        cell = ws.cell(row=next_row + 1, column=2, value=_to_excel_number(gap))
        cell.number_format = ACCOUNTING
        cell.font = FONT_VALUE
        cell.alignment = ALIGN_RIGHT
        registry.register("s_cover_sources_gap", ws.title, next_row + 1, 2)
        hint = (
            f"{_format_currency_short(gap)} of owner equity implied but not assigned to a module"
            f" — recorded as implied gap on Underwriting Summary Sources & Uses"
        )
        ws.cell(row=next_row + 2, column=1, value=hint).font = FONT_HINT
        ws.merge_cells(start_row=next_row + 2, start_column=1,
                       end_row=next_row + 2, end_column=2)
        legend_row = next_row + 4
    else:
        legend_row = next_row

    # Color legend
    section_label(ws, legend_row, "Color Legend", span_cols=2)
    ws.cell(row=legend_row + 1, column=1, value="Black text").font = FONT_VALUE
    ws.cell(row=legend_row + 1, column=2,
            value="Calculated value (derived from inputs).").font = FONT_HINT
    ws.cell(row=legend_row + 2, column=1, value="Blue text").font = FONT_INPUT
    ws.cell(row=legend_row + 2, column=2,
            value="User input (assumption that drives the model).").font = FONT_HINT
    ws.cell(row=legend_row + 3, column=1, value="Green underlined text").font = FONT_LINK
    ws.cell(row=legend_row + 3, column=2,
            value="Cross-sheet link or external reference (click to follow).").font = FONT_HINT
    ws.cell(row=legend_row + 4, column=1, value="Gold bold").font = FONT_HERO_VALUE
    ws.cell(row=legend_row + 4, column=2,
            value="Headline KPI on Underwriting Summary (TPC, IRR, NOI, etc.).").font = FONT_HINT

    tl_start = legend_row + 6
    _write_cover_timeline(ws, tl_start, projects, ctx.get("milestones", {}))

    freeze_top(ws, row=3)
    print_landscape(ws)


# ── Aggregation helpers (commit 2) ────────────────────────────────────────────


def _period_to_year(period: int) -> int:
    """Year-bucket convention from plan §5.3.

    Period 0 = acquisition close → Y0. Periods 1-12 = Y1, etc. Y0 carries
    capital events (acquisition outflows, partial-year operations) that
    aren't visible if rolled into Y1.
    """
    if period == 0:
        return 0
    return (period - 1) // 12 + 1


def _max_year(rows: list[CashFlow]) -> int:
    if not rows:
        return 0
    return max(_period_to_year(cf.period) for cf in rows)


def _aggregate_annual(monthly: list[CashFlow]) -> dict[int, dict[str, Decimal]]:
    """Aggregate per-period CashFlow rows into annual buckets.

    Returns ``{year: {field: Decimal}}`` for the standard cashflow fields.
    Skipping ``cumulative_cash_flow`` because it's a balance series, not
    additive — the consumers compute their own running totals.
    """
    fields = (
        "gross_revenue",
        "vacancy_loss",
        "effective_gross_income",
        "operating_expenses",
        "capex_reserve",
        "noi",
        "debt_service",
        "net_cash_flow",
    )
    out: dict[int, dict[str, Decimal]] = {}
    for cf in monthly:
        year = _period_to_year(cf.period)
        bucket = out.setdefault(year, {f: Decimal(0) for f in fields})
        for field in fields:
            bucket[field] += _coerce_decimal(getattr(cf, field, 0) or 0)
    return out


def _annual_line_items(
    items: list[CashFlowLineItem],
) -> dict[int, dict[str, Decimal]]:
    """Aggregate CashFlowLineItem rows by (year, label) for the Pro Forma sheet.

    OpEx categories like "Real Estate Taxes" / "Insurance" / "Property Mgmt"
    show up here as separate rows. Capital events (acquisition outflows,
    sale proceeds) likewise — the cash-flow sheet picks those out by
    label prefix.
    """
    out: dict[int, dict[str, Decimal]] = {}
    for li in items:
        year = _period_to_year(li.period)
        bucket = out.setdefault(year, {})
        bucket[li.label] = bucket.get(li.label, Decimal(0)) + _coerce_decimal(
            li.net_amount or 0
        )
    return out


def _waterfall_by_tier(
    rollup: list[dict],
) -> dict[str, dict[str, Decimal]]:
    """Aggregate the waterfall rollup into ``{tier_type: totals}``.

    Each tier-type bucket carries ``cash_total`` and ``module_count`` (unique
    Capital Modules that received distributions through this tier).
    """
    out: dict[str, dict[str, Decimal]] = {}
    seen_modules: dict[str, set[str]] = {}
    for row in rollup:
        tier = row.get("tier_type") or "unknown"
        bucket = out.setdefault(tier, {"cash_total": Decimal(0)})
        bucket["cash_total"] += _coerce_decimal(row.get("cash_distributed") or 0)
        module_id = row.get("capital_module_id")
        if module_id:
            seen_modules.setdefault(tier, set()).add(module_id)
    for tier, modules in seen_modules.items():
        out[tier]["module_count"] = Decimal(len(modules))
    return out


def _aggregate_scenario_annual(
    cash_flows_by_project: dict[UUID, list[CashFlow]],
) -> dict[int, dict[str, Decimal]]:
    """Sum all projects' annual cashflow buckets into scenario totals."""
    combined: dict[int, dict[str, Decimal]] = {}
    for cf_list in cash_flows_by_project.values():
        per_year = _aggregate_annual(cf_list)
        for year, fields in per_year.items():
            bucket = combined.setdefault(year, {})
            for field, value in fields.items():
                bucket[field] = bucket.get(field, Decimal(0)) + value
    return combined


# ── Deal Health helpers ───────────────────────────────────────────────────────

_RAG_FILLS = {
    "green":  FILL_RAG_GREEN,
    "yellow": FILL_RAG_YELLOW,
    "red":    FILL_RAG_RED,
}
_RAG_SYMBOL = {"green": "✓", "yellow": "!", "red": "✗"}


# Default discount rate (hurdle) by deal type.  Aligns with archetype IRR
# bands from CRE best-practice corpus: acquisition = LP pref threshold (8%),
# value-add / conversion = mid-range IRR target (10%), ground-up = development
# spread minimum (12%).  NULL scenario.discount_rate_pct falls back here.
_DISCOUNT_RATE_DEFAULTS: dict[str, Decimal] = {
    "acquisition":      Decimal("8.0"),
    "value_add":        Decimal("10.0"),
    "conversion":       Decimal("10.0"),
    "new_construction": Decimal("12.0"),
}


async def _resolve_discount_rate_default(
    scenario: Any,
    deal: Any,
    session: AsyncSession,
) -> Decimal:
    """Resolve the discount rate / hurdle when scenario.discount_rate_pct is NULL.

    Resolution order:
      1. Org/User default for ``irr_hurdle_pct_tier1`` (same concept: required
         return for NPV and waterfall). Requires both a creator user and deal
         org to look up.
      2. Per-deal-type hardcoded fallback (8% / 10% / 12%).
    """
    user_id = getattr(scenario, "created_by_user_id", None)
    org_id = getattr(deal, "org_id", None) if deal is not None else None
    if user_id is not None and org_id is not None:
        try:
            from app.settings.resolver import resolve_default as _resolve_one
            val = await _resolve_one("irr_hurdle_pct_tier1", user_id, org_id, session)
            if val is not None and str(val).strip() != "":
                return Decimal(str(val))
        except Exception:
            pass
    return _DISCOUNT_RATE_DEFAULTS.get(
        str(getattr(scenario, "project_type", "") or "").replace("ProjectType.", "").lower(),
        Decimal("8.0"),
    )


# Default RAG thresholds by deal type.  Yellow band = green − 5pp (pct metrics)
# or green − 0.10× (DSCR).  Red = below yellow.
HEALTH_THRESHOLD_DEFAULTS: dict[str, dict[str, float]] = {
    "acquisition":      {"occ_green": 93.0, "oer_green": 45.0, "dscr_green": 1.25, "margin_green": 10.0},
    "value_add":        {"occ_green": 90.0, "oer_green": 50.0, "dscr_green": 1.20, "margin_green":  8.0},
    "conversion":       {"occ_green": 90.0, "oer_green": 50.0, "dscr_green": 1.20, "margin_green":  8.0},
    "new_construction": {"occ_green": 88.0, "oer_green": 50.0, "dscr_green": 1.20, "margin_green":  8.0},
}
_HEALTH_THRESHOLD_FALLBACK = HEALTH_THRESHOLD_DEFAULTS["acquisition"]


def _resolve_thresholds(scenario: Any) -> dict[str, float]:
    """Return effective RAG thresholds: scenario override or deal-type default."""
    stored: dict | None = getattr(scenario, "health_thresholds", None)
    deal_type = str(getattr(scenario.project_type, "value", scenario.project_type) or "")
    defaults = HEALTH_THRESHOLD_DEFAULTS.get(deal_type, _HEALTH_THRESHOLD_FALLBACK)
    if not stored:
        return defaults
    return {**defaults, **{k: float(v) for k, v in stored.items() if v is not None}}


def _occ_rag(v: float, green: float = 93.0) -> str:
    return "green" if v >= green else "yellow" if v >= green - 5.0 else "red"


def _oer_rag(v: float, green: float = 45.0) -> str:
    return "green" if v <= green else "yellow" if v <= green + 10.0 else "red"


def _dscr_rag(v: float, green: float = 1.25) -> str:
    return "green" if v >= green else "yellow" if v >= green - 0.10 else "red"


def _margin_rag(v: float, green: float = 10.0) -> str:
    return "green" if v >= green else "yellow" if v >= green - 5.0 else "red"


def _compute_deal_health(ctx: dict) -> dict[str, Any]:
    """Compute 4-pillar health signals and archetype + IRR band classification."""
    scenario = ctx["scenario"]
    thresholds = _resolve_thresholds(scenario)
    per_project: list[dict] = (ctx.get("rollup_summary") or {}).get("per_project") or []
    totals: dict = (ctx.get("rollup_summary") or {}).get("totals") or {}
    inputs_by_project: dict = ctx.get("operational_inputs") or {}
    cash_flows_by_project: dict = ctx.get("cash_flows") or {}

    # Pillar 1: Stabilized Occupancy — NOI-weighted avg across projects.
    # Use explicit stabilized_occupancy_pct when set; fall back to EGI ÷
    # gross_revenue from stabilized cashflow rows when the field is blank.
    _proj_implied_occ: dict[UUID, Decimal] = {}
    for _pid, _cfs in cash_flows_by_project.items():
        _gr = Decimal(0)
        _egi_occ = Decimal(0)
        for _cf in _cfs:
            if str(getattr(_cf.period_type, "value", _cf.period_type) or "") == "stabilized":
                _gr      += _coerce_decimal(_cf.gross_revenue or 0)
                _egi_occ += _coerce_decimal(_cf.effective_gross_income or 0)
        if _gr > Decimal(0):
            _proj_implied_occ[_pid] = _egi_occ / _gr * Decimal(100)

    def _occ_for_project(p: dict) -> Decimal:
        pid = UUID(p["project_id"]) if p.get("project_id") else None
        inp = inputs_by_project.get(pid) if pid else None
        explicit = _coerce_decimal(getattr(inp, "stabilized_occupancy_pct", None) or 0)
        return explicit if explicit > Decimal(0) else _proj_implied_occ.get(pid, Decimal(0))

    total_noi = _sum_per_project_field(per_project, "noi_stabilized")
    combined_occ: Decimal | None
    if total_noi > Decimal(0):
        occ_sum = sum(
            _coerce_decimal(p.get("noi_stabilized") or 0) * _occ_for_project(p)
            for p in per_project
            if p.get("project_id")
        )
        combined_occ = occ_sum / total_noi
    else:
        occs = [_occ_for_project(p) for p in per_project if p.get("project_id")]
        combined_occ = sum(occs, Decimal(0)) / len(occs) if occs else None

    # Pillars 2 + 4: OER and Post-Debt CF Margin from stabilized CashFlow rows
    total_opex = Decimal(0)
    total_egi = Decimal(0)
    total_ncf = Decimal(0)
    for cfs in cash_flows_by_project.values():
        for cf in cfs:
            if str(getattr(cf.period_type, "value", cf.period_type) or "") == "stabilized":
                total_opex += _coerce_decimal(cf.operating_expenses or 0)
                total_egi  += _coerce_decimal(cf.effective_gross_income or 0)
                total_ncf  += _coerce_decimal(cf.net_cash_flow or 0)
    combined_oer: Decimal | None = (
        (total_opex / total_egi * Decimal(100)) if total_egi > Decimal(0) else None
    )
    ncf_margin: Decimal | None = (
        (total_ncf / total_egi * Decimal(100)) if total_egi > Decimal(0) else None
    )

    # Pillar 3: DSCR
    dscr = _combined_dscr(per_project)

    # Archetype: deal_type primary; occupancy + OER split acquisition into Core / Core-Plus
    deal_type = str(getattr(scenario.project_type, "value", scenario.project_type) or "")
    occ_f = float(combined_occ) if combined_occ is not None else 0.0
    oer_f = float(combined_oer) if combined_oer is not None else 50.0

    if deal_type in ("conversion", "new_construction"):
        archetype, irr_lo, irr_hi = "Opportunistic", 15.0, None
    elif deal_type == "value_add":
        archetype, irr_lo, irr_hi = "Value-Add", 11.0, 15.0
    elif deal_type == "acquisition":
        if occ_f >= 92.0 and oer_f <= 42.0:
            archetype, irr_lo, irr_hi = "Core", 6.0, 8.0
        else:
            archetype, irr_lo, irr_hi = "Core-Plus", 8.0, 11.0
    else:
        archetype, irr_lo, irr_hi = "Value-Add", 11.0, 15.0

    combined_irr = float(_coerce_decimal(totals.get("combined_irr_pct") or 0))
    irr_flag = combined_irr < irr_lo or (irr_hi is not None and combined_irr > irr_hi)

    return {
        "occupancy_pct": combined_occ,
        "oer_pct": combined_oer,
        "dscr": dscr,
        "ncf_margin_pct": ncf_margin,
        "archetype": archetype,
        "irr_lo": irr_lo,
        "irr_hi": irr_hi,
        "irr_flag": irr_flag,
        "combined_irr_pct": Decimal(str(combined_irr)),
        "thresholds": thresholds,
    }


def _write_deal_health_section(
    ws, row: int, registry: CellRegistry, health: dict[str, Any]
) -> int:
    """Write Deal Health block starting at `row`. Returns first row after block."""
    section_label(ws, row, "Deal Health", span_cols=4)
    row += 1

    def _pillar(
        label: str,
        value: Any,
        fmt: str,
        rag: str | None,
        note: str,
        name: str,
    ) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = FONT_LABEL
        ws.cell(row=row, column=1).alignment = ALIGN_LEFT
        if value is None:
            ws.cell(row=row, column=2, value="—").font = FONT_VALUE
        else:
            cv = ws.cell(row=row, column=2)
            cv.value = _to_excel_number(value)
            cv.number_format = fmt
            cv.font = FONT_VALUE
            cv.alignment = ALIGN_RIGHT
        if rag is not None:
            cs = ws.cell(row=row, column=3, value=_RAG_SYMBOL[rag])
            cs.fill = _RAG_FILLS[rag]
            cs.font = FONT_VALUE
            cs.alignment = ALIGN_RIGHT
        ws.cell(row=row, column=4, value=note).font = FONT_HINT
        registry.register(name, ws.title, row, 2)
        row += 1

    thr = health.get("thresholds") or _HEALTH_THRESHOLD_FALLBACK
    occ_g    = thr["occ_green"]
    oer_g    = thr["oer_green"]
    dscr_g   = thr["dscr_green"]
    margin_g = thr["margin_green"]

    occ = health["occupancy_pct"]
    _pillar(
        "Stabilized Occupancy",
        _coerce_pct(occ) if occ is not None else None,
        PCT,
        _occ_rag(float(occ), occ_g) if occ is not None else None,
        f"≥ {occ_g:.0f}% green · {occ_g - 5:.0f}–{occ_g:.0f}% yellow · <{occ_g - 5:.0f}% red",
        "dh_occupancy",
    )
    oer = health["oer_pct"]
    _pillar(
        "Operating Expense Ratio",
        _coerce_pct(oer) if oer is not None else None,
        PCT,
        _oer_rag(float(oer), oer_g) if oer is not None else None,
        f"≤ {oer_g:.0f}% green · {oer_g:.0f}–{oer_g + 10:.0f}% yellow · >{oer_g + 10:.0f}% red",
        "dh_oer",
    )
    dscr_v = _coerce_decimal(health["dscr"]) if health["dscr"] is not None else None
    _pillar(
        "DSCR",
        float(dscr_v) if dscr_v is not None else None,
        "0.00",
        _dscr_rag(float(dscr_v), dscr_g) if dscr_v is not None else None,
        f"≥ {dscr_g:.2f}× green · {dscr_g - 0.10:.2f}–{dscr_g:.2f}× yellow · <{dscr_g - 0.10:.2f}× red",
        "dh_dscr",
    )
    margin = health["ncf_margin_pct"]
    _pillar(
        "Post-Debt CF Margin (NCF / EGI)",
        _coerce_pct(margin) if margin is not None else None,
        PCT,
        _margin_rag(float(margin), margin_g) if margin is not None else None,
        f"≥ {margin_g:.0f}% green · {margin_g - 5:.0f}–{margin_g:.0f}% yellow · <{margin_g - 5:.0f}% red",
        "dh_ncf_margin",
    )

    row += 1  # blank separator before archetype row

    arch = health["archetype"]
    irr_lo = health["irr_lo"]
    irr_hi = health["irr_hi"]
    irr_pct = float(health["combined_irr_pct"])
    irr_flag = health["irr_flag"]
    band_str = f"{irr_lo:.0f}–{irr_hi:.0f}%" if irr_hi is not None else f"≥{irr_lo:.0f}%"
    irr_note = (
        f"⚠ IRR {irr_pct:.1f}% outside {arch} band ({band_str})"
        if irr_flag
        else f"✓ IRR {irr_pct:.1f}% within {arch} band ({band_str})"
    )
    irr_rag = "red" if irr_flag else "green"

    ws.cell(row=row, column=1, value="Archetype").font = FONT_LABEL
    ws.cell(row=row, column=1).alignment = ALIGN_LEFT
    av = ws.cell(row=row, column=2, value=arch)
    av.font = FONT_VALUE
    av.alignment = ALIGN_LEFT
    cs = ws.cell(row=row, column=3, value=_RAG_SYMBOL[irr_rag])
    cs.fill = _RAG_FILLS[irr_rag]
    cs.font = FONT_VALUE
    cs.alignment = ALIGN_RIGHT
    ws.cell(row=row, column=4, value=irr_note).font = FONT_HINT
    registry.register("dh_archetype", ws.title, row, 2)
    row += 1

    return row + 1  # trailing blank row after block


# ── Underwriting Summary sheet ────────────────────────────────────────────────


def _build_uw_summary(ws, registry: CellRegistry, ctx: dict) -> None:
    """Underwriting Summary: hero KPIs + scenario S&U + per-project mini-table.

    KPI sources reference ``rollup_summary`` totals + ``rollup_irr``. The
    per-project mini-table uses ``=HYPERLINK("#'P1 Liberty'!A1", ...)``
    Excel syntax to navigate to per-project sheets — those sheets land in
    commit 3, so the hyperlink is already present and resolves once those
    sheets exist.
    """
    set_widths(ws, [32, 24, 18, 14, 14, 14, 14])
    summary = ctx.get("rollup_summary") or {}
    totals = summary.get("totals") or {}
    per_project = summary.get("per_project") or []
    rollup_waterfall: list[dict] = ctx.get("rollup_waterfall") or []
    projects: list[Project] = ctx["projects"]
    capital_modules: list[CapitalModule] = ctx["capital_modules"]
    junctions: list[CapitalModuleProject] = ctx["junctions"]
    use_lines_by_project: dict[UUID, list[UseLine]] = ctx["use_lines"]

    # ── Deal Health block (top of sheet — first thing LP sees) ───────────────
    health = _compute_deal_health(ctx)
    row = _write_deal_health_section(ws, 1, registry, health)

    # ── Primary KPI block ──────────────────────────────────────────────────
    section_label(ws, row, "Primary KPIs", span_cols=2)
    row += 1
    kv_row(
        ws, row, "Total Project Cost",
        _coerce_decimal(totals.get("total_project_cost") or 0),
        name="s_total_project_cost", registry=registry,
        fmt=ACCOUNTING, hero=True,
    ); row += 1
    # Phase 4: Total Uses references the S&U sheet's total directly so
    # any LP edit to a Use line on Sources & Uses ripples through to
    # this KPI without re-running the engine. Both values are defined
    # as "sum of UseLine.amount where phase != exit" (see engine
    # rollup_summary.totals.total_uses and the S&U sheet's uses_total
    # cell), so the formula is semantically equivalent to the prior
    # engine-derived scalar.
    kv_row(
        ws, row, "Total Uses",
        "=s_su_uses_total",
        name="s_total_uses", registry=registry,
        fmt=ACCOUNTING, hero=True,
    ); row += 1
    kv_row(
        ws, row, "Equity Required",
        _coerce_decimal(totals.get("equity_required") or 0),
        name="s_equity_required", registry=registry,
        fmt=ACCOUNTING, hero=True,
    ); row += 1
    # Combined DSCR = Σ NOI / Σ DS across projects. Per LP feedback the
    # right number for a Primary-KPI block is a singular combined coverage
    # figure, not the weakest project's DSCR. Engine doesn't store DS as
    # a per-project scalar, so derive it from per-project (NOI ÷ DSCR).
    # Phase 4 completion: prefer the Pro Forma Y1 NOI / Debt Service
    # cells (formula chain back to Block F/G inputs and Debt Schedule)
    # whenever both are positive — LP edits then ripple here. Engine
    # fallback baked into the IF false branch covers pre-operational Y1
    # scenarios where the formula chain would resolve to 0.
    combined_dscr = _combined_dscr(per_project)
    _dscr_fallback = float(combined_dscr) if combined_dscr else 0
    _dscr_formula = (
        f"=IF(AND(s_pf_noi_y1>0,s_pf_debt_service_y1>0),"
        f"s_pf_noi_y1/s_pf_debt_service_y1,{_dscr_fallback})"
    )
    kv_row(
        ws, row, "Stabilized DSCR (combined)",
        _dscr_formula,
        name="s_combined_dscr", registry=registry,
        fmt="0.000", hero=True,
    ); row += 1
    _noi_fallback = float(_sum_per_project_field(per_project, "noi_stabilized") or 0)
    _noi_formula = f"=IF(s_pf_noi_y1>0,s_pf_noi_y1,{_noi_fallback})"
    kv_row(
        ws, row, "Combined Stabilized NOI (DSCR basis)",
        _noi_formula,
        name="s_combined_noi", registry=registry,
        fmt=ACCOUNTING, hero=True,
    ); row += 1
    # Phase 4 completion: Combined Levered IRR runs IRR() directly on the
    # Underwriting Cash Flow levered row (r_uw_cf_levered). Same formula
    # the Investor Returns sheet uses for s_returns_combined_irr — both
    # KPIs compute against the single underlying range, so an upstream
    # cash-flow edit ripples to both identically. Direct reference (vs
    # aliasing s_returns_combined_irr) keeps the formula self-contained
    # for the lender profile, which renders UW Summary but not Returns.
    kv_row(
        ws, row, "Combined Levered IRR",
        "=IFERROR(IRR(r_uw_cf_levered),0)",
        name="s_combined_irr", registry=registry,
        fmt=PCT, hero=True,
    ); row += 1
    # Hold = max of milestone chain (engine writes total_timeline_months as
    # the count of generated cashflow rows = sum of phase durations from
    # acquisition close through divestment, or stabilized when no divestment
    # exists). This is the actual modeled horizon, distinct from the
    # OperationalInputs.hold_period_years input on the Assumptions sheet
    # (which represents the user's intent for *stabilized* hold only).
    # Phase 4 completion: Modeled Duration = MAX across each project's
    # Assumptions phase-plan total horizon. Each project's phase block
    # registers `p{idx}_total_horizon_months` (see
    # `_emit_phase_plan_block`), so an LP edit to any phase duration
    # ripples to the longest-hold KPI without re-running the engine.
    # Single-project = literal ref; multi-project = MAX. Engine value
    # (`_longest_hold_months`) is the fallback when no projects exist.
    _horizon_names = [f"p{idx}_total_horizon_months" for idx in range(1, len(projects) + 1)]
    if len(_horizon_names) == 1:
        _duration_value = f"=IFERROR({_horizon_names[0]},0)"
    elif len(_horizon_names) > 1:
        _duration_value = f"=IFERROR(MAX({','.join(_horizon_names)}),0)"
    else:
        _duration_value = _longest_hold_months(per_project)
    kv_row(
        ws, row, "Total Modeled Duration (months)",
        _duration_value,
        name="s_modeled_duration_months", registry=registry,
        fmt=INT_COMMA, hero=True,
    ); row += 1
    # Combined Unlevered IRR — computed from per-project unlevered CF series
    # summed by period (NCF + DS = NOI − capital_outflows + capital_inflows),
    # then XIRR. Sibling to Combined Levered IRR for the standard
    # leverage-amplification read.
    unlevered_irr = _combined_unlevered_irr(ctx["cash_flows"])
    kv_row(
        ws, row, "Combined Unlevered IRR",
        unlevered_irr,
        name="s_combined_unlevered_irr", registry=registry,
        fmt=PCT, hero=True,
    ); row += 1
    # Equity multiples — EM, WEM, CoC. Require a non-trivial equity basis.
    # Compute _equity_req first so the EM fallback and CoC fallback can gate
    # on it: a 100%-debt deal (equity_required < $1) suppresses all three
    # rather than showing multiples against a near-zero denominator.
    _disc_rate = _coerce_decimal(ctx.get("discount_rate_pct") or Decimal("8.0"))
    _equity_req = _coerce_decimal(totals.get("equity_required") or 0)
    # Formula-conversion plan §4.3 (commit 8): EM is now a SUMIF over the
    # Underwriting Cash Flow levered row — positive flows are distributions,
    # negative flows are equity calls. LP edits to NOI, debt service, or
    # capital events on the upstream cells re-derive EM via Excel's calc
    # engine. IFERROR guards the zero-equity / degenerate case.
    _kv_row_optional(
        ws, row, "Combined Equity Multiple",
        '=IFERROR(SUMIF(r_uw_cf_levered,">0")/(-SUMIF(r_uw_cf_levered,"<0")),0)',
        name="s_combined_equity_multiple", registry=registry,
        fmt='0.00"×"', hero=True,
    ); row += 1
    w_em = _weighted_em_calc(rollup_waterfall, capital_modules, _equity_req, _disc_rate)
    _kv_row_optional(
        ws, row, f"Weighted Equity Multiple ({_disc_rate:.2f}% hurdle)",
        w_em,
        name="s_weighted_equity_multiple", registry=registry,
        fmt="0.00\\x", hero=True,
    ); row += 1
    coc_y1 = _coc_year_one(rollup_waterfall, capital_modules)
    if coc_y1 is None and _equity_req > Decimal(1):
        # Auto-funded deals have $0 equity module commitments so _coc_year_one
        # returns None (denominator = 0). Fall back: use scenario equity_required
        # as the denominator, sum Y1 cash_distributed across all waterfall tiers.
        _y1_dists = sum(
            _coerce_decimal(row.get("cash_distributed") or 0)
            for row in rollup_waterfall
            if 1 <= (row.get("period") or 0) <= 12
        )
        if _y1_dists > 0:
            coc_y1 = _y1_dists / _equity_req
    _kv_row_optional(
        ws, row, "Cash-on-Cash (Year 1)",
        coc_y1,
        name="s_coc_year_one", registry=registry,
        fmt=PCT, hero=True,
    ); row += 1

    # ── Spread Stack ──────────────────────────────────────────────────────────
    # Three rows anchoring risk-adjusted return context:
    #   RFR → Cap Rate (unlevered going-in premium over T-bill)
    #   Cap Rate → Levered IRR (execution / leverage premium)
    row += 1
    section_label(ws, row, "Spread Stack", span_cols=2)
    row += 1
    _rfr = _coerce_decimal(ctx.get("risk_free_rate_pct") or Decimal("4.25"))
    kv_row(
        ws, row, "Risk-Free Rate (10Y Treasury)", _coerce_pct(_rfr),
        name="s_rfr_pct", registry=registry, fmt=PCT, hero=True,
    ); row += 1
    # Phase 4: Spread Stack cells become formulas referencing the named
    # KPI cells above (s_combined_noi, s_total_project_cost,
    # s_combined_irr, s_rfr_pct). When LP edits revenue / OpEx / Use
    # lines, NOI + TPC re-derive upstream and the Spread Stack rows
    # follow without re-running the engine. IFERROR guards a missing
    # TPC (0% / unconfigured deal).
    kv_row(
        ws, row, "Cap Rate on Cost (stabilized Y1 NOI ÷ TPC)",
        "=IFERROR(s_combined_noi/s_total_project_cost,0)",
        name="s_spread_cap_pct", registry=registry, fmt=PCT, hero=True,
    ); row += 1
    kv_row(
        ws, row, "Cap Rate Spread (vs RFR)",
        "=IFERROR(s_spread_cap_pct-s_rfr_pct,0)",
        name="s_cap_rate_spread", registry=registry, fmt=PCT, hero=True,
    ); row += 1
    kv_row(
        ws, row, "Levered IRR Spread (vs Cap Rate)",
        "=IFERROR(s_combined_irr-s_spread_cap_pct,0)",
        name="s_irr_spread", registry=registry, fmt=PCT, hero=True,
    ); row += 1
    kv_row(
        ws, row, "Levered IRR Spread (vs RFR)",
        "=IFERROR(s_combined_irr-s_rfr_pct,0)",
        name="s_irr_rfr_spread", registry=registry, fmt=PCT, hero=True,
    ); row += 1

    # ── Scenario Sources & Uses ────────────────────────────────────────────
    su_row = row + 2
    section_label(ws, su_row, "Scenario Sources & Uses", span_cols=4)
    header_row(ws, su_row + 1, ["Side", "Label", "Amount", "Notes"])
    line = su_row + 2

    # Uses — sum across projects, by cost category
    uses_by_cat: dict[str, Decimal] = {}
    for pid, uls in use_lines_by_project.items():
        for ul in uls:
            phase = str(getattr(ul.phase, "value", ul.phase) or "")
            if phase == "exit":
                continue
            cat = str(ul.cost_category or "soft")
            uses_by_cat[cat] = uses_by_cat.get(cat, Decimal(0)) + _coerce_decimal(ul.amount or 0)
    for cat in USE_COST_CATEGORIES:
        amount = uses_by_cat.get(cat, Decimal(0))
        cat_label = USE_CATEGORY_LABELS.get(cat, cat.title())
        ws.cell(row=line, column=1, value="Use").font = FONT_VALUE
        ws.cell(row=line, column=2, value=cat_label).font = FONT_VALUE
        ws.cell(row=line, column=3, value=_to_excel_number(amount)).number_format = ACCOUNTING
        ws.cell(row=line, column=4, value="(all projects)").font = FONT_HINT
        line += 1
    uses_total = sum(uses_by_cat.values(), Decimal(0))
    ws.cell(row=line, column=1, value="Use").font = FONT_LABEL
    ws.cell(row=line, column=2, value="Total Uses (excl. exit)").font = FONT_LABEL
    # Reference the dedicated S&U sheet's grand total so this stays in sync
    # when LP edits Block C principals. The S&U sheet registers the name.
    _c = ws.cell(row=line, column=3, value="=s_su_uses_total")
    _c.number_format = ACCOUNTING
    _c.font = FONT_LABEL
    _c.alignment = ALIGN_RIGHT
    line += 2

    # Sources — capital modules, deduplicated for shared modules via junctions
    junction_amount: dict[UUID, Decimal] = {}
    for j in junctions:
        junction_amount[j.capital_module_id] = junction_amount.get(
            j.capital_module_id, Decimal(0)
        ) + _coerce_decimal(j.amount or 0)
    sources_total = Decimal(0)
    for module in capital_modules:
        amount = junction_amount.get(module.id) or _coerce_decimal(
            (module.source or {}).get("amount") or 0
        )
        if amount <= Decimal(1) and _funder_class(module) == "Equity":
            continue
        ws.cell(row=line, column=1, value="Source").font = FONT_VALUE
        ws.cell(row=line, column=2, value=module.label or _funder_type_label(module)).font = FONT_VALUE
        ws.cell(row=line, column=3, value=_to_excel_number(amount)).number_format = ACCOUNTING
        ws.cell(row=line, column=4, value=_funder_type_label(module)).font = FONT_HINT
        sources_total += amount
        line += 1

    # Auto-funded equity (implied gap): equity modules sized as Uses − explicit Sources.
    # Show the implied contribution as a funded source line so Total Sources = Total Uses
    # and lenders don't flag a gap that is intentionally covered by sponsor equity.
    _implied_equity = uses_total - sources_total
    if _implied_equity > Decimal(1):
        ws.cell(row=line, column=1, value="Source").font = FONT_VALUE
        ws.cell(row=line, column=2, value="Owner Equity (implied gap)").font = FONT_VALUE
        ws.cell(row=line, column=3, value=_to_excel_number(_implied_equity)).number_format = ACCOUNTING
        ws.cell(row=line, column=4, value="Auto-funded equity — residual after debt").font = FONT_HINT
        sources_total += _implied_equity
        line += 1

    ws.cell(row=line, column=1, value="Source").font = FONT_LABEL
    ws.cell(row=line, column=2, value="Total Sources").font = FONT_LABEL
    _c = ws.cell(row=line, column=3, value="=s_su_sources_total")
    _c.number_format = ACCOUNTING
    _c.font = FONT_LABEL
    _c.alignment = ALIGN_RIGHT
    line += 1

    ws.cell(row=line, column=1, value="Δ").font = FONT_LABEL
    ws.cell(row=line, column=2, value="Sources Gap (Uses − Sources)").font = FONT_LABEL
    _c = ws.cell(row=line, column=3, value="=s_su_uses_total-s_su_sources_total")
    _c.number_format = ACCOUNTING
    _c.font = FONT_LABEL
    _c.alignment = ALIGN_RIGHT
    registry.register("s_sources_gap", ws.title, line, 3)
    line += 2

    # ── Per-project mini-summary ───────────────────────────────────────────
    pp_row = line + 1
    section_label(ws, pp_row, "Per-Project Mini-Summary", span_cols=7)
    _pp_col3_header = "Equity Req'd" if _equity_req > Decimal(1) else "Cap on Cost"
    header_row(
        ws, pp_row + 1,
        ["Project", "TPC", _pp_col3_header, "Stabilized NOI", "DSCR", "Levered IRR", "Sheet"],
    )
    pp_data = pp_row + 2
    for idx, project in enumerate(projects, start=1):
        proj_id = str(project.id)
        record = next(
            (p for p in per_project if str(p.get("project_id") or "") == proj_id),
            {},
        )
        ws.cell(row=pp_data, column=1, value=project.name or f"Project {idx}").font = FONT_VALUE
        ws.cell(row=pp_data, column=2, value=_to_excel_number(record.get("total_project_cost"))).number_format = ACCOUNTING
        if _equity_req > Decimal(1):
            ws.cell(row=pp_data, column=3, value=_to_excel_number(record.get("equity_required"))).number_format = ACCOUNTING
        else:
            _pp_noi = _coerce_decimal(record.get("noi_stabilized") or 0)
            _pp_tpc = _coerce_decimal(record.get("total_project_cost") or 0)
            ws.cell(row=pp_data, column=3, value=_to_excel_number(_pp_noi / _pp_tpc if _pp_tpc > 0 else None)).number_format = PCT
        ws.cell(row=pp_data, column=4, value=_to_excel_number(record.get("noi_stabilized"))).number_format = ACCOUNTING
        ws.cell(row=pp_data, column=5, value=_to_excel_number(record.get("dscr"))).number_format = "0.000"
        levered = record.get("project_irr_levered")
        ws.cell(row=pp_data, column=6, value=_to_excel_number(_coerce_pct(levered) if levered is not None else None)).number_format = PCT
        # Sheet hyperlink: for single-project scenarios the per-project
        # sheet is suppressed (plan §5, commit 8) so point at the
        # combined Pro Forma instead. Multi-project keeps the per-P
        # sheet target.
        if len(projects) == 1:
            sheet_label = "Underwriting Pro Forma"
        else:
            sheet_label = _project_sheet_name(idx, project.name)
        ws.cell(
            row=pp_data, column=7,
            value=f'=HYPERLINK("#\'{sheet_label}\'!A1", "→ open")',
        ).font = FONT_LINK
        pp_data += 1

    # ── Property Valuation ─────────────────────────────────────────────────
    # The previous "Valuation Reconciliation" block compared two methods
    # that the engine computes identically (sale_proceeds = stab_NOI /
    # exit_cap; Direct Cap is also stab_NOI / exit_cap), so Δ was always
    # $0 — tautological per V2-D in the Subject Model Review.
    #
    # The reframed block presents three distinct valuations the LP can
    # actually act on:
    #   - Yield on Cost = NOI / TPC: the asset's unlevered earnings rate
    #     against what it cost to build/buy. The headline "is this deal
    #     reasonable on its own?" check.
    #   - Going-In Cap Value = NOI / Going-In Cap: the market valuation
    #     at acquisition based on the analyst's going-in cap input.
    #   - Exit Cap Value = NOI / Exit Cap: the market valuation at exit.
    #     Differs from going-in only when the analyst has set the two
    #     caps differently (cap-rate compression / decompression).
    # Cap Spread (Yield on Cost − Going-In Cap) shows the yield premium
    # — positive means buying below market cap, negative means above.
    val_row = pp_data + 1
    section_label(ws, val_row, "Property Valuation", span_cols=3)
    header_row(ws, val_row + 1, ["Method", "Value", "Notes"])

    default_inputs = (
        ctx["operational_inputs"].get(projects[0].id) if projects else None
    )
    exit_cap_pct_raw = _coerce_decimal(
        getattr(default_inputs, "exit_cap_rate_pct", 0) or 0
    )
    going_in_cap_pct_raw = _coerce_decimal(
        getattr(default_inputs, "going_in_cap_rate_pct", 0) or 0
    )
    combined_noi = _sum_per_project_field(per_project, "noi_stabilized")
    combined_tpc = _coerce_decimal(totals.get("total_project_cost") or 0)

    # Exit-year NOI: sum monthly NOI for the last 12 months of the modeled hold
    # (capped at Y10), matching the _per_year_irr_series convention.
    # Yield-on-cost and Going-In Cap Value intentionally keep stabilized (Y1) NOI.
    _cf_map: dict = ctx["cash_flows"]
    _pnoi: dict[int, Decimal] = {}
    for _cfl in _cf_map.values():
        for _cf in _cfl:
            _pnoi[_cf.period] = _pnoi.get(_cf.period, Decimal(0)) + _coerce_decimal(_cf.noi or 0)
    _ann = _aggregate_scenario_annual(_cf_map)
    _exit_yr = min(max(_ann) if _ann else 10, 10)
    _exit_noi = sum(
        (_pnoi.get(p, Decimal(0)) for p in range(_exit_yr * 12 - 11, _exit_yr * 12 + 1)),
        Decimal(0),
    )
    exit_year_noi = _exit_noi if _exit_noi > 0 else combined_noi

    # yield_on_cost is still needed for the Cap Spread hint text ("Yield
    # premium / Yield discount"); the cell itself is a formula.
    # going_in_value / exit_value were removed after the Property Valuation
    # block was converted to live formulas — see Appendix F in
    # docs/FINANCIAL_MODEL.md for the catalog of formula-driven cells.
    yield_on_cost = (combined_noi / combined_tpc) if combined_tpc > 0 else None

    cur = val_row + 2

    # Row 1: Yield on Cost — Phase D: formula-driven so it tracks LP edits
    # to revenue/OpEx assumptions (through s_combined_noi on this sheet)
    # and Use-line edits (through s_su_uses_total on the S&U sheet). The
    # IFERROR fallback covers the s_su_uses_total == 0 / unconfigured
    # case, so the formula is safe regardless of seed data quality.
    ws.cell(row=cur, column=1, value="Yield on Cost (NOI ÷ TPC)").font = FONT_LABEL
    formula = "=IFERROR(s_combined_noi/s_su_uses_total,\"\")"
    cell = ws.cell(row=cur, column=2, value=formula)
    cell.number_format = PCT
    cell.font = FONT_VALUE
    cell.alignment = ALIGN_RIGHT
    registry.register("s_yield_on_cost", ws.title, cur, 2)
    ws.cell(
        row=cur, column=3,
        value="Unlevered earnings rate vs cost basis",
    ).font = FONT_HINT
    cur += 1

    # Row 2: Going-In Cap Value — formula tracks LP edits to NOI
    # (via s_combined_noi) and Going-In Cap rate (s_going_in_cap_rate).
    ws.cell(
        row=cur, column=1, value="Going-In Cap Value (NOI ÷ Going-In Cap)"
    ).font = FONT_LABEL
    cell = ws.cell(
        row=cur, column=2,
        value='=IFERROR(s_combined_noi/s_going_in_cap_rate,"")',
    )
    cell.font = FONT_VALUE
    cell.alignment = ALIGN_RIGHT
    cell.number_format = ACCOUNTING
    registry.register("s_going_in_cap_value", ws.title, cur, 2)
    ws.cell(
        row=cur, column=3,
        value=f"Market value at acquisition ({going_in_cap_pct_raw}% cap)",
    ).font = FONT_HINT
    cur += 1

    # Row 3: Exit Cap Value (= Direct Cap, kept name for back-compat).
    # Formula tracks exit-year NOI (s_exit_year_noi) + Exit Cap (s_exit_cap_rate).
    ws.cell(
        row=cur, column=1, value="Exit Cap Value (NOI ÷ Exit Cap)"
    ).font = FONT_LABEL
    cell = ws.cell(
        row=cur, column=2,
        value='=IFERROR(s_exit_year_noi/s_exit_cap_rate,"")',
    )
    cell.font = FONT_VALUE
    cell.alignment = ALIGN_RIGHT
    cell.number_format = ACCOUNTING
    registry.register("s_direct_cap_value", ws.title, cur, 2)
    ws.cell(
        row=cur, column=3,
        value=f"Market value at exit ({exit_cap_pct_raw}% cap, Y{_exit_yr} NOI)",
    ).font = FONT_HINT
    cur += 1

    # Row 4: Cap Spread (Yield on Cost − Going-In Cap) — both operands
    # are named cells, so spread re-derives on LP edit.
    ws.cell(row=cur, column=1, value="Cap Spread (Yield − Going-In Cap)").font = FONT_LABEL
    cell = ws.cell(
        row=cur, column=2,
        value='=IFERROR(s_yield_on_cost-s_going_in_cap_rate,"")',
    )
    cell.font = FONT_VALUE
    cell.alignment = ALIGN_RIGHT
    cell.number_format = PCT
    registry.register("s_cap_spread", ws.title, cur, 2)
    if yield_on_cost is not None and going_in_cap_pct_raw > 0:
        cap_spread = yield_on_cost - (going_in_cap_pct_raw / Decimal(100))
        ws.cell(
            row=cur, column=3,
            value=("Yield premium" if cap_spread > 0 else "Yield discount"),
        ).font = FONT_HINT
    cur += 1

    # Row 5: DCF NPV — PV of levered cash flows less equity invested.
    # Reconciles against Direct Cap: NPV > 0 means the asset generates
    # returns above the hurdle; NPV < 0 means it doesn't clear the bar
    # at the stated price.  Pairs with Weighted EM above.
    _disc_rate_pct = _coerce_decimal(ctx.get("discount_rate_pct") or Decimal("8.0"))
    _equity_req_val = _coerce_decimal(totals.get("equity_required") or 0)
    npv_lev = _npv_levered(rollup_waterfall, capital_modules, _equity_req_val, _disc_rate_pct)
    ws.cell(
        row=cur, column=1,
        value=f"DCF NPV ({_disc_rate_pct:.2f}% hurdle)",
    ).font = FONT_LABEL
    if npv_lev is not None:
        registry.write(
            ws, cur, 2, npv_lev,
            name="s_dcf_npv", fmt=ACCOUNTING,
            font=FONT_VALUE, align=ALIGN_RIGHT,
        )
        ws.cell(
            row=cur, column=3,
            value=("Value created above hurdle" if npv_lev > 0 else "Return below hurdle"),
        ).font = FONT_HINT
    else:
        ws.cell(row=cur, column=2, value=_DASH).font = FONT_VALUE
        ws.cell(
            row=cur, column=3,
            value="(set equity + Discount Rate in Deal Settings)",
        ).font = FONT_HINT
        registry.register("s_dcf_npv", ws.title, cur, 2)
    cur += 1

    # ── Per-Year Returns Matrix ────────────────────────────────────────────
    # BIW pattern (Building_I_Want v5): a year-by-year grid of the metrics
    # an LP uses to size up a deal at a glance. Each "year N" column is the
    # metric AS OF year N, computed two ways:
    #   - Cash-based metrics (NOI, OpEx, OER, Levered/Unlevered CF):
    #     just that year's value.
    #   - Cumulative / IRR metrics (Cumulative CF, Lev/Unlev IRR-if-exit):
    #     computed over the cash-flow window from period 0 through end of
    #     year N, with a simulated exit in the last period equal to
    #     year-N NOI ÷ exit cap rate.
    # The IRR-if-exit columns are particularly useful — they show "if you
    # bailed at Y3, your IRR would be X%" so the LP can see when the deal
    # crosses its hurdle.
    matrix_row = cur + 2
    cur = _build_per_year_returns_matrix(
        ws, registry, matrix_row, ctx, per_project=per_project, totals=totals,
    )

    freeze_top(ws, row=2)
    print_landscape(ws)


def _build_per_year_returns_matrix(
    ws,
    registry: CellRegistry,
    start_row: int,
    ctx: dict,
    *,
    per_project: list[dict],
    totals: dict,
) -> int:
    """Render the BIW-style per-year matrix and return the next-free row.

    Columns are Y1, Y2, … up to the hold horizon (capped at 10 for a
    Underwriting-Summary skim view; the full series is on the Cash Flow
    sheet). Rows split into two visual groups: per-year cash metrics
    (NOI / OER / Levered / Unlevered) and cumulative-through-this-year
    metrics (Cumulative CF, IRR-if-exit-at-Y_N).
    """
    cash_flows: dict[UUID, list[CashFlow]] = ctx["cash_flows"]
    cash_flow_items: dict[UUID, list[CashFlowLineItem]] = ctx["cash_flow_items"]
    operational_inputs: dict[UUID, OperationalInputs] = ctx["operational_inputs"]
    projects: list[Project] = ctx["projects"]

    annual = _aggregate_scenario_annual(cash_flows)
    if not annual:
        ws.cell(
            row=start_row, column=1,
            value="(no compute output — Per-Year Returns Matrix populates after Compute)",
        ).font = FONT_HINT
        return start_row + 1

    max_year_modeled = max(annual)
    # Skip Y0 (acquisition stub) — investor-friendly read starts at Y1.
    year_cols = [y for y in range(1, max_year_modeled + 1) if y <= 10]
    if not year_cols:
        return start_row

    combined_tpc = _coerce_decimal(totals.get("total_project_cost") or 0)
    default_inputs = operational_inputs.get(projects[0].id) if projects else None
    exit_cap_pct = _coerce_decimal(
        getattr(default_inputs, "exit_cap_rate_pct", 0) or 0
    )

    # Period-totals cache for IRR computations (sum across projects per period).
    period_ncf: dict[int, Decimal] = {}
    period_ds: dict[int, Decimal] = {}
    period_noi: dict[int, Decimal] = {}
    for cf_list in cash_flows.values():
        for cf in cf_list:
            p = cf.period
            period_ncf[p] = period_ncf.get(p, Decimal(0)) + _coerce_decimal(cf.net_cash_flow or 0)
            period_ds[p] = period_ds.get(p, Decimal(0)) + _coerce_decimal(cf.debt_service or 0)
            period_noi[p] = period_noi.get(p, Decimal(0)) + _coerce_decimal(cf.noi or 0)
    # Equity calls (capital_outflow) by period — needed for Cash-on-Cash
    # denominators. Sum signed capital events per project across periods.
    period_signed_events = _signed_capital_events_by_year(cash_flow_items)

    set_widths(ws, [30, *([14] * len(year_cols))])
    section_label(
        ws, start_row, "Per-Year Returns Matrix (BIW-style)",
        span_cols=len(year_cols) + 1,
    )
    header_row(ws, start_row + 1, ["Metric", *[f"Y{y}" for y in year_cols]])

    cur = start_row + 2

    def write_row(label: str, values: dict[int, Decimal | None], range_name: str | None,
                  fmt: str = ACCOUNTING) -> None:
        nonlocal cur
        ws.cell(row=cur, column=1, value=label).font = FONT_LABEL
        for col_offset, year in enumerate(year_cols):
            value = values.get(year)
            cell = ws.cell(
                row=cur, column=2 + col_offset,
                value=_to_excel_number(value) if value is not None else _DASH,
            )
            if value is not None:
                cell.number_format = fmt
            cell.font = FONT_VALUE
            cell.alignment = ALIGN_RIGHT
        if range_name and year_cols:
            registry.register_range(
                range_name, ws.title, cur, cur, col=2,
                end_col=1 + len(year_cols),
            )
        cur += 1

    # Row 1: NOI per year
    noi_per_year = {y: annual.get(y, {}).get("noi", Decimal(0)) for y in year_cols}
    write_row("NOI", noi_per_year, "r_uw_matrix_noi")

    # Row 2: Cap on Cost = NOI[Y] / TPC per year
    cap_on_cost = {
        y: (noi_per_year[y] / combined_tpc) if combined_tpc > 0 else None
        for y in year_cols
    }
    write_row("Cap on Cost (NOI ÷ TPC)", cap_on_cost, "r_uw_matrix_cap_on_cost", fmt=PCT)

    # Row 3: OER per year
    oer_per_year = {}
    for y in year_cols:
        opex = annual.get(y, {}).get("operating_expenses", Decimal(0))
        egi = annual.get(y, {}).get("effective_gross_income", Decimal(0))
        oer_per_year[y] = (opex / egi) if egi > 0 else None
    write_row("OER (OpEx ÷ EGI)", oer_per_year, "r_uw_matrix_oer", fmt=PCT)

    # Row 4: Levered Cash Flow per year
    levered_per_year = {y: annual.get(y, {}).get("net_cash_flow", Decimal(0)) for y in year_cols}
    write_row("Levered Cash Flow", levered_per_year, "r_uw_matrix_levered_cf")

    # Row 5: Unlevered CF per year (= NCF + DS)
    unlevered_per_year = {
        y: annual.get(y, {}).get("net_cash_flow", Decimal(0))
           + annual.get(y, {}).get("debt_service", Decimal(0))
        for y in year_cols
    }
    write_row("Unlevered Cash Flow", unlevered_per_year, "r_uw_matrix_unlevered_cf")

    # Row 6: Cumulative Levered CF through year N
    cumulative_levered = {}
    running = Decimal(0)
    for y in year_cols:
        running += levered_per_year.get(y, Decimal(0))
        cumulative_levered[y] = running
    write_row("Cumulative Levered CF", cumulative_levered, "r_uw_matrix_cumulative_levered")

    # Row 7: Levered IRR-if-exit-at-Y_N
    # Build once per year: NCF[0..N*12] with the last period augmented by
    # (NOI[Y_N] * 12 / exit_cap) — a simulated sale at year-N's stabilized
    # cap value. Engine NCF already nets the actual debt payoff at exit
    # (when project actually exits at Y_N), but for years before exit we
    # have to simulate.
    levered_irr_per_year = _per_year_irr_series(
        period_ncf, period_noi, year_cols, exit_cap_pct,
    )
    write_row("Levered IRR (if exit at Y)", levered_irr_per_year, "r_uw_matrix_levered_irr", fmt=PCT)

    # Row 8: Unlevered IRR-if-exit-at-Y_N — same but using NCF + DS
    period_unlev = {p: period_ncf.get(p, Decimal(0)) + period_ds.get(p, Decimal(0))
                    for p in period_ncf}
    unlevered_irr_per_year = _per_year_irr_series(
        period_unlev, period_noi, year_cols, exit_cap_pct,
    )
    write_row("Unlevered IRR (if exit at Y)", unlevered_irr_per_year, "r_uw_matrix_unlevered_irr", fmt=PCT)

    # Suppress an unused-local lint flag while keeping the variable
    # documented for future Cash-on-Cash extensions.
    _ = period_signed_events

    return cur + 1


def _per_year_irr_series(
    period_cf: dict[int, Decimal],
    period_noi: dict[int, Decimal],
    year_cols: list[int],
    exit_cap_pct: Decimal,
) -> dict[int, Decimal | None]:
    """Compute IRR-if-exit-at-Y_N for each year in ``year_cols``.

    For each year N, take the period cash flow series from period 0
    through period N×12, replace the last period's value with
    ``cf + simulated_exit`` where ``simulated_exit = NOI(Y_N) ÷ exit_cap``.
    Returns the IRR as a fraction (PCT-format ready), or None when the
    series has no sign change / no exit cap configured / pyxirr unavailable.
    """
    from app.engines.cashflow import _compute_xirr  # late import — keep module imports lean

    out: dict[int, Decimal | None] = {}
    if exit_cap_pct <= 0:
        return {y: None for y in year_cols}

    for year_n in year_cols:
        max_period = year_n * 12
        # Year-N annualized NOI: sum of monthly NOI in months (year_n*12 - 11)..year_n*12
        ytd_noi_annual = sum(
            (period_noi.get(p, Decimal(0)) for p in range(max_period - 11, max_period + 1)),
            Decimal(0),
        )
        if ytd_noi_annual <= 0:
            out[year_n] = None
            continue
        simulated_exit = ytd_noi_annual * Decimal(100) / exit_cap_pct

        # Build clipped + augmented series
        series: list[Decimal] = []
        for p in sorted(period_cf):
            if p > max_period:
                break
            value = period_cf[p]
            if p == max_period:
                value = value + simulated_exit
            series.append(value)
        if not series:
            out[year_n] = None
            continue
        pct_whole = _compute_xirr(series)
        if pct_whole == 0:
            out[year_n] = None
        else:
            out[year_n] = pct_whole / Decimal(100)
    return out


# ── Underwriting Pro Forma sheet ──────────────────────────────────────────────


def _build_uw_proforma(ws, registry: CellRegistry, ctx: dict) -> None:
    """Annual P&L summed across projects: Y0 → Y10 (or longest hold)."""
    cash_flows: dict[UUID, list[CashFlow]] = ctx["cash_flows"]
    cash_flow_items: dict[UUID, list[CashFlowLineItem]] = ctx["cash_flow_items"]

    annual = _aggregate_scenario_annual(cash_flows)
    max_year = min(max(annual) if annual else 0, 10)
    year_cols = list(range(0, max(max_year, 1) + 1))

    set_widths(ws, [30, *([14] * (len(year_cols) + 1))])

    section_label(ws, 1, "Pro Forma — Annual P&L (combined across projects)", span_cols=len(year_cols) + 1)
    header_row(ws, 2, ["Line Item", *[f"Y{y}" for y in year_cols]])

    rows: list[tuple[str, str, str | None]] = [
        ("Gross Revenue", "gross_revenue", "r_uw_gross_revenue"),
        ("Vacancy Loss", "vacancy_loss", None),
        ("Effective Gross Income", "effective_gross_income", "r_uw_egi"),
        ("Operating Expenses", "operating_expenses", "r_uw_opex"),
        ("CapEx Reserve", "capex_reserve", None),
        ("NOI", "noi", "r_uw_noi"),
        # Phase C: Asset Mgmt Fee shown explicitly so the LP sees the
        # sponsor's fee drag against NOI. Always rendered as a formula
        # ``=IFERROR(-EGI*s_asset_mgmt_fee,0)`` — engine doesn't push this
        # value to the annual rollup, so the row is purely informational
        # and never carries an engine seed.
        # No range name registered — informational row; named ranges on this
        # sheet are reserved for engine-derived KPIs.
        ("Asset Mgmt Fee", "asset_mgmt_fee", None),
        ("Debt Service", "debt_service", "r_uw_debt_service"),
        ("Net Cash Flow", "net_cash_flow", "r_uw_net_cash_flow"),
    ]

    from openpyxl.utils import get_column_letter

    # Formula-conversion plan §4.1 (commit 3): the rows whose math is a
    # direct sum/difference of other rows on this sheet become formulas
    # so LP edits to a single Use line / OpEx category propagate to the
    # downstream KPIs. ``effective_gross_income`` = GrossRev + Vacancy
    # (vacancy negative); ``noi`` = EGI - OpEx. Net Cash Flow + Debt
    # Service stay as engine values until commit 4 wires the Debt
    # Schedule sheet, then they convert too.
    _DERIVED_FORMULA_FIELDS: dict[str, tuple[str, ...]] = {
        # field name -> (sign, operand_field) pairs. Sign is "+" or "-"
        # and applied to the operand field's cell on this same row block.
        "effective_gross_income": ("+gross_revenue", "+vacancy_loss"),
        "noi": ("+effective_gross_income", "-operating_expenses"),
        # Phase E: Net Cash Flow is derived so the new Debt Service formula
        # propagates through to the LP's bottom-line cash row.
        "net_cash_flow": ("+noi", "-debt_service"),
    }
    # Phase B: rows whose Y_n>=2 values are a growth-rate chain on the
    # prior year. Y0/Y1 stay as engine-computed seed values; Y2+ become
    # ``=prev * (1 + s_opex_growth_rate)`` so a single Assumptions edit
    # ripples through every downstream year. CapEx Reserve uses the same
    # OpEx growth rate (no separate reserve-growth assumption today).
    _GROWTH_CHAIN_FIELDS: dict[str, str] = {
        "gross_revenue": "s_revenue_growth_rate",
        "operating_expenses": "s_opex_growth_rate",
        "capex_reserve": "s_opex_growth_rate",
    }
    # Phase E: Debt Service Y1+ becomes ``=SUM(s_loan_{i}_annual_pi, ...)``
    # over the PMT-eligible loan named ranges so an LP changing principal
    # or rate on the Debt Schedule sees the Pro Forma debt service flow.
    # Only meaningful when the profile renders Debt Schedule AND at least
    # one ``pi``-carry loan with rate + principal exists. Y0 stays at the
    # engine value (construction-phase debt service often differs from the
    # stabilized PMT). Approximation caveat: assumes the PMT loans are
    # active in every year of the chain — fine for the common
    # single-perm-debt stack; over-states debt service in years where a
    # perm loan hasn't funded yet on a construction-to-perm stack.
    _capital_modules: list[CapitalModule] = ctx["capital_modules"]
    _profile = ctx.get("_profile")
    _pmt_indices = (
        _pmt_loan_indices(_capital_modules)
        if _profile in {"internal", "lender"}
        else []
    )
    # Loans whose Debt Schedule row registered an
    # ``s_loan_<n>_perm_origination_month`` cell. PI contributions for
    # these loans are additionally gated on the year being past perm
    # origination — closes the construction-to-perm overstatement gap
    # called out below.
    _perm_gated_loan_idxs: set[int] = (
        set(_perm_origination_loan_idxs(ctx).keys())
        if _profile in {"internal", "lender"}
        else set()
    )

    # Phase 2: Y1 Gross Revenue + Y1 OpEx become formulas referencing
    # Assumptions Block F / G named cells. LP edits to a stream's rent
    # or an OpEx line's annual amount now ripple forward via the Y2+
    # growth chain that already exists. Y0 (pre-op) stays at engine
    # value — construction-phase gross/opex are governed by the
    # cashflow engine, not the stabilized inputs.
    _rev_slug_list: list[str] = list(_all_revenue_slugs(ctx).values())
    _opex_slug_list: list[str] = list(_all_opex_slugs(ctx).values())

    def _gross_revenue_y1_formula() -> str | None:
        """SUM of every stream's Y1 monthly cell × 12.

        Returns None when no streams exist (defensive — Pro Forma still
        shows engine value, which will be 0 anyway). Wrapped in
        parentheses so the ``*12`` annualization applies to the whole
        SUM rather than only the last term.
        """
        if not _rev_slug_list:
            return None
        refs = ",".join(f"s_rev_{s}_y1_monthly" for s in _rev_slug_list)
        return f"=SUM({refs})*12"

    def _opex_y1_formula() -> str | None:
        """SUM of every OpEx line's annual cell."""
        if not _opex_slug_list:
            return None
        refs = ",".join(f"s_opex_{s}_annual" for s in _opex_slug_list)
        return f"=SUM({refs})"

    def _debt_service_formula_for_year(y: int) -> str | None:
        """Per-year Debt Service SUM, gated by term + perm origination.

        For year Y (1-indexed), each PI loan contributes its annual P&I
        only when both conditions hold:

          1. Loan's hold term covers the end of the year
             (``s_loan_{i}_term_months >= Y*12``) — a 10-year hold loan
             drops out of the SUM starting in Y11 instead of overstating
             debt service forever.

          2. Year-end is past perm origination
             (``Y*12 >= s_loan_{i}_perm_origination_month``) — for
             construction-to-perm loans, PI doesn't accrue during the
             construction window. Skipped for loans without a registered
             perm cell (pure-perm loans, single-project workbooks, or
             pure-acquisition scenarios) — those keep the legacy
             term-only gate.

        Loans whose term cell is em-dash / blank fail the >= test
        silently and contribute 0, which is the safe graceful-degradation
        outcome.
        """
        if not _pmt_indices:
            return None
        year_end_months = y * 12
        terms = []
        for i in _pmt_indices:
            if i in _perm_gated_loan_idxs:
                terms.append(
                    f"IF(AND(s_loan_{i}_term_months>={year_end_months},"
                    f"{year_end_months}>=s_loan_{i}_perm_origination_month),"
                    f"s_loan_{i}_annual_pi,0)"
                )
            else:
                terms.append(
                    f"IF(s_loan_{i}_term_months>={year_end_months},"
                    f"s_loan_{i}_annual_pi,0)"
                )
        return "=" + "+".join(terms)
    # Track each engine-value row's coord so derived formulas can reference
    # the actual cells. col=2 is Y0; each year_cols entry adds one column.
    field_row: dict[str, int] = {}

    cur_row = 3
    for label, field, range_name in rows:
        ws.cell(row=cur_row, column=1, value=label).font = FONT_LABEL
        field_row[field] = cur_row
        derived = _DERIVED_FORMULA_FIELDS.get(field)
        growth_name = _GROWTH_CHAIN_FIELDS.get(field)
        for col_offset, year in enumerate(year_cols):
            col_idx = 2 + col_offset
            if derived:
                # Build per-year formula like ``=B3+B4`` referencing the
                # corresponding column on the operand rows. Operand rows
                # are guaranteed to be written before any derived row
                # because the rows table orders inputs before derivations.
                operands: list[str] = []
                for spec in derived:
                    sign, operand_field = spec[0], spec[1:]
                    operand_row = field_row.get(operand_field)
                    if operand_row is None:
                        # Defensive: operand not yet written. Fall back to
                        # engine value so we don't emit a broken formula.
                        operands = []
                        break
                    col_letter = get_column_letter(col_idx)
                    operands.append(f"{sign}{col_letter}{operand_row}")
                if operands:
                    formula = "=" + "".join(operands).lstrip("+")
                    cell = ws.cell(row=cur_row, column=col_idx, value=formula)
                else:
                    value = annual.get(year, {}).get(field, Decimal(0))
                    cell = ws.cell(
                        row=cur_row, column=col_idx,
                        value=_to_excel_number(value),
                    )
            elif (
                field == "debt_service"
                and _pmt_indices
                and col_offset >= 1
            ):
                # Phase E: Y2+ debt service = SUM of per-loan annual P&I
                # named ranges, gated by each loan's term_months so loans
                # past their hold term drop out instead of overstating.
                ds_formula = _debt_service_formula_for_year(year)
                cell = ws.cell(
                    row=cur_row, column=col_idx, value=ds_formula,
                )
            elif field == "asset_mgmt_fee":
                # Phase C: Asset Mgmt Fee = -EGI * s_asset_mgmt_fee, every
                # year. IFERROR guards a missing s_asset_mgmt_fee gracefully.
                egi_r = field_row.get("effective_gross_income")
                if egi_r is not None:
                    col_letter = get_column_letter(col_idx)
                    formula = (
                        f"=IFERROR(-{col_letter}{egi_r}*s_asset_mgmt_fee,0)"
                    )
                    cell = ws.cell(row=cur_row, column=col_idx, value=formula)
                else:
                    cell = ws.cell(
                        row=cur_row, column=col_idx,
                        value=_to_excel_number(Decimal(0)),
                    )
            elif (
                field == "gross_revenue"
                and col_offset == 1
                and _gross_revenue_y1_formula() is not None
                and Decimal(annual.get(year, {}).get("gross_revenue", 0) or 0) > 0
            ):
                # Phase 2: Y1 Gross Revenue = SUM(Block F y1_monthly cells) * 12,
                # but only when the engine shows Y1 is operationally active
                # (gross_revenue > 0). When Y1 is still in a construction
                # phase the engine returns 0; overriding with the stabilized
                # input formula would misrepresent phase semantics, so we
                # keep the engine seed (the growth chain picks up at the
                # first stabilized year via the Y2+ branch).
                cell = ws.cell(
                    row=cur_row, column=col_idx,
                    value=_gross_revenue_y1_formula(),
                )
            elif (
                field == "operating_expenses"
                and col_offset == 1
                and _opex_y1_formula() is not None
                and Decimal(annual.get(year, {}).get("operating_expenses", 0) or 0) > 0
            ):
                # Phase 2: Y1 OpEx = SUM(Block G annual cells), same
                # phase-gating contract as gross_revenue above — engine
                # value wins when Y1 is pre-operational.
                cell = ws.cell(
                    row=cur_row, column=col_idx,
                    value=_opex_y1_formula(),
                )
            elif growth_name and col_offset >= 2:
                # Y2+ for growth-chain fields: reference the prior-year
                # cell on the same row and multiply by (1 + growth). Y0
                # (pre-op) keeps engine values as the chain's seed; Y1
                # is now also formula-driven via the input-reference
                # branches above (gross_revenue, operating_expenses).
                prev_col = get_column_letter(col_idx - 1)
                formula = f"={prev_col}{cur_row}*(1+{growth_name})"
                cell = ws.cell(row=cur_row, column=col_idx, value=formula)
            else:
                value = annual.get(year, {}).get(field, Decimal(0))
                cell = ws.cell(
                    row=cur_row, column=col_idx,
                    value=_to_excel_number(value),
                )
            cell.number_format = ACCOUNTING
            cell.font = FONT_VALUE
            cell.alignment = ALIGN_RIGHT
        if range_name and year_cols:
            registry.register_range(
                range_name,
                ws.title,
                cur_row,
                cur_row,
                col=2,
                end_col=1 + len(year_cols),
            )
        # Phase 3: emit one indented breakout row per IncomeStream
        # directly below the Gross Revenue total, mirroring the OpEx
        # breakout block. Y0 blank (construction-phase revenue is
        # engine-governed). Y1 = the stream's `s_rev_<slug>_y1_monthly`
        # × 12, Y2+ = prior-year × (1 + per-stream
        # `s_rev_<slug>_escalation_pct`). Per-stream escalation so a
        # rent-controlled 1BR and market-rate 2BR can ramp at their own
        # rates instead of collapsing onto `s_revenue_growth_rate`.
        if field == "gross_revenue" and _rev_slug_list:
            rev_streams = _all_revenue_streams_ordered(ctx)
            rev_slug_map = _all_revenue_slugs(ctx)
            for stream in rev_streams:
                slug = rev_slug_map.get(stream.id)
                if slug is None:
                    continue
                cur_row += 1
                label = f"   • {stream.label or 'Income Stream'}"
                ws.cell(row=cur_row, column=1, value=label).font = FONT_HINT
                for col_offset, _year in enumerate(year_cols):
                    col_idx = 2 + col_offset
                    if col_offset == 0:
                        cell = ws.cell(row=cur_row, column=col_idx, value=None)
                    elif col_offset == 1:
                        cell = ws.cell(
                            row=cur_row, column=col_idx,
                            value=f"=s_rev_{slug}_y1_monthly*12",
                        )
                    else:
                        prev_col = get_column_letter(col_idx - 1)
                        cell = ws.cell(
                            row=cur_row, column=col_idx,
                            value=(
                                f"={prev_col}{cur_row}"
                                f"*(1+s_rev_{slug}_escalation_pct)"
                            ),
                        )
                    cell.number_format = ACCOUNTING
                    cell.font = FONT_HINT
                    cell.alignment = ALIGN_RIGHT
        # Phase B follow-up: expose Y1 OpEx as a workbook-scoped name so
        # the S&U Operating Reserve UseLine formula resolves.
        if field == "operating_expenses" and len(year_cols) >= 2:
            registry.register("s_y1_opex", ws.title, cur_row, 3)
        # Phase 4 completion: expose Y1 Gross Revenue + Y1 NOI cells as
        # workbook-scoped names so UW Summary's Combined Stabilized NOI
        # (and any future input-derived KPI) can reference the actual
        # formula chain — Block F/G edits ripple through Pro Forma Y1 →
        # these names → downstream KPIs.
        if field == "gross_revenue" and len(year_cols) >= 2:
            registry.register("s_pf_gross_revenue_y1", ws.title, cur_row, 3)
        if field == "noi" and len(year_cols) >= 2:
            registry.register("s_pf_noi_y1", ws.title, cur_row, 3)
        if field == "debt_service" and len(year_cols) >= 2:
            registry.register("s_pf_debt_service_y1", ws.title, cur_row, 3)
        # Phase 3: emit one indented breakout row per OpEx line directly
        # below the Operating Expenses total. Each row references the
        # corresponding Block G cell so an LP can trace any single
        # expense back to its input — and edit any one of them to see
        # the total row (formula-driven since Phase 2) follow. Total row
        # itself is left untouched; this is a transparency block. Y0
        # stays blank because construction-phase OpEx is governed by
        # the engine, not these stabilized inputs.
        if field == "operating_expenses" and _opex_slug_list:
            opex_lines = _all_opex_lines_ordered(ctx)
            opex_slug_map = _all_opex_slugs(ctx)
            for line in opex_lines:
                slug = opex_slug_map.get(line.id)
                if slug is None:
                    continue
                cur_row += 1
                label = f"   • {line.label or 'Operating Expense'}"
                ws.cell(row=cur_row, column=1, value=label).font = FONT_HINT
                for col_offset, _year in enumerate(year_cols):
                    col_idx = 2 + col_offset
                    if col_offset == 0:
                        # Y0 blank — construction-phase OpEx differs from
                        # stabilized inputs.
                        cell = ws.cell(row=cur_row, column=col_idx, value=None)
                    elif col_offset == 1:
                        cell = ws.cell(
                            row=cur_row, column=col_idx,
                            value=f"=s_opex_{slug}_annual",
                        )
                    else:
                        prev_col = get_column_letter(col_idx - 1)
                        cell = ws.cell(
                            row=cur_row, column=col_idx,
                            value=(
                                f"={prev_col}{cur_row}"
                                f"*(1+s_opex_{slug}_escalation_pct)"
                            ),
                        )
                    cell.number_format = ACCOUNTING
                    cell.font = FONT_HINT
                    cell.alignment = ALIGN_RIGHT
        # Phase D follow-up: expose the exit-year NOI cell (last column
        # of the NOI row) so the UW Summary Exit Cap Value formula can
        # reference a single named cell instead of duplicating the
        # ``last 12 months of NOI`` aggregation engine-side.
        if field == "noi" and year_cols:
            exit_col = 1 + len(year_cols)
            registry.register("s_exit_year_noi", ws.title, cur_row, exit_col)
        cur_row += 1

    # OER (Operating Expense Ratio) = OpEx / EGI per year. Standard CRE
    # operating-efficiency metric — typical multifamily targets 35-45%; a
    # number above 50% is a yellow flag for the LP. Rendered as a derived
    # ratio row immediately below CapEx Reserve, before the NOI line.
    # Phase B: formula-driven so OpEx growth-chain edits flow into OER.
    ws.cell(row=cur_row, column=1, value="OER (OpEx ÷ EGI)").font = FONT_LABEL
    opex_r = field_row.get("operating_expenses")
    egi_r = field_row.get("effective_gross_income")
    for col_offset, _year in enumerate(year_cols):
        col_idx = 2 + col_offset
        if opex_r is not None and egi_r is not None:
            col_letter = get_column_letter(col_idx)
            formula = f"=IFERROR({col_letter}{opex_r}/{col_letter}{egi_r},\"\")"
            cell = ws.cell(row=cur_row, column=col_idx, value=formula)
        else:
            cell = ws.cell(row=cur_row, column=col_idx, value=_DASH)
        cell.number_format = PCT
        cell.font = FONT_VALUE
        cell.alignment = ALIGN_RIGHT
    if year_cols:
        registry.register_range(
            "r_uw_oer", ws.title, cur_row, cur_row,
            col=2, end_col=1 + len(year_cols),
        )
    cur_row += 1

    # Revenue + OpEx breakouts: separate tables driven by line items.
    # Category-aware aggregation — Revenue from `income` line items
    # (per-stream labels, Option C: same label across projects = one row),
    # OpEx from `expense` line items (per-category labels). Capital events
    # show on the Underwriting Cash Flow sheet, not here.
    by_category = _aggregate_scenario_line_items_by_category(cash_flow_items)

    cur_row += 1
    cur_row = _write_breakout_table(
        ws, registry, cur_row,
        title="Revenue Breakout (by stream)",
        rows=by_category.get("income", {}),
        year_cols=year_cols,
        empty_hint="(no revenue line items recorded — run Compute to populate)",
    )

    cur_row += 1
    cur_row = _write_breakout_table(
        ws, registry, cur_row,
        title="OpEx Breakout (by category)",
        rows=by_category.get("expense", {}),
        year_cols=year_cols,
        empty_hint="(no OpEx line items recorded — run Compute to populate)",
        always_show=ALWAYS_SHOWN_OPEX_CATEGORIES,
    )

    freeze_top(ws, row=3)
    print_landscape(ws)


def _write_breakout_table(
    ws,
    registry: CellRegistry,
    start_row: int,
    *,
    title: str,
    rows: dict[int, dict[str, Decimal]],
    year_cols: list[int],
    empty_hint: str,
    always_show: tuple[str, ...] = (),
) -> int:
    """Render one labelled annual-buckets table and return the row after it.

    Shared by the Pro Forma sheet's Revenue and OpEx sections so they
    stay visually identical. ``rows`` shape mirrors the per-category
    output from ``_aggregate_scenario_line_items_by_category``: a
    ``{year: {label: amount}}`` dict for the chosen category.

    ``always_show`` is a tuple of canonical labels that must appear even
    when their year totals are zero. Used by the OpEx breakout to surface
    universal multifamily categories (Real Estate Taxes, Insurance,
    Property Management) so a missing line is *visible* — a CRE LP
    immediately notices a $0 Property Tax row and asks; an *absent*
    Property Tax row is silent and dangerous.
    """
    section_label(ws, start_row, title, span_cols=len(year_cols) + 1)
    cur = start_row + 1

    # Keep a label if (a) it's in the always-show list, OR (b) any of its
    # years has a non-zero amount. Drops typo placeholder rows ("$0 across
    # the board, non-canonical name") while keeping universal-vocabulary
    # rows visible even when missing data.
    always_set = set(always_show)
    labels = sorted({
        label
        for year_data in rows.values()
        for label in year_data
        if label in always_set
        or any(rows.get(y, {}).get(label, Decimal(0)) != 0 for y in year_cols)
    } | always_set)
    if not labels:
        ws.cell(row=cur, column=1, value=empty_hint).font = FONT_HINT
        return cur + 1

    for label in labels:
        ws.cell(row=cur, column=1, value=label).font = FONT_VALUE
        for col_offset, year in enumerate(year_cols):
            value = rows.get(year, {}).get(label, Decimal(0))
            cell = ws.cell(row=cur, column=2 + col_offset, value=_to_excel_number(value))
            cell.number_format = ACCOUNTING
            cell.font = FONT_VALUE
            cell.alignment = ALIGN_RIGHT
        cur += 1
    _ = registry  # accepted for future per-row named ranges; not used today
    return cur


# ── Underwriting Cash Flow sheet ──────────────────────────────────────────────


def _build_uw_cashflow(ws, registry: CellRegistry, ctx: dict) -> None:
    """Annual cash flow: NOI / Capital Events / Levered / Unlevered / DS / DSCR / Cum LCF."""
    from openpyxl.utils import get_column_letter

    cash_flows: dict[UUID, list[CashFlow]] = ctx["cash_flows"]
    cash_flow_items: dict[UUID, list[CashFlowLineItem]] = ctx["cash_flow_items"]

    annual = _aggregate_scenario_annual(cash_flows)
    # Signed capital events — outflows negative, inflows positive — so the
    # row reads correctly for an investor (Y0 acquisition shows -$X, exit
    # shows +$Y). See _signed_capital_events_by_year docstring.
    capital_events_by_year = _signed_capital_events_by_year(cash_flow_items)
    max_year = min(max(annual) if annual else 0, 10)
    year_cols = list(range(0, max(max_year, 1) + 1))

    set_widths(ws, [30, *([14] * len(year_cols))])
    section_label(
        ws, 1, "Cash Flow — Annual (scenario-wide)", span_cols=len(year_cols) + 1
    )
    header_row(ws, 2, ["Line Item", *[f"Y{y}" for y in year_cols]])

    cur_row = 3
    # Track row index per series key so derived rows (DSCR, Unlevered,
    # Cumulative, Levered) can reference the actual cells by column letter
    # rather than emit hard-coded engine values. Keys are local labels
    # (noi, capital_events, debt_proceeds, debt_service, levered, etc.).
    series_row: dict[str, int] = {}

    def write_series(
        label: str,
        source: dict[int, Decimal],
        range_name: str | None,
        fmt: str = ACCOUNTING,
        series_key: str | None = None,
    ) -> None:
        nonlocal cur_row
        ws.cell(row=cur_row, column=1, value=label).font = FONT_LABEL
        for col_offset, year in enumerate(year_cols):
            value = source.get(year, Decimal(0))
            cell = ws.cell(row=cur_row, column=2 + col_offset, value=_to_excel_number(value))
            cell.number_format = fmt
            cell.font = FONT_VALUE
            cell.alignment = ALIGN_RIGHT
        if range_name and year_cols:
            registry.register_range(
                range_name, ws.title, cur_row, cur_row, col=2,
                end_col=1 + len(year_cols),
            )
        if series_key:
            series_row[series_key] = cur_row
        cur_row += 1

    def write_formula_series(
        label: str,
        formula_fn,
        range_name: str | None,
        fmt: str = ACCOUNTING,
        series_key: str | None = None,
    ) -> None:
        """Write a derived row where each year cell is a formula string.

        ``formula_fn(col_letter)`` returns the formula text (including the
        leading ``=``) for one column. Mirrors write_series for layout
        (label col, year cols, registry range, optional series_key) but
        emits Excel formulas instead of engine scalars.
        """
        nonlocal cur_row
        ws.cell(row=cur_row, column=1, value=label).font = FONT_LABEL
        for col_offset, _year in enumerate(year_cols):
            col_idx = 2 + col_offset
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(row=cur_row, column=col_idx, value=formula_fn(col_letter))
            cell.number_format = fmt
            cell.font = FONT_VALUE
            cell.alignment = ALIGN_RIGHT
        if range_name and year_cols:
            registry.register_range(
                range_name, ws.title, cur_row, cur_row, col=2,
                end_col=1 + len(year_cols),
            )
        if series_key:
            series_row[series_key] = cur_row
        cur_row += 1

    noi_series = {y: annual.get(y, {}).get("noi", Decimal(0)) for y in year_cols}
    debt_series = {y: annual.get(y, {}).get("debt_service", Decimal(0)) for y in year_cols}
    ncf_series = {y: annual.get(y, {}).get("net_cash_flow", Decimal(0)) for y in year_cols}

    write_series("NOI", noi_series, "r_uw_cf_noi", series_key="noi")
    write_series(
        "Capital Events (acq + exit)", capital_events_by_year,
        "r_uw_cf_capital_events", series_key="capital_events",
    )

    # Debt proceeds drawn at acquisition close (Y0) — show as explicit inflow so
    # the LP can see: Acquisition Cost − Debt Proceeds = Net Equity Deployed.
    _junc_dp: dict = {}
    for _j in ctx.get("junctions", []):
        _junc_dp[_j.capital_module_id] = _junc_dp.get(
            _j.capital_module_id, Decimal(0)
        ) + _coerce_decimal(_j.amount or 0)
    _debt_y0 = Decimal(0)
    for _m in ctx.get("capital_modules", []):
        if _funder_class(_m) == "Debt":
            _debt_y0 += _junc_dp.get(_m.id) or _coerce_decimal(
                (_m.source or {}).get("amount") or 0
            )
    if _debt_y0 > Decimal(1):
        write_series(
            "Debt Proceeds (Y0 draws)", {0: _debt_y0},
            "r_uw_cf_debt_proceeds", series_key="debt_proceeds",
        )

    write_series(
        "Debt Service", debt_series, "r_uw_cf_debt_service",
        series_key="debt_service",
    )

    # Formula-conversion plan §4.2 (commit 4): Levered, Unlevered, DSCR,
    # and Cumulative are clean sum/diff/ratio derivations of the engine
    # rows written above. Wire them as formulas so an LP edit to NOI
    # propagates through the deal-level cash flow downstream.
    #
    # Levered CF = NOI + Capital Events [+ Debt Proceeds] - Debt Service.
    # Engine writes this as net_cash_flow already; building it from the
    # other rows here means an LP can see the arithmetic and edit any
    # operand to recompute.
    noi_row = series_row["noi"]
    capevt_row = series_row["capital_events"]
    debt_proceeds_row = series_row.get("debt_proceeds")
    ds_row = series_row["debt_service"]

    def _levered_formula(col_letter: str) -> str:
        terms = [f"{col_letter}{noi_row}", f"+{col_letter}{capevt_row}"]
        if debt_proceeds_row is not None:
            terms.append(f"+{col_letter}{debt_proceeds_row}")
        terms.append(f"-{col_letter}{ds_row}")
        return "=" + "".join(terms)

    write_formula_series(
        "Levered Cash Flow", _levered_formula, "r_uw_cf_levered",
        series_key="levered",
    )

    # Unlevered = Levered + Debt Service. Matches the engine path used by
    # _combined_unlevered_irr (NCF + DS). Reads as "cash flow before debt".
    lev_row = series_row["levered"]

    def _unlevered_formula(col_letter: str) -> str:
        return f"={col_letter}{lev_row}+{col_letter}{ds_row}"

    write_formula_series(
        "Unlevered Cash Flow", _unlevered_formula, "r_uw_cf_unlevered",
    )

    # DSCR = NOI / Debt Service per column. IFERROR guards Y0 / zero-DS
    # rows so a divide-by-zero shows as 0 instead of #DIV/0! in the
    # workbook — matches the engine behavior which returns 0 when ds is 0.
    def _dscr_formula(col_letter: str) -> str:
        return (
            f"=IFERROR({col_letter}{noi_row}/{col_letter}{ds_row},0)"
        )

    write_formula_series(
        "DSCR (annual)", _dscr_formula, "r_uw_cf_dscr", fmt="0.000",
    )

    # Cumulative = running SUM of Levered CF from Y0 through current col.
    # SUM($B$<lev_row>:<col_letter><lev_row>) lets Excel propagate edits
    # without us needing to redo the prefix sum here.
    def _cumulative_formula(col_letter: str) -> str:
        return f"=SUM($B${lev_row}:{col_letter}{lev_row})"

    write_formula_series(
        "Cumulative Cash Flow", _cumulative_formula, "r_uw_cf_cumulative",
    )

    freeze_top(ws, row=3)
    print_landscape(ws)


# ── Sensitivity sheet ─────────────────────────────────────────────────────────


def _render_sensitivity_grid(
    ws,
    registry: CellRegistry,
    matrix: dict,
    grid_values: list[list[float | None]],
    metric_spec: dict,
    start_row: int,
    range_name: str | None = None,
) -> int:
    """Render one 5×5 sensitivity grid starting at ``start_row``.

    Returns the row number immediately after the last note line so the
    caller can stack a second grid below with appropriate spacing.

    Rows emitted (relative to start_row):
      +0: subtitle (axes → metric)
      +2: corner label + x-axis header
      +3 … +7: y-axis labels + data cells  (GRID_SIZE rows)
      +9: base-case readout
      +11 … +14: notes bullet points
    """
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    axis_x_label = matrix["axis_x"]["label"]
    axis_y_label = matrix["axis_y"]["label"]
    metric_label = metric_spec["label"]
    metric_fmt   = metric_spec.get("format", "pct")
    mode_label = "combined deal-level" if matrix.get("mode") == "combined" else "first-project"

    subtitle_row = start_row
    ws.cell(
        row=subtitle_row, column=1,
        value=f"{axis_y_label} × {axis_x_label} → {metric_label} ({mode_label})",
    ).font = FONT_SUBTITLE

    grid_top = start_row + 2
    corner = ws.cell(
        row=grid_top, column=1,
        value=f"{axis_y_label} ↓ / {axis_x_label} →",
    )
    corner.font = FONT_LABEL
    corner.alignment = ALIGN_WRAP

    x_values = matrix["axis_x"]["values"]
    y_values = matrix["axis_y"]["values"]
    base_x_idx = matrix.get("base_x_index")
    base_y_idx = matrix.get("base_y_index")
    base_fill = PatternFill("solid", fgColor=BRAND["fog"])

    for xi, xv in enumerate(x_values):
        cell = ws.cell(row=grid_top, column=2 + xi, value=float(xv))
        cell.number_format = "0.00\"%\""
        cell.font = FONT_LABEL
        cell.alignment = ALIGN_RIGHT

    grid_first_data_col = 2
    grid_last_data_col  = 1 + len(x_values)
    grid_first_data_row = grid_top + 1
    grid_last_data_row  = grid_top + len(y_values)

    is_pct      = metric_fmt == "pct"
    is_multiple = metric_fmt == "multiple"
    num_fmt     = PCT if is_pct else ('0.00"×"' if is_multiple else ACCOUNTING)

    for yi, yv in enumerate(y_values):
        row = grid_first_data_row + yi
        y_cell = ws.cell(row=row, column=1, value=float(yv))
        y_cell.number_format = "0.00\"%\""
        y_cell.font = FONT_LABEL
        y_cell.alignment = ALIGN_RIGHT

        for xi in range(len(x_values)):
            col = grid_first_data_col + xi
            v = (
                grid_values[yi][xi]
                if yi < len(grid_values) and xi < len(grid_values[yi])
                else None
            )
            # IRR/pct values stored as percent-magnitude (e.g. 12.5 = 12.5%);
            # divide by 100 so Excel % format renders correctly.
            # Multiple values (EM) are already ratios (e.g. 1.85).
            display_v = (float(v) / 100.0 if is_pct else float(v)) if v is not None else None
            cell = ws.cell(row=row, column=col, value=display_v)
            cell.number_format = num_fmt
            cell.alignment = ALIGN_RIGHT
            if (
                base_x_idx is not None
                and base_y_idx is not None
                and yi == base_y_idx
                and xi == base_x_idx
            ):
                cell.fill = base_fill
                cell.font = FONT_HERO_VALUE
            else:
                cell.font = FONT_VALUE

    if range_name:
        registry.register_range(
            range_name,
            ws.title,
            grid_first_data_row,
            grid_last_data_row,
            grid_first_data_col,
            end_col=grid_last_data_col,
        )

    first_col_letter = get_column_letter(grid_first_data_col)
    last_col_letter  = get_column_letter(grid_last_data_col)
    grid_ref = (
        f"{first_col_letter}{grid_first_data_row}"
        f":{last_col_letter}{grid_last_data_row}"
    )
    ws.conditional_formatting.add(
        grid_ref,
        ColorScaleRule(
            start_type="min",        start_color="F8696B",
            mid_type="percentile",   mid_value=50, mid_color="FFEB84",
            end_type="max",          end_color="63BE7B",
        ),
    )

    readout_row = grid_last_data_row + 2
    if base_x_idx is not None and base_y_idx is not None:
        base_x_val = x_values[base_x_idx]
        base_y_val = y_values[base_y_idx]
        base_v = grid_values[base_y_idx][base_x_idx] if grid_values else None
        if base_v is not None:
            v_text = (
                f"{base_v:.2f}%" if is_pct
                else f"{base_v:.2f}×" if is_multiple
                else f"${base_v:,.0f}"
            )
        else:
            v_text = "—"
        ws.cell(
            row=readout_row, column=1,
            value=(
                f"Base case: {axis_y_label} = {base_y_val:.2f}%, "
                f"{axis_x_label} = {base_x_val:.2f}%  →  {metric_label} = {v_text}"
            ),
        ).font = FONT_LABEL

    notes_row = readout_row + 2
    ws.cell(row=notes_row, column=1, value="Notes").font = FONT_LABEL
    notes = [
        "Each cell re-runs the full cashflow engine with the column/row "
        "values substituted into every project's OperationalInputs.",
        f"{metric_label} is the deal-level rollup over summed monthly NCF.",
        "Color scale: red = lowest, green = highest. Base case cell "
        "highlighted with brand fog fill.",
        "Blank cells indicate engine errors (e.g. infeasible debt sizing).",
    ]
    for offset, text in enumerate(notes):
        cell = ws.cell(row=notes_row + 1 + offset, column=1, value=f"• {text}")
        cell.font = FONT_HINT
        cell.alignment = ALIGN_WRAP

    return notes_row + 1 + len(notes)


def _build_sensitivity(
    ws, registry: CellRegistry, ctx: dict, matrix: dict
) -> None:
    """Two-way sensitivity sheet: stacked grids for each metric in ``matrix``.

    Grid 1 (primary metric, default Levered IRR): rows 1–~16.
    Grid 2 (secondary metric, default Equity Multiple): rows ~19–~35.
    Each grid: subtitle → axes header → 5×5 data → base-case readout → notes.
    """
    from openpyxl.utils import get_column_letter  # noqa: F401 — used by helper

    set_widths(ws, [18, 14, 14, 14, 14, 14, 30])

    ws.cell(row=1, column=1, value="Two-Way Sensitivity").font = FONT_TITLE

    next_row = _render_sensitivity_grid(
        ws, registry, matrix,
        grid_values=matrix["values"],
        metric_spec=matrix["metric"],
        start_row=2,
        range_name="r_sensitivity_grid",
    )

    _summary_s = ctx.get("rollup_summary") or {}
    _eq_req_s = _coerce_decimal((_summary_s.get("totals") or {}).get("equity_required") or 0)
    _has_equity_s = _eq_req_s > Decimal(1) or any(
        _funder_class(m) == "Equity"
        and _coerce_decimal((m.source or {}).get("amount") or 0) > Decimal(1)
        for m in (ctx.get("capital_modules") or [])
    )
    if _has_equity_s and matrix.get("values_secondary") and matrix.get("metric_secondary"):
        _render_sensitivity_grid(
            ws, registry, matrix,
            grid_values=matrix["values_secondary"],
            metric_spec=matrix["metric_secondary"],
            start_row=next_row + 2,
            range_name="r_sensitivity_grid_em",
        )

    freeze_top(ws, row=5)  # freeze through x-axis header of first grid (row 4)
    print_landscape(ws)


# ── Investor Returns sheet ────────────────────────────────────────────────────


def _funder_class(module_or_vt) -> str:
    """Return one of `Debt` / `Equity` / `Grant` / `Forgivable Loan` / `Other` for display.

    Accepts a CapitalModule ORM object or a raw vehicle_type string.
    """
    if hasattr(module_or_vt, "vehicle_type"):
        # ORM object — read vehicle_type directly
        vt = str(getattr(module_or_vt, "vehicle_type", "") or "").replace("VehicleType.", "").lower()
    else:
        vt = (str(getattr(module_or_vt, "value", module_or_vt)) or "").lower()
    if vt == "debt":
        return "Debt"
    if vt == "equity":
        return "Equity"
    if vt == "grant":
        return "Grant"
    if vt == "forgivable_loan":
        return "Forgivable Loan"
    return "Other"


_DASH = "—"  # rendered when a column is not meaningful for a row's funder class


def _kv_row_optional(
    ws,
    row: int,
    key: str,
    value,
    *,
    name: str,
    registry: CellRegistry,
    fmt: str,
    hero: bool = False,
) -> None:
    """kv_row variant that writes em-dash for None values without applying
    the numeric format. Mirrors ``_write_optional`` but lays out as a
    label/value pair (column 1 = key, column 2 = value) instead of a
    bare cell. Used on Underwriting Summary KPIs where the metric is
    meaningful only when its denominator is non-zero (Equity Multiple,
    Cash-on-Cash Year 1, etc.) — emitting "—" instead of leaving the
    cell blank tells the LP "no equity stack to compute against",
    matching the per-class column semantics on Source Returns."""
    if value is None:
        ws.cell(row=row, column=1, value=key).font = FONT_LABEL
        ws.cell(row=row, column=1).alignment = ALIGN_LEFT
        cell = ws.cell(row=row, column=2, value=_DASH)
        cell.font = FONT_HERO_VALUE if hero else FONT_VALUE
        if hero:
            cell.fill = FILL_HERO
        cell.alignment = ALIGN_RIGHT
        registry.register(name, ws.title, row, 2)
    else:
        kv_row(ws, row, key, value, name=name, registry=registry, fmt=fmt, hero=hero)


def _write_optional(ws, row, col, value, registry: CellRegistry, *, name: str, fmt: str) -> None:
    """Write a numeric value at (row, col) if non-None, else write the
    em-dash ``"—"`` literal. Either way the named range is registered so
    workbook structure stays stable for downstream formulas; the cell value
    is the dash string when data is missing instead of a misleading $0."""
    if value is None:
        cell = ws.cell(row=row, column=col, value=_DASH)
        cell.font = FONT_VALUE
        cell.alignment = ALIGN_RIGHT
        registry.register(name, ws.title, row, col)
    else:
        registry.write(
            ws, row, col, value,
            name=name, fmt=fmt,
            font=FONT_VALUE, align=ALIGN_RIGHT,
        )


def _build_investor_returns(ws, registry: CellRegistry, ctx: dict) -> None:
    """Source Returns — per-CapitalModule view with per-class column semantics.

    Layout: ``Module | Funder Type | Class | Principal | Rate | Total DS |
    Distributions | Return $ | Return %``.

    Per-class fill: only the columns meaningful to a row's funder class
    carry numeric values; the rest render the em-dash ``"—"`` so the LP
    can tell at a glance that "$0" never means "missing data".

      Debt rows: Rate = ``source.interest_rate_pct``; Total DS = sum of
        ``WaterfallResult.cash_distributed`` for ``debt_service``-tier rows
        on this module (or "—" when no waterfall is computed); Distributions
        = "—" (debt doesn't receive distributions); Return $ = Total DS −
        Principal (= lifetime interest paid, or "—" when no DS data);
        Return % = "—" (cost of capital is the Rate column; restating it
        here as a Return mislabels rate as realized yield).
      Equity rows: Rate = pref rate from carry config (or "—"); Total DS =
        "—"; Distributions = ``cumulative_distributed`` from waterfall;
        Return $ = Distributions − Principal; Return % = the linked
        project's ``project_irr_levered`` (weighted by junction amount when
        the module spans multiple projects). The waterfall's
        ``party_irr_pct`` is the scenario-wide LP/GP IRR — same value on
        every equity module — so it doesn't differentiate per-module.
      Grant / tax-credit / other: Principal only — every other column "—".

    Duplicate-label disambiguation: when two modules share the exact same
    label (the engine creates one ``Owner Equity`` per project, so a
    2-project deal renders two visually-identical rows), the displayed
    label is rewritten to ``"<label> (<project_name>)"`` looked up via
    the ``junctions`` table. Keeps the LP from reading two rows as one
    duplicated row.
    """
    rollup: list[dict] = ctx.get("rollup_waterfall") or []
    summary = ctx.get("rollup_summary") or {}
    totals = summary.get("totals") or {}
    capital_modules: list[CapitalModule] = ctx["capital_modules"]
    module_slugs: dict = ctx.get("module_slugs") or _compute_module_slugs(capital_modules)
    junctions: list[CapitalModuleProject] = ctx["junctions"]
    projects_by_id: dict[UUID, Project] = {p.id: p for p in ctx["projects"]}
    outputs_by_project: dict[UUID, OperationalOutputs] = ctx.get("outputs") or {}
    cash_flows_by_project: dict = ctx.get("cash_flows") or {}

    set_widths(ws, [30, 18, 10, 16, 10, 16, 16, 16, 12])

    section_label(ws, 1, "Source Returns — Per Capital Module", span_cols=9)
    header_row(
        ws, 2,
        ["Module", "Funder Type", "Class", "Principal", "Rate",
         "Total DS", "Distributions", "Return ($)", "Return (%)"],
    )

    # Junction-aggregated principals (one shared debt module covering N
    # projects has its principal split across N junction rows; the
    # module-level principal is their sum).
    junction_principal: dict[UUID, Decimal] = {}
    junction_projects: dict[UUID, list[UUID]] = {}
    junction_amount_by_project: dict[UUID, dict[UUID, Decimal]] = {}
    for j in junctions:
        junction_principal[j.capital_module_id] = junction_principal.get(
            j.capital_module_id, Decimal(0)
        ) + _coerce_decimal(j.amount or 0)
        junction_projects.setdefault(j.capital_module_id, []).append(j.project_id)
        junction_amount_by_project.setdefault(j.capital_module_id, {})[j.project_id] = (
            junction_amount_by_project.get(j.capital_module_id, {}).get(j.project_id, Decimal(0))
            + _coerce_decimal(j.amount or 0)
        )

    def _module_per_project_irr(module_id: UUID) -> Decimal | None:
        """Per-module Levered IRR = principal-weighted average of the linked
        projects' ``project_irr_levered``. Returns None when none of the
        linked projects have a computed IRR.
        """
        amounts = junction_amount_by_project.get(module_id) or {}
        if not amounts:
            return None
        total_amt = sum(amounts.values(), Decimal(0))
        if total_amt <= 0:
            return None
        weighted = Decimal(0)
        seen = False
        for pid, amt in amounts.items():
            o = outputs_by_project.get(pid)
            if o is None:
                continue
            irr_raw = getattr(o, "project_irr_levered", None)
            if irr_raw is None:
                continue
            seen = True
            weighted += _coerce_decimal(irr_raw) * amt
        if not seen:
            return None
        # project_irr_levered is stored as percent (e.g. 5.34 = 5.34%); the
        # _write_optional call below uses _coerce_pct via PCT format which
        # expects a fraction. Divide by 100 to convert.
        return (weighted / total_amt) / Decimal(100)

    # Pre-aggregate waterfall: per-module cumulative distributions, latest
    # party IRR, and per-module debt-service totals (debt_service tier rows).
    module_distributions: dict[str, Decimal] = {}
    module_irr: dict[str, Decimal] = {}
    module_latest_period: dict[str, int] = {}
    module_debt_service_total: dict[str, Decimal] = {}
    for row in rollup:
        mid = row.get("capital_module_id")
        if not mid:
            continue
        cum = _coerce_decimal(row.get("cumulative_distributed") or 0)
        if cum > module_distributions.get(mid, Decimal(0)):
            module_distributions[mid] = cum
        period = row.get("period") or 0
        if period >= module_latest_period.get(mid, -1):
            module_latest_period[mid] = period
            irr = row.get("party_irr_pct")
            if irr is not None:
                module_irr[mid] = _coerce_pct(irr)
        if (row.get("tier_type") or "") == "debt_service":
            module_debt_service_total[mid] = (
                module_debt_service_total.get(mid, Decimal(0))
                + _coerce_decimal(row.get("cash_distributed") or 0)
            )

    # Pre-walk module labels to disambiguate duplicates by project context.
    label_counts: dict[str, int] = {}
    for module in capital_modules:
        raw = module.label or _funder_type_label(module)
        label_counts[raw] = label_counts.get(raw, 0) + 1

    def _display_label(module: CapitalModule) -> str:
        raw = module.label or _funder_type_label(module)
        if label_counts.get(raw, 0) <= 1:
            return raw
        # Disambiguate via the first project in the module's junction rows.
        proj_ids = junction_projects.get(module.id) or []
        if proj_ids:
            proj = projects_by_id.get(proj_ids[0])
            if proj and proj.name:
                return f"{raw} ({proj.name})"
        return raw

    cur_row = 3
    if not capital_modules:
        ws.cell(
            row=cur_row, column=1,
            value="(no capital modules — add Sources on the Capital Stack module to populate)",
        ).font = FONT_HINT
        cur_row += 1

    for m_idx, module in enumerate(capital_modules, start=1):
        source = module.source or {}
        carry = module.carry or {}
        mid_str = str(module.id)
        slug = module_slugs.get(module.id) or f"module_{m_idx}"
        principal = junction_principal.get(module.id)
        if principal is None:
            principal = _coerce_decimal(source.get("amount") or 0)
        rate_raw = source.get("interest_rate_pct") or carry.get("io_rate_pct") or 0
        rate = _coerce_pct(rate_raw) if rate_raw else None
        funder_class = _funder_class(module)
        if funder_class == "Equity" and principal <= Decimal(1):
            continue

        # Per-class column fill — write em-dash strings where a column doesn't
        # apply, so missing data never reads as "$0" or "0%".
        if funder_class == "Debt":
            total_ds = module_debt_service_total.get(mid_str)
            if not total_ds:
                # Waterfall can't see DS (NCF is already net-of-DS), so sum
                # CashFlow.debt_service directly from covered projects.
                covered_pids = junction_projects.get(module.id, [])
                cf_ds = sum(
                    sum(_coerce_decimal(cf.debt_service or 0) for cf in cash_flows_by_project.get(pid, []))
                    for pid in covered_pids
                )
                total_ds = cf_ds if cf_ds > 0 else None
            distributions = None  # debt has no distributions
            if total_ds is not None and total_ds > 0:
                return_dollars = total_ds - principal
            else:
                return_dollars = None  # no DS data ⇒ blank, not -principal
            # Cost of capital lives in the Rate column. Restating it as a
            # Return mislabels rate-as-realized-yield, so leave Return % blank
            # for debt rows.
            return_pct = None
        elif funder_class == "Equity":
            total_ds = None
            distributions = module_distributions.get(mid_str)
            return_dollars = (distributions - principal) if distributions is not None else None
            # Per-module IRR from the linked project(s) — not the waterfall's
            # scenario-wide LP/GP IRR (which is the same value on every
            # equity row and reads as a per-module differentiator when it
            # isn't).
            return_pct = _module_per_project_irr(module.id)
        else:
            # Grant / tax_credit / other — only Principal is meaningful
            total_ds = None
            distributions = None
            return_dollars = None
            return_pct = None

        ws.cell(row=cur_row, column=1, value=_display_label(module)).font = FONT_VALUE
        ws.cell(row=cur_row, column=2, value=_funder_type_label(module)).font = FONT_VALUE
        ws.cell(row=cur_row, column=3, value=funder_class).font = FONT_VALUE

        registry.write(
            ws, cur_row, 4, principal,
            name=f"s_{slug}_principal_returns", fmt=ACCOUNTING,
            font=FONT_VALUE, align=ALIGN_RIGHT,
        )
        _write_optional(
            ws, cur_row, 5, rate, registry,
            name=f"s_{slug}_rate_returns", fmt=PCT,
        )
        _write_optional(
            ws, cur_row, 6, total_ds, registry,
            name=f"s_{slug}_total_ds", fmt=ACCOUNTING,
        )
        _write_optional(
            ws, cur_row, 7, distributions, registry,
            name=f"s_{slug}_distributions", fmt=ACCOUNTING,
        )
        # Formula-conversion plan §4.3 (commit 5): Return ($) is a clean
        # derivation of (total_ds or distributions) - principal. When the
        # operand exists we emit the formula so LP edits to the principal
        # or distribution cells propagate; when missing, fall back to the
        # em-dash via _write_optional.
        return_formula: str | None = None
        if funder_class == "Debt" and total_ds is not None and total_ds > 0:
            return_formula = f"=F{cur_row}-D{cur_row}"
        elif funder_class == "Equity" and distributions is not None:
            return_formula = f"=G{cur_row}-D{cur_row}"
        if return_formula is not None:
            cell = ws.cell(row=cur_row, column=8, value=return_formula)
            cell.number_format = ACCOUNTING
            cell.font = FONT_VALUE
            cell.alignment = ALIGN_RIGHT
            registry.register(
                f"s_{slug}_return_dollars", ws.title, cur_row, 8,
            )
        else:
            _write_optional(
                ws, cur_row, 8, return_dollars, registry,
                name=f"s_{slug}_return_dollars", fmt=ACCOUNTING,
            )
        _write_optional(
            ws, cur_row, 9, return_pct, registry,
            name=f"s_{slug}_return_pct", fmt=PCT,
        )
        cur_row += 1

    # ── Aggregate rollup (only meaningful when a waterfall is populated) ──
    cur_row += 1
    _equity_req_check = _coerce_decimal(totals.get("equity_required") or 0)
    _committed_equity = sum(
        junction_principal.get(m.id) or _coerce_decimal((m.source or {}).get("amount") or 0)
        for m in capital_modules
        if _funder_class(m) == "Equity"
    )
    _has_real_equity = _committed_equity > Decimal(1) or _equity_req_check > Decimal(1)

    section_label(ws, cur_row, "Scenario-Level Aggregates", span_cols=2)
    cur_row += 1

    # Formula-conversion plan §4.3 (commit 5): Combined Levered IRR is
    # IRR over the Levered Cash Flow row on 'Underwriting Cash Flow' —
    # which itself is a formula chain back to NOI/CapEvents/DS. Edits
    # anywhere upstream now propagate to this IRR cell via Excel's calc
    # engine. IFERROR guards the no-equity / degenerate-stream case so a
    # blank scenario shows 0% instead of #NUM!.
    kv_row(
        ws, cur_row, "Combined Levered IRR (scenario)",
        "=IFERROR(IRR(r_uw_cf_levered),0)",
        name="s_returns_combined_irr", registry=registry, fmt=PCT,
    ); cur_row += 1

    # Formula-conversion plan §4.3 (commit 8): EM is now a SUMIF over the
    # Underwriting Cash Flow levered row so it tracks LP edits to upstream
    # NOI / DS / capital events. When the scenario has no real equity
    # stack, the formula resolves to 0 via IFERROR — but we keep the cell
    # blank in that case to match the prior em-dash semantics.
    if _has_real_equity:
        kv_row(
            ws, cur_row, "Combined Equity Multiple (scenario)",
            '=IFERROR(SUMIF(r_uw_cf_levered,">0")/(-SUMIF(r_uw_cf_levered,"<0")),0)',
            name="s_returns_combined_em", registry=registry, fmt='0.00"×"',
        )
    else:
        kv_row(
            ws, cur_row, "Combined Equity Multiple (scenario)",
            None,
            name="s_returns_combined_em", registry=registry, fmt='0.00"×"',
        )
    cur_row += 1

    by_tier = _waterfall_by_tier(rollup)
    pref_total = by_tier.get("pref_return", {}).get("cash_total", Decimal(0))
    catch_up_total = by_tier.get("catch_up", {}).get("cash_total", Decimal(0))
    residual_total = by_tier.get("residual", {}).get("cash_total", Decimal(0))
    promote_total = residual_total + catch_up_total
    # "GP Promote" assumes a promote-above-pref structure. With no pref tier
    # paid, residual is straight pro-rata distribution to equity, not promote.
    # Rename so the LP doesn't read residual cash as carried interest.
    has_pref_tier = pref_total > 0
    promote_label = (
        "GP Promote $ (catch-up + residual)"
        if has_pref_tier
        else "Residual Distributions to GP (no pref tier)"
    )
    kv_row(
        ws, cur_row, promote_label,
        promote_total,
        name="s_gp_promote_dollars", registry=registry, fmt=ACCOUNTING,
    ); cur_row += 1

    if _equity_req_check > Decimal(1) and _committed_equity <= Decimal(1):
        cur_row += 1
        note = ws.cell(
            row=cur_row, column=1,
            value=(
                f"⚠ Return multiples (EM, Weighted EM, CoC) are computed against the implied equity "
                f"basis of {_format_currency_short(_equity_req_check)}, not a formally committed equity "
                f"module. Assign equity to a capital module to lock in the contribution basis."
            ),
        )
        note.font = FONT_HINT
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=8)

    if not rollup:
        cur_row += 1
        ws.cell(
            row=cur_row, column=1,
            value=(
                "(no waterfall distributions yet — Source Returns above show "
                "principal + cost-of-capital semantics; add equity tiers + "
                "compute to populate IRR / promote.)"
            ),
        ).font = FONT_HINT
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=8)

    freeze_top(ws, row=3)
    print_landscape(ws)


# Canonical investor-waterfall tier order. When the Scenario has no
# WaterfallTier rows configured, the structure block renders these as $0
# placeholders so the LP sees the policy structure the deal *should* have.
_CANONICAL_WATERFALL_TIERS: tuple[tuple[str, str], ...] = (
    ("debt_service", "Debt Service"),
    ("return_of_equity", "Return of Equity"),
    ("pref_return", "Pref Return (LP preferred)"),
    ("catch_up", "GP Catch-Up"),
    ("irr_hurdle_split", "IRR-Hurdle Split"),
    ("deferred_developer_fee", "Deferred Developer Fee"),
    ("residual", "Residual / Promote"),
)


def _build_waterfall_structure(
    ws, registry: CellRegistry, start_row: int, ctx: dict,
) -> int:
    """Render the Waterfall Structure block.

    Two display modes:

      Configured: Scenario has ``WaterfallTier`` rows. Renders one row per
      tier in priority order with tier_type, IRR hurdle (if applicable),
      LP / GP split %, and Total Distributed (cumulative
      ``WaterfallResult.cash_distributed`` for the tier).

      Unconfigured: Scenario has zero tiers. Renders the canonical
      structure (Pref → Catch-Up → Promote etc.) with "—" / 0 placeholders
      so the LP sees the policy structure the deal *should* have. A hint
      cell calls out the placeholder state explicitly.

    Returns the next-free row.
    """
    waterfall_tiers: list[WaterfallTier] = ctx.get("waterfall_tiers") or []
    rollup: list[dict] = ctx.get("rollup_waterfall") or []

    section_label(ws, start_row, "Waterfall Structure", span_cols=8)
    header_row(
        ws, start_row + 1,
        ["Priority", "Tier Type", "IRR Hurdle", "LP Split", "GP Split",
         "Total Distributed", "LP Amount", "GP Amount"],
    )

    # Pre-aggregate distributions per tier_id from the rollup.
    dist_by_tier_id: dict[str, Decimal] = {}
    for row in rollup:
        tier_id = row.get("tier_id") or ""
        amount = _coerce_decimal(row.get("cash_distributed") or 0)
        dist_by_tier_id[tier_id] = dist_by_tier_id.get(tier_id, Decimal(0)) + amount

    cur = start_row + 2

    if waterfall_tiers:
        ordered = sorted(
            waterfall_tiers,
            key=lambda t: (int(t.priority or 999), str(t.tier_type)),
        )
        for tier_idx, tier in enumerate(ordered, start=1):
            tier_type_str = str(getattr(tier.tier_type, "value", tier.tier_type) or "")
            display_label = tier_type_str.replace("_", " ").title()
            ws.cell(row=cur, column=1, value=int(tier.priority or 0)).font = FONT_VALUE
            ws.cell(row=cur, column=2, value=display_label).font = FONT_LABEL
            hurdle = _coerce_decimal(tier.irr_hurdle_pct or 0)
            if tier_type_str == "irr_hurdle_split" and hurdle > 0:
                cell = ws.cell(row=cur, column=3, value=_to_excel_number(hurdle / Decimal(100)))
                cell.number_format = PCT
            else:
                ws.cell(row=cur, column=3, value=_DASH).font = FONT_VALUE
            cell_lp = ws.cell(
                row=cur, column=4,
                value=_to_excel_number(_coerce_decimal(tier.lp_split_pct or 0) / Decimal(100)),
            )
            cell_lp.number_format = PCT
            cell_gp = ws.cell(
                row=cur, column=5,
                value=_to_excel_number(_coerce_decimal(tier.gp_split_pct or 0) / Decimal(100)),
            )
            cell_gp.number_format = PCT
            distributed = dist_by_tier_id.get(str(tier.id), Decimal(0))
            registry.write(
                ws, cur, 6, distributed,
                name=f"s_waterfall_tier_{tier_idx}_distributed", fmt=ACCOUNTING,
                font=FONT_VALUE, align=ALIGN_RIGHT,
            )
            lp_pct_d = _coerce_decimal(tier.lp_split_pct or 0)
            gp_pct_d = _coerce_decimal(tier.gp_split_pct or 0)
            lp_amt = distributed * lp_pct_d / Decimal(100)
            gp_amt = distributed * gp_pct_d / Decimal(100)
            lp_c = ws.cell(row=cur, column=7, value=_to_excel_number(lp_amt))
            lp_c.number_format = ACCOUNTING; lp_c.font = FONT_VALUE; lp_c.alignment = ALIGN_RIGHT
            gp_c = ws.cell(row=cur, column=8, value=_to_excel_number(gp_amt))
            gp_c.number_format = ACCOUNTING; gp_c.font = FONT_VALUE; gp_c.alignment = ALIGN_RIGHT
            cur += 1
    else:
        # Unconfigured — render canonical structure with $0 placeholders.
        for tier_idx, (_tier_type, label) in enumerate(_CANONICAL_WATERFALL_TIERS, start=1):
            ws.cell(row=cur, column=1, value=tier_idx).font = FONT_VALUE
            ws.cell(row=cur, column=2, value=label).font = FONT_LABEL
            ws.cell(row=cur, column=3, value=_DASH).font = FONT_VALUE
            ws.cell(row=cur, column=4, value=_DASH).font = FONT_VALUE
            ws.cell(row=cur, column=5, value=_DASH).font = FONT_VALUE
            registry.write(
                ws, cur, 6, Decimal(0),
                name=f"s_waterfall_tier_{tier_idx}_distributed", fmt=ACCOUNTING,
                font=FONT_VALUE, align=ALIGN_RIGHT,
            )
            ws.cell(row=cur, column=7, value=_DASH).font = FONT_VALUE
            ws.cell(row=cur, column=8, value=_DASH).font = FONT_VALUE
            cur += 1
        cur += 1
        ws.cell(
            row=cur, column=1,
            value=(
                "(placeholder structure — no WaterfallTier rows configured. "
                "Configure pref / catch-up / promote tiers on the Capital Stack module "
                "to replace the placeholders with real splits and distributions.)"
            ),
        ).font = FONT_HINT
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=8)
        cur += 1

    return cur


def _build_waterfall_sheet(ws, registry: CellRegistry, ctx: dict) -> None:
    """Dedicated Waterfall sheet — tier structure and LP / GP distributions."""
    set_widths(ws, [10, 32, 12, 10, 10, 18, 18, 18])
    print_landscape(ws)
    _build_waterfall_structure(ws, registry, 1, ctx)
    freeze_top(ws, row=3)


def _build_unit_mix_sheet(ws, registry: CellRegistry, ctx: dict) -> None:
    """Unit Mix — per-project breakdown of unit types, rents, and gap."""
    projects: list[Project] = ctx["projects"]
    unit_mix_by_project: dict[UUID, list] = ctx.get("unit_mix") or {}

    set_widths(ws, [28, 7, 7, 10, 8, 16, 16, 16, 10, 22, 18, 18])
    print_landscape(ws)

    has_any = any(unit_mix_by_project.get(p.id) for p in projects)
    if not has_any:
        ws.cell(row=1, column=1, value="No unit mix configured.").font = FONT_HINT
        ws.cell(row=2, column=1, value=(
            "Add unit types in Project Setup to enable per-type rent and "
            "loss-to-lease analysis."
        )).font = FONT_HINT
        return

    row = 1
    for project in projects:
        units = unit_mix_by_project.get(project.id) or []
        if not units:
            continue

        section_label(ws, row, project.name or "Project", span_cols=12)
        row += 1
        header_row(
            ws, row,
            ["Unit Type", "Beds", "Baths", "Avg SF", "Count",
             "In-Place Rent", "Market Rent", "Rent Gap ($)", "L-t-L %",
             "Strategy", "Monthly Rev (In-Place)", "Monthly Rev (Market)"],
        )
        row += 1

        proj_ip_total = Decimal(0)
        proj_mkt_total = Decimal(0)
        proj_units_total = 0

        _STRATEGY_LABELS = {
            "base_escalation": "Base Escalation",
            "ltl_catchup": "LTL Catchup",
            "value_add_renovation": "Value-Add Renovation",
        }
        for um in sorted(units, key=lambda u: (float(u.beds or 0), float(u.baths or 0))):
            count = um.unit_count or 0
            ip_rent = _coerce_decimal(um.in_place_rent_per_unit) if um.in_place_rent_per_unit else None
            mkt_rent = _coerce_decimal(um.market_rent_per_unit) if um.market_rent_per_unit else None
            sqft = _coerce_decimal(um.avg_sqft) if um.avg_sqft else None
            gap = (mkt_rent - ip_rent) if (mkt_rent is not None and ip_rent is not None) else None
            ltl_pct = (gap / mkt_rent) if (gap is not None and mkt_rent and mkt_rent > Decimal(0)) else None
            strategy = _STRATEGY_LABELS.get(
                um.unit_strategy or "",
                (um.unit_strategy or "").replace("_", " ").title(),
            ) or "—"
            ip_rev = (ip_rent * count) if ip_rent and count else None
            mkt_rev = (mkt_rent * count) if mkt_rent and count else None

            proj_ip_total += ip_rev or Decimal(0)
            proj_mkt_total += mkt_rev or Decimal(0)
            proj_units_total += count

            ws.cell(row=row, column=1, value=um.label or "—").font = FONT_LABEL
            ws.cell(row=row, column=1).alignment = ALIGN_LEFT

            for col, val, fmt in (
                (2, float(um.beds) if um.beds is not None else None, None),
                (3, float(um.baths) if um.baths is not None else None, None),
                (4, _to_excel_number(sqft) if sqft else None, INT_COMMA),
                (5, count, INT_COMMA),
                (6, _to_excel_number(ip_rent) if ip_rent else None, ACCOUNTING),
                (7, _to_excel_number(mkt_rent) if mkt_rent else None, ACCOUNTING),
                (11, _to_excel_number(ip_rev) if ip_rev else None, ACCOUNTING),
                (12, _to_excel_number(mkt_rev) if mkt_rev else None, ACCOUNTING),
            ):
                c = ws.cell(row=row, column=col, value=val)
                c.font = FONT_VALUE
                c.alignment = ALIGN_RIGHT
                if fmt:
                    c.number_format = fmt

            gap_c = ws.cell(
                row=row, column=8,
                value=_to_excel_number(gap) if gap is not None else _DASH,
            )
            gap_c.font = FONT_VALUE
            gap_c.alignment = ALIGN_RIGHT
            if gap is not None:
                gap_c.number_format = ACCOUNTING
                gap_c.fill = FILL_RAG_GREEN if gap >= 0 else FILL_RAG_RED

            ltl_c = ws.cell(
                row=row, column=9,
                value=_to_excel_number(ltl_pct) if ltl_pct is not None else _DASH,
            )
            ltl_c.font = FONT_VALUE
            ltl_c.alignment = ALIGN_RIGHT
            if ltl_pct is not None:
                ltl_c.number_format = PCT_1
                ltl_c.fill = FILL_RAG_GREEN if ltl_pct >= 0 else FILL_RAG_RED

            ws.cell(row=row, column=10, value=strategy).font = FONT_VALUE
            ws.cell(row=row, column=10).alignment = ALIGN_LEFT
            row += 1

        ws.cell(row=row, column=1, value=f"Total — {project.name or 'Project'}").font = FONT_LABEL
        for col, val, fmt in (
            (5, proj_units_total, INT_COMMA),
            (11, _to_excel_number(proj_ip_total) if proj_ip_total else None, ACCOUNTING),
            (12, _to_excel_number(proj_mkt_total) if proj_mkt_total else None, ACCOUNTING),
        ):
            c = ws.cell(row=row, column=col, value=val)
            c.font = FONT_HERO_VALUE
            c.fill = FILL_HERO
            c.alignment = ALIGN_RIGHT
            if fmt:
                c.number_format = fmt
        row += 2

    freeze_top(ws, row=3)
    _ = registry  # reserved for future per-row named ranges


# ── Per-project sheets (commit 3) ─────────────────────────────────────────────


def _build_project_sheet(
    ws,
    registry: CellRegistry,
    ctx: dict,
    project_idx: int,
    project: Project,
) -> None:
    """One sheet per project: header → Pro Forma → Cash Flow → S&U.

    Named ranges use the ``p{n}_`` prefix from plan §4 — outputs only.
    Per-project *inputs* live on the Assumptions sheet (Block B) and use
    the same prefix; outputs are distinct names so the registry doesn't
    collide. Layout matches the underwriting rollup sheets so an LP can
    open a project sheet and read it the same way as the scenario summary.
    """
    inputs_by_project: dict[UUID, OperationalInputs] = ctx["operational_inputs"]
    use_lines_by_project: dict[UUID, list[UseLine]] = ctx["use_lines"]
    cash_flows_by_project: dict[UUID, list[CashFlow]] = ctx["cash_flows"]
    cash_flow_items_by_project: dict[UUID, list[CashFlowLineItem]] = ctx["cash_flow_items"]
    outputs_by_project: dict[UUID, "OperationalOutputs"] = ctx["outputs"]
    capital_modules: list[CapitalModule] = ctx["capital_modules"]
    junctions: list[CapitalModuleProject] = ctx["junctions"]

    inputs = inputs_by_project.get(project.id)
    use_lines = use_lines_by_project.get(project.id, [])
    cash_flows = cash_flows_by_project.get(project.id, [])
    line_items = cash_flow_items_by_project.get(project.id, [])
    outputs = outputs_by_project.get(project.id)

    annual = _aggregate_annual(cash_flows)
    # Signed per-project capital events — outflows negative, inflows
    # positive — see V2-B fix in _signed_capital_events_by_year_for_project.
    capital_events_by_year = _signed_capital_events_by_year_for_project(line_items)
    max_year = min(max(annual) if annual else 0, 10)
    year_cols = list(range(0, max(max_year, 1) + 1))

    set_widths(ws, [30, *([14] * len(year_cols))])

    # ── Project header ─────────────────────────────────────────────────────
    ws.cell(
        row=1, column=1,
        value=f"P{project_idx} — {project.name or 'Project'}",
    ).font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(year_cols) + 1)
    ws.row_dimensions[1].height = 24

    # Top-of-sheet hyperlinks back to the rollup + glossary — green per
    # the cross-sheet-link convention (Phase H3).
    ws.cell(
        row=2, column=1,
        value='=HYPERLINK("#\'Underwriting Summary\'!A1", "← Underwriting Summary")',
    ).font = FONT_LINK
    ws.cell(
        row=2, column=2,
        value='=HYPERLINK("#\'Glossary & Methodology\'!A1", "Glossary →")',
    ).font = FONT_LINK

    section_label(ws, 4, "Project KPIs", span_cols=2)
    cur = 5
    kv_row(
        ws, cur, "Project Type",
        getattr(project, "deal_type", "") or "",
        name=f"p{project_idx}_uw_project_type", registry=registry,
    ); cur += 1
    kv_row(
        ws, cur, "Total Project Cost",
        _safe_decimal(outputs, "total_project_cost"),
        name=f"p{project_idx}_total_project_cost", registry=registry,
        fmt=ACCOUNTING, hero=True,
    ); cur += 1
    kv_row(
        ws, cur, "Equity Required",
        _safe_decimal(outputs, "equity_required"),
        name=f"p{project_idx}_equity_required", registry=registry,
        fmt=ACCOUNTING, hero=True,
    ); cur += 1
    kv_row(
        ws, cur, "Stabilized NOI",
        _safe_decimal(outputs, "noi_stabilized"),
        name=f"p{project_idx}_noi_stabilized", registry=registry,
        fmt=ACCOUNTING, hero=True,
    ); cur += 1
    kv_row(
        ws, cur, "DSCR",
        _safe_decimal(outputs, "dscr"),
        name=f"p{project_idx}_dscr", registry=registry,
        fmt="0.000", hero=True,
    ); cur += 1
    kv_row(
        ws, cur, "Cap Rate on Cost",
        _pct_value(outputs, "cap_rate_on_cost_pct"),
        name=f"p{project_idx}_cap_rate", registry=registry,
        fmt=PCT, hero=True,
    ); cur += 1
    kv_row(
        ws, cur, "Debt Yield",
        _pct_value(outputs, "debt_yield_pct"),
        name=f"p{project_idx}_debt_yield", registry=registry,
        fmt=PCT, hero=True,
    ); cur += 1
    kv_row(
        ws, cur, "Levered IRR",
        _pct_value(outputs, "project_irr_levered"),
        name=f"p{project_idx}_levered_irr", registry=registry,
        fmt=PCT, hero=True,
    ); cur += 1
    kv_row(
        ws, cur, "Unlevered IRR",
        _pct_value(outputs, "project_irr_unlevered"),
        name=f"p{project_idx}_unlevered_irr", registry=registry,
        fmt=PCT, hero=True,
    ); cur += 1
    kv_row(
        ws, cur, "Total Timeline (months)",
        _safe_decimal(outputs, "total_timeline_months"),
        name=f"p{project_idx}_timeline_months", registry=registry,
        fmt=INT_COMMA,
    ); cur += 1

    # ── Phase Plan (absolute month boundaries) ─────────────────────────────
    cur = _emit_phase_plan_block(
        ws, registry, ctx,
        project_idx=project_idx, project=project,
        inputs=inputs, capital_modules=capital_modules,
        start_row=cur,
    )

    # ── Project Pro Forma ──────────────────────────────────────────────────
    pf_row = cur + 2
    section_label(
        ws, pf_row, "Project Pro Forma — Annual",
        span_cols=len(year_cols) + 1,
    )
    header_row(ws, pf_row + 1, ["Line Item", *[f"Y{y}" for y in year_cols]])
    pf_data = pf_row + 2
    pf_rows: list[tuple[str, str, str | None]] = [
        ("Gross Revenue", "gross_revenue", f"r_p{project_idx}_gross_revenue"),
        ("Vacancy Loss", "vacancy_loss", None),
        ("EGI", "effective_gross_income", f"r_p{project_idx}_egi"),
        ("Operating Expenses", "operating_expenses", f"r_p{project_idx}_opex"),
        ("CapEx Reserve", "capex_reserve", None),
        ("NOI", "noi", f"r_p{project_idx}_noi"),
    ]
    for label, field, range_name in pf_rows:
        ws.cell(row=pf_data, column=1, value=label).font = FONT_LABEL
        for col_offset, year in enumerate(year_cols):
            value = annual.get(year, {}).get(field, Decimal(0))
            cell = ws.cell(row=pf_data, column=2 + col_offset, value=_to_excel_number(value))
            cell.number_format = ACCOUNTING
            cell.font = FONT_VALUE
            cell.alignment = ALIGN_RIGHT
        if range_name and year_cols:
            registry.register_range(
                range_name, ws.title, pf_data, pf_data,
                col=2, end_col=1 + len(year_cols),
            )
        pf_data += 1

    # ── Project Cash Flow ──────────────────────────────────────────────────
    cf_row = pf_data + 1
    section_label(
        ws, cf_row, "Project Cash Flow — Annual",
        span_cols=len(year_cols) + 1,
    )
    header_row(ws, cf_row + 1, ["Line Item", *[f"Y{y}" for y in year_cols]])
    cf_data = cf_row + 2

    noi_series = {y: annual.get(y, {}).get("noi", Decimal(0)) for y in year_cols}
    debt_series = {y: annual.get(y, {}).get("debt_service", Decimal(0)) for y in year_cols}
    ncf_series = {y: annual.get(y, {}).get("net_cash_flow", Decimal(0)) for y in year_cols}

    def write_proj_series(label: str, source: dict[int, Decimal], range_name: str | None,
                          fmt: str = ACCOUNTING) -> None:
        nonlocal cf_data
        ws.cell(row=cf_data, column=1, value=label).font = FONT_LABEL
        for col_offset, year in enumerate(year_cols):
            value = source.get(year, Decimal(0))
            cell = ws.cell(row=cf_data, column=2 + col_offset, value=_to_excel_number(value))
            cell.number_format = fmt
            cell.font = FONT_VALUE
            cell.alignment = ALIGN_RIGHT
        if range_name and year_cols:
            registry.register_range(
                range_name, ws.title, cf_data, cf_data,
                col=2, end_col=1 + len(year_cols),
            )
        cf_data += 1

    write_proj_series("NOI", noi_series, f"r_p{project_idx}_cf_noi")
    write_proj_series(
        "Capital Events", capital_events_by_year, f"r_p{project_idx}_cf_capital_events"
    )
    write_proj_series("Debt Service", debt_series, f"r_p{project_idx}_cf_debt_service")
    write_proj_series("Levered Cash Flow", ncf_series, f"r_p{project_idx}_cf_levered")

    # V2-B: Unlevered = engine NCF + DS (matches IRR helper path).
    # Cumulative = running sum of NCF (capital events already inside NCF
    # via engine invariant; adding them separately would double-count).
    unlevered_series = {
        y: ncf_series.get(y, Decimal(0)) + debt_series.get(y, Decimal(0))
        for y in year_cols
    }
    write_proj_series(
        "Unlevered Cash Flow", unlevered_series, f"r_p{project_idx}_cf_unlevered"
    )

    cumulative: dict[int, Decimal] = {}
    running = Decimal(0)
    for y in year_cols:
        running += ncf_series.get(y, Decimal(0))
        cumulative[y] = running
    write_proj_series("Cumulative Cash Flow", cumulative, f"r_p{project_idx}_cf_cumulative")

    # ── Project S&U ────────────────────────────────────────────────────────
    su_row = cf_data + 1
    section_label(ws, su_row, "Project Sources & Uses", span_cols=4)
    header_row(ws, su_row + 1, ["Side", "Label", "Amount", "Notes"])
    su_data = su_row + 2

    uses_total = Decimal(0)
    by_phase: dict[str, Decimal] = {}
    for ul in use_lines:
        phase = str(getattr(ul.phase, "value", ul.phase) or "")
        if phase == "exit":
            continue
        amt = _coerce_decimal(ul.amount or 0)
        by_phase[phase] = by_phase.get(phase, Decimal(0)) + amt
        uses_total += amt
    for phase, amount in sorted(by_phase.items()):
        ws.cell(row=su_data, column=1, value="Use").font = FONT_VALUE
        ws.cell(row=su_data, column=2, value=phase.replace("_", " ").title()).font = FONT_VALUE
        ws.cell(row=su_data, column=3, value=_to_excel_number(amount)).number_format = ACCOUNTING
        su_data += 1
    ws.cell(row=su_data, column=1, value="Use").font = FONT_LABEL
    ws.cell(row=su_data, column=2, value="Total Uses").font = FONT_LABEL
    registry.write(
        ws, su_data, 3, uses_total,
        name=f"p{project_idx}_uw_total_uses", fmt=ACCOUNTING,
        font=FONT_LABEL, align=ALIGN_RIGHT,
    )
    su_data += 2

    # Sources for THIS project — junction-scoped (each capital module's
    # share for this project, not the scenario-wide commitment).
    junction_for_project: dict[UUID, Decimal] = {}
    for j in junctions:
        if j.project_id != project.id:
            continue
        junction_for_project[j.capital_module_id] = junction_for_project.get(
            j.capital_module_id, Decimal(0)
        ) + _coerce_decimal(j.amount or 0)

    sources_total = Decimal(0)
    for module in capital_modules:
        if module.id not in junction_for_project:
            continue
        amount = junction_for_project[module.id]
        ws.cell(row=su_data, column=1, value="Source").font = FONT_VALUE
        ws.cell(
            row=su_data, column=2,
            value=module.label or _funder_type_label(module),
        ).font = FONT_VALUE
        ws.cell(row=su_data, column=3, value=_to_excel_number(amount)).number_format = ACCOUNTING
        ws.cell(
            row=su_data, column=4,
            value=_funder_type_label(module),
        ).font = FONT_HINT
        sources_total += amount
        su_data += 1

    if not junction_for_project:
        ws.cell(
            row=su_data, column=1,
            value="(no capital module attached to this project)",
        ).font = FONT_HINT
        su_data += 1

    _proj_implied = uses_total - sources_total
    if _proj_implied > Decimal(1):
        ws.cell(row=su_data, column=1, value="Source").font = FONT_VALUE
        ws.cell(row=su_data, column=2, value="Owner Equity (implied gap)").font = FONT_VALUE
        ws.cell(row=su_data, column=3, value=_to_excel_number(_proj_implied)).number_format = ACCOUNTING
        ws.cell(row=su_data, column=4, value="Auto-funded equity — residual after debt").font = FONT_HINT
        sources_total += _proj_implied
        su_data += 1

    ws.cell(row=su_data, column=1, value="Source").font = FONT_LABEL
    ws.cell(row=su_data, column=2, value="Total Sources").font = FONT_LABEL
    registry.write(
        ws, su_data, 3, sources_total,
        name=f"p{project_idx}_uw_total_sources", fmt=ACCOUNTING,
        font=FONT_LABEL, align=ALIGN_RIGHT,
    )
    su_data += 1

    gap = uses_total - sources_total
    ws.cell(row=su_data, column=1, value="Δ").font = FONT_LABEL
    ws.cell(row=su_data, column=2, value="Gap (Uses − Sources)").font = FONT_LABEL
    registry.write(
        ws, su_data, 3, gap,
        name=f"p{project_idx}_uw_gap", fmt=ACCOUNTING,
        font=FONT_LABEL, align=ALIGN_RIGHT,
    )

    # Suppress the inputs param when truthy via a no-op reference — keeps
    # the function signature stable for future per-project pulls without
    # ruff flagging the unused local.
    _ = inputs

    freeze_top(ws, row=4)
    print_landscape(ws)


# ── Sheet-builder support helpers (commit 2) ──────────────────────────────────


def _aggregate_scenario_line_items(
    items_by_project: dict[UUID, list[CashFlowLineItem]],
) -> dict[int, dict[str, Decimal]]:
    combined: dict[int, dict[str, Decimal]] = {}
    for items in items_by_project.values():
        per_year = _annual_line_items(items)
        for year, by_label in per_year.items():
            bucket = combined.setdefault(year, {})
            for label, amount in by_label.items():
                bucket[label] = bucket.get(label, Decimal(0)) + amount
    return combined


def _aggregate_scenario_line_items_by_category(
    items_by_project: dict[UUID, list[CashFlowLineItem]],
) -> dict[str, dict[int, dict[str, Decimal]]]:
    """Returns ``{category: {year: {label: amount}}}``.

    Aggregates across projects per LP feedback Option C: same exact label
    across projects → one combined row, no project-name suffixing. Labels
    are stripped of leading/trailing whitespace defensively so e.g.
    ``"CapEx Reserve"`` and ``"CapEx Reserve "`` collapse into one row.

    Categories follow ``LineItemCategory``: ``income`` / ``expense`` /
    ``debt_service`` / ``capex_reserve`` / ``capital_event``. The Pro Forma
    splits this into separate Revenue (income) and OpEx (expense) tables;
    capital events are summed for the Cash Flow sheet's "Capital Events"
    row.
    """
    out: dict[str, dict[int, dict[str, Decimal]]] = {}
    for items in items_by_project.values():
        for li in items:
            year = _period_to_year(li.period)
            category = str(getattr(li.category, "value", li.category) or "")
            label = (li.label or "").strip()
            # Expense labels get folded onto the canonical OpEx vocabulary
            # so legacy free-text duplicates ("Water / Sewer" vs "Water/Sewer",
            # "Property Tax" vs "Real Estate Taxes") collapse into one row.
            # Income labels stay as-is — Phase B1 Option C dedup already
            # handles cross-project name collisions for revenue streams.
            if category == "expense":
                label = normalize_opex_label(label)
            cat_dict = out.setdefault(category, {})
            year_dict = cat_dict.setdefault(year, {})
            year_dict[label] = year_dict.get(label, Decimal(0)) + _coerce_decimal(
                li.net_amount or 0
            )
    return out


_CAPITAL_EVENT_PREFIXES = ("Refi —", "Acquisition", "Sale", "Prepay", "Exit", "Purchase Price", "Closing Costs")


def _capital_events_by_year(
    annual_items: dict[int, dict[str, Decimal]],
) -> dict[int, Decimal]:
    """Legacy unsigned capital-event sum. Deprecated — use
    ``_signed_capital_events_by_year`` for sheet display.

    The engine writes line-item ``net_amount`` as a positive number with
    the sign convention encoded in ``adjustments['direction']``. This
    helper sums the bare amounts and so produces a positive value for
    Y0 acquisition costs — wrong sign for an investor read where outflows
    must be negative. Retained only because some legacy callers still
    consume this shape; new callers should use the signed variant.
    """
    out: dict[int, Decimal] = {}
    for year, by_label in annual_items.items():
        total = Decimal(0)
        for label, amount in by_label.items():
            if any(p in label for p in _CAPITAL_EVENT_PREFIXES):
                total += amount
        out[year] = total
    return out


def _signed_capital_events_by_year_for_project(
    items: list[CashFlowLineItem],
) -> dict[int, Decimal]:
    """Per-project signed capital events, respecting direction metadata.

    The engine writes ``CashFlowLineItem.adjustments['direction']`` with
    ``"outflow"`` or ``"inflow"``; ``net_amount`` is always a positive
    magnitude. This helper applies the sign so outflows render negative
    and inflows render positive in the export — matching the engine's
    own ``net_cash_flow`` invariant
    (``NCF = NOI - DS - capital_outflow + capital_inflow``).

    Without this fix, the Cash Flow sheet's Y0 Capital Events row showed
    ``+$5M`` for a $5M acquisition outflow, and the derived Unlevered CF
    row inherited the wrong sign — see Subject Model Review V2-B for the
    full diagnosis.
    """
    out: dict[int, Decimal] = {}
    for li in items:
        label = (li.label or "").strip()
        if not any(p in label for p in _CAPITAL_EVENT_PREFIXES):
            continue
        amount = _coerce_decimal(li.net_amount or 0)
        adjustments = li.adjustments or {}
        if adjustments.get("direction") == "outflow":
            amount = -amount
        year = _period_to_year(li.period)
        out[year] = out.get(year, Decimal(0)) + amount
    return out


def _signed_capital_events_by_year(
    items_by_project: dict[UUID, list[CashFlowLineItem]],
) -> dict[int, Decimal]:
    """Scenario-wide signed capital events: sum the per-project signed
    series. Outflows negative, inflows positive — see
    ``_signed_capital_events_by_year_for_project`` for the per-project
    rationale."""
    out: dict[int, Decimal] = {}
    for items in items_by_project.values():
        per_project = _signed_capital_events_by_year_for_project(items)
        for year, amount in per_project.items():
            out[year] = out.get(year, Decimal(0)) + amount
    return out


def _worst_dscr(per_project: list[dict]) -> Decimal | None:
    """Lowest non-null DSCR across projects (covenant binds at the weakest one).

    Retained for callers that want the per-loan worst-case view; the
    Underwriting Summary now leads with the combined DSCR instead (see
    ``_combined_dscr``) per LP feedback.
    """
    candidates = [
        _coerce_decimal(p.get("dscr"))
        for p in per_project
        if p.get("dscr") is not None
    ]
    return min(candidates) if candidates else None


def _format_currency_short(amount: Decimal | None) -> str:
    """Render a Decimal dollar amount as a compact human-readable string
    for use in hint text — ``$7.85M``, ``$869K``, ``$1.2K``, ``$0``."""
    if amount is None:
        return "$0"
    abs_amount = abs(amount)
    sign = "-" if amount < 0 else ""
    if abs_amount >= 1_000_000:
        return f"{sign}${abs_amount / Decimal('1000000'):.2f}M"
    if abs_amount >= 1_000:
        return f"{sign}${abs_amount / Decimal('1000'):.0f}K"
    return f"{sign}${abs_amount:.0f}"


def _compute_sources_gap(ctx: dict) -> tuple[Decimal, Decimal, Decimal]:
    """Compute scenario-wide ``(uses_total, sources_total, gap)``.

    ``gap = uses_total − sources_total``: positive means deal is undersized
    (Uses exceed funded Sources), negative means surplus.

    Mirrors the Underwriting Summary's S&U math so the Cover banner reads
    the same number the LP sees on the rollup. Pure aggregation — no DB
    roundtrip; reads ``use_lines`` (per project) + ``junctions``
    (junction-aggregated source principals) from ctx.
    """
    use_lines_by_project: dict[UUID, list[UseLine]] = ctx["use_lines"]
    junctions: list[CapitalModuleProject] = ctx["junctions"]
    capital_modules: list[CapitalModule] = ctx["capital_modules"]

    uses_total = Decimal(0)
    for uls in use_lines_by_project.values():
        for ul in uls:
            phase = str(getattr(ul.phase, "value", ul.phase) or "")
            if phase == "exit":
                continue
            uses_total += _coerce_decimal(ul.amount or 0)

    junction_amount: dict[UUID, Decimal] = {}
    for j in junctions:
        junction_amount[j.capital_module_id] = junction_amount.get(
            j.capital_module_id, Decimal(0)
        ) + _coerce_decimal(j.amount or 0)
    sources_total = Decimal(0)
    for module in capital_modules:
        amount = junction_amount.get(module.id) or _coerce_decimal(
            (module.source or {}).get("amount") or 0
        )
        sources_total += amount

    return uses_total, sources_total, uses_total - sources_total


def _combined_dscr(per_project: list[dict]) -> Decimal | None:
    """Combined DSCR = Σ NOI_stabilized / Σ Debt Service across projects.

    The engine doesn't materialize a per-project ``debt_service`` scalar;
    we reverse-derive ``ds_per_project = noi_stabilized / dscr`` and sum
    both sides to get a coverage figure that's a singular ratio rather
    than a worst-case across loans. Returns None when nothing has a
    non-zero DSCR (compute hasn't run, or no debt is sized).
    """
    total_noi = Decimal(0)
    total_ds = Decimal(0)
    for p in per_project:
        noi = _coerce_decimal(p.get("noi_stabilized") or 0)
        dscr = _coerce_decimal(p.get("dscr") or 0)
        if dscr <= 0 or noi <= 0:
            continue
        total_noi += noi
        total_ds += noi / dscr
    return (total_noi / total_ds) if total_ds > 0 else None


def _combined_unlevered_irr(
    cash_flows_by_project: dict[UUID, list[CashFlow]],
) -> Decimal | None:
    """Combined unlevered IRR — XIRR over per-period unlevered CF totals.

    Sums each project's unlevered cash flow (NCF + DS) per period, then
    runs the engine's pyxirr helper. Mirrors the rollup engine's path
    for the levered version (``rollup_irr``) but reverses out debt
    service so the result represents the asset-level return.
    Returns None on the typical no-pyxirr / no-sign-change cases.
    """
    from app.engines.cashflow import _compute_xirr  # late import — keep this module's imports lean

    period_totals: dict[int, Decimal] = {}
    for cf_list in cash_flows_by_project.values():
        for cf in cf_list:
            ncf = _coerce_decimal(cf.net_cash_flow or 0)
            ds = _coerce_decimal(cf.debt_service or 0)
            unlevered = ncf + ds
            period_totals[cf.period] = period_totals.get(cf.period, Decimal(0)) + unlevered
    if not period_totals:
        return None
    series = [period_totals[p] for p in sorted(period_totals)]
    pct_whole = _compute_xirr(series)
    if pct_whole == 0:
        return None
    # _compute_xirr returns percent as whole number (e.g. 12.34 = 12.34%);
    # PCT format wants a fraction.
    return pct_whole / Decimal(100)


def _coc_year_one(
    rollup_waterfall: list[dict],
    capital_modules: list[CapitalModule],
) -> Decimal | None:
    """Cash-on-Cash Year 1 = Σ equity distributions in periods 1-12 ÷ contributions.

    Per CRE convention "year 1" = first 12 months from deal close. If the
    deal is still mid-construction during year 1, this comes out 0 or
    negative — that's the honest number; the LP reads it in context.
    Returns None when there's no equity stack with non-zero commitments.
    """
    y1_per_module: dict[str, Decimal] = {}
    for row in rollup_waterfall:
        mid = row.get("capital_module_id")
        period = row.get("period") or 0
        if not mid or period < 1 or period > 12:
            continue
        amount = _coerce_decimal(row.get("cash_distributed") or 0)
        y1_per_module[mid] = y1_per_module.get(mid, Decimal(0)) + amount

    total_y1_dist = Decimal(0)
    total_contrib = Decimal(0)
    for module in capital_modules:
        if _funder_class(module) != "Equity":
            continue
        commitment = _coerce_decimal((module.source or {}).get("amount") or 0)
        if commitment <= 0:
            continue
        total_contrib += commitment
        total_y1_dist += y1_per_module.get(str(module.id), Decimal(0))
    return (total_y1_dist / total_contrib) if total_contrib > 0 else None


def _sum_per_project_field(per_project: list[dict], field: str) -> Decimal:
    return sum(
        (_coerce_decimal(p.get(field) or 0) for p in per_project),
        Decimal(0),
    )


def _longest_hold_months(per_project: list[dict]) -> int | None:
    candidates = [
        int(p.get("total_timeline_months") or 0)
        for p in per_project
        if p.get("total_timeline_months")
    ]
    return max(candidates) if candidates else None


def _project_sheet_name(idx: int, project_name: str | None) -> str:
    """Build the per-project sheet name (commit 3 will create these sheets).

    Format ``P{n} {Name}`` truncated to Excel's 31-char ceiling. The exact
    rule comes from plan §2: prefix is `P` + 1- or 2-digit ordinal + space
    (4 chars max), then up to ``PROJECT_SHEET_NAME_BUDGET`` chars of name.
    """
    name = (project_name or "").strip()
    truncated = name[:PROJECT_SHEET_NAME_BUDGET].rstrip()
    return f"P{idx} {truncated}".rstrip()


def _is_lp_funder(module_or_role) -> bool:
    """True when the module/equity_role indicates an LP investor."""
    if hasattr(module_or_role, "equity_role"):
        er = str(getattr(module_or_role, "equity_role", "") or "").replace("EquityRole.", "").lower()
        if er:
            return er == "lp"
        # fall back to vehicle_type check — equity with no role treated as LP
        return _funder_class(module_or_role) == "Equity"
    label = (str(getattr(module_or_role, "value", module_or_role)) or "").lower()
    return "common_equity" in label or "preferred" in label or "lp" in label


def _is_gp_funder(module_or_role) -> bool:
    """True when the module/equity_role indicates a GP investor."""
    if hasattr(module_or_role, "equity_role"):
        er = str(getattr(module_or_role, "equity_role", "") or "").replace("EquityRole.", "").lower()
        return er == "gp"
    label = (str(getattr(module_or_role, "value", module_or_role)) or "").lower()
    return "owner_equity" in label or label == "gp" or "developer" in label


def _lp_gp_irr_from_rollup(
    rollup: list[dict], capital_modules: list[CapitalModule]
) -> tuple[Decimal | None, Decimal | None]:
    """Pull LP and GP IRR percentages from the latest waterfall row per module.

    Returns (LP IRR fraction, GP IRR fraction) — None when no eligible rows.
    """
    by_module: dict[str, dict] = {}
    for row in rollup:
        mid = row.get("capital_module_id")
        if not mid:
            continue
        prior = by_module.get(mid)
        if prior is None or (row.get("period") or 0) > (prior.get("period") or 0):
            by_module[mid] = row
    module_index = {str(m.id): m for m in capital_modules}

    lp_vals: list[Decimal] = []
    gp_vals: list[Decimal] = []
    for mid, row in by_module.items():
        module = module_index.get(mid)
        if module is None or row.get("party_irr_pct") is None:
            continue
        irr_fraction = _coerce_pct(row.get("party_irr_pct"))
        if _is_lp_funder(module):
            lp_vals.append(irr_fraction)
        elif _is_gp_funder(module):
            gp_vals.append(irr_fraction)

    lp_irr = sum(lp_vals, Decimal(0)) / Decimal(len(lp_vals)) if lp_vals else None
    gp_irr = sum(gp_vals, Decimal(0)) / Decimal(len(gp_vals)) if gp_vals else None
    return lp_irr, gp_irr


def _equity_multiples_from_rollup(
    rollup: list[dict], capital_modules: list[CapitalModule]
) -> tuple[Decimal | None, Decimal | None]:
    """Compute LP / GP equity multiples from cumulative distributed totals.

    EM = total distributions ÷ total contributions. We don't have direct
    contribution data here, so use ``cumulative_distributed`` as the
    numerator and the module's source amount as the denominator (the
    committed amount is the contribution proxy for equity modules).
    """
    by_module: dict[str, Decimal] = {}
    for row in rollup:
        mid = row.get("capital_module_id")
        if not mid:
            continue
        cum = _coerce_decimal(row.get("cumulative_distributed") or 0)
        prev = by_module.get(mid)
        if prev is None or cum > prev:
            by_module[mid] = cum

    lp_dist = lp_contrib = Decimal(0)
    gp_dist = gp_contrib = Decimal(0)
    for module in capital_modules:
        commitment = _coerce_decimal((module.source or {}).get("amount") or 0)
        cum = by_module.get(str(module.id), Decimal(0))
        if _is_lp_funder(module):
            lp_dist += cum
            lp_contrib += commitment
        elif _is_gp_funder(module):
            gp_dist += cum
            gp_contrib += commitment

    lp_em = (lp_dist / lp_contrib) if lp_contrib > 0 else None
    gp_em = (gp_dist / gp_contrib) if gp_contrib > 0 else None
    return lp_em, gp_em


def _to_excel_number(value):
    """Coerce a Decimal/None to a plain float-or-blank for openpyxl cells.

    Returns "" for None so empty cells render blank, not as the literal
    string "None". Mirrors ``_workbook_helpers.to_excel_value`` but is
    inlined here for the hot per-cell path.
    """
    if value is None:
        return ""
    if isinstance(value, Decimal):
        f = float(value)
        return int(f) if f == int(f) else round(f, 6)
    return value


# ── Debt Schedule sheet (Phase H2) ────────────────────────────────────────────


def _build_debt_schedule(ws, registry: CellRegistry, ctx: dict) -> None:
    """Debt Schedule — per-module loan terms + perm-loan amortization table.

    Two sections:

      Loan Summary: one row per debt-class CapitalModule with the contractual
      terms an LP / lender reads at first scan — principal (junction-aggregated),
      rate, term in months (loan's active window), amort years, IO months,
      carry type, annual P&I payment, balloon balance at term end.

      Perm Loan Amortization: year-by-year balance / payment / interest /
      principal table for the *largest* permanent-debt module on the stack
      (or the largest senior-debt module if no permanent debt is present).
      One row per year of the amort term, capped at 30 years for readability.

    Bridge / construction / pre-dev loans don't get amort tables — they're
    typically interest-only and short-term, so the summary table is enough.
    """
    capital_modules: list[CapitalModule] = ctx["capital_modules"]
    junctions: list[CapitalModuleProject] = ctx["junctions"]

    set_widths(ws, [28, 16, 14, 10, 10, 10, 12, 18, 13, 14, 14])

    section_label(ws, 1, "Loan Summary — Per Capital Module", span_cols=11)
    header_row(
        ws, 2,
        ["Module", "Funder Type", "Principal", "Rate", "Term (mo)",
         "Amort (yrs)", "IO Months", "Carry Type", "Day Count", "Annual P&I", "Balloon"],
    )

    # Junction-aggregated principal per module (mirrors the Investor Returns
    # path — one shared debt module covering N projects has its principal
    # split across N junctions; the loan's headline principal is the sum).
    junction_principal: dict[UUID, Decimal] = {}
    for j in junctions:
        junction_principal[j.capital_module_id] = junction_principal.get(
            j.capital_module_id, Decimal(0)
        ) + _coerce_decimal(j.amount or 0)

    debt_modules = [m for m in capital_modules if _funder_class(m) == "Debt"]

    cur_row = 3
    if not debt_modules:
        ws.cell(
            row=cur_row, column=1,
            value="(no debt modules — Loan Summary populates when debt is added to the Capital Stack)",
        ).font = FONT_HINT
        cur_row += 1

    # Track the Loan Summary row of the perm-loan winner so the amort
    # table below can absolute-ref principal/rate/amort/IO cells.
    perm_candidate: tuple[CapitalModule, Decimal, int] | None = None  # (module, principal, loan_row)
    for m_idx, module in enumerate(debt_modules, start=1):
        source = module.source or {}
        carry = module.carry or {}
        principal = junction_principal.get(module.id) or _coerce_decimal(
            source.get("amount") or 0
        )
        rate_raw = source.get("interest_rate_pct") or carry.get("io_rate_pct") or 0
        rate = _coerce_pct(rate_raw) if rate_raw else None
        amort_years = source.get("amort_term_years") or 30
        io_months = source.get("io_months") or 0
        carry_type = _resolve_carry_type(carry)
        term_months = _loan_active_term_months(module, ctx)

        # Annual P&I — only meaningful for amortizing carry types
        annual_pi: Decimal | None = None
        if carry_type == "pi" and rate_raw:
            from app.engines.cashflow import _monthly_pmt
            monthly = _monthly_pmt(principal, float(rate_raw), int(amort_years))
            annual_pi = monthly * Decimal(12)

        # Balloon balance at end of term
        balloon: Decimal | None = None
        if rate_raw and term_months and term_months > 0:
            from app.engines.cashflow import _balloon_balance
            balloon = _balloon_balance(
                principal, float(rate_raw), int(amort_years),
                int(term_months), io_months=int(io_months),
            )

        # Day-count convention — read from carry schema, default 30/360.
        # Labels match lender term-sheet wording per CRE best practice.
        _DC_LABELS = {"30_360": "30/360", "actual_365": "Actual/365", "actual_360": "Actual/360"}
        day_count_label = _DC_LABELS.get(carry.get("day_count") or "30_360", "30/360")

        ws.cell(row=cur_row, column=1, value=module.label or _funder_type_label(module)).font = FONT_VALUE
        ws.cell(row=cur_row, column=2, value=_funder_type_label(module)).font = FONT_VALUE
        registry.write(
            ws, cur_row, 3, principal,
            name=f"s_loan_{m_idx}_principal", fmt=ACCOUNTING,
            font=FONT_VALUE, align=ALIGN_RIGHT,
        )
        _write_optional(ws, cur_row, 4, rate, registry,
                        name=f"s_loan_{m_idx}_rate", fmt=PCT)
        # Phase E refinement: register Term (months) as a named cell so the
        # Pro Forma Debt Service per-year SUM can gate each loan by its
        # hold term — a 10-year hold loan stops contributing P&I after Y10
        # instead of overstating debt service across the entire chain.
        _write_optional(
            ws, cur_row, 5, int(term_months) if term_months else None,
            registry, name=f"s_loan_{m_idx}_term_months", fmt=INT_COMMA,
        )
        ws.cell(row=cur_row, column=6, value=int(amort_years)).font = FONT_VALUE
        ws.cell(row=cur_row, column=7, value=int(io_months)).font = FONT_VALUE
        ws.cell(
            row=cur_row, column=8,
            value=carry_type.replace("_", " ").title() if carry_type else _DASH,
        ).font = FONT_VALUE
        ws.cell(row=cur_row, column=9, value=day_count_label).font = FONT_VALUE
        # Formula-conversion plan §4.4 (commit 6): Annual P&I for ``pi``
        # carry-type loans is a clean PMT derivation of principal / rate /
        # amort term. Wire it as a formula referencing the named-range
        # inputs from the Loan Summary row (principal col C, rate col D)
        # plus the amort years in col F, so an LP can change any of the
        # three and watch Annual P&I recompute. Other carry types
        # (io_only, interest_reserve, capitalized_interest) leave the
        # cell as the engine value (None → em-dash) since their annual
        # outlay isn't a simple PMT.
        if carry_type == "pi" and rate_raw and principal > 0 and amort_years:
            # PMT(rate/12, amort*12, -principal) * 12 yields a positive
            # annual payment. Reference cells: principal=C<row>,
            # rate=D<row>, amort_years=F<row>. amort_years is an int
            # in col F (header "Amort (yrs)"). IFERROR guards against
            # degenerate inputs (zero rate, zero term).
            pmt_formula = (
                f"=IFERROR(PMT(D{cur_row}/12,F{cur_row}*12,-C{cur_row})*12,0)"
            )
            cell = ws.cell(row=cur_row, column=10, value=pmt_formula)
            cell.number_format = ACCOUNTING
            cell.font = FONT_VALUE
            cell.alignment = ALIGN_RIGHT
            registry.register(
                f"s_loan_{m_idx}_annual_pi", ws.title, cur_row, 10,
            )
        else:
            _write_optional(
                ws, cur_row, 10, annual_pi, registry,
                name=f"s_loan_{m_idx}_annual_pi", fmt=ACCOUNTING,
            )
        _write_optional(
            ws, cur_row, 11, balloon, registry,
            name=f"s_loan_{m_idx}_balloon", fmt=ACCOUNTING,
        )

        # Track largest debt module for the amort table below.
        ft_vt = str(getattr(module, "vehicle_type", "") or "").replace("VehicleType.", "")
        if ft_vt == "debt":
            if perm_candidate is None or principal > perm_candidate[1]:
                perm_candidate = (module, principal, cur_row)
        cur_row += 1

    # ── Sizing & Carry-Type Notes (graceful-degradation disclosure) ───────
    # Surface engine-driven approximations and binding constraints that the
    # LP would otherwise have to infer from missing context. Each note is
    # only emitted when the underlying condition is present, so the block
    # stays empty (no header) for a vanilla perm-debt deal.
    notes: list[str] = []
    projects_list = ctx.get("projects") or []
    default_inputs_dn = None
    if projects_list:
        default_inputs_dn = (ctx.get("operational_inputs") or {}).get(
            projects_list[0].id
        )
    sizing_mode = getattr(default_inputs_dn, "debt_sizing_mode", None)
    if sizing_mode == "dscr_capped":
        notes.append(
            "Debt sizing: DSCR-capped. Principals reflect the maximum loan "
            "amount consistent with the minimum DSCR constraint. Editing "
            "principal or rate in this workbook does NOT re-solve the DSCR "
            "cap — recompute in the app to update."
        )
    elif sizing_mode == "dual_constraint":
        notes.append(
            "Debt sizing: dual constraint (DSCR-capped AND LTV-capped). "
            "Principals reflect the binding of both constraints; the "
            "tighter of the two governs."
        )
    carry_types_seen = {
        _resolve_carry_type(m.carry or {}) for m in debt_modules
    }
    if "interest_reserve" in carry_types_seen:
        notes.append(
            "Interest reserve: shown values use the average-draw "
            "approximation (draws-to-date ÷ 2). Engine cashflow uses "
            "day-precise period-by-period accrual; small variances "
            "between this sheet and the per-period schedule are expected."
        )
    if "capitalized_interest" in carry_types_seen:
        notes.append(
            "Capitalized interest (PIK): balance grows monthly at "
            "rate ÷ 12. Debt service = 0 during the PIK window; "
            "accrued interest is repaid at sale."
        )
    if notes:
        cur_row += 1
        section_label(ws, cur_row, "Notes", span_cols=6)
        cur_row += 1
        for note in notes:
            cell = ws.cell(row=cur_row, column=1, value=note)
            cell.font = FONT_HINT
            cell.alignment = ALIGN_WRAP
            ws.merge_cells(
                start_row=cur_row, start_column=1,
                end_row=cur_row, end_column=8,
            )
            ws.row_dimensions[cur_row].height = 30
            cur_row += 1

    # ── Perm Loan Amortization Table ──────────────────────────────────────
    if perm_candidate is None:
        return

    perm_module, perm_principal, perm_loan_row = perm_candidate
    perm_source = perm_module.source or {}
    perm_rate_raw = perm_source.get("interest_rate_pct") or 0
    perm_amort_yrs = int(perm_source.get("amort_term_years") or 30)

    if not perm_rate_raw:
        return

    cur_row += 2
    section_label(
        ws, cur_row,
        f"Amortization — {perm_module.label or _funder_type_label(perm_module)}",
        span_cols=6,
    )
    cur_row += 1
    header_row(
        ws, cur_row,
        ["Year", "Beginning Balance", "Annual Payment", "Interest", "Principal", "Ending Balance"],
    )
    cur_row += 1

    # Cap at 30 years for readability — the LP doesn't need a 40-year amort
    # table on a Phase 1 deal review.
    display_years = min(perm_amort_yrs, 30)

    # Formula-conversion plan §4.4 (commit 7): the amort table is wired as
    # CUMIPMT / CUMPRINC over the Loan Summary's principal/rate/amort/IO
    # cells so LP edits flow through. Cell refs:
    #   $C${perm_loan_row} = principal
    #   $D${perm_loan_row} = rate (fraction, e.g. 0.065)
    #   $F${perm_loan_row} = amort years (int)
    #   $G${perm_loan_row} = IO months (int)
    # CUMIPMT returns a negative number under Excel's sign convention; we
    # negate to keep the amort table values positive. IO branch: payment =
    # principal * rate (annual interest), principal_paid = 0.
    PR = f"$C${perm_loan_row}"   # principal
    RT = f"$D${perm_loan_row}"   # rate
    AY = f"$F${perm_loan_row}"   # amort yrs
    IO = f"$G${perm_loan_row}"   # io months

    for year in range(1, display_years + 1):
        start_period = (year - 1) * 12 + 1
        end_period = year * 12

        # Beginning balance: first year pulls principal; later years reach
        # back to the prior row's End Balance cell (col 6).
        if year == 1:
            beg_formula = f"={PR}"
        else:
            beg_formula = f"=F{cur_row - 1}"

        annual_pmt_formula = (
            f"=IF({IO}>={end_period},{PR}*{RT},"
            f"IFERROR(-PMT({RT}/12,{AY}*12,{PR})*12,0))"
        )
        interest_formula = (
            f"=IF({IO}>={end_period},{PR}*{RT},"
            f"IFERROR(-CUMIPMT({RT}/12,{AY}*12,{PR},{start_period},{end_period},0),0))"
        )
        principal_formula = f"=C{cur_row}-D{cur_row}"  # AnnualPmt − Interest
        end_formula = f"=B{cur_row}-E{cur_row}"  # Beg − Principal

        ws.cell(row=cur_row, column=1, value=year).font = FONT_VALUE
        for col, formula in enumerate(
            (beg_formula, annual_pmt_formula, interest_formula,
             principal_formula, end_formula),
            start=2,
        ):
            cell = ws.cell(row=cur_row, column=col, value=formula)
            cell.number_format = ACCOUNTING
            cell.font = FONT_VALUE
            cell.alignment = ALIGN_RIGHT
        cur_row += 1

    # ── Construction-to-Perm Status (per loan) ────────────────────────────
    # First slice of construction-to-perm formula gating: per-loan scalar
    # cells that surface "when does this loan's permanent tranche
    # originate" and "has perm origination occurred by loan term-end" as
    # formulas referencing the per-project phase-plan cells registered by
    # _build_project_sheet. Future debt-service formula gating can
    # consume these named ranges directly.
    _build_c2p_status_block(ws, registry, ctx, debt_modules, start_row=cur_row + 2)

    freeze_top(ws, row=3)
    print_landscape(ws)


def _build_c2p_status_block(
    ws,
    registry: CellRegistry,
    ctx: dict,
    debt_modules: list,
    *,
    start_row: int,
) -> None:
    """Render the Construction-to-Perm Status section at the bottom of the
    Debt Schedule sheet. One row per debt module that funds at least one
    project with a registered perm-origination month.

    For each such row the section writes:

      ``s_loan_<n>_perm_origination_month`` — formula pulling
      ``MAX(p<projidx>_perm_origination_month, ...)`` across the projects
      this loan funds (multi-project loans take the latest perm switch,
      most conservative for "is the loan active in operations").

      ``s_loan_<n>_active_in_operations`` — boolean formula:
      ``=IF(s_loan_<n>_term_months >= s_loan_<n>_perm_origination_month,
      TRUE, FALSE)``. Tells an LP whether the loan's active term extends
      past perm origination — i.e. whether its operations-phase carry
      ever applies in-model.

    Skipped entirely when no eligible projects exist (e.g. pure
    acquisition scenarios with no construction-side phase).
    """
    if not debt_modules:
        return

    loan_proj_idxs = _perm_origination_loan_idxs(ctx)
    if not loan_proj_idxs:
        return

    section_label(ws, start_row, "Construction-to-Perm Status — Per Loan", span_cols=6)
    header_row(
        ws, start_row + 1,
        ["Module", "Funds Project(s)", "Perm Origination Month",
         "Loan Term (mo)", "Active in Operations", "Notes"],
    )
    row = start_row + 2
    any_written = False
    for m_idx, module in enumerate(debt_modules, start=1):
        eligible_proj_idxs = loan_proj_idxs.get(m_idx)
        if not eligible_proj_idxs:
            continue

        perm_name = f"s_loan_{m_idx}_perm_origination_month"
        active_name = f"s_loan_{m_idx}_active_in_operations"
        term_name = f"s_loan_{m_idx}_term_months"
        per_proj_refs = ",".join(
            f"p{idx}_perm_origination_month" for idx in eligible_proj_idxs
        )
        proj_display = ", ".join(f"P{idx}" for idx in eligible_proj_idxs)

        ws.cell(row=row, column=1, value=module.label or _funder_type_label(module)).font = FONT_VALUE
        ws.cell(row=row, column=2, value=proj_display).font = FONT_VALUE
        # MAX across projects = the latest perm switch (most conservative
        # bound for "loan is past perm origination").
        perm_formula = f"=IFERROR(MAX({per_proj_refs}),\"\")"
        registry.write(
            ws, row, 3, perm_formula,
            name=perm_name, fmt=INT_COMMA,
            font=FONT_VALUE, align=ALIGN_RIGHT,
        )
        # Term column: cross-ref to the existing per-loan term cell so an
        # LP edit to active window flows through automatically.
        ws.cell(row=row, column=4, value=f"={term_name}").font = FONT_VALUE
        ws.cell(row=row, column=4).number_format = INT_COMMA
        ws.cell(row=row, column=4).alignment = ALIGN_RIGHT
        # Active-in-operations boolean: TRUE iff the loan's term extends
        # past perm origination. Returns FALSE (not #N/A) when either
        # input is missing — keeps the section readable.
        active_formula = (
            f"=IFERROR(IF(AND(ISNUMBER({term_name}),"
            f"ISNUMBER({perm_name}),{term_name}>={perm_name}),"
            f"TRUE,FALSE),FALSE)"
        )
        registry.write(
            ws, row, 5, active_formula,
            name=active_name,
            font=FONT_VALUE, align=ALIGN_RIGHT,
        )
        # Notes column consumes the active-in-operations boolean as a
        # modeling sanity check — flags loans that retire before perm
        # origination (likely a misconfigured active window: a loan
        # cannot retire in construction and then re-originate as perm).
        notes_formula = (
            f"=IF({active_name},\"Active in ops by term end\","
            f"\"⚠ Retires before perm origination — check active window\")"
        )
        ws.cell(row=row, column=6, value=notes_formula).font = FONT_HINT
        any_written = True
        row += 1

    # Defensive: if no rows ended up written (e.g. every debt module funds
    # only ineligible projects), leave a hint instead of a header with no
    # rows under it.
    if not any_written:
        ws.cell(
            row=row, column=1,
            value="(no debt modules fund a project with a construction-side phase)",
        ).font = FONT_HINT


def _emit_phase_plan_block(
    ws,
    registry: CellRegistry,
    ctx: dict,
    *,
    project_idx: int,
    project,
    inputs,
    capital_modules: list,
    start_row: int,
) -> int:
    """Emit the per-project Phase Plan block onto ``ws`` starting at
    ``start_row``. Registers ``p{idx}_phase_*_{start,end,duration}_month``,
    ``p{idx}_perm_origination_month`` (when construction-side phases exist),
    and ``p{idx}_total_horizon_months``. Returns the next free row.

    Single source of truth for both call sites: per-project sheet
    (``_build_project_sheet``) for multi-project scenarios, and the
    Assumptions sheet (``_build_assumptions``) for single-project
    scenarios where the per-project sheet is suppressed.
    """
    milestones_by_project: dict = ctx.get("milestones") or {}
    project_milestones = milestones_by_project.get(project.id, [])
    scenario_obj = ctx.get("scenario")
    raw_pt = getattr(scenario_obj, "project_type", None)
    project_type_name = str(getattr(raw_pt, "value", raw_pt) or "")
    phase_windows: list = []
    if inputs is not None and project_type_name:
        try:
            phase_windows = build_project_phase_windows(
                project_type=project_type_name,
                inputs=inputs,
                milestones=project_milestones,
                capital_modules=capital_modules,
            )
        except ValueError:
            phase_windows = []
    if not phase_windows:
        return start_row

    cur = start_row
    section_label(ws, cur + 1, "Phase Plan (months)", span_cols=2)
    cur += 2
    for window in phase_windows:
        phase_name = window.period_type.value
        kv_row(
            ws, cur, f"  {phase_name} — start",
            window.start_month,
            name=f"p{project_idx}_phase_{phase_name}_start_month",
            registry=registry, fmt=INT_COMMA,
        ); cur += 1
        kv_row(
            ws, cur, f"  {phase_name} — end",
            window.end_month,
            name=f"p{project_idx}_phase_{phase_name}_end_month",
            registry=registry, fmt=INT_COMMA,
        ); cur += 1
        kv_row(
            ws, cur, f"  {phase_name} — duration",
            window.duration_months,
            name=f"p{project_idx}_phase_{phase_name}_duration_months",
            registry=registry, fmt=INT_COMMA,
        ); cur += 1
    perm_month = perm_origination_month(phase_windows)
    if perm_month is not None:
        kv_row(
            ws, cur, "Perm origination month",
            perm_month,
            name=f"p{project_idx}_perm_origination_month",
            registry=registry, fmt=INT_COMMA,
        ); cur += 1
    kv_row(
        ws, cur, "Total horizon (months)",
        total_horizon_months(phase_windows),
        name=f"p{project_idx}_total_horizon_months",
        registry=registry, fmt=INT_COMMA,
    ); cur += 1
    return cur


def _perm_origination_loan_idxs(ctx: dict) -> dict[int, list[int]]:
    """Return ``{loan_m_idx: [project_idxs]}`` for every debt module that
    funds at least one project with a registered perm-origination cell.

    Single source of truth for two consumers:

      - ``_build_c2p_status_block`` — emits the Construction-to-Perm Status
        section using the project-idx list to build the formula's
        ``MAX(p<idx>_perm_origination_month, ...)`` expression.
      - ``_debt_service_formula_for_year`` — gates each PI loan's annual
        P&I contribution on the loan's perm origination month so PI
        doesn't accrue during construction years.

    Returns an empty dict when no project has a construction-side
    phase. The named cells these consumers reference (``p<idx>_*``)
    are written by ``_emit_phase_plan_block`` on the per-project sheet
    (multi-project) or on the Assumptions sheet (single-project), so
    they exist in both cases when a construction-side phase is present.
    """
    projects: list = ctx.get("projects") or []
    if not projects:
        return {}
    capital_modules: list[CapitalModule] = ctx.get("capital_modules") or []
    junctions: list = ctx.get("junctions") or []
    inputs_by_project: dict = ctx.get("operational_inputs") or {}
    milestones_by_project: dict = ctx.get("milestones") or {}
    scenario_obj = ctx.get("scenario")
    raw_pt = getattr(scenario_obj, "project_type", None)
    project_type_name = str(getattr(raw_pt, "value", raw_pt) or "")
    if not project_type_name:
        return {}

    project_idx_by_id: dict = {p.id: idx for idx, p in enumerate(projects, start=1)}
    perm_eligible_idx: set[int] = set()
    for idx, project in enumerate(projects, start=1):
        inputs = inputs_by_project.get(project.id)
        if inputs is None:
            continue
        try:
            windows = build_project_phase_windows(
                project_type=project_type_name,
                inputs=inputs,
                milestones=milestones_by_project.get(project.id, []),
                capital_modules=capital_modules,
            )
        except ValueError:
            continue
        if perm_origination_month(windows) is not None:
            perm_eligible_idx.add(idx)

    if not perm_eligible_idx:
        return {}

    module_projects: dict = {}
    for j in junctions:
        module_projects.setdefault(j.capital_module_id, set()).add(j.project_id)

    debt_modules = [m for m in capital_modules if _funder_class(m) == "Debt"]
    out: dict[int, list[int]] = {}
    for m_idx, module in enumerate(debt_modules, start=1):
        funded = module_projects.get(module.id, set())
        eligible = sorted(
            project_idx_by_id[pid] for pid in funded
            if pid in project_idx_by_id
            and project_idx_by_id[pid] in perm_eligible_idx
        )
        if eligible:
            out[m_idx] = eligible
    return out


def _pmt_loan_indices(capital_modules: list[CapitalModule]) -> list[int]:
    """1-based indices of debt modules whose Annual P&I cell on the Debt
    Schedule is a PMT formula (carry_type == 'pi', rate + principal +
    amort all set). Enumeration order matches ``_build_debt_schedule``'s
    ``enumerate(debt_modules, start=1)`` so the returned ints align with
    the ``s_loan_{n}_annual_pi`` named ranges that sheet registers.

    Used by ``_build_uw_proforma`` to emit the Pro Forma Debt Service
    row as a SUM over those names so an LP editing rate or principal on
    the Debt Schedule sees the Pro Forma debt service shift in lock-step.
    """
    out: list[int] = []
    debt_modules = [m for m in capital_modules if _funder_class(m) == "Debt"]
    for m_idx, module in enumerate(debt_modules, start=1):
        source = module.source or {}
        carry = module.carry or {}
        principal = _coerce_decimal(source.get("amount") or 0) or Decimal(0)
        rate_raw = source.get("interest_rate_pct") or carry.get("io_rate_pct") or 0
        amort_years = source.get("amort_term_years") or 0
        if (
            _resolve_carry_type(carry) == "pi"
            and rate_raw
            and principal > 0
            and amort_years
        ):
            out.append(m_idx)
    return out


def _resolve_carry_type(carry: dict) -> str:
    """Best-effort carry-type read from the carry JSON. Mirrors what the
    cashflow engine does (`_carry_type_for_phase`) but simpler — just pulls
    the operations-phase carry if present, else top-level carry_type, else
    "io_only" as a default."""
    if not carry:
        return "io_only"
    phases = carry.get("phases") or []
    for phase in phases:
        if phase.get("name") == "operation":
            return phase.get("carry_type") or "io_only"
    return carry.get("carry_type") or (phases[0].get("carry_type") if phases else "io_only")


def _loan_active_term_months(module: CapitalModule, ctx: dict) -> int | None:
    """Approximate term-in-months for a loan based on its active phase
    window. Returns the count of months from active_phase_start through
    the scenario's modeled horizon — close enough for a Loan Summary
    table; the engine has more precise per-loan windowing
    (`_loan_pre_op_months`) but it's not exposed in ctx today.
    """
    # Fall back to the scenario's longest project's total_timeline_months
    # since loans typically extend to exit. Bridge loans get retired earlier
    # (their ``exit_terms.vehicle`` points at the perm) but the terminal
    # value table here is illustrative.
    rollup_summary = ctx.get("rollup_summary") or {}
    per_project = rollup_summary.get("per_project") or []
    timeline_candidates = [
        int(p.get("total_timeline_months") or 0)
        for p in per_project
        if p.get("total_timeline_months")
    ]
    if not timeline_candidates:
        return None
    return max(timeline_candidates)


def _build_assumptions(ws, registry: CellRegistry, ctx: dict) -> None:
    """Assumptions sheet: scenario-level / per-project / capital-stack blocks."""
    scenario: DealModel = ctx["scenario"]
    projects: list[Project] = ctx["projects"]
    inputs_by_project: dict[UUID, OperationalInputs] = ctx["operational_inputs"]
    use_lines_by_project: dict[UUID, list[UseLine]] = ctx["use_lines"]
    unit_mix_by_project: dict[UUID, list] = ctx["unit_mix"]
    capital_modules: list[CapitalModule] = ctx["capital_modules"]
    module_slugs: dict = ctx.get("module_slugs") or _compute_module_slugs(capital_modules)
    junctions: list[CapitalModuleProject] = ctx["junctions"]

    # Layout: 1 (label) + max(MAX_PROJECTS_PER_SCENARIO, len(CAPITAL_STACK_HEADERS)-1)
    # data columns. Block B writes into cols 2..1+N_projects; Block C writes
    # into cols 1..len(CAPITAL_STACK_HEADERS). The sheet is sized for both.
    label_w = 36
    project_col_widths = [22] * MAX_PROJECTS_PER_SCENARIO
    # Block C extra column widths for the post-expansion debt-assumption
    # surface (commit 1 of the formula-conversion plan). Eight new fields
    # added: term, amort, io_months, carry_type, day_count, dscr_min, ltv,
    # prepay. Existing 6 cols stay (label, funder, principal, rate,
    # auto-sized, covers).
    block_c_extra_widths = [12, 12, 12, 14, 12, 12, 10, 12]
    set_widths(ws, [label_w, *project_col_widths, *block_c_extra_widths])

    # ── Block A: Scenario-level ────────────────────────────────────────────
    # Default project's OperationalInputs carries scenario-level conceptual
    # fields (hold years, exit cap, reserve months, etc.) since these don't
    # vary per project today. When per-project becomes meaningful, this
    # block becomes the "default project" snapshot.
    default_project = projects[0] if projects else None
    default_inputs = (
        inputs_by_project.get(default_project.id) if default_project else None
    )

    section_label(ws, 1, "A. Scenario-Level Assumptions", span_cols=2)
    row = 2
    # Scenario / NOI Basis / Project Type rows are meta — derived from the
    # Scenario record, not "inputs" the LP would tweak. Keep black/calc.
    kv_row(ws, row, "Scenario Name", scenario.name,
           name="s_assumptions_scenario_name", registry=registry); row += 1
    kv_row(ws, row, "NOI Basis", _noi_basis_label(scenario.income_mode),
           name="s_assumptions_noi_basis", registry=registry); row += 1
    # `project_type` is typed Mapped[ProjectType] but stored as String(60)
    # — SQLAlchemy doesn't auto-coerce on read, so it comes back as a bare
    # string in production. Use the same safe pattern as _funder_type_label.
    project_type_label = getattr(
        scenario.project_type, "value", scenario.project_type
    ) or ""
    kv_row(ws, row, "Project Type (default)", str(project_type_label),
           name="s_assumptions_project_type", registry=registry); row += 1
    # The remaining Block A rows are user-editable inputs — render in blue
    # per the input/output color convention so the LP can tell at a glance
    # which numbers drive the model vs which are derived.
    # Hold Period removed — now per-perm-debt CapitalModule.source.hold_term_years.
    # Show the MAX across perm-debt modules as scenario-level summary.
    _perm_holds: list[int] = []
    for _cm in capital_modules:
        _cm_vt = str(getattr(_cm, "vehicle_type", "") or "").replace("VehicleType.", "")
        if _cm_vt != "debt":
            continue
        _src = getattr(_cm, "source", None) or {}
        _h = _src.get("hold_term_years") if isinstance(_src, dict) else None
        try:
            _hi = int(_h) if _h is not None else 0
        except (TypeError, ValueError):
            _hi = 0
        if _hi > 0:
            _perm_holds.append(_hi)
    _perm_hold_display = max(_perm_holds) if _perm_holds else None
    if _perm_hold_display is not None:
        kv_row(
            ws, row, "Hold Term (years, MAX of perm debt)",
            Decimal(str(_perm_hold_display)),
            name="s_hold_years", registry=registry, fmt=INT_COMMA, style="input",
        ); row += 1
    kv_row(
        ws, row, "Exit Cap Rate",
        _pct_value(default_inputs, "exit_cap_rate_pct"),
        name="s_exit_cap_rate", registry=registry, fmt=PCT, style="input",
    ); row += 1
    # Going-In Cap Rate at scenario level — sourced from default project's
    # OperationalInputs (per-project today; promotes to scenario-level when
    # editing lands). Needed so the UW Summary Property Valuation block's
    # Going-In Cap Value can ref a single named cell instead of duplicating
    # the per-project p1_going_in_cap_rate.
    kv_row(
        ws, row, "Going-In Cap Rate",
        _pct_value(default_inputs, "going_in_cap_rate_pct"),
        name="s_going_in_cap_rate", registry=registry, fmt=PCT, style="input",
    ); row += 1
    _streams_by_project: dict[UUID, list[IncomeStream]] = ctx.get(
        "income_streams", {}
    )
    _default_streams = (
        _streams_by_project.get(default_project.id, []) if default_project else []
    )
    kv_row(
        ws, row, "Revenue Growth Rate (annual)",
        _revenue_growth_default(_default_streams),
        name="s_revenue_growth_rate", registry=registry, fmt=PCT, style="input",
    ); row += 1
    kv_row(
        ws, row, "OpEx Growth Rate (annual)",
        _pct_value(default_inputs, "expense_growth_rate_pct_annual"),
        name="s_opex_growth_rate", registry=registry, fmt=PCT, style="input",
    ); row += 1
    kv_row(
        ws, row, "Operating Reserve (months)",
        _safe_decimal(default_inputs, "operation_reserve_months"),
        name="s_operating_reserve_months", registry=registry, fmt=INT_COMMA, style="input",
    ); row += 1
    # Graceful-degradation anchor date: lets the LP overlay their own
    # reporting calendar on the relative Y0/Y1/Y2 grid. Defaulted to the
    # scenario's creation date (most useful "as-of" anchor for a fresh
    # underwriting); downstream sheets can pick this up wherever calendar
    # dating is needed (DCF dating, milestone date-stamping) in future
    # phases. Today: input-only, no downstream wiring.
    _anchor_default = (
        scenario.created_at.date()
        if scenario.created_at is not None
        else None
    )
    kv_row(
        ws, row, "Anchor Date (Y0 as-of)",
        _anchor_default.isoformat() if _anchor_default else "—",
        name="s_anchor_date", registry=registry, style="input",
    ); row += 1
    kv_row(
        ws, row, "Initial Occupancy",
        _pct_value(default_inputs, "initial_occupancy_pct"),
        name="s_initial_occupancy", registry=registry, fmt=PCT, style="input",
    ); row += 1
    kv_row(
        ws, row, "Asset Mgmt Fee",
        _pct_value(default_inputs, "asset_mgmt_fee_pct"),
        name="s_asset_mgmt_fee", registry=registry, fmt=PCT, style="input",
    ); row += 1

    # Formula-conversion plan §3.1 — new Block A inputs that downstream
    # formulas need. All sourced from OperationalInputs (per-project today;
    # the "default project" snapshot represents scenario-level intent until
    # per-project assumption editing lands). Discount Rate is sourced from
    # the Scenario record directly since it's truly scenario-scoped.
    kv_row(
        ws, row, "Vacancy Rate (hold)",
        _pct_value(default_inputs, "hold_vacancy_rate_pct"),
        name="s_vacancy_pct", registry=registry, fmt=PCT, style="input",
    ); row += 1
    kv_row(
        ws, row, "CapEx Reserve / Unit / Yr",
        _safe_decimal(default_inputs, "capex_reserve_per_unit_annual"),
        name="s_capex_reserve_per_unit", registry=registry, fmt=ACCOUNTING,
        style="input",
    ); row += 1
    kv_row(
        ws, row, "Selling Costs",
        _pct_value(default_inputs, "selling_costs_pct"),
        name="s_selling_costs_pct", registry=registry, fmt=PCT, style="input",
    ); row += 1
    _discount_rate = ctx.get("discount_rate_pct")
    kv_row(
        ws, row, "Discount Rate / Hurdle",
        (_discount_rate / Decimal(100)) if isinstance(_discount_rate, Decimal)
        else None,
        name="s_discount_rate", registry=registry, fmt=PCT, style="input",
    ); row += 1

    # ── Block B: Per-project ───────────────────────────────────────────────
    block_b_row = row + 2
    section_label(
        ws, block_b_row, "B. Per-Project Assumptions",
        span_cols=1 + max(len(projects), 1),
    )
    header_row(ws, block_b_row + 1, ["Concept", *_project_column_labels(projects)])

    metrics = _per_project_metric_specs()
    for offset, (label, key, fmt, prefix) in enumerate(metrics):
        r = block_b_row + 2 + offset
        ws.cell(row=r, column=1, value=label).font = FONT_LABEL
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        # Numeric metric rows are user inputs (acquisition_price, unit
        # counts, occupancy %, cap rates, hold years, etc.) — render in
        # blue per the input/output color convention. Meta rows where
        # ``fmt`` is None (project_name, project_type) stay black.
        cell_font = FONT_INPUT if fmt is not None else FONT_VALUE
        for proj_idx, project in enumerate(projects, start=1):
            value = _per_project_value(
                key,
                project,
                inputs_by_project.get(project.id),
                use_lines_by_project.get(project.id, []),
                unit_mix_by_project.get(project.id, []),
            )
            registry.write(
                ws,
                r,
                1 + proj_idx,
                value,
                name=f"p{proj_idx}_{prefix}",
                fmt=fmt,
                font=cell_font,
                align=ALIGN_RIGHT,
            )
    next_row = block_b_row + 2 + len(metrics)

    # ── Block C: Capital Stack ─────────────────────────────────────────────
    # Expanded for formula-conversion (plan §3.1 / commit 1): every field
    # that drives carry-cost math is surfaced as a named input cell so the
    # Debt Schedule + Cash Flow formulas in later commits can reference
    # them by name. Eight new columns added after Rate: Term, Amort, IO,
    # Carry Type, Day Count, DSCR Min, LTV, Prepay. Original Auto-Sized?
    # and Covers retain their position at the right edge.
    block_c_row = next_row + 2
    section_label(ws, block_c_row, "C. Capital Stack", span_cols=len(CAPITAL_STACK_HEADERS))
    header_row(ws, block_c_row + 1, list(CAPITAL_STACK_HEADERS))

    junction_count_by_module: dict[UUID, int] = {}
    junction_principal_by_module: dict[UUID, Decimal] = {}
    for j in junctions:
        junction_count_by_module[j.capital_module_id] = (
            junction_count_by_module.get(j.capital_module_id, 0) + 1
        )
        junction_principal_by_module[j.capital_module_id] = (
            junction_principal_by_module.get(j.capital_module_id, Decimal(0))
            + _coerce_decimal(j.amount or 0)
        )

    for m_idx, module in enumerate(capital_modules, start=1):
        r = block_c_row + 1 + m_idx
        source = module.source or {}
        carry = module.carry or {}
        _junc_p = junction_principal_by_module.get(module.id)
        principal = _junc_p if _junc_p is not None else _coerce_decimal(source.get("amount") or 0)
        rate = source.get("interest_rate_pct") or carry.get("io_rate_pct") or 0
        auto_size = bool(source.get("auto_size"))
        is_shared = junction_count_by_module.get(module.id, 0) > 1
        slug = module_slugs.get(module.id) or f"module_{m_idx}"

        # Col 1: Label (display).
        ws.cell(row=r, column=1, value=module.label or "—").font = FONT_VALUE
        # Col 2: Funder Type (display).
        ws.cell(row=r, column=2, value=_funder_type_label(module)).font = FONT_VALUE
        # Cols 3-12: editable inputs (blue). Principal + Rate are existing;
        # Term through Prepay are added in commit 1 of formula-conversion.
        registry.write(
            ws, r, 3, _coerce_decimal(principal),
            name=f"s_{slug}_principal", fmt=ACCOUNTING,
            font=FONT_INPUT, align=ALIGN_RIGHT,
        )
        registry.write(
            ws, r, 4, _coerce_pct(rate),
            name=f"s_{slug}_rate", fmt=PCT,
            font=FONT_INPUT, align=ALIGN_RIGHT,
        )
        registry.write(
            ws, r, 5, _safe_int(source.get("hold_term_years")),
            name=f"s_{slug}_term_years", fmt=INT_COMMA,
            font=FONT_INPUT, align=ALIGN_RIGHT,
        )
        registry.write(
            ws, r, 6, _safe_int(carry.get("amort_term_years")),
            name=f"s_{slug}_amort_years", fmt=INT_COMMA,
            font=FONT_INPUT, align=ALIGN_RIGHT,
        )
        registry.write(
            ws, r, 7, _safe_int(carry.get("io_period_months")),
            name=f"s_{slug}_io_months", fmt=INT_COMMA,
            font=FONT_INPUT, align=ALIGN_RIGHT,
        )
        registry.write(
            ws, r, 8, str(carry.get("carry_type") or ""),
            name=f"s_{slug}_carry_type",
            font=FONT_INPUT, align=ALIGN_RIGHT,
        )
        registry.write(
            ws, r, 9, str(carry.get("day_count") or "30_360"),
            name=f"s_{slug}_day_count",
            font=FONT_INPUT, align=ALIGN_RIGHT,
        )
        # Block C numeric inputs may be None on debt modules that don't
        # populate every field (e.g. an IO bridge loan without dscr_min /
        # ltv / prepay terms). Guard against _coerce_pct(None) crashing —
        # write the value through unchanged when present, blank otherwise.
        _dscr_min = source.get("dscr_min")
        _ltv = source.get("ltv_pct")
        _prepay = source.get("prepay_penalty_pct")
        registry.write(
            ws, r, 10,
            _coerce_pct(_dscr_min) if _dscr_min is not None else None,
            name=f"s_{slug}_dscr_min", fmt="0.00",
            font=FONT_INPUT, align=ALIGN_RIGHT,
        )
        registry.write(
            ws, r, 11,
            _coerce_pct(_ltv) if _ltv is not None else None,
            name=f"s_{slug}_ltv_pct", fmt=PCT,
            font=FONT_INPUT, align=ALIGN_RIGHT,
        )
        registry.write(
            ws, r, 12,
            _coerce_pct(_prepay) if _prepay is not None else None,
            name=f"s_{slug}_prepay_pct", fmt=PCT,
            font=FONT_INPUT, align=ALIGN_RIGHT,
        )
        # Cols 13-14: display-only meta.
        ws.cell(row=r, column=13, value="Yes" if auto_size else "No").font = FONT_VALUE
        ws.cell(
            row=r, column=14,
            value=("shared (covers " + str(junction_count_by_module.get(module.id, 0)) + ")")
            if is_shared else "single project",
        ).font = FONT_VALUE

    if not capital_modules:
        ws.cell(
            row=block_c_row + 2, column=1,
            value="(no capital modules configured)",
        ).font = FONT_HINT

    # ── Block D: Waterfall Hurdles ─────────────────────────────────────────
    # Formula-conversion plan §3.1 — exposes waterfall tier inputs (priority,
    # type, IRR hurdle, LP/GP split) as named cells so the Waterfall sheet
    # formulas in commit 5 can reference them. One row per tier in the
    # scenario's WaterfallTier rows.
    block_d_row = block_c_row + 2 + max(len(capital_modules), 1) + 2
    waterfall_tiers: list = ctx.get("waterfall_tiers", [])
    section_label(ws, block_d_row, "D. Waterfall Hurdles", span_cols=6)
    header_row(
        ws,
        block_d_row + 1,
        ["Tier", "Type", "IRR Hurdle", "LP Split", "GP Split", "Notes"],
    )
    if waterfall_tiers:
        for t_idx, tier in enumerate(
            sorted(waterfall_tiers, key=lambda t: t.priority), start=1
        ):
            r = block_d_row + 1 + t_idx
            ws.cell(row=r, column=1, value=tier.priority).font = FONT_VALUE
            tier_type_label = getattr(tier.tier_type, "value", tier.tier_type) or ""
            ws.cell(row=r, column=2, value=str(tier_type_label)).font = FONT_VALUE
            registry.write(
                ws, r, 3, _coerce_pct(tier.irr_hurdle_pct),
                name=f"s_tier_{t_idx}_irr_hurdle", fmt=PCT,
                font=FONT_INPUT, align=ALIGN_RIGHT,
            )
            registry.write(
                ws, r, 4, _coerce_pct(tier.lp_split_pct),
                name=f"s_tier_{t_idx}_lp_split", fmt=PCT,
                font=FONT_INPUT, align=ALIGN_RIGHT,
            )
            registry.write(
                ws, r, 5, _coerce_pct(tier.gp_split_pct),
                name=f"s_tier_{t_idx}_gp_split", fmt=PCT,
                font=FONT_INPUT, align=ALIGN_RIGHT,
            )
            ws.cell(row=r, column=6, value=tier.description or "").font = FONT_HINT
    else:
        ws.cell(
            row=block_d_row + 2, column=1,
            value="(no waterfall tiers configured)",
        ).font = FONT_HINT

    # ── Block E: Phase Plan (single-project only) ──────────────────────────
    # Multi-project scenarios emit their per-project phase plan blocks on
    # each P{n} sheet inside _build_project_sheet. Single-project
    # scenarios suppress the P1 sheet (noise reduction), so emit the
    # phase plan here instead — keeps p1_perm_origination_month /
    # p1_total_horizon_months / p1_phase_*_{start,end,duration}_month
    # cells available so the Debt Schedule's Construction-to-Perm Status
    # block and the perm-gated Pro Forma debt service formula can
    # reference them.
    cur_row = block_d_row + 1 + max(len(waterfall_tiers), 1) + 2
    if len(projects) == 1:
        single_project = projects[0]
        single_inputs = inputs_by_project.get(single_project.id)
        next_row = _emit_phase_plan_block(
            ws, registry, ctx,
            project_idx=1, project=single_project,
            inputs=single_inputs, capital_modules=capital_modules,
            start_row=cur_row,
        )
        # Advance cursor only when the block actually emitted rows
        # (returns start_row unchanged when no phase windows exist).
        if next_row != cur_row:
            cur_row = next_row + 2  # spacer before next block

    # ── Block F: Revenue inputs (per stream) ──────────────────────────────
    next_row = _build_assumptions_revenue_block(
        ws, registry, ctx, start_row=cur_row,
    )
    if next_row != cur_row:
        cur_row = next_row + 2

    # ── Block G: Operating Expense inputs (per line) ──────────────────────
    next_row = _build_assumptions_opex_block(
        ws, registry, ctx, start_row=cur_row,
    )
    if next_row != cur_row:
        cur_row = next_row + 2

    freeze_top(ws, row=2)
    print_landscape(ws)


# Block C header row — defined at module scope so tests and the sheet
# builder share one source of truth.
CAPITAL_STACK_HEADERS: tuple[str, ...] = (
    "Module",          # col 1  — display
    "Funder Type",     # col 2  — display
    "Principal",       # col 3  — input
    "Rate",            # col 4  — input
    "Term (yrs)",      # col 5  — input
    "Amort (yrs)",     # col 6  — input
    "IO (mo)",         # col 7  — input
    "Carry Type",      # col 8  — input
    "Day Count",       # col 9  — input
    "DSCR Min",        # col 10 — input
    "LTV",             # col 11 — input
    "Prepay",          # col 12 — input
    "Auto-Sized?",     # col 13 — display
    "Covers",          # col 14 — display
)


def _safe_int(raw) -> int | None:
    """Coerce raw JSONB value to int, returning None on missing/invalid."""
    if raw is None:
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


def _glossary_term_tokens(name: str) -> list[str]:
    """Extract searchable tokens from a metric name.

    Handles two conventions used in FINANCIAL_MODEL.md:
    - ``ABBREV (Full Name)`` → ["ABBREV", "Full Name", full string]
    - ``Full Name (ABBREV)`` → ["Full Name", "ABBREV", full string]
    - ``Plain Name``         → [full string]
    """
    tokens = [name]
    m = re.search(r"\((.+)\)", name)
    if m:
        inner = m.group(1).strip()
        prefix = name[: name.index("(")].strip()
        tokens.extend([prefix, inner])
    return tokens


def _glossary_metric_used(name: str, written: set[str]) -> bool:
    """Return True if any token from this metric name appears in written cell strings."""
    written_lower = {s.lower() for s in written}
    for token in _glossary_term_tokens(name):
        if any(token.lower() in w for w in written_lower):
            return True
    return False


def _build_su_sheet(
    ws,
    registry: CellRegistry,
    ctx: dict,
) -> None:
    """Dedicated Sources & Uses sheet: per-project line detail + category summary + sources.

    Formula-conversion plan §4.6 (commit 2): every summable cell is a
    formula, not a Decimal. Per-project Use-line subtotals use
    ``=SUM(start:end)`` over the contiguous Use lines above them.
    Per-project totals sum the subtotals. The Category Summary block
    sums each category across projects. Sources reference the
    per-module Principal cells on the Assumptions sheet directly via
    defined names. ``s_su_gap = s_su_uses_total - s_su_sources_total``.

    Editing a Use-line amount in the per-project section ripples up to
    its subtotal, project total, category summary, grand total, and
    the Cover sheet's Total Uses cell — without re-running the engine.
    """
    set_widths(ws, [42, 22])

    use_lines_by_project: dict = ctx.get("use_lines", {})
    projects: list[Project] = ctx["projects"]
    capital_modules: list[CapitalModule] = ctx["capital_modules"]
    module_slugs: dict = ctx.get("module_slugs") or _compute_module_slugs(capital_modules)
    junctions: list[CapitalModuleProject] = ctx["junctions"]

    line = 1
    section_label(ws, line, "Sources & Uses", span_cols=2)
    line += 2

    # ── Per-project sections ───────────────────────────────────────────────
    # Track row addresses of per-project category subtotals so the
    # Category Summary block can build cross-section sum formulas.
    # Shape: {category: [row_1, row_2, ...]} one row per project that had
    # any lines in that category.
    cat_subtotal_rows: dict[str, list[int]] = {cat: [] for cat in USE_COST_CATEGORIES}
    # Track per-project total rows for any future per-project rollup formulas.
    _proj_total_rows: list[int] = []

    for idx, project in enumerate(projects, start=1):
        pid = project.id
        uls = use_lines_by_project.get(pid, [])

        section_label(ws, line, f"P{idx} — {project.name}", span_cols=2)
        line += 1

        # Track per-project subtotal cells so Total Uses P{n} can SUM them
        # (subtotals are not contiguous across categories, so use a list
        # of cell refs rather than a single SUM range).
        proj_subtotal_refs: list[str] = []

        for cat in USE_COST_CATEGORIES:
            cat_label = USE_CATEGORY_LABELS.get(cat, cat.title())
            cat_lines = [
                ul for ul in uls
                if str(ul.cost_category or "soft") == cat
                and str(getattr(ul.phase, "value", ul.phase) or "") != "exit"
            ]
            if not cat_lines:
                continue
            ws.cell(row=line, column=1, value=cat_label).font = FONT_LABEL
            line += 1

            first_line_row = line
            for ul in cat_lines:
                amt = _coerce_decimal(ul.amount or 0)
                ws.cell(row=line, column=1, value=f"  {ul.label or ''}").font = FONT_VALUE
                # Operating Reserve is a derived Use: months × Y1 OpEx ÷ 12.
                # When the workbook includes a Pro Forma sheet (every profile
                # that has S&U also has a Pro Forma) the Y1 OpEx cell is
                # registered as ``s_y1_opex`` so this formula resolves at
                # open-time. Assumptions sheet supplies ``s_operating_reserve_months``.
                label_norm = (ul.label or "").strip().lower()
                if "operating reserve" in label_norm or label_norm == "op reserve":
                    formula = "=s_operating_reserve_months*s_y1_opex/12"
                    cell = ws.cell(row=line, column=2, value=formula)
                    cell.number_format = ACCOUNTING
                else:
                    ws.cell(row=line, column=2, value=_to_excel_number(amt)).number_format = ACCOUNTING
                line += 1
            last_line_row = line - 1

            # Formula: SUM the contiguous Use-line cells just written.
            # Cells live in column B (col 2).
            subtotal_formula = f"=SUM(B{first_line_row}:B{last_line_row})"
            ws.cell(row=line, column=1, value=f"  Subtotal {cat_label}").font = FONT_LABEL
            cell = ws.cell(row=line, column=2, value=subtotal_formula)
            cell.number_format = ACCOUNTING
            cell.font = FONT_LABEL
            cat_subtotal_rows[cat].append(line)
            proj_subtotal_refs.append(f"B{line}")
            line += 1

        # Per-project Total Uses: formula summing this project's category
        # subtotals (non-contiguous so list-of-refs rather than SUM range).
        proj_total_formula = (
            "=" + "+".join(proj_subtotal_refs) if proj_subtotal_refs else "=0"
        )
        ws.cell(row=line, column=1, value=f"Total Uses P{idx}").font = FONT_LABEL
        cell = ws.cell(row=line, column=2, value=proj_total_formula)
        cell.number_format = ACCOUNTING
        cell.font = FONT_LABEL
        _proj_total_rows.append(line)
        line += 2

    # ── Category summary (all projects) ───────────────────────────────────
    section_label(ws, line, "Category Summary (All Projects)", span_cols=2)
    line += 1

    _su_cat_names = {
        "acquisition": "s_su_acq_total",
        "soft":        "s_su_soft_total",
        "hard":        "s_su_hard_total",
    }
    cat_summary_rows: dict[str, int] = {}
    for cat in USE_COST_CATEGORIES:
        cat_label = USE_CATEGORY_LABELS.get(cat, cat.title())
        # Formula sums the per-project subtotal cells for this category.
        subtotal_refs = [f"B{r}" for r in cat_subtotal_rows[cat]]
        formula = "=" + "+".join(subtotal_refs) if subtotal_refs else "=0"
        ws.cell(row=line, column=1, value=cat_label).font = FONT_VALUE
        cell = ws.cell(row=line, column=2, value=formula)
        cell.number_format = ACCOUNTING
        cell.font = FONT_VALUE
        cell.alignment = ALIGN_RIGHT
        registry.register(_su_cat_names[cat], ws.title, line, 2)
        cat_summary_rows[cat] = line
        line += 1

    # Total Uses: sum the category-total cells.
    cat_sum_refs = [f"B{cat_summary_rows[c]}" for c in USE_COST_CATEGORIES]
    uses_total_formula = "=" + "+".join(cat_sum_refs) if cat_sum_refs else "=0"
    ws.cell(row=line, column=1, value="Total Uses").font = FONT_LABEL
    cell = ws.cell(row=line, column=2, value=uses_total_formula)
    cell.number_format = ACCOUNTING
    cell.font = FONT_LABEL
    cell.alignment = ALIGN_RIGHT
    registry.register("s_su_uses_total", ws.title, line, 2)
    uses_total_row = line
    line += 2

    # ── Sources ───────────────────────────────────────────────────────────
    # Each non-equity source pulls its principal directly from the
    # Assumptions sheet's Block C ``s_<slug>_principal`` defined name,
    # so editing the principal there ripples into Sources here.
    section_label(ws, line, "Sources", span_cols=2)
    line += 1

    junction_amount: dict = {}
    for j in junctions:
        junction_amount[j.capital_module_id] = junction_amount.get(
            j.capital_module_id, Decimal(0)
        ) + _coerce_decimal(j.amount or 0)

    source_refs: list[str] = []
    for m_idx, module in enumerate(capital_modules, start=1):
        amount = junction_amount.get(module.id) or _coerce_decimal(
            (module.source or {}).get("amount") or 0
        )
        if amount <= Decimal(1) and _funder_class(module) == "Equity":
            continue
        ws.cell(row=line, column=1, value=module.label or _funder_type_label(module)).font = FONT_VALUE
        # Reference the Assumptions-sheet Principal cell by defined name.
        # Workbook-scoped defined names resolve without a sheet qualifier.
        slug = module_slugs.get(module.id) or f"module_{m_idx}"
        cell = ws.cell(row=line, column=2, value=f"=s_{slug}_principal")
        cell.number_format = ACCOUNTING
        source_refs.append(f"B{line}")
        line += 1

    # Implied equity: Uses − Sources. Computed as a formula so it tracks
    # edits to either side.
    if source_refs:
        implied_equity_formula = (
            f"=B{uses_total_row}-(" + "+".join(source_refs) + ")"
        )
    else:
        implied_equity_formula = f"=B{uses_total_row}"
    ws.cell(row=line, column=1, value="Owner Equity (implied gap)").font = FONT_VALUE
    cell = ws.cell(row=line, column=2, value=implied_equity_formula)
    cell.number_format = ACCOUNTING
    source_refs.append(f"B{line}")
    line += 1

    # Total Sources: sum of all source rows (including implied equity).
    sources_total_formula = "=" + "+".join(source_refs) if source_refs else "=0"
    ws.cell(row=line, column=1, value="Total Sources").font = FONT_LABEL
    cell = ws.cell(row=line, column=2, value=sources_total_formula)
    cell.number_format = ACCOUNTING
    cell.font = FONT_LABEL
    cell.alignment = ALIGN_RIGHT
    registry.register("s_su_sources_total", ws.title, line, 2)
    sources_total_row = line
    line += 1

    # Gap: Uses − Sources. Pure formula referencing the totals above.
    ws.cell(row=line, column=1, value="Δ Sources Gap (Uses − Sources)").font = FONT_LABEL
    cell = ws.cell(
        row=line, column=2,
        value=f"=B{uses_total_row}-B{sources_total_row}",
    )
    cell.number_format = ACCOUNTING
    cell.font = FONT_LABEL
    cell.alignment = ALIGN_RIGHT
    registry.register("s_su_gap", ws.title, line, 2)


def _build_glossary(
    ws,
    registry: CellRegistry,
    ctx: dict,
    written_strings: set[str] | None = None,
) -> None:
    """Glossary & Methodology sheet — driven by FINANCIAL_MODEL.md.

    When ``written_strings`` is provided (set of all string cell values
    already written to the workbook), filters to metrics whose name or
    abbreviation appears in the rendered output — so each export profile
    shows only the terms it actually uses.

    Filters parsed metrics down to ``audience='investor'`` per plan §3.8.
    The bidirectional doc/export validator verifies every named range
    in the workbook traces to a row here.
    """
    set_widths(ws, [28, 60, 50, 36])
    section_label(ws, 1, "Glossary & Methodology", span_cols=4)
    header_row(ws, 2, ["Term", "Definition", "Calculation", "Reference"])

    report = parse_doc()
    all_investor = sorted(
        report.for_audience("investor"),
        key=lambda m: m.name.lower(),
    )
    investor_metrics = (
        [m for m in all_investor if _glossary_metric_used(m.name, written_strings)]
        if written_strings
        else all_investor
    )

    for row_offset, metric in enumerate(investor_metrics):
        r = 3 + row_offset
        definition, calc = _split_definition_and_calc(metric)
        ws.cell(row=r, column=1, value=metric.name).font = FONT_LABEL
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        ws.cell(row=r, column=1).border = THIN_BORDER

        ws.cell(row=r, column=2, value=definition).font = FONT_VALUE
        ws.cell(row=r, column=2).alignment = ALIGN_WRAP
        ws.cell(row=r, column=2).border = THIN_BORDER

        ws.cell(row=r, column=3, value=calc).font = FONT_VALUE
        ws.cell(row=r, column=3).alignment = ALIGN_WRAP
        ws.cell(row=r, column=3).border = THIN_BORDER

        # GitHub-anchored hyperlink → opens the doc heading in a browser.
        # Friendly label first, URL behind the click — most LPs won't have
        # local repo access but anyone with a web browser can follow it.
        anchor = _github_anchor_for(metric)
        link_url = f"{_FINANCIAL_MODEL_URL}#{anchor}"
        link_label = f"FINANCIAL_MODEL.md § {metric.name}"
        # Escape any double quotes in the label to keep the formula valid.
        safe_label = link_label.replace('"', '""')
        ws.cell(
            row=r,
            column=4,
            value=f'=HYPERLINK("{link_url}","{safe_label}")',
        ).font = FONT_LINK
        ws.cell(row=r, column=4).alignment = ALIGN_LEFT
        ws.cell(row=r, column=4).border = THIN_BORDER

        ws.row_dimensions[r].height = 60

    # Footer caption documenting the contract
    foot = 3 + len(investor_metrics) + 1
    ws.cell(
        row=foot, column=1,
        value=(
            "Doc-driven glossary. Source of truth is docs/FINANCIAL_MODEL.md; "
            "the investor-export build runs a bidirectional validator that fails "
            "if any named range here lacks a doc entry or vice versa."
        ),
    ).font = FONT_HINT
    ws.merge_cells(start_row=foot, start_column=1, end_row=foot, end_column=4)

    freeze_top(ws, row=3)
    print_landscape(ws)


# ── Per-project assumptions metric specs ──────────────────────────────────────


def _per_project_metric_specs() -> list[tuple[str, str, str | None, str]]:
    """Rows for Block B of the Assumptions sheet.

    Each tuple: (label, lookup_key, number_format, named-range suffix).
    """
    return [
        ("Project Name", "project_name", None, "project_name"),
        ("Project Type", "project_type", None, "project_type"),
        ("Acquisition Price", "acquisition_price", ACCOUNTING, "acquisition_price"),
        ("Unit Count (existing)", "unit_count_existing", INT_COMMA, "unit_count_existing"),
        ("Unit Count (new)", "unit_count_new", INT_COMMA, "unit_count_new"),
        ("Avg In-Place Rent", "avg_in_place_rent", ACCOUNTING, "avg_in_place_rent"),
        ("Avg Market Rent", "avg_market_rent", ACCOUNTING, "avg_market_rent"),
        ("Stabilized Occupancy", "stabilized_occupancy_pct", PCT, "stabilized_occupancy"),
        ("Going-In Cap Rate", "going_in_cap_rate_pct", PCT, "going_in_cap_rate"),
        ("Exit Cap Rate", "exit_cap_rate_pct", PCT, "exit_cap_rate"),
        ("Construction Months", "construction_months", INT_COMMA, "construction_months"),
        ("Lease-Up Months", "lease_up_months", INT_COMMA, "lease_up_months"),
    ]


def _per_project_value(
    key: str,
    project: Project,
    inputs: OperationalInputs | None,
    use_lines: list[UseLine],
    unit_mix: list,
):
    raw = _per_project_value_raw(key, project, inputs, use_lines, unit_mix)
    # The DB stores percentages as whole numbers ("5.5" for 5.5%). The Block B
    # cell uses Excel's PCT format which expects fractions ("0.055"). Without
    # this divide-by-100 the per-project Exit Cap Rate column displays "5"
    # which Excel renders as "500.00%" — silently wrong. Block A's kv_row uses
    # _pct_value for this same conversion; this is the per-project equivalent.
    if isinstance(raw, Decimal) and key.endswith("_pct"):
        return raw / Decimal(100)
    return raw


def _per_project_value_raw(
    key: str,
    project: Project,
    inputs: OperationalInputs | None,
    use_lines: list[UseLine],
    unit_mix: list,
):
    if key == "project_name":
        return project.name or ""
    if key == "project_type":
        return getattr(project, "deal_type", "") or ""
    if key == "acquisition_price":
        # Heuristic: sum acquisition-phase Use lines. Commit 4 will switch
        # to Project.acquisition_price once the schema refactor lands.
        return sum(
            (_coerce_decimal(ul.amount) for ul in use_lines if _is_acquisition_phase(ul)),
            Decimal(0),
        ) or None
    if key == "avg_in_place_rent":
        return _weighted_avg_rent(unit_mix, "in_place_rent_per_unit")
    if key == "avg_market_rent":
        return _weighted_avg_rent(unit_mix, "market_rent_per_unit")
    if key == "unit_count_existing":
        return sum((um.unit_count or 0) for um in unit_mix) or None
    if inputs is None:
        return None
    return _safe_decimal(inputs, key)


def _weighted_avg_rent(unit_mix: list, field: str) -> Decimal | None:
    total_units = 0
    weighted = Decimal(0)
    for um in unit_mix:
        rent = getattr(um, field, None)
        units = um.unit_count or 0
        if rent is None or units <= 0:
            continue
        weighted += _coerce_decimal(rent) * Decimal(units)
        total_units += units
    if total_units <= 0:
        return None
    return weighted / Decimal(total_units)


def _is_acquisition_phase(ul: UseLine) -> bool:
    phase = str(getattr(ul.phase, "value", ul.phase) or "")
    return phase == "acquisition"


def _project_column_labels(projects: list[Project]) -> list[str]:
    labels: list[str] = []
    for idx, project in enumerate(projects, start=1):
        labels.append(f"P{idx} {(project.name or '').strip() or '—'}")
    # Pad out to MAX_PROJECTS_PER_SCENARIO so the header row width is stable.
    while len(labels) < MAX_PROJECTS_PER_SCENARIO:
        labels.append("")
    return labels


def _funder_type_label(module: CapitalModule) -> str:
    """Return a human-readable label for the module based on vehicle_type + equity_role."""
    vt = str(getattr(module, "vehicle_type", "") or "").replace("VehicleType.", "")
    er = str(getattr(module, "equity_role", "") or "").replace("EquityRole.", "")
    if vt == "equity":
        if er == "gp":
            return "GP Equity"
        if er == "lp":
            return "LP Equity"
        return "Equity"
    if vt == "debt":
        return "Debt"
    if vt == "grant":
        return "Grant"
    if vt == "forgivable_loan":
        return "Forgivable Loan"
    # Legacy fallback — label contains the loan type in Deal Setup
    lbl = (getattr(module, "label", "") or "").strip()
    return lbl or "Capital Source"


# ── Validator-driven glossary helpers ─────────────────────────────────────────


_BOLD_DEFINITION = re.compile(r"\*\*Definition\.\*\*\s*(.*?)(?=\n\s*\n|\Z)", re.DOTALL)
_BOLD_CALC = re.compile(
    r"\*\*Calculation\.\*\*\s*(?:```[\w]*\n(.*?)```|(.*?)(?=\n\s*\n|\Z))",
    re.DOTALL,
)


def _split_definition_and_calc(metric: MetricEntry) -> tuple[str, str]:
    """Pull the labelled paragraphs out of a metric body.

    Falls back to the first paragraph (definition) and "" (calc) when the
    body doesn't follow the structured shape. Lenient on purpose so the
    bidirectional validator can grow before every entry is fully shaped.
    """
    body = metric.body
    definition_match = _BOLD_DEFINITION.search(body)
    if definition_match:
        definition = _collapse_whitespace(definition_match.group(1))
    else:
        definition = _collapse_whitespace(body.split("\n\n", 1)[0])

    calc_match = _BOLD_CALC.search(body)
    calc = ""
    if calc_match:
        calc = (calc_match.group(1) or calc_match.group(2) or "").strip()
    return definition, calc


def _collapse_whitespace(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


# ── Coercion helpers (Decimal / pct / safe attr lookup) ───────────────────────


def _safe_decimal(obj, attr: str) -> Decimal | None:
    """Read ``obj.attr`` and coerce numerics to ``Decimal``; return None for missing."""
    if obj is None:
        return None
    value = getattr(obj, attr, None)
    if value is None:
        return None
    return _coerce_decimal(value)


def _pct_value(obj, attr: str) -> Decimal | None:
    """Read a percent-stored-as-whole-number field and convert to fraction.

    The DB stores percentages as e.g. ``5.5`` for 5.5%. Excel's PCT format
    expects fractions (0.055), so we divide by 100 here.
    """
    raw = _safe_decimal(obj, attr)
    if raw is None:
        return None
    return raw / Decimal(100)


def _revenue_growth_default(streams: list[IncomeStream]) -> Decimal:
    """Unit-count-weighted mean of stream escalation rates (as a fraction).

    Phase A: the Pro Forma's gross_revenue growth chain uses a single
    scenario-wide knob (s_revenue_growth_rate) so an LP editing one cell
    re-flows every Y2+ year. The seed value comes from the underlying
    stream-level escalation_rate_pct_annual, weighted by unit_count so a
    100-unit residential stream's growth dominates a 1-unit laundry line.
    Falls back to 3% (industry-standard rent-growth default) when no
    streams or all weights are zero.
    """
    total_weight = Decimal(0)
    weighted = Decimal(0)
    for s in streams:
        units = s.unit_count or 0
        try:
            weight = Decimal(int(units))
        except (TypeError, ValueError):
            weight = Decimal(0)
        if weight <= 0:
            weight = Decimal(1)
        rate = _coerce_decimal(getattr(s, "escalation_rate_pct_annual", None))
        if rate is None:
            continue
        weighted += weight * rate
        total_weight += weight
    if total_weight <= 0:
        return Decimal("0.03")
    return (weighted / total_weight) / Decimal(100)


def _coerce_decimal(value) -> Decimal | None:
    """Coerce to ``Decimal``. ``None`` / empty-string in, ``None`` out.

    Permissive on None because Block C / Block D / Source Returns all
    call this on optional ORM fields (``tier.irr_hurdle_pct``,
    ``source.dscr_min``, etc.). Crashing on missing data is the wrong
    behavior for a display-only exporter — surface ``None`` instead and
    let the caller's ``_write_optional`` / em-dash path render the
    missing cell as ``—``.
    """
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _coerce_pct(value) -> Decimal | None:
    d = _coerce_decimal(value)
    if d is None:
        return None
    return d / Decimal(100)


def _slugify_simple(label: str | None) -> str:
    """Lowercase, non-alphanumeric → ``_``, collapse runs, strip ends,
    cap at 40 chars. Returns empty string for None/empty input — the
    caller is responsible for substituting a fallback (e.g. record
    index) when the label is missing.
    """
    if not label:
        return ""
    s = "".join(c.lower() if c.isalnum() else "_" for c in label)
    while "__" in s:
        s = s.replace("__", "_")
    return s.strip("_")[:40]


def _stream_slugs(streams: list[IncomeStream]) -> dict[uuid.UUID, str]:
    """Build ``{stream.id: slug}`` for Assumptions Block F (Revenue).

    Collision-resolved with ``_2``, ``_3``, … suffixes so two streams
    with the same label still get unique named cells. Empty labels fall
    back to ``stream_<idx>`` where ``idx`` is the 1-based enumeration
    position — stable across re-renders of the same scenario but not
    across reordering, which matches the IncomeStream record's lack of
    a stable display-order column.
    """
    out: dict[uuid.UUID, str] = {}
    used: set[str] = set()
    for i, s in enumerate(streams, start=1):
        base = _slugify_simple(s.label) or f"stream_{i}"
        slug = base
        n = 2
        while slug in used:
            slug = f"{base}_{n}"
            n += 1
        used.add(slug)
        out[s.id] = slug
    return out


def _opex_slugs(lines: list[OperatingExpenseLine]) -> dict[uuid.UUID, str]:
    """Build ``{expense_line.id: slug}`` for Assumptions Block G (OpEx).

    Same collision-resolution + fallback pattern as :func:`_stream_slugs`.
    """
    out: dict[uuid.UUID, str] = {}
    used: set[str] = set()
    for i, line in enumerate(lines, start=1):
        base = _slugify_simple(line.label) or f"opex_{i}"
        slug = base
        n = 2
        while slug in used:
            slug = f"{base}_{n}"
            n += 1
        used.add(slug)
        out[line.id] = slug
    return out


def _all_revenue_slugs(ctx: dict) -> dict[uuid.UUID, str]:
    """Return ``{stream.id: slug}`` for every IncomeStream across all
    projects in the scenario. Computed via a SINGLE ``_stream_slugs``
    call over the flattened list so two streams with identical labels
    on different projects collision-resolve globally (gets ``_2`` /
    ``_3`` suffix) instead of silently shadowing one another.

    Single source of truth: Assumptions Block F and the Pro Forma's
    Gross Revenue formula both consume this map so the formula's
    ``s_rev_<slug>_y1_monthly`` references always resolve.
    """
    projects: list[Project] = ctx.get("projects") or []
    streams_by_project: dict = ctx.get("income_streams") or {}
    flat: list[IncomeStream] = []
    for project in projects:
        flat.extend(streams_by_project.get(project.id, []))
    return _stream_slugs(flat)


def _all_opex_slugs(ctx: dict) -> dict[uuid.UUID, str]:
    """Return ``{expense_line.id: slug}`` for every OperatingExpenseLine
    across all projects in the scenario. Same global collision-resolution
    contract as :func:`_all_revenue_slugs`.
    """
    projects: list[Project] = ctx.get("projects") or []
    lines_by_project: dict = ctx.get("expense_lines") or {}
    flat: list[OperatingExpenseLine] = []
    for project in projects:
        flat.extend(lines_by_project.get(project.id, []))
    return _opex_slugs(flat)


def _all_opex_lines_ordered(ctx: dict) -> list[OperatingExpenseLine]:
    """Return every OperatingExpenseLine across all projects in the
    same flatten order :func:`_all_opex_slugs` resolves slugs against.
    The breakout-row renderer pairs each line with its slug by walking
    both lists in lockstep, so the order MUST match the slug map's
    insertion order to avoid label/cell-reference mis-pairing.
    """
    projects: list[Project] = ctx.get("projects") or []
    lines_by_project: dict = ctx.get("expense_lines") or {}
    flat: list[OperatingExpenseLine] = []
    for project in projects:
        flat.extend(lines_by_project.get(project.id, []))
    return flat


def _all_revenue_streams_ordered(ctx: dict) -> list[IncomeStream]:
    """Return every IncomeStream across all projects in the same
    flatten order :func:`_all_revenue_slugs` resolves slugs against.
    Mirrors :func:`_all_opex_lines_ordered` — the per-stream breakout
    renderer pairs each stream with its slug by walking both lists in
    lockstep.
    """
    projects: list[Project] = ctx.get("projects") or []
    streams_by_project: dict = ctx.get("income_streams") or {}
    flat: list[IncomeStream] = []
    for project in projects:
        flat.extend(streams_by_project.get(project.id, []))
    return flat


def _build_assumptions_revenue_block(
    ws,
    registry: CellRegistry,
    ctx: dict,
    *,
    start_row: int,
) -> int:
    """Block F — Revenue inputs (per-project per-stream).

    Renders one row per ``IncomeStream`` across all projects. Each row
    registers four named cells the downstream Pro Forma / Cash Flow /
    Unit Mix sheets can reference instead of hardcoding the same
    numbers in multiple places:

      ``s_rev_<slug>_unit_count``           — int
      ``s_rev_<slug>_rent_per_unit_monthly`` — Decimal
      ``s_rev_<slug>_occupancy_pct``         — Decimal (0–100)
      ``s_rev_<slug>_escalation_pct``        — Decimal (0–100)

    A fifth column writes the computed Y1 stabilized monthly revenue
    (``unit_count × rent × occupancy/100``) as a formula referencing
    the three input cells — gives an LP an at-a-glance "what's this
    line worth?" check, and feeds future Pro Forma Gross Revenue
    formulas via ``s_rev_<slug>_y1_monthly``.

    Returns the next free row.
    """
    projects: list[Project] = ctx.get("projects") or []
    streams_by_project: dict = ctx.get("income_streams") or {}
    all_streams: list[tuple[int, IncomeStream]] = []
    for idx, project in enumerate(projects, start=1):
        for stream in streams_by_project.get(project.id, []):
            all_streams.append((idx, stream))
    if not all_streams:
        return start_row

    section_label(ws, start_row, "F. Revenue Inputs (per stream)", span_cols=7)
    header_row(
        ws, start_row + 1,
        ["Project", "Stream", "Unit Count", "Rent / Unit / Mo",
         "Occupancy %", "Escalation %", "Y1 Monthly (calc)"],
    )

    # Cross-project slug map (single source of truth — same map the
    # Pro Forma's Gross Revenue formula consumes).
    slug_by_id = _all_revenue_slugs(ctx)

    row = start_row + 2
    for project_idx, stream in all_streams:
        slug = slug_by_id[stream.id]
        ws.cell(row=row, column=1, value=f"P{project_idx}").font = FONT_VALUE
        ws.cell(row=row, column=2, value=stream.label or "(unnamed)").font = FONT_VALUE

        unit_count = stream.unit_count if stream.unit_count is not None else None
        rent = _coerce_decimal(stream.amount_per_unit_monthly) or Decimal(0)
        occ = _coerce_pct(stream.stabilized_occupancy_pct) or Decimal(0)
        esc = _coerce_pct(stream.escalation_rate_pct_annual) or Decimal(0)

        registry.write(
            ws, row, 3, unit_count,
            name=f"s_rev_{slug}_unit_count", fmt=INT_COMMA,
            font=FONT_INPUT, align=ALIGN_RIGHT,
        )
        registry.write(
            ws, row, 4, rent,
            name=f"s_rev_{slug}_rent_per_unit_monthly", fmt=ACCOUNTING,
            font=FONT_INPUT, align=ALIGN_RIGHT,
        )
        registry.write(
            ws, row, 5, occ,
            name=f"s_rev_{slug}_occupancy_pct", fmt=PCT,
            font=FONT_INPUT, align=ALIGN_RIGHT,
        )
        registry.write(
            ws, row, 6, esc,
            name=f"s_rev_{slug}_escalation_pct", fmt=PCT,
            font=FONT_INPUT, align=ALIGN_RIGHT,
        )
        # Y1 monthly formula: count × rent × occupancy. Wrapped in
        # IFERROR so a missing unit_count cell (fixed-amount streams)
        # falls back to the rent cell alone.
        y1_formula = (
            f"=IFERROR(s_rev_{slug}_unit_count*s_rev_{slug}_rent_per_unit_monthly*"
            f"s_rev_{slug}_occupancy_pct,s_rev_{slug}_rent_per_unit_monthly*"
            f"s_rev_{slug}_occupancy_pct)"
        )
        registry.write(
            ws, row, 7, y1_formula,
            name=f"s_rev_{slug}_y1_monthly", fmt=ACCOUNTING,
            font=FONT_VALUE, align=ALIGN_RIGHT,
        )
        row += 1
    return row


def _build_assumptions_opex_block(
    ws,
    registry: CellRegistry,
    ctx: dict,
    *,
    start_row: int,
) -> int:
    """Block G — Operating Expense inputs (per-project per-line).

    Renders one row per ``OperatingExpenseLine`` across all projects.
    Registers three named cells per row so future Pro Forma OpEx rows
    can reference them instead of repeating the same hardcoded
    numbers:

      ``s_opex_<slug>_annual``       — Decimal (Y1 $)
      ``s_opex_<slug>_escalation_pct`` — Decimal (0–100)
      ``s_opex_<slug>_monthly`` (formula) — convenience for Y1 monthly

    Returns the next free row.
    """
    projects: list[Project] = ctx.get("projects") or []
    lines_by_project: dict = ctx.get("expense_lines") or {}
    all_lines: list[tuple[int, OperatingExpenseLine]] = []
    for idx, project in enumerate(projects, start=1):
        for line in lines_by_project.get(project.id, []):
            all_lines.append((idx, line))
    if not all_lines:
        return start_row

    section_label(ws, start_row, "G. Operating Expense Inputs (per line)", span_cols=5)
    header_row(
        ws, start_row + 1,
        ["Project", "Line Item", "Annual ($)", "Escalation %", "Y1 Monthly (calc)"],
    )

    slug_by_id = _all_opex_slugs(ctx)

    row = start_row + 2
    for project_idx, line in all_lines:
        slug = slug_by_id[line.id]
        ws.cell(row=row, column=1, value=f"P{project_idx}").font = FONT_VALUE
        ws.cell(row=row, column=2, value=line.label or "(unnamed)").font = FONT_VALUE

        annual = _coerce_decimal(line.annual_amount) or Decimal(0)
        esc = _coerce_pct(line.escalation_rate_pct_annual) or Decimal(0)

        registry.write(
            ws, row, 3, annual,
            name=f"s_opex_{slug}_annual", fmt=ACCOUNTING,
            font=FONT_INPUT, align=ALIGN_RIGHT,
        )
        registry.write(
            ws, row, 4, esc,
            name=f"s_opex_{slug}_escalation_pct", fmt=PCT,
            font=FONT_INPUT, align=ALIGN_RIGHT,
        )
        monthly_formula = f"=s_opex_{slug}_annual/12"
        registry.write(
            ws, row, 5, monthly_formula,
            name=f"s_opex_{slug}_monthly", fmt=ACCOUNTING,
            font=FONT_VALUE, align=ALIGN_RIGHT,
        )
        row += 1
    return row


def _slugify_module_label(label: str | None, fallback_idx: int) -> str:
    """Convert a CapitalModule.label into an Excel-name-safe slug fragment.

    Lowercase, non-alphanumeric → ``_``, collapse runs, strip ends, cap at
    40 chars. Empty/None label falls back to ``module_<idx>`` so blank
    labels still produce stable, unique-per-position names.
    """
    if label:
        s = "".join(c.lower() if c.isalnum() else "_" for c in label)
        while "__" in s:
            s = s.replace("__", "_")
        s = s.strip("_")
        if s:
            return s[:40]
    return f"module_{fallback_idx}"


def _compute_module_slugs(capital_modules) -> dict:
    """Build {module.id: slug} with collision-resolved, human-readable slugs.

    On collision (two modules with the same slugified label) the second
    gets ``_2``, third ``_3``, etc. ``stack_position`` is preferred as
    the fallback index because it's stable across re-orderings; falls
    back to enumerate index when ``stack_position`` is missing.
    """
    slugs: dict = {}
    used: set[str] = set()
    for i, m in enumerate(capital_modules, start=1):
        base = _slugify_module_label(
            getattr(m, "label", None),
            getattr(m, "stack_position", None) or i,
        )
        slug = base
        n = 2
        while slug in used:
            slug = f"{base}_{n}"
            n += 1
        used.add(slug)
        slugs[m.id] = slug
    return slugs


# ── Pro Forma profile sheet builders ──────────────────────────────────────────
#
# The proforma export profile renders these sheets instead of the full
# Underwriting Pro Forma / per-project sheets.  Phase 1 = hardcoded computed
# values (same data as the UW sheets, different layout for broker/external use).
# Phase 2 = formula-driven input cells so the recipient can adjust escalation
# rates, vacancy, etc. without going back to the app.

_PF_ROWS: list[tuple[str, str]] = [
    ("Gross Revenue", "gross_revenue"),
    ("Vacancy Loss", "vacancy_loss"),
    ("Effective Gross Income (EGI)", "effective_gross_income"),
    ("Operating Expenses (OpEx)", "operating_expenses"),
    ("CapEx Reserve", "capex_reserve"),
    ("NOI (Net Operating Income)", "noi"),
    # Phase C parity with _build_uw_proforma.
    ("Asset Mgmt Fee", "asset_mgmt_fee"),
    ("Debt Service", "debt_service"),
    ("Net Cash Flow", "net_cash_flow"),
]


def _write_pf_table(
    ws,
    registry: CellRegistry,
    start_row: int,
    annual: dict,
    year_cols: list[int],
    cash_flow_items_for_scope,  # dict[UUID, list[CashFlowLineItem]]
) -> int:
    """Write the standard NOI build table + breakouts; return next free row."""
    from openpyxl.utils import get_column_letter

    # Formula-conversion parity with commit 3's _build_uw_proforma: rows
    # whose math is a direct sum/difference of other rows on this sheet
    # become Excel formulas so the proforma-profile workbook responds to
    # LP edits the same way the internal/lp/lender Underwriting Pro Forma
    # does. EGI = GrossRev + Vacancy (vacancy is signed negative); NOI =
    # EGI - OpEx - CapEx Reserve. Net Cash Flow stays engine-driven until
    # Debt Service formulas land.
    _DERIVED_FORMULA_FIELDS: dict[str, tuple[str, ...]] = {
        "effective_gross_income": ("+gross_revenue", "+vacancy_loss"),
        "noi": ("+effective_gross_income", "-operating_expenses", "-capex_reserve"),
        # Parity with _build_uw_proforma Phase E: NCF derived from NOI and
        # Debt Service so any future Debt Service formula flows through.
        "net_cash_flow": ("+noi", "-debt_service"),
    }
    # Phase A/B: gross revenue grows off Y1 via ``s_revenue_growth_rate``;
    # OpEx + CapEx Reserve via ``s_opex_growth_rate``.
    _GROWTH_CHAIN_FIELDS: dict[str, str] = {
        "gross_revenue": "s_revenue_growth_rate",
        "operating_expenses": "s_opex_growth_rate",
        "capex_reserve": "s_opex_growth_rate",
    }
    field_row: dict[str, int] = {}

    cur_row = start_row
    for label, field in _PF_ROWS:
        ws.cell(row=cur_row, column=1, value=label).font = FONT_LABEL
        field_row[field] = cur_row
        derived = _DERIVED_FORMULA_FIELDS.get(field)
        growth_name = _GROWTH_CHAIN_FIELDS.get(field)
        for col_offset, year in enumerate(year_cols):
            col_idx = 2 + col_offset
            if derived:
                operands: list[str] = []
                for spec in derived:
                    sign, operand_field = spec[0], spec[1:]
                    operand_row = field_row.get(operand_field)
                    if operand_row is None:
                        operands = []
                        break
                    col_letter = get_column_letter(col_idx)
                    operands.append(f"{sign}{col_letter}{operand_row}")
                if operands:
                    formula = "=" + "".join(operands).lstrip("+")
                    cell = ws.cell(row=cur_row, column=col_idx, value=formula)
                else:
                    value = annual.get(year, {}).get(field, Decimal(0))
                    cell = ws.cell(
                        row=cur_row, column=col_idx,
                        value=_to_excel_number(value),
                    )
            elif field == "asset_mgmt_fee":
                # Phase C parity with _build_uw_proforma.
                egi_r = field_row.get("effective_gross_income")
                if egi_r is not None:
                    col_letter = get_column_letter(col_idx)
                    formula = (
                        f"=IFERROR(-{col_letter}{egi_r}*s_asset_mgmt_fee,0)"
                    )
                    cell = ws.cell(row=cur_row, column=col_idx, value=formula)
                else:
                    cell = ws.cell(
                        row=cur_row, column=col_idx,
                        value=_to_excel_number(Decimal(0)),
                    )
            elif growth_name and col_offset >= 2:
                prev_col = get_column_letter(col_idx - 1)
                formula = f"={prev_col}{cur_row}*(1+{growth_name})"
                cell = ws.cell(row=cur_row, column=col_idx, value=formula)
            else:
                value = annual.get(year, {}).get(field, Decimal(0))
                cell = ws.cell(
                    row=cur_row, column=col_idx,
                    value=_to_excel_number(value),
                )
            cell.number_format = ACCOUNTING
            cell.font = FONT_VALUE
            cell.alignment = ALIGN_RIGHT
        # Phase B follow-up: expose Y1 OpEx as a workbook-scoped name. The
        # proforma profile calls _write_pf_table twice (combined Pro Forma
        # sheet + per-project Pro Forma sheets); only the first call should
        # register the name to avoid collision.
        if (
            field == "operating_expenses"
            and len(year_cols) >= 2
            and "s_y1_opex" not in registry._names
        ):
            registry.register("s_y1_opex", ws.title, cur_row, 3)
        cur_row += 1

    # OER derived row — formula references the OpEx and EGI cells written
    # above so it updates in lock-step with the growth-chained OpEx values.
    # IFERROR guards against EGI = 0 (pre-stabilization years).
    ws.cell(row=cur_row, column=1, value="OER (OpEx ÷ EGI)").font = FONT_LABEL
    opex_r = field_row.get("operating_expenses")
    egi_r = field_row.get("effective_gross_income")
    for col_offset, _year in enumerate(year_cols):
        col_idx = 2 + col_offset
        if opex_r is not None and egi_r is not None:
            col_letter = get_column_letter(col_idx)
            formula = f"=IFERROR({col_letter}{opex_r}/{col_letter}{egi_r},\"\")"
            cell = ws.cell(row=cur_row, column=col_idx, value=formula)
        else:
            cell = ws.cell(row=cur_row, column=col_idx, value=_DASH)
        cell.number_format = PCT
        cell.font = FONT_VALUE
        cell.alignment = ALIGN_RIGHT
    cur_row += 1

    # Revenue + OpEx breakouts
    by_category = _aggregate_scenario_line_items_by_category(cash_flow_items_for_scope)

    cur_row += 1
    cur_row = _write_breakout_table(
        ws, registry, cur_row,
        title="Revenue Breakout (by stream)",
        rows=by_category.get("income", {}),
        year_cols=year_cols,
        empty_hint="(no revenue line items — run Compute to populate)",
    )
    cur_row += 1
    cur_row = _write_breakout_table(
        ws, registry, cur_row,
        title="OpEx Breakout (by category)",
        rows=by_category.get("expense", {}),
        year_cols=year_cols,
        empty_hint="(no OpEx line items — run Compute to populate)",
        always_show=ALWAYS_SHOWN_OPEX_CATEGORIES,
    )
    return cur_row


def _build_proforma_combined(
    ws,
    registry: CellRegistry,
    ctx: dict,
) -> None:
    """Combined pro forma — all projects aggregated.

    Phase 1: hardcoded computed values matching the Underwriting Pro Forma
    sheet data.  Phase 2 will add adjustable input cells.
    """
    cash_flows: dict[UUID, list[CashFlow]] = ctx["cash_flows"]
    cash_flow_items: dict[UUID, list[CashFlowLineItem]] = ctx["cash_flow_items"]
    projects: list[Project] = ctx["projects"]

    annual = _aggregate_scenario_annual(cash_flows)
    max_year = min(max(annual) if annual else 0, 10)
    year_cols = list(range(0, max(max_year, 1) + 1))

    set_widths(ws, [32, *([14] * (len(year_cols) + 1))])

    section_label(
        ws, 1,
        f"Pro Forma — Combined ({len(projects)} project{'s' if len(projects) != 1 else ''})",
        span_cols=len(year_cols) + 2,
    )
    header_row(ws, 2, ["Line Item", *[f"Y{y}" for y in year_cols]])

    _write_pf_table(ws, registry, 3, annual, year_cols, cash_flow_items)

    freeze_top(ws, row=3)
    print_landscape(ws)


def _build_proforma_project_sheet(
    ws,
    registry: CellRegistry,
    ctx: dict,
    idx: int,
    project: Project,
) -> None:
    """Per-project pro forma.

    Phase 1: hardcoded computed values matching the existing per-project sheet
    data.  Phase 2 will add adjustable input cells.
    """
    cash_flows_by_project: dict[UUID, list[CashFlow]] = ctx["cash_flows"]
    cash_flow_items_by_project: dict[UUID, list[CashFlowLineItem]] = ctx["cash_flow_items"]
    outputs_by_project: dict[UUID, "OperationalOutputs"] = ctx["outputs"]

    project_cash_flows = cash_flows_by_project.get(project.id, [])
    project_items = cash_flow_items_by_project.get(project.id, [])
    outputs = outputs_by_project.get(project.id)

    annual = _aggregate_annual(project_cash_flows)
    max_year = min(max(annual) if annual else 0, 10)
    year_cols = list(range(0, max(max_year, 1) + 1))

    set_widths(ws, [32, *([14] * (len(year_cols) + 1))])

    # Header
    header_text = f"P{idx} — {project.name or 'Project'}"
    ws.cell(row=1, column=1, value=header_text).font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(year_cols) + 2)
    ws.row_dimensions[1].height = 24

    # Key metrics strip
    cur = 2
    kpi_pairs: list[tuple[str, object, str]] = [
        ("Stabilized NOI", _safe_decimal(outputs, "noi_stabilized"), ACCOUNTING),
        ("DSCR", _safe_decimal(outputs, "dscr"), "0.000"),
        ("Cap Rate on Cost", _pct_value(outputs, "cap_rate_on_cost_pct"), PCT),
        ("Levered IRR", _pct_value(outputs, "project_irr_levered"), PCT),
    ]
    for label, val, fmt in kpi_pairs:
        ws.cell(row=cur, column=1, value=label).font = FONT_LABEL
        cell = ws.cell(row=cur, column=2, value=_to_excel_number(val))
        cell.number_format = fmt
        cell.font = FONT_VALUE
        cur += 1

    cur += 1  # blank separator
    header_row(ws, cur, ["Line Item", *[f"Y{y}" for y in year_cols]])
    cur += 1

    # Wrap single project in a dict so _aggregate_scenario_line_items_by_category works.
    _write_pf_table(ws, registry, cur, annual, year_cols, {project.id: project_items})

    freeze_top(ws, row=cur)
    print_landscape(ws)


# Re-exports for callers that want to inline format strings without importing
# from the helpers module.
__all__ = [
    "DATE_FMT",
    "MAX_PROJECTS_PER_SCENARIO",
    "PROJECT_SHEET_NAME_BUDGET",
    "BRAND",
    "export_investor_workbook",
    "make_investor_filename",
]
