"""Excel workbook export for deal models — designed for round-trip re-import.

Sheet layout:
  Summary      — computed outputs, read-only (protected)
  Cash Flow    — monthly period table, read-only (protected)
  Uses         — UseLine rows, editable (col A = hidden key)
  Income       — IncomeStream rows, editable
  OpEx         — OperatingExpenseLine rows, editable
  Sources      — CapitalModule rows flattened, editable
  _setup       — key-value metadata for round-trip import, protected (locked)

Round-trip design:
  - Column A on editable sheets = row key: "use_line|<uuid>" or "use_line|new"
  - Column A is styled light grey to signal it's a system field; width is narrow
  - _setup sheet: locked with no password — user can unprotect but has friction
  - On import, column A is the authority for row identity; position is irrelevant
  - _setup stores structural metadata (project_type, debt_structure, milestones, etc.)
    so a file exported from this tool can be re-imported with near-zero wizard input
"""
from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal
from enum import Enum
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Protection,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.capital import CapitalModule, WaterfallTier
from app.models.cashflow import CashFlow, OperationalOutputs
from app.models.deal import (
    DealModel,
    IncomeStream,
    OperatingExpenseLine,
    OperationalInputs,
    UseLine,
)
from app.models.milestone import Milestone
from app.models.project import Project

# ── Palette ───────────────────────────────────────────────────────────────────
_CLR_HEADER_BG   = "1E3A5F"   # dark navy
_CLR_HEADER_FG   = "FFFFFF"
_CLR_KEY_BG      = "F0F0F0"   # light grey — key column
_CLR_KEY_FG      = "999999"
_CLR_EDITABLE    = "FFFFFF"   # white — editable cells
_CLR_LOCKED_BG   = "F8F8F8"   # near-white — locked cells on editable sheets
_CLR_SETUP_BG    = "FFF8E1"   # warm yellow — _setup sheet
_CLR_SETUP_LOCK  = "FFE082"   # amber — locked cells on _setup
_CLR_HINT        = "7F7F7F"

_THIN = Side(style="thin", color="D0D0D0")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

SCHEMA_VERSION = "1.1"


# ── Public entry point ────────────────────────────────────────────────────────

async def export_deal_model_workbook(deal_model_id: UUID, session: AsyncSession) -> bytes:
    """Return a round-trip-capable XLSX workbook for the requested deal model."""
    data = await _load_all(session, deal_model_id)
    if data is None:
        raise ValueError(f"DealModel {deal_model_id} was not found")

    model, project, inputs, use_lines, income_streams, expense_lines, \
        capital_modules, waterfall_tiers, cash_flows, outputs, milestones = data

    wb = Workbook()

    # Read-only sheets first
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _build_summary(ws_summary, model, outputs)

    ws_cf = wb.create_sheet("Cash Flow")
    _build_cash_flow(ws_cf, cash_flows)

    # Editable sheets
    ws_uses = wb.create_sheet("Uses")
    _build_uses(ws_uses, use_lines)

    ws_income = wb.create_sheet("Income")
    _build_income(ws_income, income_streams)

    ws_opex = wb.create_sheet("OpEx")
    _build_opex(ws_opex, expense_lines)

    ws_sources = wb.create_sheet("Sources")
    _build_sources(ws_sources, capital_modules)

    # Hidden metadata sheet (always last)
    ws_setup = wb.create_sheet("_setup")
    _build_setup(ws_setup, model, project, inputs, milestones)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_export_filename(model: DealModel) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", (model.name or "deal").lower()).strip("-") or "deal"
    return f"{slug}.xlsx"


# ── Data loader ───────────────────────────────────────────────────────────────

async def _load_all(session: AsyncSession, model_id: UUID):
    model = (await session.execute(
        select(DealModel).where(DealModel.id == model_id)
    )).scalar_one_or_none()
    if model is None:
        return None

    project = (await session.execute(
        select(Project).where(Project.scenario_id == model_id).order_by(Project.created_at).limit(1)
    )).scalar_one_or_none()

    proj_id = project.id if project else None

    inputs = (await session.execute(
        select(OperationalInputs).where(OperationalInputs.project_id == proj_id)
    )).scalar_one_or_none() if proj_id else None

    use_lines = list((await session.execute(
        select(UseLine).where(UseLine.project_id == proj_id).order_by(UseLine.phase, UseLine.label)
    )).scalars()) if proj_id else []

    income_streams = list((await session.execute(
        select(IncomeStream).where(IncomeStream.project_id == proj_id).order_by(IncomeStream.label)
    )).scalars()) if proj_id else []

    expense_lines = list((await session.execute(
        select(OperatingExpenseLine).where(OperatingExpenseLine.project_id == proj_id).order_by(OperatingExpenseLine.label)
    )).scalars()) if proj_id else []

    capital_modules = list((await session.execute(
        select(CapitalModule).where(CapitalModule.scenario_id == model_id).order_by(CapitalModule.stack_position)
    )).scalars())

    waterfall_tiers = list((await session.execute(
        select(WaterfallTier).where(WaterfallTier.scenario_id == model_id).order_by(WaterfallTier.priority)
    )).scalars())

    cash_flows = list((await session.execute(
        select(CashFlow).where(CashFlow.scenario_id == model_id).order_by(CashFlow.period)
    )).scalars())

    outputs = (await session.execute(
        select(OperationalOutputs).where(OperationalOutputs.scenario_id == model_id)
    )).scalar_one_or_none()

    milestones = list((await session.execute(
        select(Milestone).where(Milestone.project_id == proj_id)
    )).scalars()) if proj_id else []

    return (model, project, inputs, use_lines, income_streams, expense_lines,
            capital_modules, waterfall_tiers, cash_flows, outputs, milestones)


# ── Summary (read-only) ───────────────────────────────────────────────────────

def _build_summary(ws: Worksheet, model: DealModel, outputs: OperationalOutputs | None) -> None:
    _header_row(ws, ["Metric", "Value"])
    rows = [
        ("Model Name",              model.name),
        ("Model ID",                str(model.id)),
        ("Exported At",             datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
        ("Schema Version",          SCHEMA_VERSION),
        (None, None),
        ("Total Project Cost",      _fmt_num(outputs.total_project_cost if outputs else None)),
        ("Equity Required",         _fmt_num(outputs.equity_required if outputs else None)),
        ("Stabilized NOI (annual)", _fmt_num(outputs.noi_stabilized if outputs else None)),
        ("Cap Rate on Cost",        _fmt_pct(outputs.cap_rate_on_cost_pct if outputs else None)),
        ("DSCR",                    _fmt_num(outputs.dscr if outputs else None)),
        ("Levered IRR",             _fmt_pct(outputs.project_irr_levered if outputs else None)),
        ("Total Timeline (months)", _to_v(outputs.total_timeline_months if outputs else None)),
        ("Computed At",             _to_v(outputs.computed_at if outputs else None)),
    ]
    for label, value in rows:
        ws.append([label, value])
    _autofit(ws, [30, 20])
    _protect_sheet(ws)


# ── Cash Flow (read-only) ─────────────────────────────────────────────────────

def _build_cash_flow(ws: Worksheet, cash_flows: list[CashFlow]) -> None:
    headers = [
        "Period", "Period Type", "Gross Revenue", "Vacancy Loss",
        "Effective Gross Income", "Operating Expenses", "CapEx Reserve",
        "NOI", "Debt Service", "Net Cash Flow", "Cash Balance",
    ]
    _header_row(ws, headers)
    for cf in cash_flows:
        ws.append([
            cf.period, _to_v(cf.period_type),
            _to_v(cf.gross_revenue), _to_v(cf.vacancy_loss),
            _to_v(cf.effective_gross_income), _to_v(cf.operating_expenses),
            _to_v(cf.capex_reserve), _to_v(cf.noi),
            _to_v(cf.debt_service), _to_v(cf.net_cash_flow),
            _to_v(cf.cumulative_cash_flow),
        ])
    _autofit(ws)
    ws.freeze_panes = "A2"
    _protect_sheet(ws)


# ── Uses (editable) ───────────────────────────────────────────────────────────

_USES_COLS = ["_key", "Label", "Phase", "Amount", "Timing", "Deferred?", "Notes"]
_USES_WIDTHS = [18, 36, 18, 16, 14, 10, 40]
_USES_EDITABLE = [2, 3, 4, 5, 6, 7]  # 1-based col indices that are unlocked

def _build_uses(ws: Worksheet, use_lines: list[UseLine]) -> None:
    _editable_header_row(ws, _USES_COLS)
    for ul in use_lines:
        key = f"use_line|{ul.id}"
        phase = _to_v(ul.phase)
        ws.append([
            key,
            ul.label,
            phase,
            _to_v(ul.amount),
            ul.timing_type or "first_day",
            "Yes" if ul.is_deferred else "No",
            ul.notes or "",
        ])
    _add_blank_rows(ws, 10, len(_USES_COLS), "use_line|new")
    _style_editable_sheet(ws, _USES_COLS, _USES_WIDTHS, _USES_EDITABLE)
    _add_instructions(ws, "Edit Label, Phase, Amount, Timing, Deferred, Notes. Do not change column A (_key).")


_PHASE_CHOICES = "acquisition, pre_construction, construction, renovation, conversion, operation, exit, other"
_TIMING_CHOICES = "first_day, spread_evenly, last_day"


# ── Income (editable) ─────────────────────────────────────────────────────────

_INCOME_COLS = [
    "_key", "Label", "Type", "Units", "$/Unit/Month",
    "Fixed $/Month", "Occupancy %", "Escalation %/yr", "Active Phases", "Notes",
]
_INCOME_WIDTHS = [18, 30, 20, 8, 14, 14, 14, 14, 36, 40]
_INCOME_EDITABLE = [2, 3, 4, 5, 6, 7, 8, 9, 10]

_STREAM_TYPE_CHOICES = (
    "residential_rent, commercial_rent, parking, laundry, utility_water, "
    "utility_electric, utility_gas, utility_internet, storage, pet_fee, "
    "deposit_forfeit, other"
)

def _build_income(ws: Worksheet, streams: list[IncomeStream]) -> None:
    _editable_header_row(ws, _INCOME_COLS)
    for s in streams:
        key = f"income_stream|{s.id}"
        ws.append([
            key,
            s.label,
            _to_v(s.stream_type),
            s.unit_count or "",
            _to_v(s.amount_per_unit_monthly) or "",
            _to_v(s.amount_fixed_monthly) or "",
            _to_v(s.stabilized_occupancy_pct),
            _to_v(s.escalation_rate_pct_annual),
            ", ".join(s.active_in_phases or []),
            s.notes or "",
        ])
    _add_blank_rows(ws, 10, len(_INCOME_COLS), "income_stream|new")
    _style_editable_sheet(ws, _INCOME_COLS, _INCOME_WIDTHS, _INCOME_EDITABLE)
    _add_instructions(ws, f"Type choices: {_STREAM_TYPE_CHOICES}. Active Phases: comma-separated phase names.")


# ── OpEx (editable) ───────────────────────────────────────────────────────────

_OPEX_COLS = [
    "_key", "Label", "Annual Amount", "Escalation %/yr",
    "Scale w/ Lease-Up?", "Lease-Up Floor %", "Active Phases",
]
_OPEX_WIDTHS = [18, 36, 16, 16, 18, 16, 36]
_OPEX_EDITABLE = [2, 3, 4, 5, 6, 7]

def _build_opex(ws: Worksheet, expense_lines: list[OperatingExpenseLine]) -> None:
    _editable_header_row(ws, _OPEX_COLS)
    for el in expense_lines:
        key = f"expense_line|{el.id}"
        ws.append([
            key,
            el.label,
            _to_v(el.annual_amount),
            _to_v(el.escalation_rate_pct_annual),
            "Yes" if el.scale_with_lease_up else "No",
            _to_v(el.lease_up_floor_pct) or "",
            ", ".join(el.active_in_phases or []),
        ])
    _add_blank_rows(ws, 10, len(_OPEX_COLS), "expense_line|new")
    _style_editable_sheet(ws, _OPEX_COLS, _OPEX_WIDTHS, _OPEX_EDITABLE)
    _add_instructions(ws, "Active Phases: comma-separated. Typical: lease_up, stabilized")


# ── Sources (editable) ────────────────────────────────────────────────────────

_SOURCES_COLS = [
    "_key", "Label", "Type", "Stack Position",
    "Amount", "Interest Rate %", "Amort Years",
    "Carry (Construction)", "Carry (Operation)",
    "Phase Start", "Phase End", "Auto-Size?",
]
_SOURCES_WIDTHS = [18, 30, 20, 14, 16, 16, 12, 24, 24, 18, 18, 10]
_SOURCES_EDITABLE = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]

_FUNDER_TYPE_CHOICES = (
    "permanent_debt, senior_debt, mezzanine_debt, bridge, construction_loan, "
    "soft_loan, bond, preferred_equity, common_equity, owner_loan, "
    "owner_investment, grant, tax_credit"
)
_CARRY_CHOICES = "io_only, pi, capitalized_interest, accruing, none"

def _build_sources(ws: Worksheet, capital_modules: list[CapitalModule]) -> None:
    _editable_header_row(ws, _SOURCES_COLS)
    for cm in capital_modules:
        key = f"capital_module|{cm.id}"
        src = cm.source or {}
        carry = cm.carry or {}

        # Flatten carry — handle phased or flat
        carry_constr = "none"
        carry_op = "none"
        if "phases" in carry:
            for ph in carry["phases"]:
                if ph.get("name") == "construction":
                    carry_constr = ph.get("carry_type", "none")
                elif ph.get("name") == "operation":
                    carry_op = ph.get("carry_type", "none")
        else:
            ct = carry.get("carry_type", "none")
            carry_constr = ct
            carry_op = ct

        ws.append([
            key,
            cm.label,
            _to_v(cm.funder_type),
            cm.stack_position,
            _to_v(src.get("amount")) or "",
            src.get("interest_rate_pct") or "",
            src.get("amort_term_years") or "",
            carry_constr,
            carry_op,
            cm.active_phase_start or "",
            cm.active_phase_end or "",
            "Yes" if src.get("auto_size") else "No",
        ])
    _add_blank_rows(ws, 5, len(_SOURCES_COLS), "capital_module|new")
    _style_editable_sheet(ws, _SOURCES_COLS, _SOURCES_WIDTHS, _SOURCES_EDITABLE)
    _add_instructions(
        ws,
        f"Type choices: {_FUNDER_TYPE_CHOICES}. "
        f"Carry choices: {_CARRY_CHOICES}. "
        "Auto-Size: Yes = amount computed by model at Compute time."
    )


# ── _setup (protected metadata) ───────────────────────────────────────────────

def _build_setup(
    ws: Worksheet,
    model: DealModel,
    project: Project | None,
    inputs: OperationalInputs | None,
    milestones: list[Milestone],
) -> None:
    """Key-value metadata sheet. Protected (locked) to provide re-import friction."""

    # Build milestone map for date resolution
    ms_map = {m.id: m for m in milestones}

    # Header
    ws.append(["_setup — Model Metadata (do not edit unless you know what you're doing)"])
    ws["A1"].font = Font(bold=True, color=_CLR_HINT)
    ws.append([])  # blank row

    rows: list[tuple[str, Any]] = [
        ("schema_version",          SCHEMA_VERSION),
        ("model_id",                str(model.id)),
        ("model_name",              model.name),
        ("exported_at",             datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")),
        ("",                        ""),
        # Project / operational
        ("deal_type",               _to_v(getattr(project, "deal_type", None)) if project else ""),
        ("unit_count_new",          _to_v(getattr(inputs, "unit_count_new", None)) if inputs else ""),
        ("unit_count_existing",     _to_v(getattr(inputs, "unit_count_existing", None)) if inputs else ""),
        ("hold_period_years",       _to_v(getattr(inputs, "hold_period_years", None)) if inputs else ""),
        ("exit_cap_rate_pct",       _to_v(getattr(inputs, "exit_cap_rate_pct", None)) if inputs else ""),
        ("selling_costs_pct",       _to_v(getattr(inputs, "selling_costs_pct", None)) if inputs else ""),
        ("lease_up_months",         _to_v(getattr(inputs, "lease_up_months", None)) if inputs else ""),
        ("initial_occupancy_pct",   _to_v(getattr(inputs, "initial_occupancy_pct", None)) if inputs else ""),
        ("",                        ""),
        # Deal setup wizard
        ("debt_structure",          _to_v(getattr(inputs, "debt_structure", None)) if inputs else ""),
        ("debt_sizing_mode",        _to_v(getattr(inputs, "debt_sizing_mode", None)) if inputs else ""),
        ("dscr_minimum",            _to_v(getattr(inputs, "dscr_minimum", None)) if inputs else ""),
        ("construction_floor_pct",  _to_v(getattr(inputs, "construction_floor_pct", None)) if inputs else ""),
        ("operation_reserve_months",_to_v(getattr(inputs, "operation_reserve_months", None)) if inputs else ""),
        ("",                        ""),
    ]

    # Milestone rows
    for m in sorted(milestones, key=lambda x: x.sequence_order):
        mt = str(m.milestone_type).replace("MilestoneType.", "")
        start = m.computed_start(ms_map)
        rows.append((
            f"milestone.{mt}.target_date",
            start.isoformat() if start else (m.target_date.isoformat() if m.target_date else ""),
        ))
        rows.append((f"milestone.{mt}.duration_days", m.duration_days or ""))

    for key, value in rows:
        ws.append([key, value])

    # Style
    fill_setup = PatternFill("solid", fgColor=_CLR_SETUP_BG)
    fill_lock  = PatternFill("solid", fgColor=_CLR_SETUP_LOCK)
    for row in ws.iter_rows():
        for cell in row:
            cell.fill = fill_setup
            cell.protection = Protection(locked=True)
            cell.font = Font(size=10)
            if cell.column == 1 and cell.value:
                cell.font = Font(size=10, bold=True, color="5D4037")

    _autofit(ws, [36, 28])
    ws.protection.sheet = True
    ws.protection.password = ""   # no password — user can unprotect via Excel ribbon, but needs deliberate action
    ws.protection.enable()


# ── Style helpers ─────────────────────────────────────────────────────────────

def _header_row(ws: Worksheet, headers: list[str]) -> None:
    """Write a bold, dark-navy header row."""
    ws.append(headers)
    fill = PatternFill("solid", fgColor=_CLR_HEADER_BG)
    font = Font(bold=True, color=_CLR_HEADER_FG, size=10)
    for cell in ws[ws.max_row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[ws.max_row].height = 18


def _editable_header_row(ws: Worksheet, headers: list[str]) -> None:
    """Header row for editable sheets — key column styled differently."""
    ws.append(headers)
    key_fill = PatternFill("solid", fgColor="BDBDBD")
    hdr_fill = PatternFill("solid", fgColor=_CLR_HEADER_BG)
    for i, cell in enumerate(ws[ws.max_row], start=1):
        if i == 1:
            cell.fill = key_fill
            cell.font = Font(bold=True, color="616161", size=9, italic=True)
        else:
            cell.fill = hdr_fill
            cell.font = Font(bold=True, color=_CLR_HEADER_FG, size=10)
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[ws.max_row].height = 18


def _add_blank_rows(ws: Worksheet, count: int, col_count: int, key_prefix: str) -> None:
    """Add blank rows with new-record keys so users can fill them in."""
    for i in range(count):
        row = [f"{key_prefix}"] + [""] * (col_count - 1)
        ws.append(row)


def _style_editable_sheet(
    ws: Worksheet,
    headers: list[str],
    widths: list[int],
    editable_cols: list[int],
) -> None:
    """Apply fills, borders, column widths, and protection to an editable sheet."""
    key_fill  = PatternFill("solid", fgColor=_CLR_KEY_BG)
    edit_fill = PatternFill("solid", fgColor=_CLR_EDITABLE)
    lock_fill = PatternFill("solid", fgColor=_CLR_LOCKED_BG)

    # Lock all cells first, then unlock editable columns
    ws.protection.sheet = True
    ws.protection.password = ""
    ws.protection.enable()

    # Style data rows (row 2 onwards, skipping instructions row which is added last)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            col_idx = cell.column
            if col_idx == 1:
                # Key column
                cell.fill = key_fill
                cell.font = Font(size=9, color=_CLR_KEY_FG, italic=True)
                cell.protection = Protection(locked=True)
            elif col_idx in editable_cols:
                cell.fill = edit_fill
                cell.font = Font(size=10)
                cell.protection = Protection(locked=False)  # UNLOCKED
                cell.border = _THIN_BORDER
            else:
                cell.fill = lock_fill
                cell.protection = Protection(locked=True)

    # Column widths
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 20


def _add_instructions(ws: Worksheet, text: str) -> None:
    """Append a grey instructions row at the bottom of an editable sheet."""
    ws.append([])
    ws.append(["ℹ Instructions", text])
    instr_row = ws.max_row
    hint_fill = PatternFill("solid", fgColor="E8EAF6")
    hint_font = Font(size=9, color="5C6BC0", italic=True)
    for cell in ws[instr_row]:
        cell.fill = hint_fill
        cell.font = hint_font
        cell.protection = Protection(locked=True)
    ws.merge_cells(
        start_row=instr_row, start_column=2,
        end_row=instr_row, end_column=min(12, ws.max_column),
    )


def _protect_sheet(ws: Worksheet) -> None:
    """Lock all cells on a read-only sheet."""
    ws.protection.sheet = True
    ws.protection.password = ""
    ws.protection.enable()
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)


def _autofit(ws: Worksheet, col_widths: list[int] | None = None) -> None:
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return
    for col_cells in ws.columns:
        values = [str(c.value) for c in col_cells if c.value is not None]
        width = max((len(v) for v in values), default=8) + 2
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(width, 10), 40)


# ── Value helpers ─────────────────────────────────────────────────────────────

def _to_v(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        f = float(value)
        return int(f) if f == int(f) else round(f, 6)
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Enum):
        return value.value
    if isinstance(value, UUID):
        return str(value)
    return value


def _fmt_num(value: Any) -> str:
    v = _to_v(value)
    if v == "" or v is None:
        return "—"
    try:
        return f"${float(v):,.0f}"
    except (TypeError, ValueError):
        return str(v)


def _fmt_pct(value: Any) -> str:
    v = _to_v(value)
    if v == "" or v is None:
        return "—"
    try:
        return f"{float(v):.2f}%"
    except (TypeError, ValueError):
        return str(v)


__all__ = ["export_deal_model_workbook", "make_export_filename"]
